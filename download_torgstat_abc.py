# VERSION: TORGSTAT_ABC_REPORT_FINAL_20260601
"""Download Torgstat/WB ABC report and upload it to Yandex Object Storage.

Repository filename should be: download_torgstat_abc.py

Primary mode: replay a browser "Copy as cURL" export request stored in secret
TORGSTAT_ABC_CURL. The script tries to replace the period in URL/body with the
requested dates, downloads XLSX, validates that it is an ABC report, and uploads
it with the SAME Torgstat naming pattern that existing report parsing expects:

  Отчёты/ABC/wb_abc_report_goods__01.05.2026-27.05.2026__at_2026-05-28_21-30.xlsx

Downstream report_runner.py reads ABC reports from exactly Отчёты/ABC/.
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import shlex
import sys
import tempfile
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

import boto3
import requests
from openpyxl import load_workbook

VERSION = "TORGSTAT_ABC_REPORT_FINAL_20260601"
DEFAULT_REPORTS_ROOT = "Отчёты"
DEFAULT_ABC_FOLDER = "ABC"
DEFAULT_STORE = "TOPFACE"
DEFAULT_TZ_OFFSET_HOURS = 3  # Used only for date choice in GitHub Actions.

START_KEYS = {
    "datefrom", "fromdate", "begindate", "startdate", "datestart", "periodstart",
    "date_from", "from_date", "begin_date", "start_date", "date_start", "period_start",
    "from", "start", "begin", "dtfrom", "dt_from", "dfrom", "date1", "startperiod",
    "period[from]", "filter[datefrom]", "filter[from]", "filter[startdate]",
}
END_KEYS = {
    "dateto", "todate", "enddate", "dateend", "periodend",
    "date_to", "to_date", "end_date", "date_end", "period_end",
    "to", "end", "dtto", "dt_to", "dto", "date2", "endperiod",
    "period[to]", "filter[dateto]", "filter[to]", "filter[enddate]",
}
DATE_RE = re.compile(r"(?:\d{4}-\d{2}-\d{2}|\d{2}\.\d{2}\.\d{4}|\d{2}/\d{2}/\d{4})")


@dataclass
class CurlRequest:
    url: str
    method: str = "GET"
    headers: Dict[str, str] = None
    body: Optional[bytes] = None

    def __post_init__(self) -> None:
        if self.headers is None:
            self.headers = {}


def log(msg: str) -> None:
    print(msg, flush=True)


def fail(msg: str, code: int = 1) -> None:
    print(f"ERROR: {msg}", file=sys.stderr, flush=True)
    raise SystemExit(code)


def load_report_env() -> None:
    """Load KEY=VALUE lines from REPORT_ENV into os.environ if missing/empty.

    GitHub Actions passes absent secrets as empty strings if they are listed in env.
    This loader therefore treats an existing empty env var as missing and fills it
    from REPORT_ENV. It also tolerates CRLF and pasted literal \n sequences.
    """
    raw = os.environ.get("REPORT_ENV", "") or ""
    if not raw.strip():
        log("report_env: REPORT_ENV is empty/not passed")
        return

    # Some copies end up with literal \n instead of real newlines.
    normalized = raw.replace("\r\n", "\n").replace("\r", "\n")
    if "\n" not in normalized and "\\n" in normalized:
        normalized = normalized.replace("\\n", "\n")

    loaded: List[str] = []
    skipped_existing: List[str] = []
    bad_lines = 0
    for line in normalized.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[len("export ") :].strip()
        if "=" not in line:
            bad_lines += 1
            continue
        key, value = line.split("=", 1)
        key = key.strip().lstrip("\ufeff")
        value = value.strip().strip('"').strip("'")
        if not key:
            bad_lines += 1
            continue
        if (os.environ.get(key) or "").strip():
            skipped_existing.append(key)
            continue
        os.environ[key] = value
        loaded.append(key)

    interesting = [
        "YC_ACCESS_KEY_ID", "YC_SECRET_ACCESS_KEY", "YC_BUCKET_NAME", "YC_ENDPOINT_URL",
        "AWS_ACCESS_KEY_ID", "AWS_SECRET_ACCESS_KEY", "S3_BUCKET", "S3_ENDPOINT_URL",
    ]
    state = ", ".join(f"{k}={'set' if (os.environ.get(k) or '').strip() else 'empty'}" for k in interesting)
    log(f"report_env: present=yes, loaded_keys={loaded}, skipped_existing={skipped_existing}, bad_lines={bad_lines}")
    log(f"env_state: {state}")


def parse_date(s: str) -> dt.date:
    s = (s or "").strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    fail(f"Неверный формат даты: {s}. Используй YYYY-MM-DD или DD.MM.YYYY")


def fmt_dmy(d: dt.date) -> str:
    return d.strftime("%d.%m.%Y")


def fmt_iso(d: dt.date) -> str:
    return d.strftime("%Y-%m-%d")


def today_local() -> dt.date:
    # Avoid timezone dependencies in GitHub runner.
    return (dt.datetime.now(dt.UTC) + dt.timedelta(hours=DEFAULT_TZ_OFFSET_HOURS)).date()


def period_for_mode(mode: str, date_from: Optional[str], date_to: Optional[str]) -> List[Tuple[dt.date, dt.date, str]]:
    mode = (mode or "auto").lower().strip()
    today = today_local()
    yesterday = today - dt.timedelta(days=1)

    def dedupe_periods(periods: List[Tuple[dt.date, dt.date, str]]) -> List[Tuple[dt.date, dt.date, str]]:
        seen = set()
        out: List[Tuple[dt.date, dt.date, str]] = []
        for start, end, label in periods:
            key = (start, end)
            if key in seen:
                continue
            seen.add(key)
            out.append((start, end, label))
        return out

    if mode == "custom":
        if not date_from or not date_to:
            fail("mode=custom требует --date-from и --date-to")
        start, end = parse_date(date_from), parse_date(date_to)
        return [(start, end, "custom")]
    if mode == "daily":
        target = parse_date(date_from) if date_from else yesterday
        return [(target, target, "daily")]
    if mode == "weekly":
        # Previous full Monday-Sunday week relative to local today.
        last_sunday = today - dt.timedelta(days=today.weekday() + 1)
        start = last_sunday - dt.timedelta(days=6)
        return [(start, last_sunday, "weekly")]
    if mode == "mtd":
        start = yesterday.replace(day=1)
        return [(start, yesterday, "mtd")]
    if mode == "auto":
        # Daily file is useful for point checks; MTD/full-month file is required for the current-month PDF block.
        # On Mondays we additionally export the closed Monday-Sunday week for exact ABC gross profit.
        periods = [
            (yesterday, yesterday, "daily"),
            (yesterday.replace(day=1), yesterday, "mtd"),
        ]
        if today.weekday() == 0:  # Monday
            last_sunday = today - dt.timedelta(days=1)
            start = last_sunday - dt.timedelta(days=6)
            periods.append((start, last_sunday, "weekly"))
        return dedupe_periods(periods)
    fail(f"Неизвестный mode={mode}. Допустимо: auto, daily, weekly, mtd, custom")


def clean_multiline_curl(raw: str) -> str:
    raw = (raw or "").strip()
    if not raw:
        fail("TORGSTAT_ABC_CURL пустой. Сохрани Copy as cURL в GitHub Secret TORGSTAT_ABC_CURL")
    # Chrome often copies line continuations with backslash-newline.
    raw = raw.replace("\\\r\n", " ").replace("\\\n", " ")
    raw = raw.replace("\r\n", " ").replace("\n", " ")
    return raw.strip()


REQUEST_HEADER_KEYS = {
    "accept", "accept-language", "authorization", "content-type", "cookie", "origin",
    "referer", "user-agent", "x-csrf-token", "x-requested-with", "sec-ch-ua",
    "sec-ch-ua-mobile", "sec-ch-ua-platform", "sec-fetch-dest", "sec-fetch-mode",
    "sec-fetch-site", "priority",
}


def parse_devtools_headers_block(raw: str) -> Optional[CurlRequest]:
    """Parse a manually copied Chrome DevTools Headers block.

    Fallback for cases when the secret contains the visible Headers panel rather
    than Copy -> Copy as cURL. Russian Chrome shows pairs like:
      URL запроса\nhttps://...\nМетод запроса\nGET\ncookie\n...
    """
    lines = [ln.strip() for ln in (raw or "").replace("\r\n", "\n").split("\n") if ln.strip()]
    if not lines:
        return None

    url: Optional[str] = None
    method = "GET"
    headers: Dict[str, str] = {}
    scheme = "https"
    authority = ""
    path = ""

    i = 0
    while i < len(lines):
        line = lines[i]
        low = line.lower()
        nxt = lines[i + 1] if i + 1 < len(lines) else ""

        if line.startswith(("http://", "https://")):
            url = line
            i += 1
            continue

        if low in {"url запроса", "request url"} and nxt:
            url = nxt
            i += 2
            continue
        if low in {"метод запроса", "request method"} and nxt:
            method = nxt.upper()
            i += 2
            continue

        if low == ":method" and nxt:
            method = nxt.upper()
            i += 2
            continue
        if low == ":scheme" and nxt:
            scheme = nxt
            i += 2
            continue
        if low == ":authority" and nxt:
            authority = nxt
            i += 2
            continue
        if low == ":path" and nxt:
            path = nxt
            i += 2
            continue

        if low in REQUEST_HEADER_KEYS and nxt:
            headers[line] = nxt
            i += 2
            continue

        i += 1

    if not url and authority and path:
        url = f"{scheme}://{authority}{path}"

    if not url:
        return None

    clean_headers: Dict[str, str] = {}
    for k, v in headers.items():
        lk = k.lower().strip()
        if lk in REQUEST_HEADER_KEYS:
            clean_headers[k] = v
    return CurlRequest(url=url, method=method, headers=clean_headers, body=None)


def parse_curl(raw: str) -> CurlRequest:
    raw = (raw or "").strip()
    looks_like_curl = bool(re.match(r"^\s*(curl|curl\.exe)\b", raw, flags=re.I))
    if not looks_like_curl:
        parsed = parse_devtools_headers_block(raw)
        if parsed:
            log("secret_parse: parsed DevTools Headers block, not cURL")
            return parsed

    text = clean_multiline_curl(raw)
    text = re.sub(r"\s+[\^`]\s+", " ", text)
    parts = shlex.split(text, posix=True)
    if not parts:
        fail("Copy as cURL не распознан")
    if parts[0].lower() in {"curl", "curl.exe"}:
        parts = parts[1:]
    url: Optional[str] = None
    method = "GET"
    headers: Dict[str, str] = {}
    body_parts: List[str] = []
    i = 0
    while i < len(parts):
        p = parts[i]
        if p in ("-X", "--request") and i + 1 < len(parts):
            method = parts[i + 1].upper()
            i += 2
            continue
        if p in ("-H", "--header") and i + 1 < len(parts):
            h = parts[i + 1]
            if ":" in h:
                k, v = h.split(":", 1)
                headers[k.strip()] = v.strip()
            i += 2
            continue
        if p in ("-b", "--cookie", "--cookie-raw") and i + 1 < len(parts):
            headers["cookie"] = parts[i + 1].strip()
            i += 2
            continue
        if p in ("--url",) and i + 1 < len(parts):
            url = parts[i + 1]
            i += 2
            continue
        if p in ("--data", "--data-raw", "--data-binary", "--data-ascii", "-d") and i + 1 < len(parts):
            body_parts.append(parts[i + 1])
            if method == "GET":
                method = "POST"
            i += 2
            continue
        if p in ("--compressed", "--location", "-L", "--insecure", "-k", "--globoff"):
            i += 1
            continue
        if p.startswith("http://") or p.startswith("https://"):
            url = p
            i += 1
            continue
        i += 1
    if not url:
        parsed = parse_devtools_headers_block(raw)
        if parsed:
            log("secret_parse: parsed DevTools Headers block after cURL parse fallback")
            return parsed
        fail("В Copy as cURL не найден URL")
    body = "&".join(body_parts).encode("utf-8") if body_parts else None
    return CurlRequest(url=url, method=method, headers=headers, body=body)

def norm_key(k: str) -> str:
    return re.sub(r"[^a-z0-9_\[\]]+", "", str(k).lower())


def format_like(original_value: Any, new_date: dt.date) -> str:
    s = str(original_value)
    if re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", s):
        return fmt_dmy(new_date)
    if re.fullmatch(r"\d{2}/\d{2}/\d{4}", s):
        return new_date.strftime("%d/%m/%Y")
    return fmt_iso(new_date)


def replace_date_tokens_text(text: str, start: dt.date, end: dt.date) -> Tuple[str, int]:
    """Fallback: replace the first two distinct date tokens in a text blob."""
    matches = list(DATE_RE.finditer(text))
    if not matches:
        return text, 0
    distinct: List[str] = []
    for m in matches:
        v = m.group(0)
        if v not in distinct:
            distinct.append(v)
    repl: Dict[str, str] = {}
    if len(distinct) >= 1:
        repl[distinct[0]] = format_like(distinct[0], start)
    if len(distinct) >= 2:
        repl[distinct[1]] = format_like(distinct[1], end)
    # If there is only one date, set it to start for daily/custom single-date APIs.
    out = text
    for old, new in repl.items():
        out = out.replace(old, new)
    return out, len(repl)


def replace_dates_in_mapping(items: List[Tuple[str, str]], start: dt.date, end: dt.date) -> Tuple[List[Tuple[str, str]], int]:
    changed = 0
    out: List[Tuple[str, str]] = []
    for k, v in items:
        nk = norm_key(k)
        if nk in START_KEYS or "datefrom" in nk or "startdate" in nk or nk.endswith("from"):
            out.append((k, format_like(v, start)))
            changed += 1
        elif nk in END_KEYS or "dateto" in nk or "enddate" in nk or nk.endswith("to"):
            out.append((k, format_like(v, end)))
            changed += 1
        else:
            out.append((k, v))
    return out, changed


def replace_dates_json(obj: Any, start: dt.date, end: dt.date) -> Tuple[Any, int]:
    changed = 0
    if isinstance(obj, dict):
        new = {}
        for k, v in obj.items():
            nk = norm_key(k)
            if nk in START_KEYS or "datefrom" in nk or "startdate" in nk or nk.endswith("from"):
                new[k] = format_like(v, start)
                changed += 1
            elif nk in END_KEYS or "dateto" in nk or "enddate" in nk or nk.endswith("to"):
                new[k] = format_like(v, end)
                changed += 1
            else:
                new[k], c = replace_dates_json(v, start, end)
                changed += c
        return new, changed
    if isinstance(obj, list):
        arr = []
        for v in obj:
            nv, c = replace_dates_json(v, start, end)
            arr.append(nv)
            changed += c
        return arr, changed
    return obj, 0


def update_request_dates(req: CurlRequest, start: dt.date, end: dt.date) -> CurlRequest:
    url = req.url
    total_changed = 0

    parsed = urlparse(url)
    q_items = parse_qsl(parsed.query, keep_blank_values=True)
    if q_items:
        new_q_items, c = replace_dates_in_mapping(q_items, start, end)
        total_changed += c
        url = urlunparse(parsed._replace(query=urlencode(new_q_items, doseq=True)))

    body = req.body
    content_type = ""
    for hk, hv in req.headers.items():
        if hk.lower() == "content-type":
            content_type = hv.lower()
            break

    if body:
        body_text = body.decode("utf-8", errors="replace")
        body_changed = 0
        if "json" in content_type or body_text.strip().startswith(("{", "[")):
            try:
                data = json.loads(body_text)
                new_data, body_changed = replace_dates_json(data, start, end)
                if body_changed:
                    body_text = json.dumps(new_data, ensure_ascii=False, separators=(",", ":"))
            except Exception:
                body_text, body_changed = replace_date_tokens_text(body_text, start, end)
        else:
            pairs = parse_qsl(body_text, keep_blank_values=True)
            if pairs:
                new_pairs, body_changed = replace_dates_in_mapping(pairs, start, end)
                if body_changed:
                    body_text = urlencode(new_pairs, doseq=True)
                else:
                    body_text, body_changed = replace_date_tokens_text(body_text, start, end)
            else:
                body_text, body_changed = replace_date_tokens_text(body_text, start, end)
        total_changed += body_changed
        body = body_text.encode("utf-8")

    if total_changed == 0:
        # Last-resort URL regex replacement. Useful when dates are embedded in path or compact payload.
        new_url, c = replace_date_tokens_text(url, start, end)
        total_changed += c
        url = new_url

    if total_changed == 0:
        fail(
            "Не нашёл даты в Copy as cURL, поэтому не могу безопасно заменить период. "
            "Открой Torgstat, выставь любой период, нажми Скачать и скопируй именно запрос скачивания из Network."
        )

    log(f"date_replace: changed_fields={total_changed}, period={fmt_dmy(start)}-{fmt_dmy(end)}")
    return CurlRequest(url=url, method=req.method, headers=dict(req.headers), body=body)


def request_download(req: CurlRequest) -> bytes:
    headers = dict(req.headers)
    # Let requests handle compressed response; remove HTTP/2 pseudo/noise headers if copied.
    for k in list(headers.keys()):
        lk = k.lower()
        if lk.startswith(":") or lk in {"content-length", "host"}:
            headers.pop(k, None)

    header_names = {k.lower() for k in headers}
    if "cookie" not in header_names and "authorization" not in header_names:
        fail(
            "В TORGSTAT_ABC_CURL нет авторизации: не найден header cookie/authorization. "
            "Скопируй не URL, а строку Network через: Правой кнопкой по wbapi-report-goods → "
            "Копировать → Копировать как cURL (bash), либо вставь полный блок Headers с cookie."
        )

    safe_header_list = ", ".join(sorted(k.lower() for k in headers.keys()))
    log(f"request_headers: {safe_header_list}")
    session = requests.Session()
    log(f"request: {req.method} {req.url.split('?')[0]}")
    resp = session.request(
        req.method,
        req.url,
        headers=headers,
        data=req.body,
        timeout=180,
        allow_redirects=True,
    )
    ct = resp.headers.get("content-type", "")
    log(f"response: status={resp.status_code}, content-type={ct}, bytes={len(resp.content):,}")
    if resp.status_code >= 400:
        preview = resp.text[:1000] if resp.text else ""
        fail(f"Torgstat вернул HTTP {resp.status_code}. Ответ: {preview}")

    content = resp.content
    # Some Torgstat exports return JSON with a temporary download link or embedded XLSX.
    if not looks_like_xlsx(content):
        try:
            js = resp.json()
            embedded = find_xlsx_bytes_in_json(js)
            if embedded:
                log("response: found embedded XLSX in JSON")
                content = embedded
            else:
                url = find_download_url(js)
                if url:
                    log("response: found download URL in JSON, fetching XLSX")
                    r2 = session.get(url, headers=headers, timeout=180, allow_redirects=True)
                    log(f"download_url_response: status={r2.status_code}, bytes={len(r2.content):,}")
                    if r2.status_code >= 400:
                        fail(f"Download URL вернул HTTP {r2.status_code}: {r2.text[:1000]}")
                    content = r2.content
        except Exception as e:
            log(f"response: JSON export parsing skipped: {e}")
    return content

def looks_like_xlsx(content: bytes) -> bool:
    return bool(content and content[:2] == b"PK" and len(content) > 1000)


def find_download_url(obj: Any) -> Optional[str]:
    if isinstance(obj, dict):
        for k, v in obj.items():
            if isinstance(v, str) and v.startswith(("http://", "https://")) and any(t in k.lower() for t in ["url", "link", "download", "file"]):
                return v
            found = find_download_url(v)
            if found:
                return found
    elif isinstance(obj, list):
        for v in obj:
            found = find_download_url(v)
            if found:
                return found
    return None


def find_xlsx_bytes_in_json(obj: Any) -> Optional[bytes]:
    """Find embedded XLSX bytes in a JSON response, including base64 data URLs."""
    import base64

    if isinstance(obj, str):
        s = obj.strip()
        if s.startswith("data:") and ";base64," in s:
            s = s.split(",", 1)[1].strip()
        # XLSX is a ZIP file, usually encoded as base64 starting with UEsDB.
        if len(s) > 1000 and re.fullmatch(r"[A-Za-z0-9+/=_\-\s]+", s):
            compact = re.sub(r"\s+", "", s)
            for candidate in (compact, compact.replace("-", "+").replace("_", "/")):
                try:
                    raw = base64.b64decode(candidate + "=" * (-len(candidate) % 4), validate=False)
                except Exception:
                    continue
                if looks_like_xlsx(raw):
                    return raw
        return None

    if isinstance(obj, dict):
        for v in obj.values():
            found = find_xlsx_bytes_in_json(v)
            if found:
                return found
    elif isinstance(obj, list):
        for v in obj:
            found = find_xlsx_bytes_in_json(v)
            if found:
                return found
    return None


def normalize_header(value: Any) -> str:
    s = str(value or "").strip().lower().replace("ё", "е")
    s = re.sub(r"\s+", " ", s)
    return s


def validate_xlsx(content: bytes) -> Tuple[str, int, List[str]]:
    if not looks_like_xlsx(content):
        sample = content[:500].decode("utf-8", errors="replace")
        fail(f"Скачанный файл не похож на XLSX. Первые байты/текст: {sample}")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name
    try:
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        required_any_gp = ["валовая прибыль", "валов прибыль", "gross profit"]
        required_any_art = ["артикул wb", "артикул вб", "nm", "nm id", "nmid"]
        best_sheet = ""
        best_row = 0
        best_headers: List[str] = []
        for ws in wb.worksheets:
            for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 30), values_only=True), start=1):
                headers = [normalize_header(v) for v in row if v is not None]
                if not headers:
                    continue
                has_gp = any(any(token in h for token in required_any_gp) for h in headers)
                has_art = any(any(token in h for token in required_any_art) for h in headers)
                if has_gp and has_art:
                    best_sheet = ws.title
                    best_row = ridx
                    best_headers = headers
                    return best_sheet, best_row, best_headers
        fail(
            "XLSX скачан, но не похож на АБС-отчёт: не нашёл одновременно колонки "
            "'Валовая прибыль' и 'Артикул WB' в первых 30 строках."
        )
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass


def output_filename(store: str, start: dt.date, end: dt.date) -> str:
    # IMPORTANT: preserve Torgstat filename family. Existing downstream code parses this pattern.
    # Example from manual downloads:
    # wb_abc_report_goods__27.04.2026-03.05.2026__at_2026-05-19_20-26.xlsx
    ts = (dt.datetime.now(dt.UTC) + dt.timedelta(hours=DEFAULT_TZ_OFFSET_HOURS)).strftime("%Y-%m-%d_%H-%M")
    return f"wb_abc_report_goods__{fmt_dmy(start)}-{fmt_dmy(end)}__at_{ts}.xlsx"


def output_key(store: str, start: dt.date, end: dt.date, reports_root: str, abc_folder: str) -> str:
    """Return the exact Object Storage key expected by report_runner.py.

    ABC reports must be stored directly under:
        Отчёты/ABC/

    There must be no store-level folder and no weekly/monthly subfolder here.
    """
    folder = abc_folder.strip("/") or DEFAULT_ABC_FOLDER
    if folder.replace("\\", "/").strip("/") != "ABC":
        log(f"WARN: abc_folder={folder!r}; downstream report expects 'ABC'. Using 'ABC'.")
        folder = "ABC"
    return f"{reports_root.rstrip('/')}/{folder}/{output_filename(store, start, end)}"


def s3_client():
    endpoint = os.environ.get("YC_ENDPOINT_URL") or os.environ.get("S3_ENDPOINT_URL") or "https://storage.yandexcloud.net"
    region = os.environ.get("YC_REGION") or os.environ.get("AWS_DEFAULT_REGION") or "ru-central1"
    access = os.environ.get("YC_ACCESS_KEY_ID") or os.environ.get("AWS_ACCESS_KEY_ID")
    secret = os.environ.get("YC_SECRET_ACCESS_KEY") or os.environ.get("AWS_SECRET_ACCESS_KEY")
    if not access or not secret:
        fail("Нет YC_ACCESS_KEY_ID/YC_SECRET_ACCESS_KEY. Добавь secrets или положи их в REPORT_ENV")
    return boto3.client(
        "s3",
        endpoint_url=endpoint,
        region_name=region,
        aws_access_key_id=access,
        aws_secret_access_key=secret,
    )


def upload_to_s3(content: bytes, key: str) -> None:
    bucket = os.environ.get("YC_BUCKET_NAME") or os.environ.get("S3_BUCKET") or os.environ.get("AWS_BUCKET")
    if not bucket:
        fail("Нет YC_BUCKET_NAME. Добавь secret или положи его в REPORT_ENV")
    client = s3_client()
    client.put_object(
        Bucket=bucket,
        Key=key,
        Body=content,
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    log(f"uploaded: s3://{bucket}/{key}")


def run_download(raw_curl: str, store: str, start: dt.date, end: dt.date, reports_root: str, abc_folder: str, dry_run: bool = False) -> str:
    base_req = parse_curl(raw_curl)
    req = update_request_dates(base_req, start, end)
    key = output_key(store, start, end, reports_root, abc_folder)
    filename = output_filename(store, start, end)
    log(f"target_filename: {filename}")
    log(f"target_key: {key}")
    if dry_run:
        log("dry_run: request was parsed and dates were replaced; download/upload skipped")
        return key
    content = request_download(req)
    sheet, header_row, headers = validate_xlsx(content)
    log(f"xlsx_validation: OK, sheet={sheet}, header_row={header_row}, headers_sample={headers[:8]}")
    upload_to_s3(content, key)
    return key


def self_test() -> None:
    sample = "curl 'https://example.com/export?dateFrom=2026-05-01&dateTo=2026-05-27' -H 'accept: application/json' -H 'cookie: session=abc' --data-raw '{\"startDate\":\"2026-05-01\",\"endDate\":\"2026-05-27\"}'"
    req = parse_curl(sample)
    new = update_request_dates(req, dt.date(2026, 5, 10), dt.date(2026, 5, 11))
    assert "2026-05-10" in new.url and "2026-05-11" in new.url
    assert b"2026-05-10" in (new.body or b"") and b"2026-05-11" in (new.body or b"")
    assert "cookie" in {k.lower() for k in new.headers}

    raw_headers = """URL запроса
https://torgstat.ru/api/trpc/wbapi-report-goods?batch=1&input=%7B%220%22%3A%7B%22dateFrom%22%3A%222026-05-01%22%2C%22dateTo%22%3A%222026-05-27%22%7D%7D
Метод запроса
GET
content-type
application/json
cookie
__Secure-next-auth.session-token=test
referer
https://torgstat.ru/wb-seller/sales
user-agent
Mozilla/5.0
"""
    req2 = parse_curl(raw_headers)
    new2 = update_request_dates(req2, dt.date(2026, 5, 12), dt.date(2026, 5, 13))
    assert "2026-05-12" in new2.url and "2026-05-13" in new2.url
    assert "cookie" in {k.lower() for k in new2.headers}

    cmd_curl = 'curl "https://example.com/export?dateFrom=2026-05-01&dateTo=2026-05-27" ^ -H "cookie: session=abc" ^ --compressed'
    req3 = parse_curl(cmd_curl)
    assert "cookie" in {k.lower() for k in req3.headers}

    fn = output_filename("TOPFACE", dt.date(2026, 5, 1), dt.date(2026, 5, 27))
    assert fn.startswith("wb_abc_report_goods__01.05.2026-27.05.2026__at_")
    assert fn.endswith(".xlsx")
    log("self-test: OK")

def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Download Torgstat ABC report to S3")
    parser.add_argument("--mode", default="auto", choices=["auto", "daily", "weekly", "mtd", "custom"])
    parser.add_argument("--store", default=DEFAULT_STORE)
    parser.add_argument("--date-from", default="")
    parser.add_argument("--date-to", default="")
    parser.add_argument("--reports-root", default=os.environ.get("REPORTS_ROOT", DEFAULT_REPORTS_ROOT))
    parser.add_argument("--abc-folder", default=os.environ.get("ABC_FOLDER", DEFAULT_ABC_FOLDER))
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args(argv)

    log(f"VERSION: {VERSION}")
    if args.self_test:
        self_test()
        return 0

    load_report_env()
    raw_curl = os.environ.get("TORGSTAT_ABC_CURL", "")
    periods = period_for_mode(args.mode, args.date_from or None, args.date_to or None)
    log(f"mode={args.mode}, store={args.store}, periods={[(fmt_dmy(a), fmt_dmy(b), label) for a,b,label in periods]}")
    uploaded_keys = []
    for start, end, label in periods:
        if start > end:
            fail(f"Некорректный период {fmt_dmy(start)}-{fmt_dmy(end)}")
        log(f"--- download {label}: {fmt_dmy(start)}-{fmt_dmy(end)} ---")
        key = run_download(raw_curl, args.store, start, end, args.reports_root, args.abc_folder, dry_run=args.dry_run)
        uploaded_keys.append(key)
        time.sleep(1)
    log("DONE")
    for key in uploaded_keys:
        log(f"RESULT_KEY={key}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
