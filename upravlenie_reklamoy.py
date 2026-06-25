#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FIX46_CORE_RAMP_PAUSE_20260611
WB Ads Manager — decision engine for ставки, CORE-трафик, разгон, паузы и возвраты.

Что изменено по новой логике:
- ABC-рентабельность полностью исключена из принятия решений.
- Финансовая оценка использует зрелое окно: последние 7 дней, исключая последние 3 дня лага.
  Например as_of=2026-06-11 => current=2026-06-01..2026-06-07, base=2026-05-25..2026-05-31.
- Search/CPC: ставка целая, шаг 1 рубль.
- Combined/CPM/полки: минимум 80 рублей, шаг 6 рублей. Значения CPC из статистики не считаются ставкой.
- Предельная ставка считается через прогнозный ДРР 16% и среднюю цену продажи finishedPrice из Orders.
- CORE оценивается не только кликами/позициями, но и кликами на 1 заказ и CPO запроса.
- Разгон: не больше 1 CPC и 1 CPM на товарный блок одновременно, до 5000 показов, затем пауза на дозревание.
- Пауза: не просто срез расхода, а pause_and_reallocate внутри product_root+placement.

Скрипт не отправляет API-вызовы сам. Он формирует решения и payload-preview.
Дальше текущий боевой runner должен отправлять только строки action in {raise, lower, pause, start}.

Пример:
python assistant_wb_ads_manager__FIX46_CORE_RAMP_PAUSE_20260611.py \
  --ads "Реклама_2026-W24.xlsx" \
  --orders "Заказы_2026-W24.xlsx" \
  --previous-output "Итог_последнего_запуска.xlsx" \
  --bid-history "История_ставок.xlsx" \
  --pause-history "История_пауз.xlsx" \
  --as-of 2026-06-11 \
  --out "wb_ads_decisions_FIX46.xlsx"
"""

from __future__ import annotations

import argparse
import io
import json
import math
import os
import re
import tempfile
import time
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import boto3
import numpy as np
import pandas as pd
import requests
from botocore.exceptions import ClientError

SCRIPT_VERSION = "v68-brush-pdf-only-2026-06-18"
VERSION = "FIX46_CORE_RAMP_PAUSE_20260611_V68_BRUSH_PDF_ONLY"

# -------------------------
# Business constants
# -------------------------
DRR_RAISE_GATE_PCT = 10.0          # ДРР < 10 => повышение можно рассматривать, но не автоматически
DRR_PAUSE_LIMIT_PCT = 15.0         # 14д + 10000 показов + ДРР > 15 => pause candidate
DRR_FORECAST_CAP_PCT = 16.0        # legacy потолок прогнозного ДРР для не-кистей
BRUSH_BID_CAP_DRR_PCT = 9.0         # кисти: максимальная ставка считается под целевой ДРР 9%
HARD_CAP_EXTRA_STEPS_ALLOWED = 1    # железное правило: максимум +1 шаг к рассчитанной ставке
SEARCH_MIN_BID_RUB = 4             # WB min can be fetched externally; safe default for CPC
SEARCH_STEP_RUB = 1
COMBINED_MIN_BID_RUB = 80
COMBINED_STEP_RUB = 6
RAMP_TARGET_IMPRESSIONS = 5000
RAMP_MAX_ACTIVE_CPC_PER_BLOCK = 1
RAMP_MAX_ACTIVE_CPM_PER_BLOCK = 1
NEW_NO_PAUSE_DAYS = 14
NEW_DAILY_IMPRESSIONS_LOW = 700
NEW_DAILY_IMPRESSIONS_HIGH = 1000
NEW_SEARCH_MAX_BID_RUB = 14
NEW_COMBINED_MAX_BID_RUB = 250
MATURE_WINDOW_DAYS = 7
DATA_LAG_DAYS = 3
PAUSE_WINDOW_DAYS = 14
POST_PAUSE_CHECK_DAYS = 7
FLAGSHIP_MAX_PER_PRODUCT = 3
FLAGSHIP_TARGET_POSITION = 10
FLAGSHIP_BEST_POSITION = 3

TARGET_SUBJECTS = {
    "кисти косметические": "Кисти косметические",
    "помады": "Помады",
    "блески": "Блески",
    "карандаши": "Косметические карандаши",
    "косметические карандаши": "Косметические карандаши",
}

# Жёсткий контур управления: любые решения/API только по этим 4 предметам.
MANAGED_SUBJECTS_CANON = {
    "Кисти косметические",
    "Помады",
    "Блески",
    "Косметические карандаши",
}
# Кисти не паузим автоматически: пауза разрешена только этим предметам.
PAUSE_ALLOWED_SUBJECTS_CANON = {"Помады", "Блески", "Косметические карандаши"}

# Артикулы, которые полностью исключены из автоматического управления ставками/паузами.
# Они не участвуют в расчёте решений, очереди разгона и API payload.
EXCLUDED_ARTICLES_FROM_AUTOMATION = {"901/4"}


def is_managed_subject_value(value: Any) -> bool:
    return canon_subject(value) in MANAGED_SUBJECTS_CANON


def is_pause_allowed_subject_value(value: Any) -> bool:
    return canon_subject(value) in PAUSE_ALLOWED_SUBJECTS_CANON


def filter_managed_subjects(df: pd.DataFrame, label: str = "") -> pd.DataFrame:
    """Оставляет только 4 управляемые категории перед расчётом решений/API.

    Это hard guard: если предмет пустой или не входит в контур, строка не должна
    попасть ни в решения, ни тем более в API.
    """
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df.copy()
    if "subject_norm" not in df.columns:
        print(f"Диагностика фильтра категорий {label}: нет subject_norm, строка/таблица исключена", flush=True)
        return df.iloc[0:0].copy()
    out = df.copy()
    before = len(out)
    out["subject_norm"] = out["subject_norm"].map(canon_subject)
    mask = out["subject_norm"].map(lambda x: x in MANAGED_SUBJECTS_CANON)
    removed = before - int(mask.sum())
    if removed:
        removed_subjects = out.loc[~mask, "subject_norm"].astype(str).value_counts().head(20).to_dict()
        print(
            f"Диагностика фильтра категорий {label}: оставлено {int(mask.sum())} из {before}; "
            f"исключено {removed}; исключённые предметы={removed_subjects}",
            flush=True,
        )
    return out.loc[mask].copy()

# -------------------------
# Column aliases
# -------------------------
ALIASES = {
    "campaign_id": ["campaign_id", "ID кампании", "id кампании", "advertId", "advert_id"],
    "nm_id": ["nm_id", "nmId", "Артикул WB", "Номенклатура WB", "nmID"],
    "supplier_article": ["supplier_article", "supplierArticle", "Артикул продавца", "Артикул"],
    "subject": ["subject", "subject_norm", "Название предмета", "Предмет", "category", "Категория"],
    "day": ["day", "Дата", "date", "dt"],
    "campaign_name": ["Название", "name", "campaign_name", "Кампания"],
    "campaign_status": ["Статус", "campaign_status", "status"],
    "payment_type": ["Тип оплаты", "payment_type", "placement", "Тип кампании"],
    "bid_type": ["Тип ставки", "bid_type"],
    "search_bid": ["Ставка в поиске (руб)", "search_bid", "bid_search", "Ставка в поиске"],
    "reco_bid": ["Ставка в рекомендациях (руб)", "reco_bid", "bid_recommendations", "Ставка в рекомендациях"],
    "impressions": ["Показы", "impressions", "shows"],
    "clicks": ["Клики", "clicks", "Переходы", "transitions", "clicks_to_card"],
    "ctr_pct": ["CTR", "ctr", "CTR, %", "ctr_pct"],
    "cpc": ["CPC", "cpc"],
    "orders": ["Заказы", "orders", "ad_orders", "keyword_orders"],
    "spend": ["Расход", "Затраты", "spend", "ad_spend"],
    "order_sum": ["Сумма заказов", "revenue", "order_sum", "ad_order_sum"],
    "drr_pct": ["ДРР", "Доля затрат", "drr", "campaign_drr_pct", "Доля затрат, %"],
    "frequency": ["query_freq", "Частотность", "frequency", "freq"],
    "query": ["query_text", "query_text_norm", "Поисковый запрос", "search_query", "Запрос"],
    "median_position": ["median_position", "Медианная позиция", "Позиция", "position"],
    "visibility_pct": ["visibility_pct", "Видимость", "visibility"],
    "finished_price": ["finishedPrice", "finished_price", "Цена продажи", "finished price"],
    "price_with_disc": ["priceWithDisc", "price_with_disc"],
    "is_cancel": ["isCancel", "is_cancel", "Отменено"],
}


def norm_col(x: Any) -> str:
    return re.sub(r"\s+", " ", str(x).strip())


def find_col(df: pd.DataFrame, logical: str) -> Optional[str]:
    aliases = ALIASES.get(logical, [logical])
    lower_map = {norm_col(c).lower(): c for c in df.columns}
    for a in aliases:
        key = norm_col(a).lower()
        if key in lower_map:
            return lower_map[key]
    # fuzzy normalized
    normalized = {re.sub(r"[^a-zа-я0-9]+", "", norm_col(c).lower()): c for c in df.columns}
    for a in aliases:
        key = re.sub(r"[^a-zа-я0-9]+", "", norm_col(a).lower())
        if key in normalized:
            return normalized[key]
    return None


def s(df: pd.DataFrame, logical: str, default: Any = np.nan) -> pd.Series:
    col = find_col(df, logical)
    if col is None:
        return pd.Series([default] * len(df), index=df.index)
    return df[col]


def to_num(x: pd.Series) -> pd.Series:
    if x is None:
        return pd.Series(dtype=float)
    return pd.to_numeric(
        x.astype(str)
        .str.replace("\u00a0", " ", regex=False)
        .str.replace("₽", "", regex=False)
        .str.replace("%", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace(",", ".", regex=False),
        errors="coerce",
    )


def to_date(x: pd.Series) -> pd.Series:
    """Robust date parser without global ambiguous parsing.

    Supported formats:
    - ISO: YYYY-MM-DD
    - Russian/Excel exports: DD.MM.YYYY or DD/MM/YYYY
    - already parsed Excel datetimes.
    """
    if x is None:
        return pd.Series(dtype="datetime64[ns]")
    raw = x if isinstance(x, pd.Series) else pd.Series(x)
    result = pd.Series(pd.NaT, index=raw.index, dtype="datetime64[ns]")
    text = raw.astype(str).str.strip()

    iso_mask = text.str.fullmatch(r"\d{4}-\d{2}-\d{2}", na=False)
    if iso_mask.any():
        result.loc[iso_mask] = pd.to_datetime(text.loc[iso_mask], format="%Y-%m-%d", errors="coerce")

    dot_mask = result.isna() & text.str.fullmatch(r"\d{2}\.\d{2}\.\d{4}", na=False)
    if dot_mask.any():
        result.loc[dot_mask] = pd.to_datetime(text.loc[dot_mask], format="%d.%m.%Y", errors="coerce")

    slash_mask = result.isna() & text.str.fullmatch(r"\d{2}/\d{2}/\d{4}", na=False)
    if slash_mask.any():
        result.loc[slash_mask] = pd.to_datetime(text.loc[slash_mask], format="%d/%m/%Y", errors="coerce")

    remaining = result.isna() & raw.notna() & text.ne("") & text.ne("NaT") & text.ne("nan")
    if remaining.any():
        result.loc[remaining] = pd.to_datetime(raw.loc[remaining], errors="coerce")

    return result.dt.normalize()


def clean_article(v: Any) -> str:
    if pd.isna(v):
        return ""
    return str(v).strip().replace("\\", "/")


def product_root(article: Any) -> str:
    a = clean_article(article)
    if not a:
        return ""
    # PT901.F26 => 901, 155/12 => 155, 617/1 => 617
    m = re.search(r"(\d{3})", a)
    return m.group(1) if m else a.split("/")[0].split(".")[0]


def canon_subject(v: Any) -> str:
    if pd.isna(v):
        return ""
    x = str(v).strip().lower()
    x = x.replace("ё", "е")
    if "кист" in x:
        return "Кисти косметические"
    if "помад" in x:
        return "Помады"
    if "блес" in x:
        return "Блески"
    if "карандаш" in x:
        return "Косметические карандаши"
    return str(v).strip()


def placement_from_payment(v: Any, bid_type: Any = "") -> str:
    x = f"{v} {bid_type}".lower()
    if "cpc" in x or "поиск" in x:
        return "search"
    if "cpm" in x or "единая" in x or "пол" in x or "combined" in x:
        return "combined"
    # default from old code terms
    if "search" in x:
        return "search"
    return "combined"


def safe_div(a: float, b: float, default: float = np.nan) -> float:
    try:
        if b is None or pd.isna(b) or float(b) == 0:
            return default
        return float(a) / float(b)
    except Exception:
        return default


def pct(a: float, b: float) -> float:
    return safe_div(a, b, np.nan) * 100.0


def floor_to_combined_grid(value: float) -> Optional[int]:
    if pd.isna(value):
        return None
    if value < COMBINED_MIN_BID_RUB:
        return COMBINED_MIN_BID_RUB
    steps = math.floor((value - COMBINED_MIN_BID_RUB) / COMBINED_STEP_RUB)
    return int(COMBINED_MIN_BID_RUB + steps * COMBINED_STEP_RUB)


def ceil_to_combined_grid(value: float) -> Optional[int]:
    if pd.isna(value):
        return None
    if value <= COMBINED_MIN_BID_RUB:
        return COMBINED_MIN_BID_RUB
    steps = math.ceil((value - COMBINED_MIN_BID_RUB) / COMBINED_STEP_RUB)
    return int(COMBINED_MIN_BID_RUB + steps * COMBINED_STEP_RUB)


def date_windows(as_of: pd.Timestamp) -> Dict[str, pd.Timestamp]:
    as_of = pd.Timestamp(as_of).normalize()
    current_end = as_of - pd.Timedelta(days=DATA_LAG_DAYS + 1)  # 11 => 7
    current_start = current_end - pd.Timedelta(days=MATURE_WINDOW_DAYS - 1)
    base_end = current_start - pd.Timedelta(days=1)
    base_start = base_end - pd.Timedelta(days=MATURE_WINDOW_DAYS - 1)
    pause_end = as_of - pd.Timedelta(days=1)
    pause_start = pause_end - pd.Timedelta(days=PAUSE_WINDOW_DAYS - 1)
    return {
        "as_of": as_of,
        "current_start": current_start,
        "current_end": current_end,
        "base_start": base_start,
        "base_end": base_end,
        "pause_start": pause_start,
        "pause_end": pause_end,
    }


def in_window(df: pd.DataFrame, start: pd.Timestamp, end: pd.Timestamp, day_col: str = "day") -> pd.DataFrame:
    if df.empty or day_col not in df.columns:
        return df.iloc[0:0].copy()
    day = pd.to_datetime(df[day_col], errors="coerce").dt.normalize()
    return df[(day >= start) & (day <= end)].copy()


def read_sheet(path: Optional[str], preferred: Sequence[str] = ()) -> pd.DataFrame:
    if not path or not os.path.exists(path):
        return pd.DataFrame()
    xl = pd.ExcelFile(path)
    for p in preferred:
        for sh in xl.sheet_names:
            if sh.strip().lower() == p.strip().lower():
                return pd.read_excel(path, sheet_name=sh)
    # if preferred not found, first sheet
    return pd.read_excel(path, sheet_name=xl.sheet_names[0])


def expand_input_paths(spec: Optional[str]) -> List[str]:
    """Accepts one path, comma/semicolon separated paths, or glob masks."""
    if not spec:
        return []
    parts = re.split(r"[;,]", str(spec))
    paths: List[str] = []
    for part in parts:
        p = part.strip().strip('"').strip("'")
        if not p:
            continue
        if any(ch in p for ch in ["*", "?", "["]):
            paths.extend([str(x) for x in sorted(Path().glob(p))])
        else:
            paths.append(p)
    # keep existing only, preserve order and dedupe
    seen = set()
    out = []
    for p in paths:
        if p not in seen and os.path.exists(p):
            out.append(p)
            seen.add(p)
    return out


def read_previous_sheet(path: Optional[str], sheet_name: str) -> pd.DataFrame:
    if not path or not os.path.exists(path):
        return pd.DataFrame()
    try:
        xl = pd.ExcelFile(path)
        for sh in xl.sheet_names:
            if sh.strip().lower() == sheet_name.strip().lower():
                return pd.read_excel(path, sheet_name=sh)
    except Exception:
        pass
    return pd.DataFrame()


# -------------------------
# Normalizers
# -------------------------

def load_ads_daily(path: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    paths = expand_input_paths(path)
    if not paths:
        raise RuntimeError(f"Не найден ни один файл рекламы по --ads={path}")
    daily_frames: List[pd.DataFrame] = []
    campaign_frames: List[pd.DataFrame] = []
    for pth in paths:
        daily_raw = read_sheet(pth, ["Статистика_Ежедневно", "daily", "Статистика ежедневно"])
        campaigns_raw = read_sheet(pth, ["Список_кампаний", "campaigns", "Кампании"])
        if daily_raw.empty:
            continue
        daily = pd.DataFrame({
            "campaign_id": to_num(s(daily_raw, "campaign_id")).astype("Int64"),
            "nm_id": to_num(s(daily_raw, "nm_id")).astype("Int64"),
            "campaign_name": s(daily_raw, "campaign_name", "").astype(str),
            "subject_norm": s(daily_raw, "subject", "").map(canon_subject),
            "day": to_date(s(daily_raw, "day")),
            "impressions": to_num(s(daily_raw, "impressions")).fillna(0.0),
            "clicks": to_num(s(daily_raw, "clicks")).fillna(0.0),
            "ctr_pct_raw": to_num(s(daily_raw, "ctr_pct")),
            "cpc_stat": to_num(s(daily_raw, "cpc")),
            "orders": to_num(s(daily_raw, "orders")).fillna(0.0),
            "spend": to_num(s(daily_raw, "spend")).fillna(0.0),
            "order_sum": to_num(s(daily_raw, "order_sum")).fillna(0.0),
            "drr_pct_raw": to_num(s(daily_raw, "drr_pct")),
            "source_file": Path(pth).name,
        })
        daily = daily[daily["campaign_id"].notna()].copy()
        if not daily.empty:
            daily["campaign_id"] = daily["campaign_id"].astype(int)
            daily["nm_id"] = daily["nm_id"].astype("Int64")
            daily_frames.append(daily)
        if not campaigns_raw.empty:
            c = normalize_campaigns(campaigns_raw)
            c["source_file"] = Path(pth).name
            campaign_frames.append(c)
    if not daily_frames:
        raise RuntimeError(f"Не найден лист статистики рекламы в файлах: {paths}")
    out = pd.concat(daily_frames, ignore_index=True)
    # Dedupe exact day+campaign rows across overlapping exports; keep latest file in input order.
    out["_source_order"] = out["source_file"].map({Path(p).name: i for i, p in enumerate(paths)})
    out = out.sort_values("_source_order").drop_duplicates(["campaign_id", "nm_id", "day"], keep="last").drop(columns=["_source_order"])

    camp = pd.concat(campaign_frames, ignore_index=True) if campaign_frames else pd.DataFrame()
    if not camp.empty:
        camp["_source_order"] = camp["source_file"].map({Path(p).name: i for i, p in enumerate(paths)})
        camp = camp.sort_values("_source_order").drop_duplicates("campaign_id", keep="last").drop(columns=["_source_order"])
        out = out.merge(camp[["campaign_id", "supplier_article", "product_root", "placement", "campaign_status", "real_bid_rub", "search_bid", "reco_bid"]], on="campaign_id", how="left")
    else:
        out["supplier_article"] = ""
        out["product_root"] = ""
        out["placement"] = ""
        out["campaign_status"] = ""
        out["real_bid_rub"] = np.nan
        out["search_bid"] = np.nan
        out["reco_bid"] = np.nan
    out["placement"] = out["placement"].fillna("").replace("", "search")

    # Сначала сохраняем предмет из ежедневной статистики. Если в списке кампаний
    # предмет пустой/нецелевой, используем most frequent subject из daily по campaign_id.
    if "subject_norm" in out.columns:
        subject_by_campaign = (
            out[["campaign_id", "subject_norm"]]
            .dropna()
            .assign(subject_norm=lambda d: d["subject_norm"].map(canon_subject))
            .groupby("campaign_id")["subject_norm"]
            .agg(lambda s: s.value_counts().index[0] if len(s) else "")
            .to_dict()
        )
        if camp is not None and not camp.empty and "campaign_id" in camp.columns:
            camp = camp.copy()
            camp["subject_from_ads_daily"] = camp["campaign_id"].map(subject_by_campaign).fillna("")
            bad_subject = ~camp.get("subject_norm", pd.Series([""] * len(camp), index=camp.index)).map(is_managed_subject_value)
            fill_mask = bad_subject & camp["subject_from_ads_daily"].map(is_managed_subject_value)
            camp.loc[fill_mask, "subject_norm"] = camp.loc[fill_mask, "subject_from_ads_daily"].to_numpy()
        if "subject_norm" in out.columns and "subject_norm_x" not in out.columns:
            # Если после merge появились дубликаты, ниже они будут обработаны в build_campaign_base.
            pass
    out = filter_managed_subjects(out, "ads_daily")
    camp = filter_managed_subjects(camp, "campaigns") if camp is not None and not camp.empty else camp
    return out, camp


def normalize_campaigns(df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame({
        "campaign_id": to_num(s(df, "campaign_id")).astype("Int64"),
        "campaign_name": s(df, "campaign_name", "").astype(str),
        "campaign_status": s(df, "campaign_status", "").astype(str),
        "payment_type": s(df, "payment_type", "").astype(str),
        "bid_type": s(df, "bid_type", "").astype(str),
        "search_bid": to_num(s(df, "search_bid")),
        "reco_bid": to_num(s(df, "reco_bid")),
        "subject_norm": s(df, "subject", "").map(canon_subject),
        "nm_id": to_num(s(df, "nm_id")).astype("Int64"),
    })
    out = out[out["campaign_id"].notna()].copy()
    out["campaign_id"] = out["campaign_id"].astype(int)
    out["placement"] = [placement_from_payment(a, b) for a, b in zip(out["payment_type"], out["bid_type"])]
    out["supplier_article"] = out["campaign_name"].map(extract_article_from_campaign_name)
    out["product_root"] = out["supplier_article"].map(product_root)

    # Real bid: search uses search_bid, combined/CPM uses max of available bids and never below 80.
    real = []
    for _, r in out.iterrows():
        if r["placement"] == "search":
            val = r["search_bid"]
            if pd.isna(val):
                val = r["reco_bid"]
            real.append(float(round(val)) if pd.notna(val) else np.nan)
        else:
            vals = [v for v in [r["search_bid"], r["reco_bid"]] if pd.notna(v)]
            val = max(vals) if vals else np.nan
            if pd.notna(val):
                val = max(float(val), COMBINED_MIN_BID_RUB)
            real.append(val)
    out["real_bid_rub"] = real
    out["is_active"] = out["campaign_status"].str.lower().str.contains("актив|active", na=False)
    return out


def extract_article_from_campaign_name(name: Any) -> str:
    if pd.isna(name):
        return ""
    text = str(name).strip()
    # campaign names like 155/15, PT901.F26, 156/16
    m = re.search(r"(PT\s*\d{3}[\.\-/]?[A-Z]?\d*|\d{3}\s*/\s*\d+|\d{3}[\./][A-Za-z0-9]+)", text, flags=re.I)
    if m:
        return m.group(1).replace(" ", "")
    # fallback root only
    m = re.search(r"\b(\d{3})\b", text)
    return m.group(1) if m else ""


def load_orders(path: Optional[str]) -> pd.DataFrame:
    paths = expand_input_paths(path)
    frames: List[pd.DataFrame] = []
    for pth in paths:
        df = read_sheet(pth, ["Заказы", "orders"])
        if df.empty:
            continue
        out = pd.DataFrame({
            "day": to_date(s(df, "day")),
            "supplier_article": s(df, "supplier_article", "").map(clean_article),
            "nm_id": to_num(s(df, "nm_id")).astype("Int64"),
            "subject_norm": s(df, "subject", "").map(canon_subject),
            "finished_price": to_num(s(df, "finished_price")),
            "price_with_disc": to_num(s(df, "price_with_disc")),
            "is_cancel": s(df, "is_cancel", False),
            "source_file": Path(pth).name,
        })
        out["product_root"] = out["supplier_article"].map(product_root)
        cancel = out["is_cancel"].astype(str).str.lower().isin(["true", "1", "да", "yes"])
        out = out[~cancel].copy()
        frames.append(out)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def load_bid_history(path: Optional[str]) -> pd.DataFrame:
    df = read_sheet(path, ["Лист1", "История_изменений_ставок", "История ставок"])
    base_columns = [
        "campaign_id", "event_date", "run_datetime", "old_bid_rub", "new_bid_rub", "direction", "reason_code",
        "supplier_article", "subject_norm", "placement", "nm_id", "api_status", "postcheck_status",
    ]
    if df.empty:
        return pd.DataFrame(columns=base_columns)
    out = pd.DataFrame({
        "campaign_id": to_num(s(df, "campaign_id")).astype("Int64"),
        "event_date": to_date(s(df, "day", np.nan) if find_col(df, "day") else df.get("event_date", pd.Series([np.nan] * len(df)))),
        "run_datetime": pd.to_datetime(df.get("run_datetime", pd.Series([pd.NaT] * len(df))), errors="coerce"),
        "old_bid_rub": to_num(df.get("old_bid_rub", pd.Series([np.nan] * len(df)))),
        "new_bid_rub": to_num(df.get("new_bid_rub", pd.Series([np.nan] * len(df)))),
        "direction": df.get("direction", pd.Series([""] * len(df))).astype(str),
        "reason_code": df.get("reason_code", pd.Series([""] * len(df))).astype(str),
        "supplier_article": df.get("supplier_article", pd.Series([""] * len(df))).astype(str),
        "subject_norm": df.get("subject_norm", pd.Series([""] * len(df))).astype(str).map(canon_subject),
        "placement": df.get("placement", pd.Series([""] * len(df))).astype(str),
        "nm_id": df.get("nm_id", pd.Series([""] * len(df))),
        "api_status": df.get("api_status", pd.Series([""] * len(df))).astype(str),
        "postcheck_status": df.get("postcheck_status", pd.Series([""] * len(df))).astype(str),
    })
    out = out[out["campaign_id"].notna()].copy()
    out["campaign_id"] = out["campaign_id"].astype(int)
    return out[base_columns]


def load_pause_history(path: Optional[str]) -> pd.DataFrame:
    df = read_sheet(path, ["Лист1", "История_пауз", "История пауз"])
    columns = [
        "campaign_id", "pause_date", "status", "reason_code", "api_status",
        "nm_id", "placement", "supplier_article", "subject_norm", "new_bid_rub",
    ]
    if df.empty:
        return pd.DataFrame(columns=columns)
    out = pd.DataFrame({
        "campaign_id": to_num(s(df, "campaign_id")).astype("Int64"),
        "pause_date": to_date(df.get("pause_date", pd.Series([np.nan] * len(df)))),
        "status": df.get("status", pd.Series([""] * len(df))).astype(str),
        "reason_code": df.get("reason_code", pd.Series([""] * len(df))).astype(str),
        "api_status": df.get("api_status", pd.Series([""] * len(df))).astype(str),
        "nm_id": df.get("nm_id", pd.Series([""] * len(df))),
        "placement": df.get("placement", pd.Series([""] * len(df))).astype(str),
        "supplier_article": df.get("supplier_article", pd.Series([""] * len(df))).astype(str),
        "subject_norm": df.get("subject_norm", pd.Series([""] * len(df))).astype(str),
        "new_bid_rub": df.get("new_bid_rub", pd.Series([np.nan] * len(df))),
    })
    out = out[out["campaign_id"].notna()].copy()
    out["campaign_id"] = out["campaign_id"].astype(int)
    return out[columns]


def load_keywords_from_previous(path: Optional[str]) -> pd.DataFrame:
    """Loads CORE key phrases from previous output if available."""
    df = read_previous_sheet(path, "Ключевые_фразы_80")
    if df.empty:
        return pd.DataFrame()
    out = pd.DataFrame({
        "nm_id": to_num(s(df, "nm_id")).astype("Int64"),
        "supplier_article": s(df, "supplier_article", "").map(clean_article),
        "subject_norm": s(df, "subject", "").map(canon_subject),
        "query_text": s(df, "query", "").astype(str),
        "orders": to_num(s(df, "orders")).fillna(0.0),
        "frequency": to_num(s(df, "frequency")).fillna(0.0),
        "median_position": to_num(s(df, "median_position")),
        "visibility_pct": to_num(s(df, "visibility_pct")),
        "clicks": to_num(s(df, "clicks")).fillna(0.0),
        "keyword_group": df.get("keyword_group", pd.Series([""]*len(df))).astype(str),
        "orders_share": to_num(df.get("orders_share", pd.Series([np.nan]*len(df)))),
        "orders_cum_share": to_num(df.get("orders_cum_share", pd.Series([np.nan]*len(df)))),
    })
    out["product_root"] = out["supplier_article"].map(product_root)
    out = out[out["query_text"].ne("")].copy()
    # The previous output may contain tens of thousands of tail queries. For decision logic we need
    # CORE only; tail queries are used later only as diagnostics. This keeps the run fast/stable.
    if "keyword_group" in out.columns and out["keyword_group"].astype(str).str.len().gt(0).any():
        core_mask = out["keyword_group"].astype(str).str.upper().str.contains("CORE")
        if core_mask.any():
            out = out[core_mask].copy()
    elif "orders_cum_share" in out.columns:
        out = out[out["orders_cum_share"].fillna(999) <= 0.85].copy()
    # Safety cap: no more than top 30 queries per product by orders/frequency.
    if not out.empty:
        out = (out.sort_values(["product_root", "orders", "frequency"], ascending=[True, False, False])
                 .groupby("product_root", as_index=False)
                 .head(30)
                 .reset_index(drop=True))
    return out


# -------------------------
# Metrics
# -------------------------

def campaign_window_metrics(ads: pd.DataFrame, start: pd.Timestamp, end: pd.Timestamp, suffix: str) -> pd.DataFrame:
    w = in_window(ads, start, end)
    if w.empty:
        return pd.DataFrame(columns=["campaign_id"])
    grp = w.groupby("campaign_id", as_index=False).agg(
        impressions=("impressions", "sum"),
        clicks=("clicks", "sum"),
        orders=("orders", "sum"),
        spend=("spend", "sum"),
        order_sum=("order_sum", "sum"),
    )
    grp[f"ctr_pct_{suffix}"] = np.where(grp["impressions"] > 0, grp["clicks"] / grp["impressions"] * 100, np.nan)
    grp[f"cpc_{suffix}"] = np.where(grp["clicks"] > 0, grp["spend"] / grp["clicks"], np.nan)
    grp[f"cpo_{suffix}"] = np.where(grp["orders"] > 0, grp["spend"] / grp["orders"], np.nan)
    grp[f"clicks_per_order_{suffix}"] = np.where(grp["orders"] > 0, grp["clicks"] / grp["orders"], np.nan)
    grp[f"impressions_per_order_{suffix}"] = np.where(grp["orders"] > 0, grp["impressions"] / grp["orders"], np.nan)
    grp[f"drr_pct_{suffix}"] = np.where(grp["order_sum"] > 0, grp["spend"] / grp["order_sum"] * 100, np.nan)
    return grp.rename(columns={
        "impressions": f"impressions_{suffix}",
        "clicks": f"clicks_{suffix}",
        "orders": f"orders_{suffix}",
        "spend": f"spend_{suffix}",
        "order_sum": f"order_sum_{suffix}",
    })


def avg_price_by_product(orders: pd.DataFrame, start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    if orders.empty:
        return pd.DataFrame(columns=["product_root", "nm_id", "avg_finished_price"])
    w = in_window(orders, start, end)
    if w.empty:
        return pd.DataFrame(columns=["product_root", "nm_id", "avg_finished_price"])
    by_nm = w.groupby(["product_root", "nm_id"], dropna=False, as_index=False).agg(
        avg_finished_price=("finished_price", "mean"),
        orders_rows=("finished_price", "size"),
    )
    by_root = w.groupby("product_root", as_index=False).agg(
        avg_finished_price_root=("finished_price", "mean"),
        orders_rows_root=("finished_price", "size"),
    )
    return by_nm.merge(by_root, on="product_root", how="left")


def build_campaign_base(ads: pd.DataFrame, campaigns: pd.DataFrame, orders: pd.DataFrame, bid_history: pd.DataFrame, windows: Dict[str, pd.Timestamp]) -> pd.DataFrame:
    current = campaign_window_metrics(ads, windows["current_start"], windows["current_end"], "cur")
    base = campaign_window_metrics(ads, windows["base_start"], windows["base_end"], "base")
    p14 = campaign_window_metrics(ads, windows["pause_start"], windows["pause_end"], "14d")
    recent7_start = windows["as_of"] - pd.Timedelta(days=7)
    recent7_end = windows["as_of"] - pd.Timedelta(days=1)
    recent7 = campaign_window_metrics(ads, recent7_start, recent7_end, "7d")
    yesterday = campaign_window_metrics(ads, recent7_end, recent7_end, "yday")

    if campaigns.empty:
        ids = pd.concat([current[["campaign_id"]], base[["campaign_id"]], p14[["campaign_id"]], recent7[["campaign_id"]], yesterday[["campaign_id"]]], ignore_index=True).drop_duplicates()
        df = ids
    else:
        df = campaigns.copy()

    for part in [current, base, p14, recent7, yesterday]:
        df = df.merge(part, on="campaign_id", how="left")

    # First seen by ads stats/history.
    first_seen_ads = ads.groupby("campaign_id", as_index=False).agg(first_seen=("day", "min")) if not ads.empty else pd.DataFrame(columns=["campaign_id", "first_seen"])
    df = df.merge(first_seen_ads, on="campaign_id", how="left")

    # Подтягиваем предмет из daily, если список кампаний не дал корректный subject_norm.
    if ads is not None and not ads.empty and "subject_norm" in ads.columns:
        ads_subjects = (
            ads[["campaign_id", "subject_norm"]]
            .dropna()
            .assign(subject_from_ads_daily=lambda d: d["subject_norm"].map(canon_subject))
            .groupby("campaign_id")["subject_from_ads_daily"]
            .agg(lambda s: s.value_counts().index[0] if len(s) else "")
            .reset_index()
        )
        if "subject_from_ads_daily" not in ads_subjects.columns:
            ads_subjects["subject_from_ads_daily"] = ""
        df = df.merge(ads_subjects[["campaign_id", "subject_from_ads_daily"]], on="campaign_id", how="left")
        if "subject_norm" not in df.columns:
            df["subject_norm"] = ""
        if "subject_from_ads_daily" not in df.columns:
            df["subject_from_ads_daily"] = ""
        bad_subject = ~df["subject_norm"].map(is_managed_subject_value)
        fill_mask = bad_subject & df["subject_from_ads_daily"].map(is_managed_subject_value)
        df.loc[fill_mask, "subject_norm"] = df.loc[fill_mask, "subject_from_ads_daily"]
        df = df.drop(columns=["subject_from_ads_daily"], errors="ignore")

    df = filter_managed_subjects(df, "campaign_base")

    if not bid_history.empty:
        last = bid_history.dropna(subset=["event_date"]).sort_values("event_date").groupby("campaign_id", as_index=False).tail(1)
        last = last[["campaign_id", "event_date", "old_bid_rub", "new_bid_rub", "direction", "reason_code"]].rename(columns={
            "event_date": "last_bid_change_date",
            "old_bid_rub": "last_old_bid_rub",
            "new_bid_rub": "last_new_bid_rub",
            "direction": "last_bid_direction",
            "reason_code": "last_bid_reason_code",
        })
        df = df.merge(last, on="campaign_id", how="left")
    else:
        df["last_bid_change_date"] = pd.NaT
        df["last_old_bid_rub"] = np.nan
        df["last_new_bid_rub"] = np.nan
        df["last_bid_direction"] = ""
        df["last_bid_reason_code"] = ""

    # Avg prices.
    prices = avg_price_by_product(orders, windows["current_start"], windows["current_end"])
    if not prices.empty and "nm_id" in df.columns:
        df = df.merge(prices[["product_root", "nm_id", "avg_finished_price", "avg_finished_price_root"]], on=["product_root", "nm_id"], how="left")
        df["avg_finished_price"] = df["avg_finished_price"].fillna(df["avg_finished_price_root"])
    else:
        df["avg_finished_price"] = np.nan

    # Basic defaults and expected metric columns even when a window has no rows.
    metric_cols = []
    for suf in ["cur", "base", "14d", "7d", "yday"]:
        metric_cols += [
            f"impressions_{suf}", f"clicks_{suf}", f"orders_{suf}", f"spend_{suf}", f"order_sum_{suf}",
            f"ctr_pct_{suf}", f"cpc_{suf}", f"cpo_{suf}", f"clicks_per_order_{suf}",
            f"impressions_per_order_{suf}", f"drr_pct_{suf}",
        ]
    for col in metric_cols:
        if col not in df.columns:
            df[col] = np.nan
    for col in [c for c in df.columns if any(x in c for x in ["impressions_", "clicks_", "orders_", "spend_", "order_sum_"])]:
        df[col] = to_num(df[col]).fillna(0.0)

    # Derived statuses.
    df["days_since_first_seen"] = (windows["as_of"] - pd.to_datetime(df["first_seen"], errors="coerce")).dt.days
    df["is_new"] = df["days_since_first_seen"].fillna(9999) < NEW_NO_PAUSE_DAYS
    df["days_since_last_bid_change"] = (windows["as_of"] - pd.to_datetime(df["last_bid_change_date"], errors="coerce")).dt.days
    df["recent_bid_change"] = df["days_since_last_bid_change"].between(0, 1, inclusive="both")
    df["order_sum_drop_vs_base_pct"] = np.where(
        df["order_sum_base"].fillna(0) > 0,
        (df["order_sum_base"].fillna(0) - df["order_sum_cur"].fillna(0)) / df["order_sum_base"].fillna(0) * 100.0,
        np.nan,
    )
    return df


def bid_cap_target_drr_pct(subject: Any) -> float:
    """Возвращает целевой ДРР для расчёта максимальной ставки.

    По новой логике для кистей используем 9% и считаем от рекламной выручки
    order_sum_cur, а не от finishedPrice/цены покупателя. Для остальных
    предметов временно оставлен legacy-порог 16%, чтобы не менять их экономику
    без отдельного решения.
    """
    return BRUSH_BID_CAP_DRR_PCT if canon_subject(subject) == "Кисти косметические" else DRR_FORECAST_CAP_PCT


def hard_cap_step_for_placement(placement: Any) -> int:
    return SEARCH_STEP_RUB if str(placement).lower() == "search" else COMBINED_STEP_RUB


def compute_bid_caps(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    caps = []
    cap_ceilings = []
    cap_steps = []
    target_drrs = []
    economic_raws = []
    below_min_flags = []
    reasons = []
    forecast_cpo_next = []
    forecast_drr_next = []

    # category competition cap by subject+placement.
    comp = out.groupby(["subject_norm", "placement"], dropna=False).agg(category_max_bid=("real_bid_rub", "max"), category_median_bid=("real_bid_rub", "median")).reset_index()
    out = out.merge(comp, on=["subject_norm", "placement"], how="left")

    for _, r in out.iterrows():
        subject = r.get("subject_norm", "")
        target_drr_pct = bid_cap_target_drr_pct(subject)
        placement = str(r.get("placement", "") or "").lower()
        current_bid = r.get("real_bid_rub", np.nan)
        cat_cap = r.get("category_max_bid", np.nan)
        order_sum_cur = r.get("order_sum_cur", np.nan)
        orders_cur = r.get("orders_cur", np.nan)
        step = hard_cap_step_for_placement(placement)
        below_min = False
        economic_raw = np.nan

        if placement == "search":
            clicks_cur = r.get("clicks_cur", np.nan)
            if pd.notna(order_sum_cur) and float(order_sum_cur) > 0 and pd.notna(clicks_cur) and float(clicks_cur) > 0:
                # max CPC = рекламная выручка × целевой ДРР / клики.
                # Это тот же знаменатель, что используется в ДРР рекламы: spend / order_sum.
                economic_raw = float(order_sum_cur) * target_drr_pct / 100.0 / float(clicks_cur)
                max_bid = math.floor(economic_raw)
                if max_bid < SEARCH_MIN_BID_RUB:
                    below_min = True
                max_bid = max(max_bid, SEARCH_MIN_BID_RUB)
                reason = f"CAP_BY_{target_drr_pct:g}_DRR_AD_REVENUE_PER_CLICK"
            else:
                # Not enough stable revenue/click data: use category competition +2, but mark weak cap.
                max_bid = math.floor((cat_cap if pd.notna(cat_cap) else current_bid if pd.notna(current_bid) else SEARCH_MIN_BID_RUB) + 2)
                max_bid = max(max_bid, SEARCH_MIN_BID_RUB)
                reason = "CAP_BY_CATEGORY_MAX_PLUS_2_NO_STABLE_REVENUE_CLICKS"
            current_effective = int(round(current_bid)) if pd.notna(current_bid) else SEARCH_MIN_BID_RUB
            next_bid = current_effective + SEARCH_STEP_RUB
            cpo_next = (float(next_bid) * float(clicks_cur) / float(orders_cur)) if pd.notna(next_bid) and pd.notna(clicks_cur) and pd.notna(orders_cur) and float(orders_cur) > 0 else np.nan
            drr_next = (float(next_bid) * float(clicks_cur) / float(order_sum_cur) * 100.0) if pd.notna(next_bid) and pd.notna(clicks_cur) and pd.notna(order_sum_cur) and float(order_sum_cur) > 0 else np.nan
        else:
            impressions_cur = r.get("impressions_cur", np.nan)
            if pd.notna(order_sum_cur) and float(order_sum_cur) > 0 and pd.notna(impressions_cur) and float(impressions_cur) > 0:
                # max CPM = рекламная выручка × целевой ДРР × 1000 / показы.
                economic_raw = float(order_sum_cur) * target_drr_pct / 100.0 * 1000.0 / float(impressions_cur)
                below_min = economic_raw < COMBINED_MIN_BID_RUB
                max_bid = floor_to_combined_grid(economic_raw)
                reason = f"CAP_BY_{target_drr_pct:g}_DRR_AD_REVENUE_PER_1000_IMPRESSIONS"
                if below_min:
                    reason += "_BELOW_WB_MIN"
            else:
                cat = cat_cap if pd.notna(cat_cap) else current_bid if pd.notna(current_bid) else COMBINED_MIN_BID_RUB
                max_bid = floor_to_combined_grid(cat + 2 * COMBINED_STEP_RUB)
                reason = "CAP_BY_CATEGORY_MAX_PLUS_2_STEPS_NO_STABLE_REVENUE_IMPRESSIONS"
            current_effective = max(float(current_bid) if pd.notna(current_bid) else COMBINED_MIN_BID_RUB, COMBINED_MIN_BID_RUB)
            current_effective = ceil_to_combined_grid(current_effective)
            next_bid = current_effective + COMBINED_STEP_RUB
            cpo_next = (float(next_bid) * float(impressions_cur) / 1000.0 / float(orders_cur)) if pd.notna(next_bid) and pd.notna(impressions_cur) and pd.notna(orders_cur) and float(orders_cur) > 0 else np.nan
            drr_next = (float(next_bid) * float(impressions_cur) / 1000.0 / float(order_sum_cur) * 100.0) if pd.notna(next_bid) and pd.notna(impressions_cur) and pd.notna(order_sum_cur) and float(order_sum_cur) > 0 else np.nan

        allowed_ceiling = float(max_bid) + float(step) * HARD_CAP_EXTRA_STEPS_ALLOWED if pd.notna(max_bid) else np.nan
        caps.append(max_bid)
        cap_ceilings.append(allowed_ceiling)
        cap_steps.append(step)
        target_drrs.append(target_drr_pct)
        economic_raws.append(economic_raw)
        below_min_flags.append(bool(below_min))
        reasons.append(reason)
        forecast_cpo_next.append(cpo_next)
        forecast_drr_next.append(drr_next)

    out["target_drr_cap_pct"] = target_drrs
    out["economic_max_bid_raw"] = economic_raws
    out["max_allowed_bid_rub"] = caps
    out["max_allowed_ceiling_bid_rub"] = cap_ceilings
    out["hard_cap_step_rub"] = cap_steps
    out["below_target_drr_at_wb_min_bid"] = below_min_flags
    out["max_bid_reason"] = reasons
    out["forecast_cpo_next_step"] = forecast_cpo_next
    out["forecast_drr_next_step_pct"] = forecast_drr_next
    out["bid_at_calculated_max"] = np.where(
        out["placement"].astype(str).str.lower().eq("search"),
        pd.to_numeric(out["real_bid_rub"], errors="coerce") >= pd.to_numeric(out["max_allowed_bid_rub"], errors="coerce"),
        pd.to_numeric(out["real_bid_rub"], errors="coerce") >= pd.to_numeric(out["max_allowed_bid_rub"], errors="coerce"),
    )
    return out

def build_core_efficiency(keywords: pd.DataFrame, campaigns: pd.DataFrame, campaign_metrics: pd.DataFrame) -> pd.DataFrame:
    if keywords.empty:
        return pd.DataFrame()
    rows = []
    # Use product-level CPC proxy: median current CPC from search campaigns of same product.
    cpc_by_product = campaign_metrics[campaign_metrics["placement"].eq("search")].groupby("product_root", as_index=True)["cpc_cur"].median().to_dict()
    for _, r in keywords.iterrows():
        root = r.get("product_root", "")
        cpc = cpc_by_product.get(root, np.nan)
        clicks = float(r.get("clicks", 0) or 0)
        orders = float(r.get("orders", 0) or 0)
        clicks_per_order = clicks / orders if orders > 0 else np.inf if clicks > 0 else np.nan
        query_cpo = clicks_per_order * cpc if pd.notna(cpc) and np.isfinite(clicks_per_order) else np.inf if clicks > 0 and orders <= 0 else np.nan
        frequency = float(r.get("frequency", 0) or 0)
        click_share = clicks / frequency * 100 if frequency > 0 else np.nan
        # classification done after ranking
        rows.append({
            "product_root": root,
            "nm_id": r.get("nm_id", pd.NA),
            "supplier_article": r.get("supplier_article", ""),
            "subject_norm": r.get("subject_norm", ""),
            "query_text": r.get("query_text", ""),
            "keyword_group": r.get("keyword_group", ""),
            "frequency": frequency,
            "clicks": clicks,
            "orders": orders,
            "clicks_per_order": clicks_per_order,
            "proxy_cpc": cpc,
            "query_cpo": query_cpo,
            "median_position": r.get("median_position", np.nan),
            "visibility_pct": r.get("visibility_pct", np.nan),
            "our_click_share_pct": click_share,
            "orders_share": r.get("orders_share", np.nan),
            "orders_cum_share": r.get("orders_cum_share", np.nan),
        })
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    # Pick 1-3 flagships by orders, frequency and acceptable query CPO.
    out["flagship_score"] = (
        out["orders"].fillna(0) * 1000
        + np.log1p(out["frequency"].fillna(0)) * 50
        - out["clicks_per_order"].replace(np.inf, 999).fillna(999) * 5
        - out["median_position"].fillna(99)
    )
    out["query_role"] = "core_secondary"
    for root, idxs in out.groupby("product_root").groups.items():
        part = out.loc[list(idxs)].copy()
        part = part[part["orders"].fillna(0) > 0]
        top = part.sort_values("flagship_score", ascending=False).head(FLAGSHIP_MAX_PER_PRODUCT).index
        out.loc[top, "query_role"] = "flagship"
    out.loc[(out["orders"].fillna(0) <= 0) & (out["clicks"].fillna(0) > 0), "query_role"] = "bad_no_orders"
    out.loc[(out["clicks_per_order"].replace(np.inf, 999) >= 20) & out["orders"].fillna(0).gt(0), "query_role"] = "broad_expensive"
    return out


def summarize_core_by_product(core: pd.DataFrame) -> pd.DataFrame:
    if core.empty:
        return pd.DataFrame(columns=["product_root"])
    g = core.groupby("product_root", as_index=False).agg(
        flagship_queries=("query_text", lambda x: "; ".join(list(x.head(3)))),
        flagship_orders=("orders", "sum"),
        flagship_clicks=("clicks", "sum"),
        flagship_frequency=("frequency", "sum"),
        flagship_position=("median_position", "median"),
        flagship_visibility_pct=("visibility_pct", "mean"),
        flagship_cpo=("query_cpo", "median"),
        flagship_clicks_per_order=("clicks_per_order", "median"),
    )
    return g


def rank_blocks(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    # A block is product_root + placement. For brushes product_root normally 901 but individual article can still be handled by supplier_article if needed.
    out["block_key"] = out["product_root"].fillna("") + "|" + out["placement"].fillna("")
    out["cpo_cur_safe"] = out["cpo_cur"].replace([np.inf, -np.inf], np.nan)
    out["block_score"] = (
        out["orders_cur"].fillna(0) * 1000
        + out["clicks_cur"].fillna(0) * 5
        + out["ctr_pct_cur"].fillna(0) * 20
        - out["cpo_cur_safe"].fillna(999) * 2
        - out["drr_pct_cur"].fillna(999) * 5
    )
    out["block_rank"] = out.groupby("block_key")["block_score"].rank(method="first", ascending=False)
    out["active_in_block"] = out.groupby("block_key")["is_active"].transform("sum") if "is_active" in out.columns else 1
    out["is_block_leader"] = out["block_rank"].eq(1)
    out["is_block_top3"] = out["block_rank"].le(3)
    return out


# -------------------------
# Decisions
# -------------------------

def decide_all(campaigns: pd.DataFrame, core: pd.DataFrame, pause_history: pd.DataFrame, windows: Dict[str, pd.Timestamp]) -> pd.DataFrame:
    df = filter_managed_subjects(campaigns.copy(), "decide_all_input")
    df = filter_excluded_articles(df, "decide_all_excluded_articles")
    if df.empty:
        return pd.DataFrame(columns=["campaign_id", "action", "reason_code", "reason_text"])
    core_summary = summarize_core_by_product(core[core["query_role"].eq("flagship")]) if not core.empty else pd.DataFrame()
    if not core_summary.empty:
        df = df.merge(core_summary, on="product_root", how="left")
    else:
        for c in ["flagship_queries", "flagship_orders", "flagship_clicks", "flagship_frequency", "flagship_position", "flagship_visibility_pct", "flagship_cpo", "flagship_clicks_per_order"]:
            df[c] = np.nan

    df = rank_blocks(df)
    df = select_ramp_slots(df)

    decisions = []
    for _, r in df.iterrows():
        decisions.append(decide_campaign(r, pause_history, windows))
    res = pd.DataFrame(decisions)
    return df.merge(res, on="campaign_id", how="left")


def assign_flagships(df: pd.DataFrame) -> pd.DataFrame:
    """Выбирает 2-3 флагманские РК внутри товарной группы по заказам.

    Флагман — РК, которая даёт основные заказы товарной группы.
    Его не выключаем как обычную low-volume РК; он может продолжать работать
    и участвовать в разгоне.
    """
    out = df.copy()
    out["is_flagship_campaign"] = False
    out["flagship_rank"] = np.nan
    out["flagship_orders_share_pct"] = np.nan
    for root, part in out.groupby("product_root", dropna=False):
        if str(root or "") == "":
            continue
        active_part = part[part.get("is_active", pd.Series(True, index=part.index)).fillna(True)].copy()
        active_part = active_part[active_part["orders_cur"].fillna(0) > 0].copy()
        if active_part.empty:
            continue
        total_orders = float(active_part["orders_cur"].fillna(0).sum())
        active_part = active_part.sort_values(["orders_cur", "drr_pct_cur", "impressions_7d"], ascending=[False, True, False])
        selected = []
        cum_orders = 0.0
        for idx, rr in active_part.iterrows():
            if len(selected) >= FLAGSHIP_MAX_PER_PRODUCT:
                break
            selected.append(idx)
            cum_orders += float(rr.get("orders_cur", 0) or 0)
            if len(selected) >= 2 and total_orders > 0 and (cum_orders / total_orders * 100.0) >= 70.0:
                break
        for rank, idx in enumerate(selected, start=1):
            out.loc[idx, "is_flagship_campaign"] = True
            out.loc[idx, "flagship_rank"] = rank
            if total_orders > 0:
                out.loc[idx, "flagship_orders_share_pct"] = float(out.loc[idx, "orders_cur"] or 0) / total_orders * 100.0
    return out


def select_ramp_slots(df: pd.DataFrame) -> pd.DataFrame:
    out = assign_flagships(df.copy())
    out["ramp_slot_selected"] = False
    out["ramp_queue_rank"] = np.nan
    out["ramp_status"] = ""
    # Очередь разгона: сначала те, кто ближе всего к 5000 показов за неделю.
    # Отдельно по товарной группе и типу размещения.
    for (root, placement), part in out.groupby(["product_root", "placement"], dropna=False):
        if str(root or "") == "":
            continue
        max_slots = RAMP_MAX_ACTIVE_CPC_PER_BLOCK if placement == "search" else RAMP_MAX_ACTIVE_CPM_PER_BLOCK
        cand = part.copy()
        cand = cand[cand.get("is_active", pd.Series(True, index=cand.index)).fillna(True)]
        cand = cand[cand["impressions_7d"].fillna(cand["impressions_cur"]).fillna(0) < RAMP_TARGET_IMPRESSIONS]
        if cand.empty:
            continue
        cand["ramp_sort_impressions"] = cand["impressions_7d"].fillna(cand["impressions_cur"]).fillna(0)
        cand = cand.sort_values(
            ["is_flagship_campaign", "ramp_sort_impressions", "orders_cur", "ctr_pct_cur"],
            ascending=[False, False, False, False],
        )
        out.loc[cand.index, "ramp_queue_rank"] = list(range(1, len(cand) + 1))
        out.loc[cand.index, "ramp_status"] = "в очереди"
        selected = cand.head(max_slots).index
        out.loc[selected, "ramp_slot_selected"] = True
        out.loc[selected, "ramp_status"] = "в разгоне"
    return out

def _bid_grid_values(placement: Any, current_bid: Any) -> Tuple[int, int, int, int, int]:
    placement_s = str(placement or "").strip().lower()
    if placement_s == "search":
        step = SEARCH_STEP_RUB
        min_bid = SEARCH_MIN_BID_RUB
        bid_effective = int(round(float(current_bid))) if pd.notna(current_bid) else min_bid
        bid_effective = max(bid_effective, min_bid)
        next_up = bid_effective + step
        next_down = max(min_bid, bid_effective - step)
        new_abs_max = NEW_SEARCH_MAX_BID_RUB
    else:
        step = COMBINED_STEP_RUB
        min_bid = COMBINED_MIN_BID_RUB
        bid_effective = max(float(current_bid) if pd.notna(current_bid) else min_bid, min_bid)
        bid_effective = int(ceil_to_combined_grid(bid_effective))
        next_up = bid_effective + step
        next_down = max(min_bid, bid_effective - step)
        new_abs_max = NEW_COMBINED_MAX_BID_RUB
    return int(step), int(min_bid), int(bid_effective), int(next_up), int(next_down), int(new_abs_max)


def _at_min_bid(placement: Any, bid_effective: Any) -> bool:
    _, min_bid, bid_effective, _, _, _ = _bid_grid_values(placement, bid_effective)
    return float(bid_effective) <= float(min_bid)


def _min_bid_matured(r: pd.Series, min_bid: int, windows: Dict[str, pd.Timestamp]) -> bool:
    """Минимальная ставка считается зрелой, если последнее снижение до минимума было >=7 дней назад."""
    last_bid = r.get("last_new_bid_rub", np.nan)
    last_dt = pd.to_datetime(r.get("last_bid_change_date", pd.NaT), errors="coerce")
    if pd.isna(last_bid) or pd.isna(last_dt):
        return False
    try:
        if float(last_bid) > float(min_bid):
            return False
    except Exception:
        return False
    return (windows["as_of"] - last_dt).days >= MATURE_WINDOW_DAYS


def decide_campaign(r: pd.Series, pause_history: pd.DataFrame, windows: Dict[str, pd.Timestamp]) -> Dict[str, Any]:
    cid = int(r["campaign_id"])
    placement = r.get("placement", "")
    active = bool(r.get("is_active", True))
    current_bid = r.get("real_bid_rub", np.nan)
    max_bid = r.get("max_allowed_bid_rub", np.nan)
    max_ceiling = r.get("max_allowed_ceiling_bid_rub", np.nan)
    drr_cur = r.get("drr_pct_cur", np.nan)
    drr_7d = r.get("drr_pct_7d", drr_cur)
    impressions_cur = float(r.get("impressions_cur", 0.0) or 0.0)
    impressions_7d = float(r.get("impressions_7d", impressions_cur) or 0.0)
    impressions_yday = float(r.get("impressions_yday", 0.0) or 0.0)
    orders_cur = float(r.get("orders_cur", 0.0) or 0.0)
    orders_7d = float(r.get("orders_7d", orders_cur) or 0.0)
    new_status = bool(r.get("is_new", False))
    is_leader = bool(r.get("is_block_leader", False))
    is_flagship = bool(r.get("is_flagship_campaign", False))
    active_in_block = int(r.get("active_in_block", 1) or 1)
    ramp_slot = bool(r.get("ramp_slot_selected", False))
    recent_bid = bool(r.get("recent_bid_change", False))
    subject = canon_subject(r.get("subject_norm", ""))

    if subject not in MANAGED_SUBJECTS_CANON:
        safe_bid = current_bid if pd.notna(current_bid) else np.nan
        return decision(cid, "hold", safe_bid, "OUT_OF_SCOPE_SUBJECT_HOLD", f"Предмет вне контура управления: {subject}. API запрещён.")

    article_norm = _normalize_article_for_experiment(r.get("supplier_article", ""))
    if article_norm in EXCLUDED_ARTICLES_FROM_AUTOMATION:
        safe_bid = current_bid if pd.notna(current_bid) else np.nan
        return decision(cid, "hold", safe_bid, "EXCLUDED_ARTICLE_HOLD", f"Артикул {article_norm} исключён из алгоритма управления: raise/lower/pause/start/API запрещены.")

    step, min_bid, bid_effective, next_up, next_down, new_abs_max = _bid_grid_values(placement, current_bid)

    # Paused campaigns: снятие с паузы только по причине/очереди/дозреванию.
    last_pause_status = get_last_pause_status(pause_history, cid)
    if not active:
        start_decision = decide_start_candidate(r, last_pause_status, windows)
        return {"campaign_id": cid, **start_decision}

    # NEW: не применяем экономический hard-cap. Только защитный абсолютный потолок 14/250.
    if new_status:
        if impressions_7d >= RAMP_TARGET_IMPRESSIONS:
            if orders_7d <= 0:
                if subject in PAUSE_ALLOWED_SUBJECTS_CANON:
                    return decision(cid, "pause", bid_effective, "NEW_5000_IMPRESSIONS_ZERO_ORDERS_PAUSE_WAIT_MATURITY", "NEW набрала >=5000 показов и 0 заказов: сразу пауза на дозревание 7+3")
                return decision(cid, "hold", bid_effective, "BRUSH_NEW_5000_ZERO_ORDERS_TG_ONLY", "Кисти не паузим: NEW >=5000 показов и 0 заказов; проблема уйдёт в TG-уведомление")
            if pd.notna(drr_7d) and drr_7d <= DRR_FORECAST_CAP_PCT:
                return decision(cid, "hold", bid_effective, "NEW_5000_WAIT_MATURITY", f"NEW набрала >=5000 показов; ставку не разгоняем, ждём дозревание 7+3; ДРР7={drr_7d:.1f}%")
            return decision(cid, "hold", bid_effective, "NEW_5000_HIGH_DRR_WAIT_MATURITY", f"NEW набрала >=5000 показов; ДРР7={drr_7d if pd.notna(drr_7d) else 'н/д'}; пауза/возврат после дозревания")
        if impressions_yday < NEW_DAILY_IMPRESSIONS_LOW:
            if next_up <= new_abs_max:
                return decision(cid, "raise", next_up, "NEW_DAILY_IMPRESSIONS_LOW_RAISE", f"NEW<14д: вчера показов {impressions_yday:.0f}<700; цель 700-1000/день и 5000/неделю; ставка {bid_effective}->{next_up}; защитный потолок NEW={new_abs_max}")
            return decision(cid, "hold", bid_effective, "NEW_RAISE_BLOCKED_ABSOLUTE_CAP", f"NEW<14д: показов мало, но следующий шаг {next_up} выше защитного потолка {new_abs_max}; hold")
        if impressions_yday > NEW_DAILY_IMPRESSIONS_HIGH:
            return decision(cid, "lower", next_down, "NEW_DAILY_IMPRESSIONS_HIGH_LOWER", f"NEW<14д: вчера показов {impressions_yday:.0f}>1000; показы пошли, ставку снижаем {bid_effective}->{next_down}")
        return decision(cid, "hold", bid_effective, "NEW_DAILY_IMPRESSIONS_OK_HOLD", f"NEW<14д: вчера показов {impressions_yday:.0f}, это 700-1000/день; ставку держим")

    # Для обычных РК действует hard cap: не выше рассчитанного max больше чем на 1 шаг.
    if active and pd.notna(max_ceiling) and float(bid_effective) > float(max_ceiling):
        forced_bid = int(round(max_ceiling)) if str(placement).lower() == "search" else int(ceil_to_combined_grid(max_ceiling))
        forced_bid = max(min_bid, forced_bid)
        return decision(cid, "lower", forced_bid, "CAP_OVERSHOOT_GT_ONE_STEP_FORCE_LOWER", f"Жёсткий cap: текущая ставка {bid_effective} выше рассчитанного максимума {max_bid} больше чем на 1 шаг; снижаем до {forced_bid}")

    # Если 5000+ показов и 0 заказов — сразу пауза на дозревание, кроме кистей.
    if impressions_7d >= RAMP_TARGET_IMPRESSIONS and orders_7d <= 0:
        if subject in PAUSE_ALLOWED_SUBJECTS_CANON:
            return decision(cid, "pause", bid_effective, "RAMP_5000_IMPRESSIONS_ZERO_ORDERS_PAUSE_WAIT_MATURITY", "РК набрала >=5000 показов и 0 заказов: не спасаем снижением, ставим на паузу на дозревание 7+3")
        return decision(cid, "hold", bid_effective, "BRUSH_5000_ZERO_ORDERS_TG_ONLY", "Кисти не паузим: >=5000 показов и 0 заказов; проблема уйдёт в TG-уведомление")

    # Low-volume: не делаем зрелый вывод без 5000 показов.
    if impressions_7d < RAMP_TARGET_IMPRESSIONS:
        if pd.notna(drr_7d) and drr_7d < DRR_RAISE_GATE_PCT:
            if recent_bid:
                return decision(cid, "hold", bid_effective, "RAMP_WAIT_AFTER_RECENT_BID_CHANGE", f"<5000 показов и ДРР7={drr_7d:.1f}%<10%, но ставка недавно менялась; ждём следующий день")
            if can_raise(next_up, max_ceiling):
                return decision(cid, "raise", next_up, "RAMP_LT5000_DRR_LT10_RAISE", f"Показов за неделю {impressions_7d:.0f}<5000 и ДРР7={drr_7d:.1f}%<10%; добираем до 5000: {bid_effective}->{next_up}")
            return decision(cid, "hold", bid_effective, "RAMP_RAISE_BLOCKED_BY_CAP", f"<5000 и ДРР<10, но следующий шаг {next_up} выше допустимого потолка {max_ceiling}")
        if pd.notna(drr_7d) and drr_7d > DRR_RAISE_GATE_PCT:
            if is_flagship:
                return decision(cid, "hold", bid_effective, "FLAGSHIP_LT5000_HIGH_DRR_KEEP_WORKING", f"Флагман: <5000 показов и ДРР7={drr_7d:.1f}%>10; не ставим на паузу как обычную РК, продолжаем наблюдение/разгон по очереди")
            if subject in PAUSE_ALLOWED_SUBJECTS_CANON and active_in_block > 1:
                return decision(cid, "pause", bid_effective, "PAUSE_WAIT_RAMP_QUEUE_LT5000_DRR_GT10", f"<5000 показов и ДРР7={drr_7d:.1f}%>10: пауза, ждёт очередь на разгон")
            return decision(cid, "hold", bid_effective, "BRUSH_OR_LAST_LT5000_DRR_GT10_HOLD", f"<5000 и ДРР>10, но кисть/единственная РК: не паузим")
        return decision(cid, "hold", bid_effective, "RAMP_LT5000_NOT_ENOUGH_DRR_HOLD", f"Показов {impressions_7d:.0f}<5000, стабильного ДРР нет; ждём/очередь")

    # 5000+ и есть заказы: работаем по ДРР.
    if pd.notna(drr_7d) and drr_7d <= DRR_RAISE_GATE_PCT:
        if recent_bid:
            return decision(cid, "hold", bid_effective, "WAIT_AFTER_RECENT_BID_CHANGE", f"ДРР7={drr_7d:.1f}%<=10%, но ставка менялась недавно; ждём")
        if can_raise(next_up, max_ceiling):
            return decision(cid, "raise", next_up, "DRR_LT10_GE5000_RAISE", f">=5000 показов и ДРР7={drr_7d:.1f}%<10%; повышаем ставку на 1 шаг: {bid_effective}->{next_up}")
        return decision(cid, "hold", bid_effective, "RAISE_BLOCKED_BY_FORECAST_DRR_CAP", f"ДРР7={drr_7d:.1f}%<=10%, но следующий шаг {next_up} выше потолка {max_ceiling}")

    if pd.notna(drr_7d) and drr_7d > DRR_RAISE_GATE_PCT:
        if bid_effective > min_bid and not recent_bid:
            return decision(cid, "lower", next_down, "DRR_GT10_GE5000_LOWER", f">=5000 показов и ДРР7={drr_7d:.1f}%>10%; снижаем ставку {bid_effective}->{next_down}")
        if bid_effective <= min_bid:
            if _min_bid_matured(r, min_bid, windows) and subject in PAUSE_ALLOWED_SUBJECTS_CANON and not is_flagship and active_in_block > 1:
                return decision(cid, "pause", bid_effective, "PAUSE_MIN_BID_MATURED_DRR_GT10", f"Ставка минимальная полное зрелое окно, ДРР7={drr_7d:.1f}%>10; пауза")
            return decision(cid, "hold", bid_effective, "MIN_BID_WAIT_FULL_MATURE_WINDOW", f"ДРР7={drr_7d:.1f}%>10, ставка минимальная; ждём полное зрелое окно на минимуме")
        return decision(cid, "hold", bid_effective, "DRR_GT10_RECENT_CHANGE_HOLD", f"ДРР7={drr_7d:.1f}%>10, но ставка недавно менялась; ждём")

    return decision(cid, "hold", bid_effective, "NO_STABLE_DRR_HOLD", "Нет стабильного ДРР; без изменений")

def can_raise(next_bid: float, allowed_ceiling: Any) -> bool:
    """Разрешает повышение только если следующий шаг не выше max_bid + 1 шаг."""
    if pd.isna(next_bid):
        return False
    if pd.isna(allowed_ceiling):
        return False
    return float(next_bid) <= float(allowed_ceiling)


def get_last_pause_status(pause_history: pd.DataFrame, campaign_id: int) -> Dict[str, Any]:
    if pause_history.empty:
        return {}
    p = pause_history[pause_history["campaign_id"].eq(campaign_id)].copy()
    if p.empty:
        return {}
    p = p.sort_values("pause_date")
    last = p.iloc[-1]
    return last.to_dict()


def decide_start_candidate(r: pd.Series, last_pause: Dict[str, Any], windows: Dict[str, pd.Timestamp]) -> Dict[str, Any]:
    cid = int(r["campaign_id"])
    placement = r.get("placement", "")
    min_bid = SEARCH_MIN_BID_RUB if placement == "search" else COMBINED_MIN_BID_RUB
    drr_14 = r.get("drr_pct_14d", np.nan)
    orders_14 = float(r.get("orders_14d", 0.0) or 0.0)
    ramp_slot = bool(r.get("ramp_slot_selected", False))
    pause_reason = str(last_pause.get("reason_code", "") or "")
    pause_date = pd.to_datetime(last_pause.get("pause_date", pd.NaT), errors="coerce")
    days_paused = (windows["as_of"] - pause_date).days if pd.notna(pause_date) else np.nan

    if pause_reason == "PAUSE_WAIT_RAMP_QUEUE_LT5000_DRR_GT10" and ramp_slot:
        return decision(cid, "start", min_bid, "START_RAMP_QUEUE_TURN", f"Очередь разгона подошла; снимаем с паузы по минимальной ставке {min_bid}")

    if pd.notna(days_paused) and days_paused < (MATURE_WINDOW_DAYS + DATA_LAG_DAYS):
        return decision(cid, "hold_paused", min_bid, "KEEP_PAUSED_WAIT_7D_PLUS_3D", f"Пауза дозревания: прошло {days_paused:.0f} дней, нужно 7+3")

    if pd.notna(drr_14) and drr_14 <= DRR_FORECAST_CAP_PCT and orders_14 > 0:
        return decision(cid, "start", min_bid, "START_AFTER_DRR_RECOVERY", f"После дозревания: заказы={orders_14:.0f}, ДРР14={drr_14:.1f}%<=16%; запуск по минимальной ставке {min_bid}")

    return decision(cid, "hold_paused", min_bid, "KEEP_PAUSED_BAD_RAMP_RESULT", f"После дозревания возврат не подтверждён; orders14={orders_14:.0f}, drr14={drr_14 if pd.notna(drr_14) else 'н/д'}")

def decision(campaign_id: int, action: str, new_bid: Any, reason_code: str, reason_text: str) -> Dict[str, Any]:
    return {
        "campaign_id": campaign_id,
        "action": action,
        "new_bid_rub": new_bid,
        "reason_code": reason_code,
        "reason_text": reason_text,
    }


def build_payload_preview(decisions: pd.DataFrame) -> pd.DataFrame:
    if decisions.empty:
        return pd.DataFrame()
    rows = []
    for _, r in decisions.iterrows():
        action = r.get("action", "hold")
        if action not in {"raise", "lower", "pause", "start"}:
            continue
        cid = r.get("campaign_id")
        if is_excluded_article_value(r.get("supplier_article", "")):
            continue
        if action in {"raise", "lower"}:
            rows.append({
                "api_action": "PATCH_BID",
                "campaign_id": cid,
                "new_bid_rub": r.get("new_bid_rub"),
                "placement": r.get("placement"),
                "reason_code": r.get("reason_code"),
            })
        elif action == "pause":
            rows.append({
                "api_action": "PAUSE",
                "campaign_id": cid,
                "new_bid_rub": np.nan,
                "placement": r.get("placement"),
                "reason_code": r.get("reason_code"),
            })
        elif action == "start":
            rows.append({
                "api_action": "START",
                "campaign_id": cid,
                "new_bid_rub": r.get("new_bid_rub"),
                "placement": r.get("placement"),
                "reason_code": r.get("reason_code"),
            })
    return pd.DataFrame(rows)


def block_reallocation_summary(decisions: pd.DataFrame) -> pd.DataFrame:
    if decisions is None or decisions.empty:
        return pd.DataFrame()
    work = decisions.copy()
    # Rollback-only mode contains only START rows from pause history, so some engine metrics
    # such as ramp_queue_rank / current spend can be absent. Create safe defaults for report sheets.
    for col, default in {
        "product_root": "",
        "placement": "",
        "campaign_id": "",
        "spend_cur": 0.0,
        "orders_cur": 0.0,
        "clicks_cur": 0.0,
        "action": "",
    }.items():
        if col not in work.columns:
            work[col] = default
    g = work.groupby(["product_root", "placement"], dropna=False).agg(
        active_campaigns=("campaign_id", "count"),
        spend_cur=("spend_cur", "sum"),
        orders_cur=("orders_cur", "sum"),
        clicks_cur=("clicks_cur", "sum"),
        pause_count=("action", lambda x: (x == "pause").sum()),
        raise_count=("action", lambda x: (x == "raise").sum()),
        lower_count=("action", lambda x: (x == "lower").sum()),
        start_count=("action", lambda x: (x == "start").sum()),
        leaders=("campaign_id", lambda x: ",".join(map(str, list(x.head(3))))),
    ).reset_index()
    g["cpo_cur"] = np.where(g["orders_cur"] > 0, g["spend_cur"] / g["orders_cur"], np.nan)
    g["method_comment"] = "Цель: pause_and_reallocate, расход не просто срезать, а перелить в лидеров; rollback-режим может содержать только START без engine-метрик."
    return g


def summary_table(decisions: pd.DataFrame, windows: Dict[str, pd.Timestamp]) -> pd.DataFrame:
    rows = [
        ["version", VERSION],
        ["as_of", windows["as_of"].date().isoformat()],
        ["current_window", f"{windows['current_start'].date()}..{windows['current_end'].date()}"],
        ["base_window", f"{windows['base_start'].date()}..{windows['base_end'].date()}"],
        ["pause_window", f"{windows['pause_start'].date()}..{windows['pause_end'].date()}"],
        ["drr_raise_gate_pct", DRR_RAISE_GATE_PCT],
        ["drr_pause_limit_pct", DRR_PAUSE_LIMIT_PCT],
        ["forecast_drr_cap_pct", DRR_FORECAST_CAP_PCT],
        ["ramp_target_impressions", RAMP_TARGET_IMPRESSIONS],
        ["abc_profitability_used", "NO"],
    ]
    if not decisions.empty and "action" in decisions.columns:
        for action, cnt in decisions["action"].value_counts(dropna=False).items():
            rows.append([f"action_{action}", int(cnt)])
    return pd.DataFrame(rows, columns=["metric", "value"])




def _safe_pause_start_subset(decisions: pd.DataFrame, pause_cols: list[str]) -> pd.DataFrame:
    work = decisions.copy() if isinstance(decisions, pd.DataFrame) else pd.DataFrame()
    if work.empty:
        return pd.DataFrame(columns=pause_cols)
    if "action" not in work.columns:
        for col in pause_cols:
            if col not in work.columns:
                work[col] = pd.Series(index=work.index, dtype="object")
        return work.iloc[0:0][pause_cols]
    mask = work["action"].astype(str).isin(["pause", "start", "hold_paused"])
    for col in pause_cols:
        if col not in work.columns:
            work[col] = pd.Series(index=work.index, dtype="object")
    return work.loc[mask, pause_cols]

def write_outputs(path: str, decisions: pd.DataFrame, core: pd.DataFrame, payload: pd.DataFrame, windows: Dict[str, pd.Timestamp]) -> None:
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_table(decisions, windows).to_excel(writer, sheet_name="Сводка", index=False)
        cols_dec = [
            "campaign_id", "supplier_article", "product_root", "subject_norm", "placement", "campaign_status", "is_new", "days_since_first_seen",
            "is_flagship_campaign", "flagship_rank", "flagship_orders_share_pct", "ramp_status",
            "real_bid_rub", "economic_max_bid_raw", "max_allowed_bid_rub", "max_allowed_ceiling_bid_rub", "hard_cap_step_rub", "target_drr_cap_pct", "below_target_drr_at_wb_min_bid", "new_bid_rub", "action", "reason_code", "reason_text",
            "impressions_cur", "clicks_cur", "orders_cur", "spend_cur", "order_sum_cur", "drr_pct_cur", "cpo_cur", "ctr_pct_cur",
            "impressions_7d", "clicks_7d", "orders_7d", "spend_7d", "order_sum_7d", "drr_pct_7d", "ctr_pct_7d",
            "impressions_yday", "clicks_yday", "orders_yday", "spend_yday", "order_sum_yday", "drr_pct_yday",
            "impressions_14d", "orders_14d", "spend_14d", "order_sum_14d", "drr_pct_14d", "order_sum_drop_vs_base_pct",
            "avg_finished_price", "forecast_cpo_next_step", "forecast_drr_next_step_pct", "max_bid_reason",
            "flagship_queries", "flagship_position", "flagship_cpo", "flagship_clicks_per_order", "block_rank", "ramp_queue_rank", "ramp_slot_selected",
            "last_bid_change_date", "days_since_last_bid_change", "last_bid_reason_code",
        ]
        existing = [c for c in cols_dec if c in decisions.columns]
        decisions[existing].to_excel(writer, sheet_name="Решения", index=False)
        if not core.empty:
            core.to_excel(writer, sheet_name="CORE_эффективность_по_окнам", index=False)
        else:
            pd.DataFrame({"note": ["CORE source not provided. Pass --previous-output with Ключевые_фразы_80 or add search query daily source."]}).to_excel(writer, sheet_name="CORE_эффективность_по_окнам", index=False)
        cap_cols = [c for c in ["campaign_id", "supplier_article", "placement", "real_bid_rub", "economic_max_bid_raw", "max_allowed_bid_rub", "max_allowed_ceiling_bid_rub", "hard_cap_step_rub", "target_drr_cap_pct", "below_target_drr_at_wb_min_bid", "bid_at_calculated_max", "order_sum_cur", "order_sum_base", "order_sum_drop_vs_base_pct", "forecast_cpo_next_step", "forecast_drr_next_step_pct", "avg_finished_price", "clicks_per_order_cur", "impressions_per_order_cur", "max_bid_reason"] if c in decisions.columns]
        decisions[cap_cols].to_excel(writer, sheet_name="Предельные_ставки", index=False)
        pause_cols = [c for c in ["campaign_id", "supplier_article", "product_root", "placement", "action", "reason_code", "reason_text", "impressions_7d", "orders_7d", "drr_pct_7d", "impressions_14d", "drr_pct_14d", "active_in_block", "is_block_leader", "is_flagship_campaign", "is_new"] if c in decisions.columns]
        _safe_pause_start_subset(decisions, pause_cols).to_excel(writer, sheet_name="Паузы_и_возвраты", index=False)
        ramp_cols = [c for c in ["campaign_id", "supplier_article", "product_root", "placement", "ramp_queue_rank", "ramp_slot_selected", "ramp_status", "is_flagship_campaign", "flagship_rank", "action", "reason_code", "impressions_7d", "impressions_yday", "ctr_pct_cur", "orders_cur", "drr_pct_7d", "real_bid_rub", "new_bid_rub", "max_allowed_bid_rub", "max_allowed_ceiling_bid_rub"] if c in decisions.columns]
        # In rollback-only mode there are no ramp columns. Sort only by columns that exist.
        sort_cols = [c for c in ["product_root", "placement", "ramp_queue_rank"] if c in decisions.columns]
        ramp_source = decisions.sort_values(sort_cols, na_position="last") if sort_cols else decisions.copy()
        ramp_source[ramp_cols].to_excel(writer, sheet_name="Разгон_очередь", index=False)
        block_reallocation_summary(decisions).to_excel(writer, sheet_name="Блоки_перелива", index=False)
        payload.to_excel(writer, sheet_name="API_payload_preview", index=False)



# -------------------------
# Local CLI + GitHub/S3 runner
# -------------------------

# These marker constants are intentionally present for the workflow guard.
EXPERIMENT_1_REASON_CODE = "EXPERIMENT_1_NIGHT_MIN_BID_MSK_1_5"
EXPERIMENT_2_REASON_CODE = "EXPERIMENT_2_NIGHT_COMBINED_DRR_GT_15_PAUSE"
EXPERIMENT_2_START_REASON_CODE = "EXPERIMENT_2_NIGHT_WINDOW_END_START"
EXPERIMENT_1_RESTORE_REASON_CODE = "EXPERIMENT_1_NIGHT_RESTORE_5MSK"
TECHNICAL_COMBINED_MIN_BID_80_FIX = "TECHNICAL_COMBINED_MIN_BID_80_FIX"
EXPERIMENT_1_ARTICLES = {"901/6", "901/2", "901/8"}

STORE_NAME = "TOPFACE"
SERVICE_PREFIX = "Служебные файлы/Ассистент WB/TOPFACE/"
ADS_MAIN_KEY = "Отчёты/Реклама/TOPFACE/Анализ рекламы.xlsx"
ADS_WEEKLY_PREFIX = "Отчёты/Реклама/TOPFACE/Недельные/"
ORDERS_WEEKLY_PREFIX = "Отчёты/Заказы/TOPFACE/Недельные/"
RUN_OUTPUT_KEY = SERVICE_PREFIX + "Итог_последнего_запуска.xlsx"
PREVIEW_OUTPUT_KEY = SERVICE_PREFIX + "Предпросмотр_последнего_запуска.xlsx"
SUMMARY_JSON_KEY = SERVICE_PREFIX + "Сводка_последнего_запуска.json"
API_LOG_KEY = SERVICE_PREFIX + "Лог_API.xlsx"
BID_HISTORY_KEY = SERVICE_PREFIX + "История_ставок.xlsx"
PAUSE_HISTORY_KEY = SERVICE_PREFIX + "История_пауз.xlsx"
BRUSH_TG_ALERT_KEY = SERVICE_PREFIX + "Проблемные_кисти_TG.xlsx"
BRUSH_TG_PDF_KEY = SERVICE_PREFIX + "Проблемные_кисти_WB_Ads.pdf"
BRUSH_TG_LOCK_PREFIX = SERVICE_PREFIX + "locks/brush_tg"
WB_ADVERT_BASE_URL = "https://advert-api.wildberries.ru"
WB_BIDS_ENDPOINT = "/api/advert/v1/bids"
WB_PAUSE_ENDPOINT = "/adv/v0/pause"
WB_START_ENDPOINT = "/adv/v0/start"
WB_BIDS_MIN_ENDPOINT = "/api/advert/v1/bids/min"

# Разовый откат ошибочных пауз, созданных слишком широкой v47-версией FIX46-runner.
# Откатываем только новые reason_code из v47/v50-контура, не трогаем старые штатные паузы и ночные эксперименты.
WRONG_FIX46_PAUSE_REASON_CODES = {
    "PAUSE_HIGH_DRR_14D_10000_REALLOCATE",
    "PAUSE_TO_RAMP_QUEUE_LOW_VOLUME",
}
ROLLBACK_WRONG_FIX46_PAUSE_REASON_CODE = "ROLLBACK_WRONG_FIX46_PAUSE_START"


@dataclass
class RunnerConfig:
    yc_access_key_id: str
    yc_secret_access_key: str
    yc_bucket_name: str
    wb_promo_key: str
    s3_endpoint_url: str = "https://storage.yandexcloud.net"
    wb_base_url: str = WB_ADVERT_BASE_URL


def _env_required(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise RuntimeError(f"Не задан обязательный env/secret: {name}")
    return value


def load_runner_config() -> RunnerConfig:
    return RunnerConfig(
        yc_access_key_id=_env_required("YC_ACCESS_KEY_ID"),
        yc_secret_access_key=_env_required("YC_SECRET_ACCESS_KEY"),
        yc_bucket_name=_env_required("YC_BUCKET_NAME"),
        wb_promo_key=_env_required("WB_PROMO_KEY_TOPFACE"),
    )


def make_s3_client(config: RunnerConfig):
    return boto3.client(
        "s3",
        endpoint_url=config.s3_endpoint_url,
        aws_access_key_id=config.yc_access_key_id,
        aws_secret_access_key=config.yc_secret_access_key,
    )


def s3_key_exists(s3_client, bucket: str, key: str) -> bool:
    try:
        s3_client.head_object(Bucket=bucket, Key=key)
        return True
    except ClientError as exc:
        code = str(exc.response.get("Error", {}).get("Code", ""))
        if code in {"404", "NoSuchKey", "NotFound"}:
            return False
        raise


def read_s3_bytes(s3_client, bucket: str, key: str) -> bytes:
    return s3_client.get_object(Bucket=bucket, Key=key)["Body"].read()


def upload_s3_bytes(s3_client, bucket: str, key: str, payload: bytes, content_type: Optional[str] = None) -> None:
    kwargs: Dict[str, Any] = {"Bucket": bucket, "Key": key, "Body": payload}
    if content_type:
        kwargs["ContentType"] = content_type
    s3_client.put_object(**kwargs)


def list_s3_keys(s3_client, bucket: str, prefix: str) -> List[str]:
    keys: List[str] = []
    token: Optional[str] = None
    while True:
        kwargs: Dict[str, Any] = {"Bucket": bucket, "Prefix": prefix}
        if token:
            kwargs["ContinuationToken"] = token
        resp = s3_client.list_objects_v2(**kwargs)
        for item in resp.get("Contents", []):
            key = item.get("Key", "")
            if key:
                keys.append(key)
        if not resp.get("IsTruncated"):
            break
        token = resp.get("NextContinuationToken")
    return keys


def latest_excel_keys(s3_client, bucket: str, prefix: str, limit: int = 4) -> List[str]:
    keys = [k for k in list_s3_keys(s3_client, bucket, prefix) if k.lower().endswith((".xlsx", ".xlsm")) and "~$" not in k]
    keys = sorted(set(keys), reverse=True)
    return keys[:limit]


def download_key_to_dir(s3_client, bucket: str, key: str, workdir: Path) -> str:
    safe_name = Path(key).name or (uuid.uuid4().hex + ".bin")
    path = workdir / safe_name
    payload = read_s3_bytes(s3_client, bucket, key)
    path.write_bytes(payload)
    return str(path)


def maybe_download_key_to_dir(s3_client, bucket: str, key: str, workdir: Path) -> Optional[str]:
    if not s3_key_exists(s3_client, bucket, key):
        return None
    return download_key_to_dir(s3_client, bucket, key, workdir)


def wb_headers(config: RunnerConfig) -> Dict[str, str]:
    return {"Authorization": config.wb_promo_key, "Content-Type": "application/json"}


def _clean_int(value: Any) -> Optional[int]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    if re.fullmatch(r"\d+", text):
        return int(text)
    try:
        val = int(float(text))
        return val if val > 0 else None
    except Exception:
        return None


def _normalize_article_for_experiment(value: Any) -> str:
    text = clean_article(value).upper().replace("_", "/").replace(" ", "")
    text = re.sub(r"^PT", "", text)
    m = re.search(r"(901)[\./\-/]?F?(\d+)", text)
    if m:
        return f"{m.group(1)}/{int(m.group(2))}"
    m = re.search(r"(901)\s*/\s*(\d+)", text)
    if m:
        return f"{m.group(1)}/{int(m.group(2))}"
    return text


def is_excluded_article_value(value: Any) -> bool:
    return _normalize_article_for_experiment(value) in EXCLUDED_ARTICLES_FROM_AUTOMATION


def filter_excluded_articles(df: pd.DataFrame, label: str = "") -> pd.DataFrame:
    """Полностью исключает выбранные артикулы из контура автоматического управления."""
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df.copy()
    if "supplier_article" not in df.columns:
        return df.copy()
    out = df.copy()
    article_norm = out["supplier_article"].map(_normalize_article_for_experiment)
    mask = ~article_norm.isin(EXCLUDED_ARTICLES_FROM_AUTOMATION)
    removed = int((~mask).sum())
    if removed:
        removed_articles = article_norm.loc[~mask].value_counts().to_dict()
        print(
            f"Диагностика исключения артикулов {label}: исключено {removed}; артикулы={removed_articles}",
            flush=True,
        )
    return out.loc[mask].copy()


def _api_log_row(method: str, endpoint: str, payload: Any, status: str, response_text: str, campaign_id: Any = "", nm_id: Any = "", placement: Any = "") -> Dict[str, Any]:
    return {
        "run_datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "method": method,
        "endpoint": endpoint,
        "campaign_id": campaign_id,
        "nm_id": nm_id,
        "placement": placement,
        "payload": json.dumps(payload, ensure_ascii=False) if payload not in (None, "") else "",
        "api_status": status,
        "response_text": str(response_text)[:2000],
    }


def _normal_bid_for_api(placement: Any, bid: Any) -> Optional[int]:
    if pd.isna(bid):
        return None
    placement_s = str(placement or "").strip().lower()
    value = float(bid)
    if placement_s == "combined":
        value = max(value, COMBINED_MIN_BID_RUB)
        value = ceil_to_combined_grid(value)
        return int(value) if value is not None else None
    return int(round(max(value, SEARCH_MIN_BID_RUB)))


def build_wb_bid_payload(row: pd.Series) -> Optional[Dict[str, Any]]:
    advert_id = _clean_int(row.get("campaign_id"))
    nm_id = _clean_int(row.get("nm_id"))
    placement = str(row.get("placement", "") or "").strip().lower()
    if placement not in {"search", "combined", "recommendations", "recommendation"}:
        placement = "combined" if "combined" in placement else "search"
    if placement == "recommendation":
        placement = "recommendations"
    new_bid = _normal_bid_for_api(placement, row.get("new_bid_rub"))
    if advert_id is None or nm_id is None or new_bid is None:
        return None
    return {
        "bids": [{
            "advert_id": int(advert_id),
            "nm_bids": [{
                "nm_id": int(nm_id),
                "bid_kopecks": int(new_bid * 100),
                "placement": placement,
            }],
        }]
    }



def _placement_for_min_endpoint(value: Any) -> str:
    placement = str(value or "").strip().lower()
    if placement == "recommendations":
        return "recommendation"
    if placement in {"search", "combined", "recommendation"}:
        return placement
    return "combined" if "combined" in placement else "search"


def _payment_type_for_min(value: Any) -> str:
    placement = str(value or "").strip().lower()
    return "cpm" if placement == "combined" else "cpc"


def fetch_wb_min_bid_for_row(row: pd.Series, config: RunnerConfig) -> Tuple[Optional[float], List[Dict[str, Any]]]:
    """Точечная проверка минимальной ставки WB перед PATCH ставки.

    Это защита от ошибок WB вида "bid value must be no less than ...".
    Если минимум WB выше расчётной ставки, PATCH не отправляем, а пишем SKIP в лог.
    """
    advert_id = _clean_int(row.get("campaign_id"))
    nm_id = _clean_int(row.get("nm_id"))
    placement = str(row.get("placement", "") or "").strip().lower()
    logs: List[Dict[str, Any]] = []
    if advert_id is None or nm_id is None:
        return None, logs
    payload = {
        "advert_id": int(advert_id),
        "nm_ids": [int(nm_id)],
        "payment_type": _payment_type_for_min(placement),
        "placement_types": [_placement_for_min_endpoint(placement)],
    }
    try:
        resp = requests.post(config.wb_base_url.rstrip("/") + WB_BIDS_MIN_ENDPOINT, headers=wb_headers(config), json=payload, timeout=60)
        logs.append(_api_log_row("POST", WB_BIDS_MIN_ENDPOINT, payload, str(resp.status_code), resp.text, advert_id, nm_id, placement))
        if not (200 <= resp.status_code < 300):
            return None, logs
        try:
            data = resp.json()
        except Exception:
            return None, logs
        values: List[float] = []
        for item in data.get("bids", []) or []:
            item_nm = _clean_int(item.get("nm_id"))
            if item_nm is not None and item_nm != int(nm_id):
                continue
            for bid_item in item.get("bids", []) or []:
                value_kopecks = pd.to_numeric(bid_item.get("value"), errors="coerce")
                if pd.notna(value_kopecks) and float(value_kopecks) > 0:
                    values.append(float(value_kopecks) / 100.0)
        if values:
            return max(values), logs
    except Exception as exc:
        logs.append(_api_log_row("POST", WB_BIDS_MIN_ENDPOINT, payload, "exception", repr(exc), advert_id, nm_id, placement))
    return None, logs


def build_wrong_fix46_pause_rollback_decisions(pause_history: pd.DataFrame) -> pd.DataFrame:
    """Формирует START только для ошибочных пауз, созданных v47/v50 FIX46-контуром.

    Логика безопасная:
    - берём только reason_code из WRONG_FIX46_PAUSE_REASON_CODES;
    - по campaign_id смотрим последнюю запись истории;
    - если последняя запись уже started — не трогаем;
    - категорийный guard намеренно не применяется: v47 могла ошибочно поставить на паузу товары вне 4 категорий.
    """
    columns = [
        "campaign_id", "nm_id", "supplier_article", "subject_norm", "placement", "action", "new_bid_rub",
        "reason_code", "reason_text", "product_root", "campaign_status", "is_active", "spend_cur", "orders_cur", "clicks_cur",
    ]
    if pause_history is None or pause_history.empty or "campaign_id" not in pause_history.columns:
        return pd.DataFrame(columns=columns)
    ph = pause_history.copy()
    for col in ["status", "reason_code", "pause_date", "nm_id", "placement", "supplier_article", "subject_norm"]:
        if col not in ph.columns:
            ph[col] = ""
    ph["campaign_id_int"] = ph["campaign_id"].map(_clean_int)
    ph = ph[ph["campaign_id_int"].notna()].copy()
    if ph.empty:
        return pd.DataFrame(columns=columns)
    ph["pause_dt"] = pd.to_datetime(ph["pause_date"], errors="coerce")
    ph["_row_order"] = range(len(ph))
    ph = ph.sort_values(["campaign_id_int", "pause_dt", "_row_order"], na_position="first")
    latest = ph.drop_duplicates("campaign_id_int", keep="last").copy()
    status = latest["status"].astype(str).str.strip().str.lower()
    reason = latest["reason_code"].astype(str).str.strip()
    candidates = latest[status.eq("paused") & reason.isin(WRONG_FIX46_PAUSE_REASON_CODES)].copy()
    rows: List[Dict[str, Any]] = []
    for _, r in candidates.iterrows():
        placement = str(r.get("placement", "") or "").strip().lower() or "combined"
        new_bid = COMBINED_MIN_BID_RUB if placement == "combined" else SEARCH_MIN_BID_RUB
        article = clean_article(r.get("supplier_article", ""))
        if is_excluded_article_value(article):
            continue
        rows.append({
            "campaign_id": int(r["campaign_id_int"]),
            "nm_id": r.get("nm_id", ""),
            "supplier_article": article,
            "subject_norm": r.get("subject_norm", ""),
            "placement": placement,
            "action": "start",
            "new_bid_rub": new_bid,
            "reason_code": ROLLBACK_WRONG_FIX46_PAUSE_REASON_CODE,
            "reason_text": f"Разовый откат ошибочной паузы v47/FIX46: последняя пауза reason={r.get('reason_code', '')}; запускаем обратно",
            "product_root": product_root(article),
            "campaign_status": "paused",
            "is_active": False,
            "spend_cur": 0.0,
            "orders_cur": 0.0,
            "clicks_cur": 0.0,
        })
    return pd.DataFrame(rows, columns=columns)

def _now_msk() -> datetime:
    return datetime.now(ZoneInfo("Europe/Moscow"))


def _is_exact_night_api_window(slot: str) -> bool:
    """Night API guard: 01:05 / 05:05 МСК с запасом на задержку GitHub."""
    slot = str(slot or "").strip().lower()
    now = _now_msk()
    minutes = now.hour * 60 + now.minute
    if slot == "start":
        return (1 * 60) <= minutes <= (1 * 60 + 35)
    if slot == "end":
        return (5 * 60) <= minutes <= (5 * 60 + 35)
    return False


def _is_main_api_window() -> bool:
    """Основной запуск ставок/пауз разрешён строго с 18:00 до 23:59 МСК."""
    now = _now_msk()
    return 18 <= now.hour <= 23


def _bid_api_allowed(night_experiment_only: bool, night_experiment_slot: str) -> Tuple[bool, str]:
    slot = str(night_experiment_slot or "").strip().lower()
    if bool(night_experiment_only):
        if slot not in {"start", "end"}:
            return False, "BID_API_BLOCKED_EMPTY_OR_UNKNOWN_NIGHT_SLOT"
        if not _is_exact_night_api_window(slot):
            return False, f"BID_API_BLOCKED_OUTSIDE_NIGHT_MSK_WINDOW_{_now_msk().strftime('%H:%M:%S')}"
        return True, "BID_API_ALLOWED_NIGHT_WINDOW"
    if _is_main_api_window():
        return True, "BID_API_ALLOWED_MAIN_18_00_23_59_MSK"
    return False, f"BID_API_BLOCKED_OUTSIDE_MAIN_WINDOW_{_now_msk().strftime('%H:%M:%S')}"

def _write_api_allowed(night_experiment_only: bool, night_experiment_slot: str) -> Tuple[bool, str]:
    return _bid_api_allowed(night_experiment_only, night_experiment_slot)


def apply_api_actions(decisions: pd.DataFrame, config: RunnerConfig, mode: str, dry_run: bool, apply_pause: bool = False, apply_start: bool = False, bypass_subject_guard: bool = False, night_experiment_only: bool = False, night_experiment_slot: str = "") -> Tuple[pd.DataFrame, pd.DataFrame]:
    if decisions is None or decisions.empty:
        return pd.DataFrame(), pd.DataFrame()
    logs: List[Dict[str, Any]] = []
    successful: List[Dict[str, Any]] = []
    for _, row in decisions.iterrows():
        action = str(row.get("action", "hold") or "hold").strip().lower()
        if action not in {"raise", "lower", "pause", "start"}:
            continue
        campaign_id = row.get("campaign_id", "")
        nm_id = row.get("nm_id", "")
        placement = row.get("placement", "")
        subject = canon_subject(row.get("subject_norm", ""))
        if (not bypass_subject_guard) and subject not in MANAGED_SUBJECTS_CANON:
            logs.append(_api_log_row("SKIP", "managed_subject_guard", {}, "blocked_out_of_scope", f"Предмет вне контура: {subject}", campaign_id, nm_id, placement))
            continue
        if is_excluded_article_value(row.get("supplier_article", "")):
            article_norm = _normalize_article_for_experiment(row.get("supplier_article", ""))
            logs.append(_api_log_row("SKIP", "excluded_article_guard", {}, "blocked_excluded_article", f"Артикул исключён из алгоритма: {article_norm}", campaign_id, nm_id, placement))
            continue
        if (not bypass_subject_guard) and action == "pause" and subject not in PAUSE_ALLOWED_SUBJECTS_CANON:
            logs.append(_api_log_row("SKIP", "pause_subject_guard", {}, "blocked_pause_subject", f"Автопауза запрещена для предмета: {subject}", campaign_id, nm_id, placement))
            continue
        if action == "pause" and not apply_pause:
            logs.append(_api_log_row("SKIP", WB_PAUSE_ENDPOINT, {"id": campaign_id}, "not_sent_apply_pause_false", "Пауза не отправлена: нужен --apply-pause", campaign_id, nm_id, placement))
            continue
        if action == "start" and not apply_start:
            logs.append(_api_log_row("SKIP", WB_START_ENDPOINT, {"id": campaign_id}, "not_sent_apply_start_false", "Запуск не отправлен: нужен --apply-start", campaign_id, nm_id, placement))
            continue
        if action in {"raise", "lower"}:
            allowed_bid_api, bid_api_guard_reason = _bid_api_allowed(night_experiment_only, night_experiment_slot)
            if not allowed_bid_api:
                logs.append(_api_log_row(
                    "SKIP",
                    WB_BIDS_ENDPOINT,
                    {},
                    "blocked_bid_api_time_guard",
                    f"Изменение ставки запрещено time guard: {bid_api_guard_reason}. Основной контур 18:00-23:59 МСК; ночной эксперимент 01:05/05:05 МСК.",
                    campaign_id,
                    nm_id,
                    placement,
                ))
                continue
            endpoint = WB_BIDS_ENDPOINT
            payload = build_wb_bid_payload(row)
            if payload is None:
                logs.append(_api_log_row("PATCH", endpoint, {}, "payload_error", "Не удалось собрать payload", campaign_id, nm_id, placement))
                continue
            min_bid, min_logs = fetch_wb_min_bid_for_row(row, config)
            logs.extend(min_logs)
            target_bid = _normal_bid_for_api(placement, row.get("new_bid_rub"))
            if min_bid is not None and target_bid is not None and float(target_bid) < float(min_bid):
                logs.append(_api_log_row("SKIP", WB_BIDS_ENDPOINT, payload, "skip_below_wb_min_bid", f"Расчётная ставка {target_bid} ₽ ниже минимальной WB {min_bid:.2f} ₽; PATCH не отправлен", campaign_id, nm_id, placement))
                continue
            if mode == "preview" or dry_run:
                status = "preview_no_call" if mode == "preview" else "dry_run_no_call"
                logs.append(_api_log_row("PATCH", endpoint, payload, status, "API-вызов не отправлялся", campaign_id, nm_id, placement))
                continue
            try:
                resp = requests.patch(config.wb_base_url.rstrip("/") + endpoint, headers=wb_headers(config), json=payload, timeout=60)
                logs.append(_api_log_row("PATCH", endpoint, payload, str(resp.status_code), resp.text, campaign_id, nm_id, placement))
                if 200 <= resp.status_code < 300:
                    item = row.to_dict()
                    item["api_status"] = str(resp.status_code)
                    successful.append(item)
            except Exception as exc:
                logs.append(_api_log_row("PATCH", endpoint, payload, "exception", repr(exc), campaign_id, nm_id, placement))
            time.sleep(0.2)
        elif action in {"pause", "start"}:
            allowed_write_api, write_api_guard_reason = _write_api_allowed(night_experiment_only, night_experiment_slot)
            if not allowed_write_api and mode != "preview" and not dry_run:
                logs.append(_api_log_row(
                    "SKIP",
                    "pause_start_time_guard",
                    {},
                    "blocked_write_api_time_guard",
                    f"PAUSE/START запрещён time guard: {write_api_guard_reason}. Основной контур 18:00-23:59 МСК; ночной эксперимент 01:05/05:05 МСК.",
                    campaign_id,
                    nm_id,
                    placement,
                ))
                continue
            endpoint = WB_PAUSE_ENDPOINT if action == "pause" else WB_START_ENDPOINT
            cid = _clean_int(campaign_id)
            if cid is None:
                logs.append(_api_log_row("GET", endpoint, {}, "payload_error", "Нет campaign_id", campaign_id, nm_id, placement))
                continue
            params = {"id": int(cid)}
            if mode == "preview" or dry_run:
                status = "preview_no_call" if mode == "preview" else "dry_run_no_call"
                logs.append(_api_log_row("GET", endpoint, params, status, "API-вызов не отправлялся", campaign_id, nm_id, placement))
                continue
            try:
                resp = requests.get(config.wb_base_url.rstrip("/") + endpoint, headers=wb_headers(config), params=params, timeout=60)
                logs.append(_api_log_row("GET", endpoint, params, str(resp.status_code), resp.text, campaign_id, nm_id, placement))
                if 200 <= resp.status_code < 300:
                    item = row.to_dict()
                    item["api_status"] = str(resp.status_code)
                    successful.append(item)
            except Exception as exc:
                logs.append(_api_log_row("GET", endpoint, params, "exception", repr(exc), campaign_id, nm_id, placement))
            time.sleep(0.2)
    return pd.DataFrame(successful), pd.DataFrame(logs)


def append_excel(existing_path: Optional[str], additions: pd.DataFrame, default_columns: Optional[List[str]] = None) -> pd.DataFrame:
    if existing_path and os.path.exists(existing_path):
        try:
            base = pd.read_excel(existing_path)
        except Exception:
            base = pd.DataFrame(columns=default_columns or [])
    else:
        base = pd.DataFrame(columns=default_columns or [])
    if additions is None or additions.empty:
        return base
    return pd.concat([base, additions], ignore_index=True, sort=False)


def record_successful_events(successful: pd.DataFrame, bid_history_path: Optional[str], pause_history_path: Optional[str]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    bid_rows: List[Dict[str, Any]] = []
    pause_rows: List[Dict[str, Any]] = []
    if successful is not None and not successful.empty:
        for _, row in successful.iterrows():
            action = str(row.get("action", "") or "").lower()
            if action in {"raise", "lower"}:
                bid_rows.append({
                    "event_id": str(uuid.uuid4()),
                    "run_datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "event_date": datetime.now().date().isoformat(),
                    "campaign_id": row.get("campaign_id", ""),
                    "nm_id": row.get("nm_id", ""),
                    "supplier_article": row.get("supplier_article", ""),
                    "subject_norm": row.get("subject_norm", ""),
                    "placement": row.get("placement", ""),
                    "old_bid_rub": row.get("real_bid_rub", ""),
                    "new_bid_rub": row.get("new_bid_rub", ""),
                    "direction": "raise" if action == "raise" else "lower",
                    "reason_code": row.get("reason_code", ""),
                    "spend_before": row.get("spend_cur", ""),
                    "revenue_before": row.get("order_sum_cur", ""),
                    "orders_before": row.get("orders_cur", ""),
                    "impressions_before": row.get("impressions_cur", ""),
                    "clicks_before": row.get("clicks_cur", ""),
                    "drr_before": row.get("drr_pct_cur", ""),
                    "postcheck_status": "pending",
                    "final_verdict": "",
                    "api_status": row.get("api_status", ""),
                })
            elif action in {"pause", "start"}:
                pause_rows.append({
                    "pause_event_id": str(uuid.uuid4()),
                    "pause_date": datetime.now().date().isoformat(),
                    "campaign_id": row.get("campaign_id", ""),
                    "nm_id": row.get("nm_id", ""),
                    "placement": row.get("placement", ""),
                    "supplier_article": row.get("supplier_article", ""),
                    "subject_norm": row.get("subject_norm", ""),
                    "reason_code": row.get("reason_code", ""),
                    "impressions_before_pause": row.get("impressions_14d", row.get("impressions_cur", "")),
                    "clicks_before_pause": row.get("clicks_14d", row.get("clicks_cur", "")),
                    "spend_before_pause": row.get("spend_14d", row.get("spend_cur", "")),
                    "revenue_before_pause": row.get("order_sum_14d", row.get("order_sum_cur", "")),
                    "orders_before_pause": row.get("orders_14d", row.get("orders_cur", "")),
                    "drr_before_pause": row.get("drr_pct_14d", row.get("drr_pct_cur", "")),
                    "status": "paused" if action == "pause" else "started",
                    "next_check_date": (datetime.now().date() + timedelta(days=POST_PAUSE_CHECK_DAYS)).isoformat() if action == "pause" else "",
                    "api_status": row.get("api_status", ""),
                })
    bid_history = append_excel(bid_history_path, pd.DataFrame(bid_rows))
    pause_history = append_excel(pause_history_path, pd.DataFrame(pause_rows))
    return bid_history, pause_history


def count_api_errors(api_log: pd.DataFrame) -> int:
    if api_log is None or api_log.empty or "api_status" not in api_log.columns:
        return 0
    statuses = api_log["api_status"].astype(str).str.strip()
    numeric = pd.to_numeric(statuses, errors="coerce")
    return int(((numeric >= 400) | statuses.str.contains("exception|error|payload_error", case=False, na=False)).sum())

def make_summary_json(mode: str, decisions: pd.DataFrame, successful: pd.DataFrame, api_log: pd.DataFrame, windows: Dict[str, pd.Timestamp], args: argparse.Namespace) -> Dict[str, Any]:
    actions = decisions["action"].value_counts(dropna=False).to_dict() if decisions is not None and not decisions.empty and "action" in decisions.columns else {}
    return {
        "Режим": mode,
        "Версия": SCRIPT_VERSION,
        "Дата формирования": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Всего рекомендаций": int(len(decisions)) if decisions is not None else 0,
        "Изменённых ставок": int(successful[successful.get("action", pd.Series(dtype=str)).astype(str).str.lower().isin(["raise", "lower"])].shape[0]) if successful is not None and not successful.empty else 0,
        "Кандидатов на паузу": int((decisions.get("action", pd.Series(dtype=str)).astype(str).str.lower() == "pause").sum()) if decisions is not None and not decisions.empty else 0,
        "Поставлено на паузу": int((successful.get("action", pd.Series(dtype=str)).astype(str).str.lower() == "pause").sum()) if successful is not None and not successful.empty else 0,
        "Кандидатов на запуск": int((decisions.get("action", pd.Series(dtype=str)).astype(str).str.lower() == "start").sum()) if decisions is not None and not decisions.empty else 0,
        "Запущено обратно": int((successful.get("action", pd.Series(dtype=str)).astype(str).str.lower() == "start").sum()) if successful is not None and not successful.empty else 0,
        "Действия": {str(k): int(v) for k, v in actions.items()},
        "Текущее окно с": windows["current_start"].date().isoformat(),
        "Текущее окно по": windows["current_end"].date().isoformat(),
        "База с": windows["base_start"].date().isoformat(),
        "База по": windows["base_end"].date().isoformat(),
        "Окно паузы с": windows["pause_start"].date().isoformat(),
        "Окно паузы по": windows["pause_end"].date().isoformat(),
        "ABC-рентабельность используется": "нет",
        "Режим только ночных экспериментов": "нет",
        "Ночной слот YAML": "",
        "API ставок: разрешённое окно": "основной 18:00-23:59 МСК; ночной эксперимент 01:05/05:05 МСК",
        "Текущее время МСК для API guard": _now_msk().strftime("%Y-%m-%d %H:%M:%S"),
        "Эксперимент 1: строк минимальной ночной ставки": int((decisions.get("reason_code", pd.Series(dtype=str)).astype(str) == EXPERIMENT_1_REASON_CODE).sum()) if decisions is not None and not decisions.empty else 0,
        "Разовый откат ошибочных пауз": "да" if getattr(args, "rollback_wrong_pauses_only", False) else "нет",
        "Разовый откат: кандидатов на start": int((decisions.get("reason_code", pd.Series(dtype=str)).astype(str) == ROLLBACK_WRONG_FIX46_PAUSE_REASON_CODE).sum()) if decisions is not None and not decisions.empty else 0,
        "Hard cap: принудительных снижений": int((decisions.get("reason_code", pd.Series(dtype=str)).astype(str) == "CAP_OVERSHOOT_GT_ONE_STEP_FORCE_LOWER").sum()) if decisions is not None and not decisions.empty else 0,
        "Исключённые артикулы": ", ".join(sorted(EXCLUDED_ARTICLES_FROM_AUTOMATION)),
        "Ошибок API": count_api_errors(api_log) if api_log is not None and not api_log.empty else 0,
    }


def _telegram_env() -> Tuple[str, str]:
    token = os.getenv("TELEGRAM_BOT_TOKEN") or os.getenv("TG_BOT_TOKEN") or os.getenv("WB_TG_BOT_TOKEN") or os.getenv("WB_TELEGRAM_BOT_TOKEN") or ""
    chat_id = os.getenv("TELEGRAM_CHAT_ID") or os.getenv("TG_CHAT_ID") or os.getenv("WB_TG_CHAT_ID") or os.getenv("WB_TELEGRAM_CHAT_ID") or ""
    return token.strip(), chat_id.strip()



def _brush_campaign_article_fallback() -> Dict[int, str]:
    """Fallback for old rows where campaign_id -> supplier_article was lost upstream.

    This is only a safety net for the current WB Ads report mapping issue. The normal path must
    still fill supplier_article before the Telegram/PDF report is built.
    """
    return {
        29651217: "901/6",
        29656882: "901/20",
        33303545: "901/16",
        33303562: "901/6",
        33303580: "901/10",
        33303619: "901/7",
        33303650: "901/20",
        33303804: "901/19",
        33937828: "901/3",
    }


def _restore_brush_article(row: pd.Series) -> str:
    article = str(row.get("supplier_article", "") or "").strip()
    if article and article.lower() not in {"nan", "none", "без артикула"}:
        return article
    try:
        campaign_id = int(float(row.get("campaign_id")))
    except Exception:
        return ""
    return _brush_campaign_article_fallback().get(campaign_id, "")


def _brush_campaign_type_label(value: Any) -> str:
    v = str(value or "").strip().lower()
    if v in {"search", "cpc", "поиск"}:
        return "ПОИСК"
    return "ПОЛКИ"


def _brush_article_sort_key(article: Any) -> Tuple[int, int, str]:
    s = str(article or "").strip()
    # 901/14, 901.14, PT901.F24, 901/F24
    m = re.search(r"(?:PT)?901[/.]?([A-Z]?)(\d+)", s, re.I)
    if "/" in s and s.split("/")[-1].isdigit():
        return (0, int(s.split("/")[-1]), s)
    if m:
        prefix = 1 if m.group(1) else 0
        return (prefix, int(m.group(2)), s)
    return (9, 9999, s)


def _short_brush_reason(reason: Any) -> str:
    s = str(reason or "").strip()
    replacements = {
        ">=5000 показов, ставка минимальная, ДРР >10%": "min ставка, ДРР >10%",
        "<5000 показов и ставка уже максимальная": "<5000 показов, max ставка",
        "достиг max ставки, сумма заказов упала >15%": "max ставка, заказы -15%",
        "ставка уже максимальная": "max ставка",
    }
    for old, new in replacements.items():
        s = s.replace(old, new)
    s = s.replace(";  ", "; ").strip("; ").strip()
    return s


def build_brush_problem_alerts(decisions: pd.DataFrame) -> pd.DataFrame:
    if decisions is None or decisions.empty:
        return pd.DataFrame()
    d = decisions.copy()
    if "subject_norm" not in d.columns:
        return pd.DataFrame()
    d["subject_norm"] = d["subject_norm"].map(canon_subject)
    d = d[d["subject_norm"].eq("Кисти косметические")].copy()
    if d.empty:
        return pd.DataFrame()

    numeric_cols = [
        "impressions_7d", "clicks_7d", "spend_7d", "order_sum_7d", "orders_7d",
        "drr_pct_7d", "ctr_pct_7d", "real_bid_rub", "max_allowed_bid_rub",
        "order_sum_cur", "order_sum_base", "order_sum_drop_vs_base_pct",
    ]
    for col in numeric_cols:
        if col not in d.columns:
            d[col] = np.nan
        d[col] = pd.to_numeric(d[col], errors="coerce")

    d["supplier_article"] = d.apply(_restore_brush_article, axis=1)
    # Final TG/PDF must not contain "без артикула". If mapping is still absent, keep it out of the report.
    d = d[d["supplier_article"].astype(str).str.strip().ne("")].copy()
    if d.empty:
        return pd.DataFrame()

    d["Тип кампании"] = d["placement"].map(_brush_campaign_type_label)
    d["is_min_bid_now"] = np.where(
        d["placement"].astype(str).str.lower().eq("search"),
        d["real_bid_rub"].fillna(999999) <= SEARCH_MIN_BID_RUB,
        d["real_bid_rub"].fillna(999999) <= COMBINED_MIN_BID_RUB,
    )
    d["is_max_bid_now"] = d["real_bid_rub"].fillna(-1) >= d["max_allowed_bid_rub"].fillna(10**9)

    cond_low_traffic_at_max = (d["impressions_7d"].fillna(0) < RAMP_TARGET_IMPRESSIONS) & d["is_max_bid_now"]
    cond_high_traffic_min_high_drr = (
        (d["impressions_7d"].fillna(0) >= RAMP_TARGET_IMPRESSIONS)
        & d["is_min_bid_now"]
        & (d["drr_pct_7d"].fillna(-1) > DRR_RAISE_GATE_PCT)
    )
    cond_max_bid_orders_drop = (
        d["is_max_bid_now"]
        & (d["order_sum_base"].fillna(0) > 0)
        & (d["order_sum_drop_vs_base_pct"].fillna(0) > 15.0)
    )

    d["tg_problem_reason"] = ""
    d.loc[cond_low_traffic_at_max, "tg_problem_reason"] = "<5000 показов, max ставка"
    d.loc[cond_high_traffic_min_high_drr, "tg_problem_reason"] = "min ставка, ДРР >10%"
    base_reason = d.loc[cond_max_bid_orders_drop, "tg_problem_reason"].astype(str)
    d.loc[cond_max_bid_orders_drop, "tg_problem_reason"] = np.where(
        base_reason.str.len() > 0,
        base_reason + "; max ставка, заказы -15%",
        "max ставка, заказы -15%",
    )

    out = d[d["tg_problem_reason"].astype(str).str.len() > 0].copy()
    if out.empty:
        return out

    out["tg_problem_reason"] = out["tg_problem_reason"].map(_short_brush_reason)
    out["_type_sort"] = np.where(out["Тип кампании"].eq("ПОИСК"), 0, 1)
    out["_article_sort"] = out["supplier_article"].map(_brush_article_sort_key)
    cols = [
        "supplier_article", "campaign_id", "placement", "Тип кампании",
        "impressions_7d", "clicks_7d", "ctr_pct_7d", "spend_7d", "order_sum_7d",
        "orders_7d", "drr_pct_7d", "real_bid_rub", "max_allowed_bid_rub",
        "order_sum_base", "order_sum_drop_vs_base_pct", "tg_problem_reason",
        "_type_sort", "_article_sort",
    ]
    cols = [c for c in cols if c in out.columns]
    out = out[cols].sort_values(["_type_sort", "_article_sort", "supplier_article", "campaign_id"]).drop(columns=["_type_sort", "_article_sort"], errors="ignore")
    return out


def _fmt_int_pdf(v: Any) -> str:
    try:
        x = float(v)
        if math.isfinite(x):
            return f"{int(round(x)):,}".replace(",", " ")
    except Exception:
        pass
    return "—"


def _fmt_money_pdf(v: Any) -> str:
    try:
        x = float(v)
        if math.isfinite(x):
            return f"{int(round(x)):,} ₽".replace(",", " ")
    except Exception:
        pass
    return "—"


def _fmt_pct_pdf(v: Any) -> str:
    try:
        x = float(v)
        if math.isfinite(x):
            return f"{x:.2f}%"
    except Exception:
        pass
    return "—"


def _fmt_bid_pdf(v: Any) -> str:
    try:
        x = float(v)
        if math.isfinite(x):
            return f"{int(round(x))} ₽"
    except Exception:
        pass
    return "—"


def build_brush_problem_pdf(alerts: pd.DataFrame, pdf_path: Path, period_label: str = "") -> Path:
    """Build PDF report exactly for Telegram: only the agreed table, no extra KPI blocks."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    font_regular = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    font_bold = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
    try:
        if "DejaVuSans" not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont("DejaVuSans", font_regular))
        if "DejaVuSans-Bold" not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", font_bold))
        regular_name, bold_name = "DejaVuSans", "DejaVuSans-Bold"
    except Exception:
        regular_name, bold_name = "Helvetica", "Helvetica-Bold"

    data = alerts.copy() if alerts is not None else pd.DataFrame()
    if not data.empty:
        data["Тип кампании"] = data.get("Тип кампании", data.get("placement", "")).map(_brush_campaign_type_label)
        data["_type_sort"] = np.where(data["Тип кампании"].eq("ПОИСК"), 0, 1)
        data["_article_sort"] = data["supplier_article"].map(_brush_article_sort_key)
        data = data.sort_values(["_type_sort", "_article_sort", "supplier_article", "campaign_id"]).drop(columns=["_type_sort", "_article_sort"], errors="ignore")

    page_w, page_h = landscape(A4)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Cell", fontName=regular_name, fontSize=6.7, leading=8.2, textColor=colors.HexColor("#222222"), alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="CellCenter", fontName=regular_name, fontSize=6.7, leading=8.2, textColor=colors.HexColor("#222222"), alignment=TA_CENTER))
    styles.add(ParagraphStyle(name="HeaderCell", fontName=bold_name, fontSize=6.25, leading=7.5, textColor=colors.HexColor("#333333"), alignment=TA_CENTER))
    styles.add(ParagraphStyle(name="Reason", fontName=regular_name, fontSize=6.2, leading=7.4, textColor=colors.HexColor("#222222"), alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="ReportTitle", fontName=bold_name, fontSize=15, leading=17, textColor=colors.HexColor("#111111"), alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="Subtitle", fontName=regular_name, fontSize=8, leading=10, textColor=colors.HexColor("#565656"), alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="Section", fontName=bold_name, fontSize=10.5, leading=13, textColor=colors.HexColor("#111111"), alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="Empty", fontName=regular_name, fontSize=8, leading=10, textColor=colors.HexColor("#565656"), alignment=TA_LEFT))

    def par(text: Any, style: str = "Cell") -> Paragraph:
        return Paragraph(str(text if text is not None else "—"), styles[style])

    def make_table(rows: pd.DataFrame) -> Table:
        header = [
            "Артикул", "Тип кампании", "Показы", "Клики", "CTR", "Расход",
            "Сумма заказов", "ДРР", "Ставка текущая", "Макс ставка рассчитанная", "Причина",
        ]
        table_data = [[par(h, "HeaderCell") for h in header]]
        for _, r in rows.iterrows():
            table_data.append([
                par(r.get("supplier_article"), "Cell"),
                par(r.get("Тип кампании"), "CellCenter"),
                par(_fmt_int_pdf(r.get("impressions_7d")), "CellCenter"),
                par(_fmt_int_pdf(r.get("clicks_7d")), "CellCenter"),
                par(_fmt_pct_pdf(r.get("ctr_pct_7d")), "CellCenter"),
                par(_fmt_money_pdf(r.get("spend_7d")), "CellCenter"),
                par(_fmt_money_pdf(r.get("order_sum_7d")), "CellCenter"),
                par(_fmt_pct_pdf(r.get("drr_pct_7d")), "CellCenter"),
                par(_fmt_bid_pdf(r.get("real_bid_rub")), "CellCenter"),
                par(_fmt_bid_pdf(r.get("max_allowed_bid_rub")), "CellCenter"),
                par(_short_brush_reason(r.get("tg_problem_reason")), "Reason"),
            ])
        col_widths = [43, 58, 48, 38, 38, 55, 70, 39, 56, 74, 188]
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), regular_name),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F3F5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#333333")),
            ("LINEABOVE", (0, 0), (-1, 0), 0.8, colors.HexColor("#D9DDE3")),
            ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.HexColor("#D9DDE3")),
            ("LINEBELOW", (0, 1), (-1, -1), 0.35, colors.HexColor("#E5E7EB")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FBFBFC")]),
        ]))
        return tbl

    def on_page(canvas, doc):
        canvas.saveState()
        canvas.setFont(regular_name, 8)
        canvas.setFillColor(colors.HexColor("#111111"))
        canvas.drawString(15 * mm, page_h - 12 * mm, "topface / WB Ads")
        canvas.setFont(regular_name, 7)
        canvas.setFillColor(colors.HexColor("#777777"))
        canvas.drawRightString(page_w - 15 * mm, page_h - 12 * mm, f"Страница {doc.page}")
        canvas.setStrokeColor(colors.HexColor("#E6E8EC"))
        canvas.setLineWidth(0.5)
        canvas.line(15 * mm, page_h - 15.5 * mm, page_w - 15 * mm, page_h - 15.5 * mm)
        canvas.restoreState()

    doc = SimpleDocTemplate(
        str(pdf_path), pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=21 * mm, bottomMargin=12 * mm,
    )

    subtitle = period_label or "проблемные кисти / сначала ПОИСК, затем ПОЛКИ"
    story: List[Any] = []
    groups = [("ПОИСК", data[data["Тип кампании"].eq("ПОИСК")])] if not data.empty else [("ПОИСК", pd.DataFrame())]
    groups.append(("ПОЛКИ", data[data["Тип кампании"].eq("ПОЛКИ")] if not data.empty else pd.DataFrame()))
    for idx, (section, rows) in enumerate(groups):
        if idx:
            story.append(PageBreak())
        story.append(par("Проблемные кисти WB Ads", "ReportTitle"))
        story.append(Spacer(1, 2.5 * mm))
        story.append(par(subtitle, "Subtitle"))
        story.append(Spacer(1, 5 * mm))
        story.append(par(section, "Section"))
        story.append(Spacer(1, 2.5 * mm))
        if rows.empty:
            story.append(par("Проблемных строк нет", "Empty"))
        else:
            story.append(make_table(rows))

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    return pdf_path


def maybe_send_brush_tg_alert(
    s3_client,
    bucket: str,
    decisions: pd.DataFrame,
    force: bool = False,
    schedule_only: bool = False,
    pdf_path: Optional[Path] = None,
    period_label: str = "",
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    alerts = build_brush_problem_alerts(decisions)
    now = _now_msk()
    lock_key = f"{BRUSH_TG_LOCK_PREFIX}/brush_tg_{now.date().isoformat()}.json"
    result = {
        "tg_attempted": False,
        "tg_sent": False,
        "tg_status": "not_requested",
        "tg_lock_key": lock_key,
        "tg_rows": int(len(alerts)),
        "tg_pdf_path": str(pdf_path) if pdf_path else "",
    }

    if pdf_path is not None:
        try:
            build_brush_problem_pdf(alerts, Path(pdf_path), period_label=period_label)
            result["tg_pdf_created"] = True
        except Exception as exc:
            result["tg_status"] = "pdf_build_exception"
            result["tg_response"] = repr(exc)[:500]
            result["tg_pdf_created"] = False
            return alerts, result

    if schedule_only:
        # Monday 19:05 MSK or later. After midnight it becomes Tuesday and is blocked by weekday guard.
        if now.weekday() != 0 or (now.hour * 60 + now.minute) < (19 * 60 + 5):
            result["tg_status"] = "blocked_by_monday_1905_guard"
            return alerts, result

    if not force:
        try:
            s3_client.head_object(Bucket=bucket, Key=lock_key)
            result["tg_status"] = "already_sent_today"
            return alerts, result
        except ClientError as exc:
            code = str(exc.response.get("Error", {}).get("Code", ""))
            if code not in {"404", "NoSuchKey", "NotFound"}:
                raise

    token, chat_id = _telegram_env()
    if not token or not chat_id:
        result["tg_status"] = "missing_telegram_env"
        return alerts, result
    if pdf_path is None or not Path(pdf_path).exists():
        result["tg_status"] = "missing_pdf_file"
        return alerts, result

    result["tg_attempted"] = True
    try:
        with open(pdf_path, "rb") as fh:
            resp = requests.post(
                f"https://api.telegram.org/bot{token}/sendDocument",
                data={"chat_id": chat_id},
                files={"document": (Path(pdf_path).name, fh, "application/pdf")},
                timeout=90,
            )
        result["tg_status"] = str(resp.status_code)
        if 200 <= resp.status_code < 300:
            result["tg_sent"] = True
            payload = {
                "sent_at_msk": now.strftime("%Y-%m-%d %H:%M:%S"),
                "rows": int(len(alerts)),
                "force": bool(force),
                "pdf_only": True,
            }
            s3_client.put_object(
                Bucket=bucket,
                Key=lock_key,
                Body=json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8"),
                ContentType="application/json",
            )
        else:
            result["tg_response"] = resp.text[:500]
    except Exception as exc:
        result["tg_status"] = "exception"
        result["tg_response"] = repr(exc)[:500]
    return alerts, result


def compute_engine(args: argparse.Namespace) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict[str, pd.Timestamp]]:
    as_of = pd.Timestamp(args.as_of).normalize() if args.as_of else pd.Timestamp(datetime.now().date())
    windows = date_windows(as_of)
    ads, campaigns = load_ads_daily(args.ads)
    orders = load_orders(args.orders)
    bid_history = load_bid_history(args.bid_history)
    pause_history = load_pause_history(args.pause_history)
    keywords = load_keywords_from_previous(args.previous_output)
    campaign_base = build_campaign_base(ads, campaigns, orders, bid_history, windows)
    campaign_base = filter_excluded_articles(campaign_base, "campaign_base_excluded_articles")
    campaign_base = compute_bid_caps(campaign_base)
    core = build_core_efficiency(keywords, campaigns, campaign_base)
    decisions = decide_all(campaign_base, core, pause_history, windows)
    payload = build_payload_preview(decisions)
    return decisions, core, payload, windows


def run_local(args: argparse.Namespace) -> int:
    decisions, core, payload, windows = compute_engine(args)
    write_outputs(args.out, decisions, core, payload, windows)
    if args.print_summary:
        print(summary_table(decisions, windows).to_string(index=False))
        print(f"Output: {args.out}")
        print(f"API actions preview: {len(payload)}")
    return 0


def _filter_night_decisions(decisions: pd.DataFrame, pause_history: pd.DataFrame, bid_history: pd.DataFrame, slot: str) -> pd.DataFrame:
    if decisions is None:
        return pd.DataFrame()
    if slot == "start":
        rows: List[Dict[str, Any]] = []
        if not decisions.empty:
            active = decisions[decisions.get("is_active", pd.Series([True] * len(decisions), index=decisions.index)).fillna(True).astype(bool)].copy()
            for _, r in active.iterrows():
                article_norm = _normalize_article_for_experiment(r.get("supplier_article", ""))
                if article_norm in EXCLUDED_ARTICLES_FROM_AUTOMATION:
                    continue
                placement = str(r.get("placement", "") or "").lower()
                if article_norm in EXPERIMENT_1_ARTICLES:
                    current = float(r.get("real_bid_rub", np.nan)) if pd.notna(r.get("real_bid_rub", np.nan)) else np.nan
                    target = SEARCH_MIN_BID_RUB if placement == "search" else COMBINED_MIN_BID_RUB
                    if pd.isna(current) or int(round(current)) != int(target):
                        item = r.to_dict()
                        item["action"] = "raise" if pd.isna(current) or current < target else "lower"
                        item["new_bid_rub"] = int(target)
                        item["reason_code"] = EXPERIMENT_1_REASON_CODE
                        item["reason_text"] = f"Ночной эксперимент 01:00-05:00 МСК: привести {article_norm} / {placement} к минимальной ставке {target} ₽"
                        rows.append(item)
                if placement == "combined" and float(r.get("drr_pct_14d", np.nan) if pd.notna(r.get("drr_pct_14d", np.nan)) else -1) > DRR_PAUSE_LIMIT_PCT:
                    item = r.to_dict()
                    item["action"] = "pause"
                    item["new_bid_rub"] = np.nan
                    item["reason_code"] = EXPERIMENT_2_REASON_CODE
                    item["reason_text"] = f"Ночной эксперимент: combined с ДРР14 {r.get('drr_pct_14d')}% > {DRR_PAUSE_LIMIT_PCT}% на паузу до 05:00 МСК"
                    rows.append(item)
        return pd.DataFrame(rows)
    if slot == "end":
        rows: List[Dict[str, Any]] = []
        today_msk = _now_msk().date().isoformat()

        # Experiment 1 restore: return bids changed at 01:00 back to their old values at 05:00.
        if bid_history is not None and not bid_history.empty and "reason_code" in bid_history.columns and "campaign_id" in bid_history.columns:
            bh = bid_history.copy()
            bh = bh[bh["reason_code"].astype(str).eq(EXPERIMENT_1_REASON_CODE)].copy()
            if not bh.empty:
                bh["event_date_str"] = pd.to_datetime(bh.get("event_date", pd.Series(pd.NaT, index=bh.index)), errors="coerce").dt.date.astype(str)
                bh = bh[bh["event_date_str"].eq(today_msk)].copy()
                bh["old_bid_rub"] = to_num(bh.get("old_bid_rub", pd.Series([np.nan] * len(bh))))
                bh["new_bid_rub"] = to_num(bh.get("new_bid_rub", pd.Series([np.nan] * len(bh))))
                bh = bh[bh["old_bid_rub"].notna() & bh["new_bid_rub"].notna()].copy()
                bh = bh[bh["old_bid_rub"].round(2).ne(bh["new_bid_rub"].round(2))].copy()
                if not bh.empty:
                    sort_col = "run_datetime" if "run_datetime" in bh.columns else "event_date"
                    bh = bh.sort_values(sort_col).drop_duplicates("campaign_id", keep="last")
                    for _, b in bh.iterrows():
                        cid = _clean_int(b.get("campaign_id"))
                        if cid is None:
                            continue
                        article_norm = _normalize_article_for_experiment(b.get("supplier_article", ""))
                        if article_norm in EXCLUDED_ARTICLES_FROM_AUTOMATION:
                            continue
                        old_bid = int(round(float(b.get("old_bid_rub"))))
                        current_bid = float(b.get("new_bid_rub"))
                        placement = str(b.get("placement", "") or "search").strip().lower()
                        item = {
                            "campaign_id": cid,
                            "nm_id": b.get("nm_id", ""),
                            "placement": placement,
                            "supplier_article": b.get("supplier_article", ""),
                            "subject_norm": canon_subject(b.get("subject_norm", "")),
                            "product_root": product_root(b.get("supplier_article", "")),
                            "real_bid_rub": current_bid,
                            "action": "raise" if old_bid > current_bid else "lower",
                            "new_bid_rub": old_bid,
                            "reason_code": EXPERIMENT_1_RESTORE_REASON_CODE,
                            "reason_text": "Ночной эксперимент: вернуть ставку, изменённую в 01:00 МСК, обратно в 05:00 МСК",
                        }
                        rows.append(item)

        # Experiment 2 restore: start campaigns paused at 01:00.
        if pause_history is not None and not pause_history.empty:
            ph = pause_history.copy()
            if "reason_code" in ph.columns and "campaign_id" in ph.columns:
                ph = ph[ph["reason_code"].astype(str).eq(EXPERIMENT_2_REASON_CODE)].copy()
                if not ph.empty:
                    ph["pause_dt"] = pd.to_datetime(ph.get("pause_date", pd.Series(pd.NaT, index=ph.index)), errors="coerce")
                    ph = ph[ph["pause_dt"].dt.date.astype(str).eq(today_msk)].copy()
                    ph = ph.sort_values("pause_dt").drop_duplicates("campaign_id", keep="last")
                    for _, p in ph.iterrows():
                        cid = _clean_int(p.get("campaign_id"))
                        if cid is None:
                            continue
                        base = decisions[decisions["campaign_id"].astype(str).eq(str(cid))].tail(1).to_dict("records") if decisions is not None and not decisions.empty and "campaign_id" in decisions.columns else []
                        item = base[0] if base else {"campaign_id": cid, "nm_id": p.get("nm_id", ""), "placement": p.get("placement", "combined"), "supplier_article": p.get("supplier_article", ""), "subject_norm": p.get("subject_norm", "")}
                        item["action"] = "start"
                        item["new_bid_rub"] = COMBINED_MIN_BID_RUB if str(item.get("placement", "")).lower() == "combined" else SEARCH_MIN_BID_RUB
                        item["reason_code"] = EXPERIMENT_2_START_REASON_CODE
                        item["reason_text"] = "Ночной эксперимент: вернуть кампанию после окна 01:00-05:00 МСК"
                        rows.append(item)
        return pd.DataFrame(rows)
    return decisions


def run_s3_legacy(args: argparse.Namespace) -> int:
    mode = args.command
    config = load_runner_config()
    s3 = make_s3_client(config)
    bucket = config.yc_bucket_name
    with tempfile.TemporaryDirectory(prefix="wb_ads_fix46_") as tmp:
        workdir = Path(tmp)
        ads_paths: List[str] = []
        if s3_key_exists(s3, bucket, ADS_MAIN_KEY):
            ads_paths.append(download_key_to_dir(s3, bucket, ADS_MAIN_KEY, workdir))
        else:
            for key in latest_excel_keys(s3, bucket, ADS_WEEKLY_PREFIX, limit=4):
                ads_paths.append(download_key_to_dir(s3, bucket, key, workdir))
        if not ads_paths:
            raise RuntimeError(f"Не найден рекламный отчёт: {ADS_MAIN_KEY} или {ADS_WEEKLY_PREFIX}")

        order_paths = [download_key_to_dir(s3, bucket, key, workdir) for key in latest_excel_keys(s3, bucket, ORDERS_WEEKLY_PREFIX, limit=4)]
        previous_output_path = maybe_download_key_to_dir(s3, bucket, RUN_OUTPUT_KEY, workdir)
        bid_history_path = maybe_download_key_to_dir(s3, bucket, BID_HISTORY_KEY, workdir)
        pause_history_path = maybe_download_key_to_dir(s3, bucket, PAUSE_HISTORY_KEY, workdir)

        local_out = workdir / ("Предпросмотр_последнего_запуска.xlsx" if mode == "preview" else "Итог_последнего_запуска.xlsx")
        ph_for_rollback = load_pause_history(pause_history_path)
        if getattr(args, "rollback_wrong_pauses_only", False):
            windows = date_windows(pd.Timestamp(datetime.now().date()))
            decisions = build_wrong_fix46_pause_rollback_decisions(ph_for_rollback)
            core = pd.DataFrame()
            payload = build_payload_preview(decisions)
            print(f"Разовый откат ошибочных пауз: кандидатов на START={len(decisions)}", flush=True)
        else:
            engine_args = argparse.Namespace(
                ads=";".join(ads_paths),
                orders=";".join(order_paths) if order_paths else None,
                previous_output=previous_output_path,
                bid_history=bid_history_path,
                pause_history=pause_history_path,
                as_of=None,
                out=str(local_out),
                print_summary=True,
            )
            decisions, core, payload, windows = compute_engine(engine_args)
            if getattr(args, "night_experiment_only", False):
                bh_for_restore = load_bid_history(bid_history_path)
                decisions = _filter_night_decisions(decisions, ph_for_rollback, bh_for_restore, getattr(args, "night_experiment_slot", "") or "")
                payload = build_payload_preview(decisions)
        write_outputs(str(local_out), decisions, core, payload, windows)
        successful, api_log = apply_api_actions(
            decisions,
            config,
            mode,
            bool(args.dry_run),
            bool(getattr(args, "apply_pause", False)),
            bool(getattr(args, "apply_start", False) or getattr(args, "rollback_wrong_pauses_only", False)),
            bool(getattr(args, "rollback_wrong_pauses_only", False)),
            bool(getattr(args, "night_experiment_only", False)),
            getattr(args, "night_experiment_slot", "") or "",
        )
        bid_history, pause_history = record_successful_events(successful, bid_history_path, pause_history_path)

        # Append API log to existing log if any.
        api_log_path = maybe_download_key_to_dir(s3, bucket, API_LOG_KEY, workdir)
        full_api_log = append_excel(api_log_path, api_log)
        summary = make_summary_json(mode, decisions, successful, full_api_log, windows, args)

        brush_tg_alerts = pd.DataFrame()
        brush_tg_result: Dict[str, Any] = {"tg_status": "not_requested"}
        brush_tg_pdf_out = workdir / "Проблемные_кисти_WB_Ads.pdf"
        if bool(getattr(args, "send_brush_tg", False)):
            period_label = f"{windows['current_start'].date().strftime('%d.%m')}-{windows['current_end'].date().strftime('%d.%m.%Y')} / сначала ПОИСК, затем ПОЛКИ"
            brush_tg_alerts, brush_tg_result = maybe_send_brush_tg_alert(
                s3,
                bucket,
                decisions,
                force=bool(getattr(args, "force_brush_tg", False)),
                schedule_only=bool(getattr(args, "brush_tg_schedule_only", False)),
                pdf_path=brush_tg_pdf_out,
                period_label=period_label,
            )
            summary["TG кисти: статус"] = brush_tg_result.get("tg_status", "")
            summary["TG кисти: строк"] = int(brush_tg_result.get("tg_rows", 0) or 0)
            summary["TG кисти: PDF"] = "да" if brush_tg_result.get("tg_pdf_created") else "нет"
            summary["TG кисти: отправлено"] = "да" if brush_tg_result.get("tg_sent") else "нет"

        summary_path = workdir / "Сводка_последнего_запуска.json"
        summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        bid_history_out = workdir / "История_ставок.xlsx"
        pause_history_out = workdir / "История_пауз.xlsx"
        api_log_out = workdir / "Лог_API.xlsx"
        brush_tg_out = workdir / "Проблемные_кисти_TG.xlsx"
        bid_history.to_excel(bid_history_out, index=False)
        pause_history.to_excel(pause_history_out, index=False)
        full_api_log.to_excel(api_log_out, index=False)
        brush_tg_alerts.to_excel(brush_tg_out, index=False)

        upload_s3_bytes(s3, bucket, PREVIEW_OUTPUT_KEY if mode == "preview" else RUN_OUTPUT_KEY, local_out.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        upload_s3_bytes(s3, bucket, SUMMARY_JSON_KEY, summary_path.read_bytes(), "application/json")
        upload_s3_bytes(s3, bucket, API_LOG_KEY, api_log_out.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        upload_s3_bytes(s3, bucket, BID_HISTORY_KEY, bid_history_out.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        upload_s3_bytes(s3, bucket, PAUSE_HISTORY_KEY, pause_history_out.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        upload_s3_bytes(s3, bucket, BRUSH_TG_ALERT_KEY, brush_tg_out.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        if brush_tg_pdf_out.exists():
            upload_s3_bytes(s3, bucket, BRUSH_TG_PDF_KEY, brush_tg_pdf_out.read_bytes(), "application/pdf")

        # Copy to repository workspace for GitHub artifacts.
        Path(local_out.name).write_bytes(local_out.read_bytes())
        Path(summary_path.name).write_bytes(summary_path.read_bytes())
        Path(api_log_out.name).write_bytes(api_log_out.read_bytes())
        Path(bid_history_out.name).write_bytes(bid_history_out.read_bytes())
        Path(pause_history_out.name).write_bytes(pause_history_out.read_bytes())
        Path(brush_tg_out.name).write_bytes(brush_tg_out.read_bytes())
        if brush_tg_pdf_out.exists():
            Path(brush_tg_pdf_out.name).write_bytes(brush_tg_pdf_out.read_bytes())

        print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
    return 0


def build_local_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="WB Ads Manager FIX46 decision engine")
    p.add_argument("--ads", required=True, help="WB advertising Excel: Реклама_YYYY-WNN.xlsx")
    p.add_argument("--orders", required=False, default=None, help="WB orders Excel for avg finishedPrice")
    p.add_argument("--previous-output", required=False, default=None, help="Previous output workbook with Ключевые_фразы_80")
    p.add_argument("--bid-history", required=False, default=None, help="История_ставок.xlsx")
    p.add_argument("--pause-history", required=False, default=None, help="История_пауз.xlsx")
    p.add_argument("--as-of", required=False, default=None, help="Decision date YYYY-MM-DD. Example: 2026-06-11")
    p.add_argument("--out", required=False, default=f"wb_ads_decisions_{VERSION}.xlsx")
    p.add_argument("--print-summary", action="store_true")
    return p


def build_legacy_runner_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="WB Ads Manager FIX46 working S3/API runner")
    p.add_argument("command", choices=["run", "preview"])
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--skip-price", action="store_true", help="Accepted for compatibility; price contour is absent in FIX46")
    p.add_argument("--apply-pause", action="store_true")
    p.add_argument("--apply-start", action="store_true")
    p.add_argument("--night-experiment-only", action="store_true")
    p.add_argument("--night-experiment-slot", choices=["start", "end", ""], default="")
    p.add_argument("--rollback-wrong-pauses-only", action="store_true", help="Разово запустить обратно кампании, ошибочно поставленные на паузу FIX46 v47")
    p.add_argument("--send-brush-tg", action="store_true", help="Сформировать и отправить TG по проблемным кистям")
    p.add_argument("--force-brush-tg", action="store_true", help="Отправить TG по кистям без дневного lock, для ручного запуска")
    p.add_argument("--brush-tg-schedule-only", action="store_true", help="TG по кистям только если понедельник >=19:05 МСК")
    return p


def main(argv: Optional[Sequence[str]] = None) -> int:
    argv = list(argv) if argv is not None else list(os.sys.argv[1:])
    if argv and argv[0] in {"run", "preview"}:
        parser = build_legacy_runner_parser()
        args = parser.parse_args(argv)
        return run_s3_legacy(args)
    parser = build_local_arg_parser()
    args = parser.parse_args(argv)
    return run_local(args)



# =========================
# V69 OVERRIDES: GP + CORE + POSTCHECK
# =========================

SCRIPT_VERSION = "v70-gp-core-postcheck-night-guard-fix-2026-06-25"
VERSION = "FIX46_CORE_RAMP_PAUSE_20260611_V70_GP_CORE_POSTCHECK"

DRR_FORECAST_CAP_PCT = 15.0
BRUSH_BID_CAP_DRR_PCT = 15.0
POSTCHECK_MAX_DAYS = 3
CORE_FLAGSHIP_TARGET_POSITION = 5
CORE_CLICK_LIFT_GOOD_PCT = 10.0
CORE_CLICK_LIFT_WEAK_PCT = 5.0
STORAGE_PCT_OF_SUM = 0.5

SEARCH_WEEKLY_PREFIX_CANDIDATES = [
    "Отчёты/Позиции по Ключам/TOPFACE/Недельные/",
    "Отчёты/Поисковые запросы/TOPFACE/Недельные/",
    "Отчёты/Поиск/TOPFACE/Недельные/",
]
ECONOMICS_KEY_CANDIDATES = [
    "Отчёты/Финансовые показатели/TOPFACE/Экономика.xlsx",
    "Отчёты/Финансовые показатели/TOPFACE/Юнит экономика.xlsx",
]

# Extend aliases for weekly keyword files and economics.
ALIASES["frequency"] = list(dict.fromkeys(ALIASES.get("frequency", []) + ["Частота запросов", "Частота за неделю"]))
ALIASES["visibility_pct"] = list(dict.fromkeys(ALIASES.get("visibility_pct", []) + ["Видимость %"]))
ALIASES["clicks"] = list(dict.fromkeys(ALIASES.get("clicks", []) + ["Переходы в карточку"]))
ALIASES["orders"] = list(dict.fromkeys(ALIASES.get("orders", []) + ["Заказы"]))
ALIASES["supplier_article"] = list(dict.fromkeys(ALIASES.get("supplier_article", []) + ["Артикул продавца"]))
ALIASES["subject"] = list(dict.fromkeys(ALIASES.get("subject", []) + ["Название предмета"]))
ALIASES["median_position"] = list(dict.fromkeys(ALIASES.get("median_position", []) + ["Медианная позиция"]))
ALIASES["campaign_status"] = list(dict.fromkeys(ALIASES.get("campaign_status", []) + ["Статус"]))
ALIASES["search_bid"] = list(dict.fromkeys(ALIASES.get("search_bid", []) + ["Ставка в поиске (руб)"]))
ALIASES["reco_bid"] = list(dict.fromkeys(ALIASES.get("reco_bid", []) + ["Ставка в рекомендациях (руб)"]))


def iso_week_label(ts: Any) -> str:
    if pd.isna(ts):
        return ""
    dt = pd.Timestamp(ts).to_pydatetime()
    iso = dt.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def parse_week_label_from_filename(path: Any) -> str:
    text = str(path or "")
    m = re.search(r"(\d{4}-W\d{2})", text)
    return m.group(1) if m else ""


def _last_non_null(series: pd.Series) -> Any:
    s = series.dropna()
    return s.iloc[-1] if len(s) else np.nan




def load_keywords_weekly(path: Optional[str]) -> pd.DataFrame:
    from openpyxl import load_workbook

    paths = expand_input_paths(path)
    frames: List[pd.DataFrame] = []
    wanted = [
        "Дата", "Поисковый запрос", "Артикул WB", "Предмет", "Артикул продавца",
        "Частота запросов", "Частота за неделю", "Медианная позиция", "Переходы в карточку",
        "Заказы", "Видимость %"
    ]
    for pth in paths:
        try:
            wb = load_workbook(pth, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            header = next(rows)
            header_map = {str(v).strip(): i for i, v in enumerate(header) if v is not None}
            idx = {name: header_map.get(name) for name in wanted}
            recs = []
            for row in rows:
                article = clean_article(row[idx["Артикул продавца"]] if idx["Артикул продавца"] is not None else "")
                query = str(row[idx["Поисковый запрос"]]).strip() if idx["Поисковый запрос"] is not None and row[idx["Поисковый запрос"]] is not None else ""
                if not article or not query:
                    continue
                recs.append({
                    "day": row[idx["Дата"]] if idx["Дата"] is not None else None,
                    "nm_id": row[idx["Артикул WB"]] if idx["Артикул WB"] is not None else None,
                    "supplier_article": article,
                    "subject_norm": canon_subject(row[idx["Предмет"]] if idx["Предмет"] is not None else ""),
                    "query_text": query,
                    "frequency": row[idx["Частота запросов"]] if idx["Частота запросов"] is not None else row[idx["Частота за неделю"]] if idx["Частота за неделю"] is not None else None,
                    "median_position": row[idx["Медианная позиция"]] if idx["Медианная позиция"] is not None else None,
                    "visibility_pct": row[idx["Видимость %"]] if idx["Видимость %"] is not None else None,
                    "clicks": row[idx["Переходы в карточку"]] if idx["Переходы в карточку"] is not None else None,
                    "orders": row[idx["Заказы"]] if idx["Заказы"] is not None else None,
                    "source_file": Path(pth).name,
                    "week_label": parse_week_label_from_filename(pth),
                })
            raw = pd.DataFrame(recs)
        except Exception:
            raw = pd.DataFrame()
        if raw.empty:
            continue
        raw["day"] = to_date(raw["day"])
        raw["nm_id"] = to_num(raw["nm_id"]).astype("Int64")
        raw["frequency"] = to_num(raw["frequency"]).fillna(0.0)
        raw["median_position"] = to_num(raw["median_position"])
        raw["visibility_pct"] = to_num(raw["visibility_pct"])
        raw["clicks"] = to_num(raw["clicks"]).fillna(0.0)
        raw["orders"] = to_num(raw["orders"]).fillna(0.0)
        raw["product_root"] = raw["supplier_article"].map(product_root)
        if raw["week_label"].eq("").all() and raw["day"].notna().any():
            raw["week_label"] = raw["day"].map(iso_week_label)
        raw = raw.sort_values(["supplier_article", "query_text", "day", "source_file"])
        raw = raw.drop_duplicates(["supplier_article", "query_text", "day"], keep="last")
        frames.append(raw)
    if not frames:
        return pd.DataFrame()
    raw = pd.concat(frames, ignore_index=True)
    rows: List[Dict[str, Any]] = []
    grp_cols = ["week_label", "supplier_article", "product_root", "subject_norm", "query_text"]
    for key, part in raw.groupby(grp_cols, dropna=False):
        part = part.sort_values("day")
        latest = part.iloc[-1]
        freq_latest = latest.get("frequency", np.nan)
        vis_latest = latest.get("visibility_pct", np.nan)
        rows.append({
            "week_label": key[0],
            "supplier_article": key[1],
            "product_root": key[2],
            "subject_norm": key[3],
            "query_text": key[4],
            "nm_id": _last_non_null(part["nm_id"]),
            "last_day": latest.get("day", pd.NaT),
            "frequency_latest": freq_latest,
            "frequency_avg": part["frequency"].mean(),
            "median_position_latest": latest.get("median_position", np.nan),
            "median_position_avg": part["median_position"].mean(),
            "visibility_pct_latest": vis_latest,
            "visibility_pct_avg": part["visibility_pct"].mean(),
            "clicks_sum": part["clicks"].sum(),
            "orders_sum": part["orders"].sum(),
            "days_seen": part["day"].nunique(),
            "impressions_proxy_latest": safe_div(freq_latest * vis_latest, 100.0, np.nan) if pd.notna(freq_latest) and pd.notna(vis_latest) else np.nan,
        })
    return pd.DataFrame(rows)


def load_economics(path: Optional[str]) -> pd.DataFrame:
    from openpyxl import load_workbook

    paths = expand_input_paths(path)
    frames: List[pd.DataFrame] = []
    wanted = [
        "Неделя", "Артикул WB", "Артикул продавца", "Предмет", "Средняя цена продажи",
        "Средняя цена покупателя", "СПП, %", "Комиссия WB, %", "Эквайринг, %",
        "Логистика прямая, руб/ед", "Логистика обратная, руб/ед", "Хранение, руб/ед",
        "Себестоимость, руб", "Валовая прибыль, руб/ед", "Чистые продажи, шт"
    ]
    for pth in paths:
        try:
            wb = load_workbook(pth, read_only=True, data_only=True)
            target = next((sh for sh in wb.sheetnames if sh.strip().lower() == "юнит экономика"), wb.sheetnames[0])
            ws = wb[target]
            rows = ws.iter_rows(values_only=True)
            header = next(rows)
            header_map = {str(v).strip(): i for i, v in enumerate(header) if v is not None}
            idx = {name: header_map.get(name) for name in wanted}
            recs = []
            for row in rows:
                week_label = str(row[idx["Неделя"]]).strip() if idx["Неделя"] is not None and row[idx["Неделя"]] is not None else ""
                article = clean_article(row[idx["Артикул продавца"]] if idx["Артикул продавца"] is not None else "")
                if not week_label or not article:
                    continue
                recs.append({
                    "week_label": week_label,
                    "nm_id": row[idx["Артикул WB"]] if idx["Артикул WB"] is not None else None,
                    "supplier_article": article,
                    "subject_norm": canon_subject(row[idx["Предмет"]] if idx["Предмет"] is not None else ""),
                    "avg_sale_price": row[idx["Средняя цена продажи"]] if idx["Средняя цена продажи"] is not None else None,
                    "avg_buyer_price": row[idx["Средняя цена покупателя"]] if idx["Средняя цена покупателя"] is not None else None,
                    "spp_pct": row[idx["СПП, %"]] if idx["СПП, %"] is not None else None,
                    "commission_pct": row[idx["Комиссия WB, %"]] if idx["Комиссия WB, %"] is not None else None,
                    "acquiring_pct": row[idx["Эквайринг, %"]] if idx["Эквайринг, %"] is not None else None,
                    "logistics_direct_unit": row[idx["Логистика прямая, руб/ед"]] if idx["Логистика прямая, руб/ед"] is not None else None,
                    "logistics_reverse_unit": row[idx["Логистика обратная, руб/ед"]] if idx["Логистика обратная, руб/ед"] is not None else None,
                    "storage_unit": row[idx["Хранение, руб/ед"]] if idx["Хранение, руб/ед"] is not None else None,
                    "cogs_unit": row[idx["Себестоимость, руб"]] if idx["Себестоимость, руб"] is not None else None,
                    "gross_profit_unit": row[idx["Валовая прибыль, руб/ед"]] if idx["Валовая прибыль, руб/ед"] is not None else None,
                    "net_sales_qty": row[idx["Чистые продажи, шт"]] if idx["Чистые продажи, шт"] is not None else None,
                })
            out = pd.DataFrame(recs)
        except Exception:
            out = pd.DataFrame()
        if out.empty:
            continue
        for c in ["nm_id","avg_sale_price","avg_buyer_price","spp_pct","commission_pct","acquiring_pct","logistics_direct_unit","logistics_reverse_unit","storage_unit","cogs_unit","gross_profit_unit","net_sales_qty"]:
            out[c] = to_num(out[c]) if c in out.columns else np.nan
        out["nm_id"] = out["nm_id"].astype("Int64")
        out["product_root"] = out["supplier_article"].map(product_root)
        out["logistics_unit"] = out[["logistics_direct_unit", "logistics_reverse_unit"]].fillna(0).sum(axis=1)
        out = out[out["week_label"].ne("") & out["supplier_article"].ne("")].copy()
        frames.append(out)
    if not frames:
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True)
    out = out.sort_values(["week_label", "supplier_article"]).drop_duplicates(["week_label", "supplier_article"], keep="last")
    return out

def _build_econ_lookups(econ: pd.DataFrame, week_label: str) -> Tuple[Dict[str, Dict[str, float]], Dict[str, Dict[str, float]]]:
    if econ is None or econ.empty:
        return {}, {}
    e = econ[econ["week_label"].astype(str).eq(str(week_label))].copy()
    if e.empty:
        return {}, {}
    article_map: Dict[str, Dict[str, float]] = {}
    for _, r in e.iterrows():
        article_map[clean_article(r["supplier_article"])] = r.to_dict()
    group_cols = ["spp_pct", "commission_pct", "acquiring_pct", "logistics_unit", "cogs_unit", "avg_sale_price", "avg_buyer_price"]
    group_avg = e.groupby("product_root", as_index=False)[group_cols].mean(numeric_only=True)
    group_map = {str(r["product_root"]): r.to_dict() for _, r in group_avg.iterrows()}
    return article_map, group_map


def _profitability_components(order_sum: float, spend: float, orders: float, subject: str, article: str, root: str,
                              article_rec: Dict[str, Any], group_rec: Dict[str, Any]) -> Dict[str, float]:
    order_sum = float(order_sum or 0.0)
    spend = float(spend or 0.0)
    orders = float(orders or 0.0)

    commission_pct = pd.to_numeric((article_rec or {}).get("commission_pct"), errors="coerce")
    if pd.isna(commission_pct):
        commission_pct = pd.to_numeric((group_rec or {}).get("commission_pct"), errors="coerce")
    acquiring_pct = pd.to_numeric((article_rec or {}).get("acquiring_pct"), errors="coerce")
    if pd.isna(acquiring_pct):
        acquiring_pct = pd.to_numeric((group_rec or {}).get("acquiring_pct"), errors="coerce")
    spp_pct = pd.to_numeric((group_rec or {}).get("spp_pct"), errors="coerce")
    if pd.isna(spp_pct):
        spp_pct = pd.to_numeric((article_rec or {}).get("spp_pct"), errors="coerce")
    article_cogs = pd.to_numeric((article_rec or {}).get("cogs_unit"), errors="coerce")
    group_cogs = pd.to_numeric((group_rec or {}).get("cogs_unit"), errors="coerce")
    article_log = pd.to_numeric((article_rec or {}).get("logistics_unit"), errors="coerce")
    group_log = pd.to_numeric((group_rec or {}).get("logistics_unit"), errors="coerce")

    if canon_subject(subject) == "Кисти косметические":
        cogs_unit = np.nanmean([article_cogs, group_cogs]) if pd.notna(article_cogs) or pd.notna(group_cogs) else np.nan
        logistics_unit = group_log if pd.notna(group_log) else article_log
    else:
        cogs_unit = article_cogs if pd.notna(article_cogs) else group_cogs
        logistics_unit = article_log if pd.notna(article_log) else group_log

    commission_rub = order_sum * (float(commission_pct) / 100.0) if pd.notna(commission_pct) else 0.0
    acquiring_rub = order_sum * (float(acquiring_pct) / 100.0) if pd.notna(acquiring_pct) else 0.0
    storage_rub = order_sum * STORAGE_PCT_OF_SUM / 100.0
    logistics_rub = orders * float(logistics_unit) if pd.notna(logistics_unit) else 0.0
    cogs_rub = orders * float(cogs_unit) if pd.notna(cogs_unit) else 0.0
    vat_rub = order_sum * (1.0 - (float(spp_pct) / 100.0 if pd.notna(spp_pct) else 0.0)) * 7.0 / 107.0

    gp_before_ads = order_sum - commission_rub - acquiring_rub - storage_rub - logistics_rub - cogs_rub - vat_rub
    gp = gp_before_ads - spend
    return {
        "commission_pct_used": float(commission_pct) if pd.notna(commission_pct) else np.nan,
        "acquiring_pct_used": float(acquiring_pct) if pd.notna(acquiring_pct) else np.nan,
        "spp_pct_used": float(spp_pct) if pd.notna(spp_pct) else np.nan,
        "logistics_unit_used": float(logistics_unit) if pd.notna(logistics_unit) else np.nan,
        "cogs_unit_used": float(cogs_unit) if pd.notna(cogs_unit) else np.nan,
        "commission_rub": commission_rub,
        "acquiring_rub": acquiring_rub,
        "storage_rub": storage_rub,
        "logistics_rub": logistics_rub,
        "cogs_rub": cogs_rub,
        "vat_rub": vat_rub,
        "gross_profit_before_ads_rub": gp_before_ads,
        "gross_profit_rub": gp,
        "gross_margin_pct": safe_div(gp, order_sum, np.nan) * 100.0 if order_sum > 0 else np.nan,
        "gross_profit_before_ads_per_order_rub": safe_div(gp_before_ads, orders, np.nan),
        "gross_profit_per_order_rub": safe_div(gp, orders, np.nan),
    }


def apply_profitability_metrics(df: pd.DataFrame, econ: pd.DataFrame, windows: Dict[str, pd.Timestamp]) -> pd.DataFrame:
    out = df.copy()
    current_week = iso_week_label(windows["current_end"])
    base_week = iso_week_label(windows["base_end"])
    article_cur, group_cur = _build_econ_lookups(econ, current_week)
    article_base, group_base = _build_econ_lookups(econ, base_week)

    gp_cur_rows: List[Dict[str, float]] = []
    gp_base_rows: List[Dict[str, float]] = []
    for _, r in out.iterrows():
        article = clean_article(r.get("supplier_article", ""))
        root = product_root(article)
        subject = canon_subject(r.get("subject_norm", ""))
        cur = _profitability_components(
            r.get("order_sum_cur", 0.0),
            r.get("spend_cur", 0.0),
            r.get("orders_cur", 0.0),
            subject,
            article,
            root,
            article_cur.get(article, {}),
            group_cur.get(root, {}),
        )
        base = _profitability_components(
            r.get("order_sum_base", 0.0),
            r.get("spend_base", 0.0),
            r.get("orders_base", 0.0),
            subject,
            article,
            root,
            article_base.get(article, {}),
            group_base.get(root, {}),
        )
        gp_cur_rows.append(cur)
        gp_base_rows.append(base)

    gp_cur_df = pd.DataFrame(gp_cur_rows).add_suffix("_cur")
    gp_base_df = pd.DataFrame(gp_base_rows).add_suffix("_base")
    out = pd.concat([out.reset_index(drop=True), gp_cur_df, gp_base_df], axis=1)
    out["gross_profit_delta_rub"] = out["gross_profit_rub_cur"] - out["gross_profit_rub_base"]
    out["gross_profit_delta_pct"] = np.where(
        out["gross_profit_rub_base"].abs() > 0,
        (out["gross_profit_rub_cur"] - out["gross_profit_rub_base"]) / out["gross_profit_rub_base"].abs() * 100.0,
        np.nan,
    )
    out["gross_profit_drop_vs_base_pct"] = np.where(
        out["gross_profit_rub_base"].abs() > 0,
        (out["gross_profit_rub_base"] - out["gross_profit_rub_cur"]) / out["gross_profit_rub_base"].abs() * 100.0,
        np.nan,
    )
    return out


def summarize_profitability_by_article(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=["supplier_article"])
    work = df.copy()
    cols_num = [
        "gross_profit_rub_cur", "gross_profit_rub_base", "gross_profit_before_ads_rub_cur", "gross_profit_before_ads_rub_base",
        "order_sum_cur", "order_sum_base", "orders_cur", "orders_base", "spend_cur", "spend_base",
        "commission_rub_cur", "acquiring_rub_cur", "storage_rub_cur", "logistics_rub_cur", "cogs_rub_cur", "vat_rub_cur",
        "commission_rub_base", "acquiring_rub_base", "storage_rub_base", "logistics_rub_base", "cogs_rub_base", "vat_rub_base",
    ]
    agg_map = {c: "sum" for c in cols_num if c in work.columns}
    agg_map.update({"product_root": "first", "subject_norm": "first"})
    g = work.groupby("supplier_article", as_index=False).agg(agg_map)
    g["gp_delta_rub_article"] = g["gross_profit_rub_cur"] - g["gross_profit_rub_base"]
    g["gp_delta_pct_article"] = np.where(
        g["gross_profit_rub_base"].abs() > 0,
        (g["gross_profit_rub_cur"] - g["gross_profit_rub_base"]) / g["gross_profit_rub_base"].abs() * 100.0,
        np.nan,
    )
    g["gp_before_ads_per_order_cur_article"] = np.where(
        g["orders_cur"] > 0,
        g["gross_profit_before_ads_rub_cur"] / g["orders_cur"],
        np.nan,
    )
    g["gp_per_order_cur_article"] = np.where(
        g["orders_cur"] > 0,
        g["gross_profit_rub_cur"] / g["orders_cur"],
        np.nan,
    )
    g["avg_order_value_cur_article"] = np.where(
        g["orders_cur"] > 0,
        g["order_sum_cur"] / g["orders_cur"],
        np.nan,
    )
    return g


def build_core_efficiency(keywords: pd.DataFrame, campaigns: pd.DataFrame, campaign_metrics: pd.DataFrame, windows: Optional[Dict[str, pd.Timestamp]] = None) -> pd.DataFrame:
    if keywords is None or keywords.empty or windows is None:
        return pd.DataFrame()

    current_week = iso_week_label(windows["current_end"])
    base_week = iso_week_label(windows["base_end"])
    cur = keywords[keywords["week_label"].astype(str).eq(current_week)].copy()
    base = keywords[keywords["week_label"].astype(str).eq(base_week)].copy()

    if cur.empty and base.empty:
        return pd.DataFrame()

    cur = cur.rename(columns={
        "frequency_latest": "frequency_cur",
        "median_position_latest": "median_position_cur",
        "visibility_pct_latest": "visibility_pct_cur",
        "clicks_sum": "clicks_cur",
        "orders_sum": "orders_cur",
        "impressions_proxy_latest": "impressions_proxy_cur",
        "last_day": "last_day_cur",
    })
    base = base.rename(columns={
        "frequency_latest": "frequency_base",
        "median_position_latest": "median_position_base",
        "visibility_pct_latest": "visibility_pct_base",
        "clicks_sum": "clicks_base",
        "orders_sum": "orders_base",
        "impressions_proxy_latest": "impressions_proxy_base",
        "last_day": "last_day_base",
    })
    key_cols = ["supplier_article", "product_root", "subject_norm", "query_text"]
    merged = cur[key_cols + [c for c in cur.columns if c not in key_cols and c != "week_label"]].merge(
        base[key_cols + [c for c in base.columns if c not in key_cols and c != "week_label"]],
        on=key_cols,
        how="outer",
    )

    # campaign/article proxies
    article_gp = summarize_profitability_by_article(campaign_metrics)
    search_only = campaign_metrics[campaign_metrics["placement"].astype(str).eq("search")].copy()
    if search_only.empty:
        cpc_by_article = {}
    else:
        tmp = search_only.groupby("supplier_article", as_index=False).agg(clicks_cur=("clicks_cur", "sum"), spend_cur=("spend_cur", "sum"))
        tmp["search_cpc_proxy"] = np.where(tmp["clicks_cur"] > 0, tmp["spend_cur"] / tmp["clicks_cur"], np.nan)
        cpc_by_article = tmp.set_index("supplier_article")["search_cpc_proxy"].to_dict()
    avg_order_value_by_article = article_gp.set_index("supplier_article")["avg_order_value_cur_article"].to_dict() if not article_gp.empty else {}
    gp_before_ads_per_order_by_article = article_gp.set_index("supplier_article")["gp_before_ads_per_order_cur_article"].to_dict() if not article_gp.empty else {}

    merged["proxy_cpc_cur"] = merged["supplier_article"].map(cpc_by_article)
    merged["avg_order_value_cur"] = merged["supplier_article"].map(avg_order_value_by_article)
    merged["gp_before_ads_per_order_cur"] = merged["supplier_article"].map(gp_before_ads_per_order_by_article)
    merged["query_spend_cur"] = merged["clicks_cur"].fillna(0.0) * merged["proxy_cpc_cur"].fillna(0.0)
    merged["query_spend_base"] = merged["clicks_base"].fillna(0.0) * merged["proxy_cpc_cur"].fillna(0.0)
    merged["query_cpo_cur"] = np.where(merged["orders_cur"].fillna(0) > 0, merged["query_spend_cur"] / merged["orders_cur"], np.nan)
    merged["query_cpo_base"] = np.where(merged["orders_base"].fillna(0) > 0, merged["query_spend_base"] / merged["orders_base"], np.nan)
    merged["query_order_sum_cur_est"] = merged["orders_cur"].fillna(0.0) * merged["avg_order_value_cur"].fillna(0.0)
    merged["query_order_sum_base_est"] = merged["orders_base"].fillna(0.0) * merged["avg_order_value_cur"].fillna(0.0)

    def pct_delta(cur_v: pd.Series, base_v: pd.Series) -> pd.Series:
        return np.where(base_v.fillna(0).abs() > 0, (cur_v.fillna(0) - base_v.fillna(0)) / base_v.abs() * 100.0, np.nan)

    merged["frequency_delta_pct"] = pct_delta(merged["frequency_cur"], merged["frequency_base"])
    merged["clicks_delta_pct"] = pct_delta(merged["clicks_cur"], merged["clicks_base"])
    merged["orders_delta_pct"] = pct_delta(merged["orders_cur"], merged["orders_base"])
    merged["visibility_delta_pct"] = pct_delta(merged["visibility_pct_cur"], merged["visibility_pct_base"])
    merged["adj_click_lift_vs_freq_pct"] = merged["clicks_delta_pct"].fillna(0.0) - np.where(merged["frequency_delta_pct"].fillna(0.0) > 0, merged["frequency_delta_pct"].fillna(0.0), 0.0)

    # One flagship product per query.
    merged["is_query_flagship_article"] = False
    for query_text, idxs in merged.groupby("query_text").groups.items():
        part = merged.loc[list(idxs)].copy()
        part = part.sort_values(
            ["orders_cur", "clicks_cur", "median_position_cur", "visibility_pct_cur", "frequency_cur"],
            ascending=[False, False, True, False, False],
            na_position="last",
        )
        if not part.empty:
            merged.loc[part.index[0], "is_query_flagship_article"] = True

    merged["query_role"] = np.where(merged["is_query_flagship_article"], "flagship_query_article", "secondary")
    merged.loc[(merged["orders_cur"].fillna(0) <= 0) & (merged["clicks_cur"].fillna(0) > 0), "query_role"] = "no_orders"
    merged["too_expensive_vs_margin"] = (
        merged["query_cpo_cur"].fillna(np.inf) > merged["gp_before_ads_per_order_cur"].fillna(-np.inf)
    )

    return merged


def summarize_core_by_product(core: pd.DataFrame) -> pd.DataFrame:
    if core is None or core.empty:
        return pd.DataFrame(columns=["supplier_article"])
    work = core[core["is_query_flagship_article"].fillna(False)].copy()
    if work.empty:
        return pd.DataFrame(columns=["supplier_article"])
    g = work.groupby("supplier_article", as_index=False).agg(
        product_root=("product_root", "first"),
        subject_norm=("subject_norm", "first"),
        flagship_queries=("query_text", lambda x: "; ".join(sorted(set(map(str, list(x)))))),
        flagship_query_count=("query_text", "count"),
        core_frequency_cur=("frequency_cur", "sum"),
        core_frequency_base=("frequency_base", "sum"),
        core_clicks_cur=("clicks_cur", "sum"),
        core_clicks_base=("clicks_base", "sum"),
        core_orders_cur=("orders_cur", "sum"),
        core_orders_base=("orders_base", "sum"),
        core_position_cur=("median_position_cur", "median"),
        core_position_base=("median_position_base", "median"),
        core_visibility_cur=("visibility_pct_cur", "mean"),
        core_visibility_base=("visibility_pct_base", "mean"),
        core_query_cpo_cur=("query_cpo_cur", "median"),
        core_query_cpo_base=("query_cpo_base", "median"),
        core_adj_click_lift_vs_freq_pct=("adj_click_lift_vs_freq_pct", "median"),
        core_impressions_proxy_cur=("impressions_proxy_cur", "sum"),
        core_impressions_proxy_base=("impressions_proxy_base", "sum"),
        expensive_query_count=("too_expensive_vs_margin", "sum"),
    )
    g["core_frequency_delta_pct"] = np.where(
        g["core_frequency_base"].fillna(0).abs() > 0,
        (g["core_frequency_cur"].fillna(0) - g["core_frequency_base"].fillna(0)) / g["core_frequency_base"].abs() * 100.0,
        np.nan,
    )
    g["core_clicks_delta_pct"] = np.where(
        g["core_clicks_base"].fillna(0).abs() > 0,
        (g["core_clicks_cur"].fillna(0) - g["core_clicks_base"].fillna(0)) / g["core_clicks_base"].abs() * 100.0,
        np.nan,
    )
    g["core_orders_delta_pct"] = np.where(
        g["core_orders_base"].fillna(0).abs() > 0,
        (g["core_orders_cur"].fillna(0) - g["core_orders_base"].fillna(0)) / g["core_orders_base"].abs() * 100.0,
        np.nan,
    )
    return g


def load_bid_history(path: Optional[str]) -> pd.DataFrame:
    df = read_sheet(path, ["Лист1", "История_изменений_ставок", "История ставок"])
    base_columns = [
        "campaign_id", "event_date", "run_datetime", "old_bid_rub", "new_bid_rub", "direction", "reason_code",
        "supplier_article", "subject_norm", "placement", "nm_id", "api_status", "postcheck_status",
        "baseline_core_frequency", "baseline_core_clicks", "baseline_core_position", "baseline_core_visibility",
        "baseline_core_orders", "baseline_core_cpo", "baseline_core_impressions_proxy",
        "baseline_gp_rub", "baseline_gp_before_ads_per_order_rub", "baseline_clicks_cur", "baseline_orders_cur",
        "baseline_order_sum_cur", "baseline_cpo_cur", "baseline_drr_pct_7d", "baseline_visibility_reason",
    ]
    if df.empty:
        return pd.DataFrame(columns=base_columns)
    out = pd.DataFrame({
        "campaign_id": to_num(s(df, "campaign_id")).astype("Int64"),
        "event_date": to_date(s(df, "day", np.nan) if find_col(df, "day") else df.get("event_date", pd.Series([np.nan] * len(df)))),
        "run_datetime": pd.to_datetime(df.get("run_datetime", pd.Series([pd.NaT] * len(df))), errors="coerce"),
        "old_bid_rub": to_num(df.get("old_bid_rub", pd.Series([np.nan] * len(df)))),
        "new_bid_rub": to_num(df.get("new_bid_rub", pd.Series([np.nan] * len(df)))),
        "direction": df.get("direction", pd.Series([""] * len(df))).astype(str),
        "reason_code": df.get("reason_code", pd.Series([""] * len(df))).astype(str),
        "supplier_article": df.get("supplier_article", pd.Series([""] * len(df))).astype(str),
        "subject_norm": df.get("subject_norm", pd.Series([""] * len(df))).astype(str).map(canon_subject),
        "placement": df.get("placement", pd.Series([""] * len(df))).astype(str),
        "nm_id": df.get("nm_id", pd.Series([""] * len(df))),
        "api_status": df.get("api_status", pd.Series([""] * len(df))).astype(str),
        "postcheck_status": df.get("postcheck_status", pd.Series([""] * len(df))).astype(str),
        "baseline_core_frequency": to_num(df.get("baseline_core_frequency", pd.Series([np.nan] * len(df)))),
        "baseline_core_clicks": to_num(df.get("baseline_core_clicks", pd.Series([np.nan] * len(df)))),
        "baseline_core_position": to_num(df.get("baseline_core_position", pd.Series([np.nan] * len(df)))),
        "baseline_core_visibility": to_num(df.get("baseline_core_visibility", pd.Series([np.nan] * len(df)))),
        "baseline_core_orders": to_num(df.get("baseline_core_orders", pd.Series([np.nan] * len(df)))),
        "baseline_core_cpo": to_num(df.get("baseline_core_cpo", pd.Series([np.nan] * len(df)))),
        "baseline_core_impressions_proxy": to_num(df.get("baseline_core_impressions_proxy", pd.Series([np.nan] * len(df)))),
        "baseline_gp_rub": to_num(df.get("baseline_gp_rub", pd.Series([np.nan] * len(df)))),
        "baseline_gp_before_ads_per_order_rub": to_num(df.get("baseline_gp_before_ads_per_order_rub", pd.Series([np.nan] * len(df)))),
        "baseline_clicks_cur": to_num(df.get("baseline_clicks_cur", pd.Series([np.nan] * len(df)))),
        "baseline_orders_cur": to_num(df.get("baseline_orders_cur", pd.Series([np.nan] * len(df)))),
        "baseline_order_sum_cur": to_num(df.get("baseline_order_sum_cur", pd.Series([np.nan] * len(df)))),
        "baseline_cpo_cur": to_num(df.get("baseline_cpo_cur", pd.Series([np.nan] * len(df)))),
        "baseline_drr_pct_7d": to_num(df.get("baseline_drr_pct_7d", pd.Series([np.nan] * len(df)))),
        "baseline_visibility_reason": df.get("baseline_visibility_reason", pd.Series([""] * len(df))).astype(str),
    })
    out = out[out["campaign_id"].notna()].copy()
    out["campaign_id"] = out["campaign_id"].astype(int)
    return out[base_columns]


def build_campaign_base(ads: pd.DataFrame, campaigns: pd.DataFrame, orders: pd.DataFrame, bid_history: pd.DataFrame, windows: Dict[str, pd.Timestamp]) -> pd.DataFrame:
    # keep original implementation but with postcheck horizon up to 3 days
    current = campaign_window_metrics(ads, windows["current_start"], windows["current_end"], "cur")
    base = campaign_window_metrics(ads, windows["base_start"], windows["base_end"], "base")
    p14 = campaign_window_metrics(ads, windows["pause_start"], windows["pause_end"], "14d")
    recent7_start = windows["as_of"] - pd.Timedelta(days=7)
    recent7_end = windows["as_of"] - pd.Timedelta(days=1)
    recent7 = campaign_window_metrics(ads, recent7_start, recent7_end, "7d")
    yesterday = campaign_window_metrics(ads, recent7_end, recent7_end, "yday")

    if campaigns.empty:
        ids = pd.concat([current[["campaign_id"]], base[["campaign_id"]], p14[["campaign_id"]], recent7[["campaign_id"]], yesterday[["campaign_id"]]], ignore_index=True).drop_duplicates()
        df = ids
    else:
        df = campaigns.copy()

    for part in [current, base, p14, recent7, yesterday]:
        df = df.merge(part, on="campaign_id", how="left")

    first_seen_ads = ads.groupby("campaign_id", as_index=False).agg(first_seen=("day", "min")) if not ads.empty else pd.DataFrame(columns=["campaign_id", "first_seen"])
    df = df.merge(first_seen_ads, on="campaign_id", how="left")

    if ads is not None and not ads.empty and "subject_norm" in ads.columns:
        ads_subjects = (
            ads[["campaign_id", "subject_norm"]]
            .dropna()
            .assign(subject_from_ads_daily=lambda d: d["subject_norm"].map(canon_subject))
            .groupby("campaign_id")["subject_from_ads_daily"]
            .agg(lambda s: s.value_counts().index[0] if len(s) else "")
            .reset_index()
        )
        if "subject_from_ads_daily" not in ads_subjects.columns:
            ads_subjects["subject_from_ads_daily"] = ""
        df = df.merge(ads_subjects[["campaign_id", "subject_from_ads_daily"]], on="campaign_id", how="left")
        if "subject_norm" not in df.columns:
            df["subject_norm"] = ""
        bad_subject = ~df["subject_norm"].map(is_managed_subject_value)
        fill_mask = bad_subject & df["subject_from_ads_daily"].map(is_managed_subject_value)
        df.loc[fill_mask, "subject_norm"] = df.loc[fill_mask, "subject_from_ads_daily"]
        df = df.drop(columns=["subject_from_ads_daily"], errors="ignore")

    df = filter_managed_subjects(df, "campaign_base")

    if not bid_history.empty:
        last = bid_history.dropna(subset=["event_date"]).sort_values(["event_date", "run_datetime"]).groupby("campaign_id", as_index=False).tail(1)
        rename_map = {
            "event_date": "last_bid_change_date",
            "old_bid_rub": "last_old_bid_rub",
            "new_bid_rub": "last_new_bid_rub",
            "direction": "last_bid_direction",
            "reason_code": "last_bid_reason_code",
            "baseline_core_frequency": "baseline_core_frequency",
            "baseline_core_clicks": "baseline_core_clicks",
            "baseline_core_position": "baseline_core_position",
            "baseline_core_visibility": "baseline_core_visibility",
            "baseline_core_orders": "baseline_core_orders",
            "baseline_core_cpo": "baseline_core_cpo",
            "baseline_core_impressions_proxy": "baseline_core_impressions_proxy",
            "baseline_gp_rub": "baseline_gp_rub",
            "baseline_gp_before_ads_per_order_rub": "baseline_gp_before_ads_per_order_rub",
            "baseline_clicks_cur": "baseline_clicks_cur",
            "baseline_orders_cur": "baseline_orders_cur",
            "baseline_order_sum_cur": "baseline_order_sum_cur",
            "baseline_cpo_cur": "baseline_cpo_cur",
            "baseline_drr_pct_7d": "baseline_drr_pct_7d",
            "baseline_visibility_reason": "baseline_visibility_reason",
        }
        need_cols = ["campaign_id"] + [c for c in rename_map.keys() if c in last.columns]
        last = last[need_cols].rename(columns=rename_map)
        df = df.merge(last, on="campaign_id", how="left")
    else:
        for c in [
            "last_bid_change_date","last_old_bid_rub","last_new_bid_rub","last_bid_direction","last_bid_reason_code",
            "baseline_core_frequency","baseline_core_clicks","baseline_core_position","baseline_core_visibility",
            "baseline_core_orders","baseline_core_cpo","baseline_core_impressions_proxy","baseline_gp_rub",
            "baseline_gp_before_ads_per_order_rub","baseline_clicks_cur","baseline_orders_cur","baseline_order_sum_cur",
            "baseline_cpo_cur","baseline_drr_pct_7d","baseline_visibility_reason",
        ]:
            df[c] = np.nan if c not in {"last_bid_direction","last_bid_reason_code","baseline_visibility_reason"} else ""

    prices = avg_price_by_product(orders, windows["current_start"], windows["current_end"])
    if not prices.empty and "nm_id" in df.columns:
        df = df.merge(prices[["product_root", "nm_id", "avg_finished_price", "avg_finished_price_root"]], on=["product_root", "nm_id"], how="left")
        df["avg_finished_price"] = df["avg_finished_price"].fillna(df["avg_finished_price_root"])
    else:
        df["avg_finished_price"] = np.nan

    metric_cols = []
    for suf in ["cur", "base", "14d", "7d", "yday"]:
        metric_cols += [
            f"impressions_{suf}", f"clicks_{suf}", f"orders_{suf}", f"spend_{suf}", f"order_sum_{suf}",
            f"ctr_pct_{suf}", f"cpc_{suf}", f"cpo_{suf}", f"clicks_per_order_{suf}",
            f"impressions_per_order_{suf}", f"drr_pct_{suf}",
        ]
    for col in metric_cols:
        if col not in df.columns:
            df[col] = np.nan
    for col in [c for c in df.columns if any(x in c for x in ["impressions_", "clicks_", "orders_", "spend_", "order_sum_"])]:
        df[col] = to_num(df[col]).fillna(0.0)

    df["days_since_first_seen"] = (windows["as_of"] - pd.to_datetime(df["first_seen"], errors="coerce")).dt.days
    df["is_new"] = df["days_since_first_seen"].fillna(9999) < NEW_NO_PAUSE_DAYS
    df["days_since_last_bid_change"] = (windows["as_of"] - pd.to_datetime(df["last_bid_change_date"], errors="coerce")).dt.days
    df["recent_bid_change"] = df["days_since_last_bid_change"].between(0, POSTCHECK_MAX_DAYS, inclusive="both")
    df["order_sum_drop_vs_base_pct"] = np.where(
        df["order_sum_base"].fillna(0) > 0,
        (df["order_sum_base"].fillna(0) - df["order_sum_cur"].fillna(0)) / df["order_sum_base"].fillna(0) * 100.0,
        np.nan,
    )
    return df


def bid_cap_target_drr_pct(subject: Any) -> float:
    return DRR_FORECAST_CAP_PCT


def _traffic_need_from_row(r: pd.Series) -> bool:
    pos_cur = pd.to_numeric(r.get("core_position_cur"), errors="coerce")
    vis_cur = pd.to_numeric(r.get("core_visibility_cur"), errors="coerce")
    vis_base = pd.to_numeric(r.get("core_visibility_base"), errors="coerce")
    freq_cur = pd.to_numeric(r.get("core_frequency_cur"), errors="coerce")
    freq_base = pd.to_numeric(r.get("core_frequency_base"), errors="coerce")
    clicks_cur = pd.to_numeric(r.get("core_clicks_cur"), errors="coerce")
    clicks_base = pd.to_numeric(r.get("core_clicks_base"), errors="coerce")

    freq_ok = (pd.notna(freq_base) and pd.notna(freq_cur) and float(freq_cur) >= float(freq_base) * 0.9) or (pd.notna(freq_cur) and pd.isna(freq_base))
    clicks_weak = (pd.notna(clicks_base) and pd.notna(clicks_cur) and float(clicks_cur) < float(clicks_base) * 0.9) or (pd.isna(clicks_base) and pd.notna(clicks_cur) and float(clicks_cur) <= 0)
    pos_bad = pd.notna(pos_cur) and float(pos_cur) > CORE_FLAGSHIP_TARGET_POSITION
    vis_weak = pd.notna(vis_cur) and ((pd.notna(vis_base) and float(vis_cur) < float(vis_base) * 0.9) or float(vis_cur) < 50.0)
    return bool(freq_ok and (pos_bad or vis_weak or clicks_weak))


def _expensive_traffic_from_row(r: pd.Series) -> bool:
    gp_before_ads = pd.to_numeric(r.get("gross_profit_before_ads_per_order_rub_cur"), errors="coerce")
    query_cpo = pd.to_numeric(r.get("core_query_cpo_cur"), errors="coerce")
    pos_cur = pd.to_numeric(r.get("core_position_cur"), errors="coerce")
    clicks_cur = pd.to_numeric(r.get("core_clicks_cur"), errors="coerce")
    clicks_base = pd.to_numeric(r.get("core_clicks_base"), errors="coerce")
    return bool(
        pd.notna(gp_before_ads) and pd.notna(query_cpo) and float(query_cpo) > float(gp_before_ads) and
        (pd.isna(pos_cur) or float(pos_cur) <= CORE_FLAGSHIP_TARGET_POSITION + 1) and
        (pd.isna(clicks_base) or pd.isna(clicks_cur) or float(clicks_cur) >= float(clicks_base) * 0.9)
    )


def _postcheck_decision_from_row(r: pd.Series) -> Optional[Dict[str, Any]]:
    days = pd.to_numeric(r.get("days_since_last_bid_change"), errors="coerce")
    direction = str(r.get("last_bid_direction", "") or "").strip().lower()
    if pd.isna(days) or days < 1 or days > POSTCHECK_MAX_DAYS or direction not in {"raise", "lower"}:
        return None

    cid = int(r["campaign_id"])
    placement = r.get("placement", "")
    step, min_bid, bid_effective, next_up, next_down, new_abs_max = _bid_grid_values(placement, r.get("real_bid_rub", np.nan))

    base_freq = pd.to_numeric(r.get("baseline_core_frequency"), errors="coerce")
    cur_freq = pd.to_numeric(r.get("core_frequency_cur"), errors="coerce")
    base_clicks = pd.to_numeric(r.get("baseline_core_clicks"), errors="coerce")
    cur_clicks = pd.to_numeric(r.get("core_clicks_cur"), errors="coerce")
    base_pos = pd.to_numeric(r.get("baseline_core_position"), errors="coerce")
    cur_pos = pd.to_numeric(r.get("core_position_cur"), errors="coerce")
    base_vis = pd.to_numeric(r.get("baseline_core_visibility"), errors="coerce")
    cur_vis = pd.to_numeric(r.get("core_visibility_cur"), errors="coerce")
    base_gp = pd.to_numeric(r.get("baseline_gp_rub"), errors="coerce")
    cur_gp = pd.to_numeric(r.get("gross_profit_rub_cur"), errors="coerce")

    freq_delta_pct = safe_div((cur_freq - base_freq), abs(base_freq), np.nan) * 100.0 if pd.notna(base_freq) else np.nan
    click_delta_pct = safe_div((cur_clicks - base_clicks), abs(base_clicks), np.nan) * 100.0 if pd.notna(base_clicks) else np.nan
    adj_click_lift = (click_delta_pct if pd.notna(click_delta_pct) else 0.0) - (freq_delta_pct if pd.notna(freq_delta_pct) and freq_delta_pct > 0 else 0.0)
    pos_improved = (base_pos - cur_pos) if pd.notna(base_pos) and pd.notna(cur_pos) else np.nan
    vis_improved = (cur_vis - base_vis) if pd.notna(base_vis) and pd.notna(cur_vis) else np.nan

    if direction == "raise":
        if (
            (pd.notna(adj_click_lift) and adj_click_lift >= CORE_CLICK_LIFT_GOOD_PCT)
            or (pd.notna(pos_improved) and pos_improved >= 1.0)
            or (pd.notna(vis_improved) and vis_improved >= 3.0)
        ):
            return decision(cid, "hold", bid_effective, "POSTCHECK_RAISE_POSITIVE_HOLD", f"post-check +{int(days)}д: после повышения есть эффект по CORE (adj_click_lift={adj_click_lift:.1f}%, pos_delta={pos_improved if pd.notna(pos_improved) else 'н/д'}, vis_delta={vis_improved if pd.notna(vis_improved) else 'н/д'})")
        if pd.notna(adj_click_lift) and adj_click_lift <= CORE_CLICK_LIFT_WEAK_PCT and (pd.isna(pos_improved) or pos_improved < 1.0) and (pd.isna(vis_improved) or vis_improved < 3.0):
            rollback_bid = pd.to_numeric(r.get("last_old_bid_rub"), errors="coerce")
            if pd.notna(rollback_bid) and float(rollback_bid) < float(bid_effective):
                return decision(cid, "lower", int(round(rollback_bid)) if str(placement).lower()=="search" else int(ceil_to_combined_grid(float(rollback_bid))), "POSTCHECK_RAISE_NO_CLICK_LIFT_ROLLBACK", f"post-check +{int(days)}д: после повышения клики по CORE не выросли при сопоставимой частотности; откат к {rollback_bid}")
            return decision(cid, "hold", bid_effective, "POSTCHECK_RAISE_NO_CLICK_LIFT_HOLD", f"post-check +{int(days)}д: эффекта от повышения не видно; не продолжаем повышать")
        return decision(cid, "hold", bid_effective, "POSTCHECK_RAISE_WAIT", f"post-check +{int(days)}д: собираем ещё данные до 3 дней")
    if direction == "lower":
        clicks_dropped = pd.notna(click_delta_pct) and click_delta_pct < -10.0
        pos_worse = pd.notna(pos_improved) and pos_improved < -1.0
        vis_worse = pd.notna(vis_improved) and vis_improved < -3.0
        gp_worse = pd.notna(base_gp) and pd.notna(cur_gp) and cur_gp < base_gp * 0.85
        if (clicks_dropped and (pos_worse or vis_worse)) or gp_worse:
            restore_bid = pd.to_numeric(r.get("last_old_bid_rub"), errors="coerce")
            if pd.notna(restore_bid) and float(restore_bid) > float(bid_effective):
                return decision(cid, "raise", int(round(restore_bid)) if str(placement).lower()=="search" else int(ceil_to_combined_grid(float(restore_bid))), "POSTCHECK_LOWER_NEGATIVE_RESTORE", f"post-check +{int(days)}д: после снижения трафик/видимость просели, ВП ухудшилась; возвращаем {restore_bid}")
            return decision(cid, "hold", bid_effective, "POSTCHECK_LOWER_NEGATIVE_HOLD", f"post-check +{int(days)}д: после снижения ухудшение, но restore_bid недоступен")
        return decision(cid, "hold", bid_effective, "POSTCHECK_LOWER_OK_HOLD", f"post-check +{int(days)}д: после снижения критичного провала нет; держим")
    return None


def decide_campaign(r: pd.Series, pause_history: pd.DataFrame, windows: Dict[str, pd.Timestamp]) -> Dict[str, Any]:
    cid = int(r["campaign_id"])
    placement = r.get("placement", "")
    active = bool(r.get("is_active", True))
    current_bid = r.get("real_bid_rub", np.nan)
    max_bid = r.get("max_allowed_bid_rub", np.nan)
    max_ceiling = r.get("max_allowed_ceiling_bid_rub", np.nan)
    drr_cur = r.get("drr_pct_cur", np.nan)
    drr_7d = r.get("drr_pct_7d", drr_cur)
    impressions_cur = float(r.get("impressions_cur", 0.0) or 0.0)
    impressions_7d = float(r.get("impressions_7d", impressions_cur) or 0.0)
    impressions_yday = float(r.get("impressions_yday", 0.0) or 0.0)
    clicks_cur = float(r.get("clicks_cur", 0.0) or 0.0)
    clicks_base = float(r.get("clicks_base", 0.0) or 0.0)
    orders_cur = float(r.get("orders_cur", 0.0) or 0.0)
    orders_7d = float(r.get("orders_7d", orders_cur) or 0.0)
    new_status = bool(r.get("is_new", False))
    is_leader = bool(r.get("is_block_leader", False))
    is_flagship = bool(r.get("is_flagship_campaign", False))
    active_in_block = int(r.get("active_in_block", 1) or 1)
    recent_bid = bool(r.get("recent_bid_change", False))
    subject = canon_subject(r.get("subject_norm", ""))

    gp_cur = pd.to_numeric(r.get("gross_profit_rub_cur"), errors="coerce")
    gp_base = pd.to_numeric(r.get("gross_profit_rub_base"), errors="coerce")
    gp_drop_pct = pd.to_numeric(r.get("gross_profit_drop_vs_base_pct"), errors="coerce")
    gp_before_ads_per_order = pd.to_numeric(r.get("gross_profit_before_ads_per_order_rub_cur"), errors="coerce")
    core_position_cur = pd.to_numeric(r.get("core_position_cur"), errors="coerce")
    core_visibility_cur = pd.to_numeric(r.get("core_visibility_cur"), errors="coerce")
    core_frequency_cur = pd.to_numeric(r.get("core_frequency_cur"), errors="coerce")
    core_clicks_cur = pd.to_numeric(r.get("core_clicks_cur"), errors="coerce")
    core_query_cpo_cur = pd.to_numeric(r.get("core_query_cpo_cur"), errors="coerce")

    if subject not in MANAGED_SUBJECTS_CANON:
        safe_bid = current_bid if pd.notna(current_bid) else np.nan
        return decision(cid, "hold", safe_bid, "OUT_OF_SCOPE_SUBJECT_HOLD", f"Предмет вне контура управления: {subject}. API запрещён.")

    article_norm = _normalize_article_for_experiment(r.get("supplier_article", ""))
    if article_norm in EXCLUDED_ARTICLES_FROM_AUTOMATION:
        safe_bid = current_bid if pd.notna(current_bid) else np.nan
        return decision(cid, "hold", safe_bid, "EXCLUDED_ARTICLE_HOLD", f"Артикул {article_norm} исключён из алгоритма управления: raise/lower/pause/start/API запрещены.")

    step, min_bid, bid_effective, next_up, next_down, new_abs_max = _bid_grid_values(placement, current_bid)

    last_pause_status = get_last_pause_status(pause_history, cid)
    if not active:
        start_decision = decide_start_candidate(r, last_pause_status, windows)
        return {"campaign_id": cid, **start_decision}

    if new_status:
        # keep NEW logic intact, but cap by 15% max + 1 step still applies after 14 days
        if impressions_7d >= RAMP_TARGET_IMPRESSIONS:
            if orders_7d <= 0:
                if subject in PAUSE_ALLOWED_SUBJECTS_CANON:
                    return decision(cid, "pause", bid_effective, "NEW_5000_IMPRESSIONS_ZERO_ORDERS_PAUSE_WAIT_MATURITY", "NEW набрала >=5000 показов и 0 заказов: сразу пауза на дозревание 7+3")
                return decision(cid, "hold", bid_effective, "BRUSH_NEW_5000_ZERO_ORDERS_TG_ONLY", "Кисти не паузим: NEW >=5000 показов и 0 заказов; проблема уйдёт в TG-уведомление")
            return decision(cid, "hold", bid_effective, "NEW_5000_WAIT_MATURITY", f"NEW набрала >=5000 показов; ждём дозревание 7+3")
        if impressions_yday < NEW_DAILY_IMPRESSIONS_LOW:
            if next_up <= new_abs_max:
                return decision(cid, "raise", next_up, "NEW_DAILY_IMPRESSIONS_LOW_RAISE", f"NEW<14д: вчера показов {impressions_yday:.0f}<700; цель 700-1000/день и 5000/неделю; ставка {bid_effective}->{next_up}; защитный потолок NEW={new_abs_max}")
            return decision(cid, "hold", bid_effective, "NEW_RAISE_BLOCKED_ABSOLUTE_CAP", f"NEW<14д: показов мало, но следующий шаг {next_up} выше защитного потолка {new_abs_max}; hold")
        if impressions_yday > NEW_DAILY_IMPRESSIONS_HIGH:
            return decision(cid, "lower", next_down, "NEW_DAILY_IMPRESSIONS_HIGH_LOWER", f"NEW<14д: вчера показов {impressions_yday:.0f}>1000; показы пошли, ставку снижаем {bid_effective}->{next_down}")
        return decision(cid, "hold", bid_effective, "NEW_DAILY_IMPRESSIONS_OK_HOLD", f"NEW<14д: вчера показов {impressions_yday:.0f}, это 700-1000/день; ставку держим")

    if active and pd.notna(max_ceiling) and float(bid_effective) > float(max_ceiling):
        forced_bid = int(round(max_ceiling)) if str(placement).lower() == "search" else int(ceil_to_combined_grid(max_ceiling))
        forced_bid = max(min_bid, forced_bid)
        return decision(cid, "lower", forced_bid, "CAP_OVERSHOOT_GT_ONE_STEP_FORCE_LOWER", f"Жёсткий cap: текущая ставка {bid_effective} выше рассчитанного максимума {max_bid} больше чем на 1 шаг; снижаем до {forced_bid}")

    # recent bid change: post-check after 1..3 days
    pc = _postcheck_decision_from_row(r)
    if pc is not None:
        return pc

    if impressions_7d >= RAMP_TARGET_IMPRESSIONS and orders_7d <= 0:
        if subject in PAUSE_ALLOWED_SUBJECTS_CANON:
            return decision(cid, "pause", bid_effective, "RAMP_5000_IMPRESSIONS_ZERO_ORDERS_PAUSE_WAIT_MATURITY", "РК набрала >=5000 показов и 0 заказов: не спасаем снижением, ставим на паузу на дозревание 7+3")
        return decision(cid, "hold", bid_effective, "BRUSH_5000_ZERO_ORDERS_TG_ONLY", "Кисти не паузим: >=5000 показов и 0 заказов; проблема уйдёт в TG-уведомление")

    need_traffic = _traffic_need_from_row(r)
    expensive_traffic = _expensive_traffic_from_row(r)

    # Flagship / CORE-driven rules first.
    if pd.notna(core_position_cur) and core_position_cur > CORE_FLAGSHIP_TARGET_POSITION and need_traffic and can_raise(next_up, max_ceiling):
        if pd.notna(gp_before_ads_per_order) and gp_before_ads_per_order > 0:
            return decision(cid, "raise", next_up, "CORE_FLAGSHIP_POSITION_RAISE", f"Флагман/CORE: позиция {core_position_cur:.1f} хуже топ-{CORE_FLAGSHIP_TARGET_POSITION}; при сопоставимой частотности трафика не хватает, повышаем {bid_effective}->{next_up}")

    if pd.notna(gp_cur) and gp_cur < 0 and bid_effective > min_bid and not recent_bid:
        return decision(cid, "lower", next_down, "GP_NEGATIVE_LOWER", f"ВП отрицательная ({gp_cur:.0f} ₽); снижаем ставку {bid_effective}->{next_down}")

    if pd.notna(gp_drop_pct) and gp_drop_pct > 15:
        if need_traffic and can_raise(next_up, max_ceiling) and pd.notna(gp_before_ads_per_order) and gp_before_ads_per_order > 0:
            return decision(cid, "raise", next_up, "GP_DROP_TRAFFIC_GAP_RAISE", f"ВП упала >15%, но причина в недостатке трафика по CORE: позиция/видимость/клики слабые, повышаем {bid_effective}->{next_up}")
        if bid_effective > min_bid and not recent_bid:
            return decision(cid, "lower", next_down, "GP_DROP_NOT_TRAFFIC_LOWER", f"ВП упала >15% и проблема не выглядит как дефицит трафика; снижаем {bid_effective}->{next_down}")

    if expensive_traffic and bid_effective > min_bid and not recent_bid:
        return decision(cid, "lower", next_down, "CORE_CPO_TOO_HIGH_LOWER", f"Трафик дорогой: query CPO {core_query_cpo_cur:.1f} выше маржи до рекламы на заказ {gp_before_ads_per_order:.1f}; снижаем {bid_effective}->{next_down}")

    if impressions_7d < RAMP_TARGET_IMPRESSIONS:
        if pd.notna(drr_7d) and drr_7d < DRR_RAISE_GATE_PCT:
            if can_raise(next_up, max_ceiling):
                return decision(cid, "raise", next_up, "RAMP_LT5000_DRR_LT10_RAISE", f"Показов за неделю {impressions_7d:.0f}<5000 и ДРР7={drr_7d:.1f}%<10%; добираем до 5000: {bid_effective}->{next_up}")
            return decision(cid, "hold", bid_effective, "RAMP_RAISE_BLOCKED_BY_CAP", f"<5000 и ДРР<10, но следующий шаг {next_up} выше допустимого потолка {max_ceiling}")
        if pd.notna(drr_7d) and drr_7d > DRR_RAISE_GATE_PCT:
            if is_flagship and need_traffic and can_raise(next_up, max_ceiling):
                return decision(cid, "raise", next_up, "FLAGSHIP_LT5000_TRAFFIC_GAP_RAISE", f"Флагман: <5000 показов и ДРР7={drr_7d:.1f}>10, но по CORE трафика не хватает; повышаем {bid_effective}->{next_up}")
            if is_flagship:
                return decision(cid, "hold", bid_effective, "FLAGSHIP_LT5000_HIGH_DRR_KEEP_WORKING", f"Флагман: <5000 показов и ДРР7={drr_7d:.1f}%>10; продолжаем наблюдение")
            if subject in PAUSE_ALLOWED_SUBJECTS_CANON and active_in_block > 1:
                return decision(cid, "pause", bid_effective, "PAUSE_WAIT_RAMP_QUEUE_LT5000_DRR_GT10", f"<5000 показов и ДРР7={drr_7d:.1f}%>10: пауза, ждёт очередь на разгон")
            return decision(cid, "hold", bid_effective, "BRUSH_OR_LAST_LT5000_DRR_GT10_HOLD", f"<5000 и ДРР>10, но кисть/единственная РК: не паузим")
        return decision(cid, "hold", bid_effective, "RAMP_LT5000_NOT_ENOUGH_DRR_HOLD", f"Показов {impressions_7d:.0f}<5000, стабильного ДРР нет; ждём/очередь")

    if need_traffic and can_raise(next_up, max_ceiling) and (pd.isna(gp_cur) or gp_cur >= 0 or pd.notna(gp_before_ads_per_order) and gp_before_ads_per_order > 0):
        return decision(cid, "raise", next_up, "CORE_TRAFFIC_GAP_RAISE", f"CORE: позиция/видимость/клики слабые при сопоставимой частотности; повышаем {bid_effective}->{next_up}")

    if pd.notna(drr_7d) and drr_7d <= DRR_RAISE_GATE_PCT and can_raise(next_up, max_ceiling) and not recent_bid:
        return decision(cid, "raise", next_up, "DRR_LT10_GE5000_RAISE", f">=5000 показов и ДРР7={drr_7d:.1f}%<10%; повышаем ставку на 1 шаг: {bid_effective}->{next_up}")

    if pd.notna(drr_7d) and drr_7d > DRR_RAISE_GATE_PCT:
        if bid_effective > min_bid and not recent_bid:
            return decision(cid, "lower", next_down, "DRR_GT10_GE5000_LOWER", f">=5000 показов и ДРР7={drr_7d:.1f}%>10%; снижаем ставку {bid_effective}->{next_down}")
        if bid_effective <= min_bid:
            if _min_bid_matured(r, min_bid, windows) and subject in PAUSE_ALLOWED_SUBJECTS_CANON and not is_flagship and active_in_block > 1:
                return decision(cid, "pause", bid_effective, "PAUSE_MIN_BID_MATURED_DRR_GT10", f"Ставка минимальная полное зрелое окно, ДРР7={drr_7d:.1f}%>10; пауза")
            return decision(cid, "hold", bid_effective, "MIN_BID_WAIT_FULL_MATURE_WINDOW", f"ДРР7={drr_7d:.1f}%>10, ставка минимальная; ждём полное зрелое окно на минимуме")

    return decision(cid, "hold", bid_effective, "NO_STRONG_SIGNAL_HOLD", "Нет сильного сигнала по CORE/ВП; без изменений")


def decide_all(campaigns: pd.DataFrame, core: pd.DataFrame, pause_history: pd.DataFrame, windows: Dict[str, pd.Timestamp]) -> pd.DataFrame:
    df = filter_managed_subjects(campaigns.copy(), "decide_all_input")
    df = filter_excluded_articles(df, "decide_all_excluded_articles")
    if df.empty:
        return pd.DataFrame(columns=["campaign_id", "action", "reason_code", "reason_text"])

    article_gp = summarize_profitability_by_article(df)
    if not article_gp.empty:
        df = df.merge(article_gp[[
            "supplier_article", "gp_delta_rub_article", "gp_delta_pct_article",
            "gp_before_ads_per_order_cur_article", "gp_per_order_cur_article", "avg_order_value_cur_article"
        ]], on="supplier_article", how="left")

    core_summary = summarize_core_by_product(core) if core is not None and not core.empty else pd.DataFrame()
    if not core_summary.empty:
        df = df.merge(core_summary, on="supplier_article", how="left")
    else:
        for c in [
            "flagship_queries","flagship_query_count","core_frequency_cur","core_frequency_base","core_clicks_cur","core_clicks_base",
            "core_orders_cur","core_orders_base","core_position_cur","core_position_base","core_visibility_cur","core_visibility_base",
            "core_query_cpo_cur","core_query_cpo_base","core_adj_click_lift_vs_freq_pct","core_impressions_proxy_cur","core_impressions_proxy_base",
            "core_frequency_delta_pct","core_clicks_delta_pct","core_orders_delta_pct"
        ]:
            df[c] = np.nan

    df = rank_blocks(df)
    df = select_ramp_slots(df)

    decisions = []
    for _, r in df.iterrows():
        decisions.append(decide_campaign(r, pause_history, windows))
    res = pd.DataFrame(decisions)
    return df.merge(res, on="campaign_id", how="left")


def build_postcheck_report(decisions: pd.DataFrame, bid_history: pd.DataFrame, windows: Dict[str, pd.Timestamp]) -> pd.DataFrame:
    if decisions is None or decisions.empty or bid_history is None or bid_history.empty:
        return pd.DataFrame()
    bh = bid_history.copy()
    bh["days_since_event"] = (windows["as_of"] - pd.to_datetime(bh["event_date"], errors="coerce")).dt.days
    bh = bh[bh["days_since_event"].between(1, POSTCHECK_MAX_DAYS, inclusive="both")].copy()
    if bh.empty:
        return pd.DataFrame()
    bh = bh.sort_values(["campaign_id", "event_date", "run_datetime"]).drop_duplicates("campaign_id", keep="last")
    d = decisions.copy()
    merged = bh.merge(
        d[[
            "campaign_id","supplier_article","placement","real_bid_rub","last_bid_direction","last_old_bid_rub","last_new_bid_rub",
            "core_frequency_cur","core_clicks_cur","core_position_cur","core_visibility_cur","core_orders_cur","core_query_cpo_cur",
            "core_impressions_proxy_cur","gross_profit_rub_cur","gross_profit_before_ads_per_order_rub_cur","clicks_cur","orders_cur",
            "order_sum_cur","cpo_cur","drr_pct_7d","reason_code","reason_text"
        ]].drop_duplicates("campaign_id"),
        on="campaign_id",
        how="left",
        suffixes=("_hist", "_cur"),
    )
    merged["postcheck_day"] = merged["days_since_event"]
    merged["freq_delta_pct"] = np.where(
        merged["baseline_core_frequency"].abs() > 0,
        (merged["core_frequency_cur"] - merged["baseline_core_frequency"]) / merged["baseline_core_frequency"].abs() * 100.0,
        np.nan,
    )
    merged["click_delta_pct"] = np.where(
        merged["baseline_core_clicks"].abs() > 0,
        (merged["core_clicks_cur"] - merged["baseline_core_clicks"]) / merged["baseline_core_clicks"].abs() * 100.0,
        np.nan,
    )
    merged["adj_click_lift_vs_freq_pct"] = merged["click_delta_pct"].fillna(0.0) - np.where(merged["freq_delta_pct"].fillna(0.0) > 0, merged["freq_delta_pct"].fillna(0.0), 0.0)
    merged["position_delta"] = merged["baseline_core_position"] - merged["core_position_cur"]
    merged["visibility_delta"] = merged["core_visibility_cur"] - merged["baseline_core_visibility"]
    merged["gp_delta_vs_baseline_pct"] = np.where(
        merged["baseline_gp_rub"].abs() > 0,
        (merged["gross_profit_rub_cur"] - merged["baseline_gp_rub"]) / merged["baseline_gp_rub"].abs() * 100.0,
        np.nan,
    )

    def _pc_reason(row: pd.Series) -> str:
        direction = str(row.get("direction", "") or "").lower()
        if direction == "raise":
            if (pd.notna(row.get("adj_click_lift_vs_freq_pct")) and row["adj_click_lift_vs_freq_pct"] >= CORE_CLICK_LIFT_GOOD_PCT) or \
               (pd.notna(row.get("position_delta")) and row["position_delta"] >= 1.0) or \
               (pd.notna(row.get("visibility_delta")) and row["visibility_delta"] >= 3.0):
                return "после повышения есть реальный прирост по CORE"
            if pd.notna(row.get("adj_click_lift_vs_freq_pct")) and row["adj_click_lift_vs_freq_pct"] <= CORE_CLICK_LIFT_WEAK_PCT:
                return "после повышения клики по CORE не выросли при сопоставимой частотности"
            return "после повышения ждём данные до 3 дней"
        if direction == "lower":
            if (pd.notna(row.get("click_delta_pct")) and row["click_delta_pct"] < -10.0) and \
               ((pd.notna(row.get("position_delta")) and row["position_delta"] < -1.0) or (pd.notna(row.get("visibility_delta")) and row["visibility_delta"] < -3.0)):
                return "после снижения трафик/видимость просели"
            return "после снижения критичного провала нет"
        return ""

    merged["postcheck_reason"] = merged.apply(_pc_reason, axis=1)
    keep_cols = [
        "campaign_id","supplier_article_hist","placement_hist","event_date","postcheck_day","direction",
        "old_bid_rub","new_bid_rub","baseline_core_frequency","core_frequency_cur","freq_delta_pct",
        "baseline_core_clicks","core_clicks_cur","click_delta_pct","adj_click_lift_vs_freq_pct",
        "baseline_core_position","core_position_cur","position_delta","baseline_core_visibility",
        "core_visibility_cur","visibility_delta","baseline_core_orders","core_orders_cur","baseline_core_cpo","core_query_cpo_cur",
        "baseline_gp_rub","gross_profit_rub_cur","gp_delta_vs_baseline_pct","postcheck_reason"
    ]
    rename = {
        "supplier_article_hist": "supplier_article",
        "placement_hist": "placement",
    }
    existing = [c for c in keep_cols if c in merged.columns]
    out = merged[existing].rename(columns=rename)
    return out


def write_outputs(path: str, decisions: pd.DataFrame, core: pd.DataFrame, payload: pd.DataFrame, windows: Dict[str, pd.Timestamp], postcheck: Optional[pd.DataFrame] = None) -> None:
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_table(decisions, windows).to_excel(writer, sheet_name="Сводка", index=False)

        cols_dec = [
            "campaign_id", "supplier_article", "product_root", "subject_norm", "placement", "campaign_status", "is_new", "days_since_first_seen",
            "is_flagship_campaign", "flagship_rank", "flagship_orders_share_pct", "ramp_status",
            "real_bid_rub", "economic_max_bid_raw", "max_allowed_bid_rub", "max_allowed_ceiling_bid_rub", "hard_cap_step_rub", "target_drr_cap_pct", "below_target_drr_at_wb_min_bid", "new_bid_rub", "action", "reason_code", "reason_text",
            "impressions_cur", "clicks_cur", "orders_cur", "spend_cur", "order_sum_cur", "drr_pct_cur", "cpo_cur", "ctr_pct_cur",
            "impressions_7d", "clicks_7d", "orders_7d", "spend_7d", "order_sum_7d", "drr_pct_7d", "ctr_pct_7d",
            "gross_profit_rub_cur", "gross_profit_rub_base", "gross_profit_delta_rub", "gross_profit_delta_pct", "gross_profit_drop_vs_base_pct",
            "gross_profit_before_ads_rub_cur", "gross_profit_before_ads_per_order_rub_cur", "gross_margin_pct_cur",
            "commission_rub_cur", "acquiring_rub_cur", "storage_rub_cur", "logistics_rub_cur", "cogs_rub_cur", "vat_rub_cur",
            "core_frequency_cur", "core_frequency_base", "core_frequency_delta_pct",
            "core_impressions_proxy_cur", "core_impressions_proxy_base", "core_clicks_cur", "core_clicks_base", "core_clicks_delta_pct",
            "core_orders_cur", "core_orders_base", "core_orders_delta_pct", "core_position_cur", "core_position_base",
            "core_visibility_cur", "core_visibility_base", "core_query_cpo_cur", "core_query_cpo_base",
            "core_adj_click_lift_vs_freq_pct", "flagship_queries", "flagship_query_count",
            "last_bid_change_date", "days_since_last_bid_change", "last_bid_reason_code",
        ]
        existing = [c for c in cols_dec if c in decisions.columns]
        decisions[existing].to_excel(writer, sheet_name="Решения", index=False)

        if core is not None and not core.empty:
            core.to_excel(writer, sheet_name="CORE_запросы", index=False)
            core_901 = core[(core["product_root"].astype(str).eq("901")) & (core["is_query_flagship_article"].fillna(False))].copy()
            core_901.to_excel(writer, sheet_name="CORE_флагманы_901", index=False)
        else:
            pd.DataFrame({"note": ["CORE source not provided. Provide weekly query files."]}).to_excel(writer, sheet_name="CORE_запросы", index=False)
            pd.DataFrame({"note": ["CORE flagships not built."]}).to_excel(writer, sheet_name="CORE_флагманы_901", index=False)

        gp_cols = [c for c in [
            "campaign_id","supplier_article","placement","orders_cur","order_sum_cur","spend_cur",
            "commission_pct_used_cur","acquiring_pct_used_cur","spp_pct_used_cur","logistics_unit_used_cur","cogs_unit_used_cur",
            "commission_rub_cur","acquiring_rub_cur","storage_rub_cur","logistics_rub_cur","cogs_rub_cur","vat_rub_cur",
            "gross_profit_before_ads_rub_cur","gross_profit_rub_cur","gross_margin_pct_cur","gross_profit_per_order_rub_cur","gross_profit_before_ads_per_order_rub_cur",
            "gross_profit_rub_base","gross_profit_delta_rub","gross_profit_delta_pct","gross_profit_drop_vs_base_pct"
        ] if c in decisions.columns]
        decisions[gp_cols].to_excel(writer, sheet_name="ВП_по_РК", index=False)

        cap_cols = [c for c in ["campaign_id", "supplier_article", "placement", "real_bid_rub", "economic_max_bid_raw", "max_allowed_bid_rub", "max_allowed_ceiling_bid_rub", "hard_cap_step_rub", "target_drr_cap_pct", "below_target_drr_at_wb_min_bid", "bid_at_calculated_max", "order_sum_cur", "order_sum_base", "order_sum_drop_vs_base_pct", "forecast_cpo_next_step", "forecast_drr_next_step_pct", "avg_finished_price", "clicks_per_order_cur", "impressions_per_order_cur", "max_bid_reason"] if c in decisions.columns]
        decisions[cap_cols].to_excel(writer, sheet_name="Предельные_ставки", index=False)

        pause_cols = [c for c in ["campaign_id", "supplier_article", "product_root", "placement", "action", "reason_code", "reason_text", "impressions_7d", "orders_7d", "drr_pct_7d", "impressions_14d", "drr_pct_14d", "active_in_block", "is_block_leader", "is_flagship_campaign", "is_new"] if c in decisions.columns]
        _safe_pause_start_subset(decisions, pause_cols).to_excel(writer, sheet_name="Паузы_и_возвраты", index=False)

        ramp_cols = [c for c in ["campaign_id", "supplier_article", "product_root", "placement", "ramp_queue_rank", "ramp_slot_selected", "ramp_status", "is_flagship_campaign", "flagship_rank", "action", "reason_code", "impressions_7d", "impressions_yday", "ctr_pct_cur", "orders_cur", "drr_pct_7d", "real_bid_rub", "new_bid_rub", "max_allowed_bid_rub", "max_allowed_ceiling_bid_rub"] if c in decisions.columns]
        sort_cols = [c for c in ["product_root", "placement", "ramp_queue_rank"] if c in decisions.columns]
        ramp_source = decisions.sort_values(sort_cols, na_position="last") if sort_cols else decisions.copy()
        ramp_source[ramp_cols].to_excel(writer, sheet_name="Разгон_очередь", index=False)

        block_reallocation_summary(decisions).to_excel(writer, sheet_name="Блоки_перелива", index=False)
        payload.to_excel(writer, sheet_name="API_payload_preview", index=False)

        if postcheck is not None and not postcheck.empty:
            postcheck.to_excel(writer, sheet_name="PostCheck_ставок", index=False)
        else:
            pd.DataFrame({"note": ["Нет кампаний с post-check 1..3 дня или отсутствуют baseline-метрики"]}).to_excel(writer, sheet_name="PostCheck_ставок", index=False)


def record_successful_events(successful: pd.DataFrame, bid_history_path: Optional[str], pause_history_path: Optional[str]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    bid_rows: List[Dict[str, Any]] = []
    pause_rows: List[Dict[str, Any]] = []
    if successful is not None and not successful.empty:
        for _, row in successful.iterrows():
            action = str(row.get("action", "") or "").lower()
            if action in {"raise", "lower"}:
                bid_rows.append({
                    "campaign_id": row.get("campaign_id", ""),
                    "event_date": datetime.now().date().isoformat(),
                    "run_datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "old_bid_rub": row.get("real_bid_rub", ""),
                    "new_bid_rub": row.get("new_bid_rub", ""),
                    "direction": "raise" if action == "raise" else "lower",
                    "reason_code": row.get("reason_code", ""),
                    "supplier_article": row.get("supplier_article", ""),
                    "subject_norm": row.get("subject_norm", ""),
                    "placement": row.get("placement", ""),
                    "nm_id": row.get("nm_id", ""),
                    "api_status": row.get("api_status", ""),
                    "postcheck_status": "",
                    "baseline_core_frequency": row.get("core_frequency_cur", np.nan),
                    "baseline_core_clicks": row.get("core_clicks_cur", np.nan),
                    "baseline_core_position": row.get("core_position_cur", np.nan),
                    "baseline_core_visibility": row.get("core_visibility_cur", np.nan),
                    "baseline_core_orders": row.get("core_orders_cur", np.nan),
                    "baseline_core_cpo": row.get("core_query_cpo_cur", np.nan),
                    "baseline_core_impressions_proxy": row.get("core_impressions_proxy_cur", np.nan),
                    "baseline_gp_rub": row.get("gross_profit_rub_cur", np.nan),
                    "baseline_gp_before_ads_per_order_rub": row.get("gross_profit_before_ads_per_order_rub_cur", np.nan),
                    "baseline_clicks_cur": row.get("clicks_cur", np.nan),
                    "baseline_orders_cur": row.get("orders_cur", np.nan),
                    "baseline_order_sum_cur": row.get("order_sum_cur", np.nan),
                    "baseline_cpo_cur": row.get("cpo_cur", np.nan),
                    "baseline_drr_pct_7d": row.get("drr_pct_7d", np.nan),
                    "baseline_visibility_reason": row.get("reason_text", ""),
                })
            elif action in {"pause", "start"}:
                pause_rows.append({
                    "campaign_id": row.get("campaign_id", ""),
                    "pause_date": datetime.now().date().isoformat(),
                    "status": "paused" if action == "pause" else "started",
                    "reason_code": row.get("reason_code", ""),
                    "api_status": row.get("api_status", ""),
                    "nm_id": row.get("nm_id", ""),
                    "placement": row.get("placement", ""),
                    "supplier_article": row.get("supplier_article", ""),
                    "subject_norm": row.get("subject_norm", ""),
                    "new_bid_rub": row.get("new_bid_rub", np.nan),
                    "next_check_date": (datetime.now().date() + timedelta(days=POST_PAUSE_CHECK_DAYS)).isoformat() if action == "pause" else "",
                })
    bid_history = append_excel(bid_history_path, pd.DataFrame(bid_rows))
    pause_history = append_excel(pause_history_path, pd.DataFrame(pause_rows))
    return bid_history, pause_history


def compute_engine(args: argparse.Namespace) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict[str, pd.Timestamp], pd.DataFrame]:
    as_of = pd.Timestamp(args.as_of).normalize() if getattr(args, "as_of", None) else pd.Timestamp(datetime.now().date())
    windows = date_windows(as_of)
    ads, campaigns = load_ads_daily(args.ads)
    orders = load_orders(getattr(args, "orders", None))
    bid_history = load_bid_history(getattr(args, "bid_history", None))
    pause_history = load_pause_history(getattr(args, "pause_history", None))

    weekly_spec = getattr(args, "keywords_weekly", None)
    keywords = load_keywords_weekly(weekly_spec) if weekly_spec else pd.DataFrame()
    if keywords.empty:
        keywords = load_keywords_from_previous(getattr(args, "previous_output", None))

    econ_spec = getattr(args, "economics", None)
    econ = load_economics(econ_spec) if econ_spec else pd.DataFrame()

    campaign_base = build_campaign_base(ads, campaigns, orders, bid_history, windows)
    campaign_base = filter_excluded_articles(campaign_base, "campaign_base_excluded_articles")
    campaign_base = apply_profitability_metrics(campaign_base, econ, windows)
    campaign_base = compute_bid_caps(campaign_base)
    core = build_core_efficiency(keywords, campaigns, campaign_base, windows)
    decisions = decide_all(campaign_base, core, pause_history, windows)
    payload = build_payload_preview(decisions)
    postcheck = build_postcheck_report(decisions, bid_history, windows)
    return decisions, core, payload, windows, postcheck


def run_local(args: argparse.Namespace) -> int:
    decisions, core, payload, windows, postcheck = compute_engine(args)
    write_outputs(args.out, decisions, core, payload, windows, postcheck)
    if args.print_summary:
        print(summary_table(decisions, windows).to_string(index=False))
        print(f"Output: {args.out}")
        print(f"API actions preview: {len(payload)}")
    return 0


def _download_candidate_keys(s3_client, bucket: str, candidates: Sequence[str], workdir: Path) -> List[str]:
    out: List[str] = []
    for candidate in candidates:
        if candidate.lower().endswith(".xlsx"):
            if s3_key_exists(s3_client, bucket, candidate):
                out.append(download_key_to_dir(s3_client, bucket, candidate, workdir))
        else:
            for key in latest_excel_keys(s3_client, bucket, candidate, limit=6):
                out.append(download_key_to_dir(s3_client, bucket, key, workdir))
    # dedupe
    seen = set()
    final = []
    for p in out:
        if p not in seen:
            final.append(p)
            seen.add(p)
    return final


def run_s3_legacy(args: argparse.Namespace) -> int:
    mode = args.command
    config = load_runner_config()
    s3 = make_s3_client(config)
    bucket = config.yc_bucket_name
    with tempfile.TemporaryDirectory(prefix="wb_ads_fix46_") as tmp:
        workdir = Path(tmp)
        ads_paths: List[str] = []
        if s3_key_exists(s3, bucket, ADS_MAIN_KEY):
            ads_paths.append(download_key_to_dir(s3, bucket, ADS_MAIN_KEY, workdir))
        else:
            for key in latest_excel_keys(s3, bucket, ADS_WEEKLY_PREFIX, limit=4):
                ads_paths.append(download_key_to_dir(s3, bucket, key, workdir))
        if not ads_paths:
            raise RuntimeError(f"Не найден рекламный отчёт: {ADS_MAIN_KEY} или {ADS_WEEKLY_PREFIX}")

        order_paths = [download_key_to_dir(s3, bucket, key, workdir) for key in latest_excel_keys(s3, bucket, ORDERS_WEEKLY_PREFIX, limit=4)]
        keyword_weekly_paths = _download_candidate_keys(s3, bucket, SEARCH_WEEKLY_PREFIX_CANDIDATES, workdir)
        economics_paths = _download_candidate_keys(s3, bucket, ECONOMICS_KEY_CANDIDATES, workdir)
        previous_output_path = maybe_download_key_to_dir(s3, bucket, RUN_OUTPUT_KEY, workdir)
        bid_history_path = maybe_download_key_to_dir(s3, bucket, BID_HISTORY_KEY, workdir)
        pause_history_path = maybe_download_key_to_dir(s3, bucket, PAUSE_HISTORY_KEY, workdir)

        local_out = workdir / ("Предпросмотр_последнего_запуска.xlsx" if mode == "preview" else "Итог_последнего_запуска.xlsx")
        ph_for_rollback = load_pause_history(pause_history_path)
        if getattr(args, "rollback_wrong_pauses_only", False):
            windows = date_windows(pd.Timestamp(datetime.now().date()))
            decisions = build_wrong_fix46_pause_rollback_decisions(ph_for_rollback)
            core = pd.DataFrame()
            payload = build_payload_preview(decisions)
            postcheck = pd.DataFrame()
            print(f"Разовый откат ошибочных пауз: кандидатов на START={len(decisions)}", flush=True)
        else:
            engine_args = argparse.Namespace(
                ads=";".join(ads_paths),
                orders=";".join(order_paths) if order_paths else None,
                previous_output=previous_output_path,
                bid_history=bid_history_path,
                pause_history=pause_history_path,
                keywords_weekly=";".join(keyword_weekly_paths) if keyword_weekly_paths else None,
                economics=";".join(economics_paths) if economics_paths else None,
                as_of=None,
                out=str(local_out),
                print_summary=True,
            )
            decisions, core, payload, windows, postcheck = compute_engine(engine_args)
            if getattr(args, "night_experiment_only", False):
                bh_for_restore = load_bid_history(bid_history_path)
                decisions = _filter_night_decisions(decisions, ph_for_rollback, bh_for_restore, getattr(args, "night_experiment_slot", "") or "")
                payload = build_payload_preview(decisions)
        write_outputs(str(local_out), decisions, core, payload, windows, postcheck)

        successful, api_log = apply_api_actions(
            decisions,
            config,
            mode,
            bool(args.dry_run),
            bool(getattr(args, "apply_pause", False)),
            bool(getattr(args, "apply_start", False) or getattr(args, "rollback_wrong_pauses_only", False)),
            bool(getattr(args, "rollback_wrong_pauses_only", False)),
            bool(getattr(args, "night_experiment_only", False)),
            getattr(args, "night_experiment_slot", "") or "",
        )
        bid_history, pause_history = record_successful_events(successful, bid_history_path, pause_history_path)

        api_log_path = maybe_download_key_to_dir(s3, bucket, API_LOG_KEY, workdir)
        full_api_log = append_excel(api_log_path, api_log)
        summary = make_summary_json(mode, decisions, successful, full_api_log, windows, args)

        brush_tg_alerts = pd.DataFrame()
        brush_tg_result: Dict[str, Any] = {"tg_status": "not_requested"}
        brush_tg_pdf_out = workdir / "Проблемные_кисти_WB_Ads.pdf"
        if bool(getattr(args, "send_brush_tg", False)):
            period_label = f"{windows['current_start'].date().strftime('%d.%m')}-{windows['current_end'].date().strftime('%d.%m.%Y')} / сначала ПОИСК, затем ПОЛКИ"
            brush_tg_alerts, brush_tg_result = maybe_send_brush_tg_alert(
                s3,
                bucket,
                decisions,
                force=bool(getattr(args, "force_brush_tg", False)),
                schedule_only=bool(getattr(args, "brush_tg_schedule_only", False)),
                pdf_path=brush_tg_pdf_out,
                period_label=period_label,
            )
            summary["TG кисти: статус"] = brush_tg_result.get("tg_status", "")
            summary["TG кисти: строк"] = int(brush_tg_result.get("tg_rows", 0) or 0)
            summary["TG кисти: PDF"] = "да" if brush_tg_result.get("tg_pdf_created") else "нет"
            summary["TG кисти: отправлено"] = "да" if brush_tg_result.get("tg_sent") else "нет"

        summary_path = workdir / "Сводка_последнего_запуска.json"
        summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        bid_history_out = workdir / "История_ставок.xlsx"
        pause_history_out = workdir / "История_пауз.xlsx"
        api_log_out = workdir / "Лог_API.xlsx"
        brush_tg_out = workdir / "Проблемные_кисти_TG.xlsx"
        bid_history.to_excel(bid_history_out, index=False)
        pause_history.to_excel(pause_history_out, index=False)
        full_api_log.to_excel(api_log_out, index=False)
        brush_tg_alerts.to_excel(brush_tg_out, index=False)

        upload_s3_bytes(s3, bucket, PREVIEW_OUTPUT_KEY if mode == "preview" else RUN_OUTPUT_KEY, local_out.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        upload_s3_bytes(s3, bucket, SUMMARY_JSON_KEY, summary_path.read_bytes(), "application/json")
        upload_s3_bytes(s3, bucket, API_LOG_KEY, api_log_out.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        upload_s3_bytes(s3, bucket, BID_HISTORY_KEY, bid_history_out.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        upload_s3_bytes(s3, bucket, PAUSE_HISTORY_KEY, pause_history_out.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        upload_s3_bytes(s3, bucket, BRUSH_TG_ALERT_KEY, brush_tg_out.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        if brush_tg_pdf_out.exists():
            upload_s3_bytes(s3, bucket, BRUSH_TG_PDF_KEY, brush_tg_pdf_out.read_bytes(), "application/pdf")

        Path(local_out.name).write_bytes(local_out.read_bytes())
        Path(summary_path.name).write_bytes(summary_path.read_bytes())
        Path(api_log_out.name).write_bytes(api_log_out.read_bytes())
        Path(bid_history_out.name).write_bytes(bid_history_out.read_bytes())
        Path(pause_history_out.name).write_bytes(pause_history_out.read_bytes())
        Path(brush_tg_out.name).write_bytes(brush_tg_out.read_bytes())
        if brush_tg_pdf_out.exists():
            Path(brush_tg_pdf_out.name).write_bytes(brush_tg_pdf_out.read_bytes())

        print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
    return 0


def build_local_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="WB Ads Manager FIX46 decision engine")
    p.add_argument("--ads", required=True, help="WB advertising Excel: Реклама_YYYY-WNN.xlsx")
    p.add_argument("--orders", required=False, default=None, help="WB orders Excel for avg finishedPrice")
    p.add_argument("--previous-output", required=False, default=None, help="Previous output workbook fallback for old CORE sheet")
    p.add_argument("--bid-history", required=False, default=None, help="История_ставок.xlsx")
    p.add_argument("--pause-history", required=False, default=None, help="История_пауз.xlsx")
    p.add_argument("--keywords-weekly", required=False, default=None, help="Неделя YYYY-WNN.xlsx files with Позиции по Ключам")
    p.add_argument("--economics", required=False, default=None, help="Экономика.xlsx")
    p.add_argument("--as-of", required=False, default=None, help="Decision date YYYY-MM-DD. Example: 2026-06-25")
    p.add_argument("--out", required=False, default=f"wb_ads_decisions_{VERSION}.xlsx")
    p.add_argument("--print-summary", action="store_true")
    return p


def build_legacy_runner_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="WB Ads Manager FIX46 working S3/API runner")
    p.add_argument("command", choices=["run", "preview"])
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--skip-price", action="store_true", help="Accepted for compatibility; price contour is absent in FIX46")
    p.add_argument("--apply-pause", action="store_true")
    p.add_argument("--apply-start", action="store_true")
    p.add_argument("--night-experiment-only", action="store_true")
    p.add_argument("--night-experiment-slot", choices=["start", "end", ""], default="")
    p.add_argument("--rollback-wrong-pauses-only", action="store_true", help="Разово запустить обратно кампании, ошибочно поставленные на паузу FIX46 v47")
    p.add_argument("--send-brush-tg", action="store_true", help="Сформировать и отправить TG по проблемным кистям")
    p.add_argument("--force-brush-tg", action="store_true", help="Отправить TG по кистям без дневного lock, для ручного запуска")
    p.add_argument("--brush-tg-schedule-only", action="store_true", help="TG по кистям только если понедельник >=19:05 МСК")
    return p


def main(argv: Optional[Sequence[str]] = None) -> int:
    argv = list(argv) if argv is not None else list(os.sys.argv[1:])
    if argv and argv[0] in {"run", "preview"}:
        parser = build_legacy_runner_parser()
        args = parser.parse_args(argv)
        return run_s3_legacy(args)
    parser = build_local_arg_parser()
    args = parser.parse_args(argv)
    return run_local(args)


if __name__ == "__main__":
    raise SystemExit(main())
