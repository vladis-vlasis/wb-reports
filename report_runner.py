# VERSION: ORDERS_ONLY_AND_TG_FIX10_TG_DAILY_CLEANUP_20260528

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WB TOPFACE analytics report.

Main goals:
- оперативная валовая прибыль по заказам с корректным % выкупа WB = buyouts / (buyouts + cancels);
- недельный факторный анализ по товарам и артикулам;
- средние и целевые значения за 90 дней;
- анализ рекламы manual/unified;
- поисковые запросы, % трафика, ядро ключей 80% заказов, позиции и рейтинги;
- точки входа / каналы продаж;
- локализация по складам с региональными заменами;
- выводы о причинах отклонения от плана/цели.

Outputs are overwritten every run:
1. Отчёты/Объединенный отчет/TOPFACE/Объединенный_отчет_TOPFACE.xlsx
2. Отчёты/Объединенный отчет/TOPFACE/Технические_расчеты_TOPFACE.xlsx
3. Отчёты/Объединенный отчет/TOPFACE/Пример_расчета_901_TOPFACE.xlsx
4. Отчёты/Объединенный отчет/TOPFACE/Средние_и_целевые_значения_TOPFACE.xlsx
5. Отчёты/Объединенный отчет/TOPFACE/Каналы_продаж_и_реклама_TOPFACE.xlsx
6. Отчёты/Объединенный отчет/TOPFACE/Поисковые_запросы_и_позиции_TOPFACE.xlsx
7. Отчёты/Объединенный отчет/TOPFACE/Локализация_TOPFACE.xlsx
8. Отчёты/Объединенный отчет/TOPFACE/Выводы_по_причинам_TOPFACE.xlsx
"""

from __future__ import annotations

import argparse
import calendar
import io
import math
import os
import re
import hashlib
import shutil
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import boto3
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TARGET_SUBJECTS = [
    "Кисти косметические",
    "Помады",
    "Блески",
    "Косметические карандаши",
]

# Global approved product -> category mapping for all operational reports.
# This prevents liquid eyeliners / liners (405/406/552) and service/stock-only
# products from leaking into "Косметические карандаши" through nmId dictionary joins.
# 158 is a lipstick family and must stay in "Помады".
VALID_PRODUCT_CATEGORY_REFERENCE: Dict[str, str] = {
    "901": "Кисти косметические",
    "605": "Косметические карандаши",
    "611": "Косметические карандаши",
    "613": "Косметические карандаши",
    "614": "Косметические карандаши",
    "617": "Косметические карандаши",
    "618": "Косметические карандаши",
    "154": "Помады",
    "155": "Помады",
    "156": "Помады",
    "157": "Помады",
    "158": "Помады",
    "206": "Помады",
    "207": "Блески",
    "209": "Блески",
    "210": "Блески",
    "211": "Блески",
}
VALID_PRODUCT_CODES = set(VALID_PRODUCT_CATEGORY_REFERENCE)

EXCLUDE_ARTICLES_UPPER = {
    "CZ420", "CZ420БРОВИ", "CZ420ГЛАЗА", "DE49", "DE49ГЛАЗА", "PT901",
}

EXAMPLE_ARTICLES = ["901/5", "901/8", "901/14", "901/18"]

OUT_DIR = "Отчёты/Объединенный отчет/TOPFACE"
MAIN_REPORT_NAME = "Объединенный_отчет_TOPFACE.xlsx"
TECH_REPORT_NAME = "Технические_расчеты_TOPFACE.xlsx"
EXAMPLE_REPORT_NAME = "Пример_расчета_901_TOPFACE.xlsx"
POTENTIAL_REPORT_NAME = "Средние_и_целевые_значения_TOPFACE.xlsx"
CHANNEL_REPORT_NAME = "Каналы_продаж_и_реклама_TOPFACE.xlsx"
SEARCH_REPORT_NAME = "Поисковые_запросы_и_позиции_TOPFACE.xlsx"
LOCALIZATION_REPORT_NAME = "Локализация_TOPFACE.xlsx"
CONCLUSIONS_REPORT_NAME = "Выводы_по_причинам_TOPFACE.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FILL = PatternFill("solid", fgColor="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
SUBSECTION_FILL = PatternFill("solid", fgColor="EAF4FF")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WEEKDAY_RU = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
MONTH_RU = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
    7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
}

# Broad alias dictionary. The code never fails because of one missing cosmetic column: it writes diagnostics instead.
ALIASES: Dict[str, Sequence[str]] = {
    "day": ["Дата", "Дата заказа", "Дата сбора", "Дата запроса", "dt", "date", "День"],
    "nm_id": ["Артикул WB", "Артикул ВБ", "Артикул ВБ", "nmID", "nmId", "nm_id", "Артикул WB", "Номенклатура", "Артикул ВБ"],
    "supplier_article": ["Артикул продавца", "supplierArticle", "supplier_article", "Артикул", "Артикул WB продавца"],
    "subject": ["Предмет", "subject", "Название предмета", "Категория", "category"],
    "brand": ["Бренд", "brand"],
    "title": ["Название", "Название товара", "Товар", "Наименование"],
    "warehouse": ["Склад", "warehouseName", "warehouse"],
    "orders": ["Заказы", "orders", "ordersCount", "Количество заказов", "Кол-во заказов", "Заказали товаров, шт", "Заказали, шт", "Заказали"],
    "order_sum": ["Сумма заказов", "ordersSumRub", "ordersSum", "Сумма заказов, руб", "Сумма заказов со скидкой", "Сумма заказов (со скидкой)", "Заказали на сумму", "Заказали, руб", "Сумма", "Итого"],
    "is_cancel": ["isCancel", "is_cancel", "Отменен", "Отменён", "Отмена", "Отменено", "Заказ отменен", "Заказ отменён"],
    "cancel_date": ["cancelDate", "cancel_date", "Дата отмены", "Дата отмены заказа", "Дата отмены/возврата"],
    "order_status": ["Статус", "status", "Статус заказа", "Статус товара", "Состояние"],
    "open_cards": ["Открытия карточки", "openCardCount", "Переходы в карточку", "Клики", "Клики карточки"],
    "add_to_cart": ["Добавления в корзину", "addToCartCount", "Корзины", "ATBS"],
    "cart_conv": ["Конверсия в корзину", "addToCartConversion", "Конверсия в корзину %"],
    "order_conv": ["Конверсия в заказ", "cartToOrderConversion", "Конверсия в заказ %"],
    "buyouts_count": ["buyoutsCount", "Выкупы", "Выкупили, шт", "Выкупили", "Кол-во выкупов"],
    "buyout_sum": ["buyoutsSumRub", "Выкупили, руб", "Сумма выкупов"],
    "cancels_count": ["cancelCount", "cancelsCount", "Отменили, шт", "Отменили", "Отмены", "Отменено"],
    "finished_price": ["finishedPrice", "Средняя конечная цена", "Средняя цена покупателя", "Цена покупателя", "Цена с учетом всех скидок, кроме суммы по WB Кошельку"],
    "price_with_disc": ["priceWithDisc", "Цена продажи", "Средняя цена продажи", "Цена со скидкой продавца, в том числе со скидкой WB Клуба"],
    "spp": ["СПП, %", "SPP", "Скидка WB, %", "spp"],
    "spend": ["Расход", "spend", "Продвижение", "Затраты", "Расходы"],
    "impressions": ["Показы", "shows", "views", "impressions"],
    "clicks": ["Клики", "Клики РК", "clicks", "clicksCount", "Переходы в карточку"],
    "ctr": ["CTR", "CTR РК", "ctr"],
    "cpc": ["CPC", "cpc"],
    "cr": ["CR", "cr"],
    "ad_orders": ["Заказы", "Заказы РК", "Заказы из рекламы", "orders"],
    "ad_order_sum": ["Сумма заказов", "Сумма заказов РК", "ordersSumRub"],
    "drr": ["ДРР", "ДРР, %", "drr", "drr_pct"],
    "campaign_id": ["ID кампании", "advertId", "campaignId"],
    "bid_type": ["Тип ставки", "bidType"],
    "search_query": ["Поисковый запрос", "Запрос", "Ключевой запрос", "Ключевая фраза", "keyword", "query"],
    "filter": ["Фильтр"],
    "frequency": ["Частота запросов", "Частотность", "Частота", "frequency"],
    "median_position": ["Медианная позиция", "medianPosition"],
    "avg_position": ["Средняя позиция", "averagePosition"],
    "visibility": ["Видимость %", "Видимость", "visibility"],
    "rating_card": ["Рейтинг карточки"],
    "rating_reviews": ["Рейтинг отзывов"],
    "entry_section": ["Раздел", "Источник", "Группа источника"],
    "entry_point": ["Точка входа", "Канал", "Источник перехода"],
    "stock": ["Доступно для продажи", "Остаток", "Остатки", "Доступный остаток", "Остаток, шт", "Количество", "Полное количество", "Всего", "stock", "quantity", "qty"],
    "gross_profit": ["Валовая прибыль", "Валовая прибыль, руб", "Валовая прибыль, руб/ед"],
    "gross_revenue": ["Валовая выручка", "Валовая выручка, руб", "Выручка"],
    "margin_pct": ["Рентабельность, %", "Рентабельность", "margin_pct"],
    "commission_amount": ["Комиссия", "Комиссия WB", "Комиссия ВБ"],
    "acquiring_amount": ["Эквайринг", "Эквайринг WB"],
    "commission_pct": ["Комиссия WB, %", "Комиссия ВБ, %", "Комиссия, %"],
    "acquiring_pct": ["Эквайринг, %", "Эквайринг WB, %"],
    "logistics_direct": ["Логистика прямая, руб/ед", "Логистика прямая"],
    "logistics_return": ["Логистика обратная, руб/ед", "Логистика обратная"],
    "storage": ["Хранение, руб/ед", "Хранение"],
    "other_costs": ["Прочие расходы, руб/ед", "Прочие расходы"],
    "cost": ["Себестоимость, руб", "Себестоимость", "Себестоимость, руб/ед"],
    "week": ["Неделя", "week", "week_code"],
}

REGIONAL_REPLACEMENT_POOLS = {
    "ЦФО": ["КОЛЕДИНО", "ЭЛЕКТРОСТАЛ", "БЕЛАЯ ДАЧ", "ВЕШК", "ВЁШК", "РЯЗАН", "ТУЛ", "АЛЕКСИН", "ВЛАДИМИР", "КОТОВСК", "ВОРОНЕЖ"],
    "ЮГ": ["КРАСНОДАР", "НЕВИННОМЫССК", "ВОЛГОГРАД", "РОСТОВ", "АКСАЙ"],
    "ПОВОЛЖЬЕ": ["КАЗАН", "ПЕНЗ", "САРАПУЛ", "НОВОСЕМЕЙКИНО", "САМАР"],
    "СЗФО": ["ШУШАР", "САНКТ", "ПЕТЕРБУРГ", "УТКИН", "СПБ"],
    "УРАЛ": ["ЕКАТЕРИНБУРГ", "ЧЕЛЯБИНСК", "ПЕРМ"],
    "СИБИРЬ": ["НОВОСИБИРСК", "КРАСНОЯРСК", "КЕМЕРОВО"],
}

# ------------------------- helpers -------------------------
def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}", flush=True)


def normalize_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ").strip())


def norm_key(value: Any) -> str:
    text = normalize_text(value).lower().replace("ё", "е")
    text = re.sub(r"[^a-zа-я0-9%]+", " ", text, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", text).strip()


SUBJECT_CANON_MAP = {
    "кисти косметические": "Кисти косметические",
    "помады": "Помады",
    "блески": "Блески",
    "косметические карандаши": "Косметические карандаши",
}


def canonical_subject(value: Any) -> str:
    key = norm_key(value)
    return SUBJECT_CANON_MAP.get(key, normalize_text(value))


def approved_subject_for_product(product: Any) -> str:
    code = normalize_text(product).upper().replace(" ", "")
    return VALID_PRODUCT_CATEGORY_REFERENCE.get(code, "")


def is_approved_product_subject(product: Any, subject: Any) -> bool:
    code = normalize_text(product).upper().replace(" ", "")
    subj = canonical_subject(subject)
    return bool(code and VALID_PRODUCT_CATEGORY_REFERENCE.get(code) == subj)


def clean_article(value: Any) -> str:
    text = normalize_text(value)
    if text.lower() in {"", "nan", "none", "null"}:
        return ""
    return text


def article_upper(value: Any) -> str:
    return clean_article(value).upper().replace(" ", "")


def product_code(article: Any) -> str:
    text = article_upper(article).replace("_", "/")
    if not text or text in EXCLUDE_ARTICLES_UPPER:
        return ""
    m = re.match(r"^PT(\d+)", text)
    if m:
        return m.group(1)
    m = re.match(r"^(\d+)", text)
    if m:
        return m.group(1)
    m = re.match(r"^([A-ZА-Я]+\d+)", text)
    if m:
        return m.group(1)
    return text.split("/")[0].split(".")[0]


def to_number(value: Any) -> float:
    if value is None:
        return np.nan
    if isinstance(value, str):
        value = value.replace("\xa0", " ").replace(" ", "").replace("₽", "").replace("%", "").replace(",", ".")
    return pd.to_numeric(value, errors="coerce")


def num_series(series: pd.Series) -> pd.Series:
    if series is None:
        return pd.Series(dtype=float)
    return series.map(to_number)


def date_series(series: pd.Series) -> pd.Series:
    """Parse dates safely: ISO yyyy-mm-dd as ISO, Russian dd.mm.yyyy as day-first."""
    def parse_one(v: Any) -> Any:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return pd.NaT
        if isinstance(v, (pd.Timestamp, datetime, date)):
            return pd.Timestamp(v).normalize()
        txt = normalize_text(v)
        if not txt or txt.lower() in {"nan", "none", "null"}:
            return pd.NaT
        try:
            if re.match(r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}", txt):
                return pd.to_datetime(txt, errors="coerce", dayfirst=False).normalize()
            if re.match(r"^\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}", txt):
                return pd.to_datetime(txt, errors="coerce", dayfirst=True).normalize()
            return pd.to_datetime(txt, errors="coerce").normalize()
        except Exception:
            return pd.NaT
    return series.map(parse_one)


def safe_div(a: Any, b: Any, default: float = np.nan) -> float:
    a = to_number(a)
    b = to_number(b)
    if pd.isna(a) or pd.isna(b) or b == 0:
        return default
    return float(a) / float(b)


def money_format() -> str:
    return '# ##0 ₽;[Red]-# ##0 ₽;0 ₽'


def pct_format() -> str:
    return '0.0%'


def parse_week_from_name(name: str) -> Optional[str]:
    m = re.search(r"(\d{4})-W(\d{2})", name)
    if m:
        return f"{m.group(1)}-W{m.group(2)}"
    return None


def week_bounds(week_code: str) -> Tuple[Optional[pd.Timestamp], Optional[pd.Timestamp]]:
    m = re.match(r"^(\d{4})-W(\d{2})$", str(week_code))
    if not m:
        return None, None
    start = pd.Timestamp(date.fromisocalendar(int(m.group(1)), int(m.group(2)), 1))
    return start, start + pd.Timedelta(days=6)


def week_code(ts: Any) -> str:
    if pd.isna(ts):
        return ""
    d = pd.Timestamp(ts)
    iso = d.isocalendar()
    return f"{int(iso.year)}-W{int(iso.week):02d}"


def parse_period_from_name(name: str) -> Tuple[Optional[pd.Timestamp], Optional[pd.Timestamp]]:
    patterns = [
        r"(\d{2})\.(\d{2})\.(\d{4})-(\d{2})\.(\d{2})\.(\d{4})",
        r"(\d{2})-(\d{2})-(\d{4}).*?(\d{2})-(\d{2})-(\d{4})",
        r"(\d{4})-(\d{2})-(\d{2}).*?(\d{4})-(\d{2})-(\d{2})",
    ]
    for p in patterns:
        m = re.search(p, name)
        if not m:
            continue
        if len(m.group(1)) == 4:
            return pd.Timestamp(date(int(m.group(1)), int(m.group(2)), int(m.group(3)))), pd.Timestamp(date(int(m.group(4)), int(m.group(5)), int(m.group(6))))
        return pd.Timestamp(date(int(m.group(3)), int(m.group(2)), int(m.group(1)))), pd.Timestamp(date(int(m.group(6)), int(m.group(5)), int(m.group(4))))
    wk = parse_week_from_name(name)
    if wk:
        return week_bounds(wk)
    return None, None




def filter_recent_report_files(files: List[str], latest_day: pd.Timestamp, lookback_days: int = 110, keep_unknown: bool = True) -> List[str]:
    """Keep report files whose period intersects the requested lookback window.
    This avoids parsing old weekly search/stock/entry reports from S3.
    """
    if latest_day is None or pd.isna(latest_day):
        return sorted(set(files))
    cutoff = pd.Timestamp(latest_day).normalize() - pd.Timedelta(days=lookback_days)
    out: List[str] = []
    skipped_old = 0
    skipped_unknown = 0
    for f in sorted(set(files)):
        start, end = parse_period_from_name(Path(f).name)
        if start is None or end is None:
            if keep_unknown:
                out.append(f)
            else:
                skipped_unknown += 1
            continue
        if pd.Timestamp(end).normalize() >= cutoff and pd.Timestamp(start).normalize() <= pd.Timestamp(latest_day).normalize():
            out.append(f)
        else:
            skipped_old += 1
    if skipped_old or skipped_unknown:
        log(f"recent_file_filter: input={len(set(files))}, kept={len(out)}, skipped_old={skipped_old}, skipped_unknown={skipped_unknown}, cutoff={cutoff.date()}")
    return out



def limit_recent_report_files(files: List[str], max_files: int) -> List[str]:
    """Keep unknown support files and the newest dated report files by parsed period end.

    Using lexical sorting for names like "7-4-2026 ..." and "29-3-2026 ..." can pull
    old weeks instead of the latest ones. This helper uses dates parsed from the file
    name, so max-file limits reduce runtime without accidentally keeping older reports.
    """
    unique = sorted(set(files))
    if max_files is None or int(max_files) <= 0 or len(unique) <= int(max_files):
        known = []
        unknown = []
        for f in unique:
            start, end = parse_period_from_name(Path(f).name)
            if start is None or end is None:
                unknown.append(f)
            else:
                known.append((pd.Timestamp(end).normalize(), pd.Timestamp(start).normalize(), f))
        return sorted(unknown) + [f for _, _, f in sorted(known)]

    max_files = int(max_files)
    known = []
    unknown = []
    for f in unique:
        start, end = parse_period_from_name(Path(f).name)
        if start is None or end is None:
            unknown.append(f)
        else:
            known.append((pd.Timestamp(end).normalize(), pd.Timestamp(start).normalize(), f))
    known_sorted = [f for _, _, f in sorted(known)]

    # Preserve service/reference files without dates when possible; use the remaining
    # quota for the latest dated weekly/monthly files.
    if len(unknown) >= max_files:
        return sorted(unknown)[-max_files:]
    keep_known = max_files - len(unknown)
    return sorted(unknown) + known_sorted[-keep_known:]

def is_month_file(start: pd.Timestamp, end: pd.Timestamp) -> bool:
    if start is None or end is None or pd.isna(start) or pd.isna(end):
        return False
    return start.day == 1 and start.year == end.year and start.month == end.month and end.day == calendar.monthrange(start.year, start.month)[1]


def month_key(ts: Any) -> str:
    d = pd.Timestamp(ts)
    return f"{d.year:04d}-{d.month:02d}"


def unwrap_excel_bytes(data: bytes) -> bytes:
    """If uploaded report is a zip with xlsx inside, return the first xlsx bytes."""
    if data[:2] != b"PK":
        return data
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = [n for n in zf.namelist() if n.lower().endswith((".xlsx", ".xlsm")) and not Path(n).name.startswith("~$")]
            if names:
                # prefer files not inside __MACOSX
                names = sorted(names, key=lambda x: ("__MACOSX" in x, len(x)))
                return zf.read(names[0])
    except zipfile.BadZipFile:
        pass
    return data


def get_col(df: pd.DataFrame, logical_name: str) -> pd.Series:
    if logical_name in df.columns:
        return df[logical_name]
    return pd.Series([np.nan] * len(df))


def add_alias_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    col_by_key = {norm_key(c): c for c in out.columns}
    for target, variants in ALIASES.items():
        if target in out.columns:
            continue
        found = None
        for v in variants:
            key = norm_key(v)
            if key in col_by_key:
                found = col_by_key[key]
                break
        out[target] = out[found] if found is not None else np.nan
    return out


def read_excel_table(data: bytes, preferred_sheet: Optional[str] = None, header_rows: Iterable[int] = (0, 1, 2, 3, 4)) -> pd.DataFrame:
    data = unwrap_excel_bytes(data)
    xl = pd.ExcelFile(io.BytesIO(data))
    if preferred_sheet and preferred_sheet in xl.sheet_names:
        sheets = [preferred_sheet]
    else:
        sheets = xl.sheet_names
    best_df = None
    best_score = -1.0
    for sheet in sheets:
        for h in header_rows:
            try:
                df = xl.parse(sheet_name=sheet, header=h, dtype=object)
            except Exception:
                continue
            df = df.dropna(how="all").dropna(axis=1, how="all")
            if df.empty:
                continue
            df.columns = [normalize_text(c) or f"col_{i}" for i, c in enumerate(df.columns)]
            aliased = add_alias_columns(df)
            score = min(len(df.columns), 50) / 100
            for c in ["day", "nm_id", "supplier_article", "subject", "orders", "order_sum", "search_query", "entry_section", "spend", "stock", "gross_profit"]:
                if c in aliased.columns and not aliased[c].isna().all():
                    score += 1
            if score > best_score:
                best_score = score
                aliased.attrs["source_sheet"] = sheet
                aliased.attrs["header_row_0based"] = h
                aliased.attrs["header_row_excel"] = h + 1
                best_df = aliased
    if best_df is None:
        return pd.DataFrame()
    return best_df


def _source_column_name(df: pd.DataFrame, logical_name: str) -> str:
    """Return the original source column name used for a logical alias when possible."""
    if df is None or df.empty:
        return ""
    variants = ALIASES.get(logical_name, []) + [logical_name]
    by_key = {norm_key(c): c for c in df.columns}
    # Prefer non-canonical original columns, then canonical alias.
    for v in variants:
        k = norm_key(v)
        if k in by_key and by_key[k] != logical_name:
            return by_key[k]
    if logical_name in df.columns:
        return logical_name
    for v in variants:
        k = norm_key(v)
        if k in by_key:
            return by_key[k]
    return ""


def _source_cell_ref(df: pd.DataFrame, logical_name: str, row_index: Any) -> str:
    """Return Excel-like cell reference for a logical value after read_excel_table.

    Header detection is done by read_excel_table; row_index is the parsed dataframe row index.
    This is used only for audit/debug logs, not for calculations.
    """
    try:
        col_name = _source_column_name(df, logical_name)
        if not col_name or col_name not in df.columns:
            return ""
        col_no = list(df.columns).index(col_name) + 1
        h0 = int(df.attrs.get("header_row_0based", 0) or 0)
        # pandas row index starts from 0 for the first data row after the header.
        excel_row = int(row_index) + h0 + 2
        return f"{get_column_letter(col_no)}{excel_row}"
    except Exception:
        return ""


def safe_sheet_name(name: str, used: set) -> str:
    raw = re.sub(r"[\\/*?:\[\]]", "_", normalize_text(name))[:31] or "Sheet"
    name = raw
    i = 1
    while name in used:
        suffix = f"_{i}"
        name = raw[:31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def clean_warehouse_name(name: Any) -> str:
    return normalize_text(name).upper().replace("Ё", "Е")


def warehouse_pool(name: Any) -> str:
    n = clean_warehouse_name(name)
    for pool, keys in REGIONAL_REPLACEMENT_POOLS.items():
        if any(k in n for k in keys):
            return pool
    return "ДРУГИЕ"


def weighted_mean(values: pd.Series, weights: pd.Series) -> float:
    v = pd.to_numeric(values, errors="coerce")
    w = pd.to_numeric(weights, errors="coerce").fillna(0)
    mask = v.notna() & (w > 0)
    if mask.any():
        return float(np.average(v[mask], weights=w[mask]))
    if v.notna().any():
        return float(v.mean())
    return np.nan


def classify_ad_type(value: Any) -> str:
    t = norm_key(value)
    if "manual" in t or "руч" in t or "поиск" in t or "каталог" in t:
        return "manual"
    if "unified" in t or "авто" in t or "един" in t or "карточ" in t or "полк" in t or "рекомен" in t:
        return "unified"
    return "unknown"


# ------------------------- business cleanup helpers -------------------------
# Prefix exclusions are only for real excluded article families.
# PT901.Fxx are valid brush articles and must not be filtered as the raw PT901 service code.
EXCLUDE_ARTICLE_PREFIXES = ("CZ420", "DE49")
EXCLUDE_ARTICLE_EXACT = {"PT901"}
EXCLUDE_PRODUCT_PREFIXES = ("CZ420", "DE49", "FL", "PE")
WAREHOUSE_EXCLUDE_KEYWORDS = (
    "ВИРТУАЛ", "АСТАН", "АЛМАТ", "АТАКЕНТ", "КАРАГАНД", "КАЗАХ", "БЕЛАРУС", "МИНСК",
    "ДАЛЬНЕГОРСК", "МАХАЧКАЛА ВИРТ", "ВИРТУАЛЬНЫЙ",
)


def is_excluded_article(value: Any) -> bool:
    t = article_upper(value)
    if not t:
        return True
    if t in EXCLUDE_ARTICLE_EXACT:
        return True
    return any(t.startswith(prefix) for prefix in EXCLUDE_ARTICLE_PREFIXES)


def is_valid_product_code(value: Any) -> bool:
    t = normalize_text(value).upper().replace(" ", "")
    if not t:
        return False
    if any(t.startswith(prefix) for prefix in EXCLUDE_PRODUCT_PREFIXES):
        return False
    # For management reports we keep only product families approved in the global mapping.
    # This removes 405/406/552/620/622 from target reports unless explicitly added above.
    return t in VALID_PRODUCT_CODES


def canonical_warehouse_name(name: Any) -> str:
    n = clean_warehouse_name(name)
    rules = [
        ("КОЛЕДИНО", "Коледино"), ("ЭЛЕКТРОСТАЛ", "Электросталь"), ("БЕЛАЯ ДАЧ", "Белая Дача"),
        ("ВЕШК", "Вёшки"), ("ВЁШК", "Вёшки"), ("РЯЗАН", "Рязань"), ("ТУЛ", "Тула"), ("АЛЕКСИН", "Тула"),
        ("ВЛАДИМИР", "Владимир"), ("КОТОВСК", "Котовск"), ("ВОРОНЕЖ", "Воронеж"),
        ("КРАСНОДАР", "Краснодар"), ("НЕВИННОМЫССК", "Невинномысск"), ("ВОЛГОГРАД", "Волгоград"),
        ("РОСТОВ", "Ростов/Аксай"), ("АКСАЙ", "Ростов/Аксай"),
        ("КАЗАН", "Казань"), ("ПЕНЗ", "Пенза"), ("САРАПУЛ", "Сарапул"), ("НОВОСЕМЕЙКИНО", "Новосемейкино"), ("САМАР", "Самара"),
        ("ШУШАР", "СПБ Шушары"), ("УТКИН", "СПБ Уткина Заводь"), ("САНКТ", "Санкт-Петербург"), ("ПЕТЕРБУРГ", "Санкт-Петербург"),
        ("ЕКАТЕРИНБУРГ", "Екатеринбург"), ("ЧЕЛЯБИНСК", "Челябинск"), ("ПЕРМ", "Пермь"),
        ("НОВОСИБИРСК", "Новосибирск"), ("КРАСНОЯРСК", "Красноярск"), ("КЕМЕРОВО", "Кемерово"),
    ]
    for key, canon in rules:
        if key in n:
            return canon
    return normalize_text(name)


def is_relevant_warehouse(name: Any) -> bool:
    n = clean_warehouse_name(name)
    if not n:
        return False
    if any(k in n for k in WAREHOUSE_EXCLUDE_KEYWORDS):
        return False
    return True


def classify_entry_channel(section: Any, point: Any = "") -> str:
    t = norm_key(f"{section} {point}")
    if "поиск" in t or "каталог" in t:
        return "Поиск/Каталог"
    if "карточ" in t or "полк" in t or "рекомен" in t:
        return "Карточка товара / полки"
    if "реклам" in t:
        return "Реклама / прочее"
    if "внеш" in t:
        return "Внешние переходы"
    return "Другие точки входа"


def pct_gap(fact: Any, target: Any) -> float:
    f = to_number(fact)
    t = to_number(target)
    if pd.isna(f) or pd.isna(t) or t == 0:
        return np.nan
    return (f / t - 1) * 100



@dataclass
class Diagnostics:
    rows: List[Dict[str, Any]] = field(default_factory=list)

    def add(self, level: str, source: str, message: str, details: Any = "") -> None:
        self.rows.append({
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "level": level,
            "source": source,
            "message": message,
            "details": normalize_text(details),
        })

    def frame(self) -> pd.DataFrame:
        return pd.DataFrame(self.rows, columns=["timestamp", "level", "source", "message", "details"])


# ------------------------- storage -------------------------
class Storage:
    is_s3 = False
    def list_files(self, prefix: str) -> List[str]:
        raise NotImplementedError
    def read_bytes(self, key: str) -> bytes:
        raise NotImplementedError
    def write_bytes(self, key: str, data: bytes) -> None:
        raise NotImplementedError
    def delete_file(self, key: str) -> None:
        raise NotImplementedError
    def exists(self, key: str) -> bool:
        raise NotImplementedError


class LocalStorage(Storage):
    def __init__(self, root: str):
        self.root = Path(root)
    def _full(self, key: str) -> Path:
        return self.root / key
    def list_files(self, prefix: str) -> List[str]:
        prefix = prefix.replace("\\", "/").strip("/")
        start = self._full(prefix)
        base = start if start.is_dir() else start.parent
        if not base.exists():
            return []
        out = []
        for p in base.rglob("*"):
            if p.is_file():
                rel = str(p.relative_to(self.root)).replace("\\", "/")
                if rel.startswith(prefix) and not Path(rel).name.startswith("~$"):
                    out.append(rel)
        return sorted(out)
    def read_bytes(self, key: str) -> bytes:
        return self._full(key).read_bytes()
    def write_bytes(self, key: str, data: bytes) -> None:
        p = self._full(key)
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_bytes(data)
    def delete_file(self, key: str) -> None:
        p = self._full(key)
        try:
            if p.exists() and p.is_file():
                p.unlink()
        except Exception as exc:
            log(f"WARN cleanup local delete failed: {key}: {exc}")
    def exists(self, key: str) -> bool:
        return self._full(key).exists()


class S3Storage(Storage):
    is_s3 = True
    def __init__(self, bucket: str, access_key: str, secret_key: str, endpoint_url: str):
        self.bucket = bucket
        self.client = boto3.client(
            "s3",
            endpoint_url=endpoint_url,
            aws_access_key_id=access_key,
            aws_secret_access_key=secret_key,
        )
    def list_files(self, prefix: str) -> List[str]:
        out = []
        token = None
        while True:
            kwargs = {"Bucket": self.bucket, "Prefix": prefix}
            if token:
                kwargs["ContinuationToken"] = token
            resp = self.client.list_objects_v2(**kwargs)
            for item in resp.get("Contents", []):
                key = item.get("Key", "")
                if key and not key.endswith("/") and not Path(key).name.startswith("~$"):
                    out.append(key)
            if not resp.get("IsTruncated"):
                break
            token = resp.get("NextContinuationToken")
        return sorted(out)
    def read_bytes(self, key: str) -> bytes:
        return self.client.get_object(Bucket=self.bucket, Key=key)["Body"].read()
    def write_bytes(self, key: str, data: bytes) -> None:
        self.client.put_object(Bucket=self.bucket, Key=key, Body=data)
    def delete_file(self, key: str) -> None:
        self.client.delete_object(Bucket=self.bucket, Key=key)
    def exists(self, key: str) -> bool:
        try:
            self.client.head_object(Bucket=self.bucket, Key=key)
            return True
        except Exception:
            return False


def make_storage(root: str) -> Storage:
    bucket = os.getenv("YC_BUCKET_NAME", "").strip()
    access = os.getenv("YC_ACCESS_KEY_ID", "").strip()
    secret = os.getenv("YC_SECRET_ACCESS_KEY", "").strip()
    endpoint = os.getenv("YC_ENDPOINT_URL", "").strip() or "https://storage.yandexcloud.net"
    # FIX_MARKER_20260525_ENDPOINT_FALLBACK: empty YC_ENDPOINT_URL must not break boto3.
    if bucket and access and secret:
        log(f"Storage: Yandex Object Storage bucket={bucket}; endpoint={endpoint}")
        return S3Storage(bucket, access, secret, endpoint)
    log(f"Storage: local root={Path(root).resolve()}")
    return LocalStorage(root)


# ------------------------- loader -------------------------
@dataclass
class DataPack:
    orders: pd.DataFrame
    funnel: pd.DataFrame
    ads_daily: pd.DataFrame
    ads_raw: pd.DataFrame
    ads_category: pd.DataFrame
    campaigns: pd.DataFrame
    search_queries: pd.DataFrame
    entry_points: pd.DataFrame
    stock: pd.DataFrame
    abc_weekly: pd.DataFrame
    abc_monthly: pd.DataFrame
    economics: pd.DataFrame
    latest_day: pd.Timestamp
    diagnostics: Diagnostics


class Loader:
    def __init__(self, storage: Storage, reports_root: str, store: str, diagnostics: Diagnostics):
        self.storage = storage
        self.reports_root = reports_root.strip("/")
        self.store = store
        self.diag = diagnostics
        self.current_week_only = os.getenv("WB_CURRENT_WEEK_ONLY", "0").strip().lower() in {"1", "true", "yes", "y"}
        target_raw = os.getenv("WB_DAILY_TARGET_DATE", "").strip()
        if target_raw:
            self.daily_target_day = pd.Timestamp(target_raw).normalize()
        else:
            # GitHub run at 12:00 MSK should take yesterday as the latest completed day.
            self.daily_target_day = pd.Timestamp(datetime.utcnow().date()).normalize() - pd.Timedelta(days=1)

    def _filter_current_week_files(self, files: List[str], keep_unparsed: bool = False, fallback_tail: int = 1) -> List[str]:
        if not self.current_week_only:
            return files
        target = self.daily_target_day
        week_start = target - pd.Timedelta(days=int(target.weekday()))
        week_end = week_start + pd.Timedelta(days=6)
        selected, unparsed = [], []
        for key in files:
            start, end = parse_period_from_name(Path(key).name)
            if start is None or end is None:
                unparsed.append(key)
                continue
            start = pd.Timestamp(start).normalize(); end = pd.Timestamp(end).normalize()
            if start <= week_end and end >= week_start:
                selected.append(key)
        if selected:
            log(f"current_week_only: selected {len(selected)}/{len(files)} files for week {week_start.date()}..{week_end.date()}")
            return sorted(set(selected))
        if keep_unparsed and unparsed:
            log(f"current_week_only: no dated files found, using {min(fallback_tail, len(unparsed))} unparsed/latest files")
            return sorted(set(unparsed))[-fallback_tail:]
        return sorted(set(files))[-fallback_tail:]

    def path(self, *parts: str) -> str:
        return "/".join([self.reports_root, *parts]).replace("//", "/")

    def list_reports(self, *parts: str) -> List[str]:
        prefix = self.path(*parts)
        files = self.storage.list_files(prefix)
        return [f for f in files if f.lower().endswith((".xlsx", ".xlsm", ".zip")) and not Path(f).name.startswith("~$")]

    def _log(self, name: str, df: pd.DataFrame, date_col: Optional[str] = None) -> None:
        if date_col and not df.empty and date_col in df.columns:
            mn, mx = pd.to_datetime(df[date_col], errors="coerce").min(), pd.to_datetime(df[date_col], errors="coerce").max()
            log(f"{name}: rows={len(df):,}, dates={mn.date() if pd.notna(mn) else '-'}..{mx.date() if pd.notna(mx) else '-'}")
        else:
            log(f"{name}: rows={len(df):,}")

    def _read_candidates(self, paths: Iterable[str]) -> Iterable[Tuple[str, bytes]]:
        for key in sorted(set(paths)):
            try:
                yield key, self.storage.read_bytes(key)
            except Exception as exc:
                self.diag.add("ERROR", "read", f"Не удалось прочитать {key}", exc)

    def load_orders(self) -> pd.DataFrame:
        """Load WB orders as the only source of orders/order_sum.

        Hard business rule 2026-05-27:
        - order_sum is taken only from the Orders report data; if WB does not export a separate
          order_sum column, the script uses priceWithDisc from the same Orders file as sales/order amount;
          finishedPrice is buyer price and is not used for order_sum;
        - technical zero cancellation dates like 0001-01-01 are not treated as cancellations;
        - funnel is not allowed to replace orders/order_sum later.
        """
        files = self.list_reports("Заказы", self.store, "Недельные")
        files = self._filter_current_week_files(files, keep_unparsed=False, fallback_tail=1)
        frames = []

        def _cancel_mask(src: pd.DataFrame) -> pd.Series:
            mask = pd.Series(False, index=src.index)
            # Boolean/text cancel flags.
            if "is_cancel" in src.columns:
                flag = src["is_cancel"].map(normalize_text).str.lower().str.replace("ё", "е")
                mask = mask | flag.isin({"1", "true", "да", "yes", "y", "отменен", "отмена", "отменено"})
                mask = mask | flag.str.contains("отмен|cancel", na=False)
            # Cancellation date: only real non-empty business dates count.
            # WB active rows often contain technical dates like 0001-01-01T00:00:00;
            # these must not exclude the whole Orders file.
            if "cancel_date" in src.columns:
                raw = src["cancel_date"].map(normalize_text).str.lower()
                raw_bad = raw.isin({"", "nan", "none", "null", "0"}) | raw.str.contains(r"^0+|0001[-/.]01[-/.]01|1970[-/.]01[-/.]01", na=False)
                dt = date_series(src["cancel_date"])
                real_cancel_date = dt.notna() & (dt >= pd.Timestamp("2000-01-01")) & ~raw_bad
                mask = mask | real_cancel_date
            # Status field fallback.
            if "order_status" in src.columns:
                st = src["order_status"].map(normalize_text).str.lower().str.replace("ё", "е")
                mask = mask | st.str.contains("отмен|cancel", na=False)
            return mask.fillna(False)

        for key, data in self._read_candidates(files):
            try:
                df = read_excel_table(data, "Заказы")
                if df.empty:
                    continue
                df = add_alias_columns(df)
                rows_before = len(df)
                cancel_mask = _cancel_mask(df)
                df_active = df.loc[~cancel_mask].copy()
                rows_cancelled = int(cancel_mask.sum())

                orders_series = num_series(get_col(df_active, "orders"))
                if orders_series.isna().all():
                    orders_series = pd.Series(1.0, index=df_active.index)
                orders_series = orders_series.fillna(1.0)

                finished_price_series = num_series(get_col(df_active, "finished_price"))
                price_with_disc_series = num_series(get_col(df_active, "price_with_disc"))
                order_sum_series = num_series(get_col(df_active, "order_sum"))
                order_sum_col = _source_column_name(df_active, "order_sum")
                if order_sum_series.isna().all():
                    # Still Orders-only: fallback is from the same WB Orders file, never from Funnel.
                    # Business rule: priceWithDisc is the sale price / order amount.
                    # finishedPrice is buyer price and must not be used as order_sum.
                    fallback_price = price_with_disc_series.copy()
                    fallback_col = _source_column_name(df_active, "price_with_disc")
                    if fallback_price.isna().all():
                        # Last-resort diagnostic fallback only when WB Orders has no priceWithDisc at all.
                        fallback_price = finished_price_series.copy()
                        fallback_col = _source_column_name(df_active, "finished_price")
                        self.diag.add("WARN", "orders", f"В файле заказов нет priceWithDisc: {key}", "использую finishedPrice только как аварийный fallback; проверь структуру Orders")
                    if not fallback_price.isna().all():
                        order_sum_series = fallback_price * orders_series
                        order_sum_col = f"{fallback_col or 'price_with_disc'} × orders (Orders file)"
                        self.diag.add("WARN", "orders", f"В файле заказов нет отдельной колонки суммы заказов: {key}", f"использую {order_sum_col}; источник всё равно Orders")
                        log(f"orders: no explicit order_sum in {Path(key).name}; use {order_sum_col}")
                    else:
                        order_sum_col = ""
                        self.diag.add("ERROR", "orders", f"В файле заказов не распознана ни сумма, ни цена продажи priceWithDisc: {key}", "продажи из воронки запрещены")
                        log(f"orders: ERROR no order_sum/priceWithDisc column recognized in {Path(key).name}; Funnel fallback is forbidden")

                out = pd.DataFrame({
                    "day": date_series(get_col(df_active, "day")),
                    "nm_id": num_series(get_col(df_active, "nm_id")),
                    "supplier_article": get_col(df_active, "supplier_article").map(clean_article),
                    "subject": get_col(df_active, "subject").map(canonical_subject),
                    "warehouse": get_col(df_active, "warehouse").map(normalize_text),
                    "orders": orders_series,
                    "order_sum": order_sum_series,
                    "finished_price": finished_price_series,
                    "price_with_disc": price_with_disc_series,
                    "spp": num_series(get_col(df_active, "spp")),
                    "source_file": key,
                    "source_sheet": df.attrs.get("source_sheet", ""),
                    "source_order_sum_col": order_sum_col,
                })
                # Missing sums become 0 and are visible in diagnostics; sales never come from Funnel.
                missing_sum_rows = int(out["order_sum"].isna().sum())
                out["order_sum"] = out["order_sum"].fillna(0.0)
                out = out[out["day"].notna()].copy()
                frames.append(out)
                log(
                    "orders: file=%s sheet=%s rows_before=%s cancelled_excluded=%s rows_after=%s order_sum_col=%s sum=%s" % (
                        Path(key).name,
                        df.attrs.get("source_sheet", ""),
                        rows_before,
                        rows_cancelled,
                        len(out),
                        order_sum_col or "NOT_FOUND",
                        f"{pd.to_numeric(out['order_sum'], errors='coerce').fillna(0).sum():.2f}",
                    )
                )
                if missing_sum_rows:
                    self.diag.add("WARN", "orders", f"В файле заказов есть строки без суммы заказов: {key}", f"rows={missing_sum_rows}; fallback запрещён")
            except Exception as exc:
                self.diag.add("ERROR", "orders", f"Не прочитан файл заказов {key}", exc)
        out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        self._log("orders", out, "day")
        return out

    def load_funnel(self) -> pd.DataFrame:
        files = []
        direct = self.path("Воронка продаж", self.store, "Воронка продаж.xlsx")
        if self.storage.exists(direct):
            files.append(direct)
        files += self.list_reports("Воронка продаж", self.store)
        files += self.list_reports("Воронка продаж")
        files = self._filter_current_week_files(files, keep_unparsed=True, fallback_tail=1)
        frames = []
        for key, data in self._read_candidates(files):
            try:
                period_start, period_end = parse_period_from_name(Path(key).name)
                preferred = "Воронка продаж"
                df = read_excel_table(data, preferred)
                if df.empty:
                    continue
                day = date_series(get_col(df, "day"))
                if day.isna().all() and period_end is not None:
                    day = pd.Series([period_end] * len(df))
                out = pd.DataFrame({
                    "day": day,
                    "nm_id": num_series(get_col(df, "nm_id")),
                    "supplier_article": get_col(df, "supplier_article").map(clean_article),
                    "subject": get_col(df, "subject").map(canonical_subject),
                    "orders": num_series(get_col(df, "orders")),
                    "order_sum": num_series(get_col(df, "order_sum")),
                    "open_cards": num_series(get_col(df, "open_cards")),
                    "add_to_cart": num_series(get_col(df, "add_to_cart")),
                    "cart_conv_pct": num_series(get_col(df, "cart_conv")),
                    "order_conv_pct": num_series(get_col(df, "order_conv")),
                    "buyouts_count": num_series(get_col(df, "buyouts_count")),
                    "buyout_sum": num_series(get_col(df, "buyout_sum")),
                    "cancels_count": num_series(get_col(df, "cancels_count")),
                    "finished_price": num_series(get_col(df, "finished_price")),
                    "spp": num_series(get_col(df, "spp")),
                    "source_file": key,
                })
                out = out[out["nm_id"].notna()]
                out = out[out["day"].notna()]
                frames.append(out)
            except Exception as exc:
                self.diag.add("ERROR", "funnel", f"Не прочитана воронка {key}", exc)
        out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        self._log("funnel", out, "day")
        return out

    def load_ads(self) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        files = self.list_reports("Реклама", self.store, "Недельные")
        consolidated = self.path("Реклама", self.store, "Анализ рекламы.xlsx")
        if self.storage.exists(consolidated):
            files.append(consolidated)
        files = self._filter_current_week_files(files, keep_unparsed=True, fallback_tail=1)
        raw_frames, campaign_frames = [], []
        for key, data in self._read_candidates(files):
            try:
                book = pd.ExcelFile(io.BytesIO(unwrap_excel_bytes(data)))
                if "Список_кампаний" in book.sheet_names:
                    camp = read_excel_table(data, "Список_кампаний")
                    if not camp.empty:
                        cdf = pd.DataFrame({
                            "campaign_id": get_col(camp, "campaign_id").map(lambda x: str(int(x)) if pd.notna(to_number(x)) else normalize_text(x)),
                            "nm_id": num_series(get_col(camp, "nm_id")),
                            "subject": get_col(camp, "subject").map(normalize_text),
                            "bid_type_raw": get_col(camp, "bid_type").map(normalize_text),
                            "source_file": key,
                        })
                        cdf["ad_type"] = cdf["bid_type_raw"].map(classify_ad_type)
                        campaign_frames.append(cdf[cdf["campaign_id"].ne("")])
                sheets = ["Статистика_Ежедневно"] if "Статистика_Ежедневно" in book.sheet_names else book.sheet_names
                for sheet in sheets:
                    df = read_excel_table(data, sheet)
                    if df.empty or get_col(df, "spend").isna().all():
                        continue
                    out = pd.DataFrame({
                        "day": date_series(get_col(df, "day")),
                        "campaign_id": get_col(df, "campaign_id").map(lambda x: str(int(x)) if pd.notna(to_number(x)) else normalize_text(x)),
                        "nm_id": num_series(get_col(df, "nm_id")),
                        "subject": get_col(df, "subject").map(canonical_subject),
                        "impressions": num_series(get_col(df, "impressions")).fillna(0),
                        "clicks": num_series(get_col(df, "clicks")).fillna(0),
                        "orders": num_series(get_col(df, "ad_orders")).fillna(0),
                        "order_sum": num_series(get_col(df, "ad_order_sum")).fillna(0),
                        "spend": num_series(get_col(df, "spend")).fillna(0),
                        "ctr_pct_src": num_series(get_col(df, "ctr")),
                        "cpc_src": num_series(get_col(df, "cpc")),
                        "cr_pct_src": num_series(get_col(df, "cr")),
                        "drr_pct_src": num_series(get_col(df, "drr")),
                        "source_file": key,
                        "source_sheet": sheet,
                    })
                    raw_frames.append(out[out["day"].notna() & out["nm_id"].notna()])
            except Exception as exc:
                self.diag.add("ERROR", "ads", f"Не прочитана реклама {key}", exc)
        raw = pd.concat(raw_frames, ignore_index=True) if raw_frames else pd.DataFrame()
        if not raw.empty:
            before_dedup = len(raw)
            dedup_cols = [c for c in ["day", "campaign_id", "nm_id", "impressions", "clicks", "orders", "order_sum", "spend"] if c in raw.columns]
            raw = raw.drop_duplicates(subset=dedup_cols).copy()
            if len(raw) != before_dedup:
                log(f"ads: removed duplicated source rows {before_dedup - len(raw):,}; weekly/consolidated overlap")
        campaigns = pd.concat(campaign_frames, ignore_index=True) if campaign_frames else pd.DataFrame(columns=["campaign_id", "nm_id", "subject", "ad_type"])
        if not campaigns.empty:
            before_camp_dedup = len(campaigns)
            campaigns = campaigns.drop_duplicates(["campaign_id", "nm_id", "ad_type", "bid_type_raw"]).copy()
            if len(campaigns) != before_camp_dedup:
                log(f"ads: removed duplicated campaign rows {before_camp_dedup - len(campaigns):,}")
        if not raw.empty:
            if not campaigns.empty:
                cmap = campaigns.drop_duplicates(["campaign_id", "nm_id"])[["campaign_id", "nm_id", "ad_type", "bid_type_raw"]]
                raw = raw.merge(cmap, on=["campaign_id", "nm_id"], how="left")
            raw["ad_type"] = raw.get("ad_type", "unknown")
            raw["ad_type"] = raw["ad_type"].fillna("unknown").replace("", "unknown")
            raw["ctr_pct"] = np.where(raw["impressions"] > 0, raw["clicks"] / raw["impressions"] * 100, raw["ctr_pct_src"])
            raw["cpc"] = np.where(raw["clicks"] > 0, raw["spend"] / raw["clicks"], raw["cpc_src"])
            raw["cr_pct"] = np.where(raw["clicks"] > 0, raw["orders"] / raw["clicks"] * 100, raw["cr_pct_src"])
            raw["drr_pct"] = np.where(raw["order_sum"] > 0, raw["spend"] / raw["order_sum"] * 100, raw["drr_pct_src"])
            daily = raw.groupby(["day", "nm_id", "ad_type"], dropna=False, as_index=False).agg(
                impressions=("impressions", "sum"), clicks=("clicks", "sum"), orders=("orders", "sum"),
                order_sum=("order_sum", "sum"), spend=("spend", "sum"),
            )
            daily["ctr_pct"] = np.where(daily["impressions"] > 0, daily["clicks"] / daily["impressions"] * 100, np.nan)
            daily["cpc"] = np.where(daily["clicks"] > 0, daily["spend"] / daily["clicks"], np.nan)
            daily["cr_pct"] = np.where(daily["clicks"] > 0, daily["orders"] / daily["clicks"] * 100, np.nan)
            daily["drr_pct"] = np.where(daily["order_sum"] > 0, daily["spend"] / daily["order_sum"] * 100, np.nan)
        else:
            daily = pd.DataFrame()
        self._log("ads_raw", raw, "day")
        self._log("ads_daily", daily, "day")
        return raw, daily, campaigns

    def load_ads_category(self) -> pd.DataFrame:
        """Read category-level advertising spend from WB advertising report.

        This is the source of truth for current-week category/current overview spend.
        Raw campaign/article sheets can duplicate spend across articles, so they are not used
        for the first PDF page when this category sheet exists.
        """
        files = self.list_reports("Реклама", self.store, "Недельные")
        consolidated = self.path("Реклама", self.store, "Анализ рекламы.xlsx")
        if self.storage.exists(consolidated):
            files.append(consolidated)
        files = self._filter_current_week_files(files, keep_unparsed=True, fallback_tail=1)
        frames = []
        wanted_sheets = {"отчет_по_категории", "отчет по категории", "отчет_по_категории_итог", "отчет по категории итог", "отчёт_по_категории", "отчёт по категории", "отчёт_по_категории_итог", "отчёт по категории итог"}
        for key, data in self._read_candidates(files):
            try:
                raw_bytes = unwrap_excel_bytes(data)
                book = pd.ExcelFile(io.BytesIO(raw_bytes))
                period_start, period_end = parse_period_from_name(Path(key).name)
                for sheet in book.sheet_names:
                    if norm_key(sheet).replace(" ", "_") not in {s.replace(" ", "_") for s in wanted_sheets} and norm_key(sheet) not in wanted_sheets:
                        continue
                    df = read_excel_table(raw_bytes, sheet)
                    if df.empty:
                        continue
                    day = date_series(get_col(df, "day"))
                    if day.isna().all() and period_end is not None:
                        day = pd.Series([period_end] * len(df), index=df.index)
                    out = pd.DataFrame({
                        "day": day,
                        "subject": get_col(df, "subject").map(canonical_subject),
                        "spend": num_series(get_col(df, "spend")).fillna(0.0),
                        "clicks": num_series(get_col(df, "clicks")).fillna(0.0),
                        "impressions": num_series(get_col(df, "impressions")).fillna(0.0),
                        "orders": num_series(get_col(df, "ad_orders")).fillna(0.0),
                        "order_sum": num_series(get_col(df, "ad_order_sum")).fillna(0.0),
                        "source_file": key,
                        "source_sheet": sheet,
                    })
                    out = out[out["day"].notna()].copy()
                    out = out[out["subject"].isin(TARGET_SUBJECTS)].copy()
                    if out.empty:
                        continue
                    frames.append(out)
                    log(f"ads_category: file={Path(key).name} sheet={sheet} rows={len(out):,} spend={out['spend'].sum():.2f}")
            except Exception as exc:
                self.diag.add("ERROR", "ads_category", f"Не прочитан категорийный лист рекламы {key}", exc)
        out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        if not out.empty:
            out = out.groupby(["day", "subject"], dropna=False, as_index=False).agg(
                spend=("spend", "sum"),
                clicks=("clicks", "sum"),
                impressions=("impressions", "sum"),
                orders=("orders", "sum"),
                order_sum=("order_sum", "sum"),
                source_file=("source_file", lambda s: "; ".join(sorted(set(map(str, s)))[:3])),
                source_sheet=("source_sheet", lambda s: "; ".join(sorted(set(map(str, s)))[:3])),
            )
        self._log("ads_category", out, "day")
        return out

    def load_search_queries(self, latest_day: Optional[pd.Timestamp] = None) -> pd.DataFrame:
        files_all = self.list_reports("Поисковые запросы", self.store, "Недельные")
        files = filter_recent_report_files(files_all, latest_day, lookback_days=110, keep_unknown=False)
        max_files = int(os.getenv("WB_MAX_SEARCH_QUERY_FILES", "8"))
        files = limit_recent_report_files(files, max_files)
        log(f"search_queries: start, files_all={len(files_all)}, files_to_read={len(files)}")
        frames = []
        for idx, (key, data) in enumerate(self._read_candidates(files), start=1):
            log(f"search_queries: reading {idx}/{len(files)} {Path(key).name}")
            try:
                df = read_excel_table(data, "Позиции по Ключам")
                if df.empty:
                    continue
                out = pd.DataFrame({
                    "day": date_series(get_col(df, "day")),
                    "nm_id": num_series(get_col(df, "nm_id")),
                    "supplier_article": get_col(df, "supplier_article").map(clean_article),
                    "subject": get_col(df, "subject").map(canonical_subject),
                    "search_query": get_col(df, "search_query").map(normalize_text),
                    "filter": get_col(df, "filter").map(normalize_text),
                    "frequency": num_series(get_col(df, "frequency")),
                    "median_position": num_series(get_col(df, "median_position")),
                    "avg_position": num_series(get_col(df, "avg_position")),
                    "transitions": num_series(get_col(df, "open_cards")).fillna(num_series(get_col(df, "clicks"))).fillna(0),
                    "add_to_cart": num_series(get_col(df, "add_to_cart")).fillna(0),
                    "orders": num_series(get_col(df, "orders")).fillna(0),
                    "cart_conv_pct": num_series(get_col(df, "cart_conv")),
                    "order_conv_pct": num_series(get_col(df, "order_conv")),
                    "rating_card": num_series(get_col(df, "rating_card")),
                    "rating_reviews": num_series(get_col(df, "rating_reviews")),
                    "visibility_pct": num_series(get_col(df, "visibility")),
                    "source_file": key,
                })
                out = out[out["day"].notna() & out["search_query"].ne("")]
                # Deduplicate same query due to filters: frequency once, commercial metrics summed.
                group_cols = ["day", "nm_id", "supplier_article", "subject", "search_query"]
                agg = out.groupby(group_cols, dropna=False, as_index=False).agg(
                    frequency=("frequency", "max"),
                    transitions=("transitions", "sum"), add_to_cart=("add_to_cart", "sum"), orders=("orders", "sum"),
                    median_position=("median_position", "mean"), avg_position=("avg_position", "mean"),
                    cart_conv_pct=("cart_conv_pct", "mean"), order_conv_pct=("order_conv_pct", "mean"),
                    rating_card=("rating_card", "mean"), rating_reviews=("rating_reviews", "mean"),
                    visibility_pct=("visibility_pct", "mean"), source_file=("source_file", "first"),
                )
                agg["traffic_capture_pct"] = np.where(agg["frequency"] > 0, agg["transitions"] / agg["frequency"] * 100, np.nan)
                frames.append(agg)
            except Exception as exc:
                self.diag.add("ERROR", "search_queries", f"Не прочитан файл поисковых запросов {key}", exc)
        out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        self._log("search_queries", out, "day")
        return out

    def load_entry_points(self, latest_day: Optional[pd.Timestamp] = None) -> pd.DataFrame:
        files_all = []
        files_all += self.list_reports("Точки входа", self.store)
        files_all += self.list_reports("Портрет покупателя", self.store)
        # Broad fallback is expensive on S3, so it is used only when targeted folders are empty.
        if not files_all:
            files_all += [f for f in self.storage.list_files(self.reports_root) if "Точки входа" in f and f.lower().endswith((".xlsx", ".zip"))]
        files = filter_recent_report_files(files_all, latest_day, lookback_days=110, keep_unknown=True)
        max_files = int(os.getenv("WB_MAX_ENTRY_POINT_FILES", "8"))
        files = limit_recent_report_files(files, max_files)
        log(f"entry_points: start, files_all={len(files_all)}, files_to_read={len(files)}")
        frames = []
        for idx, (key, data) in enumerate(self._read_candidates(files), start=1):
            log(f"entry_points: reading {idx}/{len(files)} {Path(key).name}")
            try:
                _, period_end = parse_period_from_name(Path(key).name)
                df = read_excel_table(data, "Детализация по артикулам")
                if df.empty or get_col(df, "entry_section").isna().all():
                    continue
                day = date_series(get_col(df, "day")) if "day" in df.columns else pd.Series([period_end] * len(df))
                if day.isna().all() and period_end is not None:
                    day = pd.Series([period_end] * len(df))
                out = pd.DataFrame({
                    "day": day,
                    "entry_section": get_col(df, "entry_section").map(normalize_text),
                    "entry_point": get_col(df, "entry_point").map(normalize_text),
                    "nm_id": num_series(get_col(df, "nm_id")),
                    "supplier_article": get_col(df, "supplier_article").map(clean_article),
                    "subject": get_col(df, "subject").map(canonical_subject),
                    "impressions": num_series(get_col(df, "impressions")).fillna(0),
                    "transitions": num_series(get_col(df, "open_cards")).fillna(0),
                    "ctr_pct": num_series(get_col(df, "ctr")),
                    "add_to_cart": num_series(get_col(df, "add_to_cart")).fillna(0),
                    "cart_conv_pct": num_series(get_col(df, "cart_conv")),
                    "orders": num_series(get_col(df, "orders")).fillna(0),
                    "order_conv_pct": num_series(get_col(df, "order_conv")),
                    "source_file": key,
                })
                out = out[out["nm_id"].notna() | out["supplier_article"].ne("")]
                frames.append(out)
            except Exception as exc:
                self.diag.add("ERROR", "entry_points", f"Не прочитаны точки входа {key}", exc)
        out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        self._log("entry_points", out, "day")
        return out

    def load_stock(self, latest_day: Optional[pd.Timestamp] = None) -> pd.DataFrame:
        files_all = []
        for parts in [("Остатки", self.store), ("Остатки", self.store, "Недельные"), ("Остатки",), ("Остатки и товары в пути", self.store), ("Остатки и товары в пути",)]:
            files_all += self.list_reports(*parts)
        files = filter_recent_report_files(files_all, latest_day, lookback_days=110, keep_unknown=True)
        max_files = int(os.getenv("WB_MAX_STOCK_FILES", "10"))
        files = limit_recent_report_files(files, max_files)
        log(f"stock: start, files_all={len(files_all)}, files_to_read={len(files)}")
        frames = []
        for idx, (key, data) in enumerate(self._read_candidates(files), start=1):
            log(f"stock: reading {idx}/{len(files)} {Path(key).name}")
            try:
                _, period_end = parse_period_from_name(Path(key).name)
                df = read_excel_table(data)
                if df.empty:
                    continue
                day = date_series(get_col(df, "day"))
                if day.isna().all() and period_end is not None:
                    day = pd.Series([period_end] * len(df))
                out = pd.DataFrame({
                    "day": day,
                    "nm_id": num_series(get_col(df, "nm_id")),
                    "supplier_article": get_col(df, "supplier_article").map(clean_article),
                    "subject": get_col(df, "subject").map(canonical_subject),
                    "warehouse": get_col(df, "warehouse").map(normalize_text),
                    "stock": num_series(get_col(df, "stock")).fillna(0),
                    "source_file": key,
                })
                out = out[(out["nm_id"].notna() | out["supplier_article"].ne("")) & out["warehouse"].ne("")]
                frames.append(out)
            except Exception as exc:
                self.diag.add("ERROR", "stock", f"Не прочитаны остатки {key}", exc)
        out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        self._log("stock", out, "day")
        return out

    def load_abc(self, current_year: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
        files = self.list_reports("ABC")
        weekly_frames, monthly_frames = [], []
        for key, data in self._read_candidates(files):
            if "abc" not in key.lower():
                continue
            start, end = parse_period_from_name(Path(key).name)
            if start is None or end is None:
                continue
            try:
                df = read_excel_table(data)
                if df.empty:
                    continue
                source_sheet = df.attrs.get("source_sheet", "")
                header_row_excel = df.attrs.get("header_row_excel", "")
                source_rows = pd.Series(df.index, index=df.index).map(lambda i: int(i) + int(df.attrs.get("header_row_0based", 0) or 0) + 2)
                out = pd.DataFrame({
                    "period_start": start,
                    "period_end": end,
                    "week_code": week_code(start),
                    "week_label": f"{start.strftime('%d.%m')}-{end.strftime('%d.%m')}",
                    "month_key": month_key(start),
                    "nm_id": num_series(get_col(df, "nm_id")),
                    "supplier_article": get_col(df, "supplier_article").map(clean_article),
                    "subject": get_col(df, "subject").map(canonical_subject),
                    "gross_profit": num_series(get_col(df, "gross_profit")).fillna(0),
                    "gross_revenue": num_series(get_col(df, "gross_revenue")).fillna(0),
                    "orders": num_series(get_col(df, "orders")).fillna(0),
                    "abc_margin_pct": num_series(get_col(df, "margin_pct")),
                    "abc_drr_pct": num_series(get_col(df, "drr")),
                    "abc_commission_amount": num_series(get_col(df, "commission_amount")).fillna(0),
                    "abc_acquiring_amount": num_series(get_col(df, "acquiring_amount")).fillna(0),
                    "source_file": key,
                    "source_sheet": source_sheet,
                    "header_row_excel": header_row_excel,
                    "source_row_excel": source_rows,
                    "cell_gross_profit": [_source_cell_ref(df, "gross_profit", i) for i in df.index],
                    "cell_gross_revenue": [_source_cell_ref(df, "gross_revenue", i) for i in df.index],
                    "cell_orders": [_source_cell_ref(df, "orders", i) for i in df.index],
                    "cell_margin_pct": [_source_cell_ref(df, "margin_pct", i) for i in df.index],
                    "cell_drr_pct": [_source_cell_ref(df, "drr", i) for i in df.index],
                })
                out["product"] = out["supplier_article"].map(product_code)
                if is_month_file(start, end) and start.year == current_year:
                    monthly_frames.append(out)
                else:
                    weekly_frames.append(out)
            except Exception as exc:
                self.diag.add("ERROR", "abc", f"Не прочитан ABC {key}", exc)
        weekly = pd.concat(weekly_frames, ignore_index=True) if weekly_frames else pd.DataFrame()
        monthly = pd.concat(monthly_frames, ignore_index=True) if monthly_frames else pd.DataFrame()
        self._log("abc_weekly", weekly, "period_start")
        self._log("abc_monthly", monthly, "period_start")
        return weekly, monthly

    def load_economics(self) -> pd.DataFrame:
        candidates = [
            self.path("Финансовые показатели", self.store, "Экономика.xlsx"),
            self.path("Финансовые показатели", self.store, "Недельные", "Экономика.xlsx"),
        ]
        files = [f for f in candidates if self.storage.exists(f)]
        frames = []
        for key, data in self._read_candidates(files):
            try:
                df = read_excel_table(data, "Юнит экономика")
                if df.empty:
                    continue
                out = pd.DataFrame({
                    "week_code": get_col(df, "week").map(normalize_text),
                    "nm_id": num_series(get_col(df, "nm_id")),
                    "supplier_article": get_col(df, "supplier_article").map(clean_article),
                    "subject": get_col(df, "subject").map(canonical_subject),
                    "commission_pct": num_series(get_col(df, "commission_pct")),
                    "acquiring_pct": num_series(get_col(df, "acquiring_pct")),
                    "logistics_direct": num_series(get_col(df, "logistics_direct")),
                    "logistics_return": num_series(get_col(df, "logistics_return")),
                    "storage": num_series(get_col(df, "storage")),
                    "other_costs": num_series(get_col(df, "other_costs")),
                    "cost": num_series(get_col(df, "cost")),
                    "source_file": key,
                })
                out["product"] = out["supplier_article"].map(product_code)
                frames.append(out)
            except Exception as exc:
                self.diag.add("ERROR", "economics", f"Не прочитана экономика {key}", exc)
        out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        self._log("economics", out)
        return out

    def load_all(self) -> DataPack:
        orders = self.load_orders()
        funnel = self.load_funnel()
        ads_raw, ads_daily, campaigns = self.load_ads()
        ads_category = self.load_ads_category()

        # Preliminary latest date is known before heavy Stage 2 sources.
        # Use it to read only the recent 90-110 day window and avoid silent 20+ minute parsing of old files.
        pre_candidates = []
        for df, col in [(orders, "day"), (funnel, "day"), (ads_daily, "day")]:
            if not df.empty and col in df.columns:
                mx = pd.to_datetime(df[col], errors="coerce").max()
                if pd.notna(mx):
                    pre_candidates.append(pd.Timestamp(mx).normalize())
        preliminary_latest_day = max(pre_candidates) if pre_candidates else pd.Timestamp(datetime.today().date())
        log(f"preliminary_latest_day: {preliminary_latest_day.date()}")

        search_queries = self.load_search_queries(preliminary_latest_day)
        if self.current_week_only:
            entry_points = pd.DataFrame()
            stock = pd.DataFrame()
            economics = pd.DataFrame()
            log("current_week_only: skip entry_points/stock/economics/ABC; daily mode updates only current-week operational block")
        else:
            entry_points = self.load_entry_points(preliminary_latest_day)
            stock = self.load_stock(preliminary_latest_day)
            economics = self.load_economics()
        candidates = []
        for df, col in [(orders, "day"), (funnel, "day"), (ads_daily, "day"), (search_queries, "day"), (stock, "day")]:
            if not df.empty and col in df.columns:
                mx = pd.to_datetime(df[col], errors="coerce").max()
                if pd.notna(mx):
                    candidates.append(pd.Timestamp(mx).normalize())
        latest_day = max(candidates) if candidates else preliminary_latest_day
        if self.current_week_only:
            abc_weekly, abc_monthly = pd.DataFrame(), pd.DataFrame()
        else:
            abc_weekly, abc_monthly = self.load_abc(latest_day.year)
            if not abc_weekly.empty:
                latest_day = max(latest_day, pd.to_datetime(abc_weekly["period_end"], errors="coerce").max())
        return DataPack(
            orders=orders, funnel=funnel, ads_daily=ads_daily, ads_raw=ads_raw, ads_category=ads_category, campaigns=campaigns,
            search_queries=search_queries, entry_points=entry_points, stock=stock, abc_weekly=abc_weekly,
            abc_monthly=abc_monthly, economics=economics, latest_day=pd.Timestamp(latest_day).normalize(), diagnostics=self.diag,
        )


# ------------------------- analytics -------------------------
class AnalyticsBuilder:
    def __init__(self, pack: DataPack):
        self.pack = pack
        self.diag = pack.diagnostics
        self.latest_day = pack.latest_day
        self.cutoff_90 = self.latest_day - pd.Timedelta(days=89)
        # Last full week. If latest day is Sunday, current week is full; otherwise previous Mon-Sun.
        last_monday = self.latest_day - pd.Timedelta(days=int(self.latest_day.weekday()))
        if self.latest_day.weekday() == 6:
            self.week_start = last_monday
            self.week_end = self.latest_day
        else:
            self.week_start = last_monday - pd.Timedelta(days=7)
            self.week_end = last_monday - pd.Timedelta(days=1)
        self.dictionary = self.build_dictionary()

    def build_dictionary(self) -> pd.DataFrame:
        frames = []
        for name in ["orders", "funnel", "ads_raw", "search_queries", "entry_points", "stock", "abc_weekly", "abc_monthly", "economics"]:
            df = getattr(self.pack, name)
            if df is None or df.empty:
                continue
            x = df.copy()
            for col in ["subject", "product", "supplier_article", "nm_id"]:
                if col not in x.columns:
                    x[col] = "" if col != "nm_id" else np.nan
            x["supplier_article"] = x["supplier_article"].map(clean_article)
            x["subject"] = x["subject"].map(canonical_subject)
            x["nm_id"] = num_series(x["nm_id"])
            x["product"] = x["supplier_article"].map(product_code).where(x["product"].map(normalize_text).eq(""), x["product"].map(normalize_text))
            x["product"] = x["product"].map(lambda v: normalize_text(v).upper().replace(" ", ""))
            x["source"] = name
            frames.append(x[["subject", "product", "supplier_article", "nm_id", "source"]])
        if not frames:
            return pd.DataFrame(columns=["subject", "product", "supplier_article", "nm_id", "source"])
        d = pd.concat(frames, ignore_index=True)
        d = d[d["subject"].isin(TARGET_SUBJECTS)]
        d = d[d["supplier_article"].ne("") & d["product"].ne("")]
        d = d[~d["supplier_article"].map(is_excluded_article)]
        d = d[d["product"].map(is_valid_product_code)]
        d = d[d.apply(lambda r: is_approved_product_subject(r.get("product"), r.get("subject")), axis=1)]
        d = d.drop_duplicates(["supplier_article", "nm_id"])
        log(f"dictionary: rows={len(d):,}, articles={d['supplier_article'].nunique():,}, nm_ids={d['nm_id'].nunique(dropna=True):,}")
        return d

    def enrich(self, df: pd.DataFrame, source: str = "") -> pd.DataFrame:
        """
        Add subject/product/supplier_article to source rows using nm_id dictionary.

        HARD FIX 2026-05-26: this function must never crash with
        KeyError: 'product' on advertising sources. Some sources arrive without
        product/supplier_article and some pandas merge/filter paths can leave
        the canonical columns in an unsafe state. We therefore build canonical
        Series first, remove any old/duplicate canonical columns, reinsert clean
        columns, and use local Series for all filters.
        """
        if df is None or df.empty:
            return pd.DataFrame(columns=list(df.columns) if isinstance(df, pd.DataFrame) else [])

        out = df.copy()

        def _series(frame: pd.DataFrame, name: str, default="") -> pd.Series:
            """Return a 1-D Series even if duplicate column labels exist."""
            if name in frame.columns:
                value = frame[name]
                if isinstance(value, pd.DataFrame):
                    value = value.iloc[:, 0]
                return value
            return pd.Series(default, index=frame.index, dtype="object")

        # Ensure merge keys exist before dictionary merge.
        if "nm_id" not in out.columns:
            out["nm_id"] = np.nan
        if "subject" not in out.columns:
            out["subject"] = ""
        if "product" not in out.columns:
            out["product"] = ""
        if "supplier_article" not in out.columns:
            out["supplier_article"] = ""

        if not self.dictionary.empty:
            d_nm = (
                self.dictionary.dropna(subset=["nm_id"])
                .drop_duplicates("nm_id")[["nm_id", "subject", "product", "supplier_article"]]
            )
            out = out.merge(d_nm, on="nm_id", how="left", suffixes=("", "_dict"))

        # Canonical values are calculated as local Series. Never use direct
        # out["product"] in filters because it can fail when the label is missing
        # or duplicated in rare pandas states.
        subject_raw = _series(out, "subject").map(canonical_subject)
        subject_dict = _series(out, "subject_dict").map(canonical_subject)
        subject_clean = subject_raw.where(subject_raw.ne(""), subject_dict)

        article_raw = _series(out, "supplier_article").map(clean_article)
        article_dict = _series(out, "supplier_article_dict").map(clean_article)
        article_clean = article_raw.where(article_raw.ne(""), article_dict)

        product_raw = _series(out, "product").map(lambda v: normalize_text(v).upper().replace(" ", ""))
        product_dict = _series(out, "product_dict").map(lambda v: normalize_text(v).upper().replace(" ", ""))
        product_clean = product_raw.where(product_raw.ne(""), product_dict)
        product_clean = product_clean.where(product_clean.ne(""), article_clean.map(product_code))
        product_clean = product_clean.map(lambda v: normalize_text(v).upper().replace(" ", ""))
        ref_subject = product_clean.map(approved_subject_for_product)
        # If the source subject is empty, use the approved reference category.
        # If the source subject is present, it must still match the approved pair below.
        subject_clean = subject_clean.where(subject_clean.ne(""), ref_subject)

        # Drop old canonical/dict columns, including duplicates, and reinsert clean canonical columns.
        drop_cols = [c for c in out.columns if str(c) in {"subject", "product", "supplier_article", "subject_dict", "product_dict", "supplier_article_dict"}]
        out = out.drop(columns=drop_cols, errors="ignore")
        out.insert(0, "subject", subject_clean.values)
        out.insert(1, "product", product_clean.values)
        out.insert(2, "supplier_article", article_clean.values)

        mask = (
            subject_clean.isin(TARGET_SUBJECTS)
            & article_clean.ne("")
            & product_clean.ne("")
            & (~article_clean.map(is_excluded_article))
            & product_clean.map(is_valid_product_code)
            & product_clean.combine(subject_clean, lambda p, s: VALID_PRODUCT_CATEGORY_REFERENCE.get(str(p)) == s)
        )
        out = out.loc[mask.values].copy()
        return out

    def buyout_rates(self) -> pd.DataFrame:
        f = self.enrich(self.pack.funnel, "funnel")
        if f.empty:
            return pd.DataFrame()
        f90 = f[(f["day"] >= self.cutoff_90) & (f["day"] <= self.latest_day)].copy()
        g = f90.groupby(["subject", "product", "supplier_article", "nm_id"], dropna=False, as_index=False).agg(
            orders_90=("orders", "sum"), buyouts_90=("buyouts_count", "sum"), cancels_90=("cancels_count", "sum"),
        )
        g["resolved_90"] = g["buyouts_90"].fillna(0) + g["cancels_90"].fillna(0)
        g["buyout_pct_90"] = np.where(g["resolved_90"] > 0, g["buyouts_90"] / g["resolved_90"], np.nan)
        g["buyout_pct_wrong_orders"] = np.where(g["orders_90"] > 0, g["buyouts_90"] / g["orders_90"], np.nan)
        # fallback by product, then category.
        prod = g.groupby(["subject", "product"], as_index=False).agg(prod_buyouts=("buyouts_90", "sum"), prod_cancels=("cancels_90", "sum"))
        prod["product_buyout_pct_90"] = np.where(prod["prod_buyouts"] + prod["prod_cancels"] > 0, prod["prod_buyouts"] / (prod["prod_buyouts"] + prod["prod_cancels"]), np.nan)
        cat = g.groupby(["subject"], as_index=False).agg(cat_buyouts=("buyouts_90", "sum"), cat_cancels=("cancels_90", "sum"))
        cat["category_buyout_pct_90"] = np.where(cat["cat_buyouts"] + cat["cat_cancels"] > 0, cat["cat_buyouts"] / (cat["cat_buyouts"] + cat["cat_cancels"]), np.nan)
        g = g.merge(prod[["subject", "product", "product_buyout_pct_90"]], on=["subject", "product"], how="left").merge(cat[["subject", "category_buyout_pct_90"]], on="subject", how="left")
        g["used_buyout_pct_90"] = g["buyout_pct_90"].fillna(g["product_buyout_pct_90"]).fillna(g["category_buyout_pct_90"]).fillna(1.0)
        g["used_buyout_pct_90"] = g["used_buyout_pct_90"].clip(0, 1)
        return g

    def order_prices_daily(self) -> pd.DataFrame:
        orders = self.enrich(self.pack.orders, "orders")
        if orders.empty:
            return pd.DataFrame()
        orders = orders[(orders["day"] >= self.cutoff_90) & (orders["day"] <= self.latest_day)].copy()
        if "order_sum" not in orders.columns:
            orders["order_sum"] = 0.0
        g = orders.groupby(["day", "subject", "product", "supplier_article", "nm_id"], dropna=False, as_index=False).agg(
            orders=("orders", "sum"),
            orders_rows=("orders", "sum"),
            order_sum=("order_sum", "sum"),
            finished_price=("finished_price", "mean"), price_with_disc=("price_with_disc", "mean"), spp=("spp", "mean"),
        )
        return g

    def funnel_daily(self) -> pd.DataFrame:
        f = self.enrich(self.pack.funnel, "funnel")
        if f.empty:
            return pd.DataFrame()
        f = f[(f["day"] >= self.cutoff_90) & (f["day"] <= self.latest_day)].copy()
        g = f.groupby(["day", "subject", "product", "supplier_article", "nm_id"], dropna=False, as_index=False).agg(
            orders=("orders", "sum"), order_sum=("order_sum", "sum"), open_cards=("open_cards", "sum"),
            add_to_cart=("add_to_cart", "sum"), buyouts_count=("buyouts_count", "sum"), cancels_count=("cancels_count", "sum"),
            finished_price_funnel=("finished_price", "mean"), spp_funnel=("spp", "mean"),
        )
        g["cart_conv_pct"] = np.where(g["open_cards"] > 0, g["add_to_cart"] / g["open_cards"] * 100, np.nan)
        g["order_conv_pct"] = np.where(g["add_to_cart"] > 0, g["orders"] / g["add_to_cart"] * 100, np.nan)
        return g

    def ads_daily_pivot(self) -> pd.DataFrame:
        ads = self.enrich(self.pack.ads_daily, "ads")
        if ads.empty:
            return pd.DataFrame()
        ads = ads[(ads["day"] >= self.cutoff_90) & (ads["day"] <= self.latest_day)].copy()
        # pivot manual/unified/unknown into columns.
        grouped = ads.groupby(["day", "subject", "product", "supplier_article", "nm_id", "ad_type"], dropna=False, as_index=False).agg(
            impressions=("impressions", "sum"), clicks=("clicks", "sum"), orders=("orders", "sum"),
            order_sum=("order_sum", "sum"), spend=("spend", "sum"),
        )
        rows = []
        for keys, part in grouped.groupby(["day", "subject", "product", "supplier_article", "nm_id"], dropna=False):
            rec = dict(zip(["day", "subject", "product", "supplier_article", "nm_id"], keys))
            for typ in ["manual", "unified", "unknown"]:
                p = part[part["ad_type"] == typ]
                imps, clicks, orders, order_sum, spend = [float(p[c].sum()) for c in ["impressions", "clicks", "orders", "order_sum", "spend"]]
                rec[f"{typ}_impressions"] = imps
                rec[f"{typ}_clicks"] = clicks
                rec[f"{typ}_orders"] = orders
                rec[f"{typ}_order_sum"] = order_sum
                rec[f"{typ}_spend"] = spend
                rec[f"{typ}_ctr_pct"] = clicks / imps * 100 if imps else np.nan
                rec[f"{typ}_cpc"] = spend / clicks if clicks else np.nan
                rec[f"{typ}_cr_pct"] = orders / clicks * 100 if clicks else np.nan
                rec[f"{typ}_drr_pct"] = spend / order_sum * 100 if order_sum else np.nan
            rows.append(rec)
        out = pd.DataFrame(rows)
        return out

    def search_daily_summary(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        q = self.enrich(self.pack.search_queries, "search_queries")
        if q.empty:
            return pd.DataFrame(), pd.DataFrame()
        q = q[(q["day"] >= self.cutoff_90) & (q["day"] <= self.latest_day)].copy()
        # Query data is already deduped by loader. Article-day summary.
        summary = q.groupby(["day", "subject", "product", "supplier_article", "nm_id"], dropna=False, as_index=False).agg(
            search_frequency=("frequency", "sum"), search_transitions=("transitions", "sum"),
            search_add_to_cart=("add_to_cart", "sum"), search_orders=("orders", "sum"),
            search_avg_position=("avg_position", lambda s: weighted_mean(s, q.loc[s.index, "orders"].fillna(0) + q.loc[s.index, "frequency"].fillna(0) / 1000)),
            search_median_position=("median_position", "mean"),
            rating_card=("rating_card", "mean"), rating_reviews=("rating_reviews", "mean"), visibility_pct=("visibility_pct", "mean"),
        )
        summary["search_traffic_capture_pct"] = np.where(summary["search_frequency"] > 0, summary["search_transitions"] / summary["search_frequency"] * 100, np.nan)
        # Core queries that give 80%+ orders over period.
        core_rows = []
        for keys, part in q.groupby(["subject", "product", "supplier_article", "nm_id"], dropna=False):
            part2 = part.groupby("search_query", as_index=False).agg(
                frequency=("frequency", "sum"), transitions=("transitions", "sum"), add_to_cart=("add_to_cart", "sum"), orders=("orders", "sum"),
                avg_position=("avg_position", lambda s: weighted_mean(s, part.loc[s.index, "orders"].fillna(0) + part.loc[s.index, "frequency"].fillna(0) / 1000)),
                median_position=("median_position", "mean"), rating_card=("rating_card", "mean"), rating_reviews=("rating_reviews", "mean"),
                visibility_pct=("visibility_pct", "mean"),
            ).sort_values("orders", ascending=False)
            total_orders = part2["orders"].sum()
            part2["orders_share_pct"] = np.where(total_orders > 0, part2["orders"] / total_orders * 100, 0)
            part2["cum_orders_share_pct"] = part2["orders_share_pct"].cumsum()
            if total_orders > 0:
                core = part2[(part2["cum_orders_share_pct"] <= 80) | (part2["orders_share_pct"] == part2["orders_share_pct"].max())].copy()
                # Ensure the row crossing 80 is included.
                crossing = part2[part2["cum_orders_share_pct"] > 80].head(1)
                core = pd.concat([core, crossing], ignore_index=True).drop_duplicates("search_query")
            else:
                core = part2.head(10).copy()
            for _, r in core.iterrows():
                rec = dict(zip(["subject", "product", "supplier_article", "nm_id"], keys))
                rec.update(r.to_dict())
                rec["traffic_capture_pct"] = r["transitions"] / r["frequency"] * 100 if r["frequency"] else np.nan
                core_rows.append(rec)
        core = pd.DataFrame(core_rows)
        return summary, core

    def search_unique_demand(self) -> pd.DataFrame:
        """Unique WB demand by level from raw search queries.

        Важно: спрос на уровне категории/товара нельзя считать суммой спроса по артикулам,
        потому что один и тот же поисковый запрос встречается у нескольких карточек.
        Поэтому считаем частотность один раз на уровне: день + уровень + нормализованный запрос.
        Для частотности берём max по дублям, для переходов/заказов — sum.
        """
        q = self.enrich(self.pack.search_queries, "search_queries")
        if q.empty:
            return pd.DataFrame()
        q = q[(q["day"] >= self.cutoff_90) & (q["day"] <= self.latest_day)].copy()
        if q.empty or "search_query" not in q.columns:
            return pd.DataFrame()
        q["query_norm"] = (
            q["search_query"].astype(str).str.lower().str.replace("ё", "е", regex=False)
            .str.replace(r"\s+", " ", regex=True).str.strip()
        )
        q = q[q["query_norm"].ne("")].copy()
        for col in ["frequency", "transitions", "add_to_cart", "orders"]:
            if col not in q.columns:
                q[col] = 0
            q[col] = pd.to_numeric(q[col], errors="coerce").fillna(0)

        def build(level: str, keys: List[str]) -> pd.DataFrame:
            cols = ["day"] + keys + ["query_norm"]
            x = q.groupby(cols, dropna=False, as_index=False).agg(
                unique_frequency=("frequency", "max"),
                query_rows=("frequency", "size"),
                transitions=("transitions", "sum"),
                add_to_cart=("add_to_cart", "sum"),
                orders=("orders", "sum"),
            )
            g = x.groupby(["day"] + keys, dropna=False, as_index=False).agg(
                unique_search_frequency=("unique_frequency", "sum"),
                unique_search_queries=("query_norm", "nunique"),
                raw_query_rows=("query_rows", "sum"),
                transitions=("transitions", "sum"),
                add_to_cart=("add_to_cart", "sum"),
                orders=("orders", "sum"),
            )
            g["duplicate_query_rows_removed"] = g["raw_query_rows"] - g["unique_search_queries"]
            g["level"] = level
            return g

        frames = [
            build("category", ["subject"]),
            build("product", ["subject", "product"]),
            build("article", ["subject", "product", "supplier_article", "nm_id"]),
        ]
        out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        return out

    def entry_points_summary(self) -> pd.DataFrame:
        e = self.enrich(self.pack.entry_points, "entry_points")
        if e.empty:
            return pd.DataFrame()
        e = e[(e["day"].isna()) | ((e["day"] >= self.cutoff_90) & (e["day"] <= self.latest_day))].copy()
        g = e.groupby(["subject", "product", "supplier_article", "nm_id", "entry_section", "entry_point"], dropna=False, as_index=False).agg(
            impressions=("impressions", "sum"), transitions=("transitions", "sum"), add_to_cart=("add_to_cart", "sum"), orders=("orders", "sum"),
        )
        g["ctr_pct"] = np.where(g["impressions"] > 0, g["transitions"] / g["impressions"] * 100, np.nan)
        g["cart_conv_pct"] = np.where(g["transitions"] > 0, g["add_to_cart"] / g["transitions"] * 100, np.nan)
        g["order_conv_pct"] = np.where(g["add_to_cart"] > 0, g["orders"] / g["add_to_cart"] * 100, np.nan)
        totals = g.groupby(["subject", "product", "supplier_article", "nm_id"], dropna=False)["orders"].sum().rename("orders_total").reset_index()
        g = g.merge(totals, on=["subject", "product", "supplier_article", "nm_id"], how="left")
        g["orders_share_pct"] = np.where(g["orders_total"] > 0, g["orders"] / g["orders_total"] * 100, np.nan)
        return g.sort_values(["subject", "product", "supplier_article", "orders"], ascending=[True, True, True, False])

    def localization(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        stock = self.enrich(self.pack.stock, "stock")
        orders = self.enrich(self.pack.orders, "orders")
        if stock.empty or orders.empty:
            return pd.DataFrame(), pd.DataFrame()
        orders = orders[(orders["day"] >= self.cutoff_90) & (orders["day"] <= self.latest_day)].copy()
        orders["warehouse"] = orders["warehouse"].map(canonical_warehouse_name)
        orders = orders[orders["warehouse"].map(is_relevant_warehouse)].copy()
        stock["warehouse"] = stock["warehouse"].map(canonical_warehouse_name)
        stock = stock[stock["warehouse"].map(is_relevant_warehouse)].copy()
        if stock.empty or orders.empty:
            return pd.DataFrame(), pd.DataFrame()
        weights = orders.groupby(["subject", "product", "supplier_article", "nm_id", "warehouse"], dropna=False, as_index=False).agg(orders_90=("orders", "sum"))
        totals = weights.groupby(["subject", "product", "supplier_article", "nm_id"], dropna=False)["orders_90"].sum().rename("orders_total").reset_index()
        weights = weights.merge(totals, on=["subject", "product", "supplier_article", "nm_id"], how="left")
        weights["warehouse_weight_pct"] = np.where(weights["orders_total"] > 0, weights["orders_90"] / weights["orders_total"] * 100, 0)
        # Оставляем ключевые склады, которые суммарно дают 97% заказов артикула.
        weights = weights.sort_values(["subject", "product", "supplier_article", "nm_id", "warehouse_weight_pct"], ascending=[True, True, True, True, False])
        weights["cum_weight_pct"] = weights.groupby(["subject", "product", "supplier_article", "nm_id"], dropna=False)["warehouse_weight_pct"].cumsum()
        weights = weights[(weights["cum_weight_pct"] <= 97) | (weights["warehouse_weight_pct"] >= 0.5)].copy()
        weights["avg_daily_orders_wh"] = weights["orders_90"] / 90.0
        weights["needed_stock_2d"] = weights["avg_daily_orders_wh"] * 2
        weights["warehouse_pool"] = weights["warehouse"].map(warehouse_pool)
        # Last stock date per article/warehouse.
        stock = stock[stock["day"].notna()].copy()
        stock = stock.sort_values("day").groupby(["subject", "product", "supplier_article", "nm_id", "warehouse"], dropna=False, as_index=False).tail(1)
        st = stock.groupby(["subject", "product", "supplier_article", "nm_id", "warehouse"], dropna=False, as_index=False).agg(stock_qty=("stock", "sum"), stock_day=("day", "max"))
        detail = weights.merge(st, on=["subject", "product", "supplier_article", "nm_id", "warehouse"], how="left")
        detail["stock_qty"] = detail["stock_qty"].fillna(0)
        detail["is_direct_covered"] = detail["stock_qty"] >= detail["needed_stock_2d"]
        # Replacement: same regional pool has enough stock in aggregate.
        pool_stock = detail.groupby(["subject", "product", "supplier_article", "nm_id", "warehouse_pool"], dropna=False)["stock_qty"].sum().rename("pool_stock_qty").reset_index()
        pool_need = detail.groupby(["subject", "product", "supplier_article", "nm_id", "warehouse_pool"], dropna=False)["needed_stock_2d"].sum().rename("pool_need_qty").reset_index()
        detail = detail.merge(pool_stock, on=["subject", "product", "supplier_article", "nm_id", "warehouse_pool"], how="left").merge(pool_need, on=["subject", "product", "supplier_article", "nm_id", "warehouse_pool"], how="left")
        detail["is_covered_with_replacement"] = detail["is_direct_covered"] | (detail["pool_stock_qty"] >= detail["needed_stock_2d"])
        detail["direct_coverage_weight_pct"] = np.where(detail["is_direct_covered"], detail["warehouse_weight_pct"], 0)
        detail["replacement_coverage_weight_pct"] = np.where(detail["is_covered_with_replacement"], detail["warehouse_weight_pct"], 0)
        summary = detail.groupby(["subject", "product", "supplier_article", "nm_id"], dropna=False, as_index=False).agg(
            direct_localization_pct=("direct_coverage_weight_pct", "sum"),
            localization_with_replacements_pct=("replacement_coverage_weight_pct", "sum"),
            stock_qty_total=("stock_qty", "sum"),
            key_warehouses=("warehouse", "nunique"),
            stock_day=("stock_day", "max"),
        )
        summary["localization_status"] = np.select(
            [summary["localization_with_replacements_pct"] >= 85, summary["localization_with_replacements_pct"] >= 60, summary["localization_with_replacements_pct"] >= 30],
            ["Норма", "Риск", "Плохая локализация"], default="Критично",
        )
        uncovered = detail[~detail["is_covered_with_replacement"]].groupby(["supplier_article", "nm_id"], dropna=False)["warehouse"].apply(lambda s: "; ".join(s.astype(str).head(8))).rename("uncovered_warehouses").reset_index()
        summary = summary.merge(uncovered, on=["supplier_article", "nm_id"], how="left")
        return detail, summary

    def gross_profit_potential(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        abc = self.enrich(self.pack.abc_weekly, "abc_weekly")
        if abc.empty:
            return pd.DataFrame(), pd.DataFrame()
        abc = abc[(abc["period_end"] >= self.cutoff_90) & (abc["period_start"] <= self.latest_day)].copy()
        abc["days_in_period"] = (pd.to_datetime(abc["period_end"]) - pd.to_datetime(abc["period_start"])).dt.days + 1
        abc["gp_per_day"] = np.where(abc["days_in_period"] > 0, abc["gross_profit"] / abc["days_in_period"], np.nan)
        weekly = abc[["subject", "product", "supplier_article", "nm_id", "week_code", "week_label", "period_start", "period_end", "gross_profit", "gp_per_day", "orders", "gross_revenue"]].copy()
        rows = []
        prev_month = (self.latest_day.to_period("M") - 1).strftime("%Y-%m")
        cur_month = self.latest_day.to_period("M").strftime("%Y-%m")
        for keys, part in weekly.groupby(["subject", "product", "supplier_article", "nm_id"], dropna=False):
            values = part["gp_per_day"].dropna()
            avg = values.mean() if len(values) else np.nan
            above = values[values > avg]
            target = above.mean() if len(above) else avg
            best = values.max() if len(values) else np.nan
            prev_gp = part.loc[pd.to_datetime(part["period_start"]).dt.to_period("M").astype(str) == prev_month, "gross_profit"].sum()
            cur_gp = part.loc[pd.to_datetime(part["period_start"]).dt.to_period("M").astype(str) == cur_month, "gross_profit"].sum()
            plan = prev_gp * 1.1 if prev_gp else cur_gp
            elapsed = min(self.latest_day.day, calendar.monthrange(self.latest_day.year, self.latest_day.month)[1])
            plan_to_date = plan / calendar.monthrange(self.latest_day.year, self.latest_day.month)[1] * elapsed if plan else np.nan
            rec = dict(zip(["subject", "product", "supplier_article", "nm_id"], keys))
            rec.update({
                "weeks_count": part["week_code"].nunique(), "gross_profit_90d": part["gross_profit"].sum(),
                "avg_gp_per_day": avg, "target_gp_per_day": target, "best_week_gp_per_day": best,
                "prev_month_gross_profit": prev_gp, "plan_month": plan, "current_month_gross_profit": cur_gp,
                "plan_completion_pct": cur_gp / plan * 100 if plan else np.nan,
                "plan_to_date": plan_to_date,
                "plan_to_date_completion_pct": cur_gp / plan_to_date * 100 if plan_to_date else np.nan,
            })
            rows.append(rec)
        return weekly, pd.DataFrame(rows)

    def article_day_fact(self) -> pd.DataFrame:
        # HARD RULE: orders/order_sum come only from WB Orders file.
        # Funnel is used only for traffic/conversion/buyout context and must never override sales.
        base = self.order_prices_daily()
        funnel = self.funnel_daily()
        ads = self.ads_daily_pivot()
        search_summary, _ = self.search_daily_summary()
        loc_detail, loc_summary = self.localization()
        gp_weekly, gp_potential = self.gross_profit_potential()
        buyouts = self.buyout_rates()
        keys = ["day", "subject", "product", "supplier_article", "nm_id"]
        if base.empty:
            if funnel is None or funnel.empty:
                return pd.DataFrame()
            # Keep traffic-only rows, but sales stay zero because Orders file has no row.
            base = funnel[keys].drop_duplicates().copy()
            base["orders"] = 0.0
            base["orders_rows"] = 0.0
            base["order_sum"] = 0.0
        out = base.copy()
        funnel_metrics = pd.DataFrame()
        if funnel is not None and not funnel.empty:
            drop_sales = [c for c in ["orders", "order_sum", "finished_price"] if c in funnel.columns]
            funnel_metrics = funnel.drop(columns=drop_sales, errors="ignore").copy()
        for df in [funnel_metrics, ads, search_summary]:
            if df is not None and not df.empty:
                out = out.merge(df, on=keys, how="outer", suffixes=("", "_dup"))
                for c in [c for c in out.columns if c.endswith("_dup")]:
                    basec = c[:-4]
                    if basec in out.columns:
                        out[basec] = out[basec].fillna(out[c])
                    out = out.drop(columns=[c])
        if not buyouts.empty:
            out = out.merge(buyouts[["supplier_article", "nm_id", "used_buyout_pct_90", "buyout_pct_90", "buyout_pct_wrong_orders"]], on=["supplier_article", "nm_id"], how="left")
        if not loc_summary.empty:
            out = out.merge(loc_summary[["supplier_article", "nm_id", "direct_localization_pct", "localization_with_replacements_pct", "localization_status", "stock_qty_total"]], on=["supplier_article", "nm_id"], how="left")
        # Sales are Orders-only. Rows created from funnel/search/ad without Orders row remain 0.
        for _sales_col in ["orders", "orders_rows", "order_sum"]:
            if _sales_col not in out.columns:
                out[_sales_col] = 0.0
            out[_sales_col] = pd.to_numeric(out[_sales_col], errors="coerce").fillna(0.0)
        # Fill price fields
        out["finished_price"] = out.get("finished_price", np.nan)
        if "finished_price_funnel" in out.columns:
            out["finished_price"] = out["finished_price"].fillna(out["finished_price_funnel"])
        out["spp"] = out.get("spp", np.nan)
        if "spp_funnel" in out.columns:
            out["spp"] = out["spp"].fillna(out["spp_funnel"])
        # General traffic capture from funnel opens vs search demand.
        search_freq = out["search_frequency"].fillna(0) if "search_frequency" in out.columns else pd.Series([0] * len(out), index=out.index)
        open_cards = out["open_cards"].fillna(0) if "open_cards" in out.columns else pd.Series([0] * len(out), index=out.index)
        # % поиска по управленческой логике = все открытия карточки / спрос WB.
        # Переходы из поиска отдельно остаются в данных, но не используются как % поиска.
        out["total_traffic_capture_pct"] = np.where(search_freq > 0, open_cards / search_freq * 100, np.nan)
        out["search_traffic_capture_pct"] = out["total_traffic_capture_pct"]
        # Gross profit forecast using economics and buyout.
        econ = self.enrich(self.pack.economics, "economics")
        if not econ.empty:
            econ_latest = econ.sort_values("week_code").drop_duplicates(["supplier_article", "nm_id"], keep="last")
            out = out.merge(econ_latest[["supplier_article", "nm_id", "commission_pct", "acquiring_pct", "logistics_direct", "logistics_return", "storage", "other_costs", "cost"]], on=["supplier_article", "nm_id"], how="left")
        for c in ["orders", "order_sum", "open_cards", "add_to_cart", "manual_spend", "unified_spend", "unknown_spend"]:
            if c not in out.columns:
                out[c] = 0
        if "used_buyout_pct_90" not in out.columns:
            out["used_buyout_pct_90"] = 1.0
        out["used_buyout_pct_90"] = pd.to_numeric(out["used_buyout_pct_90"], errors="coerce").fillna(1.0)
        out["buyout_qty_model"] = out["orders"].fillna(0) * out["used_buyout_pct_90"]
        # Formula: revenue = order_sum * buyout pct. Direct logistics from all orders, others as agreed.
        out["revenue_model"] = out["order_sum"].fillna(0) * out["used_buyout_pct_90"]
        for c in ["commission_pct", "acquiring_pct", "logistics_direct", "logistics_return", "storage", "other_costs", "cost"]:
            if c not in out.columns:
                out[c] = 0
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0)
        out["commission_model"] = out["revenue_model"] * out["commission_pct"] / 100
        out["acquiring_model"] = out["revenue_model"] * out["acquiring_pct"] / 100
        out["logistics_direct_model"] = out["orders"].fillna(0) * out["logistics_direct"]
        out["logistics_return_model"] = out["buyout_qty_model"] * out["logistics_return"]
        out["storage_model"] = out["buyout_qty_model"] * out["storage"]
        out["other_costs_model"] = out["buyout_qty_model"] * out["other_costs"]
        out["cost_model"] = out["buyout_qty_model"] * out["cost"]
        ad_spend_cols = [c for c in ["manual_spend", "unified_spend", "unknown_spend"] if c in out.columns]
        out["ad_spend_model"] = out[ad_spend_cols].sum(axis=1) if ad_spend_cols else 0
        out["gross_profit_model"] = out["revenue_model"] - out["commission_model"] - out["acquiring_model"] - out["logistics_direct_model"] - out["logistics_return_model"] - out["storage_model"] - out["other_costs_model"] - out["cost_model"] - out["ad_spend_model"]
        out = out.sort_values(["subject", "product", "supplier_article", "day"])
        return out

    @staticmethod
    def target_value(series: pd.Series) -> float:
        s = pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
        if s.empty:
            return np.nan
        avg = s.mean()
        above = s[s > avg]
        return float(above.mean()) if not above.empty else float(avg)

    def metrics_summary(self, daily: pd.DataFrame) -> pd.DataFrame:
        if daily.empty:
            return pd.DataFrame()
        metric_defs = [
            ("orders", "Заказы в день"), ("order_sum", "Сумма заказов"), ("gross_profit_model", "Валовая прибыль модель"),
            ("open_cards", "Открытия карточки / клики"), ("add_to_cart", "Добавления в корзину"),
            ("cart_conv_pct", "Конверсия в корзину, %"), ("order_conv_pct", "Конверсия в заказ, %"),
            ("price_with_disc", "priceWithDisc / цена продажи"), ("finished_price", "finishedPrice / цена покупателя"), ("spp", "СПП, %"),
            ("manual_impressions", "manual показы"), ("manual_clicks", "manual клики"), ("manual_ctr_pct", "manual CTR, %"), ("manual_cpc", "manual CPC"), ("manual_drr_pct", "manual ДРР, %"),
            ("unified_impressions", "unified показы"), ("unified_clicks", "unified клики"), ("unified_ctr_pct", "unified CTR, %"), ("unified_cpc", "unified CPC"), ("unified_drr_pct", "unified ДРР, %"),
            ("search_frequency", "Спрос / частотность"), ("search_transitions", "Переходы из поиска"), ("search_traffic_capture_pct", "% поискового трафика"), ("total_traffic_capture_pct", "% общего захвата спроса"),
            ("search_avg_position", "Средняя позиция"), ("rating_card", "Рейтинг карточки"), ("rating_reviews", "Рейтинг отзывов"),
            ("direct_localization_pct", "Прямая локализация, %"), ("localization_with_replacements_pct", "Локализация с заменами, %"),
        ]
        rows = []
        for keys, part in daily.groupby(["subject", "product", "supplier_article", "nm_id"], dropna=False):
            avg_order_sum = pd.to_numeric(part.get("order_sum", pd.Series(dtype=float)), errors="coerce").mean()
            best_days = part[pd.to_numeric(part.get("order_sum", 0), errors="coerce") > avg_order_sum].copy() if pd.notna(avg_order_sum) else part.iloc[0:0]
            last_week = part[(part["day"] >= self.week_start) & (part["day"] <= self.week_end)].copy()
            for col, label in metric_defs:
                if col not in part.columns:
                    continue
                s = pd.to_numeric(part[col], errors="coerce")
                nonzero = s.replace(0, np.nan)
                rec = dict(zip(["subject", "product", "supplier_article", "nm_id"], keys))
                rec.update({
                    "metric": label,
                    "avg_90d_all_days": s.mean(),
                    "avg_90d_nonzero_days": nonzero.mean(),
                    "target_above_mean_90d": self.target_value(nonzero),
                    "best_days_avg": pd.to_numeric(best_days[col], errors="coerce").mean() if col in best_days.columns and not best_days.empty else np.nan,
                    "last_full_week_avg": pd.to_numeric(last_week[col], errors="coerce").mean() if col in last_week.columns and not last_week.empty else np.nan,
                    "last_full_week_sum": pd.to_numeric(last_week[col], errors="coerce").sum() if col in last_week.columns and not last_week.empty else np.nan,
                    "days_count": part["day"].nunique(),
                    "best_days_count": best_days["day"].nunique(),
                })
                target = rec["target_above_mean_90d"]
                fact = rec["last_full_week_avg"]
                rec["gap_to_target_pct"] = (fact / target - 1) * 100 if pd.notna(target) and target != 0 and pd.notna(fact) else np.nan
                rows.append(rec)
        return pd.DataFrame(rows)

    def best_days(self, daily: pd.DataFrame) -> pd.DataFrame:
        rows = []
        if daily.empty:
            return pd.DataFrame()
        for keys, part in daily.groupby(["subject", "product", "supplier_article", "nm_id"], dropna=False):
            avg_order_sum = pd.to_numeric(part["order_sum"], errors="coerce").mean() if "order_sum" in part.columns else np.nan
            best = part[pd.to_numeric(part.get("order_sum", 0), errors="coerce") > avg_order_sum].copy() if pd.notna(avg_order_sum) else part.head(0)
            best = best.sort_values("order_sum", ascending=False).head(15)
            for _, r in best.iterrows():
                rows.append({
                    "subject": keys[0], "product": keys[1], "supplier_article": keys[2], "nm_id": keys[3],
                    "day": r.get("day"), "orders": r.get("orders"), "order_sum": r.get("order_sum"), "gross_profit_model": r.get("gross_profit_model"),
                    "price_with_disc": r.get("price_with_disc"), "finished_price": r.get("finished_price"), "spp": r.get("spp"), "open_cards": r.get("open_cards"), "add_to_cart": r.get("add_to_cart"),
                    "cart_conv_pct": r.get("cart_conv_pct"), "order_conv_pct": r.get("order_conv_pct"),
                    "manual_ctr_pct": r.get("manual_ctr_pct"), "manual_drr_pct": r.get("manual_drr_pct"),
                    "unified_ctr_pct": r.get("unified_ctr_pct"), "unified_drr_pct": r.get("unified_drr_pct"),
                    "search_frequency": r.get("search_frequency"), "search_traffic_capture_pct": r.get("search_traffic_capture_pct"),
                    "total_traffic_capture_pct": r.get("total_traffic_capture_pct"), "search_avg_position": r.get("search_avg_position"),
                    "rating_reviews": r.get("rating_reviews"), "localization_with_replacements_pct": r.get("localization_with_replacements_pct"),
                })
        return pd.DataFrame(rows)

    def price_ranges(self, daily: pd.DataFrame) -> pd.DataFrame:
        if daily.empty or ("price_with_disc" not in daily.columns and "order_sum" not in daily.columns):
            return pd.DataFrame()
        df = daily.copy()
        if "price_with_disc" not in df.columns:
            df["price_with_disc"] = np.nan
        # priceWithDisc is the seller sale price. When it is absent on aggregated rows,
        # fallback to Orders-only order_sum / orders; do not use finishedPrice for sale-price ranges.
        orders_num = pd.to_numeric(df.get("orders", 0), errors="coerce").replace(0, np.nan)
        order_sum_num = pd.to_numeric(df.get("order_sum", 0), errors="coerce")
        df["sale_price_for_range"] = pd.to_numeric(df["price_with_disc"], errors="coerce").fillna(order_sum_num / orders_num)
        def bucket(p):
            if pd.isna(p):
                return "нет цены"
            step = 10 if p < 300 else 20 if p < 700 else 50
            lo = math.floor(p / step) * step
            hi = lo + step
            return f"{lo:.0f}-{hi:.0f}"
        df["price_range"] = df["sale_price_for_range"].map(bucket)
        g = df.groupby(["subject", "product", "supplier_article", "nm_id", "price_range"], dropna=False, as_index=False).agg(
            days=("day", "nunique"), order_sum=("order_sum", "sum"), orders=("orders", "sum"),
            avg_finished_price=("finished_price", "mean"), avg_price_with_disc=("sale_price_for_range", "mean"), avg_gross_profit=("gross_profit_model", "mean"),
            avg_drr_manual=("manual_drr_pct", "mean"), avg_drr_unified=("unified_drr_pct", "mean"),
            avg_cart_conv_pct=("cart_conv_pct", "mean"), avg_order_conv_pct=("order_conv_pct", "mean"),
        )
        # Mark recommended range: max order_sum per article.
        g["is_recommended"] = False
        idx = g.groupby(["supplier_article", "nm_id"], dropna=False)["order_sum"].idxmax()
        g.loc[idx.dropna().astype(int), "is_recommended"] = True
        return g.sort_values(["subject", "product", "supplier_article", "order_sum"], ascending=[True, True, True, False])

    def channel_summary(self, daily: pd.DataFrame) -> pd.DataFrame:
        rows = []
        if daily.empty:
            return pd.DataFrame()
        # средний чек нужен, чтобы оценить сумму заказов по точкам входа, где WB отдаёт только шт.
        avg_check = daily.copy()
        avg_check["avg_order_value"] = np.where(avg_check.get("orders", 0).fillna(0) > 0, avg_check.get("order_sum", 0).fillna(0) / avg_check.get("orders", 0).replace(0, np.nan), np.nan)
        avg_check_map = avg_check.groupby(["subject", "product", "supplier_article", "nm_id"], dropna=False)["avg_order_value"].mean().to_dict()
        ad_sums = {}
        for keys, part in daily.groupby(["subject", "product", "supplier_article", "nm_id"], dropna=False):
            ad_sums[keys] = {}
            for typ, name in [("manual", "Реклама manual: Поиск/Каталог"), ("unified", "Реклама unified: Карточка товара/полки"), ("unknown", "Реклама без типа ставки")]:
                imps = part.get(f"{typ}_impressions", pd.Series(dtype=float)).sum()
                clicks = part.get(f"{typ}_clicks", pd.Series(dtype=float)).sum()
                orders = part.get(f"{typ}_orders", pd.Series(dtype=float)).sum()
                order_sum = part.get(f"{typ}_order_sum", pd.Series(dtype=float)).sum()
                spend = part.get(f"{typ}_spend", pd.Series(dtype=float)).sum()
                ad_sums[keys][typ] = {"spend": spend, "impressions": imps, "clicks": clicks, "orders": orders, "order_sum": order_sum}
                rows.append({
                    "subject": keys[0], "product": keys[1], "supplier_article": keys[2], "nm_id": keys[3], "channel": name,
                    "channel_group": "Реклама", "impressions": imps, "clicks": clicks,
                    "ctr_pct": clicks / imps * 100 if imps else np.nan,
                    "orders": orders, "order_sum": order_sum, "estimated_order_sum": np.nan, "spend": spend,
                    "cpc": spend / clicks if clicks else np.nan,
                    "cr_pct": orders / clicks * 100 if clicks else np.nan,
                    "drr_pct": spend / order_sum * 100 if order_sum else np.nan,
                    "comment": "Факт по рекламному отчёту",
                })
        entry = self.entry_points_summary()
        entry_agg_rows = []
        if not entry.empty:
            entry["entry_channel"] = entry.apply(lambda r: classify_entry_channel(r.get("entry_section"), r.get("entry_point")), axis=1)
            # Детальные точки входа
            for _, r in entry.iterrows():
                keys = (r["subject"], r["product"], r["supplier_article"], r["nm_id"])
                aov = avg_check_map.get(keys, np.nan)
                est_sum = r["orders"] * aov if pd.notna(aov) else np.nan
                rows.append({
                    "subject": r["subject"], "product": r["product"], "supplier_article": r["supplier_article"], "nm_id": r["nm_id"],
                    "channel": f"Точка входа: {r['entry_section']} / {r['entry_point']}", "channel_group": r["entry_channel"],
                    "impressions": r["impressions"], "clicks": r["transitions"], "ctr_pct": np.nan,
                    "orders": r["orders"], "order_sum": np.nan, "estimated_order_sum": est_sum, "spend": np.nan,
                    "cpc": np.nan, "cr_pct": r["order_conv_pct"], "drr_pct": np.nan,
                    "orders_share_pct": r.get("orders_share_pct", np.nan),
                    "comment": "CTR для точек входа не сравниваем с рекламным CTR; показы/переходы могут быть разными сущностями",
                })
            # Агрегация каналов точек входа + привязка расходов рекламы для оценки ДРР канала
            ep = entry.groupby(["subject", "product", "supplier_article", "nm_id", "entry_channel"], dropna=False, as_index=False).agg(
                impressions=("impressions", "sum"), transitions=("transitions", "sum"), add_to_cart=("add_to_cart", "sum"), orders=("orders", "sum")
            )
            for _, r in ep.iterrows():
                keys = (r["subject"], r["product"], r["supplier_article"], r["nm_id"])
                aov = avg_check_map.get(keys, np.nan)
                est_sum = r["orders"] * aov if pd.notna(aov) else np.nan
                spend = np.nan
                if r["entry_channel"] == "Поиск/Каталог":
                    spend = ad_sums.get(keys, {}).get("manual", {}).get("spend", np.nan)
                    channel = "Канал Поиск/Каталог: заказы точки входа + расход manual"
                elif r["entry_channel"] == "Карточка товара / полки":
                    spend = ad_sums.get(keys, {}).get("unified", {}).get("spend", np.nan)
                    channel = "Канал Карточка товара/полки: заказы точки входа + расход unified"
                else:
                    channel = f"Канал {r['entry_channel']}: точки входа"
                rows.append({
                    "subject": r["subject"], "product": r["product"], "supplier_article": r["supplier_article"], "nm_id": r["nm_id"],
                    "channel": channel, "channel_group": r["entry_channel"],
                    "impressions": r["impressions"], "clicks": r["transitions"], "ctr_pct": np.nan,
                    "orders": r["orders"], "order_sum": np.nan, "estimated_order_sum": est_sum, "spend": spend,
                    "cpc": np.nan, "cr_pct": r["orders"] / r["transitions"] * 100 if r["transitions"] else np.nan,
                    "drr_pct": spend / est_sum * 100 if pd.notna(spend) and pd.notna(est_sum) and est_sum else np.nan,
                    "orders_share_pct": np.nan,
                    "comment": "ДРР канала оценочный: сумма заказов канала = заказы канала × средний чек артикула",
                })
        out = pd.DataFrame(rows)
        return out.sort_values(["subject", "product", "supplier_article", "orders"], ascending=[True, True, True, False]) if not out.empty else out

    def best_day_factors(self, daily: pd.DataFrame) -> pd.DataFrame:
        """Сравнение обычных дней и дней, где сумма заказов выше среднего."""
        if daily.empty:
            return pd.DataFrame()
        factors = [
            ("orders", "Заказы"), ("order_sum", "Сумма заказов"), ("open_cards", "Открытия карточки / клики"),
            ("add_to_cart", "Добавления в корзину"), ("cart_conv_pct", "Конверсия в корзину, %"), ("order_conv_pct", "Конверсия в заказ, %"),
            ("price_with_disc", "priceWithDisc / цена продажи"), ("finished_price", "finishedPrice / цена покупателя"), ("spp", "СПП, %"),
            ("manual_impressions", "Показы manual"), ("manual_clicks", "Клики manual"), ("manual_ctr_pct", "CTR manual, %"), ("manual_drr_pct", "ДРР manual, %"),
            ("unified_impressions", "Показы unified"), ("unified_clicks", "Клики unified"), ("unified_ctr_pct", "CTR unified, %"), ("unified_drr_pct", "ДРР unified, %"),
            ("search_frequency", "Спрос / частотность"), ("search_transitions", "Переходы из поиска"), ("search_traffic_capture_pct", "% поискового трафика"),
            ("total_traffic_capture_pct", "% общего захвата спроса"), ("search_avg_position", "Средняя позиция"),
            ("rating_reviews", "Рейтинг отзывов"), ("localization_with_replacements_pct", "Локализация с заменами, %"),
        ]
        rows = []
        for keys, part in daily.groupby(["subject", "product", "supplier_article", "nm_id"], dropna=False):
            order_sum = pd.to_numeric(part.get("order_sum", 0), errors="coerce")
            avg_order_sum = order_sum.mean()
            best = part[order_sum > avg_order_sum].copy() if pd.notna(avg_order_sum) else part.iloc[0:0]
            normal = part[~part.index.isin(best.index)].copy()
            if best.empty:
                continue
            for col, label in factors:
                if col not in part.columns:
                    continue
                best_avg = pd.to_numeric(best[col], errors="coerce").replace([np.inf, -np.inf], np.nan).mean()
                norm_avg = pd.to_numeric(normal[col], errors="coerce").replace([np.inf, -np.inf], np.nan).mean()
                diff = pct_gap(best_avg, norm_avg)
                if pd.isna(diff):
                    conclusion = "недостаточно данных"
                elif label in ["Средняя позиция", "ДРР manual, %", "ДРР unified, %"]:
                    conclusion = "лучше в сильные дни" if diff < -5 else "хуже в сильные дни" if diff > 5 else "примерно без изменений"
                else:
                    conclusion = "выше в сильные дни" if diff > 5 else "ниже в сильные дни" if diff < -5 else "примерно без изменений"
                rows.append({
                    "subject": keys[0], "product": keys[1], "supplier_article": keys[2], "nm_id": keys[3],
                    "factor": label, "normal_days_avg": norm_avg, "best_days_avg": best_avg,
                    "difference_pct": diff, "best_days_count": best["day"].nunique(), "conclusion": conclusion,
                })
        return pd.DataFrame(rows)

    def conclusions(self, summary: pd.DataFrame, daily: pd.DataFrame, loc_summary: pd.DataFrame) -> pd.DataFrame:
        if summary.empty:
            return pd.DataFrame()
        piv = summary.pivot_table(index=["subject", "product", "supplier_article", "nm_id"], columns="metric", values=["last_full_week_avg", "target_above_mean_90d", "gap_to_target_pct"], aggfunc="first")
        piv.columns = [f"{a}__{b}" for a, b in piv.columns]
        piv = piv.reset_index()
        if loc_summary is not None and not loc_summary.empty:
            piv = piv.merge(loc_summary[["supplier_article", "nm_id", "direct_localization_pct", "localization_with_replacements_pct", "localization_status", "uncovered_warehouses"]], on=["supplier_article", "nm_id"], how="left")
        rows = []
        for _, r in piv.iterrows():
            sales_gap = r.get("gap_to_target_pct__Сумма заказов", np.nan)
            if pd.notna(sales_gap) and sales_gap >= 5:
                status = "Опережаем целевой уровень"
            elif pd.notna(sales_gap) and sales_gap < -10:
                status = "Отстаём от целевого уровня"
            else:
                status = "Около целевого уровня"
            candidates = []
            def add_factor(name, severity, evidence, recommendation):
                if pd.notna(severity):
                    candidates.append((abs(float(severity)), name, evidence, recommendation))
            loc = r.get("localization_with_replacements_pct", np.nan)
            if status.startswith("Отстаём"):
                if pd.notna(loc) and loc < 85:
                    add_factor("Локализация / остатки", 85 - loc, f"локализация с заменами {loc:.1f}%", "восстановить остатки на ключевых складах и в региональных заменителях")
                traffic_gap = r.get("gap_to_target_pct__% поискового трафика", np.nan)
                if pd.notna(traffic_gap) and traffic_gap < -10:
                    add_factor("Забираем меньше поискового трафика", traffic_gap, f"% поискового трафика ниже цели на {abs(traffic_gap):.1f}%", "проверить позиции, SEO, ставки manual и релевантность ключей")
                demand_gap = r.get("gap_to_target_pct__Спрос / частотность", np.nan)
                if pd.notna(demand_gap) and demand_gap < -10:
                    add_factor("Общий спрос на WB ниже", demand_gap, f"частотность ниже цели на {abs(demand_gap):.1f}%", "сравнить с категорией и не завышать план на период низкого спроса")
                clicks_gap = r.get("gap_to_target_pct__Открытия карточки / клики", np.nan)
                if pd.notna(clicks_gap) and clicks_gap < -10:
                    add_factor("Меньше открытий карточки", clicks_gap, f"клики ниже цели на {abs(clicks_gap):.1f}%", "проверить выдачу, рекламу, CTR и карточку")
                cart_gap = r.get("gap_to_target_pct__Конверсия в корзину, %", np.nan)
                if pd.notna(cart_gap) and cart_gap < -10:
                    add_factor("Просела конверсия в корзину", cart_gap, f"конверсия в корзину ниже цели на {abs(cart_gap):.1f}%", "проверить фото, первый экран, цену, отзывы и УТП")
                order_gap = r.get("gap_to_target_pct__Конверсия в заказ, %", np.nan)
                if pd.notna(order_gap) and order_gap < -10:
                    add_factor("Просела конверсия в заказ", order_gap, f"конверсия в заказ ниже цели на {abs(order_gap):.1f}%", "проверить конечную цену, доставку, остатки и рейтинг")
                rating_gap = r.get("gap_to_target_pct__Рейтинг отзывов", np.nan)
                if pd.notna(rating_gap) and rating_gap < -2:
                    add_factor("Ухудшился рейтинг отзывов", rating_gap, f"рейтинг отзывов ниже нормы на {abs(rating_gap):.1f}%", "проверить свежие отзывы и причины снижения доверия")
            else:
                if status.startswith("Опережаем"):
                    add_factor("Сумма заказов выше цели", sales_gap if pd.notna(sales_gap) else 0, f"сумма заказов выше цели на {sales_gap:.1f}%" if pd.notna(sales_gap) else "сумма заказов выше цели", "зафиксировать условия лучших дней: цену, СПП, каналы, локализацию")
                else:
                    add_factor("Факт близок к цели", 1, "ключевые показатели около целевого уровня", "держать показатели не ниже средней планки")
            candidates = sorted(candidates, reverse=True)
            if candidates:
                main = candidates[0]
                secondary = candidates[1] if len(candidates) > 1 else None
                main_reason = f"{main[1]}: {main[2]}"
                second_reason = f"{secondary[1]}: {secondary[2]}" if secondary else "нет явной вторичной причины"
                recommendation = main[3]
                if secondary and secondary[3] != main[3]:
                    recommendation = recommendation + "; " + secondary[3]
            else:
                main_reason = "отклонение без явного единственного фактора"
                second_reason = "нужно смотреть блок факторов лучших дней"
                recommendation = "сравнить лучшие дни с обычными: цена, трафик, реклама, локализация"
            rec = r.to_dict()
            rec.update({
                "status": status,
                "main_reason": main_reason,
                "secondary_reason": second_reason,
                "recommendation": recommendation,
            })
            rows.append(rec)
        out = pd.DataFrame(rows)
        first = ["subject", "product", "supplier_article", "nm_id", "status", "main_reason", "secondary_reason", "recommendation", "localization_with_replacements_pct", "localization_status", "uncovered_warehouses"]
        # When Orders/localization are empty, these localization columns may not exist.
        # Keep report generation stable and leave the values blank instead of crashing.
        for col in first:
            if col not in out.columns:
                out[col] = np.nan
        rest = [c for c in out.columns if c not in first]
        return out[first + rest]

    def build_all(self) -> Dict[str, pd.DataFrame]:
        daily = self.article_day_fact()
        metrics = self.metrics_summary(daily)
        best = self.best_days(daily)
        best_factors = self.best_day_factors(daily)
        price = self.price_ranges(daily)
        channel = self.channel_summary(daily)
        search_summary, core_queries = self.search_daily_summary()
        search_unique_demand = self.search_unique_demand()
        entry = self.entry_points_summary()
        loc_detail, loc_summary = self.localization()
        gp_weekly, gp_potential = self.gross_profit_potential()
        buyout = self.buyout_rates()
        conclusions = self.conclusions(metrics, daily, loc_summary)
        return {
            "article_day_fact": daily,
            "metrics_summary_90d": metrics,
            "best_days": best,
            "best_day_factors": best_factors,
            "price_ranges": price,
            "channel_summary": channel,
            # Raw ad source for PDF current-week advertising truth.
            # article_day_fact can repeat campaign spend across articles; this source is grouped from the ad report itself.
            "ads_category_source": self.pack.ads_category.copy() if getattr(self.pack, "ads_category", pd.DataFrame()) is not None else pd.DataFrame(),
            "ads_raw_source": self.enrich(self.pack.ads_raw, "ads_raw"),
            "ads_daily_source": self.enrich(self.pack.ads_daily, "ads_daily"),
            "search_daily_summary": search_summary,
            "core_queries_80": core_queries,
            "search_unique_demand": search_unique_demand,
            "entry_points_summary": entry,
            "localization_detail": loc_detail,
            "localization_summary": loc_summary,
            "gp_potential_weekly_90d": gp_weekly,
            "gp_potential_90d": gp_potential,
            "buyout_validation": buyout,
            "conclusions": conclusions,
            "dictionary": self.dictionary,
            "diagnostics": self.diag.frame(),
        }




COLUMN_RU = {
    "subject": "Категория", "product": "Товар", "supplier_article": "Артикул продавца", "nm_id": "Артикул WB", "day": "Дата",
    "metric": "Показатель", "avg_90d_all_days": "Среднее за 90 дней", "avg_90d_nonzero_days": "Среднее по активным дням",
    "target_above_mean_90d": "Целевое значение", "best_days_avg": "Среднее в лучшие дни", "last_full_week_avg": "Среднее за последнюю полную неделю",
    "last_full_week_sum": "Сумма за последнюю полную неделю", "gap_to_target_pct": "Отклонение от цели, %", "days_count": "Дней в анализе", "best_days_count": "Лучших дней",
    "orders": "Заказы", "orders_rows": "Строк заказов", "order_sum": "Сумма заказов", "gross_profit_model": "Валовая прибыль модель",
    "open_cards": "Открытия карточки / клики", "add_to_cart": "Добавления в корзину", "cart_conv_pct": "Конверсия в корзину, %", "order_conv_pct": "Конверсия в заказ, %",
    "finished_price": "finishedPrice / цена покупателя", "price_with_disc": "priceWithDisc / цена продажи", "spp": "СПП, %",
    "manual_impressions": "Показы manual", "manual_clicks": "Клики manual", "manual_orders": "Заказы manual", "manual_order_sum": "Сумма заказов manual", "manual_spend": "Расход manual", "manual_ctr_pct": "CTR manual, %", "manual_cpc": "CPC manual, ₽", "manual_cr_pct": "CR manual, %", "manual_drr_pct": "ДРР manual, %",
    "unified_impressions": "Показы unified", "unified_clicks": "Клики unified", "unified_orders": "Заказы unified", "unified_order_sum": "Сумма заказов unified", "unified_spend": "Расход unified", "unified_ctr_pct": "CTR unified, %", "unified_cpc": "CPC unified, ₽", "unified_cr_pct": "CR unified, %", "unified_drr_pct": "ДРР unified, %",
    "unknown_impressions": "Показы без типа", "unknown_clicks": "Клики без типа", "unknown_orders": "Заказы без типа", "unknown_order_sum": "Сумма заказов без типа", "unknown_spend": "Расход без типа", "unknown_ctr_pct": "CTR без типа, %", "unknown_cpc": "CPC без типа, ₽", "unknown_cr_pct": "CR без типа, %", "unknown_drr_pct": "ДРР без типа, %",
    "search_frequency": "Спрос / частотность", "search_transitions": "Переходы из поиска", "search_add_to_cart": "Добавления из поиска", "search_orders": "Заказы из поиска", "search_traffic_capture_pct": "% поискового трафика", "total_traffic_capture_pct": "% общего захвата спроса",
    "search_avg_position": "Средняя позиция", "search_median_position": "Медианная позиция", "visibility_pct": "Видимость, %", "rating_card": "Рейтинг карточки", "rating_reviews": "Рейтинг отзывов",
    "direct_localization_pct": "Прямая локализация, %", "localization_with_replacements_pct": "Локализация с заменами, %", "localization_status": "Статус локализации", "stock_qty_total": "Остаток всего", "uncovered_warehouses": "Непокрытые склады", "key_warehouses": "Ключевых складов", "stock_day": "Дата остатков",
    "price_range": "Диапазон priceWithDisc", "days": "Дней", "avg_finished_price": "Средний finishedPrice / цена покупателя", "avg_price_with_disc": "Средний priceWithDisc / цена продажи", "avg_gross_profit": "Средняя валовая прибыль", "avg_drr_manual": "Средний ДРР manual, %", "avg_drr_unified": "Средний ДРР unified, %", "avg_cart_conv_pct": "Средняя конверсия в корзину, %", "avg_order_conv_pct": "Средняя конверсия в заказ, %", "is_recommended": "Рекомендуемый диапазон",
    "channel": "Канал", "channel_group": "Группа канала", "impressions": "Показы", "clicks": "Клики / переходы", "ctr_pct": "CTR, %", "spend": "Расход", "cpc": "CPC, ₽", "cr_pct": "CR, %", "drr_pct": "ДРР, %", "orders_share_pct": "Доля заказов, %", "estimated_order_sum": "Оценочная сумма заказов", "comment": "Комментарий",
    "entry_section": "Раздел", "entry_point": "Точка входа", "transitions": "Переходы", "frequency": "Частотность", "search_query": "Поисковый запрос", "traffic_capture_pct": "% трафика", "orders_share_pct": "Доля заказов, %", "cum_orders_share_pct": "Накопленная доля заказов, %", "avg_position": "Средняя позиция", "median_position": "Медианная позиция",
    "warehouse": "Склад", "orders_90": "Заказы за 90 дней", "orders_total": "Заказы всего", "warehouse_weight_pct": "Вес склада, %", "cum_weight_pct": "Накопленный вес, %", "avg_daily_orders_wh": "Средние заказы склада в день", "needed_stock_2d": "Нужно остатка на 2 дня", "warehouse_pool": "Региональный пул", "stock_qty": "Остаток", "is_direct_covered": "Покрыт напрямую", "is_covered_with_replacement": "Покрыт с заменой", "pool_stock_qty": "Остаток пула", "pool_need_qty": "Потребность пула", "direct_coverage_weight_pct": "Вклад прямого покрытия, %", "replacement_coverage_weight_pct": "Вклад покрытия с заменой, %",
    "week_code": "Неделя", "week_label": "Период недели", "period_start": "Начало периода", "period_end": "Конец периода", "gross_profit": "Валовая прибыль", "gross_revenue": "Валовая выручка", "gp_per_day": "ВП в день", "weeks_count": "Недель в анализе", "gross_profit_90d": "ВП за 90 дней", "avg_gp_per_day": "Средняя ВП/день", "target_gp_per_day": "Целевая ВП/день", "best_week_gp_per_day": "Лучшая неделя ВП/день", "prev_month_gross_profit": "ВП прошлого месяца", "plan_month": "План месяца", "current_month_gross_profit": "ВП текущего месяца", "plan_completion_pct": "Выполнение плана, %", "plan_to_date": "План на дату", "plan_to_date_completion_pct": "Выполнение плана на дату, %",
    "orders_90": "Заказали за 90 дней", "buyouts_90": "Выкупили за 90 дней", "cancels_90": "Отменили за 90 дней", "resolved_90": "Завершённые заказы", "buyout_pct_90": "% выкупа правильный", "buyout_pct_wrong_orders": "% выкупа старый ошибочный", "product_buyout_pct_90": "% выкупа товара", "category_buyout_pct_90": "% выкупа категории", "used_buyout_pct_90": "Использованный % выкупа",
    "factor": "Фактор", "normal_days_avg": "Обычные дни", "best_days_avg": "Лучшие дни", "difference_pct": "Разница, %", "conclusion": "Вывод",
    "status": "Статус", "main_reason": "Главная причина", "secondary_reason": "Вторичная причина", "recommendation": "Рекомендация",
    "source": "Источник", "source_file": "Файл-источник", "timestamp": "Время", "level": "Уровень", "message": "Сообщение", "details": "Детали",
}


def translate_col_name(col: Any) -> str:
    c = str(col)
    if c in COLUMN_RU:
        return COLUMN_RU[c]
    if "__" in c:
        left, right = c.split("__", 1)
        left_ru = COLUMN_RU.get(left, left)
        return f"{left_ru}: {right}"
    # Остаточные служебные имена переводим по частям.
    out = c
    replacements = {
        "pct": "%", "avg": "среднее", "target": "цель", "last_full_week": "последняя полная неделя",
        "order_sum": "сумма заказов", "gross_profit": "валовая прибыль", "localization": "локализация",
    }
    for a, b in replacements.items():
        out = out.replace(a, b)
    return out


def translate_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    x = df.copy()
    x = x.rename(columns={c: translate_col_name(c) for c in x.columns})
    return x

# ------------------------- export -------------------------
def autofit(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for cell in ws[letter]:
            max_len = max(max_len, min(len(str(cell.value)) if cell.value is not None else 0, 60))
        ws.column_dimensions[letter].width = min(max_len + 2, 42)


def style_sheet(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = BORDER
            if isinstance(cell.value, (int, float, np.number)) and not isinstance(cell.value, bool):
                header = str(ws.cell(1, cell.column).value or "")
                if "%" in header or "pct" in header.lower() or "ДРР" in header or "CTR" in header:
                    cell.number_format = '0.0'
                elif "цена" in header.lower() or "сум" in header.lower() or "ВП" in header or "profit" in header.lower() or "spend" in header.lower() or "выруч" in header.lower() or "order_sum" in header.lower():
                    cell.number_format = money_format()
                else:
                    cell.number_format = '# ##0.00'
    ws.freeze_panes = "A2"
    autofit(ws)


def write_df_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name[:31])
    if df is None or df.empty:
        ws.cell(1, 1, "Нет данных")
        return
    x = translate_df(df.copy())
    for c in x.columns:
        if pd.api.types.is_datetime64_any_dtype(x[c]):
            x[c] = x[c].dt.strftime("%Y-%m-%d")
    ws.append(list(x.columns))
    for row in x.itertuples(index=False, name=None):
        ws.append(list(row))
    style_sheet(ws)


def write_product_blocks(path: Path, title: str, outputs: Dict[str, pd.DataFrame], sections: List[Tuple[str, str, Optional[List[str]]]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    # Determine products from metrics/conclusions/daily.
    base = None
    for key in ["metrics_summary_90d", "conclusions", "article_day_fact", "channel_summary"]:
        if key in outputs and outputs[key] is not None and not outputs[key].empty:
            base = outputs[key]
            break
    if base is None or base.empty:
        ws = wb.create_sheet("Нет данных")
        ws.cell(1, 1, "Нет данных")
        wb.save(path)
        return
    used = set()
    products = base[["subject", "product"]].drop_duplicates().sort_values(["subject", "product"]).itertuples(index=False, name=None)
    for subject, product in products:
        ws = wb.create_sheet(safe_sheet_name(str(product), used))
        row = 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        c = ws.cell(row, 1, f"{title}: товар {product} / {subject}")
        c.fill = TITLE_FILL
        c.font = Font(color="FFFFFF", bold=True, size=14)
        c.alignment = Alignment(horizontal="center")
        row += 2
        articles = base[(base["subject"] == subject) & (base["product"].astype(str) == str(product))]["supplier_article"].dropna().drop_duplicates().astype(str).sort_values().tolist()
        for art in articles:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            ac = ws.cell(row, 1, f"Артикул {art}")
            ac.fill = SECTION_FILL
            ac.font = Font(bold=True, size=12)
            row += 1
            for section_title, key, cols in sections:
                df = outputs.get(key, pd.DataFrame())
                if df is None or df.empty or "supplier_article" not in df.columns:
                    continue
                part = df[(df["subject"] == subject) & (df["product"].astype(str) == str(product)) & (df["supplier_article"].astype(str) == art)].copy()
                if part.empty:
                    continue
                if cols:
                    keep = [c for c in cols if c in part.columns]
                    part = part[keep]
                if len(part) > 60:
                    part = part.head(60)
                part = translate_df(part)
                ws.cell(row, 1, section_title).fill = SUBSECTION_FILL
                ws.cell(row, 1).font = Font(bold=True)
                row += 1
                for col_idx, col in enumerate(part.columns, start=1):
                    cell = ws.cell(row, col_idx, col)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = BORDER
                row += 1
                for rec in part.to_dict("records"):
                    for col_idx, col in enumerate(part.columns, start=1):
                        val = rec.get(col)
                        if isinstance(val, pd.Timestamp):
                            val = val.strftime("%Y-%m-%d")
                        cell = ws.cell(row, col_idx, val)
                        cell.border = BORDER
                        if isinstance(val, (int, float, np.number)) and not isinstance(val, bool):
                            if "%" in col or "pct" in col.lower() or "ДРР" in col or "CTR" in col:
                                cell.number_format = '0.0'
                            elif "цена" in col.lower() or "сум" in col.lower() or "profit" in col.lower() or "spend" in col.lower() or "ВП" in col:
                                cell.number_format = money_format()
                            else:
                                cell.number_format = '# ##0.00'
                    row += 1
                row += 2
            row += 1
        ws.freeze_panes = "A3"
        autofit(ws)
    wb.save(path)


def export_outputs(outputs: Dict[str, pd.DataFrame], local_dir: Path) -> List[Path]:
    local_dir.mkdir(parents=True, exist_ok=True)
    paths = []
    # Main concise report
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    ws.cell(1, 1, "Сводка по причинам и целям")
    ws.cell(1, 1).fill = TITLE_FILL
    ws.cell(1, 1).font = Font(color="FFFFFF", bold=True, size=14)
    cons = outputs.get("conclusions", pd.DataFrame())
    if not cons.empty:
        small_cols = [c for c in ["subject", "product", "supplier_article", "status", "main_reason", "recommendation", "localization_with_replacements_pct"] if c in cons.columns]
        x = translate_df(cons[small_cols].copy())
        ws.append(list(x.columns))
        for row in x.itertuples(index=False, name=None):
            ws.append(list(row))
        style_sheet(ws)
    p = local_dir / MAIN_REPORT_NAME
    wb.save(p)
    paths.append(p)
    # Technical report
    wb = Workbook()
    wb.remove(wb.active)
    for name in ["article_day_fact", "metrics_summary_90d", "best_days", "best_day_factors", "price_ranges", "channel_summary", "ads_category_source", "ads_raw_source", "ads_daily_source", "core_queries_80", "search_unique_demand", "entry_points_summary", "localization_summary", "localization_detail", "gp_potential_90d", "buyout_validation", "dictionary", "diagnostics"]:
        write_df_sheet(wb, name[:31], outputs.get(name, pd.DataFrame()))
    p = local_dir / TECH_REPORT_NAME
    wb.save(p)
    paths.append(p)
    # Example 901
    wb = Workbook()
    wb.remove(wb.active)
    for name in ["article_day_fact", "metrics_summary_90d", "best_days", "price_ranges", "core_queries_80"]:
        df = outputs.get(name, pd.DataFrame())
        if df is not None and not df.empty and "supplier_article" in df.columns:
            df = df[df["supplier_article"].isin(EXAMPLE_ARTICLES)].copy()
        write_df_sheet(wb, name[:31], df)
    p = local_dir / EXAMPLE_REPORT_NAME
    wb.save(p)
    paths.append(p)
    # Product block reports
    potential_sections = [
        ("Средние и целевые значения", "metrics_summary_90d", ["metric", "avg_90d_all_days", "avg_90d_nonzero_days", "target_above_mean_90d", "best_days_avg", "last_full_week_avg", "gap_to_target_pct"]),
        ("Лучшие дни по сумме заказов", "best_days", None),
        ("Факторы лучших дней", "best_day_factors", ["factor", "normal_days_avg", "best_days_avg", "difference_pct", "best_days_count", "conclusion"]),
        ("Рекомендуемый ценовой диапазон", "price_ranges", ["price_range", "days", "order_sum", "orders", "avg_price_with_disc", "avg_finished_price", "avg_gross_profit", "avg_drr_manual", "avg_drr_unified", "avg_cart_conv_pct", "avg_order_conv_pct", "is_recommended"]),
        ("Потенциал валовой прибыли по ABC", "gp_potential_90d", ["gross_profit_90d", "avg_gp_per_day", "target_gp_per_day", "best_week_gp_per_day", "prev_month_gross_profit", "plan_month", "current_month_gross_profit", "plan_to_date_completion_pct"]),
    ]
    p = local_dir / POTENTIAL_REPORT_NAME
    write_product_blocks(p, "Средние и целевые значения", outputs, potential_sections)
    paths.append(p)
    channel_sections = [
        ("Каналы продаж и реклама", "channel_summary", None),
        ("Точки входа", "entry_points_summary", None),
    ]
    p = local_dir / CHANNEL_REPORT_NAME
    write_product_blocks(p, "Каналы продаж и реклама", outputs, channel_sections)
    paths.append(p)
    search_sections = [
        ("Ключи, которые дают 80%+ заказов", "core_queries_80", None),
        ("Поиск по дням", "search_daily_summary", None),
    ]
    p = local_dir / SEARCH_REPORT_NAME
    write_product_blocks(p, "Поисковые запросы и позиции", outputs, search_sections)
    paths.append(p)
    loc_sections = [
        ("Сводка локализации", "localization_summary", None),
        ("Детализация складов", "localization_detail", None),
    ]
    p = local_dir / LOCALIZATION_REPORT_NAME
    write_product_blocks(p, "Локализация", outputs, loc_sections)
    paths.append(p)
    cons_sections = [
        ("Выводы", "conclusions", None),
        ("Факторы лучших дней", "best_day_factors", ["factor", "normal_days_avg", "best_days_avg", "difference_pct", "conclusion"]),
        ("Средние и целевые значения", "metrics_summary_90d", ["metric", "last_full_week_avg", "target_above_mean_90d", "gap_to_target_pct"]),
    ]
    p = local_dir / CONCLUSIONS_REPORT_NAME
    write_product_blocks(p, "Выводы по причинам", outputs, cons_sections)
    paths.append(p)
    return paths


def upload_to_storage(storage: Storage, local_paths: List[Path], root: str) -> None:
    for path in local_paths:
        rel = str(path.relative_to(Path(root))).replace("\\", "/") if Path(root) in path.parents or Path(root) == path.parent else f"{OUT_DIR}/{path.name}"
        try:
            storage.write_bytes(rel, path.read_bytes())
            log(f"Saved: {rel}")
        except Exception as exc:
            log(f"WARN: failed to save {rel}: {exc}")



# ------------------------- factor money bridge + PDF + Telegram -------------------------
FACTOR_REPORT_NAME = "Факторный_мост_ВП_TOPFACE.xlsx"
PDF_REPORT_NAME = "Управленческий_отчет_TOPFACE.pdf"

REPORT_MONTH_GENITIVE_RU = {
    1: "января", 2: "февраля", 3: "марта", 4: "апреля", 5: "мая", 6: "июня",
    7: "июля", 8: "августа", 9: "сентября", 10: "октября", 11: "ноября", 12: "декабря",
}

def _max_report_day_from_outputs(outputs: Dict[str, Any]) -> pd.Timestamp:
    """Return the latest real date represented in report outputs for PDF naming/caption."""
    max_day = pd.NaT
    for name in ["article_day_fact", "ads_daily_source", "search_unique_demand", "ads_raw_source"]:
        df = outputs.get(name) if isinstance(outputs, dict) else None
        if not isinstance(df, pd.DataFrame) or df.empty or "day" not in df.columns:
            continue
        days = pd.to_datetime(df["day"], errors="coerce").dropna()
        if days.empty:
            continue
        d = days.max().normalize()
        if pd.isna(max_day) or d > max_day:
            max_day = d
    if pd.isna(max_day):
        max_day = pd.Timestamp.today().normalize()
    return pd.Timestamp(max_day).normalize()

def report_date_label_ru(outputs: Dict[str, Any]) -> str:
    d = _max_report_day_from_outputs(outputs)
    return f"{int(d.day)} {REPORT_MONTH_GENITIVE_RU.get(int(d.month), d.strftime('%m'))}"

def sales_pdf_report_name(outputs: Dict[str, Any]) -> str:
    return f"Отчет по продажам Влад {report_date_label_ru(outputs)}.pdf"


PDF_ONLY_SHEETS = {
    TECH_REPORT_NAME: ["article_day_fact", "search_unique_demand", "ads_category_source", "ads_raw_source", "ads_daily_source", "gp_potential_90d"],
    FACTOR_REPORT_NAME: ["optimal_benchmarks", "factor_bridge", "entry_points_bridge", "factor_summary_for_pdf"],
}

# Exact reverse mapping for PDF-only mode. Excel files are exported with Russian headers,
# while generate_management_pdf expects the internal English column names.
PDF_ONLY_RU_TO_EN = {v: k for k, v in COLUMN_RU.items()}
PDF_ONLY_RU_TO_EN.update({
    "Уровень": "level",
    "Категория": "subject",
    "Товар": "product",
    "Артикул продавца": "supplier_article",
    "Артикул WB": "nm_id",
    "Дата": "day",
    "Фактор": "factor",
    "Комментарий": "comment",
    "Раздел": "entry_section",
    "Точка входа": "entry_point",
    "Переходы": "transitions",
    "Добавления в корзину": "add_to_cart",
    "Заказы": "orders",
    "Показы": "impressions",
    "CTR, %": "ctr_pct",
    "Конверсия в корзину, %": "cart_conv_pct",
    "Конверсия в заказ, %": "order_conv_pct",
    "Доля заказов, %": "orders_share_pct",
    "Заказы всего": "orders_total",
    "Валовая прибыль модель": "gross_profit_model",
    "Сумма заказов": "order_sum",
    "Открытия карточки / клики": "open_cards",
    "Расход manual": "manual_spend",
    "Расход unified": "unified_spend",
    "Расход без типа": "unknown_spend",
    "Клики manual": "manual_clicks",
    "Клики unified": "unified_clicks",
    "Клики без типа": "unknown_clicks",
    "Показы manual": "manual_impressions",
    "Показы unified": "unified_impressions",
    "Показы без типа": "unknown_impressions",
    "Заказы manual": "manual_orders",
    "Заказы unified": "unified_orders",
    "Заказы без типа": "unknown_orders",
    "Сумма заказов manual": "manual_order_sum",
    "Сумма заказов unified": "unified_order_sum",
    "Сумма заказов без типа": "unknown_order_sum",
    "Спрос / частотность": "search_frequency",
    "Переходы из поиска": "search_transitions",
    "Добавления из поиска": "search_add_to_cart",
    "Заказы из поиска": "search_orders",
    "% поискового трафика": "search_traffic_capture_pct",
    "% общего захвата спроса": "total_traffic_capture_pct",
    "Уникальный спрос WB": "unique_search_frequency",
    "Уникальные запросы": "unique_search_queries",
    "Дублей снято": "duplicate_query_rows_removed",
    "Локализация с заменами, %": "localization_with_replacements_pct",
    "Прямая локализация, %": "direct_localization_pct",
    "Статус локализации": "localization_status",
    "Остаток всего": "stock_qty_total",
    "Непокрытые склады": "uncovered_warehouses",
    "effect_type": "effect_type",
    "zone": "zone",
    "effect_gp_rub": "effect_gp_rub",
    "abs_effect_gp_rub": "abs_effect_gp_rub",
    "total_abs_effect": "total_abs_effect",
    "effect_weight_%": "effect_weight_pct",
})


def _normalize_pdf_only_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    out = out.rename(columns={c: PDF_ONLY_RU_TO_EN.get(str(c), str(c)) for c in out.columns})
    for c in ["day", "period_start", "period_end", "stock_day"]:
        if c in out.columns:
            out[c] = pd.to_datetime(out[c], errors="coerce")
    for c in ["subject", "product", "supplier_article"]:
        if c in out.columns:
            out[c] = out[c].astype(str).replace({"nan": "", "None": ""})
    return out


def _read_existing_report_bytes(storage: Storage, local_dir: Path, file_name: str) -> bytes:
    local_path = local_dir / file_name
    if local_path.exists():
        log(f"pdf_only: using local {local_path}")
        return local_path.read_bytes()
    local_root_fallback = Path(file_name)
    if local_root_fallback.exists():
        log(f"pdf_only: using local {local_root_fallback}")
        return local_root_fallback.read_bytes()
    if storage.is_s3:
        key = f"{OUT_DIR}/{file_name}"
        log(f"pdf_only: reading s3://*/{key}")
        return storage.read_bytes(key)
    raise FileNotFoundError(f"Не найден файл для PDF-only режима: {local_path}")


def load_existing_outputs_for_pdf(storage: Storage, local_dir: Path) -> Dict[str, pd.DataFrame]:
    """Load previously generated Excel outputs and rebuild the minimal outputs dict for PDF.

    This mode avoids the expensive S3 ingestion/recalculation stage. It expects that the full
    report has already generated these files in OUT_DIR either locally or in Yandex Object Storage:
    - Технические_расчеты_TOPFACE.xlsx / article_day_fact
    - Факторный_мост_ВП_TOPFACE.xlsx / factor_* sheets
    """
    outputs: Dict[str, pd.DataFrame] = {}
    for file_name, sheets in PDF_ONLY_SHEETS.items():
        data = _read_existing_report_bytes(storage, local_dir, file_name)
        xls = pd.ExcelFile(io.BytesIO(data))
        for sheet_name in sheets:
            if sheet_name not in xls.sheet_names:
                log(f"WARN pdf_only: sheet {sheet_name} missing in {file_name}")
                outputs[sheet_name] = pd.DataFrame()
                continue
            df = pd.read_excel(xls, sheet_name=sheet_name)
            outputs[sheet_name] = _normalize_pdf_only_df(df)
            log(f"pdf_only: loaded {file_name}/{sheet_name}: rows={len(outputs[sheet_name]):,}, cols={len(outputs[sheet_name].columns):,}")
    daily = outputs.get("article_day_fact", pd.DataFrame())
    if daily.empty:
        raise RuntimeError("PDF-only режим невозможен: пустой article_day_fact в техническом файле")
    required = ["day", "subject", "product", "supplier_article", "nm_id", "order_sum", "gross_profit_model"]
    missing = [c for c in required if c not in daily.columns]
    if missing:
        raise RuntimeError(f"PDF-only режим невозможен: в article_day_fact нет колонок {missing}")
    return outputs


def _period_bounds_from_daily(daily: pd.DataFrame) -> Tuple[pd.Timestamp, pd.Timestamp, pd.Timestamp, pd.Timestamp]:
    if daily is None or daily.empty or "day" not in daily.columns:
        today = pd.Timestamp(datetime.today().date())
        last_monday = today - pd.Timedelta(days=int(today.weekday()))
        if today.weekday() == 6:
            week_start, week_end = last_monday, today
        else:
            week_start, week_end = last_monday - pd.Timedelta(days=7), last_monday - pd.Timedelta(days=1)
        return week_start, week_end, week_start - pd.Timedelta(days=7), week_start - pd.Timedelta(days=1)
    mx = pd.to_datetime(daily["day"], errors="coerce").max()
    if pd.isna(mx):
        mx = pd.Timestamp(datetime.today().date())
    mx = pd.Timestamp(mx).normalize()
    last_monday = mx - pd.Timedelta(days=int(mx.weekday()))
    if mx.weekday() == 6:
        week_start, week_end = last_monday, mx
    else:
        week_start, week_end = last_monday - pd.Timedelta(days=7), last_monday - pd.Timedelta(days=1)
    return week_start, week_end, week_start - pd.Timedelta(days=7), week_start - pd.Timedelta(days=1)


def _agg_daily_for_bridge(daily: pd.DataFrame, start: pd.Timestamp, end: pd.Timestamp, group_cols: List[str]) -> pd.DataFrame:
    base_sum_cols = [
        "orders", "order_sum", "open_cards", "add_to_cart", "buyouts_count", "cancels_count",
        "manual_impressions", "manual_clicks", "manual_spend", "manual_orders", "manual_order_sum",
        "unified_impressions", "unified_clicks", "unified_spend", "unified_orders", "unified_order_sum",
        "unknown_impressions", "unknown_clicks", "unknown_spend", "unknown_orders", "unknown_order_sum",
        "search_frequency", "search_transitions", "search_add_to_cart", "search_orders",
        "ad_spend_model", "gross_profit_model", "buyout_qty_model", "revenue_model",
        "commission_model", "acquiring_model", "logistics_direct_model", "logistics_return_model",
        "storage_model", "other_costs_model", "cost_model"
    ]
    base_mean_cols = ["finished_price", "price_with_disc", "spp", "finished_price_funnel", "spp_funnel", "rating_reviews", "direct_localization_pct", "localization_with_replacements_pct"]
    derived_cols = ["ad_spend_total", "ad_clicks_total", "ad_impressions_total", "drr_pct", "cpc", "ctr_pct", "cart_conv_pct", "order_conv_pct", "card_to_order_pct", "search_traffic_capture_pct", "avg_order_price"]

    if daily is None or daily.empty:
        return pd.DataFrame(columns=group_cols + base_sum_cols + base_mean_cols + derived_cols)
    x = daily.copy()
    for c in group_cols:
        if c not in x.columns:
            x[c] = ""
    sum_cols = [c for c in base_sum_cols if c in x.columns]
    mean_cols = [c for c in base_mean_cols if c in x.columns]
    empty_cols = group_cols + sum_cols + mean_cols + derived_cols

    x["day"] = pd.to_datetime(x["day"], errors="coerce").dt.normalize()
    x = x[(x["day"] >= start) & (x["day"] <= end)].copy()
    if x.empty:
        return pd.DataFrame(columns=empty_cols)

    agg = {c: (c, "sum") for c in sum_cols}
    for c in mean_cols:
        agg[c] = (c, "mean")
    g = x.groupby(group_cols, dropna=False, as_index=False).agg(**agg)
    g["ad_spend_total"] = sum((g[c] if c in g.columns else 0) for c in ["manual_spend", "unified_spend", "unknown_spend", "ad_spend_model"])
    # Avoid double-count if ad_spend_model already includes manual+unified and channels exist.
    channel_spend = sum((g[c] if c in g.columns else 0) for c in ["manual_spend", "unified_spend", "unknown_spend"])
    if "ad_spend_model" in g.columns and float(pd.to_numeric(channel_spend, errors="coerce").fillna(0).sum()) > 0:
        g["ad_spend_total"] = channel_spend
    g["ad_clicks_total"] = sum((g[c] if c in g.columns else 0) for c in ["manual_clicks", "unified_clicks", "unknown_clicks"])
    g["ad_impressions_total"] = sum((g[c] if c in g.columns else 0) for c in ["manual_impressions", "unified_impressions", "unknown_impressions"])

    zero = pd.Series(0.0, index=g.index)
    def _num_col(col: str) -> pd.Series:
        return pd.to_numeric(g[col], errors="coerce").fillna(0) if col in g.columns else zero

    order_sum = _num_col("order_sum")
    orders = _num_col("orders")
    open_cards = _num_col("open_cards")
    add_to_cart = _num_col("add_to_cart")
    search_frequency = _num_col("search_frequency")
    search_transitions = _num_col("search_transitions")

    g["drr_pct"] = np.where(order_sum > 0, g["ad_spend_total"] / order_sum * 100, np.nan)
    g["cpc"] = np.where(g["ad_clicks_total"] > 0, g["ad_spend_total"] / g["ad_clicks_total"], np.nan)
    g["ctr_pct"] = np.where(g["ad_impressions_total"] > 0, g["ad_clicks_total"] / g["ad_impressions_total"] * 100, np.nan)
    g["cart_conv_pct"] = np.where(open_cards > 0, add_to_cart / open_cards * 100, np.nan)
    g["order_conv_pct"] = np.where(add_to_cart > 0, orders / add_to_cart * 100, np.nan)
    g["card_to_order_pct"] = np.where(open_cards > 0, orders / open_cards * 100, np.nan)
    g["search_traffic_capture_pct"] = np.where(search_frequency > 0, search_transitions / search_frequency * 100, np.nan)
    g["avg_order_price"] = np.where(orders > 0, order_sum / orders, np.nan)
    return g


def _abc_gp_for_period(builder: AnalyticsBuilder, start: pd.Timestamp, end: pd.Timestamp, group_cols: List[str]) -> pd.DataFrame:
    """Strict ABC gross-profit lookup for PDF/reporting.

    Important: do NOT use overlapping weekly ABC files for arbitrary periods.
    The previous implementation silently summed every ABC week overlapping the requested
    period; for monthly / partial periods this double-counted days outside the period and
    made PDF gross profit disagree with ABC. Now ABC fact is used only for exact weekly
    periods, or exact closed monthly periods. Other periods must use the model and be
    labelled as calculated GP.
    """
    if builder is None:
        return pd.DataFrame(columns=group_cols + ["gp_fact", "gross_revenue_fact", "sales_qty_fact"])
    start = pd.Timestamp(start).normalize()
    end = pd.Timestamp(end).normalize()
    # There is no daily ABC fact in the source files; never try to merge weekly ABC into day rows.
    if "day" in group_cols:
        return pd.DataFrame(columns=group_cols + ["gp_fact", "gross_revenue_fact", "sales_qty_fact", "gp_source"])

    frames = []
    # Exact weekly fact.
    abc_w = builder.enrich(builder.pack.abc_weekly, "abc_weekly")
    if abc_w is not None and not abc_w.empty:
        w = abc_w.copy()
        w["period_start"] = pd.to_datetime(w["period_start"], errors="coerce").dt.normalize()
        w["period_end"] = pd.to_datetime(w["period_end"], errors="coerce").dt.normalize()
        w = w[(w["period_start"] == start) & (w["period_end"] == end)].copy()
        if not w.empty:
            w["gp_source"] = "ABC_weekly_exact"
            frames.append(w)
    # Exact monthly fact for closed full months.
    abc_m = builder.enrich(builder.pack.abc_monthly, "abc_monthly")
    if abc_m is not None and not abc_m.empty:
        m = abc_m.copy()
        m["period_start"] = pd.to_datetime(m["period_start"], errors="coerce").dt.normalize()
        m["period_end"] = pd.to_datetime(m["period_end"], errors="coerce").dt.normalize()
        m = m[(m["period_start"] == start) & (m["period_end"] == end)].copy()
        if not m.empty:
            m["gp_source"] = "ABC_monthly_exact"
            frames.append(m)
    if not frames:
        return pd.DataFrame(columns=group_cols + ["gp_fact", "gross_revenue_fact", "sales_qty_fact", "gp_source"])
    exact = pd.concat(frames, ignore_index=True)
    for c in group_cols:
        if c not in exact.columns:
            exact[c] = ""
    g = exact.groupby(group_cols, dropna=False, as_index=False).agg(
        gp_fact=("gross_profit", "sum"),
        gross_revenue_fact=("gross_revenue", "sum"),
        sales_qty_fact=("orders", "sum"),
        gp_source=("gp_source", "first"),
    )
    return g


def _merge_cur_prev(cur: pd.DataFrame, prev: pd.DataFrame, keys: List[str]) -> pd.DataFrame:
    cur = cur.copy() if cur is not None else pd.DataFrame(columns=keys)
    prev = prev.copy() if prev is not None else pd.DataFrame(columns=keys)
    for k in keys:
        if k not in cur.columns:
            cur[k] = ""
        if k not in prev.columns:
            prev[k] = ""
    value_cols = sorted(set([c for c in cur.columns if c not in keys]) | set([c for c in prev.columns if c not in keys]))
    for c in value_cols:
        if c not in cur.columns:
            cur[c] = np.nan
        if c not in prev.columns:
            prev[c] = np.nan
    out = cur.merge(prev, on=keys, how="outer", suffixes=("", "_prev"))
    for c in value_cols:
        if c not in out.columns:
            out[c] = np.nan
        pc = f"{c}_prev"
        if pc not in out.columns:
            out[pc] = np.nan
    return out.fillna(0)


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return default
        return float(value)
    except Exception:
        return default


def _factor_row(level: str, keys: Dict[str, Any], factor: str, was: Any, now: Any, change: Any, effect: float, zone: str, comment: str) -> Dict[str, Any]:
    rec = {"level": level, **keys}
    rec.update({
        "factor": factor, "was": was, "now": now, "change": change,
        "effect_gp_rub": float(effect) if pd.notna(effect) else 0.0,
        "effect_type": "плюс" if effect > 0 else ("минус" if effect < 0 else "нейтрально"),
        "zone": zone, "comment": comment,
    })
    return rec


def _entity_factor_rows(level: str, g: pd.DataFrame, keys: List[str]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for _, r in g.iterrows():
        keyvals = {k: r.get(k, "") for k in keys}
        cur_sum = _safe_float(r.get("order_sum"))
        prev_sum = _safe_float(r.get("order_sum_prev"))
        cur_gp = _safe_float(r.get("gp_fact"), _safe_float(r.get("gross_profit_model")))
        prev_gp = _safe_float(r.get("gp_fact_prev"), _safe_float(r.get("gross_profit_model_prev")))
        cur_margin = cur_gp / cur_sum if cur_sum else 0.0
        prev_margin = prev_gp / prev_sum if prev_sum else cur_margin
        cur_orders = _safe_float(r.get("orders"))
        avg_price = cur_sum / cur_orders if cur_orders else _safe_float(r.get("avg_order_price"), 0)
        # 1) Volume/order sum effect.
        volume_effect = (cur_sum - prev_sum) * prev_margin
        rows.append(_factor_row(level, keyvals, "Объём / сумма заказов", prev_sum, cur_sum, cur_sum - prev_sum, volume_effect, "рынок + управление", "Сколько ВП изменилось из-за изменения суммы заказов при прежней рентабельностинальности."))
        # 2) Margin total effect.
        margin_effect = cur_sum * (cur_margin - prev_margin)
        rows.append(_factor_row(level, keyvals, "Рентабельность", prev_margin * 100, cur_margin * 100, (cur_margin - prev_margin) * 100, margin_effect, "экономика", "Изменение ВП из-за изменения рентабельностинальности после расходов и ABC-факта."))
        # 3) Advertising/Drr effect.
        cur_drr = _safe_float(r.get("drr_pct"), np.nan)
        prev_drr = _safe_float(r.get("drr_pct_prev"), np.nan)
        if pd.notna(cur_drr) and pd.notna(prev_drr):
            drr_effect = -cur_sum * ((cur_drr - prev_drr) / 100.0)
            rows.append(_factor_row(level, keyvals, "ДРР / рекламная нагрузка", prev_drr, cur_drr, cur_drr - prev_drr, drr_effect, "управляемый", "Сколько ВП забрал или добавил сдвиг ДРР относительно прошлой недели."))
        # 4) Price effect.
        cur_price = _safe_float(r.get("avg_order_price"), _safe_float(r.get("finished_price")))
        prev_price = _safe_float(r.get("avg_order_price_prev"), _safe_float(r.get("finished_price_prev")))
        if cur_orders and prev_price:
            price_effect = cur_orders * (cur_price - prev_price) * prev_margin
            rows.append(_factor_row(level, keyvals, "Цена продажи", prev_price, cur_price, cur_price - prev_price, price_effect, "управляемый", "Эффект изменения продажной цены при текущем объёме заказов."))
        # 5) WB buyer price / SPP.
        cur_spp = _safe_float(r.get("spp"), _safe_float(r.get("spp_funnel")))
        prev_spp = _safe_float(r.get("spp_prev"), _safe_float(r.get("spp_funnel_prev")))
        if cur_spp or prev_spp:
            # Higher SPP usually makes buyer price better; this is an explanatory external/partly external factor.
            spp_effect = cur_sum * ((cur_spp - prev_spp) / 100.0) * 0.30 * prev_margin
            rows.append(_factor_row(level, keyvals, "СПП / цена покупателя", prev_spp, cur_spp, cur_spp - prev_spp, spp_effect, "WB / внешний", "Оценка влияния изменения СПП на привлекательность цены и ВП."))
        # 6) Unit expenses.
        qty = _safe_float(r.get("buyout_qty_model"), _safe_float(r.get("sales_qty_fact"), cur_orders))
        for title, col, zone in [
            ("Комиссия WB/шт", "commission_model", "WB / экономика"),
            ("Эквайринг/шт", "acquiring_model", "WB / экономика"),
            ("Логистика/шт", "logistics_direct_model", "WB / логистика"),
            ("Обратная логистика/шт", "logistics_return_model", "WB / логистика"),
            ("Хранение/шт", "storage_model", "WB / логистика"),
            ("Себестоимость/шт", "cost_model", "управляемый"),
            ("Прочие расходы/шт", "other_costs_model", "экономика"),
        ]:
            cur_total = _safe_float(r.get(col))
            prev_total = _safe_float(r.get(f"{col}_prev"))
            cur_unit = cur_total / qty if qty else 0.0
            prev_qty = _safe_float(r.get("buyout_qty_model_prev"), _safe_float(r.get("sales_qty_fact_prev"), _safe_float(r.get("orders_prev"))))
            prev_unit = prev_total / prev_qty if prev_qty else 0.0
            if cur_unit or prev_unit:
                effect = -qty * (cur_unit - prev_unit)
                rows.append(_factor_row(level, keyvals, title, prev_unit, cur_unit, cur_unit - prev_unit, effect, zone, "Эффект изменения расхода на единицу."))
        # 7) Demand, traffic share and conversions.
        demand_cur = _safe_float(r.get("search_frequency"))
        demand_prev = _safe_float(r.get("search_frequency_prev"))
        capture_cur = _safe_float(r.get("search_traffic_capture_pct")) / 100.0
        capture_prev = _safe_float(r.get("search_traffic_capture_pct_prev")) / 100.0
        cart_cur = _safe_float(r.get("cart_conv_pct")) / 100.0
        cart_prev = _safe_float(r.get("cart_conv_pct_prev")) / 100.0
        order_cur = _safe_float(r.get("order_conv_pct")) / 100.0
        order_prev = _safe_float(r.get("order_conv_pct_prev")) / 100.0
        if demand_cur or demand_prev:
            demand_effect = (demand_cur - demand_prev) * capture_prev * cart_prev * order_prev * avg_price * prev_margin
            rows.append(_factor_row(level, keyvals, "Спрос WB", demand_prev, demand_cur, demand_cur - demand_prev, demand_effect, "внешний", "Эффект изменения общего поискового спроса WB."))
        if demand_cur and (capture_cur or capture_prev):
            traffic_effect = demand_cur * (capture_cur - capture_prev) * cart_prev * order_prev * avg_price * prev_margin
            rows.append(_factor_row(level, keyvals, "% поискового трафика", capture_prev * 100, capture_cur * 100, (capture_cur - capture_prev) * 100, traffic_effect, "управляемый", "Эффект изменения доли спроса, которую забрала карточка."))
        opens = _safe_float(r.get("open_cards"))
        if opens and (cart_cur or cart_prev):
            cart_effect = opens * (cart_cur - cart_prev) * order_prev * avg_price * prev_margin
            rows.append(_factor_row(level, keyvals, "Конверсия в корзину", cart_prev * 100, cart_cur * 100, (cart_cur - cart_prev) * 100, cart_effect, "карточка", "Сколько ВП изменилось из-за входной конверсии карточки."))
        if opens and (order_cur or order_prev):
            order_effect = opens * cart_cur * (order_cur - order_prev) * avg_price * prev_margin
            rows.append(_factor_row(level, keyvals, "Корзина -> заказ", order_prev * 100, order_cur * 100, (order_cur - order_prev) * 100, order_effect, "карточка / цена / доставка", "Сколько ВП изменилось из-за дожима из корзины в заказ."))
    return rows


def compute_optimal_benchmarks(outputs: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    daily = outputs.get("article_day_fact", pd.DataFrame())
    if daily is None or daily.empty:
        return pd.DataFrame()
    x = daily.copy()
    x["day"] = pd.to_datetime(x["day"], errors="coerce").dt.normalize()
    rows = []
    group_cols = ["subject", "product", "supplier_article", "nm_id"]
    for keys, part in x.groupby(group_cols, dropna=False):
        p = part.sort_values("day").copy()
        if p.empty:
            continue
        positive = p[p["order_sum"].fillna(0) > 0].copy()
        if positive.empty:
            positive = p.copy()
        threshold = positive["order_sum"].quantile(0.80) if len(positive) >= 5 else positive["order_sum"].mean()
        best = positive[positive["order_sum"] >= threshold].copy()
        if best.empty:
            best = positive.nlargest(min(3, len(positive)), "order_sum")
        def ratio(num, den, mul=1.0):
            den_sum = pd.to_numeric(best.get(den, 0), errors="coerce").sum()
            if not den_sum:
                return np.nan
            return pd.to_numeric(best.get(num, 0), errors="coerce").sum() / den_sum * mul
        ad_spend = best[[c for c in ["manual_spend", "unified_spend", "unknown_spend", "ad_spend_model"] if c in best.columns]].sum(axis=1)
        clicks = best[[c for c in ["manual_clicks", "unified_clicks", "unknown_clicks"] if c in best.columns]].sum(axis=1)
        impressions = best[[c for c in ["manual_impressions", "unified_impressions", "unknown_impressions"] if c in best.columns]].sum(axis=1)
        order_sum = pd.to_numeric(best.get("order_sum", 0), errors="coerce").sum()
        rec = dict(zip(group_cols, keys))
        rec.update({
            "best_days_count": len(best),
            "optimal_order_sum_day": pd.to_numeric(best.get("order_sum", 0), errors="coerce").mean(),
            "optimal_orders_day": pd.to_numeric(best.get("orders", 0), errors="coerce").mean(),
            "optimal_drr_pct": ad_spend.sum() / order_sum * 100 if order_sum else np.nan,
            "optimal_cpc": ad_spend.sum() / clicks.sum() if clicks.sum() else np.nan,
            "optimal_ctr_pct": clicks.sum() / impressions.sum() * 100 if impressions.sum() else np.nan,
            "optimal_cart_conv_pct": ratio("add_to_cart", "open_cards", 100),
            "optimal_order_conv_pct": ratio("orders", "add_to_cart", 100),
            "optimal_search_capture_pct": ratio("search_transitions", "search_frequency", 100),
            "optimal_price_sale": ratio("order_sum", "orders", 1),
            "optimal_spp": pd.to_numeric(best.get("spp", best.get("spp_funnel", np.nan)), errors="coerce").mean(),
        })
        rows.append(rec)
    return pd.DataFrame(rows)


def compute_entry_points_bridge(builder: AnalyticsBuilder, week_start: pd.Timestamp, week_end: pd.Timestamp, prev_start: pd.Timestamp, prev_end: pd.Timestamp) -> pd.DataFrame:
    e = builder.enrich(builder.pack.entry_points, "entry_points")
    if e is None or e.empty:
        return pd.DataFrame()
    e = e.copy()
    e["day"] = pd.to_datetime(e["day"], errors="coerce").dt.normalize()
    group_cols = ["subject", "product", "supplier_article", "nm_id", "entry_section", "entry_point"]
    metric_cols = ["impressions", "transitions", "add_to_cart", "orders", "ctr_pct", "cart_conv_pct", "order_conv_pct"]
    def agg(start, end):
        x = e[(e["day"] >= start) & (e["day"] <= end)].copy()
        if x.empty:
            return pd.DataFrame(columns=group_cols + metric_cols)
        g = x.groupby(group_cols, dropna=False, as_index=False).agg(
            impressions=("impressions", "sum"), transitions=("transitions", "sum"),
            add_to_cart=("add_to_cart", "sum"), orders=("orders", "sum"),
        )
        g["ctr_pct"] = np.where(g["impressions"] > 0, g["transitions"] / g["impressions"] * 100, np.nan)
        g["cart_conv_pct"] = np.where(g["transitions"] > 0, g["add_to_cart"] / g["transitions"] * 100, np.nan)
        g["order_conv_pct"] = np.where(g["add_to_cart"] > 0, g["orders"] / g["add_to_cart"] * 100, np.nan)
        return g
    cur = agg(week_start, week_end)
    prev = agg(prev_start, prev_end)
    out = _merge_cur_prev(cur, prev, group_cols)
    for c in metric_cols:
        if c not in out.columns:
            out[c] = 0.0
        if f"{c}_prev" not in out.columns:
            out[f"{c}_prev"] = 0.0
        out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0)
        out[f"{c}_prev"] = pd.to_numeric(out[f"{c}_prev"], errors="coerce").fillna(0)
    # Join article margin and avg price from daily data for ₽ effect.
    daily = outputs_global_for_bridge.get("article_day_fact", pd.DataFrame()) if "outputs_global_for_bridge" in globals() else pd.DataFrame()
    if daily is not None and not daily.empty:
        ag = _agg_daily_for_bridge(daily, week_start, week_end, ["subject", "product", "supplier_article", "nm_id"])
        gp_cur = _abc_gp_for_period(builder, week_start, week_end, ["subject", "product", "supplier_article", "nm_id"])
        ag = ag.merge(gp_cur, on=["subject", "product", "supplier_article", "nm_id"], how="left")
        ag["gp_use"] = ag["gp_fact"].fillna(ag.get("gross_profit_model", 0))
        ag["margin_use"] = np.where(ag["order_sum"] > 0, ag["gp_use"] / ag["order_sum"], 0)
        ag["avg_price_use"] = np.where(ag["orders"] > 0, ag["order_sum"] / ag["orders"], 0)
        out = out.merge(ag[["subject", "product", "supplier_article", "nm_id", "margin_use", "avg_price_use"]], on=["subject", "product", "supplier_article", "nm_id"], how="left")
    else:
        out["margin_use"] = 0.0
        out["avg_price_use"] = 0.0
    out["delta_transitions"] = out["transitions"] - out["transitions_prev"]
    out["delta_orders"] = out["orders"] - out["orders_prev"]
    out["delta_cart_conv_pp"] = out["cart_conv_pct"] - out["cart_conv_pct_prev"]
    out["delta_order_conv_pp"] = out["order_conv_pct"] - out["order_conv_pct_prev"]
    out["effect_gp_rub"] = out["delta_orders"] * out["avg_price_use"].fillna(0) * out["margin_use"].fillna(0)
    totals = out.groupby(["subject", "product", "supplier_article", "nm_id"], as_index=False)["orders"].sum().rename(columns={"orders": "orders_total"})
    out = out.merge(totals, on=["subject", "product", "supplier_article", "nm_id"], how="left")
    out["orders_share_pct"] = np.where(out["orders_total"] > 0, out["orders"] / out["orders_total"] * 100, np.nan)
    return out.sort_values(["subject", "product", "supplier_article", "effect_gp_rub"], ascending=[True, True, True, False])


def build_factor_outputs(builder: AnalyticsBuilder, outputs: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    global outputs_global_for_bridge
    outputs_global_for_bridge = outputs
    daily = outputs.get("article_day_fact", pd.DataFrame())
    week_start, week_end, prev_start, prev_end = _period_bounds_from_daily(daily)
    optimal = compute_optimal_benchmarks(outputs)
    factor_rows: List[Dict[str, Any]] = []
    for level, keys in [
        ("category", ["subject"]),
        ("product", ["subject", "product"]),
        ("article", ["subject", "product", "supplier_article", "nm_id"]),
    ]:
        cur = _agg_daily_for_bridge(daily, week_start, week_end, keys)
        prev = _agg_daily_for_bridge(daily, prev_start, prev_end, keys)
        cur_gp = _abc_gp_for_period(builder, week_start, week_end, keys)
        prev_gp = _abc_gp_for_period(builder, prev_start, prev_end, keys).rename(columns={"gp_fact": "gp_fact_prev", "gross_revenue_fact": "gross_revenue_fact_prev", "sales_qty_fact": "sales_qty_fact_prev"})
        g = _merge_cur_prev(cur, prev, keys)
        if not cur_gp.empty:
            g = g.merge(cur_gp, on=keys, how="left")
        if not prev_gp.empty:
            g = g.merge(prev_gp, on=keys, how="left")
        for c in ["gp_fact", "gp_fact_prev", "gross_revenue_fact", "gross_revenue_fact_prev", "sales_qty_fact", "sales_qty_fact_prev"]:
            if c not in g.columns:
                g[c] = np.nan
        factor_rows.extend(_entity_factor_rows(level, g, keys))
    factor_bridge = pd.DataFrame(factor_rows)
    if not factor_bridge.empty:
        factor_bridge["abs_effect_gp_rub"] = factor_bridge["effect_gp_rub"].abs()
        totals = factor_bridge.groupby(["level", "subject", "product", "supplier_article", "nm_id"], dropna=False)["abs_effect_gp_rub"].sum().rename("total_abs_effect").reset_index()
        factor_bridge = factor_bridge.merge(totals, on=["level", "subject", "product", "supplier_article", "nm_id"], how="left")
        factor_bridge["effect_weight_pct"] = np.where(factor_bridge["total_abs_effect"] > 0, factor_bridge["abs_effect_gp_rub"] / factor_bridge["total_abs_effect"] * 100, 0)
        factor_bridge = factor_bridge.sort_values(["level", "subject", "product", "supplier_article", "abs_effect_gp_rub"], ascending=[True, True, True, True, False])
    entry_bridge = compute_entry_points_bridge(builder, week_start, week_end, prev_start, prev_end)
    # Human-readable summary for PDF: top-4 money factors per entity.
    summary_rows = []
    if not factor_bridge.empty:
        for keys, part in factor_bridge.groupby(["level", "subject", "product", "supplier_article", "nm_id"], dropna=False):
            lvl, subject, product, art, nm_id = keys
            p = part[part["abs_effect_gp_rub"] > 50].sort_values("abs_effect_gp_rub", ascending=False).head(4)
            if p.empty:
                text = "Критичных денежных факторов не выделено: изменение ВП находится в рабочем диапазоне."
            else:
                phrases = []
                for _, r in p.iterrows():
                    val = float(r["effect_gp_rub"])
                    sign = "добавил" if val > 0 else "забрал"
                    phrases.append(f"{r['factor']}: {sign} около {abs(val):,.0f} ₽ ВП".replace(",", " "))
                text = "; ".join(phrases) + "."
            summary_rows.append({
                "level": lvl, "subject": subject, "product": product, "supplier_article": art, "nm_id": nm_id,
                "period": f"{week_start.strftime('%d.%m')}-{week_end.strftime('%d.%m.%Y')}",
                "compare_period": f"{prev_start.strftime('%d.%m')}-{prev_end.strftime('%d.%m.%Y')}",
                "summary_text": text,
            })
    factor_summary = pd.DataFrame(summary_rows)
    return {
        "optimal_benchmarks": optimal,
        "factor_bridge": factor_bridge,
        "entry_points_bridge": entry_bridge,
        "factor_summary_for_pdf": factor_summary,
    }


def write_factor_report(path: Path, factor_outputs: Dict[str, pd.DataFrame]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, df in factor_outputs.items():
        write_df_sheet(wb, sheet_name[:31], df if df is not None else pd.DataFrame())
    wb.save(path)


def _fmt_rub(x: Any, short: bool = False) -> str:
    try:
        val = float(x)
    except Exception:
        return "-"
    if pd.isna(val):
        return "-"
    if short and abs(val) >= 1000:
        return f"{val/1000:.0f}к ₽".replace(".", ",")
    return f"{val:,.0f} ₽".replace(",", " ")


def _fmt_num_pdf(x: Any) -> str:
    try:
        val = float(x)
    except Exception:
        return "-"
    if pd.isna(val):
        return "-"
    return f"{val:,.0f}".replace(",", " ")


def _fmt_pct_pdf(x: Any, digits: int = 1) -> str:
    try:
        val = float(x)
    except Exception:
        return "-"
    if pd.isna(val):
        return "-"
    return f"{val:.{digits}f}%".replace(".", ",")


def _fmt_cpc_pdf(x: Any) -> str:
    try:
        val = float(x)
    except Exception:
        return "-"
    if pd.isna(val):
        return "-"
    return f"{val:.1f} ₽".replace(".", ",")


def _delta_pct(cur: Any, prev: Any) -> Optional[float]:
    try:
        cur = float(cur); prev = float(prev)
        if pd.isna(cur) or pd.isna(prev) or abs(prev) < 1e-9:
            return None
        return (cur / prev - 1) * 100
    except Exception:
        return None


def _register_topface_fonts():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    candidates = [
        (os.getenv("TOPFACE_FONT_REGULAR"), os.getenv("TOPFACE_FONT_BOLD"), os.getenv("TOPFACE_FONT_BLACK")),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf", "/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf", "/usr/share/fonts/truetype/noto/NotoSans-Black.ttf"),
    ]
    for reg, bold, black in candidates:
        if reg and bold and black and Path(reg).exists() and Path(bold).exists() and Path(black).exists():
            pdfmetrics.registerFont(TTFont("TFReg", reg))
            pdfmetrics.registerFont(TTFont("TFBold", bold))
            pdfmetrics.registerFont(TTFont("TFBlack", black))
            return "TFReg", "TFBold", "TFBlack"
    return "Helvetica", "Helvetica-Bold", "Helvetica-Bold"




PRODUCT_GROUP_AUDIT_NAME = "Проверить_товарные_группы_TOPFACE.xlsx"
PDF_CALC_TRACE_NAME = "Лог_расчетов_PDF_TOPFACE.xlsx"

# Строгий справочник для PDF. Он важнее автоматического product_code():
# если товар/артикул не подтвержден здесь, в управленческий PDF он не попадает.
# Так подводки 405/406 не будут попадать в "Карандаши", а случайные товары вроде 552 уйдут в техлист.
PDF_PRODUCT_CATEGORY_REFERENCE: Dict[str, str] = VALID_PRODUCT_CATEGORY_REFERENCE.copy()

PDF_EXCLUDED_PRODUCT_REASONS: Dict[str, str] = {
    "405": "Подводки/лайнеры: не включать в категорию 'Косметические карандаши' для PDF",
    "406": "Подводки/лайнеры: не включать в категорию 'Косметические карандаши' для PDF",
}

# Products can remain in top category totals, but must not receive detailed PDF pages.
# Defaults reflect the current management rule: low-tail products 206/207/209/210/211
# are not detailed objects in the PDF. Override with PDF_FORCE_EXCLUDE_PRODUCTS if needed.
PDF_FORCE_EXCLUDE_DETAIL_PRODUCTS = set(
    p.strip() for p in os.getenv("PDF_FORCE_EXCLUDE_PRODUCTS", "206,207,209,210,211").split(",") if p.strip()
)

# Products that must appear in the detailed PDF when they have current-week sales/GP.
# This prevents useful pencil groups 605/611/613 from disappearing due to global 90% trimming.
PDF_FORCE_INCLUDE_DETAIL_PRODUCTS = set(
    p.strip() for p in os.getenv("PDF_FORCE_INCLUDE_PRODUCTS", "901,605,611,613,614,617,618,154,155,156,157").split(",") if p.strip()
)


def _pdf_product_code_from_value(value: Any) -> str:
    """Return a strict product code candidate for PDF category validation."""
    text = normalize_text(value).upper().replace(" ", "")
    if not text or text in {"NAN", "NONE", "NULL", "-"}:
        return ""
    text = text.replace("_", "/")
    # PT901.F25 -> 901; PT156.001 -> 156; 901/22 -> 901; 405/черный -> 405.
    m = re.match(r"^PT(\d+)", text)
    if m:
        return m.group(1)
    m = re.match(r"^(\d+)", text)
    if m:
        return m.group(1)
    return ""


def _pdf_resolve_product_category(article: Any = "", product: Any = "") -> Tuple[str, str, bool, str]:
    """Resolve PDF product/category only through the approved reference.

    Returns: (canonical_subject, canonical_product, keep, reason).
    """
    code = _pdf_product_code_from_value(article) or _pdf_product_code_from_value(product)
    if not code:
        return "", "", False, "Не удалось распознать товар по артикулу/товару"
    if code in PDF_PRODUCT_CATEGORY_REFERENCE:
        return PDF_PRODUCT_CATEGORY_REFERENCE[code], code, True, ""
    if code in PDF_EXCLUDED_PRODUCT_REASONS:
        return "", code, False, PDF_EXCLUDED_PRODUCT_REASONS[code]
    return "", code, False, "Товар отсутствует в утвержденном справочнике PDF"


def _filter_df_by_pdf_product_reference(df: pd.DataFrame, source_name: str, rejects: List[Dict[str, Any]]) -> pd.DataFrame:
    """Keep only rows whose product/category is approved for the PDF report.

    Rows without article/product are treated as category-level rows and are kept only for target subjects.
    Rows with product/article are validated by PDF_PRODUCT_CATEGORY_REFERENCE; canonical subject/product override source values.
    """
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    for col in ["subject", "product", "supplier_article", "nm_id"]:
        if col not in out.columns:
            out[col] = "" if col != "nm_id" else np.nan
    keep_mask: List[bool] = []
    canonical_subjects: List[Any] = []
    canonical_products: List[Any] = []
    rejected_count = 0
    for _, row in out.iterrows():
        subject_src = canonical_subject(row.get("subject", ""))
        product_src = normalize_text(row.get("product", ""))
        article_src = clean_article(row.get("supplier_article", ""))
        has_product_level = bool(_pdf_product_code_from_value(article_src) or _pdf_product_code_from_value(product_src))
        if not has_product_level:
            keep = canonical_subject(subject_src) in TARGET_SUBJECTS
            keep_mask.append(keep)
            canonical_subjects.append(subject_src)
            canonical_products.append(product_src)
            if not keep:
                rejected_count += 1
                rejects.append({
                    "source": source_name,
                    "subject_src": subject_src,
                    "product_src": product_src,
                    "supplier_article": article_src,
                    "nm_id": row.get("nm_id", ""),
                    "resolved_product": "",
                    "reason": "Строка без товара/артикула и вне целевых категорий PDF",
                    "rows": 1,
                })
            continue
        resolved_subject, canonical_product, keep, reason = _pdf_resolve_product_category(article_src, product_src)
        keep_mask.append(keep)
        canonical_subjects.append(resolved_subject if keep else subject_src)
        canonical_products.append(canonical_product if keep else product_src)
        if not keep:
            rejected_count += 1
            rejects.append({
                "source": source_name,
                "subject_src": subject_src,
                "product_src": product_src,
                "supplier_article": article_src,
                "nm_id": row.get("nm_id", ""),
                "resolved_product": canonical_product,
                "reason": reason,
                "rows": 1,
            })
    filtered = out.loc[pd.Series(keep_mask, index=out.index)].copy()
    if not filtered.empty:
        filtered.loc[:, "subject"] = pd.Series(canonical_subjects, index=out.index).loc[filtered.index].values
        filtered.loc[:, "product"] = pd.Series(canonical_products, index=out.index).loc[filtered.index].values
    if rejected_count:
        log(f"PDF product category filter: {source_name}: kept={len(filtered):,}, rejected={rejected_count:,}")
    return filtered


def _write_pdf_product_group_audit(rejects: List[Dict[str, Any]], audit_path: Path) -> Optional[Path]:
    cols = ["source", "subject_src", "product_src", "supplier_article", "nm_id", "resolved_product", "reason", "rows"]
    if rejects:
        audit = pd.DataFrame(rejects, columns=cols)
        group_cols = [c for c in cols if c != "rows"]
        audit = audit.groupby(group_cols, dropna=False, as_index=False).agg(rows=("rows", "sum"))
        audit = audit.sort_values(["reason", "resolved_product", "supplier_article"], na_position="last")
    else:
        audit = pd.DataFrame(columns=cols)
    audit_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(audit_path, engine="openpyxl") as writer:
        audit.to_excel(writer, sheet_name="Проверить_группы", index=False)
    log(f"Saved product group audit: {audit_path} rows={len(audit):,}")
    return audit_path


def _filter_outputs_by_pdf_product_reference(outputs: Dict[str, pd.DataFrame], audit_dir: Path) -> Dict[str, pd.DataFrame]:
    rejects: List[Dict[str, Any]] = []
    filtered: Dict[str, pd.DataFrame] = dict(outputs)
    for name in ["article_day_fact", "factor_bridge", "entry_points_bridge", "optimal_benchmarks", "factor_summary_for_pdf"]:
        df = filtered.get(name, pd.DataFrame())
        if df is not None and not df.empty:
            filtered[name] = _filter_df_by_pdf_product_reference(df, name, rejects)
    _write_pdf_product_group_audit(rejects, audit_dir / PRODUCT_GROUP_AUDIT_NAME)
    return filtered


def generate_management_pdf(outputs: Dict[str, pd.DataFrame], path: Path) -> Optional[Path]:
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.colors import HexColor
        from reportlab.pdfbase.pdfmetrics import stringWidth
    except Exception as exc:
        log(f"WARN: reportlab недоступен, PDF не создан: {exc}")
        return None
    F_REG, F_BOLD, F_BLACK = _register_topface_fonts()
    W, H = 1600, 900
    RED = HexColor("#c90022")
    RED_DARK = HexColor("#a50019")
    WHITE = colors.white
    SOFT = HexColor("#fff4f5")
    BLACK = HexColor("#111111")
    GRAY = HexColor("#555555")
    GREEN = HexColor("#087a38")
    BAD = HexColor("#b00020")
    outputs = _filter_outputs_by_pdf_product_reference(outputs, path.parent)
    daily = outputs.get("article_day_fact", pd.DataFrame())
    week_start, week_end, prev_start, prev_end = _period_bounds_from_daily(daily)
    # Current incomplete week: from Monday to latest available day in daily.
    latest = pd.to_datetime(daily["day"], errors="coerce").max() if daily is not None and not daily.empty else week_end
    cur_monday = latest - pd.Timedelta(days=int(latest.weekday()))
    cur_week_end = cur_monday + pd.Timedelta(days=6)
    c = canvas.Canvas(str(path), pagesize=(W, H))
    bookmarks: Dict[str, str] = {}
    page_num = 0

    def bg(title: str, subtitle: str = "", section: str = ""):
        nonlocal page_num
        page_num += 1
        c.setFillColor(RED); c.rect(0, 0, W, H, fill=1, stroke=0)
        c.setFillColor(WHITE); c.setFont(F_REG, 34); c.drawString(70, 835, "topface")
        c.setFont(F_BLACK, 46); c.drawString(70, 765, title)
        if subtitle:
            c.setFont(F_BOLD, 20); c.drawString(70, 725, subtitle)
        if section:
            c.setFont(F_BOLD, 14); c.drawRightString(W-75, 710, section)
        c.setFont(F_BOLD, 13); c.drawRightString(W-75, 38, f"Страница {page_num}")

    def button(x, y, w, label, target=None):
        c.setFillColor(WHITE); c.roundRect(x, y, w, 44, 14, fill=1, stroke=0)
        c.setFillColor(RED_DARK); c.setFont(F_BOLD, 13); c.drawCentredString(x+w/2, y+17, label)
        if target:
            c.linkRect("", str(target), (x, y, x+w, y+44), relative=0)

    def top_nav(active=""):
        labels = [("cur", "Текущая"), ("prev", "Прошлая"), ("month", "Месяц"), ("closed", "Закр. месяц"), ("summary", "Сводка")]
        x = 880
        for key, lab in labels:
            button(x, 800, 128, lab, key)
            x += 142

    def card(x, y, w, h, value, label, sub1="", sub2="", metric="good"):
        c.setFillColor(WHITE); c.roundRect(x, y, w, h, 14, fill=1, stroke=0)
        c.setFillColor(BLACK); c.setFont(F_BLACK, 28); c.drawCentredString(x+w/2, y+h-42, str(value))
        c.setFillColor(GRAY); c.setFont(F_REG, 14); c.drawCentredString(x+w/2, y+h-72, label)
        if sub1:
            color = GREEN if "↑" in sub1 else BAD if "↓" in sub1 else GRAY
            c.setFillColor(color); c.setFont(F_BOLD, 11); c.drawCentredString(x+w/2, y+34, sub1)
        if sub2:
            color = GREEN if "↑" in sub2 else BAD if "↓" in sub2 else GRAY
            c.setFillColor(color); c.setFont(F_BOLD, 11); c.drawCentredString(x+w/2, y+18, sub2)

    def table_box(x, y, w, h, headers, rows, col_widths=None, font_size=13, row_h=48, first_col_red=True):
        if col_widths is None:
            col_widths = [w/len(headers)]*len(headers)
        c.setFillColor(WHITE); c.roundRect(x, y, w, h, 16, fill=1, stroke=0)
        c.setFillColor(RED_DARK); c.roundRect(x, y+h-54, w, 54, 12, fill=1, stroke=0)
        xx=x
        c.setFillColor(WHITE); c.setFont(F_BOLD, font_size)
        for i, head in enumerate(headers):
            c.drawCentredString(xx+col_widths[i]/2, y+h-34, str(head))
            xx += col_widths[i]
        yy=y+h-54-row_h
        for ridx, row in enumerate(rows):
            c.setFillColor(WHITE if ridx%2==0 else SOFT); c.rect(x, yy, w, row_h, fill=1, stroke=0)
            xx=x
            for i, val in enumerate(row):
                if i == 0 and first_col_red:
                    c.setFillColor(SOFT); c.roundRect(xx+8, yy+6, col_widths[i]-16, row_h-12, 8, fill=1, stroke=0)
                    c.setFillColor(RED_DARK); c.setFont(F_BOLD, font_size)
                else:
                    c.setFillColor(BLACK); c.setFont(F_BOLD, font_size)
                # Manual multi-line support.
                text = str(val)
                lines = text.split("\n")
                line_y = yy + row_h/2 + (len(lines)-1)*8
                for line in lines:
                    c.drawCentredString(xx+col_widths[i]/2, line_y, line[:44])
                    line_y -= 16
                xx += col_widths[i]
            yy -= row_h

    def delta_text(cur, prev, metric_lower_better=False):
        d = _delta_pct(cur, prev)
        if d is None:
            return ""
        arrow = "↑" if d > 0 else "↓" if d < 0 else "→"
        # visual sign only; color is handled by arrow in card; compact.
        return f"{arrow} {abs(d):.1f}%".replace(".", ",")

    # Page 1 current week by days.
    bookmarks["cur"] = "cur"; c.bookmarkPage("cur")
    bg("Текущая неделя", f"{cur_monday.strftime('%d.%m')}-{cur_week_end.strftime('%d.%m.%Y')} / оперативно: дни и план", "Текущая неделя")
    top_nav("cur")
    cur_period = _agg_daily_for_bridge(daily, cur_monday, latest, ["subject"])
    prev_same = _agg_daily_for_bridge(daily, cur_monday-pd.Timedelta(days=7), latest-pd.Timedelta(days=7), ["subject"])
    cur_total = cur_period.sum(numeric_only=True)
    prev_total = prev_same.sum(numeric_only=True) if not prev_same.empty else pd.Series(dtype=float)
    card(70, 610, 260, 120, _fmt_rub(cur_total.get("order_sum", 0)), "Сумма заказов", delta_text(cur_total.get("order_sum",0), prev_total.get("order_sum",0)))
    card(360, 610, 260, 120, _fmt_rub(cur_total.get("gross_profit_model", 0)), "ВП расч.", delta_text(cur_total.get("gross_profit_model",0), prev_total.get("gross_profit_model",0)))
    card(650, 610, 260, 120, _fmt_pct_pdf(cur_total.get("ad_spend_total",0)/cur_total.get("order_sum",1)*100 if cur_total.get("order_sum",0) else 0), "ДРР", delta_text(cur_total.get("ad_spend_total",0)/cur_total.get("order_sum",1), prev_total.get("ad_spend_total",0)/prev_total.get("order_sum",1) if prev_total.get("order_sum",0) else None))
    card(940, 610, 260, 120, _fmt_rub(cur_total.get("ad_spend_total", 0)), "Расход РК", delta_text(cur_total.get("ad_spend_total",0), prev_total.get("ad_spend_total",0)))
    plan_day = max(0, float(cur_total.get("order_sum",0)) / max(1, (latest-cur_monday).days+1) * 1.1)
    card(1230, 610, 260, 120, _fmt_rub(plan_day), "План/день", "по сумме")
    # Day table.
    headers = ["Категория", "Пн\n"+cur_monday.strftime("%d.%m"), "Вт\n"+(cur_monday+pd.Timedelta(days=1)).strftime("%d.%m"), "Ср\n"+(cur_monday+pd.Timedelta(days=2)).strftime("%d.%m"), "Чт\n"+(cur_monday+pd.Timedelta(days=3)).strftime("%d.%m"), "Пт\n"+(cur_monday+pd.Timedelta(days=4)).strftime("%d.%m"), "Сб\n"+(cur_monday+pd.Timedelta(days=5)).strftime("%d.%m"), "Вс\n"+(cur_monday+pd.Timedelta(days=6)).strftime("%d.%m"), "План/день"]
    rows=[]
    cats = ["Кисти косметические", "Косметические карандаши", "Помады", "Блески"]
    cat_short = {"Кисти косметические":"Кисти", "Косметические карандаши":"Карандаши", "Помады":"Помады", "Блески":"Блески"}
    day_agg = _agg_daily_for_bridge(daily, cur_monday, cur_week_end, ["day", "subject"])
    for cat in cats:
        left = f"{cat_short.get(cat,cat)}\nСумма\nВП\nРасх. РК\nДРР"
        vals=[left]
        for i in range(7):
            day = cur_monday + pd.Timedelta(days=i)
            p = day_agg[(day_agg["day"] == day) & (day_agg["subject"] == cat)] if not day_agg.empty and "day" in day_agg.columns else pd.DataFrame()
            if p.empty or day > latest:
                vals.append("-\n-\n-\n-")
            else:
                rr=p.iloc[0]
                vals.append(f"{_fmt_rub(rr.get('order_sum',0), True)}\n{_fmt_rub(rr.get('gross_profit_model',0), True)}\n{_fmt_rub(rr.get('ad_spend_total',0), True)}\n{_fmt_pct_pdf(rr.get('drr_pct',0))}")
        vals.append(f"{_fmt_rub(plan_day/4, True)}\n-\n-\n-")
        rows.append(vals)
    table_box(70, 80, 1460, 470, headers, rows, col_widths=[190]+[145]*7+[255], font_size=12, row_h=92)
    c.showPage()

    # Page 2 current categories.
    bookmarks["cur_cat"] = "cur_cat"; c.bookmarkPage("cur_cat")
    bg("Текущая неделя: категории", f"{cur_monday.strftime('%d.%m')}-{latest.strftime('%d.%m.%Y')} / переход по категории", "Текущая неделя")
    top_nav("cur")
    cat_cur = _merge_cur_prev(cur_period, prev_same, ["subject"])
    rows=[]
    for _, r in cat_cur.sort_values("order_sum", ascending=False).iterrows():
        rows.append([cat_short.get(r.get("subject"), r.get("subject")), _fmt_rub(r.get("order_sum")), delta_text(r.get("order_sum"), r.get("order_sum_prev")), _fmt_rub(r.get("gross_profit_model")), _fmt_pct_pdf(r.get("drr_pct")), _fmt_rub(r.get("ad_spend_total")), _fmt_cpc_pdf(r.get("cpc"))])
    table_box(80, 300, 1360, 330, ["Категория", "Сумма", "Δ", "ВП расч.", "ДРР", "Расход РК", "CPC"], rows, col_widths=[230,210,120,210,160,220,160], font_size=14, row_h=66)
    c.showPage()

    # Page 3 previous week summary.
    bookmarks["prev"] = "prev"; c.bookmarkPage("prev")
    bg("Прошлая неделя", f"{week_start.strftime('%d.%m')}-{week_end.strftime('%d.%m.%Y')} / сравнение с {prev_start.strftime('%d.%m')}-{prev_end.strftime('%d.%m.%Y')}", "Прошлая неделя")
    top_nav("prev")
    cat = _agg_daily_for_bridge(daily, week_start, week_end, ["subject"])
    cat_prev = _agg_daily_for_bridge(daily, prev_start, prev_end, ["subject"])
    cat_gp = _abc_gp_for_period(builder_global_for_pdf, week_start, week_end, ["subject"]) if "builder_global_for_pdf" in globals() else pd.DataFrame()
    cat_gp_prev = _abc_gp_for_period(builder_global_for_pdf, prev_start, prev_end, ["subject"]).rename(columns={"gp_fact":"gp_fact_prev"}) if "builder_global_for_pdf" in globals() else pd.DataFrame()
    cat_sum = _merge_cur_prev(cat, cat_prev, ["subject"])
    if not cat_gp.empty: cat_sum = cat_sum.merge(cat_gp, on="subject", how="left")
    if not cat_gp_prev.empty: cat_sum = cat_sum.merge(cat_gp_prev[["subject","gp_fact_prev"]], on="subject", how="left")
    total = cat_sum.sum(numeric_only=True)
    card(70, 610, 260, 120, _fmt_rub(total.get("order_sum",0)), "Сумма заказов", delta_text(total.get("order_sum",0), total.get("order_sum_prev",0)))
    card(360, 610, 260, 120, _fmt_rub(total.get("gp_fact", total.get("gross_profit_model",0))), "ВП факт ABC", delta_text(total.get("gp_fact",0), total.get("gp_fact_prev",0)))
    card(650, 610, 260, 120, _fmt_pct_pdf(total.get("ad_spend_total",0)/total.get("order_sum",1)*100 if total.get("order_sum",0) else 0), "ДРР", delta_text(total.get("ad_spend_total",0)/total.get("order_sum",1), total.get("ad_spend_total_prev",0)/total.get("order_sum_prev",1) if total.get("order_sum_prev",0) else None))
    card(940, 610, 260, 120, _fmt_rub(total.get("ad_spend_total",0)), "Расход РК", delta_text(total.get("ad_spend_total",0), total.get("ad_spend_total_prev",0)))
    card(1230, 610, 260, 120, _fmt_cpc_pdf(total.get("ad_spend_total",0)/total.get("ad_clicks_total",1) if total.get("ad_clicks_total",0) else 0), "CPC")
    rows=[]
    for _, r in cat_sum.sort_values("order_sum", ascending=False).iterrows():
        gp = r.get("gp_fact", r.get("gross_profit_model", 0))
        mar = gp / r.get("order_sum",1)*100 if r.get("order_sum",0) else 0
        rows.append([cat_short.get(r.get("subject"), r.get("subject")), _fmt_rub(r.get("order_sum")), delta_text(r.get("order_sum"), r.get("order_sum_prev")), _fmt_rub(gp), delta_text(gp, r.get("gp_fact_prev", r.get("gross_profit_model_prev",0))), _fmt_pct_pdf(mar), _fmt_pct_pdf(r.get("drr_pct")), _fmt_cpc_pdf(r.get("cpc"))])
    table_box(70, 90, 1460, 420, ["Категория", "Сумма", "Δ", "ВП", "Δ", "Рент.", "ДРР", "CPC"], rows, col_widths=[220,200,100,200,100,150,150,150], font_size=13, row_h=66)
    c.showPage()

    # Month placeholders simplified.
    for key, title, sub in [("month", "Текущий месяц", "месяц неполный / темп к плану"), ("closed", "Последний закрытый месяц", "факт по доступным данным"), ("summary", "Сводка по месяцам", "категории / без лишней детализации")]:
        bookmarks[key] = key; c.bookmarkPage(key)
        bg(title, sub, title)
        top_nav(key)
        c.setFillColor(WHITE); c.roundRect(100, 300, 1400, 180, 20, fill=1, stroke=0)
        c.setFillColor(BLACK); c.setFont(F_BOLD, 24); c.drawCentredString(800, 390, "Данные раздела формируются в Excel-расчёте; PDF использует этот блок как навигационный уровень.")
        c.showPage()

    # Category pages for previous week.
    factor_summary = outputs.get("factor_summary_for_pdf", pd.DataFrame())
    factor_bridge = outputs.get("factor_bridge", pd.DataFrame())
    opt = outputs.get("optimal_benchmarks", pd.DataFrame())
    detail_articles = []
    cat_rows = cat_sum.sort_values("order_sum", ascending=False)
    for _, catr in cat_rows.iterrows():
        subject = catr.get("subject")
        cat_name = cat_short.get(subject, subject)
        cat_book = f"cat_{cat_name}"
        bookmarks[cat_book] = cat_book; c.bookmarkPage(cat_book)
        # Article/product rows from bridge/daily.
        a_cur = _agg_daily_for_bridge(daily, week_start, week_end, ["subject", "product", "supplier_article", "nm_id"])
        a_prev = _agg_daily_for_bridge(daily, prev_start, prev_end, ["subject", "product", "supplier_article", "nm_id"])
        a = _merge_cur_prev(a_cur, a_prev, ["subject", "product", "supplier_article", "nm_id"])
        agp = _abc_gp_for_period(builder_global_for_pdf, week_start, week_end, ["subject", "product", "supplier_article", "nm_id"]) if "builder_global_for_pdf" in globals() else pd.DataFrame()
        if not agp.empty: a = a.merge(agp, on=["subject", "product", "supplier_article", "nm_id"], how="left")
        a = a[a["subject"] == subject].copy()
        if a.empty:
            continue
        a["gp_use"] = a["gp_fact"].fillna(a.get("gross_profit_model", 0)) if "gp_fact" in a.columns else a.get("gross_profit_model", 0)
        # ABC-80: only details contributing 80% positive GP; category table still shows visible rows.
        apos = a[a["gp_use"] > 0].sort_values("gp_use", ascending=False).copy()
        total_gp = apos["gp_use"].sum()
        if total_gp > 0:
            apos["cum_share"] = apos["gp_use"].cumsum() / total_gp
            detail_articles += apos[apos["cum_share"] <= 0.80][["subject","product","supplier_article","nm_id"]].to_dict("records")
        pages = [a.sort_values("order_sum", ascending=False).iloc[i:i+7] for i in range(0, len(a), 7)]
        for pi, part in enumerate(pages, start=1):
            bg(f"Категория: {cat_name}", f"{week_start.strftime('%d.%m')}-{week_end.strftime('%d.%m.%Y')} / динамика к прошлой неделе / лист {pi} из {len(pages)}", "Категория")
            button(1280, 800, 180, "← прошлая", "prev")
            rows=[]
            for _, r in part.iterrows():
                gp = r.get("gp_use", 0); mar = gp / r.get("order_sum",1)*100 if r.get("order_sum",0) else 0
                rows.append([r.get("supplier_article"), _fmt_rub(r.get("order_sum")), delta_text(r.get("order_sum"), r.get("order_sum_prev")), _fmt_rub(gp), _fmt_pct_pdf(mar), _fmt_pct_pdf(r.get("drr_pct")), _fmt_cpc_pdf(r.get("cpc")), _fmt_pct_pdf(r.get("search_traffic_capture_pct")), _fmt_pct_pdf(r.get("localization_with_replacements_pct"))])
            table_box(80, 170, 1400, 520, ["Артикул", "Сумма", "Δ", "ВП", "Рент.", "ДРР", "CPC", "% поиска", "Локал."], rows, col_widths=[190,180,90,180,140,130,120,140,140], font_size=13, row_h=62)
            c.showPage()

    # Article detail pages from ABC-80 only.
    seen = set()
    for rec in detail_articles:
        key_tuple = (rec.get("subject"), str(rec.get("product")), str(rec.get("supplier_article")), rec.get("nm_id"))
        if key_tuple in seen:
            continue
        seen.add(key_tuple)
        subject, product, art, nm_id = key_tuple
        a_cur = _agg_daily_for_bridge(daily, week_start, week_end, ["subject", "product", "supplier_article", "nm_id"])
        a_prev = _agg_daily_for_bridge(daily, prev_start, prev_end, ["subject", "product", "supplier_article", "nm_id"])
        a = _merge_cur_prev(a_cur, a_prev, ["subject", "product", "supplier_article", "nm_id"])
        agp = _abc_gp_for_period(builder_global_for_pdf, week_start, week_end, ["subject", "product", "supplier_article", "nm_id"]) if "builder_global_for_pdf" in globals() else pd.DataFrame()
        if not agp.empty: a = a.merge(agp, on=["subject", "product", "supplier_article", "nm_id"], how="left")
        rr = a[(a["subject"] == subject) & (a["supplier_article"].astype(str) == str(art))]
        if rr.empty:
            continue
        r = rr.iloc[0]
        page1 = f"article_{art}_1"; page2=f"article_{art}_2"
        bookmarks[page1]=page1; c.bookmarkPage(page1)
        bg(f"Артикул: {art}", f"{cat_short.get(subject,subject)} / товар {product} / {week_start.strftime('%d.%m')}-{week_end.strftime('%d.%m.%Y')}", "Артикул 1/2")
        button(1240, 800, 170, "← категория", f"cat_{cat_short.get(subject,subject)}")
        button(1430, 800, 100, "стр.2", page2)
        c.setFillColor(RED_DARK); c.roundRect(80, 690, 1440, 44, 10, fill=1, stroke=0); c.setFillColor(WHITE); c.setFont(F_BLACK, 20); c.drawString(105, 705, "Блок 1. Продажи и экономика")
        gp = _safe_float(r.get("gp_fact"), _safe_float(r.get("gross_profit_model")))
        mar = gp / _safe_float(r.get("order_sum"),1)*100 if _safe_float(r.get("order_sum")) else 0
        cards1 = [
            (_fmt_rub(r.get("order_sum")), "Сумма заказов", delta_text(r.get("order_sum"), r.get("order_sum_prev"))),
            (_fmt_rub(gp), "ВП факт ABC", ""),
            (_fmt_pct_pdf(mar), "Рентабельность", ""),
            (_fmt_rub(r.get("avg_order_price")), "Цена продажи", delta_text(r.get("avg_order_price"), r.get("avg_order_price_prev"))),
            (_fmt_pct_pdf(r.get("spp", r.get("spp_funnel",0))), "СПП", delta_text(r.get("spp",0), r.get("spp_prev",0))),
            (_fmt_rub(r.get("commission_model",0)/max(_safe_float(r.get("buyout_qty_model"),1),1)), "Комиссия/шт", ""),
            (_fmt_rub((r.get("logistics_direct_model",0)+r.get("logistics_return_model",0))/max(_safe_float(r.get("buyout_qty_model"),1),1)), "Логистика/шт", ""),
            (_fmt_rub(r.get("storage_model",0)/max(_safe_float(r.get("buyout_qty_model"),1),1)), "Хранение/шт", ""),
            (_fmt_rub(r.get("acquiring_model",0)/max(_safe_float(r.get("buyout_qty_model"),1),1)), "Эквайринг/шт", ""),
            (_fmt_rub(r.get("cost_model",0)/max(_safe_float(r.get("buyout_qty_model"),1),1)), "Себест./шт", ""),
            (_fmt_rub(r.get("other_costs_model",0)/max(_safe_float(r.get("buyout_qty_model"),1),1)), "Прочие/шт", ""),
        ]
        for idx, it in enumerate(cards1[:6]): card(80+idx*240, 570, 220, 96, *it)
        for idx, it in enumerate(cards1[6:]): card(80+idx*240, 455, 220, 96, *it)
        c.setFillColor(RED_DARK); c.roundRect(80, 385, 1440, 44, 10, fill=1, stroke=0); c.setFillColor(WHITE); c.setFont(F_BLACK, 20); c.drawString(105, 400, "Блок 2. Реклама, спрос и конверсии")
        cards2 = [
            (_fmt_rub(r.get("ad_spend_total")), "Расход РК", delta_text(r.get("ad_spend_total"), r.get("ad_spend_total_prev"))),
            (_fmt_pct_pdf(r.get("drr_pct")), "ДРР", delta_text(r.get("drr_pct"), r.get("drr_pct_prev"))),
            (_fmt_cpc_pdf(r.get("cpc")), "CPC", delta_text(r.get("cpc"), r.get("cpc_prev"))),
            (_fmt_num_pdf(r.get("ad_impressions_total")), "Показы РК", delta_text(r.get("ad_impressions_total"), r.get("ad_impressions_total_prev"))),
            (_fmt_num_pdf(r.get("ad_clicks_total")), "Клики РК", delta_text(r.get("ad_clicks_total"), r.get("ad_clicks_total_prev"))),
            (_fmt_num_pdf(r.get("open_cards")), "Открытия", delta_text(r.get("open_cards"), r.get("open_cards_prev"))),
            (_fmt_pct_pdf(r.get("cart_conv_pct")), "Конв. в корзину", delta_text(r.get("cart_conv_pct"), r.get("cart_conv_pct_prev"))),
            (_fmt_pct_pdf(r.get("order_conv_pct")), "Корзина -> заказ", delta_text(r.get("order_conv_pct"), r.get("order_conv_pct_prev"))),
            (_fmt_num_pdf(r.get("search_frequency")), "Спрос WB", delta_text(r.get("search_frequency"), r.get("search_frequency_prev"))),
            (_fmt_pct_pdf(r.get("search_traffic_capture_pct")), "% поиска", delta_text(r.get("search_traffic_capture_pct"), r.get("search_traffic_capture_pct_prev"))),
            (_fmt_pct_pdf(r.get("localization_with_replacements_pct")), "Локализация", ""),
        ]
        for idx, it in enumerate(cards2[:6]): card(80+idx*240, 260, 220, 96, *it)
        for idx, it in enumerate(cards2[6:]): card(80+idx*240, 145, 220, 96, *it)
        c.showPage()
        bookmarks[page2]=page2; c.bookmarkPage(page2)
        bg(f"Артикул: {art}", f"{cat_short.get(subject,subject)} / товар {product} / точки входа и выводы", "Артикул 2/2")
        button(1240, 800, 170, "← категория", f"cat_{cat_short.get(subject,subject)}")
        button(1430, 800, 100, "стр.1", page1)
        eb = outputs.get("entry_points_bridge", pd.DataFrame())
        ep_rows=[]
        if eb is not None and not eb.empty:
            part = eb[(eb["subject"]==subject) & (eb["supplier_article"].astype(str)==str(art))].sort_values("orders", ascending=False).head(7)
            for _, ebr in part.iterrows():
                ep_rows.append([f"{ebr.get('entry_section','')} / {ebr.get('entry_point','')}", _fmt_num_pdf(ebr.get("transitions")), delta_text(ebr.get("transitions"), ebr.get("transitions_prev")), _fmt_num_pdf(ebr.get("orders")), delta_text(ebr.get("orders"), ebr.get("orders_prev")), _fmt_rub(ebr.get("effect_gp_rub"))])
        if not ep_rows:
            ep_rows=[['-', '-', '-', '-', '-', '-']]
        table_box(90, 380, 1420, 340, ["Точка входа", "Переходы", "Δ", "Заказы", "Δ", "Вклад ВП"], ep_rows, col_widths=[570,160,100,140,100,180], font_size=13, row_h=42, first_col_red=False)
        # Factor summary money.
        fs = factor_summary[(factor_summary["level"]=="article") & (factor_summary["supplier_article"].astype(str)==str(art))] if factor_summary is not None and not factor_summary.empty else pd.DataFrame()
        txt = fs.iloc[0]["summary_text"] if not fs.empty else "Факторный мост не выделил значимых денежных причин."
        c.setFillColor(WHITE); c.roundRect(90, 130, 1420, 190, 18, fill=1, stroke=0)
        c.setFillColor(RED_DARK); c.setFont(F_BLACK, 22); c.drawString(120, 280, "Факторный вывод в деньгах")
        c.setFillColor(BLACK); c.setFont(F_BOLD, 18)
        # simple wrap
        words = str(txt).split()
        lines=[]; line=""
        for w0 in words:
            cand = (line + " " + w0).strip()
            if stringWidth(cand, F_BOLD, 18) > 1320:
                lines.append(line); line=w0
            else:
                line=cand
        if line: lines.append(line)
        yy=245
        for line in lines[:6]:
            c.drawString(120, yy, line); yy -= 26
        c.showPage()
    c.save()
    return path



# ============================================================================
# STRICT PDF REPORT OVERRIDES 2026-05-26
# Пересборка управленческого PDF строго по чек-листу:
# - PDF-only без полного пересчета;
# - утвержденный справочник товаров;
# - отбор товаров по 90% стабильной ВП на уровне товара;
# - ДРР сравнивается только с оптимальным ДРР по лучшим дням;
# - факторные выводы строятся по денежному эффекту факторов, а не по сухим %;
# - точки входа выводятся с динамикой и денежным вкладом;
# - страницы артикула показывают два блока: экономика/продажи и спрос/точки входа/конверсии.
# ============================================================================

# PDF-only должен читать новые листы, но старые файлы не ломаются: отсутствующие листы = WARN + пустой df.
try:
    PDF_ONLY_SHEETS[FACTOR_REPORT_NAME] = [
        "optimal_benchmarks", "factor_bridge", "entry_points_bridge",
        "factor_summary_for_pdf", "selected_pdf_products", "product_stability_for_pdf",
    ]
except Exception:
    pass


_FACTOR_ECONOMY_KEYWORDS = (
    "рентабельностин", "дрр", "реклам", "цена", "спп", "комис", "логист", "хран", "эквайр", "себест", "прочие", "расход",
)
_FACTOR_DEMAND_KEYWORDS = (
    "спрос", "поиск", "трафик", "конвер", "корзина", "заказ", "открыт", "ctr", "cpc",
)
_LOWER_BETTER_METRICS = {
    "drr_pct", "cpc", "ad_spend_total", "manual_spend", "unified_spend", "unknown_spend",
    "commission_unit", "logistics_unit", "storage_unit", "acquiring_unit", "cost_unit", "other_costs_unit",
}
_GOOD_GROWTH_METRICS = {
    "order_sum", "gross_profit", "gp_fact", "gross_profit_model", "orders", "open_cards", "add_to_cart",
    "ad_impressions_total", "ad_clicks_total", "search_frequency", "search_traffic_capture_pct",
    "cart_conv_pct", "order_conv_pct", "localization_with_replacements_pct", "margin_pct",
}


def _pdf_text(value: Any) -> str:
    return normalize_text(value)


def _pdf_num(value: Any, default: float = 0.0) -> float:
    try:
        if value is None:
            return default
        v = float(value)
        if pd.isna(v):
            return default
        return v
    except Exception:
        return default


def _pdf_color_delta_value(cur: Any, prev: Any, lower_better: bool = False) -> Tuple[str, str]:
    d = _delta_pct(cur, prev)
    if d is None:
        return "", "neutral"
    arrow = "↑" if d > 0 else "↓" if d < 0 else "→"
    # color semantic, not just direction
    if abs(d) < 1e-9:
        tone = "neutral"
    elif lower_better:
        tone = "bad" if d > 0 else "good"
    else:
        tone = "good" if d > 0 else "bad"
    return f"{arrow} {abs(d):.1f}%".replace(".", ","), tone


def _pdf_signed_rub(x: Any) -> str:
    val = _pdf_num(x, 0.0)
    sign = "+" if val > 0 else "-" if val < 0 else ""
    return f"{sign}{abs(val):,.0f} ₽".replace(",", " ")


def _pdf_short_factor_name(name: Any) -> str:
    n = _pdf_text(name)
    repl = {
        "Объём / сумма заказов": "объём заказов",
        "Рентабельность": "рентабельность",
        "ДРР / рекламная нагрузка": "рекламная нагрузка",
        "Цена продажи": "цена продажи",
        "СПП / цена покупателя": "СПП / цена покупателя",
        "% поискового трафика": "доля поискового трафика",
        "Конверсия в корзину": "конверсия в корзину",
        "Корзина -> заказ": "корзина → заказ",
        "Спрос WB": "спрос WB",
    }
    return repl.get(n, n.lower())


def _factor_block_name(factor: Any, zone: Any = "") -> str:
    f = norm_key(factor)
    z = norm_key(zone)
    if any(k in f or k in z for k in ["спрос", "поиск", "трафик", "конверс", "корзин", "открыт", "ctr", "cpc"]):
        return "Спрос и точки входа / конверсии"
    return "Экономика и продажи"


def _selected_products_by_stable_gp(outputs: Dict[str, pd.DataFrame], threshold: float = 0.90) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Select only product groups that form 90% stable GP by category.

    Selection is done at product level, not article level. Products with non-positive or unstable GP are excluded.
    """
    daily = outputs.get("article_day_fact", pd.DataFrame())
    if daily is None or daily.empty:
        return pd.DataFrame(columns=["subject", "product", "selected_for_pdf"]), pd.DataFrame()
    x = daily.copy()
    for col in ["subject", "product", "supplier_article"]:
        if col not in x.columns:
            x[col] = ""
    x["day"] = pd.to_datetime(x.get("day"), errors="coerce").dt.normalize()
    latest = x["day"].max()
    if pd.notna(latest):
        x = x[x["day"] >= latest - pd.Timedelta(days=89)].copy()
    x["week"] = x["day"].map(lambda v: week_code(v) if pd.notna(v) else "")
    gp_col = "gp_fact" if "gp_fact" in x.columns else "gross_profit_model"
    if gp_col not in x.columns:
        x[gp_col] = 0.0
    x[gp_col] = pd.to_numeric(x[gp_col], errors="coerce").fillna(0)
    if "order_sum" in x.columns:
        x["order_sum"] = pd.to_numeric(x["order_sum"], errors="coerce").fillna(0)
    else:
        x["order_sum"] = 0.0
    if "orders" in x.columns:
        x["orders"] = pd.to_numeric(x["orders"], errors="coerce").fillna(0)
    else:
        x["orders"] = 0.0
    week_gp = x.groupby(["subject", "product", "week"], dropna=False, as_index=False).agg(week_gp=(gp_col, "sum"), week_orders=("orders", "sum"))
    agg = x.groupby(["subject", "product"], dropna=False, as_index=False).agg(
        gp_90=(gp_col, "sum"), order_sum_90=("order_sum", "sum"), orders_90=("orders", "sum"),
        active_days=("day", "nunique"), articles=("supplier_article", "nunique"),
    )
    pos = week_gp.groupby(["subject", "product"], dropna=False, as_index=False).agg(
        active_weeks=("week", "nunique"), positive_weeks=("week_gp", lambda s: int((pd.to_numeric(s, errors="coerce") > 0).sum())),
        negative_weeks=("week_gp", lambda s: int((pd.to_numeric(s, errors="coerce") < 0).sum())),
    )
    agg = agg.merge(pos, on=["subject", "product"], how="left")
    agg["selected_for_pdf"] = False
    agg["selection_reason"] = ""
    selected_rows = []
    for subject, part in agg.groupby("subject", dropna=False):
        p = part.copy()
        p = p[p["subject"].isin(TARGET_SUBJECTS)].copy()
        p = p[p["gp_90"] > 0].copy()
        p = p[(p["positive_weeks"].fillna(0) >= 2) | (p["active_days"].fillna(0) >= 5)].copy()
        if p.empty:
            continue
        p = p.sort_values("gp_90", ascending=False).copy()
        total = p["gp_90"].sum()
        p["gp_share_pct"] = np.where(total > 0, p["gp_90"] / total * 100, 0)
        p["cum_gp_share_pct"] = p["gp_share_pct"].cumsum()
        # Include products until cumulative reaches threshold; always keep at least top product.
        keep = (p["cum_gp_share_pct"] <= threshold * 100) | (p.index == p.index[0])
        # Also include the first product that crosses the threshold, because it is part of the 90% bucket.
        if (~keep).any():
            first_cross_pos = list(p.index).index(p[~keep].index[0]) if len(p[~keep].index) else None
            if first_cross_pos is not None and first_cross_pos > 0:
                keep.loc[p.index[first_cross_pos]] = True
        p.loc[:, "selected_for_pdf"] = keep
        p.loc[:, "selection_reason"] = np.where(keep, "Входит в 90% стабильной ВП категории", "Остаток вне 90% ВП категории")
        selected_rows.append(p)
    result = pd.concat(selected_rows, ignore_index=True) if selected_rows else pd.DataFrame(columns=list(agg.columns) + ["gp_share_pct", "cum_gp_share_pct"])
    audit = agg.merge(result[["subject", "product", "selected_for_pdf", "selection_reason", "gp_share_pct", "cum_gp_share_pct"]], on=["subject", "product"], how="left", suffixes=("", "_sel")) if not result.empty else agg
    if "selected_for_pdf_sel" in audit.columns:
        audit["selected_for_pdf"] = audit["selected_for_pdf_sel"].fillna(False)
        audit["selection_reason"] = audit["selection_reason_sel"].fillna("Не входит в 90% стабильной ВП категории / неприбыльный товар")
        audit = audit.drop(columns=[c for c in ["selected_for_pdf_sel", "selection_reason_sel"] if c in audit.columns])
    return result[result.get("selected_for_pdf", False) == True].copy() if not result.empty else result, audit


def _filter_to_selected_products(outputs: Dict[str, pd.DataFrame], selected: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    if selected is None or selected.empty:
        return outputs
    keyset = set((str(r["subject"]), str(r["product"])) for _, r in selected.iterrows())
    filtered = dict(outputs)
    for name in ["article_day_fact", "factor_bridge", "entry_points_bridge", "optimal_benchmarks", "factor_summary_for_pdf"]:
        df = filtered.get(name, pd.DataFrame())
        if df is None or df.empty or not {"subject", "product"}.issubset(df.columns):
            continue
        # Keep pure category-level rows, and product/article rows only when product is selected.
        if "product" in df.columns:
            mask = []
            for _, r in df.iterrows():
                prod = _pdf_text(r.get("product", ""))
                subj = _pdf_text(r.get("subject", ""))
                if not prod:
                    mask.append(subj in TARGET_SUBJECTS)
                else:
                    mask.append((subj, prod) in keyset)
            filtered[name] = df.loc[pd.Series(mask, index=df.index)].copy()
    return filtered


def compute_optimal_benchmarks(outputs: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Compute optimal benchmark values from best-order days, not hard limits.

    The report must not use a DRR limit. It compares current DRR with the DRR observed on the best days
    when the entity received the highest number of orders / order revenue.
    """
    daily = outputs.get("article_day_fact", pd.DataFrame())
    if daily is None or daily.empty:
        return pd.DataFrame()
    x = daily.copy()
    x["day"] = pd.to_datetime(x.get("day"), errors="coerce").dt.normalize()
    levels = [
        ("category", ["subject"]),
        ("product", ["subject", "product"]),
        ("article", ["subject", "product", "supplier_article", "nm_id"]),
    ]
    rows: List[Dict[str, Any]] = []
    for level, keys in levels:
        for k in keys:
            if k not in x.columns:
                x[k] = ""
        for key_vals, part in x.groupby(keys, dropna=False):
            if not isinstance(key_vals, tuple):
                key_vals = (key_vals,)
            d = part.copy()
            if d.empty:
                continue
            day = _agg_daily_for_bridge(d, d["day"].min(), d["day"].max(), ["day"])
            if day.empty:
                continue
            day["score"] = pd.to_numeric(day.get("orders", 0), errors="coerce").fillna(0) * 0.65 + (pd.to_numeric(day.get("order_sum", 0), errors="coerce").fillna(0) / 1000.0) * 0.35
            day = day[day["score"] > 0].copy()
            if day.empty:
                continue
            if len(day) >= 10:
                threshold = day["score"].quantile(0.80)
                best = day[day["score"] >= threshold].copy()
            else:
                best = day.nlargest(min(3, len(day)), "score").copy()
            if best.empty:
                continue
            ad_spend = pd.to_numeric(best.get("ad_spend_total", 0), errors="coerce").fillna(0)
            order_sum = pd.to_numeric(best.get("order_sum", 0), errors="coerce").fillna(0)
            clicks = pd.to_numeric(best.get("ad_clicks_total", 0), errors="coerce").fillna(0)
            imps = pd.to_numeric(best.get("ad_impressions_total", 0), errors="coerce").fillna(0)
            orders = pd.to_numeric(best.get("orders", 0), errors="coerce").fillna(0)
            open_cards = pd.to_numeric(best.get("open_cards", 0), errors="coerce").fillna(0)
            add_to_cart = pd.to_numeric(best.get("add_to_cart", 0), errors="coerce").fillna(0)
            search_freq = pd.to_numeric(best.get("search_frequency", 0), errors="coerce").fillna(0)
            search_trans = pd.to_numeric(best.get("search_transitions", 0), errors="coerce").fillna(0)
            rec = {"level": level}
            rec.update(dict(zip(keys, key_vals)))
            for k in ["subject", "product", "supplier_article", "nm_id"]:
                rec.setdefault(k, "")
            rec.update({
                "best_days_count": int(len(best)),
                "optimal_order_sum_day": float(order_sum.mean()) if len(order_sum) else np.nan,
                "optimal_orders_day": float(orders.mean()) if len(orders) else np.nan,
                "optimal_drr_pct": float(ad_spend.sum() / order_sum.sum() * 100) if order_sum.sum() else np.nan,
                "optimal_cpc": float(ad_spend.sum() / clicks.sum()) if clicks.sum() else np.nan,
                "optimal_ctr_pct": float(clicks.sum() / imps.sum() * 100) if imps.sum() else np.nan,
                "optimal_cart_conv_pct": float(add_to_cart.sum() / open_cards.sum() * 100) if open_cards.sum() else np.nan,
                "optimal_order_conv_pct": float(orders.sum() / add_to_cart.sum() * 100) if add_to_cart.sum() else np.nan,
                "optimal_search_capture_pct": float(search_trans.sum() / search_freq.sum() * 100) if search_freq.sum() else np.nan,
                "optimal_price_sale": float(order_sum.sum() / orders.sum()) if orders.sum() else np.nan,
                "optimal_spp": float(pd.to_numeric(best.get("spp", best.get("spp_funnel", np.nan)), errors="coerce").mean()),
            })
            rows.append(rec)
    return pd.DataFrame(rows)


def _append_optimal_factor_rows(level: str, g: pd.DataFrame, keys: List[str], optimal: pd.DataFrame) -> List[Dict[str, Any]]:
    if g is None or g.empty or optimal is None or optimal.empty:
        return []
    rows: List[Dict[str, Any]] = []
    opt = optimal[optimal.get("level", "") == level].copy() if "level" in optimal.columns else optimal.copy()
    if opt.empty:
        return []
    join_cols = [k for k in keys if k in opt.columns and k in g.columns]
    if not join_cols:
        return []
    m = g.merge(opt, on=join_cols, how="left", suffixes=("", "_opt"))
    for _, r in m.iterrows():
        keyvals = {k: r.get(k, "") for k in keys}
        cur_sum = _pdf_num(r.get("order_sum"), 0.0)
        cur_orders = _pdf_num(r.get("orders"), 0.0)
        cur_gp = _pdf_num(r.get("gp_fact"), _pdf_num(r.get("gross_profit_model"), 0.0))
        margin = cur_gp / cur_sum if cur_sum else 0.0
        avg_price = cur_sum / cur_orders if cur_orders else _pdf_num(r.get("avg_order_price"), 0.0)
        # Against optimal DRR from best days.
        cur_drr = _pdf_num(r.get("drr_pct"), np.nan)
        opt_drr = _pdf_num(r.get("optimal_drr_pct"), np.nan)
        if pd.notna(cur_drr) and pd.notna(opt_drr) and opt_drr > 0 and cur_sum:
            effect = -cur_sum * ((cur_drr - opt_drr) / 100.0)
            rows.append(_factor_row(level, keyvals, "ДРР против оптимума", opt_drr, cur_drr, cur_drr - opt_drr, effect, "оптимум", "Сколько ВП отличается от лучших дней из-за текущего ДРР."))
        # Demand/traffic/conversions vs best-day benchmark.
        opens = _pdf_num(r.get("open_cards"), 0.0)
        cart_cur = _pdf_num(r.get("cart_conv_pct"), np.nan) / 100.0
        cart_opt = _pdf_num(r.get("optimal_cart_conv_pct"), np.nan) / 100.0
        order_cur = _pdf_num(r.get("order_conv_pct"), np.nan) / 100.0
        order_opt = _pdf_num(r.get("optimal_order_conv_pct"), np.nan) / 100.0
        if opens and pd.notna(cart_cur) and pd.notna(cart_opt) and pd.notna(order_cur):
            effect = opens * (cart_cur - cart_opt) * order_cur * avg_price * margin
            rows.append(_factor_row(level, keyvals, "Конверсия в корзину против оптимума", cart_opt * 100, cart_cur * 100, (cart_cur - cart_opt) * 100, effect, "оптимум", "Потеря/прирост ВП относительно конверсии в корзину на лучших днях."))
        if opens and pd.notna(order_cur) and pd.notna(order_opt) and pd.notna(cart_cur):
            effect = opens * cart_cur * (order_cur - order_opt) * avg_price * margin
            rows.append(_factor_row(level, keyvals, "Корзина -> заказ против оптимума", order_opt * 100, order_cur * 100, (order_cur - order_opt) * 100, effect, "оптимум", "Потеря/прирост ВП относительно дожима корзина→заказ на лучших днях."))
        capture_cur = _pdf_num(r.get("search_traffic_capture_pct"), np.nan) / 100.0
        capture_opt = _pdf_num(r.get("optimal_search_capture_pct"), np.nan) / 100.0
        demand = _pdf_num(r.get("search_frequency"), 0.0)
        if demand and pd.notna(capture_cur) and pd.notna(capture_opt) and pd.notna(cart_cur) and pd.notna(order_cur):
            effect = demand * (capture_cur - capture_opt) * cart_cur * order_cur * avg_price * margin
            rows.append(_factor_row(level, keyvals, "% поиска против оптимума", capture_opt * 100, capture_cur * 100, (capture_cur - capture_opt) * 100, effect, "оптимум", "Потеря/прирост ВП относительно доли поискового трафика на лучших днях."))
    for row in rows:
        row["comparison"] = "optimal_best_days"
    return rows


def _build_factor_summary_df(factor_bridge: pd.DataFrame) -> pd.DataFrame:
    if factor_bridge is None or factor_bridge.empty:
        return pd.DataFrame(columns=["level", "subject", "product", "supplier_article", "nm_id", "summary_text"])
    fb = factor_bridge.copy()
    for col in ["level", "subject", "product", "supplier_article", "nm_id", "factor", "zone"]:
        if col not in fb.columns:
            fb[col] = ""
    fb["effect_gp_rub"] = pd.to_numeric(fb.get("effect_gp_rub"), errors="coerce").fillna(0.0)
    fb["abs_effect_gp_rub"] = fb["effect_gp_rub"].abs()
    if "comparison" not in fb.columns:
        fb["comparison"] = "prev_week"
    rows: List[Dict[str, Any]] = []
    for keys, part in fb.groupby(["level", "subject", "product", "supplier_article", "nm_id"], dropna=False):
        lvl, subject, product, art, nm = keys
        p = part[part["abs_effect_gp_rub"] > 100].sort_values("abs_effect_gp_rub", ascending=False).head(6)
        if p.empty:
            main = "Существенных денежных факторов не выделено: изменение ВП в рабочем диапазоне."
            econ = "Экономика/продажи: значимых отклонений в деньгах нет."
            demand = "Спрос/точки входа/конверсии: значимых отклонений в деньгах нет."
        else:
            labels = ["Основная причина", "Вторичная причина", "Дополнительный фактор"]
            lines = []
            for i, (_, r) in enumerate(p.head(3).iterrows()):
                val = float(r["effect_gp_rub"])
                verb = "добавила" if val > 0 else "забрала"
                compare = "к оптимуму" if str(r.get("comparison", "")) == "optimal_best_days" else "к прошлой неделе"
                change = r.get("change", "")
                try:
                    ch = float(change)
                    ch_txt = f", изменение {ch:+.1f} п.п." if "конверс" in norm_key(r.get("factor")) or "дрр" in norm_key(r.get("factor")) or "поиск" in norm_key(r.get("factor")) else ""
                    ch_txt = ch_txt.replace(".", ",")
                except Exception:
                    ch_txt = ""
                lines.append(f"{i+1}. {labels[i]}: {_pdf_short_factor_name(r.get('factor'))} {verb} {abs(val):,.0f} ₽ ВП {compare}{ch_txt}.".replace(",", " "))
            main = " ".join(lines)
            econ_part = p[p.apply(lambda r: _factor_block_name(r.get("factor"), r.get("zone")) == "Экономика и продажи", axis=1)].head(3)
            dem_part = p[p.apply(lambda r: _factor_block_name(r.get("factor"), r.get("zone")) != "Экономика и продажи", axis=1)].head(3)
            def block_text(title: str, df: pd.DataFrame) -> str:
                if df.empty:
                    return f"{title}: крупных денежных факторов нет."
                pieces = []
                for _, rr in df.iterrows():
                    val = float(rr["effect_gp_rub"])
                    pieces.append(f"{_pdf_short_factor_name(rr.get('factor'))}: {'+' if val > 0 else '-'}{abs(val):,.0f} ₽".replace(",", " "))
                return f"{title}: " + "; ".join(pieces) + "."
            econ = block_text("Экономика и продажи", econ_part)
            demand = block_text("Спрос и точки входа / конверсии", dem_part)
        rows.append({
            "level": lvl, "subject": subject, "product": product, "supplier_article": art, "nm_id": nm,
            "summary_text": main,
            "economy_sales_text": econ,
            "demand_entry_conversion_text": demand,
            "top_negative_effect_rub": float(p[p["effect_gp_rub"] < 0]["effect_gp_rub"].min()) if not p[p["effect_gp_rub"] < 0].empty else 0.0,
            "top_positive_effect_rub": float(p[p["effect_gp_rub"] > 0]["effect_gp_rub"].max()) if not p[p["effect_gp_rub"] > 0].empty else 0.0,
        })
    return pd.DataFrame(rows)


def build_factor_outputs(builder: AnalyticsBuilder, outputs: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    global outputs_global_for_bridge
    outputs_global_for_bridge = outputs
    daily = outputs.get("article_day_fact", pd.DataFrame())
    week_start, week_end, prev_start, prev_end = _period_bounds_from_daily(daily)
    optimal = compute_optimal_benchmarks(outputs)
    selected, product_stability = _selected_products_by_stable_gp(outputs, threshold=0.90)
    factor_rows: List[Dict[str, Any]] = []
    for level, keys in [
        ("category", ["subject"]),
        ("product", ["subject", "product"]),
        ("article", ["subject", "product", "supplier_article", "nm_id"]),
    ]:
        cur = _agg_daily_for_bridge(daily, week_start, week_end, keys)
        prev = _agg_daily_for_bridge(daily, prev_start, prev_end, keys)
        cur_gp = _abc_gp_for_period(builder, week_start, week_end, keys)
        prev_gp = _abc_gp_for_period(builder, prev_start, prev_end, keys).rename(columns={"gp_fact": "gp_fact_prev", "gross_revenue_fact": "gross_revenue_fact_prev", "sales_qty_fact": "sales_qty_fact_prev"})
        g = _merge_cur_prev(cur, prev, keys)
        if not cur_gp.empty:
            g = g.merge(cur_gp, on=keys, how="left")
        if not prev_gp.empty:
            g = g.merge(prev_gp, on=keys, how="left")
        for c0 in ["gp_fact", "gp_fact_prev", "gross_revenue_fact", "gross_revenue_fact_prev", "sales_qty_fact", "sales_qty_fact_prev"]:
            if c0 not in g.columns:
                g[c0] = np.nan
        base_rows = _entity_factor_rows(level, g, keys)
        for row in base_rows:
            row["comparison"] = "prev_week"
            row["factor_block"] = _factor_block_name(row.get("factor"), row.get("zone"))
        factor_rows.extend(base_rows)
        opt_rows = _append_optimal_factor_rows(level, g, keys, optimal)
        for row in opt_rows:
            row["factor_block"] = _factor_block_name(row.get("factor"), row.get("zone"))
        factor_rows.extend(opt_rows)
    factor_bridge = pd.DataFrame(factor_rows)
    if not factor_bridge.empty:
        for c0 in ["subject", "product", "supplier_article", "nm_id"]:
            if c0 not in factor_bridge.columns:
                factor_bridge[c0] = ""
        factor_bridge["abs_effect_gp_rub"] = pd.to_numeric(factor_bridge["effect_gp_rub"], errors="coerce").fillna(0).abs()
        totals = factor_bridge.groupby(["level", "subject", "product", "supplier_article", "nm_id"], dropna=False)["abs_effect_gp_rub"].sum().rename("total_abs_effect").reset_index()
        factor_bridge = factor_bridge.merge(totals, on=["level", "subject", "product", "supplier_article", "nm_id"], how="left")
        factor_bridge["effect_weight_pct"] = np.where(factor_bridge["total_abs_effect"] > 0, factor_bridge["abs_effect_gp_rub"] / factor_bridge["total_abs_effect"] * 100, 0)
        factor_bridge = factor_bridge.sort_values(["level", "subject", "product", "supplier_article", "abs_effect_gp_rub"], ascending=[True, True, True, True, False])
    entry_bridge = compute_entry_points_bridge(builder, week_start, week_end, prev_start, prev_end)
    factor_summary = _build_factor_summary_df(factor_bridge)
    return {
        "optimal_benchmarks": optimal,
        "factor_bridge": factor_bridge,
        "entry_points_bridge": entry_bridge,
        "factor_summary_for_pdf": factor_summary,
        "selected_pdf_products": selected,
        "product_stability_for_pdf": product_stability,
    }


def _wrap_pdf_lines(text: Any, font_name: str, font_size: int, max_width: float, max_lines: int = 99) -> List[str]:
    try:
        from reportlab.pdfbase.pdfmetrics import stringWidth
    except Exception:
        return str(text).split("\n")[:max_lines]
    words = str(text).replace("\n", " ").split()
    lines: List[str] = []
    line = ""
    for w0 in words:
        cand = (line + " " + w0).strip()
        if stringWidth(cand, font_name, font_size) > max_width and line:
            lines.append(line)
            line = w0
            if len(lines) >= max_lines:
                break
        else:
            line = cand
    if line and len(lines) < max_lines:
        lines.append(line)
    return lines[:max_lines]


def _find_optimal_row(opt: pd.DataFrame, level: str, subject: Any = "", product: Any = "", article: Any = "") -> pd.Series:
    if opt is None or opt.empty:
        return pd.Series(dtype=object)
    x = opt.copy()
    if "level" in x.columns:
        x = x[x["level"].astype(str).eq(level)]
    if subject and "subject" in x.columns:
        x = x[x["subject"].astype(str).eq(str(subject))]
    if product and "product" in x.columns:
        x = x[x["product"].astype(str).eq(str(product))]
    if article and "supplier_article" in x.columns:
        x = x[x["supplier_article"].astype(str).eq(str(article))]
    return x.iloc[0] if not x.empty else pd.Series(dtype=object)



# ------------------------- FINAL PDF/GROSS PROFIT FIXES 2026-05-26 -------------------------
def _load_abc_for_pdf_only(storage: Storage, reports_root: str, store: str, diagnostics: Optional[Diagnostics] = None, latest_year: Optional[int] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Load only ABC reports for PDF-only mode.

    This is much faster than the full report rebuild and prevents PDF-only from falling back to
    model GP while the captions say "ВП факт ABC".
    """
    diag = diagnostics or Diagnostics()
    try:
        loader = Loader(storage, reports_root, store, diag)
        year = int(latest_year or datetime.today().year)
        weekly, monthly = loader.load_abc(year)
        return weekly, monthly
    except Exception as exc:
        log(f"WARN pdf_only: ABC fact was not loaded, GP fact pages will use calculated model where needed: {exc}")
        return pd.DataFrame(), pd.DataFrame()


def load_existing_outputs_for_pdf(storage: Storage, local_dir: Path, reports_root: str = "Отчёты", store: str = "TOPFACE", diagnostics: Optional[Diagnostics] = None) -> Dict[str, pd.DataFrame]:
    """Load generated Excel files for PDF-only plus ABC fact files for correct GP.

    Old PDF-only mode used only article_day_fact and factor sheets; therefore gross profit in
    the PDF was often a model value while the page title said ABC fact. This function also loads
    weekly/monthly ABC from S3/local storage and stores it in outputs.
    """
    outputs: Dict[str, pd.DataFrame] = {}
    for file_name, sheets in PDF_ONLY_SHEETS.items():
        data = _read_existing_report_bytes(storage, local_dir, file_name)
        xls = pd.ExcelFile(io.BytesIO(data))
        for sheet_name in sheets:
            if sheet_name not in xls.sheet_names:
                log(f"WARN pdf_only: sheet {sheet_name} missing in {file_name}")
                outputs[sheet_name] = pd.DataFrame()
                continue
            df = pd.read_excel(xls, sheet_name=sheet_name)
            outputs[sheet_name] = _normalize_pdf_only_df(df)
            log(f"pdf_only: loaded {file_name}/{sheet_name}: rows={len(outputs[sheet_name]):,}, cols={len(outputs[sheet_name].columns):,}")
    daily = outputs.get("article_day_fact", pd.DataFrame())
    if daily.empty:
        raise RuntimeError("PDF-only режим невозможен: пустой article_day_fact в техническом файле")
    required = ["day", "subject", "product", "supplier_article", "nm_id", "order_sum", "gross_profit_model"]
    missing = [c for c in required if c not in daily.columns]
    if missing:
        raise RuntimeError(f"PDF-only режим невозможен: в article_day_fact нет колонок {missing}")
    latest = pd.to_datetime(daily.get("day"), errors="coerce").max()
    latest_year = int(pd.Timestamp(latest).year) if pd.notna(latest) else datetime.today().year
    abc_weekly, abc_monthly = _load_abc_for_pdf_only(storage, reports_root, store, diagnostics, latest_year)
    outputs["abc_weekly"] = abc_weekly
    outputs["abc_monthly"] = abc_monthly
    return outputs


def _gp_from_abc_frames(outputs: Dict[str, pd.DataFrame], start: pd.Timestamp, end: pd.Timestamp, group_cols: List[str]) -> pd.DataFrame:
    """Return exact ABC GP for weekly/monthly periods from outputs, never overlap-prorated."""
    start = pd.Timestamp(start).normalize()
    end = pd.Timestamp(end).normalize()
    if "day" in group_cols:
        return pd.DataFrame(columns=group_cols + ["gp_fact", "gross_revenue_fact", "sales_qty_fact", "gp_source"])
    frames = []
    for src_name, tag in [("abc_weekly", "ABC_weekly_exact"), ("abc_monthly", "ABC_monthly_exact")]:
        src = outputs.get(src_name, pd.DataFrame())
        if src is None or src.empty:
            continue
        x = src.copy()
        if "period_start" not in x.columns or "period_end" not in x.columns:
            continue
        x["period_start"] = pd.to_datetime(x["period_start"], errors="coerce").dt.normalize()
        x["period_end"] = pd.to_datetime(x["period_end"], errors="coerce").dt.normalize()
        x = x[(x["period_start"] == start) & (x["period_end"] == end)].copy()
        if x.empty:
            continue
        for c in group_cols:
            if c not in x.columns:
                x[c] = ""
        x["gp_source"] = tag
        frames.append(x)
    if not frames and "builder_global_for_pdf" in globals():
        try:
            return _abc_gp_for_period(builder_global_for_pdf, start, end, group_cols)
        except Exception:
            pass
    if not frames:
        return pd.DataFrame(columns=group_cols + ["gp_fact", "gross_revenue_fact", "sales_qty_fact", "gp_source"])
    exact = pd.concat(frames, ignore_index=True)
    for c0 in ["gross_profit", "gross_revenue", "orders", "abc_drr_pct", "abc_commission_amount", "abc_acquiring_amount"]:
        if c0 not in exact.columns:
            exact[c0] = np.nan if c0 == "abc_drr_pct" else 0.0
        exact[c0] = pd.to_numeric(exact[c0], errors="coerce")
    exact["_abc_ad_spend"] = np.where(exact["gross_revenue"].fillna(0) > 0, exact["gross_revenue"].fillna(0) * exact["abc_drr_pct"].fillna(0) / 100.0, 0.0)
    exact["_abc_commission_abs"] = exact["abc_commission_amount"].abs().fillna(0)
    exact["_abc_acquiring_abs"] = exact["abc_acquiring_amount"].abs().fillna(0)
    def _join_unique(s):
        vals = [normalize_text(v) for v in s.dropna().astype(str).tolist() if normalize_text(v)]
        vals = list(dict.fromkeys(vals))
        return " | ".join(vals[:20])
    g = exact.groupby(group_cols, dropna=False, as_index=False).agg(
        gp_fact=("gross_profit", "sum"),
        gross_revenue_fact=("gross_revenue", "sum"),
        sales_qty_fact=("orders", "sum"),
        abc_ad_spend_fact=("_abc_ad_spend", "sum"),
        abc_commission_amount_fact=("_abc_commission_abs", "sum"),
        abc_acquiring_amount_fact=("_abc_acquiring_abs", "sum"),
        gp_source=("gp_source", "first"),
        gp_source_file=("source_file", _join_unique),
        gp_source_sheet=("source_sheet", _join_unique),
        gp_source_rows=("source_row_excel", _join_unique),
        gp_source_cells=("cell_gross_profit", _join_unique),
    )
    rev = pd.to_numeric(g["gross_revenue_fact"], errors="coerce").fillna(0)
    g["abc_margin_pct"] = np.where(rev > 0, pd.to_numeric(g["gp_fact"], errors="coerce").fillna(0) / rev * 100, np.nan)
    g["abc_drr_pct"] = np.where(rev > 0, pd.to_numeric(g["abc_ad_spend_fact"], errors="coerce").fillna(0) / rev * 100, np.nan)
    g["abc_commission_pct"] = np.where(rev > 0, pd.to_numeric(g["abc_commission_amount_fact"], errors="coerce").fillna(0) / rev * 100, np.nan)
    g["abc_acquiring_pct"] = np.where(rev > 0, pd.to_numeric(g["abc_acquiring_amount_fact"], errors="coerce").fillna(0) / rev * 100, np.nan)
    return g


def _selected_products_by_stable_gp(outputs: Dict[str, pd.DataFrame], threshold: float = 0.90) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Select product groups for detailed PDF pages by stable GP at PRODUCT level.

    Final rule:
    - select products first, never articles directly;
    - selection is global across approved products, not category-by-category, so weak tails
      do not receive pages just because their category is small;
    - exclude products with no current sales/profit and explicit low-value tails;
    - keep products inside 90% of stable GP, subject to materiality guards.
    """
    daily = outputs.get("article_day_fact", pd.DataFrame())
    if daily is None or daily.empty:
        return pd.DataFrame(columns=["subject", "product", "selected_for_pdf"]), pd.DataFrame()
    x = daily.copy()
    for col in ["subject", "product", "supplier_article"]:
        if col not in x.columns:
            x[col] = ""
    x["day"] = pd.to_datetime(x.get("day"), errors="coerce").dt.normalize()
    latest = x["day"].max()
    if pd.notna(latest):
        x = x[x["day"] >= latest - pd.Timedelta(days=89)].copy()
    gp_col = "gross_profit_model"
    if gp_col not in x.columns:
        x[gp_col] = 0.0
    for c0 in [gp_col, "order_sum", "orders"]:
        if c0 not in x.columns:
            x[c0] = 0.0
        x[c0] = pd.to_numeric(x[c0], errors="coerce").fillna(0)
    x = x[x["subject"].astype(str).isin(TARGET_SUBJECTS)].copy()
    allowed_pairs = {(subj, prod) for prod, subj in PDF_PRODUCT_CATEGORY_REFERENCE.items()}
    x = x[x.apply(lambda r: (str(r.get("subject")), str(r.get("product"))) in allowed_pairs, axis=1)].copy()
    if x.empty:
        return pd.DataFrame(columns=["subject", "product", "selected_for_pdf"]), pd.DataFrame()
    x["week"] = x["day"].map(lambda v: week_code(v) if pd.notna(v) else "")
    cur_monday = latest - pd.Timedelta(days=int(latest.weekday())) if pd.notna(latest) else None
    week_gp = x.groupby(["subject", "product", "week"], dropna=False, as_index=False).agg(
        week_gp=(gp_col, "sum"), week_order_sum=("order_sum", "sum"), week_orders=("orders", "sum")
    )
    agg = x.groupby(["subject", "product"], dropna=False, as_index=False).agg(
        gp_90=(gp_col, "sum"), order_sum_90=("order_sum", "sum"), orders_90=("orders", "sum"),
        active_days=("day", "nunique"), articles=("supplier_article", "nunique"),
    )
    pos = week_gp.groupby(["subject", "product"], dropna=False, as_index=False).agg(
        active_weeks=("week", "nunique"),
        positive_weeks=("week_gp", lambda s: int((pd.to_numeric(s, errors="coerce") > 0).sum())),
        negative_weeks=("week_gp", lambda s: int((pd.to_numeric(s, errors="coerce") < 0).sum())),
    )
    agg = agg.merge(pos, on=["subject", "product"], how="left")
    if cur_monday is not None:
        cur_week = x[x["day"] >= cur_monday].groupby(["subject", "product"], dropna=False, as_index=False).agg(
            current_week_gp=(gp_col, "sum"), current_week_order_sum=("order_sum", "sum"), current_week_orders=("orders", "sum")
        )
        agg = agg.merge(cur_week, on=["subject", "product"], how="left")
    else:
        agg["current_week_gp"] = np.nan
        agg["current_week_order_sum"] = np.nan
        agg["current_week_orders"] = np.nan
    agg["selected_for_pdf"] = False
    agg["selection_reason"] = "Не входит в глобальные 90% стабильной ВП / хвостовой товар"
    min_share_pct = float(os.getenv("PDF_MIN_PRODUCT_GP_SHARE_PCT", "1.0") or 1.0)
    min_current_gp = float(os.getenv("PDF_MIN_PRODUCT_CURRENT_WEEK_GP", "1000") or 1000)
    min_current_order_sum = float(os.getenv("PDF_MIN_PRODUCT_CURRENT_WEEK_ORDER_SUM", "10000") or 10000)
    force_excl = set(PDF_FORCE_EXCLUDE_DETAIL_PRODUCTS)
    eligible = agg.copy()
    eligible = eligible[~eligible["product"].astype(str).isin(force_excl)].copy()
    eligible = eligible[pd.to_numeric(eligible["gp_90"], errors="coerce").fillna(0) > 0].copy()
    eligible = eligible[pd.to_numeric(eligible["current_week_gp"], errors="coerce").fillna(0) >= min_current_gp].copy()
    eligible = eligible[pd.to_numeric(eligible["current_week_order_sum"], errors="coerce").fillna(0) >= min_current_order_sum].copy()
    eligible = eligible[pd.to_numeric(eligible["current_week_orders"], errors="coerce").fillna(0) > 0].copy()
    eligible = eligible[(eligible["positive_weeks"].fillna(0) >= 3) & (eligible["negative_weeks"].fillna(0) <= 1)].copy()
    if eligible.empty:
        audit = agg.copy()
        audit.loc[audit["product"].astype(str).isin(force_excl), "selection_reason"] = "Исключен как хвостовой товар по правилу PDF_FORCE_EXCLUDE_PRODUCTS"
        return pd.DataFrame(columns=list(agg.columns) + ["gp_share_pct", "cum_gp_share_pct"]), audit
    eligible = eligible.sort_values("gp_90", ascending=False).copy()
    total = pd.to_numeric(eligible["gp_90"], errors="coerce").fillna(0).sum()
    eligible["gp_share_pct"] = np.where(total > 0, eligible["gp_90"] / total * 100, 0)
    eligible["cum_gp_share_pct"] = eligible["gp_share_pct"].cumsum()
    eligible["cum_before_pct"] = eligible["cum_gp_share_pct"] - eligible["gp_share_pct"]
    keep = (eligible["cum_before_pct"] < threshold * 100) & (eligible["gp_share_pct"] >= min_share_pct)
    if not keep.any() and not eligible.empty:
        keep.iloc[0] = True
    selected = eligible[keep].copy()
    # Force include user-confirmed product groups when they have current sales and positive current GP.
    force_incl = set(PDF_FORCE_INCLUDE_DETAIL_PRODUCTS) - set(PDF_FORCE_EXCLUDE_DETAIL_PRODUCTS)
    forced = agg[agg["product"].astype(str).isin(force_incl)].copy()
    forced = forced[(pd.to_numeric(forced.get("current_week_order_sum"), errors="coerce").fillna(0) > 0) & (pd.to_numeric(forced.get("current_week_gp"), errors="coerce").fillna(0) > 0)].copy()
    if not forced.empty:
        if "gp_share_pct" not in forced.columns:
            forced["gp_share_pct"] = np.where(total > 0, pd.to_numeric(forced["gp_90"], errors="coerce").fillna(0) / total * 100, 0)
        if "cum_gp_share_pct" not in forced.columns:
            forced["cum_gp_share_pct"] = np.nan
        selected = pd.concat([selected, forced], ignore_index=True).drop_duplicates(["subject", "product"], keep="first")
    selected["selected_for_pdf"] = True
    selected["selection_reason"] = np.where(selected["product"].astype(str).isin(force_incl), "Включен принудительно как подтвержденная товарная группа PDF", "Входит в глобальные 90% стабильной ВП товаров")
    audit = agg.merge(selected[["subject", "product", "selected_for_pdf", "selection_reason", "gp_share_pct", "cum_gp_share_pct"]], on=["subject", "product"], how="left", suffixes=("", "_sel"))
    if "selected_for_pdf_sel" in audit.columns:
        audit["selected_for_pdf"] = audit["selected_for_pdf_sel"].fillna(False)
        audit["selection_reason"] = audit["selection_reason_sel"].fillna("Не входит в глобальные 90% стабильной ВП / хвостовой товар")
        audit = audit.drop(columns=[c for c in ["selected_for_pdf_sel", "selection_reason_sel"] if c in audit.columns])
    audit.loc[audit["product"].astype(str).isin(force_excl), "selection_reason"] = "Исключен как хвостовой товар по правилу PDF_FORCE_EXCLUDE_PRODUCTS"
    return selected, audit


def generate_management_pdf(outputs: Dict[str, pd.DataFrame], path: Path) -> Optional[Path]:
    """Generate the strict management PDF report.

    This implementation intentionally avoids Excel-like tables and uses one clear red/white layout.
    It relies on monetary factor effects already calculated in factor_bridge. If old factor files are used
    in --pdf-only mode, it rebuilds the PDF summary from factor_bridge on the fly.
    """
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.colors import HexColor
        from reportlab.pdfbase.pdfmetrics import stringWidth
    except Exception as exc:
        log(f"WARN: reportlab недоступен, PDF не создан: {exc}")
        return None
    F_REG, F_BOLD, F_BLACK = _register_topface_fonts()
    W, H = 1600, 900
    RED = HexColor("#c90022")
    RED_DARK = HexColor("#9d0018")
    WHITE = colors.white
    SOFT = HexColor("#fff4f5")
    BLACK = HexColor("#111111")
    GRAY = HexColor("#555555")
    GREEN = HexColor("#087a38")
    BAD = HexColor("#b00020")
    AMBER = HexColor("#b85c00")

    outputs = _filter_outputs_by_pdf_product_reference(outputs, path.parent)
    selected, product_stability = _selected_products_by_stable_gp(outputs, threshold=0.90)
    # Do NOT filter article_day_fact globally: top pages must show category totals for all approved products.
    # The selected product set is applied only when building detailed category/product/article pages.

    daily = outputs.get("article_day_fact", pd.DataFrame()).copy()
    if daily is None or daily.empty:
        raise RuntimeError("PDF невозможен: пустой article_day_fact")
    daily["day"] = pd.to_datetime(daily["day"], errors="coerce").dt.normalize()
    for cnum in ["order_sum", "orders", "gross_profit_model", "ad_spend_total", "open_cards", "add_to_cart"]:
        if cnum in daily.columns:
            daily[cnum] = pd.to_numeric(daily[cnum], errors="coerce").fillna(0)
    # Technical files written by older versions may not have derived ad totals. Rebuild them for PDF-only.
    def _sum_existing(cols):
        s = pd.Series(0.0, index=daily.index)
        for col in cols:
            if col in daily.columns:
                s = s + pd.to_numeric(daily[col], errors="coerce").fillna(0)
        return s
    if "ad_spend_total" not in daily.columns:
        daily["ad_spend_total"] = _sum_existing(["manual_spend", "unified_spend", "unknown_spend", "ad_spend_model"])
    if "ad_clicks_total" not in daily.columns:
        daily["ad_clicks_total"] = _sum_existing(["manual_clicks", "unified_clicks", "unknown_clicks"])
    if "ad_impressions_total" not in daily.columns:
        daily["ad_impressions_total"] = _sum_existing(["manual_impressions", "unified_impressions", "unknown_impressions"])
    week_start, week_end, prev_start, prev_end = _period_bounds_from_daily(daily)
    latest = pd.to_datetime(daily["day"], errors="coerce").max()
    latest = pd.Timestamp(latest).normalize() if pd.notna(latest) else week_end
    cur_monday = latest - pd.Timedelta(days=int(latest.weekday()))
    cur_week_end = cur_monday + pd.Timedelta(days=6)

    factor_bridge = outputs.get("factor_bridge", pd.DataFrame()).copy()
    if factor_bridge is None:
        factor_bridge = pd.DataFrame()
    if not factor_bridge.empty:
        # In the PDF we compare factors to previous week / plan, not to an abstract optimum.
        if "comparison" in factor_bridge.columns:
            factor_bridge = factor_bridge[~factor_bridge["comparison"].astype(str).eq("optimal_best_days")].copy()
        if "factor" in factor_bridge.columns:
            factor_bridge = factor_bridge[~factor_bridge["factor"].astype(str).str.contains("оптим", case=False, na=False)].copy()
        factor_bridge["effect_gp_rub"] = pd.to_numeric(factor_bridge.get("effect_gp_rub"), errors="coerce").fillna(0)
        if "abs_effect_gp_rub" not in factor_bridge.columns:
            factor_bridge["abs_effect_gp_rub"] = factor_bridge["effect_gp_rub"].abs()
        if "comparison" not in factor_bridge.columns:
            factor_bridge["comparison"] = "prev_week"
        if "factor_block" not in factor_bridge.columns:
            factor_bridge["factor_block"] = factor_bridge.apply(lambda r: _factor_block_name(r.get("factor"), r.get("zone")), axis=1)
    factor_summary = outputs.get("factor_summary_for_pdf", pd.DataFrame())
    if factor_summary is None or factor_summary.empty or "economy_sales_text" not in factor_summary.columns:
        factor_summary = _build_factor_summary_df(factor_bridge)
    entry_bridge = outputs.get("entry_points_bridge", pd.DataFrame()).copy()
    if entry_bridge is None:
        entry_bridge = pd.DataFrame()
    opt = outputs.get("optimal_benchmarks", pd.DataFrame()).copy()
    if opt is None:
        opt = pd.DataFrame()

    search_unique_demand = outputs.get("search_unique_demand", pd.DataFrame()).copy()
    if search_unique_demand is None:
        search_unique_demand = pd.DataFrame()
    if not search_unique_demand.empty:
        if "day" in search_unique_demand.columns:
            search_unique_demand["day"] = pd.to_datetime(search_unique_demand["day"], errors="coerce").dt.normalize()
        if "subject_disp" not in search_unique_demand.columns:
            if "subject" in search_unique_demand.columns:
                search_unique_demand["subject_disp"] = search_unique_demand["subject"].map(_subject_disp)
            else:
                search_unique_demand["subject_disp"] = ""
        if "product_code" not in search_unique_demand.columns:
            if "product" in search_unique_demand.columns:
                search_unique_demand["product_code"] = search_unique_demand["product"].map(_prod)
            else:
                search_unique_demand["product_code"] = ""
        if "supplier_article" in search_unique_demand.columns:
            search_unique_demand["supplier_article"] = search_unique_demand["supplier_article"].map(_clean_article_local)
        search_unique_demand = _normalize_pdf_merge_keys(search_unique_demand, ["subject_disp", "product_code", "supplier_article", "nm_id"])
        for _c in ["unique_search_frequency", "unique_search_queries", "duplicate_query_rows_removed", "raw_query_rows"]:
            if _c not in search_unique_demand.columns:
                search_unique_demand[_c] = 0
            search_unique_demand[_c] = pd.to_numeric(search_unique_demand[_c], errors="coerce").fillna(0)

    trace_rows: List[Dict[str, Any]] = []
    log(f"PDF_DEBUG: article_day_fact rows={len(daily):,}, period={daily['day'].min().date() if not daily.empty else '-'}..{daily['day'].max().date() if not daily.empty else '-'}, cols={len(daily.columns):,}")
    log(f"PDF_DEBUG: abc_weekly rows={len(outputs.get('abc_weekly', pd.DataFrame())):,}, abc_monthly rows={len(outputs.get('abc_monthly', pd.DataFrame())):,}")

    def add_trace(block: str, level: str, period_start: Any, period_end: Any, metric: str, value: Any, source: str, source_columns: str, filters: str = "", formula: str = "", source_file: str = "", source_sheet: str = "", source_rows: str = "", source_cells: str = ""):
        trace_rows.append({
            "block": block,
            "level": level,
            "period_start": str(pd.Timestamp(period_start).date()) if pd.notna(period_start) else "",
            "period_end": str(pd.Timestamp(period_end).date()) if pd.notna(period_end) else "",
            "metric": metric,
            "value": value,
            "source": source,
            "source_file": source_file,
            "source_sheet": source_sheet,
            "source_rows": source_rows,
            "source_cells": source_cells,
            "source_columns": source_columns,
            "filters": filters,
            "formula": formula,
        })

    c = canvas.Canvas(str(path), pagesize=(W, H))
    bookmarks: Dict[str, str] = {}
    page_num = 0
    cat_short = {"Кисти косметические":"Кисти", "Косметические карандаши":"Карандаши", "Помады":"Помады", "Блески":"Блески"}
    cats = ["Кисти косметические", "Косметические карандаши", "Помады", "Блески"]
    cat_code = {"Кисти косметические": "brushes", "Косметические карандаши": "pencils", "Помады": "lipsticks", "Блески": "glosses"}
    def cat_bookmark(subject_value: Any) -> str:
        return "cat_" + cat_code.get(str(subject_value), re.sub(r"[^A-Za-z0-9]+", "_", str(subject_value)))
    def product_bookmark(subject_value: Any, product_value: Any) -> str:
        return "prod_" + cat_code.get(str(subject_value), re.sub(r"[^A-Za-z0-9]+", "_", str(subject_value))) + "_" + re.sub(r"[^A-Za-z0-9]+", "_", str(product_value))

    def page_bg(title: str, subtitle: str = "", section: str = ""):
        nonlocal page_num
        page_num += 1
        c.setFillColor(RED); c.rect(0, 0, W, H, fill=1, stroke=0)
        c.setFillColor(WHITE); c.setFont(F_REG, 32); c.drawString(70, 838, "topface")
        c.setFont(F_BLACK, 46); c.drawString(70, 765, title)
        if subtitle:
            c.setFont(F_BOLD, 20); c.drawString(70, 724, subtitle)
        if section:
            c.setFont(F_BOLD, 14); c.drawRightString(W-75, 712, section)
        c.setFont(F_BOLD, 13); c.drawRightString(W-75, 36, f"Страница {page_num}")

    def button(x, y, w, label, target=None):
        c.setFillColor(WHITE); c.roundRect(x, y, w, 42, 14, fill=1, stroke=0)
        c.setFillColor(RED_DARK); c.setFont(F_BOLD, 13); c.drawCentredString(x+w/2, y+16, label)
        if target:
            c.linkRect("", str(target), (x, y, x+w, y+42), relative=0)

    def top_nav(active=""):
        labels = [("cur", "Текущая"), ("prev", "Прошлая"), ("month", "Месяц"), ("closed", "Закр. месяц"), ("summary", "Сводка")]
        x = 880
        for key, lab in labels:
            button(x, 800, 126, lab, key)
            x += 140

    def dyn_text(cur, prev, lower=False):
        return _pdf_color_delta_value(cur, prev, lower)[0]

    def tone_color(tone: str):
        return GREEN if tone == "good" else BAD if tone == "bad" else GRAY

    def draw_metric_card(x, y, w, h, value, label, dyn="", tone="neutral", sub2="", sub2_tone="neutral"):
        c.setFillColor(WHITE); c.roundRect(x, y, w, h, 16, fill=1, stroke=0)
        val = str(value)
        c.setFont(F_BLACK, 27)
        val_w = stringWidth(val, F_BLACK, 27)
        base_x = x + w/2 - (val_w/2 if dyn else val_w/2)
        # When there is dynamic, reserve space to the right so the arrow is not drawn below.
        if dyn:
            dyn_w = stringWidth(str(dyn), F_BOLD, 12)
            total_w = val_w + 12 + dyn_w
            base_x = max(x + 14, x + w/2 - total_w/2)
        c.setFillColor(BLACK); c.setFont(F_BLACK, 27); c.drawString(base_x, y+h-42, val)
        if dyn:
            c.setFillColor(tone_color(tone)); c.setFont(F_BOLD, 12); c.drawString(base_x + val_w + 12, y+h-39, str(dyn))
        c.setFillColor(GRAY); c.setFont(F_REG, 14); c.drawCentredString(x+w/2, y+h-72, str(label))
        if sub2:
            c.setFillColor(tone_color(sub2_tone)); c.setFont(F_BOLD, 12); c.drawCentredString(x+w/2, y+18, sub2)

    def draw_table(x, y, w, h, headers, rows, col_widths=None, font_size=13, row_h=44, first_col_red=True, align_left_cols=None, lower_better_cols=None):
        align_left_cols = set(align_left_cols or [])
        lower_better_cols = set(lower_better_cols or [])
        if col_widths is None:
            col_widths = [w/len(headers)]*len(headers)
        c.setFillColor(WHITE); c.roundRect(x, y, w, h, 18, fill=1, stroke=0)
        c.setFillColor(RED_DARK); c.roundRect(x, y+h-50, w, 50, 12, fill=1, stroke=0)
        xx = x
        c.setFillColor(WHITE); c.setFont(F_BLACK, font_size)
        for i, head in enumerate(headers):
            lines = str(head).split("\n")
            yy = y+h-29 + (len(lines)-1)*8
            for line in lines:
                c.drawCentredString(xx+col_widths[i]/2, yy, line)
                yy -= 15
            xx += col_widths[i]
        yy = y+h-50-row_h
        for ridx, row in enumerate(rows):
            c.setFillColor(WHITE if ridx % 2 == 0 else SOFT); c.rect(x, yy, w, row_h, fill=1, stroke=0)
            xx = x
            for i, val in enumerate(row):
                text = str(val)
                lines = text.split("\n")
                if i == 0 and first_col_red:
                    c.setFillColor(SOFT); c.roundRect(xx+6, yy+5, col_widths[i]-12, row_h-10, 8, fill=1, stroke=0)
                    c.setFillColor(RED_DARK); c.setFont(F_BLACK, font_size)
                else:
                    c.setFillColor(BLACK); c.setFont(F_BOLD, font_size)
                # If the value is written as two lines: value + arrow, render the arrow to the right.
                if len(lines) >= 2 and lines[1].strip().startswith(("↑", "↓", "→")):
                    lines = [lines[0] + " " + lines[1].strip()] + lines[2:]
                line_y = yy + row_h/2 + (len(lines)-1)*8
                for line in lines:
                    raw = line[:70]
                    arrow_pos = min([p for p in [raw.find("↑"), raw.find("↓"), raw.find("→")] if p >= 0], default=-1)
                    if arrow_pos >= 0:
                        main_part = raw[:arrow_pos].rstrip()
                        delta_part = raw[arrow_pos:].strip()
                        is_lower = i in lower_better_cols
                        if delta_part.startswith("↑"):
                            dcolor = BAD if is_lower else GREEN
                        elif delta_part.startswith("↓"):
                            dcolor = GREEN if is_lower else BAD
                        else:
                            dcolor = GRAY
                        if i in align_left_cols:
                            start_x = xx + 12
                        else:
                            total_w = stringWidth(main_part, F_BOLD, font_size) + 8 + stringWidth(delta_part, F_BOLD, max(font_size-1, 9))
                            start_x = xx + col_widths[i]/2 - total_w/2
                        c.setFillColor(BLACK); c.setFont(F_BOLD, font_size); c.drawString(start_x, line_y, main_part)
                        c.setFillColor(dcolor); c.setFont(F_BOLD, max(font_size-1, 9)); c.drawString(start_x + stringWidth(main_part, F_BOLD, font_size) + 8, line_y, delta_part)
                    else:
                        if raw.strip().startswith(("↑", "↓", "→")):
                            c.setFillColor(GREEN if raw.strip().startswith("↑") else BAD if raw.strip().startswith("↓") else GRAY)
                        if i in align_left_cols:
                            c.drawString(xx+12, line_y, raw)
                        else:
                            c.drawCentredString(xx+col_widths[i]/2, line_y, raw)
                    line_y -= 16
                    c.setFillColor(BLACK)
                xx += col_widths[i]
            yy -= row_h


    def add_table_row_links(x, y, w, h, row_h, targets, header_h=50):
        """Overlay clickable areas on table body rows. Targets may point to bookmarks created later."""
        yy = y + h - header_h - row_h
        for target in targets:
            if target:
                c.linkRect("", str(target), (x, yy, x + w, yy + row_h), relative=0)
            yy -= row_h

    def agg(start, end, keys):
        return _agg_daily_for_bridge(daily, pd.Timestamp(start), pd.Timestamp(end), keys)

    def gp_week(keys, start=week_start, end=week_end):
        return _gp_from_abc_frames(outputs, start, end, keys)

    def with_gp(a: pd.DataFrame, keys: List[str], start=week_start, end=week_end) -> pd.DataFrame:
        a = a.copy() if a is not None else pd.DataFrame(columns=keys)
        period_start = pd.Timestamp(start).normalize()
        period_end = pd.Timestamp(end).normalize()
        # Use exact ABC for any exact closed week/month that exists, including the current report week
        # when WB ABC has already been exported. If exact ABC is absent, fallback is model GP.
        gp = gp_week(keys, start, end)
        if gp is not None and not gp.empty:
            a = a.merge(gp, on=keys, how="left")
        # Previous exact ABC for dynamics. For week pages this is the previous week.
        prev_len = period_end - period_start + pd.Timedelta(days=1)
        prev_period_start = period_start - prev_len
        prev_period_end = period_start - pd.Timedelta(days=1)
        gp_prev_exact = _gp_from_abc_frames(outputs, prev_period_start, prev_period_end, keys)
        if gp_prev_exact is not None and not gp_prev_exact.empty:
            rename_prev = {c: f"{c}_prev_abc" for c in gp_prev_exact.columns if c not in keys}
            a = a.merge(gp_prev_exact.rename(columns=rename_prev), on=keys, how="left")
        if "gp_fact" not in a.columns:
            a["gp_fact"] = np.nan
        if len(a.index):
            model_gp = pd.to_numeric(a.get("gross_profit_model", 0), errors="coerce").fillna(0)
            fact_gp = pd.to_numeric(a["gp_fact"], errors="coerce")
            a["gp_use"] = fact_gp.where(fact_gp.notna(), model_gp)
            a["gp_is_fact"] = fact_gp.notna()
            model_gp_prev = pd.to_numeric(a["gross_profit_model_prev"] if "gross_profit_model_prev" in a.columns else pd.Series(0.0, index=a.index), errors="coerce").fillna(0)
            fact_gp_prev = pd.to_numeric(a["gp_fact_prev_abc"] if "gp_fact_prev_abc" in a.columns else pd.Series(np.nan, index=a.index), errors="coerce")
            a["gp_use_prev"] = fact_gp_prev.where(fact_gp_prev.notna(), model_gp_prev)
            # ABC is source of truth for weekly margin and DRR when exact ABC exists.
            if "abc_drr_pct" in a.columns:
                a["drr_pct"] = pd.to_numeric(a["abc_drr_pct"], errors="coerce").where(pd.to_numeric(a["abc_drr_pct"], errors="coerce").notna(), pd.to_numeric(a["drr_pct"] if "drr_pct" in a.columns else pd.Series(np.nan, index=a.index), errors="coerce"))
            if "abc_drr_pct_prev_abc" in a.columns:
                a["drr_pct_prev"] = pd.to_numeric(a["abc_drr_pct_prev_abc"], errors="coerce").where(pd.to_numeric(a["abc_drr_pct_prev_abc"], errors="coerce").notna(), pd.to_numeric(a["drr_pct_prev"] if "drr_pct_prev" in a.columns else pd.Series(np.nan, index=a.index), errors="coerce"))
            if "abc_commission_pct" in a.columns:
                a["commission_pct_abc"] = pd.to_numeric(a["abc_commission_pct"], errors="coerce")
            if "abc_commission_pct_prev_abc" in a.columns:
                a["commission_pct_abc_prev"] = pd.to_numeric(a["abc_commission_pct_prev_abc"], errors="coerce")
            if "abc_acquiring_pct" in a.columns:
                a["acquiring_pct_abc"] = pd.to_numeric(a["abc_acquiring_pct"], errors="coerce")
            if "abc_acquiring_pct_prev_abc" in a.columns:
                a["acquiring_pct_abc_prev"] = pd.to_numeric(a["abc_acquiring_pct_prev_abc"], errors="coerce")
            order_sum_num = pd.to_numeric(a.get("order_sum", 0), errors="coerce").fillna(0)
            abc_margin = pd.to_numeric(a["abc_margin_pct"] if "abc_margin_pct" in a.columns else pd.Series(np.nan, index=a.index), errors="coerce")
            fact_rev = pd.to_numeric(a["gross_revenue_fact"] if "gross_revenue_fact" in a.columns else pd.Series(np.nan, index=a.index), errors="coerce")
            margin_calc = np.where(fact_rev.fillna(0) > 0, pd.to_numeric(a["gp_use"], errors="coerce") / fact_rev * 100, np.where(order_sum_num > 0, pd.to_numeric(a["gp_use"], errors="coerce") / order_sum_num * 100, np.nan))
            a["margin_pct"] = abc_margin.where(abc_margin.notna(), margin_calc)
            abc_margin_prev = pd.to_numeric(a["abc_margin_pct_prev_abc"] if "abc_margin_pct_prev_abc" in a.columns else pd.Series(np.nan, index=a.index), errors="coerce")
            fact_rev_prev = pd.to_numeric(a["gross_revenue_fact_prev_abc"] if "gross_revenue_fact_prev_abc" in a.columns else pd.Series(np.nan, index=a.index), errors="coerce")
            order_sum_prev_num = pd.to_numeric(a["order_sum_prev"] if "order_sum_prev" in a.columns else pd.Series(0.0, index=a.index), errors="coerce").fillna(0)
            margin_prev_calc = np.where(fact_rev_prev.fillna(0) > 0, pd.to_numeric(a["gp_use_prev"], errors="coerce") / fact_rev_prev * 100, np.where(order_sum_prev_num > 0, pd.to_numeric(a["gp_use_prev"], errors="coerce") / order_sum_prev_num * 100, np.nan))
            a["margin_pct_prev"] = abc_margin_prev.where(abc_margin_prev.notna(), margin_prev_calc)
            try:
                source_cnt = int(a["gp_is_fact"].fillna(False).sum())
                source_txt = "ABC exact" if source_cnt else "model gross_profit_model"
                log(f"PDF_GP: period={period_start.date()}..{period_end.date()} group={keys} rows={len(a):,} abc_rows={source_cnt:,} gp_sum={pd.to_numeric(a['gp_use'], errors='coerce').fillna(0).sum():,.0f} source={source_txt}")
                for _, tr in a.iterrows():
                    key_filter = "; ".join([f"{k}={tr.get(k, '')}" for k in keys])
                    if bool(tr.get("gp_is_fact", False)):
                        add_trace("gross_profit", "+".join(keys), period_start, period_end, "ВП", float(_pdf_num(tr.get("gp_use"), 0)), "ABC exact", "Валовая прибыль", key_filter, "SUM(ABC[Валовая прибыль]) по exact period_start/period_end и ключам", str(tr.get("gp_source_file", "")), str(tr.get("gp_source_sheet", "")), str(tr.get("gp_source_rows", "")), str(tr.get("gp_source_cells", "")))
                    else:
                        add_trace("gross_profit", "+".join(keys), period_start, period_end, "ВП", float(_pdf_num(tr.get("gp_use"), 0)), "article_day_fact model", "gross_profit_model", key_filter, "SUM(article_day_fact[gross_profit_model]) по периоду и ключам")
            except Exception as exc:
                log(f"WARN PDF_TRACE with_gp failed: {exc}")
        else:
            a["gp_use"] = pd.Series(dtype=float)
            a["gp_is_fact"] = pd.Series(dtype=bool)
            a["margin_pct"] = pd.Series(dtype=float)
        return a

    # ------------------------------------------------------------------
    # 1. Current week by days. Strict checklist format.
    # ------------------------------------------------------------------
    bookmarks["cur"] = "cur"; c.bookmarkPage("cur")
    page_bg("Текущая неделя", f"{cur_monday.strftime('%d.%m')}-{cur_week_end.strftime('%d.%m.%Y')} / оперативно: дни и план", "Текущая неделя")
    top_nav("cur")
    cur_period = agg(cur_monday, latest, ["subject"])
    prev_same = agg(cur_monday-pd.Timedelta(days=7), latest-pd.Timedelta(days=7), ["subject"])
    cur_period_gp = with_gp(cur_period, ["subject"], cur_monday, cur_week_end if latest >= cur_week_end else latest)
    prev_same_gp = with_gp(prev_same, ["subject"], cur_monday-pd.Timedelta(days=7), cur_week_end-pd.Timedelta(days=7) if latest >= cur_week_end else latest-pd.Timedelta(days=7))
    cur_total = cur_period.sum(numeric_only=True)
    prev_total = prev_same.sum(numeric_only=True) if not prev_same.empty else pd.Series(dtype=float)
    cur_gp_total = pd.to_numeric(cur_period_gp.get("gp_use", pd.Series(dtype=float)), errors="coerce").fillna(0).sum() if not cur_period_gp.empty else _pdf_num(cur_total.get("gross_profit_model"),0)
    prev_gp_total = pd.to_numeric(prev_same_gp.get("gp_use", pd.Series(dtype=float)), errors="coerce").fillna(0).sum() if not prev_same_gp.empty else _pdf_num(prev_total.get("gross_profit_model"),0)
    cur_gp_fact_flag = bool(cur_period_gp.get("gp_is_fact", pd.Series(dtype=bool)).fillna(False).any()) if not cur_period_gp.empty else False
    drr_cur = _pdf_num(cur_total.get("ad_spend_total"), 0) / _pdf_num(cur_total.get("order_sum"), 1) * 100 if _pdf_num(cur_total.get("order_sum"), 0) else np.nan
    drr_prev = _pdf_num(prev_total.get("ad_spend_total"), 0) / _pdf_num(prev_total.get("order_sum"), 1) * 100 if _pdf_num(prev_total.get("order_sum"), 0) else np.nan
    dt, tone = _pdf_color_delta_value(cur_total.get("order_sum",0), prev_total.get("order_sum",0), False)
    draw_metric_card(70, 610, 260, 120, _fmt_rub(cur_total.get("order_sum", 0)), "Сумма заказов", dt, tone)
    dt, tone = _pdf_color_delta_value(cur_gp_total, prev_gp_total, False)
    draw_metric_card(360, 610, 260, 120, _fmt_rub(cur_gp_total), "ВП факт ABC" if cur_gp_fact_flag else "ВП расч.", dt, tone)
    dt, tone = _pdf_color_delta_value(drr_cur, drr_prev, True)
    draw_metric_card(650, 610, 260, 120, _fmt_pct_pdf(drr_cur), "ДРР", dt, tone)
    dt, tone = _pdf_color_delta_value(cur_total.get("ad_spend_total",0), prev_total.get("ad_spend_total",0), True)
    draw_metric_card(940, 610, 260, 120, _fmt_rub(cur_total.get("ad_spend_total", 0)), "Расход РК", dt, tone)
    plan_day = max(0, _pdf_num(cur_total.get("order_sum"), 0) / max(1, (latest-cur_monday).days+1) * 1.1)
    draw_metric_card(1230, 610, 260, 120, _fmt_rub(plan_day), "План/день", "по сумме", "neutral")
    headers = ["Категория", "Пн\n"+cur_monday.strftime("%d.%m"), "Вт\n"+(cur_monday+pd.Timedelta(days=1)).strftime("%d.%m"), "Ср\n"+(cur_monday+pd.Timedelta(days=2)).strftime("%d.%m"), "Чт\n"+(cur_monday+pd.Timedelta(days=3)).strftime("%d.%m"), "Пт\n"+(cur_monday+pd.Timedelta(days=4)).strftime("%d.%m"), "Сб\n"+(cur_monday+pd.Timedelta(days=5)).strftime("%d.%m"), "Вс\n"+(cur_monday+pd.Timedelta(days=6)).strftime("%d.%m"), "План/день"]
    rows=[]
    day_agg = with_gp(agg(cur_monday, cur_week_end, ["day", "subject"]), ["day", "subject"], cur_monday, cur_week_end)
    day_prev = with_gp(agg(cur_monday-pd.Timedelta(days=7), cur_week_end-pd.Timedelta(days=7), ["day", "subject"]), ["day", "subject"], cur_monday-pd.Timedelta(days=7), cur_week_end-pd.Timedelta(days=7))
    for cat in cats:
        vals=[f"{cat_short.get(cat,cat)}\nСумма\nВП расч.\nРасх. РК\nДРР"]
        for i in range(7):
            day = cur_monday + pd.Timedelta(days=i)
            p = day_agg[(day_agg["day"] == day) & (day_agg["subject"] == cat)] if not day_agg.empty else pd.DataFrame()
            pp = day_prev[(day_prev["day"] == day-pd.Timedelta(days=7)) & (day_prev["subject"] == cat)] if not day_prev.empty else pd.DataFrame()
            if p.empty or day > latest:
                vals.append("—\n—\n—\n—")
            else:
                r = p.iloc[0]
                rp = pp.iloc[0] if not pp.empty else pd.Series(dtype=object)
                dt_sum = dyn_text(r.get("order_sum"), rp.get("order_sum"), False)
                vals.append(f"{_fmt_rub(r.get('order_sum'), True)} {dt_sum}\n{_fmt_rub(r.get('gp_use'), True)}\n{_fmt_rub(r.get('ad_spend_total'), True)}\n{_fmt_pct_pdf(r.get('drr_pct'))}")
        vals.append(f"{_fmt_rub(plan_day/4, True)}\n—\n—\n—")
        rows.append(vals)
    draw_table(70, 80, 1460, 470, headers, rows, col_widths=[190]+[145]*7+[255], font_size=12, row_h=92)
    c.showPage()

    # Current category summary.
    bookmarks["cur_cat"] = "cur_cat"; c.bookmarkPage("cur_cat")
    page_bg("Текущая неделя: категории", f"{cur_monday.strftime('%d.%m')}-{latest.strftime('%d.%m.%Y')} / переход по категории", "Текущая неделя")
    top_nav("cur")
    cat_cur = with_gp(_merge_cur_prev(cur_period, prev_same, ["subject"]), ["subject"], cur_monday, latest)
    rows=[]
    for _, r in cat_cur.sort_values("order_sum", ascending=False).iterrows():
        rows.append([
            cat_short.get(r.get("subject"), r.get("subject")),
            f"{_fmt_rub(r.get('order_sum'))}\n{dyn_text(r.get('order_sum'), r.get('order_sum_prev'))}",
            f"{_fmt_rub(r.get('gp_use'))}",
            _fmt_pct_pdf(r.get("margin_pct")),
            f"{_fmt_pct_pdf(r.get('drr_pct'))}\n{dyn_text(r.get('drr_pct'), r.get('drr_pct_prev'), True)}",
            f"{_fmt_rub(r.get('ad_spend_total'))}",
            f"{_fmt_cpc_pdf(r.get('cpc'))}\n{dyn_text(r.get('cpc'), r.get('cpc_prev'), True)}",
            _fmt_pct_pdf(r.get("search_traffic_capture_pct")),
        ])
    cur_gp_label = "ВП факт ABC" if bool(cat_cur.get("gp_is_fact", pd.Series(dtype=bool)).fillna(False).any()) else "ВП расч."
    draw_table(80, 260, 1440, 400, ["Категория", "Сумма", cur_gp_label, "Рент.", "ДРР", "Расход РК", "CPC", "% поиска"], rows, col_widths=[210,210,190,150,190,190,160,150], font_size=14, row_h=74, lower_better_cols={4,5,6})
    # Category rows are clickable: category -> products; brushes -> article list on the category page.
    cat_targets = [cat_bookmark(str(r.get("subject"))) for _, r in cat_cur.sort_values("order_sum", ascending=False).iterrows()]
    add_table_row_links(80, 260, 1440, 400, 74, cat_targets)
    c.showPage()

    # Previous week, current month, closed month, monthly summary.
    def draw_period_category_page(bookmark, title, subtitle, section, start, end, prev_s, prev_e, gp_fact_label="ВП факт ABC"):
        bookmarks[bookmark] = bookmark; c.bookmarkPage(bookmark)
        page_bg(title, subtitle, section); top_nav(bookmark)
        cur = with_gp(agg(start, end, ["subject"]), ["subject"], start, end)
        prev = with_gp(agg(prev_s, prev_e, ["subject"]), ["subject"], prev_s, prev_e)
        merged = _merge_cur_prev(cur, prev, ["subject"])
        rows = []
        for _, r in merged.sort_values("order_sum", ascending=False).iterrows():
            gp = r.get("gp_use", r.get("gross_profit_model", 0))
            gp_prev = r.get("gp_use_prev", r.get("gross_profit_model_prev", 0))
            rows.append([
                cat_short.get(r.get("subject"), r.get("subject")),
                f"{_fmt_rub(r.get('order_sum'))}\n{dyn_text(r.get('order_sum'), r.get('order_sum_prev'))}",
                f"{_fmt_rub(gp)}\n{dyn_text(gp, gp_prev)}",
                _fmt_pct_pdf(r.get("margin_pct")),
                f"{_fmt_pct_pdf(r.get('drr_pct'))}\n{dyn_text(r.get('drr_pct'), r.get('drr_pct_prev'), True)}",
                f"{_fmt_rub(r.get('ad_spend_total'))}\n{dyn_text(r.get('ad_spend_total'), r.get('ad_spend_total_prev'), True)}",
                f"{_fmt_cpc_pdf(r.get('cpc'))}\n{dyn_text(r.get('cpc'), r.get('cpc_prev'), True)}",
            ])
        draw_table(80, 235, 1440, 460, ["Категория", "Сумма", gp_fact_label, "Рент.", "ДРР", "Расход РК", "CPC"], rows, col_widths=[230,230,230,150,180,230,150], font_size=14, row_h=80, lower_better_cols={4,5,6})
        c.showPage()

    prev_report_start = cur_monday - pd.Timedelta(days=7)
    prev_report_end = cur_monday - pd.Timedelta(days=1)
    prev_report_cmp_start = cur_monday - pd.Timedelta(days=14)
    prev_report_cmp_end = cur_monday - pd.Timedelta(days=8)
    draw_period_category_page("prev", "Прошлая неделя", f"{prev_report_start.strftime('%d.%m')}-{prev_report_end.strftime('%d.%m.%Y')} / сравнение с {prev_report_cmp_start.strftime('%d.%m')}-{prev_report_cmp_end.strftime('%d.%m.%Y')}", "Прошлая неделя", prev_report_start, prev_report_end, prev_report_cmp_start, prev_report_cmp_end)

    month_start = pd.Timestamp(latest.replace(day=1))
    prev_month_end = month_start - pd.Timedelta(days=1)
    prev_month_start = pd.Timestamp(prev_month_end.replace(day=1))
    draw_period_category_page("month", "Текущий месяц", f"{month_start.strftime('%d.%m')}-{latest.strftime('%d.%m.%Y')} / месяц неполный / темп к плану", "Текущий месяц", month_start, latest, prev_month_start, prev_month_start + (latest - month_start), "ВП расч.")
    closed_start = prev_month_start
    closed_end = prev_month_end
    preclosed_end = closed_start - pd.Timedelta(days=1)
    preclosed_start = pd.Timestamp(preclosed_end.replace(day=1))
    draw_period_category_page("closed", "Последний закрытый месяц", f"{closed_start.strftime('%d.%m')}-{closed_end.strftime('%d.%m.%Y')} / сравнение с {preclosed_start.strftime('%d.%m')}-{preclosed_end.strftime('%d.%m.%Y')}", "Закрытый месяц", closed_start, closed_end, preclosed_start, preclosed_end)

    bookmarks["summary"] = "summary"; c.bookmarkPage("summary")
    page_bg("Сводка по месяцам", "Категории / без лишней детализации", "Сводка"); top_nav("summary")
    monthly_rows=[]
    mm = daily.copy()
    mm["month"] = mm["day"].dt.strftime("%m.%Y")
    m_agg = mm.groupby(["month", "subject"], as_index=False).agg(order_sum=("order_sum", "sum"), gp_model=("gross_profit_model", "sum"), ad=("ad_spend_total", "sum"))
    # Closed months must use exact ABC fact, otherwise this page disagrees with the closed-month page.
    abc_m = outputs.get("abc_monthly", pd.DataFrame())
    if abc_m is not None and not abc_m.empty:
        am = abc_m.copy()
        am["period_start"] = pd.to_datetime(am.get("period_start"), errors="coerce").dt.normalize()
        am["period_end"] = pd.to_datetime(am.get("period_end"), errors="coerce").dt.normalize()
        am = am[am["period_end"] < month_start].copy()
        if not am.empty:
            am["month"] = am["period_start"].dt.strftime("%m.%Y")
            am_gp = am.groupby(["month", "subject"], as_index=False).agg(gp_fact=("gross_profit", "sum"))
            m_agg = m_agg.merge(am_gp, on=["month", "subject"], how="left")
    if "gp_fact" not in m_agg.columns:
        m_agg["gp_fact"] = np.nan
    m_agg["gp"] = pd.to_numeric(m_agg["gp_fact"], errors="coerce").where(pd.to_numeric(m_agg["gp_fact"], errors="coerce").notna(), pd.to_numeric(m_agg["gp_model"], errors="coerce").fillna(0))
    for _, r in m_agg.sort_values(["month", "subject"]).iterrows():
        drr = _pdf_num(r.get("ad"), 0) / _pdf_num(r.get("order_sum"), 1) * 100 if _pdf_num(r.get("order_sum"), 0) else np.nan
        monthly_rows.append([r.get("month"), cat_short.get(r.get("subject"), r.get("subject")), _fmt_rub(r.get("order_sum")), _fmt_rub(r.get("gp")), _fmt_pct_pdf(drr)])
    draw_table(180, 160, 1240, 560, ["Месяц", "Категория", "Сумма", "ВП", "ДРР"], monthly_rows[-16:], col_widths=[180,280,280,280,180], font_size=15, row_h=42)
    c.showPage()

    # ------------------------------------------------------------------
    # Category and product pages.
    # ------------------------------------------------------------------
    selected_products = selected.copy() if selected is not None and not selected.empty else pd.DataFrame(columns=["subject", "product"])
    selected_set = set((str(r.get("subject")), str(r.get("product"))) for _, r in selected_products.iterrows())
    product_pages: Dict[Tuple[str, str], str] = {}
    article_pages: Dict[str, Tuple[str, str]] = {}

    product_week = with_gp(_merge_cur_prev(agg(week_start, week_end, ["subject", "product"]), agg(prev_start, prev_end, ["subject", "product"]), ["subject", "product"]), ["subject", "product"], week_start, week_end)
    article_week = with_gp(_merge_cur_prev(agg(week_start, week_end, ["subject", "product", "supplier_article", "nm_id"]), agg(prev_start, prev_end, ["subject", "product", "supplier_article", "nm_id"]), ["subject", "product", "supplier_article", "nm_id"]), ["subject", "product", "supplier_article", "nm_id"], week_start, week_end)

    for subject in cats:
        cp = product_week[product_week["subject"].astype(str).eq(subject)].copy()
        cp = cp[cp["product"].astype(str).map(lambda p: (subject, p) in selected_set)].copy() if selected_set else cp
        cat_key = cat_bookmark(subject)
        if cp.empty:
            bookmarks[cat_key] = cat_key; c.bookmarkPage(cat_key)
            page_bg(f"Категория: {cat_short.get(subject, subject)}", f"{week_start.strftime('%d.%m')}-{week_end.strftime('%d.%m.%Y')} / нет товаров для детализации по правилу отбора", "Категория")
            button(1220, 800, 240, "← текущая неделя", "cur_cat")
            draw_table(180, 330, 1240, 220, ["Статус", "Комментарий"], [["Не детализируем", "Товары категории не прошли отбор стабильной ВП / исключены как хвостовые"]], col_widths=[260, 980], font_size=16, row_h=80, first_col_red=True, align_left_cols={1})
            c.showPage()
            continue
        products = list(cp["product"].astype(str).unique())
        skip_product_level = (len(products) == 1 and products[0] == "901")
        bookmarks[cat_key] = cat_key; c.bookmarkPage(cat_key)

        def article_keep_for_product(product_value: str) -> pd.DataFrame:
            aw0 = article_week[(article_week["subject"].astype(str).eq(subject)) & (article_week["product"].astype(str).eq(product_value))].copy()
            if aw0.empty:
                return aw0
            aw0 = aw0[(pd.to_numeric(aw0.get("order_sum"), errors="coerce").fillna(0) > 0) & (pd.to_numeric(aw0.get("orders"), errors="coerce").fillna(0) > 0) & (pd.to_numeric(aw0["gp_use"], errors="coerce").fillna(0) > 0)].copy()
            if aw0.empty:
                return aw0
            total_gp0 = pd.to_numeric(aw0["gp_use"], errors="coerce").fillna(0).sum()
            aw0 = aw0.sort_values("gp_use", ascending=False).copy()
            aw0["share"] = np.where(total_gp0 > 0, aw0["gp_use"] / total_gp0 * 100, 0)
            aw0["cum"] = aw0["share"].cumsum()
            keep0 = aw0[(aw0["cum"] <= 90) | (aw0.index == aw0.index[0])].copy()
            if len(keep0) < len(aw0):
                first_out0 = aw0[~aw0.index.isin(keep0.index)].head(1)
                if not first_out0.empty and _pdf_num(first_out0.iloc[0].get("gp_use"), 0) > 1000:
                    keep0 = pd.concat([keep0, first_out0])
            return keep0.head(14)

        if skip_product_level:
            # Brushes have only one meaningful product (901), so category page is immediately an article list.
            product = products[0]
            aw_cat = article_keep_for_product(product)
            page_bg(f"Категория: {cat_short.get(subject, subject)}", f"{week_start.strftime('%d.%m')}-{week_end.strftime('%d.%m.%Y')} / артикулы товара {product} / 90% ВП", "Категория")
            button(1220, 800, 240, "← текущая неделя", "cur_cat")
            rows=[]; row_targets=[]
            for _, ar in aw_cat.iterrows():
                art = str(ar.get("supplier_article"))
                page1 = f"article_{art}_1".replace("/", "_").replace(" ", "_")
                row_targets.append(page1)
                rows.append([
                    art,
                    f"{_fmt_rub(ar.get('order_sum'))}\n{dyn_text(ar.get('order_sum'), ar.get('order_sum_prev'))}",
                    f"{_fmt_rub(ar.get('gp_use'))}\n{dyn_text(ar.get('gp_use'), ar.get('gp_use_prev'))}",
                    _fmt_pct_pdf(ar.get("margin_pct")),
                    f"{_fmt_pct_pdf(ar.get('drr_pct'))}\n{dyn_text(ar.get('drr_pct'), ar.get('drr_pct_prev'), True)}",
                    f"{_fmt_cpc_pdf(ar.get('cpc'))}\n{dyn_text(ar.get('cpc'), ar.get('cpc_prev'), True)}",
                    f"{_fmt_pct_pdf(ar.get('search_traffic_capture_pct'))}\n{dyn_text(ar.get('search_traffic_capture_pct'), ar.get('search_traffic_capture_pct_prev'))}",
                    _fmt_pct_pdf(ar.get("localization_with_replacements_pct")),
                ])
            draw_table(90, 150, 1420, 500, ["Артикул", "Сумма", "ВП", "Рент.", "ДРР", "CPC", "% поиска", "Локал."], rows, col_widths=[220,200,200,130,130,130,170,150], font_size=13, row_h=58, lower_better_cols={4,5})
            add_table_row_links(90, 150, 1420, 500, 58, row_targets)
            c.showPage()
        else:
            page_bg(f"Категория: {cat_short.get(subject, subject)}", f"{week_start.strftime('%d.%m')}-{week_end.strftime('%d.%m.%Y')} / товары 90% стабильной ВП", "Категория")
            button(1220, 800, 240, "← текущая неделя", "cur_cat")
            rows=[]; row_targets=[]
            for _, r in cp.sort_values("gp_use", ascending=False).iterrows():
                product_value = str(r.get("product"))
                prod_key_tmp = product_bookmark(subject, product_value)
                row_targets.append(prod_key_tmp)
                rows.append([
                    product_value,
                    f"{_fmt_rub(r.get('order_sum'))}\n{dyn_text(r.get('order_sum'), r.get('order_sum_prev'))}",
                    f"{_fmt_rub(r.get('gp_use'))}\n{dyn_text(r.get('gp_use'), r.get('gp_use_prev'))}",
                    _fmt_pct_pdf(r.get("margin_pct")),
                    f"{_fmt_pct_pdf(r.get('drr_pct'))}\n{dyn_text(r.get('drr_pct'), r.get('drr_pct_prev'), True)}",
                    f"{_fmt_cpc_pdf(r.get('cpc'))}\n{dyn_text(r.get('cpc'), r.get('cpc_prev'), True)}",
                    f"{_fmt_pct_pdf(r.get('search_traffic_capture_pct'))}\n{dyn_text(r.get('search_traffic_capture_pct'), r.get('search_traffic_capture_pct_prev'))}",
                    _fmt_pct_pdf(r.get("localization_with_replacements_pct")),
                ])
            draw_table(90, 150, 1420, 500, ["Товар", "Сумма", "ВП", "Рент.", "ДРР", "CPC", "% поиска", "Локал."], rows, col_widths=[190,210,210,140,180,150,180,150], font_size=14, row_h=74, lower_better_cols={4,5})
            add_table_row_links(90, 150, 1420, 500, 74, row_targets)
            c.showPage()

        for product in products:
            aw = article_keep_for_product(product)
            if aw.empty:
                continue
            aw_keep = aw.copy()

            prod_key = product_bookmark(subject, product)
            if not skip_product_level:
                product_pages[(subject, product)] = prod_key
                bookmarks[prod_key] = prod_key; c.bookmarkPage(prod_key)
                page_bg(f"Товар: {product}", f"{cat_short.get(subject,subject)} / {week_start.strftime('%d.%m')}-{week_end.strftime('%d.%m.%Y')}", "Товар")
                button(1240, 800, 190, "← категория", cat_key)
                pr = cp[cp["product"].astype(str).eq(product)].iloc[0]
                dt, tone = _pdf_color_delta_value(pr.get("order_sum"), pr.get("order_sum_prev"), False)
                draw_metric_card(90, 620, 230, 100, _fmt_rub(pr.get("order_sum")), "Сумма", dt, tone)
                dt, tone = _pdf_color_delta_value(pr.get("gp_use"), pr.get("gp_use_prev"), False)
                draw_metric_card(340, 620, 230, 100, _fmt_rub(pr.get("gp_use")), "ВП", dt, tone)
                dt, tone = _pdf_color_delta_value(pr.get("drr_pct"), pr.get("drr_pct_prev"), True)
                draw_metric_card(590, 620, 230, 100, _fmt_pct_pdf(pr.get("drr_pct")), "ДРР", dt, tone)
                dt, tone = _pdf_color_delta_value(pr.get("cpc"), pr.get("cpc_prev"), True)
                draw_metric_card(840, 620, 230, 100, _fmt_cpc_pdf(pr.get("cpc")), "CPC", dt, tone)
                rows2=[]; art_targets=[]
                for _, ar in aw_keep.iterrows():
                    art = str(ar.get("supplier_article"))
                    art_targets.append(f"article_{art}_1".replace("/", "_").replace(" ", "_"))
                    rows2.append([
                        art,
                        f"{_fmt_rub(ar.get('order_sum'))}\n{dyn_text(ar.get('order_sum'), ar.get('order_sum_prev'))}",
                        f"{_fmt_rub(ar.get('gp_use'))}\n{dyn_text(ar.get('gp_use'), ar.get('gp_use_prev'))}",
                        _fmt_pct_pdf(ar.get("margin_pct")),
                        f"{_fmt_pct_pdf(ar.get('drr_pct'))}\n{dyn_text(ar.get('drr_pct'), ar.get('drr_pct_prev'), True)}",
                        f"{_fmt_cpc_pdf(ar.get('cpc'))}\n{dyn_text(ar.get('cpc'), ar.get('cpc_prev'), True)}",
                        f"{_fmt_pct_pdf(ar.get('search_traffic_capture_pct'))}\n{dyn_text(ar.get('search_traffic_capture_pct'), ar.get('search_traffic_capture_pct_prev'))}",
                        _fmt_pct_pdf(ar.get("localization_with_replacements_pct")),
                    ])
                draw_table(90, 150, 1420, 410, ["Артикул", "Сумма", "ВП", "Рент.", "ДРР", "CPC", "% поиска", "Локал."], rows2, col_widths=[220,200,200,130,130,130,170,150], font_size=13, row_h=58, lower_better_cols={4,5})
                add_table_row_links(90, 150, 1420, 410, 58, art_targets)
                c.showPage()

            # Article pages for top 90% GP inside selected product.
            for _, ar in aw_keep.iterrows():
                art = str(ar.get("supplier_article"))
                page1 = f"article_{art}_1".replace("/", "_").replace(" ", "_")
                page2 = f"article_{art}_2".replace("/", "_").replace(" ", "_")
                article_pages[art] = (page1, page2)
                bookmarks[page1] = page1; c.bookmarkPage(page1)
                page_bg(f"Артикул: {art}", f"{cat_short.get(subject,subject)} / товар {product} / {week_start.strftime('%d.%m')}-{week_end.strftime('%d.%m.%Y')}", "Артикул 1/2")
                if not skip_product_level:
                    button(1120, 800, 150, "← товар", prod_key)
                    button(1290, 800, 150, "← категория", cat_key)
                else:
                    button(1240, 800, 180, "← категория", cat_key)
                button(1440, 800, 90, "стр.2", page2)

                def unit(col):
                    qty = max(_pdf_num(ar.get("buyout_qty_model"), _pdf_num(ar.get("orders"), 1)), 1)
                    return _pdf_num(ar.get(col), 0) / qty
                def pct_of_sales(col, suffix=""):
                    denom = _pdf_num(ar.get("order_sum" + suffix), 0)
                    return (_pdf_num(ar.get(col + suffix), 0) / denom * 100) if denom else np.nan
                gp = _pdf_num(ar.get("gp_use"), 0)
                gp_label = "ВП факт ABC" if bool(ar.get("gp_is_fact", False)) else "ВП расч."
                cards1 = [
                    (_fmt_rub(ar.get("order_sum")), "Сумма заказов", ar.get("order_sum"), ar.get("order_sum_prev"), False, ""),
                    (_fmt_rub(gp), gp_label, gp, ar.get("gp_use_prev"), False, ""),
                    (_fmt_pct_pdf(ar.get("margin_pct")), "Рентабельность", ar.get("margin_pct"), ar.get("margin_pct_prev"), False, ""),
                    (_fmt_rub(ar.get("avg_order_price")), "Цена продажи", ar.get("avg_order_price"), ar.get("avg_order_price_prev"), False, ""),
                    (_fmt_rub(ar.get("price_with_disc")), "Цена покупателя", ar.get("price_with_disc"), ar.get("price_with_disc_prev"), False, ""),
                    (_fmt_pct_pdf(ar.get("spp", ar.get("spp_funnel",0))), "СПП", ar.get("spp", ar.get("spp_funnel",0)), ar.get("spp_prev", ar.get("spp_funnel_prev",0)), False, ""),
                    (_fmt_pct_pdf(ar.get("commission_pct_abc", pct_of_sales("commission_model"))), "Комиссия, %", ar.get("commission_pct_abc", pct_of_sales("commission_model")), ar.get("commission_pct_abc_prev", pct_of_sales("commission_model", "_prev")), True, ""),
                    (_fmt_rub(unit("logistics_direct_model")+unit("logistics_return_model")), "Логистика/шт", "", "", True, ""),
                    (_fmt_rub(unit("storage_model")), "Хранение/шт", "", "", True, ""),
                    (_fmt_pct_pdf(ar.get("acquiring_pct_abc", pct_of_sales("acquiring_model"))), "Эквайринг, %", ar.get("acquiring_pct_abc", pct_of_sales("acquiring_model")), ar.get("acquiring_pct_abc_prev", pct_of_sales("acquiring_model", "_prev")), True, ""),
                    (_fmt_rub(unit("cost_model")), "Себест./шт", "", "", True, ""),
                    (_fmt_rub(unit("other_costs_model")), "Прочие/шт", "", "", True, ""),
                ]
                c.setFillColor(RED_DARK); c.roundRect(80, 650, 1440, 42, 10, fill=1, stroke=0); c.setFillColor(WHITE); c.setFont(F_BLACK, 20); c.drawString(105, 664, "Блок 1. Экономика и продажи")
                for idx, (val, lab, curv, prevv, lower, sub2) in enumerate(cards1[:6]):
                    dt, tone = _pdf_color_delta_value(curv, prevv, lower) if curv != "" else ("", "neutral")
                    draw_metric_card(80+idx*240, 535, 220, 96, val, lab, dt, tone, sub2)
                for idx, (val, lab, curv, prevv, lower, sub2) in enumerate(cards1[6:]):
                    dt, tone = _pdf_color_delta_value(curv, prevv, lower) if curv != "" else ("", "neutral")
                    draw_metric_card(80+idx*240, 420, 220, 96, val, lab, dt, tone)
                c.setFillColor(RED_DARK); c.roundRect(80, 350, 1440, 42, 10, fill=1, stroke=0); c.setFillColor(WHITE); c.setFont(F_BLACK, 20); c.drawString(105, 364, "Блок 2. Спрос, точки входа и конверсии")
                cards2 = [
                    (_fmt_rub(ar.get("ad_spend_total")), "Расход РК", ar.get("ad_spend_total"), ar.get("ad_spend_total_prev"), True, ""),
                    (_fmt_pct_pdf(ar.get("drr_pct")), "ДРР", ar.get("drr_pct"), ar.get("drr_pct_prev"), True, ""),
                    (_fmt_cpc_pdf(ar.get("cpc")), "CPC", ar.get("cpc"), ar.get("cpc_prev"), True, ""),
                    (_fmt_num_pdf(ar.get("ad_impressions_total")), "Показы РК", ar.get("ad_impressions_total"), ar.get("ad_impressions_total_prev"), False, ""),
                    (_fmt_num_pdf(ar.get("ad_clicks_total")), "Клики РК", ar.get("ad_clicks_total"), ar.get("ad_clicks_total_prev"), False, ""),
                    (_fmt_num_pdf(ar.get("open_cards")), "Открытия", ar.get("open_cards"), ar.get("open_cards_prev"), False, ""),
                    (_fmt_pct_pdf(ar.get("cart_conv_pct")), "Конв. в корзину", ar.get("cart_conv_pct"), ar.get("cart_conv_pct_prev"), False, ""),
                    (_fmt_pct_pdf(ar.get("order_conv_pct")), "Корзина → заказ", ar.get("order_conv_pct"), ar.get("order_conv_pct_prev"), False, ""),
                    (_fmt_num_pdf(ar.get("search_frequency")), "Спрос WB", ar.get("search_frequency"), ar.get("search_frequency_prev"), False, ""),
                    (_fmt_pct_pdf(ar.get("search_traffic_capture_pct")), "% поиска", ar.get("search_traffic_capture_pct"), ar.get("search_traffic_capture_pct_prev"), False, ""),
                    (_fmt_pct_pdf(ar.get("localization_with_replacements_pct")), "Локализация", ar.get("localization_with_replacements_pct"), ar.get("localization_with_replacements_pct_prev"), False, ""),
                    (_fmt_num_pdf(ar.get("rating_reviews")), "Рейтинг отзывов", ar.get("rating_reviews"), ar.get("rating_reviews_prev"), False, ""),
                ]
                for idx, (val, lab, curv, prevv, lower, sub2) in enumerate(cards2[:6]):
                    dt, tone = _pdf_color_delta_value(curv, prevv, lower)
                    draw_metric_card(80+idx*240, 205, 220, 96, val, lab, dt, tone, sub2)
                for idx, (val, lab, curv, prevv, lower, sub2) in enumerate(cards2[6:]):
                    dt, tone = _pdf_color_delta_value(curv, prevv, lower)
                    draw_metric_card(80+idx*240, 90, 220, 96, val, lab, dt, tone, sub2)
                c.showPage()

                bookmarks[page2] = page2; c.bookmarkPage(page2)
                page_bg(f"Артикул: {art}", f"{cat_short.get(subject,subject)} / товар {product} / точки входа и выводы", "Артикул 2/2")
                if not skip_product_level:
                    button(1120, 800, 150, "← товар", prod_key)
                    button(1290, 800, 150, "← категория", cat_key)
                else:
                    button(1240, 800, 180, "← категория", cat_key)
                button(1440, 800, 90, "стр.1", page1)
                ep_rows=[]
                if entry_bridge is not None and not entry_bridge.empty:
                    ep = entry_bridge[(entry_bridge["subject"].astype(str).eq(subject)) & (entry_bridge["product"].astype(str).eq(product)) & (entry_bridge["supplier_article"].astype(str).eq(art))].copy()
                    if not ep.empty:
                        ep = ep.sort_values("orders", ascending=False).head(8)
                        for _, er in ep.iterrows():
                            dord, _tone = _pdf_color_delta_value(er.get("orders"), er.get("orders_prev"), False)
                            ep_rows.append([
                                f"{er.get('entry_section','')} / {er.get('entry_point','')}",
                                f"{_fmt_num_pdf(er.get('transitions'))}\n{dyn_text(er.get('transitions'), er.get('transitions_prev'))}",
                                f"{_fmt_num_pdf(er.get('orders'))}\n{dord}",
                                _fmt_pct_pdf(er.get("cart_conv_pct")),
                                _fmt_pct_pdf(er.get("order_conv_pct")),
                                _fmt_pct_pdf(er.get("orders_share_pct")),
                                _pdf_signed_rub(er.get("effect_gp_rub")),
                            ])
                if not ep_rows:
                    ep_rows = [["—", "—", "—", "—", "—", "—", "—"]]
                draw_table(90, 425, 1420, 300, ["Канал / точка входа", "Переходы", "Заказы", "Конв.\nкорз.", "Корзина\n→ заказ", "Доля\nзаказов", "Вклад\nВП"], ep_rows, col_widths=[470,170,150,140,160,140,170], font_size=12, row_h=31, first_col_red=False, align_left_cols={0})
                fs = factor_summary[(factor_summary["level"].astype(str).eq("article")) & (factor_summary["supplier_article"].astype(str).eq(art))] if factor_summary is not None and not factor_summary.empty else pd.DataFrame()
                txt_main = fs.iloc[0].get("summary_text", "Факторный мост не выделил значимых денежных причин.") if not fs.empty else "Факторный мост не выделил значимых денежных причин."
                txt_econ = fs.iloc[0].get("economy_sales_text", "") if not fs.empty else ""
                txt_dem = fs.iloc[0].get("demand_entry_conversion_text", "") if not fs.empty else ""
                c.setFillColor(WHITE); c.roundRect(90, 105, 1420, 270, 18, fill=1, stroke=0)
                c.setFillColor(RED_DARK); c.setFont(F_BLACK, 22); c.drawString(120, 340, "Факторный вывод в деньгах")
                c.setFillColor(BLACK); c.setFont(F_BOLD, 16)
                yy = 305
                for line in _wrap_pdf_lines(txt_main, F_BOLD, 16, 1340, 4):
                    c.drawString(120, yy, line); yy -= 23
                c.setFillColor(GRAY); c.setFont(F_BOLD, 14)
                for line in _wrap_pdf_lines(txt_econ, F_BOLD, 14, 1340, 2):
                    c.drawString(120, yy, line); yy -= 20
                for line in _wrap_pdf_lines(txt_dem, F_BOLD, 14, 1340, 2):
                    c.drawString(120, yy, line); yy -= 20
                c.showPage()

    c.save()
    try:
        trace_path = path.parent / PDF_CALC_TRACE_NAME
        with pd.ExcelWriter(trace_path, engine="openpyxl") as writer:
            trace_df = pd.DataFrame(trace_rows) if trace_rows else pd.DataFrame(columns=["block", "level", "period_start", "period_end", "metric", "value", "source", "source_file", "source_sheet", "source_rows", "source_cells", "source_columns", "filters", "formula"])
            trace_df.to_excel(writer, sheet_name="Расчет_ВП", index=False)
            if selected is not None and not selected.empty:
                selected.to_excel(writer, sheet_name="Выбранные_товары", index=False)
            if product_stability is not None and not product_stability.empty:
                product_stability.to_excel(writer, sheet_name="Отбор_товаров", index=False)
        log(f"Saved PDF calc trace: {trace_path} rows={len(trace_rows):,}")
    except Exception as exc:
        log(f"WARN PDF calc trace was not saved: {exc}")
    return path



# The PDF/report logic must not introduce artificial "optimum" comparison rows.
def _append_optimal_factor_rows(level: str, g: pd.DataFrame, keys: List[str], optimal: pd.DataFrame) -> List[Dict[str, Any]]:
    return []


# ================================================================
# PDF v11: three-contour management report
# Контуры:
# 1) текущая неделя — только обзор, без провала в детализацию;
# 2) прошлая полная неделя — категория → товар → артикул;
# 3) последний закрытый месяц — категория → товар → артикул.
# ================================================================

def generate_management_pdf(outputs: Dict[str, pd.DataFrame], path: Path) -> Optional[Path]:
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.colors import HexColor
        from reportlab.pdfbase.pdfmetrics import stringWidth
    except Exception as exc:
        log(f"WARN: reportlab недоступен, PDF не создан: {exc}")
        return None

    F_REG, F_BOLD, F_BLACK = _register_topface_fonts()
    W, H = 1600, 900
    RED = HexColor("#c90022")
    RED_DARK = HexColor("#9d0018")
    WHITE = colors.white
    BLACK = HexColor("#111111")
    GRAY = HexColor("#595959")
    GREEN = HexColor("#087a38")
    BAD = HexColor("#b00020")
    SOFT = HexColor("#fff4f5")
    LINE = HexColor("#e6d8d8")

    # 1) строгий справочник товарных групп + аудит.
    outputs = _filter_outputs_by_pdf_product_reference(outputs, path.parent)

    daily = outputs.get("article_day_fact", pd.DataFrame()).copy()
    if daily is None or daily.empty:
        raise RuntimeError("PDF невозможен: пустой article_day_fact")

    # ---------- basic normalization ----------
    SUBJECT_DISPLAY = {
        "Кисти косметические": "Кисти",
        "Кисти": "Кисти",
        "Косметические карандаши": "Карандаши",
        "Карандаши": "Карандаши",
        "Помады": "Помады",
        "Блески": "Блески",
    }
    DISPLAY_TO_CANON = {
        "Кисти": "Кисти косметические",
        "Карандаши": "Косметические карандаши",
        "Помады": "Помады",
        "Блески": "Блески",
    }
    CATEGORY_ORDER = ["Кисти", "Карандаши", "Помады", "Блески"]
    PRODUCT_ORDER = {
        "Кисти": ["901"],
        "Карандаши": ["605", "611", "613", "614", "617", "618"],
        "Помады": ["154", "155", "156", "157", "206"],
        "Блески": ["207", "209", "210", "211"],
    }
    DETAIL_EXCLUDE = set(os.getenv("PDF_FORCE_EXCLUDE_PRODUCTS", "206,207,209,210,211").split(","))
    DETAIL_EXCLUDE = {x.strip() for x in DETAIL_EXCLUDE if x.strip()}

    def _num(x, default=0.0):
        try:
            if pd.isna(x):
                return default
            return float(x)
        except Exception:
            return default

    def _clean_article_local(x):
        try:
            return clean_article(x)
        except Exception:
            return "" if pd.isna(x) else str(x).strip()

    def _prod(x):
        return _pdf_product_code_from_value(x)

    def _normalize_pdf_merge_keys(df: pd.DataFrame, keys: List[str]) -> pd.DataFrame:
        """Normalize PDF grouping/merge keys so int/string nm_id cannot break late PDF generation.

        The source tables come from different Excel sheets: article_day_fact often keeps nm_id
        as numeric, while ads/search/ABC can return it as text after Excel round-trips. Pandas
        refuses to merge int64 and object keys, so every PDF-side key is normalized before
        groupby/merge.
        """
        if df is None:
            return pd.DataFrame(columns=keys)
        out = df.copy()
        for k in keys:
            if k not in out.columns:
                out[k] = ""
            if k == "nm_id":
                vals = pd.to_numeric(out[k], errors="coerce")
                # WB nm_id is an integer identifier; keep empty for missing values.
                out[k] = vals.round().astype("Int64").astype(str).replace("<NA>", "")
                out[k] = out[k].replace({"nan": "", "None": ""})
            elif k == "supplier_article":
                out[k] = out[k].map(_clean_article_local)
            else:
                out[k] = out[k].map(normalize_text)
        return out

    def _fmt_money(x):
        x = _num(x, 0)
        sign = "-" if x < 0 else ""
        return f"{sign}{int(round(abs(x))):,} ₽".replace(",", " ")

    def _fmt_money_short(x):
        x = _num(x, 0)
        sign = "-" if x < 0 else ""
        ax = abs(x)
        if ax >= 1_000_000:
            return f"{sign}{ax/1_000_000:.1f} млн ₽".replace(".", ",")
        if ax >= 1000:
            return f"{sign}{ax/1000:.0f}к ₽".replace(".", ",")
        return f"{sign}{ax:.0f} ₽".replace(".", ",")

    def _fmt_signed_money(x):
        x = _num(x, 0)
        if abs(x) < 0.5:
            return "0 ₽"
        return ("+" if x > 0 else "-") + f"{int(round(abs(x))):,} ₽".replace(",", " ")

    def _fmt_pct(x, digits=1):
        if x is None or pd.isna(x):
            return "—"
        return f"{_num(x):.{digits}f}%".replace(".", ",")

    def _fmt_num(x):
        if x is None or pd.isna(x):
            return "—"
        return f"{int(round(_num(x))):,}".replace(",", " ")

    def _fmt_rub1(x):
        if x is None or pd.isna(x):
            return "—"
        return f"{_num(x):.1f} ₽".replace(".", ",")

    def _fmt_loc_pair(direct, repl):
        # Compact localization display requested by user: direct coverage -> coverage with replacements.
        # Header stays simply "Локализация"; no explanatory text in PDF.
        d = _num(direct, np.nan)
        r = _num(repl, np.nan)
        if pd.isna(d) and pd.isna(r):
            return "—"
        if pd.isna(d):
            return f"— → {int(round(r))}%"
        if pd.isna(r):
            return f"{int(round(d))}% → —"
        return f"{int(round(d))}% → {int(round(r))}%"

    def _delta_abs(cur, prev):
        cur, prev = _num(cur), _num(prev)
        if abs(prev) < 1e-9 and abs(cur) < 1e-9:
            return 0.0
        return cur - prev

    def _arrow_money(delta, lower_bad=False):
        if delta is None or abs(_num(delta)) < 0.5:
            return "→ 0 ₽"
        d = _num(delta)
        return ("↑ +" if d > 0 else "↓ -") + f"{int(round(abs(d))):,} ₽".replace(",", " ")

    def _delta(cur, prev):
        cur, prev = _num(cur), _num(prev)
        # Управленческий отчет: не показываем мусорные +10000% / -3000%
        # при нулевой или отрицательной базе. В таких случаях в PDF выводится «—».
        if abs(prev) < 1e-9:
            if abs(cur) < 1e-9:
                return 0.0
            return None
        if prev < 0:
            return None
        d = (cur / prev - 1.0) * 100.0
        if abs(d) > 999.0:
            return None
        return d

    LOWER_BAD = {"ДРР", "CPC", "Расход РК", "Комиссия", "Эквайринг", "Логистика", "Хранение", "Себест", "Прочие", "СПП"}
    def _lower_bad(metric: str) -> bool:
        s = str(metric).lower()
        return any(k.lower() in s for k in LOWER_BAD)

    def _arrow(delta, lower_bad=False):
        if delta is None:
            return "→ 0,0%"
        if abs(delta) < 0.05:
            return "→ 0,0%"
        return ("↑ " if delta > 0 else "↓ ") + f"{abs(delta):.1f}%".replace(".", ",")

    def _tone(delta, lower_bad=False):
        if delta is None or abs(delta) < 0.05:
            return GRAY
        good = (delta < 0) if lower_bad else (delta > 0)
        return GREEN if good else BAD

    def _subject_disp(x):
        return SUBJECT_DISPLAY.get(normalize_text(x), normalize_text(x))

    def _safe_mean(s):
        s = pd.to_numeric(s, errors="coerce")
        return float(s.dropna().mean()) if len(s.dropna()) else 0.0

    for col in ["day", "subject", "product", "supplier_article", "nm_id"]:
        if col not in daily.columns:
            daily[col] = ""
    daily["day"] = pd.to_datetime(daily["day"], errors="coerce").dt.normalize()
    daily["subject_disp"] = daily["subject"].map(_subject_disp)
    daily["product_code"] = daily.apply(lambda r: _prod(r.get("product")) or _prod(r.get("supplier_article")), axis=1)
    # Keep approved products. 405/406 are removed by _filter_outputs_by_pdf_product_reference.
    daily = daily[daily["subject_disp"].isin(CATEGORY_ORDER)].copy()
    daily = _normalize_pdf_merge_keys(daily, ["subject_disp", "product_code", "supplier_article", "nm_id"])

    # Уникальный спрос WB для PDF: считаем по уникальным поисковым запросам,
    # а не суммой спроса по артикулам. Если листа нет в старом техфайле,
    # используем старый fallback, но в новом полном запуске лист будет создан.
    search_unique_demand = outputs.get("search_unique_demand", pd.DataFrame()).copy()
    if search_unique_demand is None:
        search_unique_demand = pd.DataFrame()
    if not search_unique_demand.empty:
        if "day" in search_unique_demand.columns:
            search_unique_demand["day"] = pd.to_datetime(search_unique_demand["day"], errors="coerce").dt.normalize()
        if "subject_disp" not in search_unique_demand.columns:
            if "subject" in search_unique_demand.columns:
                search_unique_demand["subject_disp"] = search_unique_demand["subject"].map(_subject_disp)
            else:
                search_unique_demand["subject_disp"] = ""
        if "product_code" not in search_unique_demand.columns:
            if "product" in search_unique_demand.columns:
                search_unique_demand["product_code"] = search_unique_demand["product"].map(_prod)
            else:
                search_unique_demand["product_code"] = ""
        if "supplier_article" in search_unique_demand.columns:
            search_unique_demand["supplier_article"] = search_unique_demand["supplier_article"].map(_clean_article_local)
        for _c in ["unique_search_frequency", "unique_search_queries", "duplicate_query_rows_removed", "raw_query_rows"]:
            if _c not in search_unique_demand.columns:
                search_unique_demand[_c] = 0
            search_unique_demand[_c] = pd.to_numeric(search_unique_demand[_c], errors="coerce").fillna(0)

    # Strict mode: the user explicitly rejected duplicated demand from article_day_fact.
    # Therefore a PDF with fallback demand is not a valid управленческий отчет.
    PDF_ALLOW_DEMAND_FALLBACK = os.getenv("PDF_ALLOW_DEMAND_FALLBACK", "0").strip().lower() in {"1", "true", "yes", "y"}
    # Default is no duplicated fallback, but also no late 30-minute crash: missing unique demand
    # is rendered as blank/diagnostic unless PDF_FAIL_ON_MISSING_DEMAND=1 is explicitly set.
    PDF_FAIL_ON_MISSING_DEMAND = os.getenv("PDF_FAIL_ON_MISSING_DEMAND", "0").strip().lower() in {"1", "true", "yes", "y"}
    if (search_unique_demand is None or search_unique_demand.empty) and not PDF_ALLOW_DEMAND_FALLBACK:
        raise RuntimeError(
            "PDF остановлен: нет листа/данных search_unique_demand. "
            "Спрос WB нельзя считать суммой по артикулам. "
            "Запусти полный пересчет источников или временно поставь PDF_ALLOW_DEMAND_FALLBACK=1 только для диагностики."
        )

    for col in ["order_sum", "orders", "gross_profit_model", "open_cards", "add_to_cart", "search_frequency", "search_traffic_capture_pct", "direct_localization_pct", "localization_with_replacements_pct", "rating_reviews", "finished_price", "price_with_disc", "spp", "commission_%", "acquiring_%", "logistics_direct", "storage", "other_costs", "cost", "cart_conv_pct", "order_conv_pct"]:
        if col not in daily.columns:
            daily[col] = 0.0
        daily[col] = pd.to_numeric(daily[col], errors="coerce").fillna(0)
    if "ad_spend_total" not in daily.columns:
        daily["ad_spend_total"] = 0.0
        for col in ["manual_spend", "unified_spend", "unknown_spend", "ad_spend_model"]:
            if col in daily.columns:
                daily["ad_spend_total"] += pd.to_numeric(daily[col], errors="coerce").fillna(0)
    if "ad_clicks_total" not in daily.columns:
        daily["ad_clicks_total"] = 0.0
        for col in ["manual_clicks", "unified_clicks", "unknown_clicks"]:
            if col in daily.columns:
                daily["ad_clicks_total"] += pd.to_numeric(daily[col], errors="coerce").fillna(0)
    if "ad_impressions_total" not in daily.columns:
        daily["ad_impressions_total"] = 0.0
        for col in ["manual_impressions", "unified_impressions", "unknown_impressions"]:
            if col in daily.columns:
                daily["ad_impressions_total"] += pd.to_numeric(daily[col], errors="coerce").fillna(0)

    # Advertising truth source. For current/incomplete periods do NOT trust repeated article_day_fact spend
    # when the raw ad report is available. It is grouped directly from Отчёты/Реклама/...
    ads_truth = outputs.get("ads_category_source", pd.DataFrame()).copy()
    ads_truth_source_name = "ads_category_source" if ads_truth is not None and not ads_truth.empty else ""
    if ads_truth is None or ads_truth.empty:
        ads_truth = outputs.get("ads_raw_source", pd.DataFrame()).copy()
        ads_truth_source_name = "ads_raw_source" if ads_truth is not None and not ads_truth.empty else ""
    if ads_truth is None or ads_truth.empty:
        ads_truth = outputs.get("ads_daily_source", pd.DataFrame()).copy()
        ads_truth_source_name = "ads_daily_source" if ads_truth is not None and not ads_truth.empty else ""
    if ads_truth is None:
        ads_truth = pd.DataFrame()
    if not ads_truth.empty:
        if "day" in ads_truth.columns:
            ads_truth["day"] = pd.to_datetime(ads_truth["day"], errors="coerce").dt.normalize()
        if "subject_disp" not in ads_truth.columns:
            ads_truth["subject_disp"] = ads_truth.get("subject", "").map(_subject_disp) if "subject" in ads_truth.columns else ""
        if "product_code" not in ads_truth.columns:
            ads_truth["product_code"] = ads_truth.apply(lambda r: _prod(r.get("product", "")) or _prod(r.get("supplier_article", "")), axis=1)
        if "supplier_article" in ads_truth.columns:
            ads_truth["supplier_article"] = ads_truth["supplier_article"].map(_clean_article_local)
        ads_truth = _normalize_pdf_merge_keys(ads_truth, ["subject_disp", "product_code", "supplier_article", "nm_id"])
        for _c in ["spend", "clicks", "impressions"]:
            if _c not in ads_truth.columns:
                ads_truth[_c] = 0.0
            ads_truth[_c] = pd.to_numeric(ads_truth[_c], errors="coerce").fillna(0.0)

    # В техфайле могут быть заготовленные строки будущих дней с нулями.
    # Их нельзя считать фактическими днями, иначе в текущей неделе появляются 0 ₽ ↓100%.
    _activity_cols = [c for c in ["order_sum", "ad_spend_total", "open_cards", "search_frequency"] if c in daily.columns]
    if _activity_cols:
        _activity = pd.Series(0.0, index=daily.index)
        for _c in _activity_cols:
            _activity = _activity + pd.to_numeric(daily[_c], errors="coerce").fillna(0).abs()
        latest = pd.to_datetime(daily.loc[_activity > 0, "day"], errors="coerce").max()
    else:
        latest = pd.to_datetime(daily["day"], errors="coerce").max()
    if pd.isna(latest):
        latest = pd.Timestamp.today().normalize()
    latest = pd.Timestamp(latest).normalize()
    cur_start = latest - pd.Timedelta(days=int(latest.weekday()))
    cur_end = cur_start + pd.Timedelta(days=6)
    cur_actual_end = latest
    prev_start = cur_start - pd.Timedelta(days=7)
    prev_end = cur_start - pd.Timedelta(days=1)
    prev2_start = cur_start - pd.Timedelta(days=14)
    prev2_end = cur_start - pd.Timedelta(days=8)
    closed_end = cur_start.replace(day=1) - pd.Timedelta(days=1)
    closed_start = closed_end.replace(day=1)
    closed_prev_end = closed_start - pd.Timedelta(days=1)
    closed_prev_start = closed_prev_end.replace(day=1)

    # ---------- ABC helpers ----------
    def _prepare_abc(src: pd.DataFrame, start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
        if src is None or src.empty:
            return pd.DataFrame()
        x = src.copy()
        if "period_start" not in x.columns or "period_end" not in x.columns:
            return pd.DataFrame()
        x["period_start"] = pd.to_datetime(x["period_start"], errors="coerce").dt.normalize()
        x["period_end"] = pd.to_datetime(x["period_end"], errors="coerce").dt.normalize()
        x = x[(x["period_start"] == pd.Timestamp(start).normalize()) & (x["period_end"] == pd.Timestamp(end).normalize())].copy()
        if x.empty:
            return x
        if "subject_disp" not in x.columns:
            x["subject_disp"] = x.get("subject", "").map(_subject_disp) if "subject" in x.columns else ""
        if "product_code" not in x.columns:
            x["product_code"] = x.apply(lambda r: _prod(r.get("product", "")) or _prod(r.get("supplier_article", "")), axis=1)
        if "supplier_article" in x.columns:
            x["supplier_article"] = x["supplier_article"].map(_clean_article_local)
        x = _normalize_pdf_merge_keys(x, ["subject_disp", "product_code", "supplier_article", "nm_id"])
        for col in ["gross_profit", "gross_revenue", "orders", "abc_drr_pct", "abc_margin_pct", "abc_commission_amount", "abc_acquiring_amount"]:
            if col not in x.columns:
                x[col] = 0.0
            x[col] = pd.to_numeric(x[col], errors="coerce").fillna(0)
        return x[x["subject_disp"].isin(CATEGORY_ORDER)].copy()

    def _abc_exact(start: pd.Timestamp, end: pd.Timestamp, keys: List[str]) -> pd.DataFrame:
        frames = []
        for nm in ["abc_weekly", "abc_monthly"]:
            src = _prepare_abc(outputs.get(nm, pd.DataFrame()), start, end)
            if not src.empty:
                frames.append(src)
        if not frames:
            return pd.DataFrame(columns=keys + ["gp_abc", "revenue_abc", "orders_abc", "abc_drr_pct", "abc_commission_pct", "abc_acquiring_pct", "abc_rows"])
        x = pd.concat(frames, ignore_index=True)
        for k in keys:
            if k not in x.columns:
                x[k] = ""
        x = _normalize_pdf_merge_keys(x, keys)
        x["_abc_ad"] = np.where(x["gross_revenue"] > 0, x["gross_revenue"] * x["abc_drr_pct"] / 100.0, 0.0)
        # Рентабельность берем из ABC-отчета, если колонка есть. Для агрегата считаем
        # weighted-average по ABC-выручке, а не пересчитываем как ВП/сумма, потому что
        # в ABC rentability может быть отдельной бизнес-метрикой.
        x["_margin_weight"] = np.where(
            x["gross_revenue"] > 0,
            x["gross_revenue"] * pd.to_numeric(x.get("abc_margin_pct", np.nan), errors="coerce") / 100.0,
            np.nan,
        )
        x["_margin_weight"] = pd.to_numeric(x["_margin_weight"], errors="coerce")
        x["_margin_rev"] = np.where(x["_margin_weight"].notna(), x["gross_revenue"], 0.0)
        x["_margin_weight"] = x["_margin_weight"].fillna(0.0)
        x["_comm_abs"] = x["abc_commission_amount"].abs()
        x["_acq_abs"] = x["abc_acquiring_amount"].abs()
        g = x.groupby(keys, dropna=False, as_index=False).agg(
            gp_abc=("gross_profit", "sum"),
            revenue_abc=("gross_revenue", "sum"),
            orders_abc=("orders", "sum"),
            abc_ad_spend=("_abc_ad", "sum"),
            abc_margin_weight=("_margin_weight", "sum"),
            abc_margin_revenue=("_margin_rev", "sum"),
            abc_commission_amount=("_comm_abs", "sum"),
            abc_acquiring_amount=("_acq_abs", "sum"),
            abc_rows=("gross_profit", "size"),
        )
        rev = pd.to_numeric(g["revenue_abc"], errors="coerce").fillna(0)
        g["abc_margin_pct_calc"] = np.where(rev > 0, g["gp_abc"] / rev * 100, np.nan)
        g["abc_margin_pct"] = np.where(
            pd.to_numeric(g.get("abc_margin_revenue", 0), errors="coerce").fillna(0) > 0,
            g["abc_margin_weight"] / g["abc_margin_revenue"] * 100,
            g["abc_margin_pct_calc"],
        )
        g["abc_drr_pct"] = np.where(rev > 0, g["abc_ad_spend"] / rev * 100, np.nan)
        g["abc_commission_pct"] = np.where(rev > 0, g["abc_commission_amount"] / rev * 100, np.nan)
        g["abc_acquiring_pct"] = np.where(rev > 0, g["abc_acquiring_amount"] / rev * 100, np.nan)
        return g

    def _level_for_keys(keys: List[str]) -> str:
        k = list(keys)
        if k == ["subject_disp"]:
            return "category"
        if k == ["subject_disp", "product_code"]:
            return "product"
        if k == ["subject_disp", "product_code", "supplier_article", "nm_id"]:
            return "article"
        return ""

    def _unique_demand_period(start: pd.Timestamp, end: pd.Timestamp, keys: List[str]) -> pd.DataFrame:
        if search_unique_demand is None or search_unique_demand.empty:
            return pd.DataFrame()
        level = _level_for_keys(keys)
        if not level:
            return pd.DataFrame()
        x = search_unique_demand[(search_unique_demand["day"] >= pd.Timestamp(start).normalize()) & (search_unique_demand["day"] <= pd.Timestamp(end).normalize())].copy()
        x = x[x.get("level", "").astype(str).eq(level)].copy() if "level" in x.columns else x
        if x.empty:
            return pd.DataFrame()
        for k in keys:
            if k not in x.columns:
                x[k] = ""
        x = _normalize_pdf_merge_keys(x, keys)
        g = x.groupby(keys, dropna=False, as_index=False).agg(
            demand_unique=("unique_search_frequency", "sum"),
            unique_queries=("unique_search_queries", "sum"),
            duplicate_query_rows_removed=("duplicate_query_rows_removed", "sum"),
            raw_query_rows=("raw_query_rows", "sum"),
        )
        return g

    def _ads_truth_period(start: pd.Timestamp, end: pd.Timestamp, keys: List[str]) -> pd.DataFrame:
        if ads_truth is None or ads_truth.empty:
            return pd.DataFrame()
        x = ads_truth[(ads_truth["day"] >= pd.Timestamp(start).normalize()) & (ads_truth["day"] <= pd.Timestamp(end).normalize())].copy()
        if x.empty:
            return pd.DataFrame()
        # Raw ad report usually has nm_id and subject. For product/article levels we can use enriched fields when present.
        for k in keys:
            if k not in x.columns:
                x[k] = ""
        x = _normalize_pdf_merge_keys(x, keys)
        g = x.groupby(keys, dropna=False, as_index=False).agg(
            ad_spend_truth=("spend", "sum"),
            clicks_truth=("clicks", "sum"),
            impressions_truth=("impressions", "sum"),
            ad_truth_rows=("spend", "size"),
        )
        return g

    def _agg_daily(start: pd.Timestamp, end: pd.Timestamp, keys: List[str]) -> pd.DataFrame:
        x = daily[(daily["day"] >= pd.Timestamp(start).normalize()) & (daily["day"] <= pd.Timestamp(end).normalize())].copy()
        for k in keys:
            if k not in x.columns:
                x[k] = ""
        x = _normalize_pdf_merge_keys(x, keys)
        if x.empty:
            return pd.DataFrame(columns=keys)
        g = x.groupby(keys, dropna=False, as_index=False).agg(
            daily_rows=("order_sum", "size"),
            active_days=("day", "nunique"),
            order_sum=("order_sum", "sum"),
            orders=("orders", "sum"),
            gp_model=("gross_profit_model", "sum"),
            ad_spend=("ad_spend_total", "sum"),
            clicks=("ad_clicks_total", "sum"),
            impressions=("ad_impressions_total", "sum"),
            opens=("open_cards", "sum"),
            carts=("add_to_cart", "sum"),
            demand=("search_frequency", "sum"),
            search_share=("search_traffic_capture_pct", _safe_mean),
            localization_direct=("direct_localization_pct", _safe_mean),
            localization=("localization_with_replacements_pct", _safe_mean),
            rating=("rating_reviews", _safe_mean),
            # buyer_price is average finishedPrice from Orders. sale price is priceWithDisc/order_sum and may later be overwritten by ABC revenue / ABC qty where available.
            buyer_price=("finished_price", _safe_mean),
            price_with_disc_avg=("price_with_disc", _safe_mean),
            spp=("spp", _safe_mean),
            commission_pct_model=("commission_%", _safe_mean),
            acquiring_pct_model=("acquiring_%", _safe_mean),
            logistics_per_unit=("logistics_direct", _safe_mean),
            storage_per_unit=("storage", _safe_mean),
            other_per_unit=("other_costs", _safe_mean),
            cost_per_unit=("cost", _safe_mean),
        )
        at = _ads_truth_period(start, end, keys)
        if at is not None and not at.empty:
            g = _normalize_pdf_merge_keys(g, keys)
            at = _normalize_pdf_merge_keys(at, keys)
            g = g.merge(at, on=keys, how="left")
            mask = pd.to_numeric(g.get("ad_truth_rows"), errors="coerce").fillna(0) > 0
            g["ad_spend_source"] = np.where(mask, ads_truth_source_name or "ads_truth", "article_day_fact")
            g["ad_spend"] = np.where(mask, pd.to_numeric(g.get("ad_spend_truth"), errors="coerce").fillna(0), g["ad_spend"])
            g["clicks"] = np.where(mask, pd.to_numeric(g.get("clicks_truth"), errors="coerce").fillna(0), g["clicks"])
            g["impressions"] = np.where(mask, pd.to_numeric(g.get("impressions_truth"), errors="coerce").fillna(0), g["impressions"])
        else:
            g["ad_spend_source"] = "article_day_fact"
            g["ad_truth_rows"] = 0
        # Store the old duplicated demand for diagnostics before replacing it with unique-query demand.
        g["demand_daily_sum"] = g["demand"]
        # Demand for category/product/article levels must be unique by search query.
        # Fallback: old article_day_fact sum only when the new search_unique_demand sheet is absent.
        du = _unique_demand_period(start, end, keys)
        if du is not None and not du.empty:
            g = _normalize_pdf_merge_keys(g, keys)
            du = _normalize_pdf_merge_keys(du, keys)
            g = g.merge(du, on=keys, how="left")
            g["demand_source"] = np.where(pd.to_numeric(g.get("demand_unique"), errors="coerce").fillna(0) > 0, "unique_queries", "daily_fallback")
            g["demand"] = np.where(pd.to_numeric(g.get("demand_unique"), errors="coerce").fillna(0) > 0, pd.to_numeric(g.get("demand_unique"), errors="coerce").fillna(0), g["demand"])
        else:
            g["demand_source"] = "daily_sum_fallback"
            g["unique_queries"] = 0
            g["duplicate_query_rows_removed"] = 0
            g["raw_query_rows"] = 0
            if not PDF_ALLOW_DEMAND_FALLBACK:
                msg = (
                    f"PDF WARN: для периода {pd.Timestamp(start):%d.%m.%Y}-{pd.Timestamp(end):%d.%m.%Y} "
                    f"и уровня {keys} нет уникального спроса search_unique_demand. "
                    "Fallback SUM(search_frequency) НЕ используется; спрос/% поиска будут пустыми."
                )
                if PDF_FAIL_ON_MISSING_DEMAND:
                    raise RuntimeError(msg)
                log(msg)
                g["demand"] = np.nan
        # % поиска = все открытия карточки / Спрос WB.
        g["search_share"] = np.where(g["demand"] > 0, g["opens"] / g["demand"] * 100, np.nan)
        g["drr_model"] = np.where(g["order_sum"] > 0, g["ad_spend"] / g["order_sum"] * 100, 0.0)
        g["cpc"] = np.where(g["clicks"] > 0, g["ad_spend"] / g["clicks"], np.nan)
        # CTR РК = все клики рекламных кампаний / все показы рекламных кампаний.
        g["ad_ctr"] = np.where(g["impressions"] > 0, g["clicks"] / g["impressions"] * 100, np.nan)
        # CPC is kept as calculated from the available ad spend/click source.
        # Do not hide low CPC values: shelf-type ads can legitimately produce CPC below bid floor.
        g["cart_conv"] = np.where(g["opens"] > 0, g["carts"] / g["opens"] * 100, 0.0)
        g["order_conv"] = np.where(g["carts"] > 0, g["orders"] / g["carts"] * 100, 0.0)
        # Управленческая конверсия в заказ: заказы / все открытия карточки.
        g["order_from_open_conv"] = np.where(g["opens"] > 0, g["orders"] / g["opens"] * 100, 0.0)
        return g

    def _metrics_period(start: pd.Timestamp, end: pd.Timestamp, prev_s: pd.Timestamp, prev_e: pd.Timestamp, keys: List[str]) -> pd.DataFrame:
        cur = _normalize_pdf_merge_keys(_agg_daily(start, end, keys), keys)
        prev = _normalize_pdf_merge_keys(_agg_daily(prev_s, prev_e, keys), keys)
        out = cur.merge(prev, on=keys, how="outer", suffixes=("", "_prev"))
        abc = _normalize_pdf_merge_keys(_abc_exact(start, end, keys), keys)
        abc_prev = _normalize_pdf_merge_keys(_abc_exact(prev_s, prev_e, keys), keys)
        out = _normalize_pdf_merge_keys(out, keys).merge(abc, on=keys, how="left")
        abc_prev = abc_prev.rename(columns={c: c + "_prev_abc" for c in abc_prev.columns if c not in keys})
        out = _normalize_pdf_merge_keys(out, keys).merge(_normalize_pdf_merge_keys(abc_prev, keys), on=keys, how="left")
        # fill numeric values
        for col in ["order_sum", "orders", "gp_model", "ad_spend", "clicks", "impressions", "ad_ctr", "opens", "carts", "demand", "demand_daily_sum", "search_share", "localization_direct", "localization", "rating", "price_sale", "buyer_price", "spp", "commission_pct_model", "acquiring_pct_model", "logistics_per_unit", "storage_per_unit", "other_per_unit", "cost_per_unit", "drr_model", "cpc", "cart_conv", "order_conv", "order_from_open_conv"]:
            if col not in out.columns: out[col] = 0.0
            if col + "_prev" not in out.columns: out[col + "_prev"] = 0.0
            out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0.0)
            out[col + "_prev"] = pd.to_numeric(out[col + "_prev"], errors="coerce").fillna(0.0)
        # ABC is source of truth for closed periods when exact ABC exists.
        out["sum_use"] = np.where(pd.to_numeric(out.get("revenue_abc", 0), errors="coerce").fillna(0) > 0, pd.to_numeric(out.get("revenue_abc", 0), errors="coerce").fillna(0), out["order_sum"])
        out["sum_prev_use"] = np.where(pd.to_numeric(out.get("revenue_abc_prev_abc", 0), errors="coerce").fillna(0) > 0, pd.to_numeric(out.get("revenue_abc_prev_abc", 0), errors="coerce").fillna(0), out["order_sum_prev"])
        out["gp_use"] = np.where(pd.to_numeric(out.get("gp_abc", 0), errors="coerce").fillna(0).abs() > 1e-9, pd.to_numeric(out.get("gp_abc", 0), errors="coerce").fillna(0), out["gp_model"])
        out["gp_prev_use"] = np.where(pd.to_numeric(out.get("gp_abc_prev_abc", 0), errors="coerce").fillna(0).abs() > 1e-9, pd.to_numeric(out.get("gp_abc_prev_abc", 0), errors="coerce").fillna(0), out["gp_model_prev"])
        out["margin"] = np.where(out["sum_use"].abs() > 1e-9, out["gp_use"] / out["sum_use"] * 100, 0.0)
        out["margin_prev"] = np.where(out["sum_prev_use"].abs() > 1e-9, out["gp_prev_use"] / out["sum_prev_use"] * 100, 0.0)
        out["drr"] = np.where(pd.to_numeric(out.get("abc_drr_pct", np.nan), errors="coerce").notna(), pd.to_numeric(out.get("abc_drr_pct", 0), errors="coerce"), out["drr_model"])
        out["drr_prev"] = np.where(pd.to_numeric(out.get("abc_drr_pct_prev_abc", np.nan), errors="coerce").notna(), pd.to_numeric(out.get("abc_drr_pct_prev_abc", 0), errors="coerce"), out["drr_model_prev"])
        out["commission_pct"] = np.where(pd.to_numeric(out.get("abc_commission_pct", np.nan), errors="coerce").notna(), pd.to_numeric(out.get("abc_commission_pct", 0), errors="coerce"), out["commission_pct_model"])
        out["commission_pct_prev"] = np.where(pd.to_numeric(out.get("abc_commission_pct_prev_abc", np.nan), errors="coerce").notna(), pd.to_numeric(out.get("abc_commission_pct_prev_abc", 0), errors="coerce"), out["commission_pct_model_prev"])
        out["acquiring_pct"] = np.where(pd.to_numeric(out.get("abc_acquiring_pct", np.nan), errors="coerce").notna(), pd.to_numeric(out.get("abc_acquiring_pct", 0), errors="coerce"), out["acquiring_pct_model"])
        out["acquiring_pct_prev"] = np.where(pd.to_numeric(out.get("abc_acquiring_pct_prev_abc", np.nan), errors="coerce").notna(), pd.to_numeric(out.get("abc_acquiring_pct_prev_abc", 0), errors="coerce"), out["acquiring_pct_model_prev"])
        out["has_abc"] = pd.to_numeric(out.get("abc_rows", 0), errors="coerce").fillna(0) > 0
        out["has_abc_prev"] = pd.to_numeric(out.get("abc_rows_prev_abc", 0), errors="coerce").fillna(0) > 0

        # ВАЖНО: если есть exact ABC за закрытый период, весь финансовый контур
        # берём из ABC, а не смешиваем ABC-ВП/ДРР с оперативными заказами/рекламой.
        # Иначе получается ложная картина: ДРР из ABC, а расход РК из дневной рекламы.
        def _ncol(name: str, default: float = 0.0) -> pd.Series:
            if name in out.columns:
                return pd.to_numeric(out[name], errors="coerce").fillna(default)
            return pd.Series(default, index=out.index, dtype="float64")
        abc_rev = _ncol("revenue_abc")
        abc_rev_prev = _ncol("revenue_abc_prev_abc")
        abc_gp = _ncol("gp_abc")
        abc_gp_prev = _ncol("gp_abc_prev_abc")
        abc_ad = _ncol("abc_ad_spend")
        abc_ad_prev = _ncol("abc_ad_spend_prev_abc")
        abc_orders = _ncol("orders_abc")
        abc_orders_prev = _ncol("orders_abc_prev_abc")

        out["sum_use"] = np.where(out["has_abc"] & (abc_rev.abs() > 1e-9), abc_rev, out["order_sum"])
        out["sum_prev_use"] = np.where(out["has_abc_prev"] & (abc_rev_prev.abs() > 1e-9), abc_rev_prev, out["order_sum_prev"])
        out["gp_use"] = np.where(out["has_abc"], abc_gp, out["gp_model"])
        out["gp_prev_use"] = np.where(out["has_abc_prev"], abc_gp_prev, out["gp_model_prev"])
        # Заказы в PDF всегда берём из отчёта WB Заказы после исключения отмен,
        # даже когда ВП/выручка для закрытого периода пришли из ABC.
        out["orders"] = pd.to_numeric(out["orders"], errors="coerce").fillna(0.0)
        out["orders_prev"] = pd.to_numeric(out["orders_prev"], errors="coerce").fillna(0.0)
        out["price_sale"] = np.where(pd.to_numeric(out["orders"], errors="coerce").fillna(0).abs() > 1e-9, out["sum_use"] / out["orders"].replace(0, np.nan), out.get("buyer_price", np.nan))
        out["price_sale_prev"] = np.where(pd.to_numeric(out["orders_prev"], errors="coerce").fillna(0).abs() > 1e-9, out["sum_prev_use"] / out["orders_prev"].replace(0, np.nan), out.get("buyer_price_prev", np.nan))

        # Расход РК для ABC-периода считаем только из ABC: Валовая выручка × ДРР ABC.
        out["ad_spend"] = np.where(out["has_abc"], abc_ad, out["ad_spend"])
        out["ad_spend_prev"] = np.where(out["has_abc_prev"], abc_ad_prev, out["ad_spend_prev"])
        out["drr"] = np.where(out["sum_use"].abs() > 1e-9, out["ad_spend"] / out["sum_use"] * 100, 0.0)
        out["drr_prev"] = np.where(out["sum_prev_use"].abs() > 1e-9, out["ad_spend_prev"] / out["sum_prev_use"] * 100, 0.0)
        margin_calc = np.where(out["sum_use"].abs() > 1e-9, out["gp_use"] / out["sum_use"] * 100, 0.0)
        margin_prev_calc = np.where(out["sum_prev_use"].abs() > 1e-9, out["gp_prev_use"] / out["sum_prev_use"] * 100, 0.0)
        abc_margin_col = _ncol("abc_margin_pct", np.nan)
        abc_margin_prev_col = _ncol("abc_margin_pct_prev_abc", np.nan)
        out["margin"] = np.where(out["has_abc"] & abc_margin_col.notna(), abc_margin_col, margin_calc)
        out["margin_prev"] = np.where(out["has_abc_prev"] & abc_margin_prev_col.notna(), abc_margin_prev_col, margin_prev_calc)
        out["cpc"] = np.where(out["clicks"] > 0, out["ad_spend"] / out["clicks"], np.nan)
        out["cpc_prev"] = np.where(out["clicks_prev"] > 0, out["ad_spend_prev"] / out["clicks_prev"], np.nan)
        out["ad_ctr"] = np.where(out["impressions"] > 0, out["clicks"] / out["impressions"] * 100, np.nan)
        out["ad_ctr_prev"] = np.where(out["impressions_prev"] > 0, out["clicks_prev"] / out["impressions_prev"] * 100, np.nan)
        # CPC is kept as calculated. Low CPC is not a data error by itself
        # because shelf-type ad formats may have much lower average CPC.

        # Комиссия/эквайринг — доля ABC-выручки, если exact ABC есть.
        out["commission_pct"] = np.where(out["has_abc"], pd.to_numeric(out.get("abc_commission_pct", np.nan), errors="coerce"), out["commission_pct"])
        out["commission_pct_prev"] = np.where(out["has_abc_prev"], pd.to_numeric(out.get("abc_commission_pct_prev_abc", np.nan), errors="coerce"), out["commission_pct_prev"])
        out["acquiring_pct"] = np.where(out["has_abc"], pd.to_numeric(out.get("abc_acquiring_pct", np.nan), errors="coerce"), out["acquiring_pct"])
        out["acquiring_pct_prev"] = np.where(out["has_abc_prev"], pd.to_numeric(out.get("abc_acquiring_pct_prev_abc", np.nan), errors="coerce"), out["acquiring_pct_prev"])
        for _c in ["commission_pct", "commission_pct_prev", "acquiring_pct", "acquiring_pct_prev"]:
            out[_c] = pd.to_numeric(out[_c], errors="coerce").fillna(0.0)
        return out

    def _current_week_prev_avg_comparison(current_df: pd.DataFrame, keys: List[str]) -> pd.DataFrame:
        """Replace current-week previous-period columns with previous-week daily average × elapsed days.

        Business rule 2026-05-28:
        - лист «Текущая неделя» is an operational partial-week report;
        - if only N days have passed in the current week, compare current fact with
          previous full week's average per day multiplied by N;
        - order quantity and order sum are from WB Orders, ad spend/clicks/impressions are
          from advertising reports;
        - do not compare 2 current days with a full previous week and do not use ABC
          revenue as the previous value for order-sum dynamics.
        """
        if current_df is None or current_df.empty:
            return current_df

        elapsed_days = int((pd.Timestamp(cur_actual_end).normalize() - pd.Timestamp(cur_start).normalize()).days) + 1
        prev_days = int((pd.Timestamp(prev_end).normalize() - pd.Timestamp(prev_start).normalize()).days) + 1
        elapsed_days = max(1, elapsed_days)
        prev_days = max(1, prev_days)
        scale = elapsed_days / prev_days

        prev_full = _metrics_period(prev_start, prev_end, prev2_start, prev2_end, keys)
        if prev_full is None or prev_full.empty:
            return current_df

        prev_full = _normalize_pdf_merge_keys(prev_full.copy(), keys)
        cur = _normalize_pdf_merge_keys(current_df.copy(), keys)

        prev_value_cols = [
            "order_sum", "sum_use", "orders", "gp_model", "gp_use", "ad_spend",
            "clicks", "impressions", "opens", "carts", "demand", "demand_daily_sum",
            "unique_queries", "duplicate_query_rows_removed", "raw_query_rows",
        ]
        prev_avg_cols = [
            "search_share", "localization_direct", "localization", "rating",
            "price_sale", "buyer_price", "price_with_disc_avg", "spp",
            "commission_pct_model", "acquiring_pct_model",
            "logistics_per_unit", "storage_per_unit", "other_per_unit", "cost_per_unit",
        ]

        keep_cols = list(dict.fromkeys(keys + [c for c in prev_value_cols + prev_avg_cols if c in prev_full.columns]))
        prev = prev_full[keep_cols].copy()
        rename_map = {c: f"__prev_avg_{c}" for c in keep_cols if c not in keys}
        prev = prev.rename(columns=rename_map)
        cur = cur.merge(prev, on=keys, how="left")

        def _series(name: str, default: float = 0.0) -> pd.Series:
            if name in cur.columns:
                return pd.to_numeric(cur[name], errors="coerce").fillna(default)
            return pd.Series(default, index=cur.index, dtype="float64")

        # Absolute previous-period values = previous full week / 7 × elapsed current days.
        scale_map = {
            "order_sum": "order_sum_prev",
            "orders": "orders_prev",
            "gp_model": "gp_model_prev",
            "gp_use": "gp_prev_use",
            "ad_spend": "ad_spend_prev",
            "clicks": "clicks_prev",
            "impressions": "impressions_prev",
            "opens": "opens_prev",
            "carts": "carts_prev",
            "demand": "demand_prev",
            "demand_daily_sum": "demand_daily_sum_prev",
            "unique_queries": "unique_queries_prev",
            "duplicate_query_rows_removed": "duplicate_query_rows_removed_prev",
            "raw_query_rows": "raw_query_rows_prev",
        }
        for src_col, dst_col in scale_map.items():
            avg_col = f"__prev_avg_{src_col}"
            if avg_col in cur.columns:
                cur[dst_col] = _series(avg_col) * scale

        # For current-week dynamics, «Сумма заказов» must stay tied to Orders priceWithDisc, not ABC revenue.
        cur["sum_prev_use"] = _series("order_sum_prev")

        # Non-absolute daily averages keep previous-week average levels for tooltip/secondary metrics.
        for src_col in prev_avg_cols:
            avg_col = f"__prev_avg_{src_col}"
            dst_col = f"{src_col}_prev"
            if avg_col in cur.columns:
                cur[dst_col] = _series(avg_col, np.nan)

        # Recalculate ratios from the scaled numerators/denominators.
        cur["drr_prev"] = np.where(_series("sum_prev_use").abs() > 1e-9, _series("ad_spend_prev") / _series("sum_prev_use") * 100, 0.0)
        cur["cpc_prev"] = np.where(_series("clicks_prev") > 0, _series("ad_spend_prev") / _series("clicks_prev"), np.nan)
        cur["ad_ctr_prev"] = np.where(_series("impressions_prev") > 0, _series("clicks_prev") / _series("impressions_prev") * 100, np.nan)
        cur["search_share_prev"] = np.where(_series("demand_prev") > 0, _series("opens_prev") / _series("demand_prev") * 100, np.nan)
        cur["cart_conv_prev"] = np.where(_series("opens_prev") > 0, _series("carts_prev") / _series("opens_prev") * 100, np.nan)
        cur["order_conv_prev"] = np.where(_series("carts_prev") > 0, _series("orders_prev") / _series("carts_prev") * 100, np.nan)
        cur["order_from_open_conv_prev"] = np.where(_series("opens_prev") > 0, _series("orders_prev") / _series("opens_prev") * 100, np.nan)
        cur["price_sale_prev"] = np.where(_series("orders_prev") > 0, _series("sum_prev_use") / _series("orders_prev"), _series("price_sale_prev", np.nan))

        helper_cols = [c for c in cur.columns if c.startswith("__prev_avg_")]
        if helper_cols:
            cur = cur.drop(columns=helper_cols)
        cur["prev_compare_mode"] = f"prev_week_daily_avg_x_{elapsed_days}_days"
        cur["prev_compare_days"] = elapsed_days
        return cur

    def _prev_week_daily_average_totals(keys: List[str]) -> Dict[str, float]:
        """Totals of previous full week divided by number of days, for day-row deltas on sheet 1."""
        prev_full = _agg_daily(prev_start, prev_end, keys)
        prev_days = int((pd.Timestamp(prev_end).normalize() - pd.Timestamp(prev_start).normalize()).days) + 1
        prev_days = max(1, prev_days)
        if prev_full is None or prev_full.empty:
            return {"order_sum": 0.0, "ad_spend": 0.0, "demand": 0.0, "opens": 0.0}
        return {
            "order_sum": _num(prev_full.get("order_sum", pd.Series(dtype="float64")).sum()) / prev_days,
            "ad_spend": _num(prev_full.get("ad_spend", pd.Series(dtype="float64")).sum()) / prev_days,
            "demand": _num(prev_full.get("demand", pd.Series(dtype="float64")).sum()) / prev_days,
            "opens": _num(prev_full.get("opens", pd.Series(dtype="float64")).sum()) / prev_days,
        }

    def _abc_periods_inside(start: pd.Timestamp, end: pd.Timestamp, keys: List[str]) -> pd.DataFrame:
        """ABC gross profit for a period without double counting.

        Business rule 2026-05-28:
        - if a Monday-uploaded current-month ABC file exists (for example 01.05-26.05),
          it is the source of truth for current-month gross profit;
        - do not add weekly ABC rows on top of that current-month ABC file;
        - order quantities and order sums for the operational page still come from Orders,
          not from ABC;
        - ad spend for the operational page comes from advertising reports, not from ABC.
        """
        start_n = pd.Timestamp(start).normalize()
        end_n = pd.Timestamp(end).normalize()
        frames = []
        for nm in ["abc_weekly", "abc_monthly"]:
            src = outputs.get(nm, pd.DataFrame()).copy()
            if src is None or src.empty or "period_start" not in src.columns or "period_end" not in src.columns:
                continue
            src["period_start"] = pd.to_datetime(src["period_start"], errors="coerce").dt.normalize()
            src["period_end"] = pd.to_datetime(src["period_end"], errors="coerce").dt.normalize()
            # Candidate periods fully contained in requested interval.
            src = src[(src["period_start"] >= start_n) & (src["period_end"] <= end_n)].copy()
            if src.empty:
                continue
            if "subject_disp" not in src.columns:
                src["subject_disp"] = src.get("subject", "").map(_subject_disp) if "subject" in src.columns else ""
            if "product_code" not in src.columns:
                src["product_code"] = src.apply(lambda r: _prod(r.get("product", "")) or _prod(r.get("supplier_article", "")), axis=1)
            if "supplier_article" in src.columns:
                src["supplier_article"] = src["supplier_article"].map(_clean_article_local)
            src = src[src["subject_disp"].isin(CATEGORY_ORDER)].copy()
            for c0 in ["gross_profit", "gross_revenue", "orders", "abc_drr_pct"]:
                if c0 not in src.columns:
                    src[c0] = 0.0
                src[c0] = pd.to_numeric(src[c0], errors="coerce").fillna(0.0)
            src["_source_bucket"] = nm
            frames.append(src)
        if not frames:
            return pd.DataFrame(columns=keys + ["gp_abc", "revenue_abc", "abc_ad_spend", "abc_rows", "abc_period_start", "abc_period_end"])
        x = pd.concat(frames, ignore_index=True)
        x = _normalize_pdf_merge_keys(x, keys)
        # Prefer one aggregated ABC report that starts exactly at the requested period start
        # and reaches the furthest period end. This covers the user's Monday current-month ABC upload.
        exact_start = x[x["period_start"].eq(start_n)].copy()
        if not exact_start.empty:
            max_end = exact_start["period_end"].max()
            preferred = exact_start[exact_start["period_end"].eq(max_end)].copy()
            # Use the preferred aggregated report when it covers more than one day or exactly matches the period.
            # This prevents double-counting weekly + MTD ABC files.
            if not preferred.empty and (max_end > start_n or (start_n == end_n)):
                x = preferred
        for k in keys:
            if k not in x.columns:
                x[k] = ""
        x = _normalize_pdf_merge_keys(x, keys)
        x["_abc_ad"] = np.where(x["gross_revenue"] > 0, x["gross_revenue"] * x["abc_drr_pct"] / 100.0, 0.0)
        return x.groupby(keys, dropna=False, as_index=False).agg(
            gp_abc=("gross_profit", "sum"),
            revenue_abc=("gross_revenue", "sum"),
            abc_ad_spend=("_abc_ad", "sum"),
            abc_rows=("gross_profit", "size"),
            abc_period_start=("period_start", "min"),
            abc_period_end=("period_end", "max"),
        )

    cur_cat = _current_week_prev_avg_comparison(
        _metrics_period(cur_start, cur_actual_end, prev_start, prev_start + (cur_actual_end-cur_start), ["subject_disp"]),
        ["subject_disp"],
    )
    prev_cat = _metrics_period(prev_start, prev_end, prev2_start, prev2_end, ["subject_disp"])
    closed_cat = _metrics_period(closed_start, closed_end, closed_prev_start, closed_prev_end, ["subject_disp"])
    current_month_cat = _metrics_period(cur_start.replace(day=1), cur_actual_end, (cur_start.replace(day=1)-pd.offsets.MonthBegin(1)).normalize(), cur_start.replace(day=1)-pd.Timedelta(days=1), ["subject_disp"])

    # Details per contour.
    prev_prod = _metrics_period(prev_start, prev_end, prev2_start, prev2_end, ["subject_disp", "product_code"])
    prev_art = _metrics_period(prev_start, prev_end, prev2_start, prev2_end, ["subject_disp", "product_code", "supplier_article", "nm_id"])
    closed_prod = _metrics_period(closed_start, closed_end, closed_prev_start, closed_prev_end, ["subject_disp", "product_code"])
    closed_art = _metrics_period(closed_start, closed_end, closed_prev_start, closed_prev_end, ["subject_disp", "product_code", "supplier_article", "nm_id"])

    def _filter_detail_products(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty: return pd.DataFrame()
        x = df.copy()
        x["product_code"] = x["product_code"].astype(str)
        x = x[x["product_code"].ne("") & ~x["product_code"].isin(DETAIL_EXCLUDE)].copy()
        x = x[(pd.to_numeric(x["sum_use"], errors="coerce").fillna(0).abs() > 0) | (pd.to_numeric(x["gp_use"], errors="coerce").fillna(0).abs() > 0) | (pd.to_numeric(x.get("sum_prev_use",0), errors="coerce").fillna(0).abs() > 0)].copy()
        # Detailed product pages are created only for products with >=10% of category GP.
        # Rows below threshold remain in the category table, but do not explode into dozens of article pages.
        gp_abs = pd.to_numeric(x["gp_use"], errors="coerce").fillna(0).clip(lower=0)
        x["_cat_gp_total"] = gp_abs.groupby(x["subject_disp"]).transform("sum").replace(0, np.nan)
        x["_gp_share_cat"] = gp_abs / x["_cat_gp_total"] * 100
        keep_detail = x["_gp_share_cat"].fillna(0) >= float(os.getenv("PDF_PRODUCT_DETAIL_MIN_GP_SHARE", "10"))
        # Mandatory important products can still be forced into detail pages by config.
        forced = set((os.getenv("PDF_FORCE_DETAIL_PRODUCTS", "901,614,617,618,154,155,156,157,158") or "").replace(";", ",").split(","))
        forced = {v.strip() for v in forced if v.strip()}
        keep_detail = keep_detail | x["product_code"].isin(forced)
        x = x[keep_detail].copy()
        x["_cat_order"] = x["subject_disp"].map({c:i for i,c in enumerate(CATEGORY_ORDER)}).fillna(99)
        return x.sort_values(["_cat_order", "gp_use", "sum_use"], ascending=[True, False, False])

    prev_prod_detail = _filter_detail_products(prev_prod)
    closed_prod_detail = _filter_detail_products(closed_prod)

    def _select_articles(df: pd.DataFrame, prod_row: pd.Series) -> pd.DataFrame:
        if df is None or df.empty: return pd.DataFrame()
        q = df[(df["subject_disp"].astype(str) == str(prod_row["subject_disp"])) & (df["product_code"].astype(str) == str(prod_row["product_code"]))].copy()
        # Для PDF-детализации больше не режем артикулы по 90% ВП/порогам: пользователь должен видеть все строки товара,
        # где есть текущая или базовая сумма/ВП. Если нужно сжать PDF, это делается отдельным FAST/brief режимом, не в боевом отчете.
        for c0 in ["sum_use", "gp_use", "sum_prev_use", "gp_prev_use"]:
            if c0 not in q.columns:
                q[c0] = 0.0
        keep = (pd.to_numeric(q["sum_use"], errors="coerce").fillna(0).abs() > 0) | (pd.to_numeric(q["gp_use"], errors="coerce").fillna(0).abs() > 0) | (pd.to_numeric(q["sum_prev_use"], errors="coerce").fillna(0).abs() > 0) | (pd.to_numeric(q["gp_prev_use"], errors="coerce").fillna(0).abs() > 0)
        q = q[keep].copy()
        if q.empty: return q
        q["_gp_sort"] = pd.to_numeric(q["gp_use"], errors="coerce").fillna(0)
        q["_sum_sort"] = pd.to_numeric(q["sum_use"], errors="coerce").fillna(0)
        return q.sort_values(["_gp_sort", "_sum_sort", "supplier_article"], ascending=[False, False, True])

    # Build contour dictionaries and planned bookmarks.
    contours = {
        "prev": {
            "label": "Прошлая неделя",
            "period": f"{prev_start:%d.%m}-{prev_end:%d.%m.%Y}",
            "start": prev_start, "end": prev_end, "prev_start": prev2_start, "prev_end": prev2_end,
            "summary_key": "prev_summary", "cat_df": prev_cat, "prod_df": prev_prod_detail, "art_df": prev_art,
            "back_label": "← прошлая неделя",
        },
        "closed": {
            "label": "Закрытый месяц",
            "period": f"{closed_start:%d.%m}-{closed_end:%d.%m.%Y}",
            "start": closed_start, "end": closed_end, "prev_start": closed_prev_start, "prev_end": closed_prev_end,
            "summary_key": "closed_summary", "cat_df": closed_cat, "prod_df": closed_prod_detail, "art_df": closed_art,
            "back_label": "← закр. месяц",
        },
    }

    def _slug(x):
        # Bookmark names must be unique. Cyrillic category names used to collapse to the same "_" slug,
        # so all category links could jump to the last category page. Use a stable hash.
        raw = str(x)
        latin = re.sub(r"[^A-Za-z0-9]+", "_", raw).strip("_")[:30]
        h = hashlib.md5(raw.encode("utf-8")).hexdigest()[:10]
        return f"{latin}_{h}" if latin else h
    def _cat_key(contour, cat): return f"{contour}_cat_{_slug(cat)}"
    def _prod_key(contour, cat, prod): return f"{contour}_prod_{_slug(cat)}_{_slug(prod)}"
    def _art_key(contour, cat, prod, art, n=1): return f"{contour}_art{n}_{_slug(cat)}_{_slug(prod)}_{_slug(art)}"
    def _cat_factor_key(contour, cat): return f"{contour}_cat_factor_{_slug(cat)}"
    def _prod_factor_key(contour, cat, prod): return f"{contour}_prod_factor_{_slug(cat)}_{_slug(prod)}"

    # Entry points are weekly only. Use them only for prev contour.
    entry_bridge = outputs.get("entry_points_bridge", pd.DataFrame()).copy()
    if entry_bridge is None: entry_bridge = pd.DataFrame()
    if not entry_bridge.empty:
        if "subject_disp" not in entry_bridge.columns:
            entry_bridge["subject_disp"] = entry_bridge.get("subject", "").map(_subject_disp) if "subject" in entry_bridge.columns else ""
        if "product_code" not in entry_bridge.columns:
            entry_bridge["product_code"] = entry_bridge.apply(lambda r: _prod(r.get("product", "")) or _prod(r.get("supplier_article", "")), axis=1)
        if "supplier_article" in entry_bridge.columns:
            entry_bridge["supplier_article"] = entry_bridge["supplier_article"].map(_clean_article_local)
        for col in ["transitions", "orders", "cart_conv_pct", "order_conv_pct", "transitions_prev", "orders_prev", "orders_share_pct", "effect_gp_rub"]:
            if col not in entry_bridge.columns:
                entry_bridge[col] = 0
            entry_bridge[col] = pd.to_numeric(entry_bridge[col], errors="coerce").fillna(0)

    # ---------- factor table ----------
    def _factor_rows(row: pd.Series, level: str) -> List[Dict[str, Any]]:
        cur_sum = _num(row.get("sum_use")); prev_sum = _num(row.get("sum_prev_use"))
        cur_gp = _num(row.get("gp_use")); prev_gp = _num(row.get("gp_prev_use"))
        cur_margin = _num(row.get("margin")); prev_margin = _num(row.get("margin_prev"))
        cur_orders = _num(row.get("orders")); prev_orders = _num(row.get("orders_prev"))
        gp_per_order = cur_gp / cur_orders if cur_orders > 0 else (prev_gp / prev_orders if prev_orders > 0 else 0)
        rows = []
        def add(factor, block, cur, prev, fmt, effect, lower=False, comment=""):
            if abs(_num(effect)) < 50 and abs(_num(cur)-_num(prev)) < 1e-9:
                return
            d = _delta(cur, prev)
            rows.append({"Фактор": factor, "Блок": block, "Текущее": fmt(cur), "База": fmt(prev), "Изменение": _arrow(d, lower), "Эффект ВП": effect, "Вывод": comment or ("потеря" if effect < 0 else "прирост" if effect > 0 else "нейтрально")})
        prev_margin_rate = prev_margin/100 if abs(prev_margin) > 1e-9 else (cur_margin/100 if abs(cur_margin)>1e-9 else 0)
        add("Объём / сумма заказов", "Экономика и продажи", cur_sum, prev_sum, _fmt_money, (cur_sum - prev_sum) * prev_margin_rate, False, "эффект изменения выручки при прежней рентабельности")
        add("Рентабельность", "Экономика и продажи", cur_margin, prev_margin, _fmt_pct, cur_sum * (cur_margin - prev_margin) / 100.0, False, "изменение рентабельности в деньгах")
        add("Расход РК", "Экономика и продажи", row.get("ad_spend"), row.get("ad_spend_prev"), _fmt_money, -(_num(row.get("ad_spend")) - _num(row.get("ad_spend_prev"))), True, "изменение рекламных расходов")
        add("CPC", "Экономика и продажи", row.get("cpc"), row.get("cpc_prev"), _fmt_rub1, -(_num(row.get("cpc")) - _num(row.get("cpc_prev"))) * max(_num(row.get("clicks")), 0), True, "изменение стоимости клика")
        add("СПП", "Экономика и продажи", row.get("spp"), row.get("spp_prev"), _fmt_pct, -cur_sum * (_num(row.get("spp_prev")) - _num(row.get("spp"))) / 100.0, True, "изменение скидки покупателя")
        add("Комиссия, %", "Экономика и продажи", row.get("commission_pct"), row.get("commission_pct_prev"), _fmt_pct, -cur_sum * (_num(row.get("commission_pct")) - _num(row.get("commission_pct_prev"))) / 100.0, True, "комиссия как доля выручки")
        add("Эквайринг, %", "Экономика и продажи", row.get("acquiring_pct"), row.get("acquiring_pct_prev"), _fmt_pct, -cur_sum * (_num(row.get("acquiring_pct")) - _num(row.get("acquiring_pct_prev"))) / 100.0, True, "эквайринг как доля выручки")
        add("Логистика/шт", "Экономика и продажи", row.get("logistics_per_unit"), row.get("logistics_per_unit_prev"), _fmt_rub1, -(_num(row.get("logistics_per_unit")) - _num(row.get("logistics_per_unit_prev"))) * max(cur_orders, 0), True, "стоимость логистики на продажу")
        add("Хранение/шт", "Экономика и продажи", row.get("storage_per_unit"), row.get("storage_per_unit_prev"), _fmt_rub1, -(_num(row.get("storage_per_unit")) - _num(row.get("storage_per_unit_prev"))) * max(cur_orders, 0), True, "стоимость хранения на продажу")
        add("Себестоимость/шт", "Экономика и продажи", row.get("cost_per_unit"), row.get("cost_per_unit_prev"), _fmt_rub1, -(_num(row.get("cost_per_unit")) - _num(row.get("cost_per_unit_prev"))) * max(cur_orders, 0), True, "изменение себестоимости")
        add("Прочие/шт", "Экономика и продажи", row.get("other_per_unit"), row.get("other_per_unit_prev"), _fmt_rub1, -(_num(row.get("other_per_unit")) - _num(row.get("other_per_unit_prev"))) * max(cur_orders, 0), True, "прочие расходы")
        demand_eff = 0.0
        if _num(row.get("demand_prev")) > 0:
            demand_eff = (_num(row.get("demand")) / _num(row.get("demand_prev")) - 1.0) * prev_sum * prev_margin_rate
        open_effect = (_num(row.get("opens")) - _num(row.get("opens_prev"))) * (_num(row.get("order_from_open_conv_prev"))/100.0) * gp_per_order
        add("Открытия карточки", "Спрос / точки входа / конверсии", row.get("opens"), row.get("opens_prev"), _fmt_num, open_effect, False, "изменение входящего карточного трафика; эффект может тянуться на следующую неделю из-за лага заказов")
        conv_effect = _num(row.get("opens")) * ((_num(row.get("order_from_open_conv")) - _num(row.get("order_from_open_conv_prev"))) / 100.0) * gp_per_order
        add("Конверсия в заказ", "Спрос / точки входа / конверсии", row.get("order_from_open_conv"), row.get("order_from_open_conv_prev"), _fmt_pct, conv_effect, False, "заказы / все открытия карточки; заменяет раздельные конверсии корзины")
        # Sort: biggest losses first, then biggest gains. Drop insignificant factors: less than 10%
        # of the biggest absolute effect. This keeps the factor page focused on money, not noise.
        if rows:
            max_abs = max(abs(_num(r.get("Эффект ВП"))) for r in rows)
            min_effect = max(max_abs * 0.05, 50.0)
            rows = [r for r in rows if abs(_num(r.get("Эффект ВП"))) >= min_effect]
        losses = sorted([r for r in rows if _num(r["Эффект ВП"]) < 0], key=lambda r: _num(r["Эффект ВП"]))
        gains = sorted([r for r in rows if _num(r["Эффект ВП"]) >= 0], key=lambda r: _num(r["Эффект ВП"]), reverse=True)
        return losses + gains

    # ---------- drawing ----------
    c = canvas.Canvas(str(path), pagesize=(W, H))
    page_num = 0
    def _link(target, rect):
        if target:
            c.linkRect("", str(target), rect, relative=0, thickness=0)

    def _draw_text(txt, x, y, max_w, font=F_REG, size=12, color=BLACK, align="left", min_size=7):
        txt = "" if txt is None else str(txt)
        s = size
        while s > min_size and stringWidth(txt, font, s) > max_w:
            s -= 0.5
        c.setFont(font, s); c.setFillColor(color)
        if align == "right": c.drawRightString(x+max_w, y, txt)
        elif align == "center": c.drawCentredString(x+max_w/2, y, txt)
        else: c.drawString(x, y, txt)

    def _wrap(txt, font, size, max_w, max_lines=5):
        words = str(txt).split()
        lines, cur = [], ""
        for w in words:
            test = (cur + " " + w).strip()
            if stringWidth(test, font, size) <= max_w:
                cur = test
            else:
                if cur: lines.append(cur)
                cur = w
            if len(lines) >= max_lines:
                break
        if cur and len(lines) < max_lines:
            lines.append(cur)
        return lines

    def _start(title, subtitle="", section="", key=None, top_menu=False, back_buttons=None):
        nonlocal page_num
        if page_num > 0:
            c.showPage()
        page_num += 1
        c.setFillColor(RED); c.rect(0,0,W,H,fill=1,stroke=0)
        if key:
            c.bookmarkPage(key)
        c.setFillColor(WHITE); c.setFont(F_REG, 32); c.drawString(70, 835, "topface")
        _draw_text(title, 70, 765, 820, F_BLACK, 46, WHITE)
        if subtitle:
            _draw_text(subtitle, 70, 722, 900, F_BOLD, 19, WHITE)
        if section:
            _draw_text(section, W-330, 720, 260, F_BOLD, 14, WHITE, align="right")
        _draw_text(f"Страница {page_num}", W-210, 36, 140, F_BOLD, 13, WHITE, align="right")
        if top_menu:
            buttons=[("Прошлая", "prev_summary"), ("Закр. месяц", "closed_summary"), ("Тек. месяц", "current_month"), ("Год", "summary")]
            bx = W - 620; by = 798
            for lab, target in buttons:
                bw = 135 if lab != "Закр. месяц" else 160
                c.setFillColor(WHITE); c.roundRect(bx, by, bw, 44, 15, fill=1, stroke=0)
                _draw_text(lab, bx+8, by+16, bw-16, F_BOLD, 12, RED_DARK, align="center")
                _link(target, (bx,by,bx+bw,by+44)); bx += bw + 18
        for b in back_buttons or []:
            bx, by, bw, label, target = b
            c.setFillColor(WHITE); c.roundRect(bx, by, bw, 44, 15, fill=1, stroke=0)
            _draw_text(label, bx+8, by+16, bw-16, F_BOLD, 12, RED_DARK, align="center")
            _link(target, (bx,by,bx+bw,by+44))

    def _metric_card(x, y, w, h, value, label, delta=None, metric="", sub=""):
        c.setFillColor(WHITE); c.roundRect(x,y,w,h,14,fill=1,stroke=0)
        dtext = _arrow(delta, _lower_bad(metric)) if delta is not None else ""
        # Значение крупно по центру, подпись внизу, динамика маленькая справа от подписи.
        _draw_text(str(value), x+10, y+h-50, w-20, F_BLACK, 30, BLACK, align="center")
        label_text = str(label)
        label_size = 12
        label_w_real = min(stringWidth(label_text, F_BOLD, label_size) + 8, w*0.68)
        lx = x + max(10, (w - (label_w_real + (stringWidth(dtext, F_BOLD, 9) + 8 if dtext else 0))) / 2)
        _draw_text(label_text, lx, y+18, label_w_real, F_BOLD, label_size, GRAY, align="left")
        if dtext:
            _draw_text(dtext, lx + label_w_real + 4, y+18, w - (lx-x) - label_w_real - 10, F_BOLD, 9, _tone(delta, _lower_bad(metric)), align="left")
        if sub:
            _draw_text(sub, x+10, y+6, w-20, F_REG, 9, GRAY, align="center")

    def _section_bar(y, text):
        c.setFillColor(RED_DARK); c.roundRect(75, y, W-150, 42, 10, fill=1, stroke=0)
        _draw_text(text, 105, y+14, W-210, F_BLACK, 20, WHITE)

    def _draw_cell_value(x, y, w, value, delta=None, metric="", size=12, align="left"):
        txt = str(value)
        val_size = size + 1
        money_metric = any(k in str(metric) for k in ["Сумма", "ВП", "Расход РК"])
        # Keep dynamic arrow next to the value, not at the far edge of the cell.
        if align == "center":
            _draw_text(txt, x, y, w*0.62 if delta is not None else w, F_BOLD, val_size, BLACK, align="center")
            vx = x + w*0.60
        else:
            _draw_text(txt, x+5, y, w*0.62 if delta is not None else w-10, F_BOLD, val_size, BLACK)
            vx = x + w*0.61
        if delta is not None:
            dtext = _arrow_money(delta, _lower_bad(metric)) if money_metric else _arrow(delta, _lower_bad(metric))
            _draw_text(dtext, vx, y, w*0.38-4, F_BOLD, max(7, size-4), _tone(delta, _lower_bad(metric)), align="left")

    def _draw_table(x, y, w, headers, widths, rows, row_h=36, font_size=11, link_col=None, max_rows=None):
        rows = rows[:max_rows] if max_rows else rows
        # Scale columns to the real table width so the right side is not empty.
        total_width = float(sum(widths)) if widths else w
        if total_width > 0 and abs(total_width - w) > 1:
            widths = [ww * w / total_width for ww in widths]
        h = 42 + row_h*len(rows)
        # Basic PDFQualityGuard: never let a table overlap the header/navigation zone.
        safe_top = 705
        safe_bottom = 62
        if y + h > safe_top:
            y = max(safe_bottom, safe_top - h)
        if y < safe_bottom:
            # If still too tall, reduce row height and font before drawing.
            available = max(120, safe_top - safe_bottom)
            if len(rows) > 0:
                row_h = max(24, int((available - 42) / max(1, len(rows))))
                font_size = max(8, min(font_size, row_h - 18))
                h = 42 + row_h*len(rows)
                y = max(safe_bottom, safe_top - h)
        c.setFillColor(WHITE); c.roundRect(x, y, w, h, 14, fill=1, stroke=0)
        c.setFillColor(RED_DARK); c.roundRect(x, y+h-42, w, 42, 11, fill=1, stroke=0)
        # column separators
        c.setStrokeColor(LINE); c.setLineWidth(0.6)
        xx = x
        for head, ww in zip(headers, widths):
            _draw_text(head, xx+4, y+h-26, ww-8, F_BOLD, max(11, font_size), WHITE, align="center")
            if xx > x + 1:
                c.line(xx, y, xx, y+h)
            xx += ww
        for ri, row in enumerate(rows):
            ry = y+h-42-(ri+1)*row_h
            c.setFillColor(SOFT if ri%2 else WHITE); c.rect(x, ry, w, row_h, fill=1, stroke=0)
            c.setStrokeColor(LINE); c.setLineWidth(0.4); c.line(x, ry, x+w, ry)
            xx=x
            target = row.get("_target") if isinstance(row, dict) else None
            cells = row.get("cells") if isinstance(row, dict) else row
            for ci, cell in enumerate(cells):
                ww = widths[ci]
                if isinstance(cell, tuple):
                    val, delta, metric = cell
                    _draw_cell_value(xx+2, ry+row_h/2-5, ww-4, val, delta, metric, font_size, align="center")
                else:
                    _draw_text(cell, xx+5, ry+row_h/2-5, ww-10, F_BOLD if ci==0 else F_REG, font_size, RED_DARK if ci==0 else BLACK, align="center")
                xx += ww
            if target:
                _link(target, (x, ry, x+w, ry+row_h))
        c.setStrokeColor(LINE); c.setLineWidth(0.8); c.roundRect(x, y, w, h, 14, fill=0, stroke=1)
        return h

    def _period_label(s,e): return f"{s:%d.%m}-{e:%d.%m.%Y}"

    def _summary_category_page(key, title, subtitle, section, df, target_contour=None):
        _start(title, subtitle, section or title, key=key, top_menu=True)
        x = df.copy()
        x["_cat_order"] = x["subject_disp"].map({c:i for i,c in enumerate(CATEGORY_ORDER)}).fillna(99)
        x = x.sort_values("_cat_order")

        def _summary_total_row(rows_df, current_only=False):
            s = float(pd.to_numeric(rows_df.get("sum_use"), errors="coerce").fillna(0).sum()) if not rows_df.empty else 0
            sp = float(pd.to_numeric(rows_df.get("sum_prev_use"), errors="coerce").fillna(0).sum()) if not rows_df.empty else 0
            ad = float(pd.to_numeric(rows_df.get("ad_spend"), errors="coerce").fillna(0).sum()) if not rows_df.empty else 0
            adp = float(pd.to_numeric(rows_df.get("ad_spend_prev"), errors="coerce").fillna(0).sum()) if not rows_df.empty else 0
            clk = float(pd.to_numeric(rows_df.get("clicks"), errors="coerce").fillna(0).sum()) if not rows_df.empty else 0
            clkp = float(pd.to_numeric(rows_df.get("clicks_prev"), errors="coerce").fillna(0).sum()) if not rows_df.empty else 0
            dem = float(pd.to_numeric(rows_df.get("demand"), errors="coerce").fillna(0).sum()) if not rows_df.empty else 0
            demp = float(pd.to_numeric(rows_df.get("demand_prev"), errors="coerce").fillna(0).sum()) if not rows_df.empty else 0
            opens = float(pd.to_numeric(rows_df.get("opens"), errors="coerce").fillna(0).sum()) if not rows_df.empty else 0
            opensp = float(pd.to_numeric(rows_df.get("opens_prev"), errors="coerce").fillna(0).sum()) if not rows_df.empty else 0
            gp = float(pd.to_numeric(rows_df.get("gp_use"), errors="coerce").fillna(0).sum()) if not rows_df.empty else 0
            gpp = float(pd.to_numeric(rows_df.get("gp_prev_use"), errors="coerce").fillna(0).sum()) if not rows_df.empty else 0
            drr = ad/s*100 if s else 0
            drrp = adp/sp*100 if sp else 0
            cpc = ad/clk if clk else 0
            cpcp = adp/clkp if clkp else 0
            ss = opens/dem*100 if dem else np.nan
            ssp = opensp/demp*100 if demp else np.nan
            rent = gp/s*100 if s else 0
            rentp = gpp/sp*100 if sp else 0
            return dict(subject_disp="ИТОГО", sum_use=s, sum_prev_use=sp, gp_use=gp, gp_prev_use=gpp, margin=rent, margin_prev=rentp, ad_spend=ad, ad_spend_prev=adp, drr=drr, drr_prev=drrp, cpc=cpc, cpc_prev=cpcp, demand=dem, demand_prev=demp, search_share=ss, search_share_prev=ssp)

        if key == "cur_categories":
            rows=[]
            for _, r in x.iterrows():
                rows.append({"cells": [
                    str(r["subject_disp"]),
                    (_fmt_money(r.get("sum_use")), _delta_abs(r.get("sum_use"), r.get("sum_prev_use")), "Сумма"),
                    (_fmt_money(r.get("ad_spend")), _delta_abs(r.get("ad_spend"), r.get("ad_spend_prev")), "Расход РК"),
                    (_fmt_pct(r.get("drr")), _delta(r.get("drr"), r.get("drr_prev")), "ДРР"),
                    (_fmt_rub1(r.get("cpc")), _delta(r.get("cpc"), r.get("cpc_prev")), "CPC"),
                    (_fmt_num(r.get("demand")), _delta(r.get("demand"), r.get("demand_prev")), "Спрос"),
                    (_fmt_pct(r.get("search_share")), _delta(r.get("search_share"), r.get("search_share_prev")), "% поиска"),
                ]})
            t = _summary_total_row(x, current_only=True)
            rows.append({"cells": ["ИТОГО", (_fmt_money(t["sum_use"]), _delta_abs(t["sum_use"], t["sum_prev_use"]), "Сумма"), (_fmt_money(t["ad_spend"]), _delta_abs(t["ad_spend"], t["ad_spend_prev"]), "Расход РК"), (_fmt_pct(t["drr"]), _delta(t["drr"], t["drr_prev"]), "ДРР"), "—", "—", "—"]})
            _draw_table(75, 330, W-150, ["Категория", "Сумма", "Расход РК", "ДРР", "CPC", "Спрос WB", "% поиска"], [250,220,220,170,160,240,200], rows, row_h=72, font_size=15)
            return

        rows=[]
        for _, r in x.iterrows():
            cat = r["subject_disp"]
            tgt = _cat_key(target_contour, cat) if target_contour and _has_category_detail(target_contour, cat) else None
            rows.append({"_target": tgt, "cells": [
                cat,
                (_fmt_money(r.get("sum_use")), _delta_abs(r.get("sum_use"), r.get("sum_prev_use")), "Сумма"),
                (_fmt_money(r.get("gp_use")), _delta_abs(r.get("gp_use"), r.get("gp_prev_use")), "ВП"),
                (_fmt_pct(r.get("margin")), _delta(r.get("margin"), r.get("margin_prev")), "Рент."),
                (_fmt_money(r.get("ad_spend")), _delta_abs(r.get("ad_spend"), r.get("ad_spend_prev")), "Расход РК"),
                (_fmt_pct(r.get("drr")), _delta(r.get("drr"), r.get("drr_prev")), "ДРР"),
                (_fmt_rub1(r.get("cpc")), _delta(r.get("cpc"), r.get("cpc_prev")), "CPC"),
                (_fmt_num(r.get("demand")), _delta(r.get("demand"), r.get("demand_prev")), "Спрос"),
                (_fmt_pct(r.get("search_share")), _delta(r.get("search_share"), r.get("search_share_prev")), "% поиска"),
            ]})
        t = _summary_total_row(x)
        rows.append({"cells": ["ИТОГО", (_fmt_money(t["sum_use"]), _delta_abs(t["sum_use"], t["sum_prev_use"]), "Сумма"), (_fmt_money(t["gp_use"]), _delta_abs(t["gp_use"], t["gp_prev_use"]), "ВП"), (_fmt_pct(t["margin"]), _delta(t["margin"], t["margin_prev"]), "Рент."), (_fmt_money(t["ad_spend"]), _delta_abs(t["ad_spend"], t["ad_spend_prev"]), "Расход РК"), (_fmt_pct(t["drr"]), _delta(t["drr"], t["drr_prev"]), "ДРР"), (_fmt_rub1(t["cpc"]), _delta(t["cpc"], t["cpc_prev"]), "CPC"), (_fmt_num(t["demand"]), _delta(t["demand"], t["demand_prev"]), "Спрос"), (_fmt_pct(t["search_share"]), _delta(t["search_share"], t["search_share_prev"]), "% поиска")]})
        _draw_table(75, 315, W-150, ["Категория", "Сумма", "ВП", "Рент.", "Расход РК", "ДРР", "CPC", "Спрос WB", "% поиска"], [190,170,160,125,170,115,100,170,140], rows, row_h=68, font_size=14)
    def _current_week_overview():
        _elapsed_compare_days = int((pd.Timestamp(cur_actual_end).normalize() - pd.Timestamp(cur_start).normalize()).days) + 1
        _start("Текущая неделя", f"{_period_label(cur_start, cur_end)} / факт за {_elapsed_compare_days} дн.; сравнение = средний день прошлой недели × {_elapsed_compare_days}", "Текущая неделя", key="cur_overview", top_menu=True)
        total = cur_cat.copy()
        total_sum = total["sum_use"].sum(); total_prev = total["sum_prev_use"].sum()
        total_ad = total["ad_spend"].sum(); total_ad_prev = total["ad_spend_prev"].sum()
        drr = total_ad/total_sum*100 if total_sum else 0; drr_prev = total_ad_prev/total_prev*100 if total_prev else 0
        demand = total["demand"].sum(); demand_prev = total["demand_prev"].sum()
        opens_total = total["opens"].sum() if "opens" in total.columns else 0
        opens_prev_total = total["opens_prev"].sum() if "opens_prev" in total.columns else 0
        search_share = opens_total / demand * 100 if demand else np.nan
        search_prev = opens_prev_total / demand_prev * 100 if demand_prev else np.nan
        cards = [
            (_fmt_money(total_sum), "Сумма заказов", _delta_abs(total_sum,total_prev), "Сумма", ""),
            (_fmt_money(total_ad), "Расход РК", _delta_abs(total_ad,total_ad_prev), "Расход РК", ""),
            (_fmt_pct(drr), "ДРР", _delta(drr,drr_prev), "ДРР", ""),
            (_fmt_num(demand), "Спрос WB", _delta(demand,demand_prev), "Спрос", ""),
            (_fmt_pct(search_share), "% поискового трафика", _delta(search_share,search_prev), "% поиска", ""),
        ]
        for i, card in enumerate(cards):
            _metric_card(75+i*295, 590, 270, 105, *card)
        # Daily overview by day: deltas compare each factual day with previous full week's average day.
        dates = pd.date_range(cur_start, cur_end)
        rows=[]
        prev_avg = _prev_week_daily_average_totals(["subject_disp"])
        psum = prev_avg.get("order_sum", 0.0)
        pad = prev_avg.get("ad_spend", 0.0)
        pdemand = prev_avg.get("demand", 0.0)
        opens_prev = prev_avg.get("opens", 0.0)
        pss = opens_prev/pdemand*100 if pdemand else np.nan
        pdrr = pad/psum*100 if psum else 0
        for dt in dates:
            cur = _agg_daily(dt, dt, ["subject_disp"])
            osum = cur["order_sum"].sum() if not cur.empty else 0
            ad = cur["ad_spend"].sum() if not cur.empty else 0
            ddemand = cur["demand"].sum() if not cur.empty else 0
            opens = cur["opens"].sum() if not cur.empty else 0
            ss = opens/ddemand*100 if ddemand else np.nan
            d = ad/osum*100 if osum else 0
            # Будущие/пустые дни не показываем как падение на 100%.
            if dt > cur_actual_end or (abs(osum) < 1e-9 and abs(ad) < 1e-9 and abs(ddemand) < 1e-9):
                rows.append({"cells": [dt.strftime("%a %d.%m"), "—", "—", "—", "—", "—"]})
            else:
                rows.append({"cells": [
                    dt.strftime("%a %d.%m"),
                    (_fmt_money(osum), _delta_abs(osum, psum), "Сумма"),
                    (_fmt_money(ad), _delta_abs(ad, pad), "Расход РК"),
                    (_fmt_pct(d), _delta(d, pdrr), "ДРР"),
                    (_fmt_num(ddemand), _delta(ddemand, pdemand), "Спрос"),
                    (_fmt_pct(ss), _delta(ss, pss), "% поиска"),
                ]})
        widths=[180,250,230,180,250,230]
        _draw_table(120, 170, W-240, ["День", "Сумма заказов", "Расход РК", "ДРР", "Спрос WB", "% поиска"], widths, rows, row_h=48, font_size=13)

    def _current_week_categories():
        _summary_category_page("cur_categories", "Текущая неделя: категории", f"{cur_start:%d.%m}-{cur_actual_end:%d.%m.%Y} / оперативный обзор", "Текущая неделя", cur_cat, target_contour=None)

    def _current_month_page():
        # Текущий месяц: ВП берём из ABC текущего месяца/вышедших ABC-отчетов,
        # заказы и сумму заказов — из Orders, рекламные расходы — из отчетов по рекламе.
        month_start = cur_start.replace(day=1)
        _start("Текущий месяц", f"{month_start:%d.%m}-{cur_actual_end:%d.%m.%Y} / план = факт ВП прошлого месяца", "Текущий месяц", key="current_month", top_menu=True)
        month_days = calendar.monthrange(month_start.year, month_start.month)[1]
        prev_month_df = _metrics_period(closed_start, closed_end, closed_prev_start, closed_prev_end, ["subject_disp"])
        # План на май/текущий месяц = факт валовой прибыли за апрель/прошлый закрытый месяц. Без +10%.
        month_plan = max(0.0, _num(prev_month_df["gp_use"].sum()) if prev_month_df is not None and not prev_month_df.empty else 0.0)
        elapsed = min((cur_actual_end - month_start).days + 1, month_days)
        plan_to_date = month_plan / month_days * elapsed if month_days else 0.0

        # Operational facts from Orders and Advertising reports only.
        mtd_orders = _agg_daily(month_start, cur_actual_end, ["subject_disp"])
        mtd_orders_qty = _num(mtd_orders["orders"].sum()) if mtd_orders is not None and not mtd_orders.empty else 0.0
        mtd_order_sum = _num(mtd_orders["order_sum"].sum()) if mtd_orders is not None and not mtd_orders.empty else 0.0
        mtd_ad = _num(mtd_orders["ad_spend"].sum()) if mtd_orders is not None and not mtd_orders.empty else 0.0

        # Gross profit from ABC only. If a current-month ABC file 01.MM..latest exists,
        # _abc_periods_inside uses it and does not add weekly pieces on top.
        mtd_abc = _abc_periods_inside(month_start, cur_actual_end, ["subject_disp"])
        mtd_gp = _num(mtd_abc["gp_abc"].sum()) if mtd_abc is not None and not mtd_abc.empty else 0.0
        pct_plan = mtd_gp / plan_to_date * 100 if plan_to_date else np.nan
        cards = [
            (_fmt_money(mtd_gp), "ВП факт ABC", None, "ВП", ""),
            (_fmt_pct(pct_plan), "% плана на дату", None, "% плана", ""),
            (_fmt_num(mtd_orders_qty), "Заказы, шт", None, "Заказы", ""),
            (_fmt_money(mtd_order_sum), "Сумма заказов", None, "Сумма", ""),
            (_fmt_money(mtd_ad), "Расход РК", None, "Расход РК", ""),
        ]
        for i, card in enumerate(cards):
            _metric_card(65+i*300, 605, 280, 110, *card)
        rows=[]
        ws = month_start
        prev_week_gp = None
        while ws <= cur_actual_end:
            we = min(ws + pd.Timedelta(days=6-int(ws.weekday())), cur_actual_end)
            wk_orders = _agg_daily(ws, we, ["subject_disp"])
            wk_abc = _abc_periods_inside(ws, we, ["subject_disp"])
            wk_gp_val = _num(wk_abc["gp_abc"].sum()) if wk_abc is not None and not wk_abc.empty else np.nan
            wk_orders_qty = _num(wk_orders["orders"].sum()) if wk_orders is not None and not wk_orders.empty else 0.0
            wk_sum = _num(wk_orders["order_sum"].sum()) if wk_orders is not None and not wk_orders.empty else 0.0
            wk_ad = _num(wk_orders["ad_spend"].sum()) if wk_orders is not None and not wk_orders.empty else 0.0
            elapsed_to_week = min((we - month_start).days + 1, month_days)
            week_plan_to_date = month_plan / month_days * elapsed_to_week if month_days else 0.0
            mtd_to_week_abc = _abc_periods_inside(month_start, we, ["subject_disp"])
            gp_to_week = _num(mtd_to_week_abc["gp_abc"].sum()) if mtd_to_week_abc is not None and not mtd_to_week_abc.empty else 0.0
            pct_to_week = gp_to_week / week_plan_to_date * 100 if week_plan_to_date else np.nan
            if pd.isna(wk_gp_val):
                gp_cell = "—"
            else:
                gp_cell = (_fmt_money(wk_gp_val), _delta_abs(wk_gp_val, prev_week_gp), "ВП")
                prev_week_gp = wk_gp_val
            rows.append({"cells":[f"{ws:%d.%m}-{we:%d.%m}", gp_cell, _fmt_pct(pct_to_week), _fmt_num(wk_orders_qty), _fmt_money(wk_sum), _fmt_money(wk_ad)]})
            ws = we + pd.Timedelta(days=1)
        _draw_table(85, 230, W-170, ["Неделя", "ВП ABC", "% плана", "Заказы", "Сумма заказов", "Расход РК"], [200,260,240,180,280,250], rows, row_h=58, font_size=16, max_rows=8)

    def _summary_page():
        _start("Помесячная динамика", f"{closed_start.year} год / ABC по 4 категориям", "Годовая динамика", key="summary", top_menu=True)
        src = outputs.get("abc_monthly", pd.DataFrame()).copy()
        rows=[]
        if src is not None and not src.empty and "period_start" in src.columns:
            src["period_start"] = pd.to_datetime(src["period_start"], errors="coerce").dt.normalize()
            src = src[src["period_start"].dt.year.eq(int(closed_start.year))].copy()
            src["subject_disp"] = src.get("subject", "").map(_subject_disp) if "subject" in src.columns else ""
            src = src[src["subject_disp"].isin(CATEGORY_ORDER)].copy()
            for col in ["gross_revenue", "gross_profit", "abc_drr_pct"]:
                if col not in src.columns:
                    src[col] = 0
                src[col] = pd.to_numeric(src[col], errors="coerce").fillna(0)
            src["_ad"] = src["gross_revenue"] * src["abc_drr_pct"] / 100.0
            mon = src.groupby(["period_start", "subject_disp"], as_index=False).agg(sum_use=("gross_revenue","sum"), gp_use=("gross_profit","sum"), ad_spend=("_ad","sum"))
            mon["_cat_order"] = mon["subject_disp"].map({c:i for i,c in enumerate(CATEGORY_ORDER)}).fillna(99)
            mon = mon.sort_values(["subject_disp", "period_start"])
            mon["sum_prev"] = mon.groupby("subject_disp")["sum_use"].shift(1)
            mon["gp_prev"] = mon.groupby("subject_disp")["gp_use"].shift(1)
            mon["ad_prev"] = mon.groupby("subject_disp")["ad_spend"].shift(1)
            mon = mon.sort_values(["period_start", "_cat_order"], ascending=[False, True])
            for _, r in mon.iterrows():
                month_name = MONTH_RU.get(int(r["period_start"].month), str(r["period_start"].month)).lower()
                rows.append({"cells":[
                    month_name,
                    str(r.get("subject_disp")),
                    (_fmt_money(r.get("sum_use")), _delta_abs(r.get("sum_use"), r.get("sum_prev")), "Сумма"),
                    (_fmt_money(r.get("gp_use")), _delta_abs(r.get("gp_use"), r.get("gp_prev")), "ВП"),
                    (_fmt_money(r.get("ad_spend")), _delta_abs(r.get("ad_spend"), r.get("ad_prev")), "Расход РК"),
                ]})
        if not rows:
            rows=[{"cells":["—","—","—","—","—"]}]
        _draw_table(95, 115, W-190, ["Месяц", "Категория", "Сумма заказов", "ВП", "Расход РК"], [180,250,330,330,330], rows, row_h=42, font_size=13, max_rows=16)

    def _children_for_category(contour: str, cat: str) -> pd.DataFrame:
        info = contours[contour]
        prod = info["prod_df"]
        if prod is None or prod.empty: return pd.DataFrame()
        q = prod[prod["subject_disp"].astype(str).eq(cat)].copy()
        return q.sort_values(["gp_use", "sum_use"], ascending=[False, False])

    def _articles_for_product(contour: str, cat: str, prod_code: str) -> pd.DataFrame:
        info = contours[contour]
        prod_df = info["prod_df"]
        prod_row = prod_df[(prod_df["subject_disp"].astype(str).eq(cat)) & (prod_df["product_code"].astype(str).eq(str(prod_code)))]
        if prod_row.empty: return pd.DataFrame()
        return _select_articles(info["art_df"], prod_row.iloc[0])

    def _has_category_detail(contour: str, cat: str) -> bool:
        if cat == "Кисти":
            return not _articles_for_product(contour, cat, "901").empty
        return not _children_for_category(contour, cat).empty

    def _draw_factor_table_page(key, title, subtitle, section, row, back_buttons):
        _start(title, subtitle, section, key=key, top_menu=False, back_buttons=back_buttons)
        factors = _factor_rows(row, section)
        # Верхние итоги показывают именно факт изменения ВП из ABC: было/стало.
        # Факторные строки ниже объясняют, но не суммируются как независимые причины,
        # потому что открытия, объем и конверсии взаимозависимы.
        actual_delta_gp = _num(row.get("gp_use")) - _num(row.get("gp_prev_use"))
        loss_total = min(actual_delta_gp, 0.0)
        gain_total = max(actual_delta_gp, 0.0)
        net_total = actual_delta_gp
        _metric_card(90, 620, 310, 90, _fmt_signed_money(loss_total), "Потери ВП ABC", None, "", "")
        _metric_card(430, 620, 310, 90, _fmt_signed_money(gain_total), "Прирост ВП ABC", None, "", "")
        _metric_card(770, 620, 310, 90, _fmt_signed_money(net_total), "Итог ABC", None, "", "")
        _draw_text("Факторы ниже не являются полностью независимыми: открытия → заказы → сумма заказов связаны. Открытия оставлены как ранний сигнал на следующую неделю.", 75, 590, W-150, F_REG, 12, WHITE)
        rows=[]
        for fr in factors:
            rows.append({"cells": [fr["Фактор"], fr["Блок"], fr["Текущее"], fr["База"], fr["Изменение"], _fmt_signed_money(fr["Эффект ВП"]), fr["Вывод"]]})
        if not rows:
            rows=[{"cells":["—", "—", "—", "—", "—", "0 ₽", "значимых денежных факторов нет"]}]
        _draw_table(75, 70, W-150, ["Фактор", "Блок", "Текущее", "Прошлая неделя", "Изм.", "Эффект ВП", "Вывод"], [235,250,140,165,110,145,455], rows, row_h=31, font_size=10, max_rows=16)

    def _draw_level_overview(title, subtitle, section, key, row, back_buttons, next_button=None):
        _start(title, subtitle, section, key=key, top_menu=False, back_buttons=back_buttons)
        if next_button:
            label, target = next_button
            bx, by, bw = 1380, 798, 120
            c.setFillColor(WHITE); c.roundRect(bx, by, bw, 44, 15, fill=1, stroke=0)
            _draw_text(label, bx+8, by+16, bw-16, F_BOLD, 12, RED_DARK, align="center")
            _link(target, (bx,by,bx+bw,by+44))
        _section_bar(640, "Блок 1. Экономика и продажи")
        cards1 = [
            (_fmt_money(row.get("sum_use")), "Сумма", _delta(row.get("sum_use"), row.get("sum_prev_use")), "Сумма", ""),
            (_fmt_num(row.get("orders")), "Заказы", _delta(row.get("orders"), row.get("orders_prev")), "Заказы", ""),
            (_fmt_money(row.get("gp_use")), "ВП ABC", _delta(row.get("gp_use"), row.get("gp_prev_use")), "ВП", ""),
            (_fmt_pct(row.get("margin")), "Рентабельность", _delta(row.get("margin"), row.get("margin_prev")), "Рент.", ""),
            (_fmt_pct(row.get("drr")), "ДРР", _delta(row.get("drr"), row.get("drr_prev")), "ДРР", ""),
            (_fmt_money(row.get("ad_spend")), "Расход РК", _delta(row.get("ad_spend"), row.get("ad_spend_prev")), "Расход РК", ""),
            (_fmt_rub1(row.get("cpc")), "CPC", _delta(row.get("cpc"), row.get("cpc_prev")), "CPC", ""),
        ]
        for i, card in enumerate(cards1):
            _metric_card(65+i*215, 520, 200, 95, *card)
        _section_bar(450, "Блок 2. Реклама, спрос и конверсии")
        cards2 = [
            (_fmt_num(row.get("demand")), "Спрос WB", _delta(row.get("demand"), row.get("demand_prev")), "Спрос", ""),
            (_fmt_pct(row.get("search_share")), "% поиска", _delta(row.get("search_share"), row.get("search_share_prev")), "% поиска", ""),
            (_fmt_pct(row.get("ad_ctr")), "CTR РК", _delta(row.get("ad_ctr"), row.get("ad_ctr_prev")), "CTR РК", ""),
            (_fmt_num(row.get("opens")), "Открытия", _delta(row.get("opens"), row.get("opens_prev")), "Открытия", ""),
            (_fmt_num(row.get("orders")), "Заказы", _delta(row.get("orders"), row.get("orders_prev")), "Заказы", ""),
            (_fmt_pct(row.get("order_from_open_conv")), "Конв. в заказ", _delta(row.get("order_from_open_conv"), row.get("order_from_open_conv_prev")), "Конверсия", ""),
            (_fmt_pct(row.get("order_conv")), "Корзина→заказ", _delta(row.get("order_conv"), row.get("order_conv_prev")), "Конверсия", ""),
        ]
        for i, card in enumerate(cards2):
            _metric_card(65+i*215, 330, 200, 95, *card)
        _section_bar(260, "Блок 3. Расходы на единицу / доля расходов")
        cards3 = [
            (_fmt_pct(row.get("commission_pct")), "Комиссия, %", _delta(row.get("commission_pct"), row.get("commission_pct_prev")), "Комиссия", ""),
            (_fmt_pct(row.get("acquiring_pct")), "Эквайринг, %", _delta(row.get("acquiring_pct"), row.get("acquiring_pct_prev")), "Эквайринг", ""),
            (_fmt_rub1(row.get("logistics_per_unit")), "Логистика/шт", _delta(row.get("logistics_per_unit"), row.get("logistics_per_unit_prev")), "Логистика", ""),
            (_fmt_rub1(row.get("cost_per_unit")), "Себест./шт", _delta(row.get("cost_per_unit"), row.get("cost_per_unit_prev")), "Себест", ""),
            (_fmt_rub1(row.get("price_sale")), "Цена продажи", _delta(row.get("price_sale"), row.get("price_sale_prev")), "Цена", ""),
            (_fmt_rub1(row.get("buyer_price")), "Цена покупателя", _delta(row.get("buyer_price"), row.get("buyer_price_prev")), "Цена", ""),
            (_fmt_loc_pair(row.get("localization_direct"), row.get("localization")), "Локализация", _delta(row.get("localization"), row.get("localization_prev")), "Локализация", ""),
        ]
        for i, card in enumerate(cards3):
            _metric_card(65+i*215, 140, 200, 95, *card)

    def _draw_listing_pages(title, subtitle, section, base_key, back_buttons, headers, widths, rows, row_h=50, font_size=14, rows_per_page=12):
        chunks = [rows[i:i+rows_per_page] for i in range(0, len(rows), rows_per_page)] or [[]]
        for idx, chunk in enumerate(chunks):
            key = base_key if idx == 0 else f"{base_key}_p{idx+1}"
            sec = section if len(chunks) == 1 else f"{section} {idx+1}/{len(chunks)}"
            bb = list(back_buttons or [])
            if idx > 0:
                bb.append((1430, 798, 90, "стр.1", base_key))
            _start(title, subtitle, sec, key=key, top_menu=False, back_buttons=bb)
            table_h = 42 + row_h*len(chunk)
            y = max(90, (H - table_h) / 2 - 20)
            _draw_table(75, y, W-150, headers, widths, chunk, row_h=row_h, font_size=font_size, max_rows=rows_per_page)

    def _draw_category_detail(contour: str, cat: str):
        info = contours[contour]
        row_df = info["cat_df"][info["cat_df"]["subject_disp"].astype(str).eq(cat)]
        if row_df.empty: return
        cat_key = _cat_key(contour, cat)
        cat_factor = _cat_factor_key(contour, cat)
        _draw_level_overview(f"Категория: {cat}", f"{info['period']} / управленческий разбор", "Категория 1/2", cat_key, row_df.iloc[0], [(1220, 798, 210, info["back_label"], info["summary_key"])], ("факторы", cat_factor))
        # children list below metric cards, separate wide table replacing lower area if needed
        # Add second content page with products/articles and factor table button.
        rows=[]
        if cat == "Кисти":
            arts = _articles_for_product(contour, cat, "901")
            for _, r in arts.iterrows():
                art = _clean_article_local(r.get("supplier_article"))
                rows.append({"_target": _art_key(contour, cat, "901", art, 1), "cells": [
                    art,
                    (_fmt_money(r.get("sum_use")), _delta_abs(r.get("sum_use"), r.get("sum_prev_use")), "Сумма"),
                    (_fmt_money(r.get("gp_use")), _delta_abs(r.get("gp_use"), r.get("gp_prev_use")), "ВП"),
                    (_fmt_money(r.get("ad_spend")), _delta_abs(r.get("ad_spend"), r.get("ad_spend_prev")), "Расход РК"),
                    (_fmt_pct(r.get("margin")), _delta(r.get("margin"), r.get("margin_prev")), "Рент."),
                    (_fmt_pct(r.get("drr")), _delta(r.get("drr"), r.get("drr_prev")), "ДРР"),
                ]})
            headers=["Артикул", "Сумма", "ВП", "Расход РК", "Рент.", "ДРР"]
            widths=[240,250,250,250,210,210]
        else:
            prods = _children_for_category(contour, cat)
            for _, r in prods.iterrows():
                prod = str(r.get("product_code"))
                rows.append({"_target": _prod_key(contour, cat, prod), "cells": [
                    prod,
                    (_fmt_money(r.get("sum_use")), _delta_abs(r.get("sum_use"), r.get("sum_prev_use")), "Сумма"),
                    (_fmt_money(r.get("gp_use")), _delta_abs(r.get("gp_use"), r.get("gp_prev_use")), "ВП"),
                    (_fmt_money(r.get("ad_spend")), _delta_abs(r.get("ad_spend"), r.get("ad_spend_prev")), "Расход РК"),
                    (_fmt_pct(r.get("margin")), _delta(r.get("margin"), r.get("margin_prev")), "Рент."),
                    (_fmt_pct(r.get("drr")), _delta(r.get("drr"), r.get("drr_prev")), "ДРР"),
                ]})
            headers=["Товар", "Сумма", "ВП", "Расход РК", "Рент.", "ДРР"]
            widths=[240,250,250,250,210,210]
        _draw_listing_pages(f"Категория: {cat}", f"{info['period']} / переход на следующий уровень", "Категория 2/2", cat_key+"_list", [(1190,798,220,info["back_label"],info["summary_key"]),(1430,798,90,"стр.1",cat_key)], headers, widths, rows, row_h=52, font_size=14, rows_per_page=12)
        _draw_factor_table_page(cat_factor, f"Категория: {cat}", f"{info['period']} / факторная таблица", "Категория факторы", row_df.iloc[0], [(1190,798,220,"← категория",cat_key+"_list"),(1430,798,90,"стр.1",cat_key)])

    def _draw_product_detail(contour: str, prod_row: pd.Series):
        info = contours[contour]
        cat = str(prod_row["subject_disp"]); prod = str(prod_row["product_code"])
        if cat == "Кисти":
            return
        pk = _prod_key(contour, cat, prod)
        pf = _prod_factor_key(contour, cat, prod)
        cat_list_key = _cat_key(contour, cat)+"_list"
        _draw_level_overview(f"Товар: {prod}", f"{cat} / {info['period']}", "Товар 1/2", pk, prod_row, [(1160,798,220,"← категория",cat_list_key)], ("факторы", pf))
        arts = _articles_for_product(contour, cat, prod)
        rows=[]
        for _, r in arts.iterrows():
            art = _clean_article_local(r.get("supplier_article"))
            rows.append({"_target": _art_key(contour, cat, prod, art, 1), "cells": [
                art,
                (_fmt_money(r.get("sum_use")), _delta_abs(r.get("sum_use"), r.get("sum_prev_use")), "Сумма"),
                (_fmt_money(r.get("gp_use")), _delta_abs(r.get("gp_use"), r.get("gp_prev_use")), "ВП"),
                (_fmt_money(r.get("ad_spend")), _delta_abs(r.get("ad_spend"), r.get("ad_spend_prev")), "Расход РК"),
                (_fmt_pct(r.get("margin")), _delta(r.get("margin"), r.get("margin_prev")), "Рент."),
                (_fmt_pct(r.get("drr")), _delta(r.get("drr"), r.get("drr_prev")), "ДРР"),
            ]})
        _draw_listing_pages(f"Товар: {prod}", f"{cat} / {info['period']} / артикулы", "Товар 2/2", pk+"_list", [(1160,798,220,"← категория",cat_list_key),(1400,798,100,"стр.1",pk)], ["Артикул", "Сумма", "ВП", "Расход РК", "Рент.", "ДРР"], [240,250,250,250,210,210], rows, row_h=52, font_size=14, rows_per_page=12)
        _draw_entity_entry_factor_page("product", pf, f"Товар: {prod}", f"{cat} / {info['period']} / точки входа и факторы", "Товар факторы", prod_row, [(1160,798,220,"← товар",pk+"_list"),(1400,798,100,"стр.1",pk)])

    def _entity_entry_rows(level: str, contour: str, row: pd.Series, max_items: int = 8) -> List[Dict[str, Any]]:
        if contour != "prev" or entry_bridge.empty:
            return []
        cat = str(row.get("subject_disp")); prod = str(row.get("product_code")); art = _clean_article_local(row.get("supplier_article"))
        q = entry_bridge[(entry_bridge["subject_disp"].astype(str).eq(cat)) & (entry_bridge["product_code"].astype(str).eq(prod))].copy()
        if level == "article":
            q = q[q["supplier_article"].astype(str).eq(art)].copy()
        if q.empty:
            return []
        group = ["entry_section", "entry_point"]
        q = q.groupby(group, dropna=False, as_index=False).agg(
            transitions=("transitions", "sum"), transitions_prev=("transitions_prev", "sum"),
            orders=("orders", "sum"), orders_prev=("orders_prev", "sum"),
            effect_gp_rub=("effect_gp_rub", "sum"),
        )
        total_orders = pd.to_numeric(q["orders"], errors="coerce").fillna(0).sum()
        q["orders_share_pct"] = np.where(total_orders > 0, pd.to_numeric(q["orders"], errors="coerce").fillna(0) / total_orders * 100, np.nan)
        q = q.sort_values(["orders", "transitions"], ascending=False).head(max_items)
        out=[]
        for _, er in q.iterrows():
            out.append({"cells": [
                f"{er.get('entry_section','')} / {er.get('entry_point','')}",
                (_fmt_num(er.get("transitions")), _delta(er.get("transitions"), er.get("transitions_prev")), "Переходы"),
                (_fmt_num(er.get("orders")), _delta(er.get("orders"), er.get("orders_prev")), "Заказы"),
                _fmt_pct(er.get("orders_share_pct")),
                _fmt_signed_money(er.get("effect_gp_rub")),
            ]})
        return out

    def _draw_entity_entry_factor_page(level: str, key: str, title: str, subtitle: str, section: str, row: pd.Series, back_buttons):
        _start(title, subtitle, section, key=key, top_menu=False, back_buttons=back_buttons)
        ep_rows = _entity_entry_rows(level, "prev", row, max_items=8)
        if not ep_rows:
            ep_rows=[{"cells":["Нет данных по точкам входа для этого периода", "—", "—", "—", "0 ₽"]}]
        _draw_table(75, 470, W-150, ["Канал / точка входа", "Переходы", "Заказы", "Доля заказов", "Вклад ВП"], [600,210,170,190,190], ep_rows, row_h=38, font_size=12, max_rows=8)
        factors = _factor_rows(row, section)
        factor_rows=[]
        for fr in factors[:9]:
            factor_rows.append({"cells": [fr["Фактор"], fr["Блок"], fr["Текущее"], fr["База"], fr["Изменение"], _fmt_signed_money(fr["Эффект ВП"])]})
        if not factor_rows:
            factor_rows=[{"cells":["—", "—", "—", "—", "→ 0,0%", "0 ₽"]}]
        _draw_table(75, 80, W-150, ["Фактор", "Блок", "Текущее", "Прошлая неделя", "Изм.", "Эффект ВП"], [280,390,180,180,150,180], factor_rows, row_h=32, font_size=11, max_rows=9)

    def _draw_article_pages(contour: str, art_row: pd.Series):
        info = contours[contour]
        cat = str(art_row["subject_disp"]); prod = str(art_row["product_code"]); art = _clean_article_local(art_row.get("supplier_article"))
        a1 = _art_key(contour, cat, prod, art, 1); a2 = _art_key(contour, cat, prod, art, 2)
        cat_list_key = _cat_key(contour, cat)+"_list"
        product_back = cat_list_key if cat == "Кисти" else _prod_key(contour, cat, prod)+"_list"
        if cat == "Кисти":
            back_buttons = [(1230,798,210,"← категория", product_back), (1480,798,80,"стр.2",a2)]
        else:
            back_buttons = [(1120,798,170,"← товар", product_back), (1310,798,150,"← категория",cat_list_key), (1480,798,80,"стр.2",a2)]
        _draw_level_overview(f"Артикул: {art}", f"{cat} / товар {prod} / {info['period']}", "Артикул 1/2", a1, art_row, back_buttons=None)
        # overwrite top buttons on the page with correct buttons (because _draw_level_overview already started page)
        for bx,by,bw,label,target in back_buttons:
            c.setFillColor(WHITE); c.roundRect(bx, by, bw, 44, 15, fill=1, stroke=0)
            _draw_text(label, bx+8, by+16, bw-16, F_BOLD, 12, RED_DARK, align="center"); _link(target, (bx,by,bx+bw,by+44))
        # Page 2: entry points and full factor table.
        if cat == "Кисти":
            page2_buttons = [(1230,798,210,"← категория", product_back),(1480,798,80,"стр.1",a1)]
        else:
            page2_buttons = [(1120,798,170,"← товар", product_back),(1310,798,150,"← категория",cat_list_key),(1480,798,80,"стр.1",a1)]
        _draw_entity_entry_factor_page("article", a2, f"Артикул: {art}", f"{cat} / товар {prod} / точки входа и факторы", "Артикул 2/2", art_row, page2_buttons)

    def _render_contour(contour: str):
        for cat in CATEGORY_ORDER:
            if contours[contour]["cat_df"][contours[contour]["cat_df"]["subject_disp"].astype(str).eq(cat)].empty:
                continue
            if not _has_category_detail(contour, cat):
                continue
            _draw_category_detail(contour, cat)
            if cat == "Кисти":
                arts = _articles_for_product(contour, cat, "901")
                for _, ar in arts.iterrows():
                    _draw_article_pages(contour, ar)
            else:
                prods = _children_for_category(contour, cat)
                for _, prow in prods.iterrows():
                    _draw_product_detail(contour, prow)
                    arts = _articles_for_product(contour, str(prow["subject_disp"]), str(prow["product_code"]))
                    for _, ar in arts.iterrows():
                        _draw_article_pages(contour, ar)

    # ---------- build pages in the requested order ----------
    # 1) Current week.
    _current_week_overview()
    # 2) Previous full week summary.
    _summary_category_page("prev_summary", "Прошлая полная неделя", f"{_period_label(prev_start, prev_end)} / категория → товар → артикул", "Прошлая неделя", prev_cat, target_contour="prev")
    # 3) Current month.
    _current_month_page()
    # 4) Last closed month summary.
    _summary_category_page("closed_summary", "Последний закрытый месяц", f"{_period_label(closed_start, closed_end)} / категория → товар → артикул", "Закрытый месяц", closed_cat, target_contour="closed")
    # 5) Current-year monthly ABC dynamics.
    _summary_page()
    # Details go after the 5 executive pages.
    _render_contour("prev")
    _render_contour("closed")

    c.save()

    # ---------- audit/trace workbook ----------
    # Делает не короткую сводку, а подробный след расчёта: откуда взялась каждая
    # метрика в PDF, какие строки/периоды были использованы и какие формулы дали
    # итоговые значения. Это нужно, чтобы быстро ловить смешение ABC/daily/search.
    try:
        trace_path = path.parent / PDF_CALC_TRACE_NAME

        trace_rows: List[Dict[str, Any]] = []
        sample_rows: List[Dict[str, Any]] = []
        demand_rows: List[Dict[str, Any]] = []
        formula_rows: List[Dict[str, Any]] = []
        factor_rows_out: List[Dict[str, Any]] = []
        source_rows: List[Dict[str, Any]] = []
        nav_rows: List[Dict[str, Any]] = []
        raw_sample_rows: List[Dict[str, Any]] = []

        sample_category = os.getenv("PDF_TRACE_SAMPLE_CATEGORY", "Кисти").strip() or "Кисти"
        sample_product = os.getenv("PDF_TRACE_SAMPLE_PRODUCT", "901").strip() or "901"
        sample_article_env = os.getenv("PDF_TRACE_SAMPLE_ARTICLE", "").strip()

        def _trace_entity_id(level: str, row: pd.Series) -> str:
            cat = str(row.get("subject_disp", ""))
            prod = str(row.get("product_code", ""))
            art = str(row.get("supplier_article", ""))
            nm = str(row.get("nm_id", ""))
            if level == "category":
                return cat
            if level == "product":
                return f"{cat} / {prod}"
            if nm and nm != "0" and nm.lower() != "nan":
                return f"{cat} / {prod} / {art} / nm={nm}"
            return f"{cat} / {prod} / {art}"

        def _trace_source(row: pd.Series, metric: str, current: bool = True) -> Tuple[str, str, str, bool, str]:
            has = bool(row.get("has_abc" if current else "has_abc_prev", False))
            suffix = "" if current else "_prev"
            if metric in ["Сумма", "ВП ABC", "Рентабельность", "Расход РК", "ДРР", "Комиссия, %", "Эквайринг, %"]:
                if has:
                    col_map = {
                        "Сумма": "gross_revenue / revenue_abc",
                        "ВП ABC": "gross_profit / gp_abc",
                        "Рентабельность": "gp_abc / revenue_abc",
                        "Расход РК": "gross_revenue * abc_drr_pct / 100 / abc_ad_spend",
                        "ДРР": "abc_ad_spend / revenue_abc",
                        "Комиссия, %": "abc_commission_amount / revenue_abc",
                        "Эквайринг, %": "abc_acquiring_amount / revenue_abc",
                    }
                    return "ABC exact", "ABC weekly/monthly from Object Storage", col_map.get(metric, ""), False, "ABC exact найден для периода"
                # fallback
                fallback_cols = {
                    "Сумма": "article_day_fact.order_sum",
                    "ВП ABC": "article_day_fact.gross_profit_model",
                    "Рентабельность": "gross_profit_model / order_sum",
                    "Расход РК": "article_day_fact.ad_spend_total",
                    "ДРР": "ad_spend_total / order_sum",
                    "Комиссия, %": "article_day_fact.commission_%",
                    "Эквайринг, %": "article_day_fact.acquiring_%",
                }
                return "daily/model fallback", "Технические_расчеты_TOPFACE.xlsx / article_day_fact", fallback_cols.get(metric, ""), True, "ABC exact не найден, использован оперативный/модельный расчёт"
            if metric == "CPC":
                spend_src = "ABC ad_spend" if has else "daily ad_spend_total"
                return "mixed allowed", "ABC/daily + article_day_fact", f"{spend_src}; clicks=article_day_fact.ad_clicks_total", False, "CPC = Расход РК выбранного источника / клики рекламы"
            if metric == "CTR РК":
                return "ads reports", "Отчёты/Реклама → ads_truth", "clicks_truth / impressions_truth * 100", False, "CTR РК = все клики РК / все показы РК"
            if metric in ["Спрос WB", "% поиска"]:
                ds = str(row.get("demand_source" + suffix, row.get("demand_source", "")))
                if ds == "unique_queries":
                    return "unique search queries", "Технические_расчеты_TOPFACE.xlsx / search_unique_demand", "unique_search_frequency, unique_search_queries, duplicate_query_rows_removed", False, "спрос дедублирован по нормализованным поисковым запросам"
                return "daily_sum_fallback", "Технические_расчеты_TOPFACE.xlsx / article_day_fact", "search_frequency", True, "нет листа search_unique_demand или нет строк уровня; спрос может быть завышен дублями"
            if metric in ["Открытия", "Конв. в заказ", "Заказы", "Логистика/шт", "Хранение/шт", "Себест./шт", "Прочие/шт", "СПП"]:
                return "daily operational", "Технические_расчеты_TOPFACE.xlsx / article_day_fact", "open_cards/add_to_cart/orders/cost fields", False, "оперативные карточные/юнит-метрики"
            return "unknown", "", "", True, "источник не классифицирован"

        def _value_pair(row: pd.Series, metric: str) -> Tuple[Any, Any, str, str]:
            mapping = {
                "Сумма": ("sum_use", "sum_prev_use", "sum_use = ABC gross_revenue если exact ABC найден, иначе SUM(order_sum)", "₽"),
                "ВП ABC": ("gp_use", "gp_prev_use", "gp_use = ABC gross_profit если exact ABC найден, иначе gross_profit_model", "₽"),
                "Рентабельность": ("margin", "margin_prev", "Рентабельность = ВП / Сумма * 100", "%"),
                "Расход РК": ("ad_spend", "ad_spend_prev", "Расход РК = ABC promotion/ABC revenue*drr если ABC, иначе SUM(ad_spend_total)", "₽"),
                "ДРР": ("drr", "drr_prev", "ДРР = Расход РК / Сумма * 100", "%"),
                "CPC": ("cpc", "cpc_prev", "CPC = Расход РК / Клики РК", "₽"),
                "CTR РК": ("ad_ctr", "ad_ctr_prev", "CTR РК = Клики РК / Показы РК * 100", "%"),
                "Спрос WB": ("demand", "demand_prev", "Спрос WB = SUM(unique_search_frequency) по уникальным запросам; fallback=SUM(search_frequency)", "шт"),
                "% поиска": ("search_share", "search_share_prev", "% поиска = Открытия карточки / Спрос WB * 100", "%"),
                "Открытия": ("opens", "opens_prev", "Открытия = SUM(open_cards)", "шт"),
                "Конв. в заказ": ("order_from_open_conv", "order_from_open_conv_prev", "Конв. в заказ = Заказы / Открытия карточки * 100", "%"),
                "Заказы": ("orders", "orders_prev", "Заказы = SUM(orders) из отчёта WB Заказы после исключения отмен", "шт"),
                "Комиссия, %": ("commission_pct", "commission_pct_prev", "Комиссия, % = ABC commission / ABC revenue * 100 если ABC, иначе commission_%", "%"),
                "Эквайринг, %": ("acquiring_pct", "acquiring_pct_prev", "Эквайринг, % = ABC acquiring / ABC revenue * 100 если ABC, иначе acquiring_%", "%"),
                "Логистика/шт": ("logistics_per_unit", "logistics_per_unit_prev", "Логистика/шт = среднее logistics_direct из daily", "₽/шт"),
                "Хранение/шт": ("storage_per_unit", "storage_per_unit_prev", "Хранение/шт = среднее storage из daily", "₽/шт"),
                "Себест./шт": ("cost_per_unit", "cost_per_unit_prev", "Себест./шт = среднее cost из daily", "₽/шт"),
                "Прочие/шт": ("other_per_unit", "other_per_unit_prev", "Прочие/шт = среднее other_costs из daily", "₽/шт"),
                "СПП": ("spp", "spp_prev", "СПП = среднее spp из daily", "%"),
            }
            cur_col, prev_col, formula, unit = mapping[metric]
            return row.get(cur_col, np.nan), row.get(prev_col, np.nan), formula, unit

        metrics_for_trace = [
            "Сумма", "Заказы", "ВП ABC", "Рентабельность", "Расход РК", "ДРР", "CPC",
            "Спрос WB", "% поиска", "CTR РК", "Открытия", "Конв. в заказ", "Заказы",
            "Комиссия, %", "Эквайринг, %", "Логистика/шт", "Хранение/шт", "Себест./шт", "Прочие/шт", "СПП",
        ]

        # Determine one sample article: env override first, otherwise top GP article inside sample product.
        sample_article = sample_article_env
        if not sample_article and isinstance(prev_art, pd.DataFrame) and not prev_art.empty:
            qsa = prev_art[(prev_art.get("subject_disp", "").astype(str).eq(sample_category)) & (prev_art.get("product_code", "").astype(str).eq(sample_product))].copy()
            if not qsa.empty:
                qsa["_gp"] = pd.to_numeric(qsa.get("gp_use"), errors="coerce").fillna(0)
                qsa = qsa.sort_values("_gp", ascending=False)
                sample_article = str(qsa.iloc[0].get("supplier_article", ""))
        if not sample_article:
            sample_article = "901/5"

        frame_specs = [
            ("current_week", "category", f"{cur_start:%d.%m}-{cur_actual_end:%d.%m.%Y}", cur_cat),
            ("current_month", "category", f"{cur_start.replace(day=1):%d.%m}-{cur_actual_end:%d.%m.%Y}", current_month_cat),
        ]
        for cname, info in contours.items():
            frame_specs.extend([
                (cname, "category", info["period"], info.get("cat_df", pd.DataFrame())),
                (cname, "product", info["period"], info.get("prod_df", pd.DataFrame())),
                (cname, "article", info["period"], info.get("art_df", pd.DataFrame())),
            ])

        def _is_sample(level: str, row: pd.Series) -> bool:
            cat = str(row.get("subject_disp", ""))
            prod = str(row.get("product_code", ""))
            art = str(row.get("supplier_article", ""))
            if level == "category":
                return cat == sample_category
            if level == "product":
                return cat == sample_category and prod == sample_product
            if level == "article":
                return cat == sample_category and prod == sample_product and art == sample_article
            return False

        for contour_name, level_name, period_label, df in frame_specs:
            if df is None or not isinstance(df, pd.DataFrame) or df.empty:
                continue
            for _, row in df.iterrows():
                row_s = row if isinstance(row, pd.Series) else pd.Series(row)
                for metric in metrics_for_trace:
                    cur_val, prev_val, formula, unit = _value_pair(row_s, metric)
                    delta = _delta(cur_val, prev_val)
                    src, src_file, src_cols, fallback, reason = _trace_source(row_s, metric, current=True)
                    prev_src, prev_src_file, prev_src_cols, prev_fallback, prev_reason = _trace_source(row_s, metric, current=False)
                    base_name = "прошлая неделя" if contour_name in ["prev", "current_week"] else "предыдущий месяц" if contour_name == "closed" else "прошлый период"
                    rec = {
                        "contour": contour_name,
                        "level": level_name,
                        "period": period_label,
                        "entity": _trace_entity_id(level_name, row_s),
                        "category": row_s.get("subject_disp", ""),
                        "product": row_s.get("product_code", ""),
                        "article": row_s.get("supplier_article", ""),
                        "nm_id": row_s.get("nm_id", ""),
                        "metric": metric,
                        "current_value": cur_val,
                        "base_value": prev_val,
                        "base_name": base_name,
                        "delta_pct": delta,
                        "unit": unit,
                        "source_current": src,
                        "source_current_file_sheet": src_file,
                        "source_current_columns": src_cols,
                        "source_base": prev_src,
                        "source_base_file_sheet": prev_src_file,
                        "source_base_columns": prev_src_cols,
                        "formula": formula,
                        "fallback_used_current": fallback,
                        "fallback_used_base": prev_fallback,
                        "source_reason_current": reason,
                        "source_reason_base": prev_reason,
                        "daily_rows_current": row_s.get("daily_rows", np.nan),
                        "daily_rows_base": row_s.get("daily_rows_prev", np.nan),
                        "active_days_current": row_s.get("active_days", np.nan),
                        "active_days_base": row_s.get("active_days_prev", np.nan),
                        "abc_rows_current": row_s.get("abc_rows", np.nan),
                        "abc_rows_base": row_s.get("abc_rows_prev_abc", np.nan),
                        "has_abc_current": row_s.get("has_abc", False),
                        "has_abc_base": row_s.get("has_abc_prev", False),
                        "demand_source_current": row_s.get("demand_source", ""),
                        "demand_source_base": row_s.get("demand_source_prev", ""),
                        "unique_queries_current": row_s.get("unique_queries", np.nan),
                        "unique_queries_base": row_s.get("unique_queries_prev", np.nan),
                        "raw_query_rows_current": row_s.get("raw_query_rows", np.nan),
                        "raw_query_rows_base": row_s.get("raw_query_rows_prev", np.nan),
                        "duplicate_query_rows_removed_current": row_s.get("duplicate_query_rows_removed", np.nan),
                        "duplicate_query_rows_removed_base": row_s.get("duplicate_query_rows_removed_prev", np.nan),
                        "daily_demand_sum_before_unique_current": row_s.get("demand_daily_sum", np.nan),
                        "daily_demand_sum_before_unique_base": row_s.get("demand_daily_sum_prev", np.nan),
                    }
                    trace_rows.append(rec)
                    if _is_sample(level_name, row_s):
                        sample_rows.append(rec)

                # Demand-specific trace row per entity.
                demand_rows.append({
                    "contour": contour_name,
                    "level": level_name,
                    "period": period_label,
                    "entity": _trace_entity_id(level_name, row_s),
                    "category": row_s.get("subject_disp", ""),
                    "product": row_s.get("product_code", ""),
                    "article": row_s.get("supplier_article", ""),
                    "demand_final": row_s.get("demand", np.nan),
                    "demand_base": row_s.get("demand_prev", np.nan),
                    "demand_source": row_s.get("demand_source", ""),
                    "demand_source_base": row_s.get("demand_source_prev", ""),
                    "unique_queries": row_s.get("unique_queries", np.nan),
                    "unique_queries_base": row_s.get("unique_queries_prev", np.nan),
                    "raw_query_rows": row_s.get("raw_query_rows", np.nan),
                    "raw_query_rows_base": row_s.get("raw_query_rows_prev", np.nan),
                    "duplicates_removed": row_s.get("duplicate_query_rows_removed", np.nan),
                    "duplicates_removed_base": row_s.get("duplicate_query_rows_removed_prev", np.nan),
                    "old_sum_by_articles_current": row_s.get("demand_daily_sum", np.nan),
                    "old_sum_by_articles_base": row_s.get("demand_daily_sum_prev", np.nan),
                    "formula": "unique demand: SUM(MAX(frequency) by date+level+normalized query); fallback: SUM(article_day_fact.search_frequency)",
                    "WARNING": "если demand_source=daily_sum_fallback, спрос может быть задвоен по артикулам",
                })

                # Formula checks for key metrics.
                sum_v = _num(row_s.get("sum_use")); gp_v = _num(row_s.get("gp_use")); ad_v = _num(row_s.get("ad_spend")); clicks_v = _num(row_s.get("clicks")); opens_v = _num(row_s.get("opens")); demand_v = _num(row_s.get("demand")); carts_v = _num(row_s.get("carts")); orders_v = _num(row_s.get("orders"))
                checks = [
                    ("Рентабельность", _num(row_s.get("margin")), gp_v / sum_v * 100 if abs(sum_v) > 1e-9 else np.nan, "gp_use / sum_use * 100"),
                    ("ДРР", _num(row_s.get("drr")), ad_v / sum_v * 100 if abs(sum_v) > 1e-9 else np.nan, "ad_spend / sum_use * 100"),
                    ("CPC", _num(row_s.get("cpc")), ad_v / clicks_v if clicks_v > 0 else 0.0, "ad_spend / clicks"),
                    ("% поиска", _num(row_s.get("search_share")), opens_v / demand_v * 100 if demand_v > 0 else np.nan, "opens / demand * 100"),
                    ("Конв. в заказ", _num(row_s.get("order_from_open_conv")), orders_v / opens_v * 100 if opens_v > 0 else np.nan, "orders / opens * 100"),
                ]
                for metric, shown, recomputed, formula in checks:
                    diff = shown - recomputed if not pd.isna(recomputed) else np.nan
                    formula_rows.append({
                        "contour": contour_name,
                        "level": level_name,
                        "period": period_label,
                        "entity": _trace_entity_id(level_name, row_s),
                        "metric": metric,
                        "shown_value": shown,
                        "recomputed_value": recomputed,
                        "diff": diff,
                        "status": "OK" if (pd.isna(diff) or abs(diff) < 0.05) else "CHECK",
                        "formula": formula,
                        "sum_use": sum_v,
                        "gp_use": gp_v,
                        "ad_spend": ad_v,
                        "clicks": clicks_v,
                        "opens": opens_v,
                        "demand": demand_v,
                        "carts": carts_v,
                        "orders": orders_v,
                    })

                if _is_sample(level_name, row_s):
                    for fr in _factor_rows(row_s, level_name):
                        rr = dict(fr)
                        rr.update({
                            "contour": contour_name,
                            "level": level_name,
                            "period": period_label,
                            "entity": _trace_entity_id(level_name, row_s),
                            "category": row_s.get("subject_disp", ""),
                            "product": row_s.get("product_code", ""),
                            "article": row_s.get("supplier_article", ""),
                        })
                        factor_rows_out.append(rr)

        # Raw daily rows for the sample category/product/article; helps find where wrong values enter.
        try:
            raw_periods = [
                ("current_week", cur_start, cur_actual_end),
                ("prev_week", prev_start, prev_end),
                ("prev_base_week", prev2_start, prev2_end),
                ("closed_month", closed_start, closed_end),
                ("closed_base_month", closed_prev_start, closed_prev_end),
            ]
            for pname, ps, pe in raw_periods:
                rx = daily[(daily["day"] >= ps) & (daily["day"] <= pe)].copy()
                for lvl, mask in [
                    ("category", rx["subject_disp"].astype(str).eq(sample_category)),
                    ("product", rx["subject_disp"].astype(str).eq(sample_category) & rx["product_code"].astype(str).eq(sample_product)),
                    ("article", rx["subject_disp"].astype(str).eq(sample_category) & rx["product_code"].astype(str).eq(sample_product) & rx["supplier_article"].astype(str).map(_clean_article_local).eq(sample_article)),
                ]:
                    part = rx[mask].copy()
                    raw_sample_rows.append({
                        "period_name": pname,
                        "level": lvl,
                        "period_start": ps,
                        "period_end": pe,
                        "sample_category": sample_category,
                        "sample_product": sample_product,
                        "sample_article": sample_article,
                        "raw_rows": len(part),
                        "active_days": part["day"].nunique() if not part.empty else 0,
                        "order_sum_sum": pd.to_numeric(part.get("order_sum", 0), errors="coerce").fillna(0).sum() if not part.empty else 0,
                        "ad_spend_total_sum": pd.to_numeric(part.get("ad_spend_total", 0), errors="coerce").fillna(0).sum() if not part.empty else 0,
                        "open_cards_sum": pd.to_numeric(part.get("open_cards", 0), errors="coerce").fillna(0).sum() if not part.empty else 0,
                        "search_frequency_sum_raw_daily": pd.to_numeric(part.get("search_frequency", 0), errors="coerce").fillna(0).sum() if not part.empty else 0,
                        "orders_sum": pd.to_numeric(part.get("orders", 0), errors="coerce").fillna(0).sum() if not part.empty else 0,
                    })
        except Exception as exc:
            raw_sample_rows.append({"error": str(exc)})

        # Sources + exact ABC usage overview.
        source_rows.extend([
            {"source": "article_day_fact", "file_sheet": "Технические_расчеты_TOPFACE.xlsx / article_day_fact", "used_for": "оперативная сумма, реклама, клики, открытия, корзины, конверсии, daily fallback", "risk": "не использовать как финансы закрытого периода при наличии ABC"},
            {"source": "search_unique_demand", "file_sheet": "Технические_расчеты_TOPFACE.xlsx / search_unique_demand", "used_for": "Спрос WB на уровнях категория/товар/артикул", "risk": "если лист отсутствует, PDF падает в fallback и спрос может быть задвоен"},
            {"source": "abc_weekly", "file_sheet": "ABC weekly from Object Storage", "used_for": "закрытая неделя: выручка, ВП, рентабельность, ДРР, комиссия, эквайринг", "risk": "должен быть exact period"},
            {"source": "abc_monthly", "file_sheet": "ABC monthly from Object Storage", "used_for": "закрытый месяц: выручка, ВП, рентабельность, ДРР, комиссия, эквайринг", "risk": "должен быть exact month"},
            {"source": "entry_points_bridge", "file_sheet": "Факторный_мост_ВП_TOPFACE.xlsx / entry_points_bridge", "used_for": "точки входа на странице артикула 2/2", "risk": "пока недельный контур"},
        ])

        for cname, info in contours.items():
            for cat in CATEGORY_ORDER:
                nav_rows.append({"contour": cname, "page_type": "category", "category": cat, "bookmark": _cat_key(cname, cat), "back_target": info["summary_key"], "list_target": _cat_key(cname, cat)+"_list", "factor_target": _cat_factor_key(cname, cat)})
            pdfp = info.get("prod_df", pd.DataFrame())
            if isinstance(pdfp, pd.DataFrame) and not pdfp.empty:
                for _, pr in pdfp.iterrows():
                    cat = str(pr.get("subject_disp", "")); prod = str(pr.get("product_code", ""))
                    nav_rows.append({"contour": cname, "page_type": "product", "category": cat, "product": prod, "bookmark": _prod_key(cname, cat, prod), "back_target": _cat_key(cname, cat)+"_list", "list_target": _prod_key(cname, cat, prod)+"_list", "factor_target": _prod_factor_key(cname, cat, prod)})

        with pd.ExcelWriter(trace_path, engine="openpyxl") as writer:
            pd.DataFrame([{
                "description": "Подробный лог расчётов PDF: каждая метрика, источник, формула, fallback и проверка формул.",
                "sample_category": sample_category,
                "sample_product": sample_product,
                "sample_article": sample_article,
                "important": "Если source_current=daily_sum_fallback для Спрос WB, значит нет search_unique_demand и спрос может быть неверным.",
            }]).to_excel(writer, sheet_name="README", index=False)
            pd.DataFrame(source_rows).to_excel(writer, sheet_name="sources", index=False)
            fix_checklist_rows = [
                {"check":"unique_demand_required", "status":"OK" if not search_unique_demand.empty else "FAIL", "how_to_fix":"полный запуск должен создать search_unique_demand; fallback запрещён"},
                {"check":"report_order", "status":"OK", "how_to_fix":"страницы идут: 1 текущая неделя → 2 прошлая неделя → 3 текущий месяц → 4 закрытый месяц → 5 год → детализация"},
                {"check":"current_month_weekly_only_gp_plan", "status":"OK", "how_to_fix":"лист current_month содержит ВП и % выполнения плана"},
                {"check":"product_sort_by_gp", "status":"OK", "how_to_fix":"товары внутри категории сортируются по gp_use desc"},
                {"check":"article_noise_filter", "status":"OK", "how_to_fix":"детализация товаров >=10% ВП категории; внутри детального товара выводятся все артикулы"},
                {"check":"ad_source_current", "status":"OK" if not ads_truth.empty else "WARN", "how_to_fix":"для текущих периодов нужен ads_raw_source из рекламного отчёта, иначе article_day_fact может задваивать расход"},
            ]
            pd.DataFrame(fix_checklist_rows).to_excel(writer, sheet_name="fix_checklist", index=False)
            pd.DataFrame(trace_rows).to_excel(writer, sheet_name="metric_trace_all", index=False)
            pd.DataFrame(sample_rows).to_excel(writer, sheet_name="metric_trace_sample", index=False)
            pd.DataFrame(demand_rows).to_excel(writer, sheet_name="demand_trace", index=False)
            pd.DataFrame(formula_rows).to_excel(writer, sheet_name="formula_checks", index=False)
            pd.DataFrame(factor_rows_out).to_excel(writer, sheet_name="factor_effects_sample", index=False)
            pd.DataFrame(raw_sample_rows).to_excel(writer, sheet_name="raw_daily_sample", index=False)
            pd.DataFrame(nav_rows).to_excel(writer, sheet_name="navigation_expected", index=False)
            # Rendered product rows for quick category/product inspection.
            for cname, info in contours.items():
                if isinstance(info.get("prod_df"), pd.DataFrame) and not info["prod_df"].empty:
                    info["prod_df"].to_excel(writer, sheet_name=(cname + "_products")[:31], index=False)
                if isinstance(info.get("cat_df"), pd.DataFrame) and not info["cat_df"].empty:
                    info["cat_df"].to_excel(writer, sheet_name=(cname + "_categories")[:31], index=False)

        log(f"Saved detailed PDF metric trace: {trace_path} trace_rows={len(trace_rows):,}; sample_rows={len(sample_rows):,}; demand_rows={len(demand_rows):,}")
        log(f"PDF TRACE SAMPLE: category={sample_category}; product={sample_product}; article={sample_article}")
    except Exception as exc:
        log(f"WARN PDF detailed calc trace was not saved: {exc}")
    log(f"PDF v11 three-contour report created: pages={page_num}")
    return path


# ------------------------- Telegram daily summary + S3 report retention -------------------------
TELEGRAM_REPORT_MONTHS_RU = {v: k for k, v in REPORT_MONTH_GENITIVE_RU.items()}


def _tg_fmt_money(x: Any) -> str:
    v = to_number(x)
    if pd.isna(v):
        return "—"
    return f"{int(round(float(v))):,} ₽".replace(",", " ")


def _tg_fmt_num(x: Any) -> str:
    v = to_number(x)
    if pd.isna(v):
        return "—"
    return f"{int(round(float(v))):,}".replace(",", " ")


def _tg_fmt_pct(x: Any, digits: int = 1) -> str:
    v = to_number(x)
    if pd.isna(v):
        return "—"
    return f"{float(v):.{digits}f}%".replace(".", ",")


def _tg_fmt_rub1(x: Any) -> str:
    v = to_number(x)
    if pd.isna(v):
        return "—"
    return f"{float(v):.1f} ₽".replace(".", ",")


def _tg_delta_pct(cur: Any, prev: Any) -> Optional[float]:
    cur_v = to_number(cur)
    prev_v = to_number(prev)
    if pd.isna(cur_v) or pd.isna(prev_v) or abs(float(prev_v)) < 1e-9 or float(prev_v) < 0:
        return None
    d = (float(cur_v) / float(prev_v) - 1.0) * 100.0
    if abs(d) > 999.0:
        return None
    return d


def _tg_arrow_pct(cur: Any, prev: Any, lower_bad: bool = False) -> str:
    d = _tg_delta_pct(cur, prev)
    if d is None or abs(d) < 0.05:
        return "→ 0,0%"
    return ("↑ " if d > 0 else "↓ ") + f"{abs(d):.1f}%".replace(".", ",")


def _tg_arrow_abs(cur: Any, prev: Any, suffix: str = "") -> str:
    cur_v = to_number(cur)
    prev_v = to_number(prev)
    if pd.isna(cur_v) or pd.isna(prev_v):
        return "→ 0" + suffix
    d = float(cur_v) - float(prev_v)
    if abs(d) < 0.5:
        return "→ 0" + suffix
    val = f"{int(round(abs(d))):,}".replace(",", " ")
    return ("↑ +" if d > 0 else "↓ -") + val + suffix


def _tg_subject_disp(value: Any) -> str:
    subj = canonical_subject(value)
    mp = {
        "Кисти косметические": "Кисти",
        "Косметические карандаши": "Карандаши",
        "Карандаши": "Карандаши",
        "Кисти": "Кисти",
        "Помады": "Помады",
        "Блески": "Блески",
    }
    return mp.get(subj, normalize_text(value))


def _tg_prepare_daily(outputs: Dict[str, Any]) -> pd.DataFrame:
    daily = outputs.get("article_day_fact", pd.DataFrame()) if isinstance(outputs, dict) else pd.DataFrame()
    if not isinstance(daily, pd.DataFrame) or daily.empty:
        return pd.DataFrame()
    rejects: List[Dict[str, Any]] = []
    daily = _filter_df_by_pdf_product_reference(daily, "telegram_article_day_fact", rejects)
    if daily.empty:
        return pd.DataFrame()
    daily = daily.copy()
    daily["day"] = pd.to_datetime(daily.get("day"), errors="coerce").dt.normalize()
    daily["subject_disp"] = daily.get("subject", "").map(_tg_subject_disp) if "subject" in daily.columns else daily.get("subject_disp", "").map(_tg_subject_disp)
    daily = daily[daily["subject_disp"].isin(["Кисти", "Карандаши", "Помады", "Блески"])].copy()
    for col in ["order_sum", "orders", "open_cards", "search_frequency", "ad_spend_total", "ad_clicks_total", "ad_impressions_total", "manual_spend", "unified_spend", "unknown_spend", "manual_clicks", "unified_clicks", "unknown_clicks", "manual_impressions", "unified_impressions", "unknown_impressions"]:
        if col not in daily.columns:
            daily[col] = 0.0
        daily[col] = pd.to_numeric(daily[col], errors="coerce").fillna(0.0)
    if "ad_spend_total" not in daily.columns or daily["ad_spend_total"].abs().sum() < 1e-9:
        daily["ad_spend_total"] = daily[[c for c in ["manual_spend", "unified_spend", "unknown_spend"] if c in daily.columns]].sum(axis=1)
    if "ad_clicks_total" not in daily.columns or daily["ad_clicks_total"].abs().sum() < 1e-9:
        daily["ad_clicks_total"] = daily[[c for c in ["manual_clicks", "unified_clicks", "unknown_clicks"] if c in daily.columns]].sum(axis=1)
    if "ad_impressions_total" not in daily.columns or daily["ad_impressions_total"].abs().sum() < 1e-9:
        daily["ad_impressions_total"] = daily[[c for c in ["manual_impressions", "unified_impressions", "unknown_impressions"] if c in daily.columns]].sum(axis=1)
    return daily


def _tg_prepare_ads_truth(outputs: Dict[str, Any]) -> Tuple[pd.DataFrame, str]:
    source_name = ""
    ads = pd.DataFrame()
    for nm in ["ads_category_source", "ads_raw_source", "ads_daily_source"]:
        cand = outputs.get(nm, pd.DataFrame()) if isinstance(outputs, dict) else pd.DataFrame()
        if isinstance(cand, pd.DataFrame) and not cand.empty:
            ads = cand.copy()
            source_name = nm
            break
    if ads.empty:
        return pd.DataFrame(), ""
    if "subject" not in ads.columns:
        if "subject_disp" in ads.columns:
            rev = {"Кисти": "Кисти косметические", "Карандаши": "Косметические карандаши", "Помады": "Помады", "Блески": "Блески"}
            ads["subject"] = ads["subject_disp"].map(lambda x: rev.get(normalize_text(x), normalize_text(x)))
        else:
            ads["subject"] = ""
    for col in ["product", "supplier_article", "nm_id"]:
        if col not in ads.columns:
            ads[col] = "" if col != "nm_id" else np.nan
    rejects: List[Dict[str, Any]] = []
    ads = _filter_df_by_pdf_product_reference(ads, f"telegram_{source_name}", rejects)
    if ads.empty:
        return pd.DataFrame(), source_name
    ads["day"] = pd.to_datetime(ads.get("day"), errors="coerce").dt.normalize()
    ads["subject_disp"] = ads.get("subject", "").map(_tg_subject_disp)
    ads = ads[ads["subject_disp"].isin(["Кисти", "Карандаши", "Помады", "Блески"])].copy()
    # Normalize names from different ad-report sheets.
    rename_candidates = {
        "ad_spend_total": "spend",
        "ad_clicks_total": "clicks",
        "ad_impressions_total": "impressions",
        "shows": "impressions",
        "views": "impressions",
    }
    for old, new in rename_candidates.items():
        if new not in ads.columns and old in ads.columns:
            ads[new] = ads[old]
    for col in ["spend", "clicks", "impressions"]:
        if col not in ads.columns:
            ads[col] = 0.0
        ads[col] = pd.to_numeric(ads[col], errors="coerce").fillna(0.0)
    return ads, source_name


def _tg_prepare_unique_demand(outputs: Dict[str, Any]) -> pd.DataFrame:
    df = outputs.get("search_unique_demand", pd.DataFrame()) if isinstance(outputs, dict) else pd.DataFrame()
    if not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame()
    x = df.copy()
    x["day"] = pd.to_datetime(x.get("day"), errors="coerce").dt.normalize()
    if "level" in x.columns:
        x = x[x["level"].astype(str).str.lower().eq("category")].copy()
    if "subject_disp" in x.columns:
        x["subject_disp"] = x["subject_disp"].map(_tg_subject_disp)
    elif "subject" in x.columns:
        x["subject_disp"] = x["subject"].map(_tg_subject_disp)
    else:
        x["subject_disp"] = ""
    x = x[x["subject_disp"].isin(["Кисти", "Карандаши", "Помады", "Блески"])].copy()
    if "unique_search_frequency" not in x.columns:
        x["unique_search_frequency"] = 0.0
    x["unique_search_frequency"] = pd.to_numeric(x["unique_search_frequency"], errors="coerce").fillna(0.0)
    return x


def _tg_sum_period(daily: pd.DataFrame, ads: pd.DataFrame, demand_df: pd.DataFrame, start: pd.Timestamp, end: pd.Timestamp) -> Dict[str, float]:
    start = pd.Timestamp(start).normalize(); end = pd.Timestamp(end).normalize()
    d = daily[(daily["day"] >= start) & (daily["day"] <= end)].copy() if not daily.empty else pd.DataFrame()
    a = ads[(ads["day"] >= start) & (ads["day"] <= end)].copy() if not ads.empty else pd.DataFrame()
    q = demand_df[(demand_df["day"] >= start) & (demand_df["day"] <= end)].copy() if not demand_df.empty else pd.DataFrame()
    order_sum = float(pd.to_numeric(d.get("order_sum", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not d.empty else 0.0
    orders = float(pd.to_numeric(d.get("orders", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not d.empty else 0.0
    opens = float(pd.to_numeric(d.get("open_cards", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not d.empty else 0.0
    fallback_demand = float(pd.to_numeric(d.get("search_frequency", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not d.empty else 0.0
    demand = float(pd.to_numeric(q.get("unique_search_frequency", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not q.empty else fallback_demand
    if not a.empty:
        spend = float(pd.to_numeric(a.get("spend", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
        clicks = float(pd.to_numeric(a.get("clicks", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
        impressions = float(pd.to_numeric(a.get("impressions", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    else:
        spend = float(pd.to_numeric(d.get("ad_spend_total", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not d.empty else 0.0
        clicks = float(pd.to_numeric(d.get("ad_clicks_total", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not d.empty else 0.0
        impressions = float(pd.to_numeric(d.get("ad_impressions_total", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not d.empty else 0.0
    return {
        "order_sum": order_sum,
        "orders": orders,
        "ad_spend": spend,
        "clicks": clicks,
        "impressions": impressions,
        "demand": demand,
        "opens": opens,
        "drr": spend / order_sum * 100.0 if order_sum else 0.0,
        "cpc": spend / clicks if clicks else np.nan,
        "ad_ctr": clicks / impressions * 100.0 if impressions else np.nan,
        "search_share": opens / demand * 100.0 if demand else np.nan,
    }


def build_telegram_daily_summary(outputs: Dict[str, Any]) -> str:
    """Build the Telegram text block matching PDF sheet 1, but for the latest concrete day."""
    daily = _tg_prepare_daily(outputs)
    if daily.empty or "day" not in daily.columns:
        return "TOPFACE: дневная сводка не сформирована — нет article_day_fact."
    meaningful = daily.copy()
    metric_cols = [c for c in ["order_sum", "orders", "open_cards", "ad_spend_total", "search_frequency"] if c in meaningful.columns]
    if metric_cols:
        m = meaningful[metric_cols].apply(pd.to_numeric, errors="coerce").fillna(0).abs().sum(axis=1) > 0
        meaningful = meaningful[m].copy()
    if meaningful.empty:
        return "TOPFACE: дневная сводка не сформирована — нет фактических строк за последний день."
    latest = pd.to_datetime(meaningful["day"], errors="coerce").dropna().max().normalize()
    cur_week_start = latest - pd.Timedelta(days=int(latest.weekday()))
    prev_start = cur_week_start - pd.Timedelta(days=7)
    prev_end = cur_week_start - pd.Timedelta(days=1)
    ads, ads_source = _tg_prepare_ads_truth(outputs)
    demand_df = _tg_prepare_unique_demand(outputs)
    cur = _tg_sum_period(daily, ads, demand_df, latest, latest)
    prev_total = _tg_sum_period(daily, ads, demand_df, prev_start, prev_end)
    prev_days = 7.0
    prev = {
        "order_sum": prev_total["order_sum"] / prev_days,
        "orders": prev_total["orders"] / prev_days,
        "ad_spend": prev_total["ad_spend"] / prev_days,
        "clicks": prev_total["clicks"] / prev_days,
        "impressions": prev_total["impressions"] / prev_days,
        "demand": prev_total["demand"] / prev_days,
        "opens": prev_total["opens"] / prev_days,
    }
    prev["drr"] = prev["ad_spend"] / prev["order_sum"] * 100.0 if prev["order_sum"] else 0.0
    prev["cpc"] = prev["ad_spend"] / prev["clicks"] if prev["clicks"] else np.nan
    prev["ad_ctr"] = prev["clicks"] / prev["impressions"] * 100.0 if prev["impressions"] else np.nan
    prev["search_share"] = prev["opens"] / prev["demand"] * 100.0 if prev["demand"] else np.nan
    label = f"{int(latest.day)} {REPORT_MONTH_GENITIVE_RU.get(int(latest.month), latest.strftime('%m'))}"
    prev_label = f"{prev_start:%d.%m}-{prev_end:%d.%m} / средний день"
    lines = [
        f"📊 TOPFACE — {label}",
        f"Сравнение: {prev_label}",
        "",
        f"💰 Сумма заказов: {_tg_fmt_money(cur['order_sum'])} {_tg_arrow_abs(cur['order_sum'], prev['order_sum'], ' ₽')}",
        f"🧾 Заказы: {_tg_fmt_num(cur['orders'])} {_tg_arrow_abs(cur['orders'], prev['orders'])}",
        f"📣 Расход РК: {_tg_fmt_money(cur['ad_spend'])} {_tg_arrow_abs(cur['ad_spend'], prev['ad_spend'], ' ₽')}",
        f"📉 ДРР: {_tg_fmt_pct(cur['drr'])} {_tg_arrow_pct(cur['drr'], prev['drr'], lower_bad=True)}",
        f"🔎 Спрос WB: {_tg_fmt_num(cur['demand'])} {_tg_arrow_pct(cur['demand'], prev['demand'])}",
        f"📌 % поиска: {_tg_fmt_pct(cur['search_share'])} {_tg_arrow_pct(cur['search_share'], prev['search_share'])}",
        f"🎯 CTR РК: {_tg_fmt_pct(cur['ad_ctr'])} {_tg_arrow_pct(cur['ad_ctr'], prev['ad_ctr'])}",
        f"💸 CPC: {_tg_fmt_rub1(cur['cpc'])} {_tg_arrow_pct(cur['cpc'], prev['cpc'], lower_bad=True)}",
    ]
    if ads_source:
        lines.append(f"Источник РК: {ads_source}")
    return "\n".join(lines)


def send_telegram_message(text: str) -> bool:
    token = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
    chat_id = os.getenv("TELEGRAM_CHAT_ID", "").strip()
    if not token or not chat_id:
        log("Telegram: TELEGRAM_BOT_TOKEN/TELEGRAM_CHAT_ID не заданы, текстовая сводка пропущена")
        return False
    if not text:
        log("Telegram: пустая текстовая сводка, отправка пропущена")
        return False
    import urllib.parse
    import urllib.request
    data = urllib.parse.urlencode({"chat_id": chat_id, "text": text}).encode("utf-8")
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    try:
        req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/x-www-form-urlencoded"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            body_resp = resp.read().decode("utf-8", errors="replace")[:500]
            ok = 200 <= resp.status < 300
            log(f"Telegram: daily summary {'sent' if ok else 'failed'} status={resp.status} response={body_resp}")
            return ok
    except Exception as exc:
        log(f"Telegram: ошибка отправки текстовой сводки: {exc}")
        return False


def _parse_sales_pdf_report_date(file_name: str, latest_day: pd.Timestamp) -> Optional[pd.Timestamp]:
    m = re.match(r"^Отчет по продажам Влад\s+(\d{1,2})\s+([а-яА-ЯёЁ]+)\.pdf$", file_name.strip())
    if not m:
        return None
    day = int(m.group(1))
    month = TELEGRAM_REPORT_MONTHS_RU.get(m.group(2).lower())
    if not month:
        return None
    year = int(pd.Timestamp(latest_day).year)
    try:
        d = pd.Timestamp(date(year, month, day)).normalize()
        # Around New Year, keep date inference sane.
        if d > pd.Timestamp(latest_day).normalize() + pd.Timedelta(days=31):
            d = pd.Timestamp(date(year - 1, month, day)).normalize()
        return d
    except Exception:
        return None


def cleanup_non_monday_sales_pdfs(storage: Storage, outputs: Dict[str, Any], keep_current_file_name: str) -> None:
    """Delete old daily PDF reports from OUT_DIR, preserving Monday reports and the current PDF."""
    enabled = os.getenv("WB_KEEP_ONLY_MONDAY_REPORTS", "1").strip().lower() in {"1", "true", "yes", "y"}
    if not enabled:
        log("cleanup: WB_KEEP_ONLY_MONDAY_REPORTS=0, старые PDF не удаляются")
        return
    if not getattr(storage, "is_s3", False):
        log("cleanup: local mode, S3 PDF cleanup skipped")
        return
    latest_day = _max_report_day_from_outputs(outputs)
    deleted = 0
    kept_monday = 0
    kept_current = 0
    for key in storage.list_files(OUT_DIR):
        name = Path(key).name
        if not name.startswith("Отчет по продажам Влад ") or not name.endswith(".pdf"):
            continue
        if name == keep_current_file_name:
            kept_current += 1
            continue
        d = _parse_sales_pdf_report_date(name, latest_day)
        if d is None:
            continue
        if d.weekday() == 0:
            kept_monday += 1
            continue
        # Do not touch reports that look newer than the current report date.
        if d > pd.Timestamp(latest_day).normalize():
            continue
        try:
            storage.delete_file(key)
            deleted += 1
            log(f"cleanup: deleted non-Monday daily PDF: {key}")
        except Exception as exc:
            log(f"WARN cleanup: не удалось удалить {key}: {exc}")
    log(f"cleanup: done, deleted={deleted}, kept_monday={kept_monday}, kept_current={kept_current}")


def send_telegram_report(outputs: Dict[str, Any], pdf_path: Path, caption: str = "", storage: Optional[Storage] = None) -> bool:
    msg_ok = send_telegram_message(build_telegram_daily_summary(outputs))
    doc_ok = send_telegram_document(pdf_path, caption)
    if doc_ok and storage is not None:
        cleanup_non_monday_sales_pdfs(storage, outputs, pdf_path.name)
    strict_msg = os.getenv("WB_TELEGRAM_STRICT_MESSAGE", "0").strip().lower() in {"1", "true", "yes", "y"}
    return bool(doc_ok and (msg_ok or not strict_msg))

def send_telegram_document(file_path: Path, caption: str = "") -> bool:
    token = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
    chat_id = os.getenv("TELEGRAM_CHAT_ID", "").strip()
    if not file_path.exists():
        log(f"Telegram: файл не найден, отправка невозможна: {file_path}")
        return False
    if not token or not chat_id:
        log("Telegram: TELEGRAM_BOT_TOKEN/TELEGRAM_CHAT_ID не заданы, отправка пропущена")
        return False
    log(f"Telegram: отправка PDF file={file_path.name} size={file_path.stat().st_size:,} bytes chat_id={chat_id}")
    import urllib.request
    import uuid
    boundary = "----WebKitFormBoundary" + uuid.uuid4().hex
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    fields = {"chat_id": chat_id, "caption": caption[:1000]}
    thread_id = os.getenv("TELEGRAM_MESSAGE_THREAD_ID", "").strip()
    if thread_id:
        fields["message_thread_id"] = thread_id
    body = bytearray()
    for name, value in fields.items():
        body.extend(f"--{boundary}\r\n".encode())
        body.extend(f'Content-Disposition: form-data; name="{name}"\r\n\r\n{value}\r\n'.encode())
    data = file_path.read_bytes()
    body.extend(f"--{boundary}\r\n".encode())
    body.extend(f'Content-Disposition: form-data; name="document"; filename="{file_path.name}"\r\n'.encode())
    body.extend(b"Content-Type: application/pdf\r\n\r\n")
    body.extend(data)
    body.extend(f"\r\n--{boundary}--\r\n".encode())
    req = urllib.request.Request(url, data=bytes(body), headers={"Content-Type": f"multipart/form-data; boundary={boundary}"})
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            body_resp = resp.read().decode("utf-8", errors="replace")[:500]
            ok = 200 <= resp.status < 300
            log(f"Telegram: {'sent' if ok else 'failed'} status={resp.status} response={body_resp}")
            return ok
    except Exception as exc:
        log(f"Telegram: ошибка отправки PDF: {exc}")
        return False




def run_smoke_test(root: str = ".") -> None:
    """Fast self-test: no S3, no heavy Excel parsing, finishes in seconds.

    Purpose: verify that the current code version imports, builds the PDF renderer,
    handles the management-report data model, writes local PDF/XLSX artifacts, and exits 0.
    It intentionally uses tiny synthetic data for March-April-May so the closed-month,
    current-week and previous-week PDF contours are all exercised.
    """
    log("SMOKE_TEST: start synthetic report test; S3 is not used")
    local_dir = Path(root) / OUT_DIR
    local_dir.mkdir(parents=True, exist_ok=True)

    latest = pd.Timestamp("2026-05-25")
    days = pd.date_range("2026-03-01", latest, freq="D")
    entities = [
        ("Кисти косметические", "901", "901/5", 110254021),
        ("Косметические карандаши", "605", "605/1", 213000001),
        ("Помады", "154", "154/1", 101000001),
        ("Блески", "207", "207/1", 102000001),
    ]
    rows = []
    for subject, product, article, nm_id in entities:
        base = 9000 if product == "901" else 5000 if subject == "Косметические карандаши" else 3500
        for d in days:
            weekday_mult = 1.15 if d.weekday() in (0, 1, 2) else 0.85
            order_sum = base * weekday_mult
            orders = max(1, int(order_sum / 450))
            ad_spend = order_sum * (0.11 if subject in {"Кисти косметические", "Косметические карандаши"} else 0.14)
            rows.append({
                "day": d,
                "subject": subject,
                "product": product,
                "supplier_article": article,
                "nm_id": nm_id,
                "orders": orders,
                "order_sum": order_sum,
                "gross_profit_model": order_sum * 0.32,
                "open_cards": orders * 28,
                "add_to_cart": orders * 7,
                "search_frequency": orders * 450,
                "search_traffic_capture_pct": 6.2,
                "direct_localization_pct": 66.0,
                "localization_with_replacements_pct": 91.0,
                "rating_reviews": 4.8,
                "finished_price": 450.0,
                "price_with_disc": 430.0,
                "spp": 12.0,
                "commission_%": 18.0,
                "acquiring_%": 1.5,
                "logistics_direct": 38.0,
                "storage": 4.0,
                "other_costs": 8.0,
                "cost": 120.0,
                "cart_conv_pct": 25.0,
                "order_conv_pct": 28.0,
                "manual_spend": ad_spend * 0.65,
                "unified_spend": ad_spend * 0.35,
                "unknown_spend": 0.0,
                "manual_clicks": max(1, int(ad_spend * 0.65 / 14)),
                "unified_clicks": max(1, int(ad_spend * 0.35 / 18)),
                "unknown_clicks": 0,
                "manual_impressions": max(1, int(ad_spend * 0.65 / 14 / 0.035)),
                "unified_impressions": max(1, int(ad_spend * 0.35 / 18 / 0.03)),
                "unknown_impressions": 0,
                "ad_spend_model": ad_spend,
            })
    daily = pd.DataFrame(rows)

    # Unique demand for all levels and all days: this prevents strict demand crashes.
    demand_rows = []
    subj_disp_map = {"Кисти косметические": "Кисти", "Косметические карандаши": "Карандаши", "Помады": "Помады", "Блески": "Блески"}
    for _, r in daily.iterrows():
        sf = float(r["search_frequency"])
        common = {"day": r["day"], "subject": r["subject"], "subject_disp": subj_disp_map[r["subject"]], "unique_search_frequency": sf, "unique_search_queries": 3, "duplicate_query_rows_removed": 1, "raw_query_rows": 4}
        demand_rows.append({**common, "level": "category"})
        demand_rows.append({**common, "level": "product", "product": r["product"], "product_code": r["product"]})
        demand_rows.append({**common, "level": "article", "product": r["product"], "product_code": r["product"], "supplier_article": r["supplier_article"], "nm_id": r["nm_id"]})
    search_unique = pd.DataFrame(demand_rows)

    ads = daily[["day", "subject", "product", "supplier_article", "nm_id"]].copy()
    ads["spend"] = daily[["manual_spend", "unified_spend", "unknown_spend"]].sum(axis=1)
    ads["clicks"] = daily[["manual_clicks", "unified_clicks", "unknown_clicks"]].sum(axis=1)
    ads["impressions"] = daily[["manual_impressions", "unified_impressions", "unknown_impressions"]].sum(axis=1)

    # Exact ABC periods for weekly and monthly contours.
    abc_rows = []
    periods = [
        (pd.Timestamp("2026-05-18"), pd.Timestamp("2026-05-24"), "weekly"),
        (pd.Timestamp("2026-05-11"), pd.Timestamp("2026-05-17"), "weekly"),
        (pd.Timestamp("2026-04-01"), pd.Timestamp("2026-04-30"), "monthly"),
        (pd.Timestamp("2026-03-01"), pd.Timestamp("2026-03-31"), "monthly"),
    ]
    for start, end, kind in periods:
        part = daily[(daily["day"] >= start) & (daily["day"] <= end)]
        for keys, g in part.groupby(["subject", "product", "supplier_article", "nm_id"], dropna=False):
            subject, product, article, nm_id = keys
            gross_revenue = float(g["order_sum"].sum())
            abc_rows.append({
                "period_start": start,
                "period_end": end,
                "week_code": week_code(start),
                "week_label": f"{start:%d.%m}-{end:%d.%m}",
                "month_key": month_key(start),
                "subject": subject,
                "product": product,
                "supplier_article": article,
                "nm_id": nm_id,
                "gross_profit": gross_revenue * 0.32,
                "gross_revenue": gross_revenue,
                "orders": float(g["orders"].sum()),
                "abc_margin_pct": 32.0,
                "abc_drr_pct": 12.0,
                "abc_commission_amount": gross_revenue * 0.18,
                "abc_acquiring_amount": gross_revenue * 0.015,
            })
    abc = pd.DataFrame(abc_rows)
    abc_weekly = abc[abc["period_start"].isin([pd.Timestamp("2026-05-18"), pd.Timestamp("2026-05-11")])].copy()
    abc_monthly = abc[abc["period_start"].isin([pd.Timestamp("2026-04-01"), pd.Timestamp("2026-03-01")])].copy()

    outputs = {
        "article_day_fact": daily,
        "search_unique_demand": search_unique,
        "ads_category_source": pd.DataFrame(),
        "ads_raw_source": ads,
        "ads_daily_source": ads,
        "abc_weekly": abc_weekly,
        "abc_monthly": abc_monthly,
        "factor_bridge": pd.DataFrame(),
        "entry_points_bridge": pd.DataFrame(),
        "optimal_benchmarks": pd.DataFrame(),
        "factor_summary_for_pdf": pd.DataFrame(),
    }
    pdf_path = local_dir / "SMOKE_TEST_Управленческий_отчет_TOPFACE.pdf"
    pdf_created = generate_management_pdf(outputs, pdf_path)
    if not pdf_created or not pdf_path.exists() or pdf_path.stat().st_size < 10_000:
        raise RuntimeError(f"SMOKE_TEST failed: PDF не создан или слишком маленький: {pdf_path}")
    tg_preview = build_telegram_daily_summary(outputs)
    if "Сумма заказов" not in tg_preview or "Расход РК" not in tg_preview:
        raise RuntimeError(f"SMOKE_TEST failed: Telegram daily summary invalid: {tg_preview[:200]}")
    log("SMOKE_TEST_OK: Telegram daily summary built")
    ok_path = local_dir / "SMOKE_TEST_OK.txt"
    ok_path.write_text(
        "SMOKE_TEST_OK\n"
        f"time={datetime.now():%Y-%m-%d %H:%M:%S}\n"
        f"pdf={pdf_path}\n"
        f"pdf_size={pdf_path.stat().st_size}\n",
        encoding="utf-8",
    )
    log(f"SMOKE_TEST_OK: PDF={pdf_path} size={pdf_path.stat().st_size:,} bytes")

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", default=".", help="Local root used for local copies and local mode")
    parser.add_argument("--reports-root", default="Отчёты")
    parser.add_argument("--store", default="TOPFACE")
    parser.add_argument("--no-pdf", action="store_true", help="Не формировать PDF")
    parser.add_argument("--pdf-only", action="store_true", help="Сформировать только PDF из уже готовых Excel-файлов без пересчета источников")
    parser.add_argument("--current-week-only", action="store_true", help="Быстро обновить только оперативный блок текущей недели из кэша/свежих данных")
    parser.add_argument("--full-refresh", action="store_true", help="Явный полный пересчет всех источников; синоним обычного запуска")
    parser.add_argument("--send-telegram", action="store_true", help="Отправить PDF в Telegram через TELEGRAM_BOT_TOKEN/TELEGRAM_CHAT_ID")
    parser.add_argument("--smoke-test", action="store_true", help="Быстрый синтетический тест PDF/кода без S3 и тяжелых Excel; должен занимать секунды")
    args = parser.parse_args()
    if args.smoke_test:
        run_smoke_test(args.root)
        return
    diagnostics = Diagnostics()
    storage = make_storage(args.root)
    local_dir = Path(args.root) / OUT_DIR
    if args.current_week_only:
        # Lightweight daily mode: load cached full-report outputs, then refresh only current-week
        # operational rows from a limited set of current-week source files. It does NOT rebuild ABC/month/year details.
        local_dir.mkdir(parents=True, exist_ok=True)
        outputs = load_existing_outputs_for_pdf(storage, local_dir, args.reports_root, args.store, diagnostics)
        os.environ["WB_CURRENT_WEEK_ONLY"] = "1"
        try:
            loader = Loader(storage, args.reports_root, args.store, diagnostics)
            pack = loader.load_all()
            builder = AnalyticsBuilder(pack)
            upd = builder.build_all()
            refreshed = upd.get("article_day_fact", pd.DataFrame())
            if refreshed is not None and not refreshed.empty and "day" in refreshed.columns:
                refreshed = refreshed.copy()
                refreshed["day"] = pd.to_datetime(refreshed["day"], errors="coerce").dt.normalize()
                latest = refreshed["day"].max()
                cw_start = latest - pd.Timedelta(days=int(pd.Timestamp(latest).weekday()))
                cw_end = cw_start + pd.Timedelta(days=6)
                log(f"current_week_only: refresh cache dates {cw_start.date()}..{cw_end.date()} using limited source load")
                for name in ["article_day_fact", "search_unique_demand", "ads_category_source", "ads_raw_source", "ads_daily_source"]:
                    new_df = upd.get(name, pd.DataFrame())
                    old_df = outputs.get(name, pd.DataFrame())
                    if new_df is None or new_df.empty or "day" not in new_df.columns:
                        continue
                    new_df = new_df.copy(); new_df["day"] = pd.to_datetime(new_df["day"], errors="coerce").dt.normalize()
                    new_df = new_df[(new_df["day"] >= cw_start) & (new_df["day"] <= cw_end)].copy()
                    if old_df is not None and not old_df.empty and "day" in old_df.columns:
                        old_df = old_df.copy(); old_df["day"] = pd.to_datetime(old_df["day"], errors="coerce").dt.normalize()
                        old_df = old_df[~((old_df["day"] >= cw_start) & (old_df["day"] <= cw_end))].copy()
                        outputs[name] = pd.concat([old_df, new_df], ignore_index=True, sort=False)
                    else:
                        outputs[name] = new_df
                    log(f"current_week_only: merged {name}, new_rows={len(new_df):,}")
            else:
                log("current_week_only: WARN no refreshed article_day_fact rows; PDF will be rebuilt from existing cache only")
        except Exception as exc:
            log(f"current_week_only: WARN limited refresh failed, using existing cache only: {exc}")
        pdf_path = local_dir / sales_pdf_report_name(outputs)
        pdf_created = generate_management_pdf(outputs, pdf_path)
        if not pdf_created or not pdf_path.exists():
            raise RuntimeError(f"PDF не создан: {pdf_path}")
        log(f"Saved local PDF current-week-only: {pdf_path}")
        if storage.is_s3:
            storage.write_bytes(f"{OUT_DIR}/{pdf_path.name}", pdf_path.read_bytes())
            log(f"Saved: {OUT_DIR}/{pdf_path.name}")
            trace_path = pdf_path.parent / PDF_CALC_TRACE_NAME
            if trace_path.exists():
                storage.write_bytes(f"{OUT_DIR}/{trace_path.name}", trace_path.read_bytes())
                log(f"Saved: {OUT_DIR}/{trace_path.name}")
        if args.send_telegram:
            caption = f"Отчет по продажам Влад {report_date_label_ru(outputs)}"
            if not send_telegram_report(outputs, pdf_path, caption, storage=storage):
                raise RuntimeError("Telegram: отчет не отправлен")
        log("Done")
        return
    if args.pdf_only:
        if args.no_pdf:
            raise SystemExit("Нельзя одновременно использовать --pdf-only и --no-pdf")
        local_dir.mkdir(parents=True, exist_ok=True)
        outputs = load_existing_outputs_for_pdf(storage, local_dir, args.reports_root, args.store, diagnostics)
        pdf_path = local_dir / sales_pdf_report_name(outputs)
        pdf_created = generate_management_pdf(outputs, pdf_path)
        if not pdf_created or not pdf_path.exists():
            raise RuntimeError(f"PDF не создан: {pdf_path}")
        log(f"Saved local PDF: {pdf_path}")
        if storage.is_s3:
            storage.write_bytes(f"{OUT_DIR}/{pdf_path.name}", pdf_path.read_bytes())
            log(f"Saved: {OUT_DIR}/{pdf_path.name}")
            audit_path = pdf_path.parent / PRODUCT_GROUP_AUDIT_NAME
            if audit_path.exists():
                storage.write_bytes(f"{OUT_DIR}/{audit_path.name}", audit_path.read_bytes())
                log(f"Saved: {OUT_DIR}/{audit_path.name}")
            trace_path = pdf_path.parent / PDF_CALC_TRACE_NAME
            if trace_path.exists():
                storage.write_bytes(f"{OUT_DIR}/{trace_path.name}", trace_path.read_bytes())
                log(f"Saved: {OUT_DIR}/{trace_path.name}")
        if args.send_telegram:
            caption = f"Отчет по продажам Влад {report_date_label_ru(outputs)}"
            if not send_telegram_report(outputs, pdf_path, caption, storage=storage):
                raise RuntimeError("Telegram: отчет не отправлен")
        log("Done")
        return
    loader = Loader(storage, args.reports_root, args.store, diagnostics)
    pack = loader.load_all()
    builder = AnalyticsBuilder(pack)
    outputs = builder.build_all()
    outputs["abc_weekly"] = builder.pack.abc_weekly
    outputs["abc_monthly"] = builder.pack.abc_monthly
    factor_outputs = build_factor_outputs(builder, outputs)
    outputs.update(factor_outputs)
    local_dir = Path(args.root) / OUT_DIR
    paths = export_outputs(outputs, local_dir)
    # Отдельный техфайл по денежному факторному мосту, чтобы не ломать основные Excel-структуры.
    factor_path = local_dir / FACTOR_REPORT_NAME
    write_factor_report(factor_path, factor_outputs)
    paths.append(factor_path)

    # Save heavy Excel outputs before PDF generation. If the PDF block fails, the 20-minute
    # source refresh is not lost and a follow-up --pdf-only/current-week run can reuse the cache.
    pre_pdf_paths = list(paths)
    if storage.is_s3:
        for p in pre_pdf_paths:
            storage.write_bytes(f"{OUT_DIR}/{p.name}", p.read_bytes())
            log(f"Saved pre-PDF: {OUT_DIR}/{p.name}")

    pdf_path = local_dir / sales_pdf_report_name(outputs)
    if not args.no_pdf:
        global builder_global_for_pdf
        builder_global_for_pdf = builder
        pdf_created = generate_management_pdf(outputs, pdf_path)
        if pdf_created:
            paths.append(pdf_path)
            audit_path = pdf_path.parent / PRODUCT_GROUP_AUDIT_NAME
            if audit_path.exists():
                paths.append(audit_path)
            trace_path = pdf_path.parent / PDF_CALC_TRACE_NAME
            if trace_path.exists():
                paths.append(trace_path)
    log(f"Saved local copies: {local_dir}")
    # Save PDF/audit/trace after PDF generation; Excel files were already saved pre-PDF.
    if storage.is_s3:
        already_saved = {str(p.resolve()) for p in pre_pdf_paths}
        for p in paths:
            if str(p.resolve()) in already_saved:
                continue
            storage.write_bytes(f"{OUT_DIR}/{p.name}", p.read_bytes())
            log(f"Saved: {OUT_DIR}/{p.name}")
    if args.send_telegram and pdf_path.exists():
        caption = f"Отчет по продажам Влад {report_date_label_ru(outputs)}"
        if not send_telegram_report(outputs, pdf_path, caption, storage=storage):
            raise RuntimeError("Telegram: отчет не отправлен")
    log("Done")


if __name__ == "__main__":
    main()
