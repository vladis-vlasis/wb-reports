#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TOPFACE: контроль поставок FBW и документов Wildberries.

Файл в репозитории: wb_supply_documents.py
Версия: WB_SUPPLY_DOCUMENTS_TOPFACE_V9_20260731

Что делает скрипт:
1. Получает все поставки FBW за скользящий период через Supplies API.
2. Для поставок получает детали и товары, считает:
   - заявлено;
   - принято;
   - излишки;
   - недостачу;
   - склад и статус.
3. Получает документы через Documents API:
   - УПД по маркировке;
   - акты приёмки;
   - акты сверки, если такая категория доступна в кабинете.
4. Сохраняет файлы и сводки в Yandex Object Storage:
   Документы по поставкам/TOPFACE/...
5. Сопоставляет УПД с поставкой сначала по точному supplyID из serviceName,
   затем — резервно по складу, дате и количеству.
6. Позволяет выбрать отправку в Telegram:
   - сообщение и XML УПД;
   - только сообщение без файла.
7. Позволяет отправить поставки только с изменениями либо за выбранный период.

GitHub Secrets / REPORT_ENV:
- WB_PROMO_KEY_TOPFACE         существующий токен TOPFACE; должен иметь категории "Поставки" и "Документы"
- YC_ACCESS_KEY_ID
- YC_SECRET_ACCESS_KEY
- YC_BUCKET_NAME
- YC_ENDPOINT_URL              необязательно, по умолчанию storage.yandexcloud.net
- TELEGRAM_BOT_TOKEN
- TELEGRAM_CHAT_ID
- TELEGRAM_MESSAGE_THREAD_ID   необязательно

Скрипт не загружает документы в Диадок и ничего не подписывает.
"""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import hashlib
import io
import json
import os
import re
import sys
import tempfile
import time
import traceback
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from dataclasses import asdict, dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path, PurePosixPath
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from zoneinfo import ZoneInfo
import xml.etree.ElementTree as ET

import boto3
import requests
from botocore.exceptions import ClientError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


VERSION = "WB_SUPPLY_DOCUMENTS_TOPFACE_V11_20260801"
STORE_NAME = "TOPFACE"
MSK = ZoneInfo("Europe/Moscow")

SUPPLIES_LIST_URL = "https://supplies-api.wildberries.ru/api/v1/supplies"
SUPPLY_DETAILS_URL = "https://supplies-api.wildberries.ru/api/v1/supplies/{supply_id}"
SUPPLY_GOODS_URL = "https://supplies-api.wildberries.ru/api/v1/supplies/{supply_id}/goods"
DOCUMENTS_LIST_URL = "https://documents-api.wildberries.ru/api/v1/documents/list"
DOCUMENT_DOWNLOAD_URL = "https://documents-api.wildberries.ru/api/v1/documents/download"

DEFAULT_ROOT_PREFIX = "Документы по поставкам/TOPFACE"
REGISTRY_RELATIVE_KEY = "_служебные файлы/registry.json"
LAST_RUN_RELATIVE_KEY = "_служебные файлы/last_run.json"
CURRENT_XLSX_RELATIVE_KEY = "_служебные файлы/Сводка_поставок_TOPFACE.xlsx"

STATUS_NAMES = {
    1: "Не запланировано",
    2: "Запланировано",
    3: "Отгрузка разрешена",
    4: "Идёт приёмка",
    5: "Принято",
    6: "Отгружено на воротах",
}
ALL_STATUS_IDS = [1, 2, 3, 4, 5, 6]
DEFAULT_NOTIFY_STATUS_IDS = {4, 5, 6}
DEFAULT_TELEGRAM_MIN_SUPPLY_QTY = 100

TELEGRAM_TOKEN_ALIASES = (
    "TELEGRAM_BOT_TOKEN",
    "TG_BOT_TOKEN",
    "TELEGRAM_TOKEN",
    "BOT_TOKEN",
    "WB_TELEGRAM_BOT_TOKEN",
    "TOPFACE_TELEGRAM_BOT_TOKEN",
)
TELEGRAM_CHAT_ALIASES = (
    "TELEGRAM_CHAT_ID",
    "TG_CHAT_ID",
    "CHAT_ID",
    "WB_TELEGRAM_CHAT_ID",
    "TOPFACE_TELEGRAM_CHAT_ID",
)
TELEGRAM_THREAD_ALIASES = (
    "TELEGRAM_MESSAGE_THREAD_ID",
    "TELEGRAM_THREAD_ID",
    "TG_MESSAGE_THREAD_ID",
    "TG_THREAD_ID",
    "MESSAGE_THREAD_ID",
    "THREAD_ID",
    "WB_TELEGRAM_MESSAGE_THREAD_ID",
    "TOPFACE_TELEGRAM_MESSAGE_THREAD_ID",
)
WB_TOKEN_ALIASES = (
    "WB_PROMO_KEY_TOPFACE",
    "WB_API_TOKEN",
    "TOPFACE_WB_API_TOKEN",
    "WB_TOKEN",
    "WILDBERRIES_API_TOKEN",
)

WAREHOUSE_RULES: Tuple[Tuple[str, str], ...] = (
    ("КОЛЕДИНО", "Коледино"),
    ("ЭЛЕКТРОСТАЛ", "Электросталь"),
    ("БЕЛАЯ ДАЧ", "Белая Дача"),
    ("ВЕШК", "Вёшки"),
    ("ВЁШК", "Вёшки"),
    ("РЯЗАН", "Рязань"),
    ("ТУЛ", "Тула"),
    ("АЛЕКСИН", "Тула"),
    ("ВЛАДИМИР", "Владимир"),
    ("КОТОВСК", "Котовск"),
    ("ВОРОНЕЖ", "Воронеж"),
    ("КРАСНОДАР", "Краснодар"),
    ("НЕВИННОМЫССК", "Невинномысск"),
    ("НЕВИННОМЫСК", "Невинномысск"),
    ("ВОЛГОГРАД", "Волгоград"),
    ("РОСТОВ", "Ростов/Аксай"),
    ("АКСАЙ", "Ростов/Аксай"),
    ("КАЗАН", "Казань"),
    ("ПЕНЗ", "Пенза"),
    ("САРАПУЛ", "Сарапул"),
    ("НОВОСЕМЕЙКИНО", "Новосемейкино"),
    ("САМАР", "Самара"),
    ("ШУШАР", "СПБ Шушары"),
    ("УТКИН", "СПБ Уткина Заводь"),
    ("САНКТ", "Санкт-Петербург"),
    ("ПЕТЕРБУРГ", "Санкт-Петербург"),
    ("ЕКАТЕРИНБУРГ", "Екатеринбург"),
    ("ПЕРСПЕКТИВН", "Екатеринбург"),
    ("ЧЕЛЯБИНСК", "Челябинск"),
    ("ПЕРМ", "Пермь"),
    ("НОВОСИБИРСК", "Новосибирск"),
    ("КРАСНОЯРСК", "Красноярск"),
    ("КЕМЕРОВО", "Кемерово"),
    ("ХАБАРОВСК", "Хабаровск"),
)


# ---------------------------------------------------------------------------
# Общие функции
# ---------------------------------------------------------------------------

def log(message: str) -> None:
    stamp = dt.datetime.now(MSK).strftime("%Y-%m-%d %H:%M:%S MSK")
    print(f"[{stamp}] {message}", flush=True)


def fail(message: str, code: int = 1) -> None:
    print(f"ERROR: {message}", file=sys.stderr, flush=True)
    raise SystemExit(code)


def now_msk() -> dt.datetime:
    return dt.datetime.now(MSK)


def today_msk() -> dt.date:
    return now_msk().date()


def parse_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    text = str(value).strip().lower()
    if not text:
        return default
    return text in {"1", "true", "yes", "y", "on", "да", "+"}


def to_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default
    try:
        text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
        if not text or text.lower() in {"nan", "none", "null"}:
            return default
        return int(Decimal(text))
    except (InvalidOperation, ValueError, TypeError):
        return default


def to_decimal(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    if value is None:
        return default
    try:
        text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
        if not text or text.lower() in {"nan", "none", "null"}:
            return default
        return Decimal(text)
    except (InvalidOperation, ValueError, TypeError):
        return default


def parse_datetime(value: Any) -> Optional[dt.datetime]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value if value.tzinfo else value.replace(tzinfo=MSK)
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time.min, tzinfo=MSK)
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return None
    try:
        parsed = dt.datetime.fromisoformat(text.replace("Z", "+00:00"))
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=MSK)
    except ValueError:
        pass
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(text[:19], fmt).replace(tzinfo=MSK)
        except ValueError:
            continue
    return None


def parse_date(value: Any) -> Optional[dt.date]:
    parsed = parse_datetime(value)
    return parsed.date() if parsed else None


def iso_or_empty(value: Optional[dt.datetime]) -> str:
    return value.isoformat() if value else ""


def first_nonempty(*values: Any) -> Any:
    for value in values:
        if value is not None and str(value).strip() not in {"", "nan", "None", "null"}:
            return value
    return ""


def normalize_text(value: Any) -> str:
    text = str(value or "").replace("ё", "е").replace("Ё", "Е")
    text = re.sub(r"[^0-9A-Za-zА-Яа-я]+", " ", text).upper()
    return re.sub(r"\s+", " ", text).strip()


def canonical_warehouse(value: Any) -> str:
    normalized = normalize_text(value)
    if not normalized:
        return ""
    for key, result in WAREHOUSE_RULES:
        if key in normalized:
            return result
    return str(value or "").strip()


def safe_filename(value: str, fallback: str = "document.bin") -> str:
    value = PurePosixPath(str(value or "")).name
    value = value.replace("\x00", "")
    value = re.sub(r"[\\/:*?\"<>|]+", "_", value)
    value = re.sub(r"\s+", " ", value).strip(" .")
    return value[:220] or fallback


def join_key(*parts: str) -> str:
    return "/".join(str(part).strip("/") for part in parts if str(part).strip("/"))


def json_dumps(data: Any) -> bytes:
    return json.dumps(data, ensure_ascii=False, indent=2, default=str).encode("utf-8")


def stable_hash(data: Any) -> str:
    payload = json.dumps(data, ensure_ascii=False, sort_keys=True, default=str, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def load_report_env() -> None:
    """Загружает KEY=VALUE из общего многострочного секрета REPORT_ENV."""
    raw = os.getenv("REPORT_ENV", "") or os.getenv("WB_REPORT_ENV", "") or ""
    if not raw.strip():
        return
    normalized = raw.replace("\r\n", "\n").replace("\r", "\n")
    if "\n" not in normalized and "\\n" in normalized:
        normalized = normalized.replace("\\n", "\n")
    loaded: List[str] = []
    for raw_line in normalized.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[7:].strip()
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip().lstrip("\ufeff")
        value = value.strip().strip('"').strip("'")
        if key and not os.getenv(key, "").strip():
            os.environ[key] = value
            loaded.append(key)
    if loaded:
        log(f"REPORT_ENV: загружены переменные: {', '.join(loaded)}")


def first_env(names: Sequence[str]) -> Tuple[str, str]:
    for name in names:
        value = os.getenv(name, "")
        if value is not None and str(value).strip():
            return str(value).strip(), name
    return "", ""


# ---------------------------------------------------------------------------
# Object Storage
# ---------------------------------------------------------------------------

class Storage:
    def exists(self, key: str) -> bool:
        raise NotImplementedError

    def read_bytes(self, key: str) -> bytes:
        raise NotImplementedError

    def write_bytes(self, key: str, data: bytes, content_type: str = "application/octet-stream") -> None:
        raise NotImplementedError


class S3Storage(Storage):
    def __init__(self, bucket: str, access_key: str, secret_key: str, endpoint: str):
        self.bucket = bucket
        self.client = boto3.client(
            "s3",
            endpoint_url=endpoint,
            aws_access_key_id=access_key,
            aws_secret_access_key=secret_key,
            region_name=os.getenv("AWS_DEFAULT_REGION", "ru-central1"),
        )

    def exists(self, key: str) -> bool:
        try:
            self.client.head_object(Bucket=self.bucket, Key=key)
            return True
        except ClientError as exc:
            code = str(exc.response.get("Error", {}).get("Code", ""))
            if code in {"404", "NoSuchKey", "NotFound"}:
                return False
            raise

    def read_bytes(self, key: str) -> bytes:
        return self.client.get_object(Bucket=self.bucket, Key=key)["Body"].read()

    def write_bytes(self, key: str, data: bytes, content_type: str = "application/octet-stream") -> None:
        self.client.put_object(Bucket=self.bucket, Key=key, Body=data, ContentType=content_type)


class LocalStorage(Storage):
    def __init__(self, root: Path):
        self.root = root

    def _path(self, key: str) -> Path:
        return self.root / Path(key)

    def exists(self, key: str) -> bool:
        return self._path(key).exists()

    def read_bytes(self, key: str) -> bytes:
        return self._path(key).read_bytes()

    def write_bytes(self, key: str, data: bytes, content_type: str = "application/octet-stream") -> None:
        path = self._path(key)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(data)


def make_storage(local_root: Optional[str] = None) -> Storage:
    load_report_env()
    access, access_name = first_env((
        "YC_ACCESS_KEY_ID", "AWS_ACCESS_KEY_ID", "S3_ACCESS_KEY_ID", "ACCESS_KEY_ID",
    ))
    secret, secret_name = first_env((
        "YC_SECRET_ACCESS_KEY", "AWS_SECRET_ACCESS_KEY", "S3_SECRET_ACCESS_KEY", "SECRET_ACCESS_KEY",
    ))
    bucket, bucket_name = first_env((
        "YC_BUCKET_NAME", "S3_BUCKET", "S3_BUCKET_NAME", "BUCKET_NAME",
    ))
    endpoint, endpoint_name = first_env((
        "YC_ENDPOINT_URL", "S3_ENDPOINT_URL", "AWS_ENDPOINT_URL",
    ))
    endpoint = endpoint or "https://storage.yandexcloud.net"
    endpoint_name = endpoint_name or "default"

    if access and secret and bucket:
        log(
            f"Object Storage: bucket={bucket}, endpoint={endpoint}; "
            f"env access={access_name}, secret={secret_name}, bucket={bucket_name}, endpoint={endpoint_name}"
        )
        return S3Storage(bucket, access, secret, endpoint)

    if local_root:
        path = Path(local_root).resolve()
        log(f"Object Storage: локальный режим {path}")
        return LocalStorage(path)

    missing = []
    if not access:
        missing.append("YC_ACCESS_KEY_ID")
    if not secret:
        missing.append("YC_SECRET_ACCESS_KEY")
    if not bucket:
        missing.append("YC_BUCKET_NAME")
    fail("Не настроено Object Storage: " + ", ".join(missing))


# ---------------------------------------------------------------------------
# Telegram
# ---------------------------------------------------------------------------

def telegram_credentials() -> Tuple[str, str, str]:
    load_report_env()
    token, _ = first_env(TELEGRAM_TOKEN_ALIASES)
    chat_id, _ = first_env(TELEGRAM_CHAT_ALIASES)
    thread_id, _ = first_env(TELEGRAM_THREAD_ALIASES)
    return token, chat_id, thread_id


def telegram_request(method: str, fields: Dict[str, str], file_field: Optional[Tuple[str, str, bytes, str]] = None) -> bool:
    token, chat_id, thread_id = telegram_credentials()
    if not token or not chat_id:
        log("Telegram: не заданы TELEGRAM_BOT_TOKEN или TELEGRAM_CHAT_ID")
        return False
    fields = dict(fields)
    fields["chat_id"] = chat_id
    if thread_id:
        fields["message_thread_id"] = thread_id
    url = f"https://api.telegram.org/bot{token}/{method}"

    try:
        if file_field is None:
            data = urllib.parse.urlencode(fields).encode("utf-8")
            request = urllib.request.Request(
                url,
                data=data,
                headers={"Content-Type": "application/x-www-form-urlencoded"},
            )
        else:
            field_name, filename, file_bytes, content_type = file_field
            boundary = "----WBSupply" + hashlib.md5(os.urandom(16)).hexdigest()
            body = bytearray()
            for name, value in fields.items():
                body.extend(f"--{boundary}\r\n".encode())
                body.extend(f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode())
                body.extend(str(value).encode("utf-8"))
                body.extend(b"\r\n")
            body.extend(f"--{boundary}\r\n".encode())
            body.extend(
                f'Content-Disposition: form-data; name="{field_name}"; filename="{safe_filename(filename)}"\r\n'.encode(
                    "utf-8"
                )
            )
            body.extend(f"Content-Type: {content_type}\r\n\r\n".encode())
            body.extend(file_bytes)
            body.extend(f"\r\n--{boundary}--\r\n".encode())
            request = urllib.request.Request(
                url,
                data=bytes(body),
                headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
            )

        with urllib.request.urlopen(request, timeout=90) as response:
            response_body = response.read().decode("utf-8", errors="replace")[:500]
            ok = 200 <= response.status < 300
            log(f"Telegram {method}: status={response.status}, ok={ok}, response={response_body}")
            return ok
    except (urllib.error.URLError, TimeoutError, OSError) as exc:
        log(f"Telegram {method}: ошибка: {exc}")
        return False


def split_telegram_text(text: str, limit: int = 3900) -> List[str]:
    if len(text) <= limit:
        return [text]
    chunks: List[str] = []
    current: List[str] = []
    current_len = 0
    for line in text.splitlines():
        line_len = len(line) + 1
        if current and current_len + line_len > limit:
            chunks.append("\n".join(current))
            current = []
            current_len = 0
        current.append(line)
        current_len += line_len
    if current:
        chunks.append("\n".join(current))
    return chunks


def send_telegram_message(text: str) -> bool:
    ok = True
    for chunk in split_telegram_text(text):
        ok = telegram_request("sendMessage", {"text": chunk}) and ok
        time.sleep(0.4)
    return ok


def send_telegram_document(filename: str, data: bytes, caption: str = "") -> bool:
    content_type = "application/octet-stream"
    lower = filename.lower()
    if lower.endswith(".xml"):
        content_type = "application/xml"
    elif lower.endswith(".zip"):
        content_type = "application/zip"
    elif lower.endswith(".xlsx"):
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return telegram_request(
        "sendDocument",
        {"caption": caption[:1000]},
        file_field=("document", filename, data, content_type),
    )


# ---------------------------------------------------------------------------
# HTTP-клиент WB с лимитами и повторами
# ---------------------------------------------------------------------------

class RateLimiter:
    def __init__(self, min_interval_seconds: float):
        self.min_interval_seconds = min_interval_seconds
        self.last_request_at = 0.0

    def wait(self) -> None:
        elapsed = time.monotonic() - self.last_request_at
        wait_for = self.min_interval_seconds - elapsed
        if wait_for > 0:
            time.sleep(wait_for)
        self.last_request_at = time.monotonic()


class WBClient:
    def __init__(self, token: str):
        self.token = token
        self.session = requests.Session()
        self.supplies_limiter = RateLimiter(float(os.getenv("WB_SUPPLIES_INTERVAL_SECONDS", "2.05")))
        self.documents_limiter = RateLimiter(float(os.getenv("WB_DOCUMENTS_INTERVAL_SECONDS", "10.1")))
        self.retry_delays = [5, 10, 20, 40, 80, 120]

    def request_json(
        self,
        method: str,
        url: str,
        *,
        params: Optional[Dict[str, Any]] = None,
        json_body: Any = None,
        limiter: Optional[RateLimiter] = None,
        allow_204: bool = True,
    ) -> Any:
        headers = {"Authorization": self.token}
        if json_body is not None:
            headers["Content-Type"] = "application/json"

        last_error = ""
        for attempt in range(1, len(self.retry_delays) + 2):
            if limiter:
                limiter.wait()
            try:
                response = self.session.request(
                    method,
                    url,
                    headers=headers,
                    params=params,
                    json=json_body,
                    timeout=180,
                )
                if response.status_code == 204 and allow_204:
                    return None
                if response.status_code == 429:
                    retry_after = response.headers.get("Retry-After", "")
                    delay = int(retry_after) if retry_after.isdigit() else self.retry_delays[min(attempt - 1, len(self.retry_delays) - 1)]
                    log(f"WB API 429: {method} {url}; попытка {attempt}; ожидание {delay} сек")
                    time.sleep(delay)
                    continue
                if response.status_code in {500, 502, 503, 504}:
                    delay = self.retry_delays[min(attempt - 1, len(self.retry_delays) - 1)]
                    log(f"WB API {response.status_code}: повтор через {delay} сек")
                    time.sleep(delay)
                    continue
                if response.status_code >= 400:
                    raise RuntimeError(
                        f"WB API HTTP {response.status_code}: {method} {url}; ответ={response.text[:1200]}"
                    )
                if not response.text.strip():
                    return None
                return response.json()
            except (requests.RequestException, ValueError, RuntimeError) as exc:
                last_error = str(exc)
                if attempt <= len(self.retry_delays):
                    delay = self.retry_delays[attempt - 1]
                    log(f"WB API: ошибка, повтор через {delay} сек: {exc}")
                    time.sleep(delay)
                    continue
                raise RuntimeError(f"WB API: запрос не выполнен: {last_error}") from exc
        raise RuntimeError(f"WB API: запрос не выполнен: {last_error}")

    @staticmethod
    def unwrap_data(data: Any) -> Any:
        if isinstance(data, dict) and "data" in data:
            return data["data"]
        if isinstance(data, dict) and "result" in data:
            return data["result"]
        return data

    @staticmethod
    def normalize_list(data: Any, keys: Sequence[str]) -> List[Dict[str, Any]]:
        data = WBClient.unwrap_data(data)
        if isinstance(data, list):
            return [item for item in data if isinstance(item, dict)]
        if isinstance(data, dict):
            for key in keys:
                value = data.get(key)
                if isinstance(value, list):
                    return [item for item in value if isinstance(item, dict)]
        return []

    def list_supplies(self, date_from: dt.date, date_to: dt.date) -> List[Dict[str, Any]]:
        """Берёт объединение по createDate, supplyDate и factDate, чтобы не терять поставки."""
        by_id: Dict[str, Dict[str, Any]] = {}
        for date_type in ("createDate", "supplyDate", "factDate"):
            offset = 0
            while True:
                body = {
                    "dates": [{"from": date_from.isoformat(), "till": date_to.isoformat(), "type": date_type}],
                    "statusIDs": ALL_STATUS_IDS,
                }
                data = self.request_json(
                    "POST",
                    SUPPLIES_LIST_URL,
                    params={"limit": 1000, "offset": offset},
                    json_body=body,
                    limiter=self.supplies_limiter,
                )
                rows = self.normalize_list(data, ("supplies", "items", "rows"))
                log(f"Поставки: date_type={date_type}, offset={offset}, получено={len(rows)}")
                for row in rows:
                    supply_id = supply_id_from(row)
                    if supply_id:
                        by_id[str(supply_id)] = {**by_id.get(str(supply_id), {}), **row}
                if len(rows) < 1000:
                    break
                offset += 1000
        result = list(by_id.values())
        result.sort(key=lambda item: parse_datetime(first_nonempty(item.get("updatedDate"), item.get("factDate"), item.get("supplyDate"), item.get("createDate"))) or dt.datetime.min.replace(tzinfo=MSK), reverse=True)
        return result

    def get_supply_details(self, supply_id: int) -> Dict[str, Any]:
        data = self.request_json(
            "GET",
            SUPPLY_DETAILS_URL.format(supply_id=supply_id),
            params={"isPreorderID": "false"},
            limiter=self.supplies_limiter,
        )
        data = self.unwrap_data(data)
        return data if isinstance(data, dict) else {}

    def get_supply_goods(self, supply_id: int) -> List[Dict[str, Any]]:
        result: List[Dict[str, Any]] = []
        offset = 0
        while True:
            data = self.request_json(
                "GET",
                SUPPLY_GOODS_URL.format(supply_id=supply_id),
                params={"limit": 1000, "offset": offset, "isPreorderID": "false"},
                limiter=self.supplies_limiter,
            )
            rows = self.normalize_list(data, ("goods", "items", "products", "rows"))
            result.extend(rows)
            if len(rows) < 1000:
                break
            offset += 1000
        return result

    def list_documents(self, date_from: dt.date, date_to: dt.date) -> List[Dict[str, Any]]:
        result: List[Dict[str, Any]] = []
        offset = 0
        while True:
            data = self.request_json(
                "GET",
                DOCUMENTS_LIST_URL,
                params={
                    "locale": "ru",
                    "beginTime": date_from.isoformat(),
                    "endTime": date_to.isoformat(),
                    "sort": "date",
                    "order": "desc",
                    "limit": 50,
                    "offset": offset,
                },
                limiter=self.documents_limiter,
            )
            raw = self.unwrap_data(data)
            rows = raw.get("documents", []) if isinstance(raw, dict) else []
            rows = [item for item in rows if isinstance(item, dict)]
            result.extend(rows)
            log(f"Документы: offset={offset}, получено={len(rows)}")
            if len(rows) < 50:
                break
            offset += 50
        return result

    def download_document(self, service_name: str, extension: str) -> Tuple[str, str, bytes]:
        data = self.request_json(
            "GET",
            DOCUMENT_DOWNLOAD_URL,
            params={"serviceName": service_name, "extension": extension},
            limiter=self.documents_limiter,
        )
        raw = self.unwrap_data(data)
        if not isinstance(raw, dict):
            raise RuntimeError(f"Документ {service_name}: неожиданный ответ {type(raw)}")
        encoded = raw.get("document")
        if not encoded:
            raise RuntimeError(f"Документ {service_name}: в ответе нет поля document")
        encoded = str(encoded)
        encoded += "=" * ((4 - len(encoded) % 4) % 4)
        content = base64.b64decode(encoded)
        file_name = safe_filename(str(raw.get("fileName") or f"{service_name}.{extension}"))
        response_extension = str(raw.get("extension") or extension or Path(file_name).suffix.lstrip(".") or "bin")
        if "." not in file_name:
            file_name = f"{file_name}.{response_extension}"
        return file_name, response_extension, content


# ---------------------------------------------------------------------------
# Модели поставок и расчёты
# ---------------------------------------------------------------------------

@dataclass
class SupplySummary:
    supply_id: int
    preorder_id: str = ""
    status_id: int = 0
    status_name: str = ""
    warehouse: str = ""
    warehouse_plan: str = ""
    warehouse_actual: str = ""
    warehouse_transit: str = ""
    create_date: str = ""
    supply_date: str = ""
    fact_date: str = ""
    updated_date: str = ""
    planned_qty: int = 0
    unloading_qty: int = 0
    accepted_qty: int = 0
    ready_qty: int = 0
    excess_qty: int = 0
    shortage_qty: int = 0
    goods_count: int = 0
    discrepancy_articles: int = 0
    source: str = "Supplies API"
    state_hash: str = ""
    upd_documents: List[Dict[str, Any]] = field(default_factory=list)

    def report_comment(self) -> str:
        if self.status_id == 5:
            if self.excess_qty == 0 and self.shortage_qty == 0 and self.accepted_qty == self.planned_qty:
                return f"Принято без расхождений {self.accepted_qty} шт."
            return (
                f"Принято {self.accepted_qty}/{self.planned_qty} шт. "
                f"Излишки: {self.excess_qty} шт. Недостача: {self.shortage_qty} шт."
            )
        if self.status_id in {4, 6}:
            return (
                f"{self.status_name}. Заявлено {self.planned_qty} шт., "
                f"выгружено {self.unloading_qty} шт., принято {self.accepted_qty} шт."
            )
        return f"Статус: {self.status_name}. Заявлено {self.planned_qty} шт."


def supply_id_from(item: Dict[str, Any]) -> Optional[int]:
    value = first_nonempty(item.get("supplyID"), item.get("supplyId"), item.get("id"))
    result = to_int(value, 0)
    return result or None


def summary_from_supply(
    shallow: Dict[str, Any],
    details: Dict[str, Any],
    goods: List[Dict[str, Any]],
    previous: Optional[Dict[str, Any]] = None,
) -> Tuple[SupplySummary, List[Dict[str, Any]]]:
    """Собирает сводку поставки.

    Итоговые количества берутся из метода деталей поставки — это агрегаты WB по
    всей поставке. Метод goods используется для детализации расхождений по SKU.
    Такой порядок защищает от частичных/нестабильных товарных строк и не меняет
    общий итог на количество строк или коробов.
    """
    merged = {**shallow, **details}
    supply_id = supply_id_from(merged)
    if not supply_id:
        raise ValueError("В данных поставки отсутствует supplyID")

    status_id = to_int(first_nonempty(merged.get("statusID"), shallow.get("statusID")), 0)
    warehouse_plan = str(first_nonempty(merged.get("warehouseName"), shallow.get("warehouseName"))).strip()
    warehouse_actual = str(first_nonempty(merged.get("actualWarehouseName"), shallow.get("actualWarehouseName"))).strip()
    warehouse_transit = str(first_nonempty(merged.get("transitWarehouseName"), shallow.get("transitWarehouseName"))).strip()
    warehouse = canonical_warehouse(first_nonempty(warehouse_actual, warehouse_plan, warehouse_transit))

    item_rows: List[Dict[str, Any]] = []
    goods_planned = goods_accepted = goods_unloading = goods_ready = 0
    excess_qty = shortage_qty = discrepancy_articles = 0
    for item in goods:
        planned = to_int(item.get("quantity"), 0)
        accepted = to_int(item.get("acceptedQuantity"), 0)
        unloading = to_int(item.get("unloadingQuantity"), 0)
        ready = to_int(item.get("readyForSaleQuantity"), 0)
        excess = max(accepted - planned, 0)
        shortage = max(planned - accepted, 0)
        goods_planned += planned
        goods_accepted += accepted
        goods_unloading += unloading
        goods_ready += ready
        excess_qty += excess
        shortage_qty += shortage
        if excess or shortage:
            discrepancy_articles += 1
            item_rows.append({
                "Номер поставки": supply_id,
                "Склад": warehouse,
                "Артикул продавца": first_nonempty(item.get("vendorCode"), item.get("supplierArticle"), item.get("article")),
                "nmID": first_nonempty(item.get("nmID"), item.get("nmId")),
                "Баркод": first_nonempty(item.get("barcode"), item.get("barCode")),
                "Размер": first_nonempty(item.get("techSize"), item.get("size")),
                "Цвет": first_nonempty(item.get("color")),
                "Нужен КИЗ": bool(item.get("needKiz")),
                "Заявлено": planned,
                "Принято": accepted,
                "Выгружено": unloading,
                "Готово к продаже": ready,
                "Излишек": excess,
                "Недостача": shortage,
            })

    def aggregate_value(field: str, goods_value: int, previous_field: str) -> int:
        if field in details and details.get(field) is not None:
            return to_int(details.get(field), 0)
        if field in merged and merged.get(field) is not None:
            return to_int(merged.get(field), 0)
        if goods:
            return goods_value
        return to_int((previous or {}).get(previous_field), 0)

    planned_qty = aggregate_value("quantity", goods_planned, "planned_qty")
    accepted_qty = aggregate_value("acceptedQuantity", goods_accepted, "accepted_qty")
    unloading_qty = aggregate_value("unloadingQuantity", goods_unloading, "unloading_qty")
    ready_qty = aggregate_value("readyForSaleQuantity", goods_ready, "ready_qty")

    # Если товарной детализации нет, расхождение считаем по общим итогам.
    if not goods:
        excess_qty = max(accepted_qty - planned_qty, 0)
        shortage_qty = max(planned_qty - accepted_qty, 0)
    elif not item_rows and accepted_qty != planned_qty:
        # Редкий случай: агрегат уже обновился, а goods ещё нет. Не скрываем разницу.
        excess_qty = max(accepted_qty - planned_qty, 0)
        shortage_qty = max(planned_qty - accepted_qty, 0)

    summary = SupplySummary(
        supply_id=supply_id,
        preorder_id=str(first_nonempty(merged.get("preorderID"), merged.get("preorderId"))),
        status_id=status_id,
        status_name=STATUS_NAMES.get(status_id, f"Статус {status_id}"),
        warehouse=warehouse,
        warehouse_plan=warehouse_plan,
        warehouse_actual=warehouse_actual,
        warehouse_transit=warehouse_transit,
        create_date=iso_or_empty(parse_datetime(merged.get("createDate"))),
        supply_date=iso_or_empty(parse_datetime(merged.get("supplyDate"))),
        fact_date=iso_or_empty(parse_datetime(merged.get("factDate"))),
        updated_date=iso_or_empty(parse_datetime(merged.get("updatedDate"))),
        planned_qty=planned_qty,
        unloading_qty=unloading_qty,
        accepted_qty=accepted_qty,
        ready_qty=ready_qty,
        excess_qty=excess_qty,
        shortage_qty=shortage_qty,
        goods_count=len(goods),
        discrepancy_articles=discrepancy_articles,
        upd_documents=list((previous or {}).get("upd_documents", [])),
    )
    summary.state_hash = stable_hash({
        "status_id": summary.status_id,
        "warehouse": summary.warehouse,
        "planned_qty": summary.planned_qty,
        "unloading_qty": summary.unloading_qty,
        "accepted_qty": summary.accepted_qty,
        "ready_qty": summary.ready_qty,
        "excess_qty": summary.excess_qty,
        "shortage_qty": summary.shortage_qty,
        "updated_date": summary.updated_date,
    })
    return summary, item_rows

def shallow_summary(item: Dict[str, Any], previous: Optional[Dict[str, Any]] = None) -> SupplySummary:
    supply_id = supply_id_from(item)
    if not supply_id:
        raise ValueError("Нет supplyID")
    previous = previous or {}
    status_id = to_int(first_nonempty(item.get("statusID"), previous.get("status_id")), 0)
    warehouse_plan = str(first_nonempty(item.get("warehouseName"), previous.get("warehouse_plan"))).strip()
    warehouse_actual = str(first_nonempty(item.get("actualWarehouseName"), previous.get("warehouse_actual"))).strip()
    warehouse_transit = str(first_nonempty(item.get("transitWarehouseName"), previous.get("warehouse_transit"))).strip()
    summary = SupplySummary(
        supply_id=supply_id,
        preorder_id=str(first_nonempty(item.get("preorderID"), item.get("preorderId"), previous.get("preorder_id"))),
        status_id=status_id,
        status_name=STATUS_NAMES.get(status_id, f"Статус {status_id}"),
        warehouse=canonical_warehouse(first_nonempty(warehouse_actual, warehouse_plan, warehouse_transit, previous.get("warehouse"))),
        warehouse_plan=warehouse_plan or str(previous.get("warehouse_plan", "")),
        warehouse_actual=warehouse_actual or str(previous.get("warehouse_actual", "")),
        warehouse_transit=warehouse_transit or str(previous.get("warehouse_transit", "")),
        create_date=iso_or_empty(parse_datetime(first_nonempty(item.get("createDate"), previous.get("create_date")))),
        supply_date=iso_or_empty(parse_datetime(first_nonempty(item.get("supplyDate"), previous.get("supply_date")))),
        fact_date=iso_or_empty(parse_datetime(first_nonempty(item.get("factDate"), previous.get("fact_date")))),
        updated_date=iso_or_empty(parse_datetime(first_nonempty(item.get("updatedDate"), previous.get("updated_date")))),
        planned_qty=to_int(first_nonempty(item.get("quantity"), previous.get("planned_qty")), 0),
        unloading_qty=to_int(first_nonempty(item.get("unloadingQuantity"), previous.get("unloading_qty")), 0),
        accepted_qty=to_int(first_nonempty(item.get("acceptedQuantity"), previous.get("accepted_qty")), 0),
        ready_qty=to_int(first_nonempty(item.get("readyForSaleQuantity"), previous.get("ready_qty")), 0),
        excess_qty=to_int(previous.get("excess_qty"), 0),
        shortage_qty=to_int(previous.get("shortage_qty"), 0),
        goods_count=to_int(previous.get("goods_count"), 0),
        discrepancy_articles=to_int(previous.get("discrepancy_articles"), 0),
        upd_documents=list(previous.get("upd_documents", [])),
    )
    summary.state_hash = stable_hash({
        "status_id": summary.status_id,
        "warehouse": summary.warehouse,
        "planned_qty": summary.planned_qty,
        "unloading_qty": summary.unloading_qty,
        "accepted_qty": summary.accepted_qty,
        "ready_qty": summary.ready_qty,
        "updated_date": summary.updated_date,
    })
    return summary


# ---------------------------------------------------------------------------
# Документы WB и разбор УПД XML
# ---------------------------------------------------------------------------

@dataclass
class UPDInfo:
    document_number: str = ""
    document_date: str = ""
    acceptance_act_number: str = ""
    warehouse_address: str = ""
    warehouse: str = ""
    total_qty: int = 0
    kiz_count: int = 0
    xml_filename: str = ""
    supply_id: Optional[int] = None
    match_score: int = 0
    match_reason: str = ""


def document_kind(doc: Dict[str, Any]) -> Optional[str]:
    text = normalize_text(" ".join([
        str(doc.get("category", "")),
        str(doc.get("name", "")),
        str(doc.get("serviceName", "")),
    ]))
    if ("УПД" in text and "МАРКИРОВ" in text) or "UPD PO MARKIROVKE" in text:
        return "upd"
    if "АКТ ПРИЕМКИ" in text or "АКТ ПРИЁМКИ" in text or "ACT INCOME" in text:
        return "acceptance"
    if "АКТ СВЕРК" in text or "ОТЧЕТ СВЕРК" in text or "ОТЧЁТ СВЕРК" in text or "RECONCILIATION" in text:
        return "reconciliation"
    return None


def kind_folder(kind: str) -> str:
    return {
        "upd": "УПД по маркировке",
        "acceptance": "Акты приёмки",
        "reconciliation": "Акты сверки/Документы WB",
    }.get(kind, "Прочие документы")


def preferred_extension(doc: Dict[str, Any], kind: str) -> str:
    extensions = [str(value).lower() for value in (doc.get("extensions") or [])]
    order = {
        "upd": ("zip", "xml", "xlsx", "pdf"),
        "acceptance": ("zip", "xlsx", "pdf", "xml"),
        "reconciliation": ("zip", "xlsx", "pdf", "xml"),
    }.get(kind, ("zip", "xlsx", "pdf", "xml"))
    for extension in order:
        if extension in extensions:
            return extension
    return extensions[0] if extensions else "zip"


def content_type_for_name(name: str) -> str:
    lower = name.lower()
    if lower.endswith(".xml"):
        return "application/xml"
    if lower.endswith(".zip"):
        return "application/zip"
    if lower.endswith(".xlsx"):
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if lower.endswith(".pdf"):
        return "application/pdf"
    if lower.endswith(".json"):
        return "application/json"
    return "application/octet-stream"


def extract_files_recursive(name: str, content: bytes, depth: int = 0, max_depth: int = 1) -> List[Tuple[str, bytes]]:
    """Возвращает исходный файл и только верхний уровень обычного ZIP.

    XLSX/DOCX не распаковываются. Вложенные ZIP (например mchd.zip) тоже
    сохраняются целиком, чтобы в Object Storage не появлялись технические XML.
    """
    safe_name = safe_filename(name)
    result: List[Tuple[str, bytes]] = [(safe_name, content)]
    if depth >= max_depth or not safe_name.lower().endswith(".zip"):
        return result
    if not zipfile.is_zipfile(io.BytesIO(content)):
        return result
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            for member in archive.infolist():
                if member.is_dir():
                    continue
                member_name = safe_filename(PurePosixPath(member.filename).name)
                if not member_name:
                    continue
                result.append((member_name, archive.read(member)))
    except (zipfile.BadZipFile, RuntimeError) as exc:
        log(f"ZIP {name}: не удалось распаковать: {exc}")

    deduped: List[Tuple[str, bytes]] = []
    seen: set[Tuple[str, str]] = set()
    for item_name, item_content in result:
        key = (item_name, hashlib.sha256(item_content).hexdigest())
        if key not in seen:
            seen.add(key)
            deduped.append((item_name, item_content))
    return deduped

def local_tag(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def first_element(root: ET.Element, tag_name: str) -> Optional[ET.Element]:
    for element in root.iter():
        if local_tag(element.tag) == tag_name:
            return element
    return None


def parse_upd_xml(xml_content: bytes, xml_filename: str = "") -> UPDInfo:
    root = ET.fromstring(xml_content)
    invoice = first_element(root, "СвСчФакт")
    total = first_element(root, "ВсегоОпл")
    basis = first_element(root, "ОснПер")

    warehouse_address = ""
    for element in root.iter():
        if local_tag(element.tag) == "ГрузПолуч":
            for child in element.iter():
                if local_tag(child.tag) == "АдрИнф":
                    warehouse_address = str(child.attrib.get("АдрТекст", "")).strip()
                    break
            if warehouse_address:
                break

    quantity_sum = Decimal("0")
    kiz_count = 0
    for element in root.iter():
        tag = local_tag(element.tag)
        if tag == "СведТов":
            quantity_sum += to_decimal(element.attrib.get("КолТов"), Decimal("0"))
        elif tag == "КИЗ":
            kiz_count += 1

    total_qty = to_int(total.attrib.get("КолНеттоВс") if total is not None else None, 0)
    if not total_qty:
        total_qty = int(quantity_sum)

    return UPDInfo(
        document_number=str(invoice.attrib.get("НомерДок", "") if invoice is not None else ""),
        document_date=str(invoice.attrib.get("ДатаДок", "") if invoice is not None else ""),
        acceptance_act_number=str(basis.attrib.get("РеквНомерДок", "") if basis is not None else ""),
        warehouse_address=warehouse_address,
        warehouse=canonical_warehouse(warehouse_address),
        total_qty=total_qty,
        kiz_count=kiz_count,
        xml_filename=xml_filename,
    )


def numeric_ids_from_text(*values: Any) -> List[int]:
    result: List[int] = []
    for value in values:
        for match in re.finditer(r"(?<!\d)(\d{6,12})(?!\d)", str(value or "")):
            number = to_int(match.group(1), 0)
            if number and number not in result:
                result.append(number)
    return result


def document_storage_id(service_name: str, response_name: str = "") -> str:
    ids = numeric_ids_from_text(service_name, response_name)
    if ids:
        return str(ids[0])
    return safe_filename(service_name or response_name, "document").replace(".", "_")


def match_upd_to_supply(
    upd: UPDInfo,
    supplies: Iterable[SupplySummary],
    service_name: str = "",
) -> UPDInfo:
    supplies_list = list(supplies)
    supplies_by_id = {item.supply_id: item for item in supplies_list}

    # В документах WB номер после UPD po markirovke-/act-income- и номер
    # основания передачи совпадают с supplyID. Это основной и точный ключ.
    direct_candidates = numeric_ids_from_text(
        upd.acceptance_act_number,
        service_name,
        upd.xml_filename,
        upd.document_number,
    )
    for candidate in direct_candidates:
        if candidate in supplies_by_id:
            upd.supply_id = candidate
            upd.match_score = 200
            upd.match_reason = f"точное совпадение с supplyID={candidate} по номеру документа/акта"
            return upd

    # Резервная эвристика нужна только для редких документов без номера поставки.
    doc_date = parse_date(upd.document_date)
    scored: List[Tuple[int, int, List[str]]] = []
    for supply in supplies_list:
        score = 0
        reasons: List[str] = []
        supply_warehouse = canonical_warehouse(supply.warehouse)
        if upd.warehouse and supply_warehouse:
            if normalize_text(upd.warehouse) == normalize_text(supply_warehouse):
                score += 45
                reasons.append("склад совпал")
            elif normalize_text(upd.warehouse) in normalize_text(supply_warehouse) or normalize_text(supply_warehouse) in normalize_text(upd.warehouse):
                score += 30
                reasons.append("склад похож")

        supply_date = parse_date(first_nonempty(supply.fact_date, supply.supply_date, supply.updated_date))
        if doc_date and supply_date:
            days = abs((doc_date - supply_date).days)
            if days == 0:
                score += 30
                reasons.append("дата совпала")
            elif days <= 2:
                score += 22
                reasons.append(f"дата ±{days} дн.")
            elif days <= 5:
                score += 10
                reasons.append(f"дата ±{days} дн.")

        if upd.total_qty and supply.accepted_qty:
            difference = abs(upd.total_qty - supply.accepted_qty)
            if difference == 0:
                score += 45
                reasons.append("количество совпало")
            elif difference <= 2:
                score += 30
                reasons.append(f"количество отличается на {difference}")
            elif difference / max(upd.total_qty, supply.accepted_qty) <= 0.01:
                score += 20
                reasons.append("количество отличается не более 1%")

        if supply.status_id == 5:
            score += 8
        if score:
            scored.append((score, supply.supply_id, reasons))

    scored.sort(reverse=True)
    if not scored:
        upd.match_reason = "кандидаты не найдены"
        return upd

    best_score, best_id, reasons = scored[0]
    second_score = scored[1][0] if len(scored) > 1 else 0
    if best_score >= 75 and best_score - second_score >= 10:
        upd.supply_id = best_id
        upd.match_score = best_score
        upd.match_reason = ", ".join(reasons)
    else:
        upd.match_score = best_score
        upd.match_reason = (
            f"совпадение неоднозначно: лучший score={best_score}, второй={second_score}; " + ", ".join(reasons)
        )
    return upd


def upd_info_from_dict(data: Dict[str, Any]) -> UPDInfo:
    allowed = {field.name for field in __import__("dataclasses").fields(UPDInfo)}
    return UPDInfo(**{key: value for key, value in (data or {}).items() if key in allowed})


def attach_upd_to_summary(
    summary: SupplySummary,
    upd: UPDInfo,
    service_name: str,
    s3_key: str = "",
) -> None:
    upd_dict = asdict(upd)
    upd_dict["service_name"] = service_name
    upd_dict["s3_key"] = s3_key
    for index, existing in enumerate(summary.upd_documents):
        if str(existing.get("service_name")) == str(service_name):
            summary.upd_documents[index] = upd_dict
            return
    summary.upd_documents.append(upd_dict)


# ---------------------------------------------------------------------------
# Реестр, Excel и уведомления
# ---------------------------------------------------------------------------

def default_registry() -> Dict[str, Any]:
    return {
        "registry_version": 1,
        "script_version": VERSION,
        "created_at": now_msk().isoformat(),
        "updated_at": "",
        "supplies": {},
        "documents": {},
    }


def load_registry(storage: Storage, key: str) -> Tuple[Dict[str, Any], bool]:
    if not storage.exists(key):
        return default_registry(), True
    try:
        data = json.loads(storage.read_bytes(key).decode("utf-8"))
        if not isinstance(data, dict):
            raise ValueError("registry не объект")
        data.setdefault("supplies", {})
        data.setdefault("documents", {})
        return data, False
    except Exception as exc:
        backup = f"{key}.broken_{now_msk().strftime('%Y%m%d_%H%M%S')}.json"
        try:
            storage.write_bytes(backup, storage.read_bytes(key), "application/json")
        except Exception:
            pass
        log(f"Реестр повреждён, создан новый: {exc}")
        return default_registry(), True


def upload_bytes(storage: Storage, key: str, data: bytes, content_type: Optional[str] = None) -> str:
    storage.write_bytes(key, data, content_type or content_type_for_name(key))
    log(f"S3: сохранён {key} ({len(data):,} байт)")
    return key


def autofit_workbook(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column in sheet.columns:
            max_length = 0
            letter = get_column_letter(column[0].column)
            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, min(len(value), 80))
                cell.alignment = Alignment(vertical="top", wrap_text=False)
            sheet.column_dimensions[letter].width = max(10, min(max_length + 2, 55))
        sheet.auto_filter.ref = sheet.dimensions


def build_supply_workbook(
    supplies: Sequence[SupplySummary],
    discrepancy_rows: Sequence[Dict[str, Any]],
    document_rows: Sequence[Dict[str, Any]],
) -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Поставки"
    headers = [
        "Номер поставки", "Статус", "Склад", "Склад плановый", "Склад фактический", "Склад транзитный",
        "Дата создания", "Дата отгрузки", "Дата фактической приёмки", "Дата обновления",
        "Заявлено", "Выгружено", "Принято", "Готово к продаже", "Излишки", "Недостача",
        "Артикулов с расхождением", "Комментарий", "УПД",
    ]
    ws.append(headers)
    for supply in sorted(supplies, key=lambda item: (item.fact_date or item.supply_date or item.create_date, item.supply_id), reverse=True):
        upd_labels = "; ".join(
            f"№{doc.get('document_number', '')} от {doc.get('document_date', '')}".strip()
            for doc in supply.upd_documents
        )
        ws.append([
            supply.supply_id,
            supply.status_name,
            supply.warehouse,
            supply.warehouse_plan,
            supply.warehouse_actual,
            supply.warehouse_transit,
            supply.create_date,
            supply.supply_date,
            supply.fact_date,
            supply.updated_date,
            supply.planned_qty,
            supply.unloading_qty,
            supply.accepted_qty,
            supply.ready_qty,
            supply.excess_qty,
            supply.shortage_qty,
            supply.discrepancy_articles,
            supply.report_comment(),
            upd_labels,
        ])

    ws2 = workbook.create_sheet("Расхождения")
    discrepancy_headers = [
        "Номер поставки", "Склад", "Артикул продавца", "nmID", "Баркод", "Размер", "Цвет", "Нужен КИЗ",
        "Заявлено", "Принято", "Выгружено", "Готово к продаже", "Излишек", "Недостача",
    ]
    ws2.append(discrepancy_headers)
    for row in discrepancy_rows:
        ws2.append([row.get(header, "") for header in discrepancy_headers])

    ws3 = workbook.create_sheet("Документы")
    document_headers = [
        "Тип", "Категория", "serviceName", "Дата создания", "Имя файла", "S3 ключ",
        "Номер УПД", "Дата УПД", "Склад", "Количество", "КИЗ", "Номер поставки", "Качество сопоставления",
    ]
    ws3.append(document_headers)
    for row in document_rows:
        ws3.append([row.get(header, "") for header in document_headers])

    autofit_workbook(workbook)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def document_linked_supply_id(
    service_name: str,
    record: Dict[str, Any],
    valid_supply_ids: set[int],
) -> Optional[int]:
    """Определяет supplyID документа по точному номеру в данных WB."""
    upd = record.get("upd") or {}
    direct = to_int(upd.get("supply_id"), 0)
    if direct and direct in valid_supply_ids:
        return direct

    file_names = [str(item.get("name", "")) for item in (record.get("files") or [])]
    candidates = numeric_ids_from_text(
        service_name,
        record.get("name", ""),
        record.get("category", ""),
        *file_names,
    )
    for candidate in candidates:
        if candidate in valid_supply_ids:
            return candidate
    return None


def final_package_key(
    acceptance_records: Sequence[Dict[str, Any]],
    upd_records: Sequence[Dict[str, Any]],
) -> str:
    """Ключ конечного комплекта: акт приёмки + УПД."""
    return stable_hash({
        "acceptance": sorted(str(item.get("service_name", "")) for item in acceptance_records),
        "upd": sorted(str(item.get("service_name", "")) for item in upd_records),
    })


def final_supply_event_text(
    supply: SupplySummary,
    acceptance_records: Sequence[Dict[str, Any]],
    upd_records: Sequence[Dict[str, Any]],
) -> str:
    """Только конечная карточка поставки: акт уже сформирован, УПД найден."""
    if supply.excess_qty == 0 and supply.shortage_qty == 0:
        discrepancy_text = "нет"
    else:
        discrepancy_text = (
            f"недостача {supply.shortage_qty} шт., "
            f"излишки {supply.excess_qty} шт."
        )

    acceptance_numbers: List[str] = []
    for record in acceptance_records:
        ids = numeric_ids_from_text(record.get("service_name", ""), record.get("name", ""))
        number = str(ids[0]) if ids else str(record.get("service_name", ""))
        if number and number not in acceptance_numbers:
            acceptance_numbers.append(number)

    destination = canonical_warehouse(supply.warehouse_plan) or supply.warehouse or "не определён"
    lines = [
        f"📦 Поставка {supply.supply_id}",
        f"Город/склад: {destination}",
        f"Заявлено: {supply.planned_qty} шт.",
        f"Принято: {supply.accepted_qty} шт.",
        f"Расхождения: {discrepancy_text}",
        f"📄 Акт приёмки: сформирован{(' №' + ', №'.join(acceptance_numbers)) if acceptance_numbers else ''}",
    ]

    for record in upd_records:
        upd = record.get("upd") or {}
        number = upd.get("document_number") or record.get("service_name") or "без номера"
        date_value = upd.get("document_date") or "без даты"
        lines.append(f"📄 УПД: №{number} от {date_value}")
    return "\n".join(lines)


def upd_xml_payload(storage: Storage, record: Dict[str, Any]) -> Optional[Tuple[str, bytes]]:
    """Читает XML УПД из Object Storage, в том числе если он скачан в прошлом запуске."""
    upd = record.get("upd") or {}
    preferred_key = str(upd.get("s3_key", "")).strip()
    candidates: List[Tuple[str, str]] = []
    for item in record.get("files") or []:
        name = str(item.get("name", ""))
        key = str(item.get("key", ""))
        if name.lower().endswith(".xml") and key:
            candidates.append((name, key))
    if preferred_key:
        preferred_name = PurePosixPath(preferred_key).name or str(upd.get("xml_filename") or "upd.xml")
        candidates.insert(0, (preferred_name, preferred_key))

    seen: set[str] = set()
    for name, key in candidates:
        if key in seen:
            continue
        seen.add(key)
        try:
            return safe_filename(name, "upd.xml"), storage.read_bytes(key)
        except Exception as exc:
            log(f"WARN УПД {record.get('service_name', '')}: не удалось прочитать {key}: {exc}")
    return None

def supply_reference_date(supply: SupplySummary) -> Optional[dt.date]:
    """Дата для отбора поставок в Telegram: дата поставки, затем факт/создание."""
    return parse_date(first_nonempty(
        supply.supply_date,
        supply.fact_date,
        supply.create_date,
        supply.updated_date,
    ))


def resolve_supply_message_period(
    mode: str,
    period_days: int,
    date_from_text: str,
    date_to_text: str,
    base_date: dt.date,
) -> Tuple[Optional[dt.date], Optional[dt.date]]:
    if mode == "changes_only":
        return None, None
    if mode != "period":
        raise ValueError(f"Неизвестный режим отправки поставок: {mode}")

    custom_from = parse_date(date_from_text) if str(date_from_text or "").strip() else None
    custom_to = parse_date(date_to_text) if str(date_to_text or "").strip() else None
    if custom_from or custom_to:
        if not custom_from or not custom_to:
            raise ValueError("Для произвольного периода нужно заполнить обе даты: from и to")
        if custom_from > custom_to:
            raise ValueError("Дата начала периода больше даты окончания")
        return custom_from, custom_to

    days = max(int(period_days or 1), 1)
    return base_date - dt.timedelta(days=days - 1), base_date


def supply_in_period(supply: SupplySummary, date_from: dt.date, date_to: dt.date) -> bool:
    value = supply_reference_date(supply)
    return bool(value and date_from <= value <= date_to)


def supply_sort_key(supply: SupplySummary) -> Tuple[dt.date, int]:
    return supply_reference_date(supply) or dt.date.min, supply.supply_id


def supply_telegram_quantity(supply: SupplySummary) -> int:
    """Количество для фильтра Telegram.

    Берём максимум из заявки, приёмки и количества связанных УПД. Это не даёт
    скрыть обычную поставку, если одно из полей Supplies API временно неполное.
    """
    upd_quantities = [to_int(item.get("total_qty"), 0) for item in supply.upd_documents]
    return max(
        [
            to_int(supply.planned_qty, 0),
            to_int(supply.unloading_qty, 0),
            to_int(supply.accepted_qty, 0),
            to_int(supply.ready_qty, 0),
            *upd_quantities,
        ],
        default=0,
    )


def supply_visible_in_telegram(supply: SupplySummary, minimum_qty: int) -> bool:
    """Поставки меньше minimum_qty не показываются и не отправляют XML в Telegram."""
    return supply_telegram_quantity(supply) >= max(to_int(minimum_qty, 0), 0)


# ---------------------------------------------------------------------------
# Основной процесс
# ---------------------------------------------------------------------------

def should_deep_refresh(
    shallow: Dict[str, Any],
    previous: Optional[Dict[str, Any]],
    deep_lookback_days: int,
) -> bool:
    status_id = to_int(first_nonempty(shallow.get("statusID"), (previous or {}).get("status_id")), 0)
    if previous is None:
        return True
    if status_id != to_int(previous.get("status_id"), 0):
        return True
    updated = str(first_nonempty(shallow.get("updatedDate"), ""))
    if updated and updated != str(previous.get("updated_date", "")):
        return True
    if status_id in {4, 5, 6}:
        date_value = parse_date(first_nonempty(shallow.get("factDate"), shallow.get("supplyDate"), shallow.get("updatedDate")))
        if date_value and (today_msk() - date_value).days <= deep_lookback_days:
            return True
    return False


def notification_required(
    summary: SupplySummary,
    previous: Optional[Dict[str, Any]],
    registry_was_new: bool,
    first_run_notify_days: int,
    notify_status_ids: set[int],
    force_notify: bool,
) -> bool:
    if summary.status_id not in notify_status_ids:
        return False
    if force_notify:
        return True
    if previous and previous.get("notified_state_hash") != summary.state_hash:
        return True
    if registry_was_new:
        relevant_date = parse_date(first_nonempty(summary.fact_date, summary.supply_date, summary.updated_date))
        return bool(relevant_date and (today_msk() - relevant_date).days <= first_run_notify_days)
    return previous is None


def run(args: argparse.Namespace) -> int:
    load_report_env()
    token, token_source = first_env(WB_TOKEN_ALIASES)
    if not token:
        fail("Не найден WB_PROMO_KEY_TOPFACE. Используется существующий GitHub Secret; токен должен иметь категории 'Поставки' и 'Документы'.")
    log(f"WB token: найден в {token_source}; значение не выводится")

    telegram_token, telegram_chat, _ = telegram_credentials()
    if not args.no_telegram and (not telegram_token or not telegram_chat):
        fail("Не настроен Telegram: нужны TELEGRAM_BOT_TOKEN и TELEGRAM_CHAT_ID")

    root_prefix = (args.root_prefix or os.getenv("WB_SUPPLY_DOCS_ROOT", DEFAULT_ROOT_PREFIX)).strip("/")
    storage = make_storage(args.local_storage)
    registry_key = join_key(root_prefix, REGISTRY_RELATIVE_KEY)
    registry, registry_was_new = load_registry(storage, registry_key)
    old_supplies: Dict[str, Dict[str, Any]] = registry.setdefault("supplies", {})
    old_documents: Dict[str, Dict[str, Any]] = registry.setdefault("documents", {})

    run_date = today_msk()
    supply_message_from, supply_message_to = resolve_supply_message_period(
        args.supply_message_mode,
        args.supply_period_days,
        args.supply_date_from,
        args.supply_date_to,
        run_date,
    )

    if args.dry_run:
        log(f"DRY RUN: root_prefix={root_prefix}")
        log(f"DRY RUN: registry_exists={not registry_was_new}")
        log(f"DRY RUN: lookback_days={args.lookback_days}; document_lookback_days={args.document_lookback_days}")
        log(f"DRY RUN: upd_send_mode={args.upd_send_mode}; supply_message_mode={args.supply_message_mode}")
        log(f"DRY RUN: telegram_min_supply_qty={args.telegram_min_supply_qty}")
        if supply_message_from and supply_message_to:
            log(f"DRY RUN: supply_message_period={supply_message_from} — {supply_message_to}")
        return 0

    client = WBClient(token)
    date_from = run_date - dt.timedelta(days=max(args.lookback_days, 1))
    date_to = run_date
    if supply_message_from:
        date_from = min(date_from, supply_message_from)
    if supply_message_to:
        date_to = max(date_to, supply_message_to)
    document_from = run_date - dt.timedelta(days=max(args.document_lookback_days, 1))
    document_to = run_date

    log(f"Период поставок: {date_from} — {date_to}")
    shallow_supplies = client.list_supplies(date_from, date_to)
    log(f"Уникальных поставок: {len(shallow_supplies)}")

    # Сначала создаём неглубокую карточку для каждой поставки.
    summaries_by_id: Dict[int, SupplySummary] = {}
    shallow_by_id: Dict[int, Dict[str, Any]] = {}
    deep_candidates: List[Tuple[dt.datetime, int]] = []
    for item in shallow_supplies:
        supply_id = supply_id_from(item)
        if not supply_id:
            continue
        shallow_by_id[supply_id] = item
        previous = old_supplies.get(str(supply_id))
        summary = shallow_summary(item, previous)
        summaries_by_id[supply_id] = summary
        if should_deep_refresh(item, previous, args.deep_lookback_days):
            rank_date = parse_datetime(first_nonempty(item.get("updatedDate"), item.get("factDate"), item.get("supplyDate"), item.get("createDate"))) or dt.datetime.min.replace(tzinfo=MSK)
            deep_candidates.append((rank_date, supply_id))

    deep_candidates.sort(reverse=True)
    if args.max_deep_supplies > 0:
        deep_candidates = deep_candidates[: args.max_deep_supplies]
    log(f"Поставок для детальной проверки: {len(deep_candidates)}")

    discrepancy_rows: List[Dict[str, Any]] = []
    deep_errors: List[str] = []
    for index, (_, supply_id) in enumerate(deep_candidates, start=1):
        try:
            log(f"Поставка {supply_id}: детали и товары ({index}/{len(deep_candidates)})")
            details = client.get_supply_details(supply_id)
            goods = client.get_supply_goods(supply_id)
            summary, item_rows = summary_from_supply(
                shallow_by_id[supply_id],
                details,
                goods,
                old_supplies.get(str(supply_id)),
            )
            summaries_by_id[supply_id] = summary
            discrepancy_rows.extend(item_rows)
        except Exception as exc:
            message = f"Поставка {supply_id}: {exc}"
            deep_errors.append(message)
            log("WARN " + message)

    # Документы.
    log(f"Период документов: {document_from} — {document_to}")
    documents = client.list_documents(document_from, document_to)
    relevant_documents = [(doc, document_kind(doc)) for doc in documents]
    relevant_documents = [(doc, kind) for doc, kind in relevant_documents if kind]
    log(f"Релевантных документов: {len(relevant_documents)}")

    new_document_candidates: List[Tuple[Dict[str, Any], str]] = []
    for doc, kind in relevant_documents:
        service_name = str(doc.get("serviceName", "")).strip()
        if not service_name:
            continue
        previous_doc = old_documents.get(service_name, {})
        if not previous_doc.get("downloaded_at"):
            new_document_candidates.append((doc, kind))

    new_document_candidates.sort(key=lambda pair: str(pair[0].get("creationTime", "")))
    new_document_candidates_total = len(new_document_candidates)
    if args.max_document_downloads > 0:
        new_document_candidates = new_document_candidates[: args.max_document_downloads]
    documents_remaining = max(new_document_candidates_total - len(new_document_candidates), 0)
    log(
        f"Новых документов: найдено={new_document_candidates_total}, "
        f"скачать сейчас={len(new_document_candidates)}, останется={documents_remaining}"
    )

    newly_downloaded: List[Dict[str, Any]] = []
    upd_file_payloads: List[Tuple[str, bytes, Dict[str, Any]]] = []
    document_rows: List[Dict[str, Any]] = []

    for index, (doc, kind) in enumerate(new_document_candidates, start=1):
        service_name = str(doc.get("serviceName", "")).strip()
        extension = preferred_extension(doc, kind)
        try:
            log(f"Документ {service_name}: скачивание {extension} ({index}/{len(new_document_candidates)})")
            response_name, response_extension, content = client.download_document(service_name, extension)
            creation_dt = parse_datetime(doc.get("creationTime")) or now_msk()
            month = creation_dt.astimezone(MSK).strftime("%Y-%m")
            folder = kind_folder(kind)
            document_id = document_storage_id(service_name, response_name)
            document_base_key = join_key(root_prefix, folder, month, document_id)
            original_key = join_key(document_base_key, "Оригиналы", response_name)
            upload_bytes(storage, original_key, content, content_type_for_name(response_name))

            extracted = extract_files_recursive(response_name, content)
            stored_files: List[Dict[str, Any]] = [{"name": response_name, "key": original_key}]
            parsed_upd: Optional[UPDInfo] = None
            seen_names: set[str] = {response_name}
            for extracted_name, extracted_content in extracted:
                # Исходник уже сохранён в Оригиналы.
                if extracted_name == response_name and hashlib.sha256(extracted_content).digest() == hashlib.sha256(content).digest():
                    continue
                final_name = safe_filename(extracted_name)
                if final_name in seen_names:
                    stem, suffix = os.path.splitext(final_name)
                    final_name = f"{stem}_{hashlib.sha1(extracted_content).hexdigest()[:8]}{suffix}"
                seen_names.add(final_name)
                extracted_key = join_key(document_base_key, final_name)
                upload_bytes(storage, extracted_key, extracted_content, content_type_for_name(final_name))
                stored_files.append({"name": final_name, "key": extracted_key})
                if kind == "upd" and final_name.lower().endswith(".xml") and parsed_upd is None:
                    try:
                        parsed_upd = parse_upd_xml(extracted_content, final_name)
                        upd_file_payloads.append((final_name, extracted_content, {"service_name": service_name}))
                    except Exception as exc:
                        log(f"WARN УПД XML {final_name}: не разобран: {exc}")

            if kind == "upd" and parsed_upd is None and response_name.lower().endswith(".xml"):
                parsed_upd = parse_upd_xml(content, response_name)
                upd_file_payloads.append((response_name, content, {"service_name": service_name}))

            if parsed_upd:
                parsed_upd = match_upd_to_supply(parsed_upd, summaries_by_id.values(), service_name)

            doc_record: Dict[str, Any] = {
                "service_name": service_name,
                "kind": kind,
                "category": str(doc.get("category", "")),
                "name": str(doc.get("name", "")),
                "creation_time": str(doc.get("creationTime", "")),
                "extension": response_extension,
                "downloaded_at": now_msk().isoformat(),
                "files": stored_files,
                "notified_at": "",
            }
            if parsed_upd:
                doc_record["upd"] = asdict(parsed_upd)
                if parsed_upd.supply_id and parsed_upd.supply_id in summaries_by_id:
                    summary = summaries_by_id[parsed_upd.supply_id]
                    xml_key = next((f["key"] for f in stored_files if str(f["name"]).lower().endswith(".xml")), original_key)
                    attach_upd_to_summary(summary, parsed_upd, service_name, xml_key)
            old_documents[service_name] = doc_record
            newly_downloaded.append(doc_record)
        except Exception as exc:
            old_documents.setdefault(service_name, {})["last_error"] = str(exc)
            old_documents[service_name]["last_error_at"] = now_msk().isoformat()
            log(f"ERROR документ {service_name}: {exc}")

    # Повторно связываем ранее скачанные УПД из registry. Это исправляет старые
    # записи без supply_id без повторного скачивания файлов из WB.
    rematched_existing = 0
    for service_name, record in old_documents.items():
        if record.get("kind") != "upd" or not isinstance(record.get("upd"), dict):
            continue
        upd_info = upd_info_from_dict(record.get("upd") or {})
        previous_supply_id = upd_info.supply_id
        upd_info = match_upd_to_supply(upd_info, summaries_by_id.values(), service_name)
        record["upd"] = asdict(upd_info)
        if upd_info.supply_id and upd_info.supply_id in summaries_by_id:
            files = record.get("files") or []
            xml_key = next((str(item.get("key", "")) for item in files if str(item.get("name", "")).lower().endswith(".xml")), "")
            attach_upd_to_summary(summaries_by_id[upd_info.supply_id], upd_info, service_name, xml_key)
            if previous_supply_id != upd_info.supply_id:
                rematched_existing += 1
    if rematched_existing:
        log(f"УПД из registry заново связаны с поставками: {rematched_existing}")

    # Формируем строки документов из реестра за текущий скользящий период.
    for service_name, record in old_documents.items():
        upd = record.get("upd") or {}
        files = record.get("files") or []
        document_rows.append({
            "Тип": {"upd": "УПД по маркировке", "acceptance": "Акт приёмки", "reconciliation": "Акт сверки"}.get(record.get("kind"), record.get("kind", "")),
            "Категория": record.get("category", ""),
            "serviceName": service_name,
            "Дата создания": record.get("creation_time", ""),
            "Имя файла": "; ".join(str(item.get("name", "")) for item in files),
            "S3 ключ": "; ".join(str(item.get("key", "")) for item in files),
            "Номер УПД": upd.get("document_number", ""),
            "Дата УПД": upd.get("document_date", ""),
            "Склад": upd.get("warehouse", ""),
            "Количество": upd.get("total_qty", ""),
            "КИЗ": upd.get("kiz_count", ""),
            "Номер поставки": upd.get("supply_id", ""),
            "Качество сопоставления": upd.get("match_reason", ""),
        })

    # Связываем акты приёмки и УПД с поставками.
    # Telegram больше не реагирует на промежуточные статусы поставки.
    valid_supply_ids = set(summaries_by_id)
    acceptance_docs_by_supply: Dict[int, List[Dict[str, Any]]] = {}
    upd_docs_by_supply: Dict[int, List[Dict[str, Any]]] = {}
    for service_name, record in old_documents.items():
        linked_supply_id = document_linked_supply_id(service_name, record, valid_supply_ids)
        if not linked_supply_id:
            continue
        record["linked_supply_id"] = linked_supply_id
        kind = str(record.get("kind", ""))
        if kind == "acceptance":
            acceptance_docs_by_supply.setdefault(linked_supply_id, []).append(record)
        elif kind == "upd":
            upd_docs_by_supply.setdefault(linked_supply_id, []).append(record)

    newly_changed_final_ids: set[int] = set()
    unmatched_new_upds: List[Dict[str, Any]] = []
    for record in newly_downloaded:
        service_name = str(record.get("service_name", ""))
        linked_supply_id = document_linked_supply_id(service_name, record, valid_supply_ids)
        if linked_supply_id and record.get("kind") in {"acceptance", "upd"}:
            newly_changed_final_ids.add(linked_supply_id)
        elif record.get("kind") == "upd":
            upd = record.get("upd") or {}
            unmatched_new_upds.append({**upd, "service_name": service_name})

    if unmatched_new_upds:
        log(
            "WARN Новых УПД без точной привязки к поставке: "
            + ", ".join(str(item.get("service_name", "")) for item in unmatched_new_upds)
        )

    period_supply_ids: set[int] = set()
    if supply_message_from and supply_message_to:
        period_supply_ids = {
            supply_id
            for supply_id, summary in summaries_by_id.items()
            if supply_in_period(summary, supply_message_from, supply_message_to)
        }
        log(
            f"Поставок в выбранном периоде {supply_message_from} — {supply_message_to}: "
            f"{len(period_supply_ids)}"
        )

    # Конечное событие существует только тогда, когда для поставки есть и акт
    # приёмки, и УПД. Изменения статуса 4/5/6 в Telegram не отправляются.
    final_supply_events: Dict[int, Dict[str, Any]] = {}
    for supply_id, summary in summaries_by_id.items():
        previous = old_supplies.get(str(supply_id)) or {}
        acceptance_records = acceptance_docs_by_supply.get(supply_id, [])
        upd_records = upd_docs_by_supply.get(supply_id, [])
        current_final_key = ""
        if acceptance_records and upd_records:
            current_final_key = final_package_key(acceptance_records, upd_records)

        selected_by_period = bool(period_supply_ids and supply_id in period_supply_ids)
        retry_pending = parse_bool(previous.get("final_notification_retry_pending"), False)
        changed_after_notification = bool(
            current_final_key
            and previous.get("final_notification_key")
            and previous.get("final_notification_key") != current_final_key
        )
        should_send = bool(
            current_final_key
            and (
                selected_by_period
                or args.force_notify
                or retry_pending
                or changed_after_notification
                or supply_id in newly_changed_final_ids
            )
        )
        if should_send:
            final_supply_events[supply_id] = {
                "acceptance_records": acceptance_records,
                "upd_records": upd_records,
                "final_key": current_final_key,
                "selected_by_period": selected_by_period,
            }

        data = asdict(summary)
        data["first_seen_at"] = previous.get("first_seen_at") or now_msk().isoformat()
        data["last_seen_at"] = now_msk().isoformat()
        data["notified_state_hash"] = previous.get("notified_state_hash", "")
        data["final_notification_key"] = previous.get("final_notification_key", "")
        data["final_notified_at"] = previous.get("final_notified_at", "")
        data["final_notification_retry_pending"] = retry_pending
        old_supplies[str(supply_id)] = data

    # Мелкие доприёмки скачиваются и остаются в S3/реестре, но Telegram молчит.
    telegram_supply_events: Dict[int, Dict[str, Any]] = {
        supply_id: event
        for supply_id, event in final_supply_events.items()
        if supply_visible_in_telegram(summaries_by_id[supply_id], args.telegram_min_supply_qty)
    }
    suppressed_small_supply_ids = sorted(set(final_supply_events) - set(telegram_supply_events))
    if suppressed_small_supply_ids:
        log(
            f"Telegram: конечные комплекты поставок меньше {args.telegram_min_supply_qty} шт. скрыты: "
            f"{suppressed_small_supply_ids}"
        )
    # Excel и JSON-снимок.
    month = run_date.strftime("%Y-%m")
    all_summaries = list(summaries_by_id.values())
    xlsx_bytes = build_supply_workbook(all_summaries, discrepancy_rows, document_rows)
    daily_xlsx_key = join_key(
        root_prefix,
        "_служебные файлы",
        "Сводки",
        month,
        f"Сверка_поставок_TOPFACE_{run_date.isoformat()}.xlsx",
    )
    upload_bytes(storage, daily_xlsx_key, xlsx_bytes, content_type_for_name(daily_xlsx_key))
    upload_bytes(storage, join_key(root_prefix, CURRENT_XLSX_RELATIVE_KEY), xlsx_bytes, content_type_for_name("x.xlsx"))

    snapshot = {
        "script_version": VERSION,
        "generated_at": now_msk().isoformat(),
        "supply_query_period": {"from": date_from.isoformat(), "to": date_to.isoformat()},
        "supply_message_period": {
            "mode": args.supply_message_mode,
            "from": supply_message_from.isoformat() if supply_message_from else "",
            "to": supply_message_to.isoformat() if supply_message_to else "",
        },
        "upd_send_mode": args.upd_send_mode,
        "telegram_min_supply_qty": args.telegram_min_supply_qty,
        "supplies": [asdict(summary) for summary in all_summaries],
        "discrepancies": discrepancy_rows,
        "new_documents": newly_downloaded,
        "errors": deep_errors,
    }
    snapshot_key = join_key(
        root_prefix,
        "_служебные файлы",
        "Сводки",
        month,
        f"Сверка_поставок_TOPFACE_{run_date.isoformat()}.json",
    )
    upload_bytes(storage, snapshot_key, json_dumps(snapshot), "application/json")

    # Telegram: только конкретные завершённые поставки. Если событий нет — не отправляем ничего.
    discrepancy_supply_count = sum(
        1 for summary in all_summaries
        if summary.excess_qty > 0 or summary.shortage_qty > 0
    )
    new_upd_count = sum(1 for record in newly_downloaded if record.get("kind") == "upd")
    new_acceptance_count = sum(1 for record in newly_downloaded if record.get("kind") == "acceptance")

    ordered_supply_ids = sorted(
        telegram_supply_events,
        key=lambda supply_id: supply_sort_key(summaries_by_id[supply_id]),
        reverse=True,
    )
    message_blocks = [
        final_supply_event_text(
            summaries_by_id[supply_id],
            telegram_supply_events[supply_id]["acceptance_records"],
            telegram_supply_events[supply_id]["upd_records"],
        )
        for supply_id in ordered_supply_ids
    ]
    message = "\n\n".join(message_blocks).strip()

    telegram_ok = True
    telegram_message_sent = False
    if not args.no_telegram and message:
        telegram_message_sent = True
        telegram_ok = send_telegram_message(message)
        if telegram_ok and args.upd_send_mode == "message_and_upd":
            sent_service_names: set[str] = set()
            for supply_id in ordered_supply_ids:
                supply = summaries_by_id[supply_id]
                for record in telegram_supply_events[supply_id]["upd_records"]:
                    service_name = str(record.get("service_name", ""))
                    if not service_name or service_name in sent_service_names:
                        continue
                    sent_service_names.add(service_name)
                    payload = upd_xml_payload(storage, record)
                    if payload is None:
                        log(f"WARN Telegram: XML для {service_name} не найден в Object Storage")
                        telegram_ok = False
                        continue
                    filename, file_bytes = payload
                    upd = record.get("upd") or {}
                    caption = (
                        f"TOPFACE | Поставка {supply_id} | {supply.warehouse or 'склад не определён'} | "
                        f"УПД №{upd.get('document_number') or 'без номера'}"
                    )
                    if not send_telegram_document(filename, file_bytes, caption):
                        telegram_ok = False
    elif not message:
        log("Telegram: новых завершённых комплектов «акт приёмки + УПД» нет; сообщение не отправлено")

    notified_at = now_msk().isoformat()
    if telegram_message_sent and telegram_ok:
        for supply_id, event in telegram_supply_events.items():
            old_supplies[str(supply_id)]["final_notification_key"] = event["final_key"]
            old_supplies[str(supply_id)]["final_notified_at"] = notified_at
            old_supplies[str(supply_id)]["final_notification_retry_pending"] = False
            old_supplies[str(supply_id)].pop("telegram_suppressed_at", None)
            old_supplies[str(supply_id)].pop("telegram_suppressed_reason", None)
    elif telegram_message_sent and not telegram_ok:
        for supply_id in telegram_supply_events:
            old_supplies[str(supply_id)]["final_notification_retry_pending"] = True
    for supply_id in suppressed_small_supply_ids:
        old_supplies[str(supply_id)]["telegram_suppressed_at"] = notified_at
        old_supplies[str(supply_id)]["telegram_suppressed_reason"] = (
            f"quantity_below_{args.telegram_min_supply_qty}"
        )

    # Отмечаем только факт обработки документа; это не означает отправку отдельного
    # Telegram-сообщения по промежуточному статусу поставки.
    for record in newly_downloaded:
        service_name = str(record.get("service_name", ""))
        if service_name in old_documents:
            old_documents[service_name]["processed_at"] = notified_at
    registry["script_version"] = VERSION
    registry["updated_at"] = now_msk().isoformat()
    registry["last_run"] = {
        "status": "ok" if telegram_ok else "telegram_error",
        "telegram_message_sent": telegram_message_sent,
        "supplies_found": len(shallow_supplies),
        "supplies_deep_checked": len(deep_candidates),
        "supplies_sent": len(telegram_supply_events),
        "supplies_suppressed_small": len(suppressed_small_supply_ids),
        "telegram_min_supply_qty": args.telegram_min_supply_qty,
        "supply_message_mode": args.supply_message_mode,
        "supply_message_from": supply_message_from.isoformat() if supply_message_from else "",
        "supply_message_to": supply_message_to.isoformat() if supply_message_to else "",
        "upd_send_mode": args.upd_send_mode,
        "new_documents": len(newly_downloaded),
        "documents_remaining": documents_remaining,
        "new_upds": new_upd_count,
        "new_acceptance_acts": new_acceptance_count,
        "supplies_with_discrepancies": discrepancy_supply_count,
        "final_supply_events": len(final_supply_events),
        "telegram_supply_events": len(telegram_supply_events),
        "unmatched_upds": len(unmatched_new_upds),
        "xlsx_key": daily_xlsx_key,
        "snapshot_key": snapshot_key,
    }
    upload_bytes(storage, registry_key, json_dumps(registry), "application/json")
    upload_bytes(storage, join_key(root_prefix, LAST_RUN_RELATIVE_KEY), json_dumps(registry["last_run"]), "application/json")

    log(
        "Готово: "
        f"поставок={len(all_summaries)}, deep={len(deep_candidates)}, "
        f"финальных_событий={len(final_supply_events)}, отправлено={len(telegram_supply_events)}, "
        f"скрыто_мелких={len(suppressed_small_supply_ids)}, сообщение={telegram_message_sent}, "
        f"новых документов={len(newly_downloaded)}, "
        f"осталось={documents_remaining}, upd_mode={args.upd_send_mode}, telegram_ok={telegram_ok}"
    )
    return 0 if telegram_ok or args.no_telegram else 2


# ---------------------------------------------------------------------------
# Self-test
# ---------------------------------------------------------------------------

def self_test() -> int:
    sample_xml = b'''<?xml version="1.0" encoding="windows-1251"?>\n<\xd4\xe0\xe9\xeb><\xc4\xee\xea\xf3\xec\xe5\xed\xf2><\xd1\xe2\xd1\xf7\xd4\xe0\xea\xf2 \xcd\xee\xec\xe5\xf0\xc4\xee\xea="40715881" \xc4\xe0\xf2\xe0\xc4\xee\xea="30.07.2026"/><\xc3\xf0\xf3\xe7\xcf\xee\xeb\xf3\xf7><\xc0\xe4\xf0\xc8\xed\xf4 \xc0\xe4\xf0\xd2\xe5\xea\xf1\xf2="\xd0\xee\xf1\xf1\xe8\xff, \xca\xee\xeb\xe5\xe4\xe8\xed\xee"/></\xc3\xf0\xf3\xe7\xcf\xee\xeb\xf3\xf7><\xd2\xe0\xe1\xeb\xd1\xf7\xd4\xe0\xea\xf2><\xd1\xe2\xe5\xe4\xd2\xee\xe2 \xca\xee\xeb\xd2\xee\xe2="556"/><\xc2\xf1\xe5\xe3\xee\xce\xef\xeb \xca\xee\xeb\xcd\xe5\xf2\xf2\xee\xc2\xf1="556"/></\xd2\xe0\xe1\xeb\xd1\xf7\xd4\xe0\xea\xf2><\xd1\xe2\xcf\xe5\xf0><\xce\xf1\xed\xcf\xe5\xf0 \xd0\xe5\xea\xe2\xcd\xee\xec\xe5\xf0\xc4\xee\xea="40715881"/></\xd1\xe2\xcf\xe5\xf0></\xc4\xee\xea\xf3\xec\xe5\xed\xf2></\xd4\xe0\xe9\xeb>'''
    info = parse_upd_xml(sample_xml, "sample.xml")
    assert info.document_number == "40715881", info
    assert info.document_date == "30.07.2026", info
    assert info.warehouse == "Коледино", info
    assert info.total_qty == 556, info

    supply = SupplySummary(
        supply_id=12345678,
        status_id=5,
        status_name="Принято",
        warehouse="Коледино",
        fact_date="2026-07-30T12:00:00+03:00",
        planned_qty=556,
        accepted_qty=556,
    )
    matched = match_upd_to_supply(info, [supply])
    assert matched.supply_id == 12345678, matched
    assert "Принято без расхождений 556 шт." == supply.report_comment()

    direct_supply = SupplySummary(supply_id=40715881, status_id=5, status_name="Принято", warehouse="Коледино")
    direct = match_upd_to_supply(parse_upd_xml(sample_xml, "sample.xml"), [direct_supply], "UPD po markirovke-40715881")
    assert direct.supply_id == 40715881 and direct.match_score == 200, direct

    xlsx = build_supply_workbook([supply], [], [])
    assert xlsx.startswith(b"PK")
    assert len(extract_files_recursive("report.xlsx", xlsx)) == 1

    outer = io.BytesIO()
    with zipfile.ZipFile(outer, "w") as archive:
        archive.writestr("act.xlsx", xlsx)
        archive.writestr("mchd.zip", b"PK\x03\x04not-a-real-nested-archive")
    extracted_names = [name for name, _ in extract_files_recursive("act.zip", outer.getvalue())]
    assert extracted_names == ["act.zip", "act.xlsx", "mchd.zip"], extracted_names

    period_from, period_to = resolve_supply_message_period("period", 7, "", "", dt.date(2026, 7, 31))
    assert period_from == dt.date(2026, 7, 25) and period_to == dt.date(2026, 7, 31)
    custom_from, custom_to = resolve_supply_message_period(
        "period", 7, "2026-07-10", "2026-07-20", dt.date(2026, 7, 31)
    )
    assert custom_from == dt.date(2026, 7, 10) and custom_to == dt.date(2026, 7, 20)

    small_supply = SupplySummary(
        supply_id=1, status_id=5, status_name="Принято", planned_qty=99, accepted_qty=99
    )
    boundary_supply = SupplySummary(
        supply_id=2, status_id=5, status_name="Принято", planned_qty=100, accepted_qty=100
    )
    assert not supply_visible_in_telegram(small_supply, 100)
    assert supply_visible_in_telegram(boundary_supply, 100)
    small_supply.upd_documents = [{"total_qty": 120}]
    assert supply_visible_in_telegram(small_supply, 100)
    acceptance_record = {"service_name": "act-income-2", "kind": "acceptance"}
    upd_record = {
        "service_name": "UPD po markirovke-2",
        "kind": "upd",
        "upd": {"document_number": "2", "document_date": "31.07.2026"},
    }
    card = final_supply_event_text(boundary_supply, [acceptance_record], [upd_record])
    assert "Заявлено: 100 шт." in card
    assert "Акт приёмки: сформирован" in card
    assert "УПД: №2" in card
    log("SELF_TEST: OK")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="TOPFACE: поставки FBW, УПД и Telegram")
    parser.add_argument("--lookback-days", type=int, default=int(os.getenv("WB_SUPPLY_LOOKBACK_DAYS", "45")))
    parser.add_argument("--document-lookback-days", type=int, default=int(os.getenv("WB_DOCUMENT_LOOKBACK_DAYS", "21")))
    parser.add_argument("--deep-lookback-days", type=int, default=int(os.getenv("WB_DEEP_LOOKBACK_DAYS", "21")))
    parser.add_argument("--max-deep-supplies", type=int, default=int(os.getenv("WB_MAX_DEEP_SUPPLIES", "180")))
    parser.add_argument("--max-document-downloads", type=int, default=int(os.getenv("WB_MAX_DOCUMENT_DOWNLOADS", "0")))
    parser.add_argument("--first-run-notify-days", type=int, default=int(os.getenv("WB_FIRST_RUN_NOTIFY_DAYS", "3")))
    parser.add_argument("--notify-status-ids", default=os.getenv("WB_NOTIFY_STATUS_IDS", "4,5,6"), help="Совместимость; статусы больше не являются триггером Telegram")
    parser.add_argument("--root-prefix", default=os.getenv("WB_SUPPLY_DOCS_ROOT", DEFAULT_ROOT_PREFIX))
    parser.add_argument(
        "--upd-send-mode",
        choices=("message_and_upd", "message_only"),
        default=os.getenv("WB_UPD_SEND_MODE", "message_and_upd"),
        help="message_and_upd — сообщение и XML; message_only — только сообщение",
    )
    parser.add_argument(
        "--supply-message-mode",
        choices=("changes_only", "period"),
        default=os.getenv("WB_SUPPLY_MESSAGE_MODE", "changes_only"),
        help="changes_only — новые конечные комплекты акт+УПД; period — готовые комплекты за выбранный период",
    )
    parser.add_argument("--supply-period-days", type=int, default=int(os.getenv("WB_SUPPLY_MESSAGE_PERIOD_DAYS", "7")))
    parser.add_argument("--supply-date-from", default=os.getenv("WB_SUPPLY_MESSAGE_DATE_FROM", ""))
    parser.add_argument("--supply-date-to", default=os.getenv("WB_SUPPLY_MESSAGE_DATE_TO", ""))
    parser.add_argument(
        "--telegram-min-supply-qty",
        type=int,
        default=int(os.getenv("WB_TELEGRAM_MIN_SUPPLY_QTY", str(DEFAULT_TELEGRAM_MIN_SUPPLY_QTY))),
        help="Поставки с меньшим количеством сохраняются, но не показываются и не отправляют XML в Telegram",
    )
    parser.add_argument("--force-notify", action="store_true")
    parser.add_argument("--no-telegram", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--local-storage", default="", help="Локальная папка вместо S3, только для теста")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    if args.self_test:
        return self_test()
    try:
        return run(args)
    except KeyboardInterrupt:
        log("Остановлено пользователем")
        return 130
    except SystemExit:
        raise
    except Exception as exc:
        log(f"FATAL: {exc}")
        traceback.print_exc()
        # Пытаемся уведомить о падении, если Telegram уже настроен.
        if not args.no_telegram:
            try:
                send_telegram_message(f"❌ TOPFACE: ошибка контроля поставок и УПД\n{str(exc)[:3000]}")
            except Exception:
                pass
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
