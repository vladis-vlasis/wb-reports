#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Wildberries: реестр поставок, продаж, текущих остатков и потенциально уничтоженного товара.

Версия: WB_WAREHOUSE_LOSS_REGISTRY_V3_20260805

Назначение
==========
1. Получить поставки FBW за выбранный период и товарную детализацию по каждой поставке.
2. Сформировать первичный реестр по строгому ключу:
       supplyID (номер поставки / giId) + barcode.
3. Получить финансовую детализацию и посчитать продажи/возвраты строго по тому же ключу.
4. Получить текущие остатки WB:
       на складах + в пути к клиенту + в пути от клиента.
5. Прочитать историю заказов и ежедневных остатков из недельных Excel-файлов,
   которые создаёт ezhednevnoe_obnovlenie_dannyh.py в Yandex Object Storage.
6. Автоматически найти склады с резкой устойчивой остановкой заказов либо использовать
   заданную вручную дату инцидента, восстановить последний остаток перед остановкой и
   оценить потенциально уничтоженное количество.
7. Сформировать отдельные листы по всем непроданным партиям и партиям старше заданного порога.
8. Сформировать Excel и отправить его в Telegram.

Важно
=====
- Продажи по поставке считаются точно по giId + barcode.
- Текущий остаток WB не содержит giId, поэтому он доступен только на уровне barcode/склад.
- Показатель потенциального уничтожения является расчётной оценкой по историческому
  снимку остатков и заказам после снимка.
- Внутренние перемещения WB между складами API не раскрывает. Поэтому колонки
  «принято на склад» и «текущий остаток на складе» нельзя трактовать как единый
  складской баланс без оговорок.

Переменные окружения
====================
Обязательные:
- YC_ACCESS_KEY_ID
- YC_SECRET_ACCESS_KEY
- YC_BUCKET_NAME
- WB_PROMO_KEY_TOPFACE              основной токен TOPFACE
- WB_FINANCE_KEY_TOPFACE            токен Finance; если не задан, используется основной
- TELEGRAM_BOT_TOKEN
- TELEGRAM_CHAT_ID

Необязательные:
- YC_ENDPOINT_URL                   по умолчанию https://storage.yandexcloud.net
- TELEGRAM_MESSAGE_THREAD_ID
- REPORT_ENV                        многострочный KEY=VALUE секрет
- WB_INCIDENT_WAREHOUSES            список через запятую
- WB_INCIDENT_DATES_JSON             JSON: {"Рязань":"2026-07-30"}
- WB_SHUTDOWN_CONFIRM_DAYS           по умолчанию 3
- WB_MIN_ACTIVE_ORDER_DAYS           по умолчанию 2
- WB_MIN_ACTIVE_ORDERS               по умолчанию 5
- WB_AGING_THRESHOLD_DAYS             порог залежалого остатка, по умолчанию 120

Примеры запуска
===============
Обычный запуск за 2026 год с отправкой файла в Telegram:

    python wb_warehouse_loss_registry_TOPFACE_V3_20260805.py

Задать период:

    python wb_warehouse_loss_registry_TOPFACE_V3_20260805.py \
      --date-from 2026-01-01 \
      --date-to 2026-08-05

Задать даты пожаров вручную:

    python wb_warehouse_loss_registry_TOPFACE_V3_20260805.py \
      --incident-dates-json '{"Рязань":"2026-07-30","Владимир":"2026-07-29"}'

Не отправлять в Telegram:

    python wb_warehouse_loss_registry_TOPFACE_V3_20260805.py --no-telegram
"""

from __future__ import annotations

import argparse
import datetime as dt
import io
import json
import os
import re
import sys
import tempfile
import time
import traceback
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple
from zoneinfo import ZoneInfo

import boto3
import pandas as pd
import requests
from botocore.client import Config
from botocore.exceptions import ClientError
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


VERSION = "WB_WAREHOUSE_LOSS_REGISTRY_V3_20260805"
MSK = ZoneInfo("Europe/Moscow")

SUPPLIES_LIST_URL = "https://supplies-api.wildberries.ru/api/v1/supplies"
SUPPLY_DETAILS_URL = "https://supplies-api.wildberries.ru/api/v1/supplies/{supply_id}"
SUPPLY_GOODS_URL = "https://supplies-api.wildberries.ru/api/v1/supplies/{supply_id}/goods"
FINANCE_SALES_DETAILED_URL = "https://finance-api.wildberries.ru/api/finance/v1/sales-reports/detailed"
WAREHOUSE_REMAINS_CREATE_URL = "https://seller-analytics-api.wildberries.ru/api/v1/warehouse_remains"
WAREHOUSE_REMAINS_STATUS_URL = (
    "https://seller-analytics-api.wildberries.ru/api/v1/warehouse_remains/tasks/{task_id}/status"
)
WAREHOUSE_REMAINS_DOWNLOAD_URL = (
    "https://seller-analytics-api.wildberries.ru/api/v1/warehouse_remains/tasks/{task_id}/download"
)

ALL_SUPPLY_STATUS_IDS = [1, 2, 3, 4, 5, 6]
STATUS_NAMES = {
    1: "Не запланировано",
    2: "Запланировано",
    3: "Отгрузка разрешена",
    4: "Идёт приёмка",
    5: "Принято",
    6: "Отгружено на воротах",
}

DEFAULT_INCIDENT_WAREHOUSES = ""
IN_WAY_TO_NAMES = {
    "В ПУТИ ДО ПОЛУЧАТЕЛЕЙ",
    "В ПУТИ К КЛИЕНТУ",
    "В ПУТИ ДО КЛИЕНТА",
}
IN_WAY_FROM_NAMES = {
    "В ПУТИ ВОЗВРАТЫ НА СКЛАД WB",
    "В ПУТИ ОТ КЛИЕНТА",
    "В ПУТИ ВОЗВРАТЫ",
}
TOTAL_WAREHOUSE_NAMES = {
    "ВСЕГО НАХОДИТСЯ НА СКЛАДАХ",
    "ИТОГО",
    "ВСЕГО",
}

STORE_MAIN_TOKEN_ENV = {
    "TOPFACE": "WB_PROMO_KEY_TOPFACE",
    "MISSTAIS": "WB_KEY_MISSTAIS",
}
STORE_FINANCE_TOKEN_ENV = {
    "TOPFACE": "WB_FINANCE_KEY_TOPFACE",
    "MISSTAIS": "WB_FINANCE_KEY_MISSTAIS",
}

WAREHOUSE_RULES: Tuple[Tuple[str, str], ...] = (
    ("КОЛЕДИНО", "Коледино"),
    ("ЭЛЕКТРОСТАЛ", "Электросталь"),
    ("БЕЛАЯ ДАЧ", "Белая Дача"),
    ("ВЕШК", "Вёшки"),
    ("ВЁШК", "Вёшки"),
    ("РЯЗАН", "Рязань"),
    ("ТУЛ", "Тула"),
    ("АЛЕКСИН", "Тула"),
    ("ВЛАДИМИР", "Владимир"),
    ("КОТОВСК", "Котовск"),
    ("ВОРОНЕЖ", "Воронеж"),
    ("КРАСНОДАР", "Краснодар"),
    ("НЕВИННОМЫССК", "Невинномысск"),
    ("НЕВИННОМЫСК", "Невинномысск"),
    ("ВОЛГОГРАД", "Волгоград"),
    ("РОСТОВ", "Ростов/Аксай"),
    ("АКСАЙ", "Ростов/Аксай"),
    ("КАЗАН", "Казань"),
    ("ПЕНЗ", "Пенза"),
    ("САРАПУЛ", "Сарапул"),
    ("НОВОСЕМЕЙКИНО", "Новосемейкино"),
    ("САМАР", "Самара"),
    ("ШУШАР", "СПБ Шушары"),
    ("УТКИН", "СПБ Уткина Заводь"),
    ("САНКТ", "Санкт-Петербург"),
    ("ПЕТЕРБУРГ", "Санкт-Петербург"),
    ("ЕКАТЕРИНБУРГ", "Екатеринбург"),
    ("ПЕРСПЕКТИВН", "Екатеринбург"),
    ("ЧЕЛЯБИНСК", "Челябинск"),
    ("ПЕРМ", "Пермь"),
    ("НОВОСИБИРСК", "Новосибирск"),
    ("КРАСНОЯРСК", "Красноярск"),
    ("КЕМЕРОВО", "Кемерово"),
    ("ХАБАРОВСК", "Хабаровск"),
)


# ---------------------------------------------------------------------------
# Общие функции
# ---------------------------------------------------------------------------


def now_msk() -> dt.datetime:
    return dt.datetime.now(MSK)


def today_msk() -> dt.date:
    return now_msk().date()


def log(message: str, level: str = "INFO") -> None:
    print(
        f"[{now_msk().strftime('%Y-%m-%d %H:%M:%S MSK')}] [{level}] {message}",
        flush=True,
    )


def fail(message: str, code: int = 1) -> None:
    print(f"ERROR: {message}", file=sys.stderr, flush=True)
    raise SystemExit(code)


def load_report_env() -> None:
    raw = os.getenv("REPORT_ENV", "") or os.getenv("WB_REPORT_ENV", "") or ""
    if not raw.strip():
        return
    normalized = raw.replace("\r\n", "\n").replace("\r", "\n")
    if "\n" not in normalized and "\\n" in normalized:
        normalized = normalized.replace("\\n", "\n")
    for raw_line in normalized.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        if line.startswith("export "):
            line = line[7:].strip()
        key, value = line.split("=", 1)
        key = key.strip().lstrip("\ufeff")
        value = value.strip().strip('"').strip("'")
        if key and not os.getenv(key, "").strip():
            os.environ[key] = value


def to_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default
    try:
        text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
        if not text or text.lower() in {"nan", "none", "null", "nat"}:
            return default
        return int(Decimal(text))
    except (InvalidOperation, ValueError, TypeError):
        return default


def to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "y", "да", "+"}


def first_nonempty(*values: Any) -> Any:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text and text.lower() not in {"nan", "none", "null", "nat"}:
            return value
    return ""


def normalize_text(value: Any) -> str:
    text = str(value or "").replace("ё", "е").replace("Ё", "Е")
    text = re.sub(r"[^0-9A-Za-zА-Яа-я]+", " ", text).upper()
    return re.sub(r"\s+", " ", text).strip()


def canonical_warehouse(value: Any) -> str:
    normalized = normalize_text(value)
    if not normalized:
        return ""
    if normalized in IN_WAY_TO_NAMES:
        return "В пути к клиенту"
    if normalized in IN_WAY_FROM_NAMES:
        return "В пути от клиента"
    if normalized in TOTAL_WAREHOUSE_NAMES:
        return "Итого WB"
    for key, result in WAREHOUSE_RULES:
        if key in normalized:
            return result
    return str(value or "").strip()


def normalize_barcode(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def normalize_srid(value: Any) -> str:
    text = str(value or "").strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def parse_date(value: Any) -> Optional[dt.date]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null", "nat"}:
        return None
    text = text.replace("Z", "+00:00")
    try:
        return dt.datetime.fromisoformat(text).date()
    except ValueError:
        pass
    for fmt in (
        "%Y-%m-%d",
        "%d.%m.%Y",
        "%Y-%m-%d %H:%M:%S",
        "%d.%m.%Y %H:%M:%S",
    ):
        try:
            return dt.datetime.strptime(text[:19], fmt).date()
        except ValueError:
            continue
    parsed = pd.to_datetime(text, errors="coerce")
    return None if pd.isna(parsed) else parsed.date()


def safe_filename(value: str, fallback: str = "file.bin") -> str:
    name = Path(str(value or "")).name
    name = re.sub(r"[\\/:*?\"<>|]+", "_", name)
    name = re.sub(r"\s+", " ", name).strip(" .")
    return name[:220] or fallback


def dataframe_column(df: pd.DataFrame, candidates: Sequence[str]) -> Optional[str]:
    normalized = {normalize_text(col): col for col in df.columns}
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
        found = normalized.get(normalize_text(candidate))
        if found is not None:
            return found
    return None


def parse_json_object(raw: str, label: str) -> Dict[str, Any]:
    if not str(raw or "").strip():
        return {}
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Некорректный JSON в {label}: {exc}") from exc
    if not isinstance(parsed, dict):
        raise ValueError(f"{label} должен быть JSON-объектом")
    return parsed


# ---------------------------------------------------------------------------
# Yandex Object Storage
# ---------------------------------------------------------------------------


class S3Storage:
    def __init__(self, access_key: str, secret_key: str, bucket: str, endpoint: str):
        self.bucket = bucket
        self.client = boto3.client(
            "s3",
            endpoint_url=endpoint,
            aws_access_key_id=access_key,
            aws_secret_access_key=secret_key,
            region_name=os.getenv("AWS_DEFAULT_REGION", "ru-central1"),
            config=Config(
                signature_version="s3v4",
                read_timeout=300,
                connect_timeout=60,
                retries={"max_attempts": 5},
            ),
        )

    def list_keys(self, prefix: str) -> List[str]:
        result: List[str] = []
        token: Optional[str] = None
        while True:
            kwargs: Dict[str, Any] = {
                "Bucket": self.bucket,
                "Prefix": prefix,
                "MaxKeys": 1000,
            }
            if token:
                kwargs["ContinuationToken"] = token
            response = self.client.list_objects_v2(**kwargs)
            for item in response.get("Contents", []) or []:
                key = str(item.get("Key", ""))
                if key:
                    result.append(key)
            if not response.get("IsTruncated"):
                break
            token = response.get("NextContinuationToken")
            if not token:
                break
        return sorted(result)

    def read_bytes(self, key: str) -> bytes:
        return self.client.get_object(Bucket=self.bucket, Key=key)["Body"].read()

    def write_bytes(self, key: str, data: bytes, content_type: str) -> None:
        self.client.put_object(
            Bucket=self.bucket,
            Key=key,
            Body=data,
            ContentType=content_type,
        )


# ---------------------------------------------------------------------------
# Telegram
# ---------------------------------------------------------------------------


def telegram_credentials() -> Tuple[str, str, str]:
    token = first_nonempty(
        os.getenv("TELEGRAM_BOT_TOKEN"),
        os.getenv("TG_BOT_TOKEN"),
        os.getenv("TELEGRAM_TOKEN"),
    )
    chat_id = first_nonempty(
        os.getenv("TELEGRAM_CHAT_ID"),
        os.getenv("TG_CHAT_ID"),
        os.getenv("CHAT_ID"),
    )
    thread_id = first_nonempty(
        os.getenv("TELEGRAM_MESSAGE_THREAD_ID"),
        os.getenv("TELEGRAM_THREAD_ID"),
        os.getenv("TG_MESSAGE_THREAD_ID"),
    )
    return str(token), str(chat_id), str(thread_id)


def send_telegram_document(filename: str, data: bytes, caption: str) -> bool:
    token, chat_id, thread_id = telegram_credentials()
    if not token or not chat_id:
        log("Telegram: отсутствует токен или chat_id", "ERROR")
        return False

    fields: Dict[str, str] = {
        "chat_id": chat_id,
        "caption": caption[:1000],
    }
    if thread_id:
        fields["message_thread_id"] = thread_id

    boundary = "----WBRegistry" + os.urandom(12).hex()
    body = bytearray()
    for name, value in fields.items():
        body.extend(f"--{boundary}\r\n".encode())
        body.extend(f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode())
        body.extend(str(value).encode("utf-8"))
        body.extend(b"\r\n")

    body.extend(f"--{boundary}\r\n".encode())
    body.extend(
        f'Content-Disposition: form-data; name="document"; filename="{safe_filename(filename)}"\r\n'.encode(
            "utf-8"
        )
    )
    body.extend(
        b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
    )
    body.extend(data)
    body.extend(f"\r\n--{boundary}--\r\n".encode())

    request = urllib.request.Request(
        f"https://api.telegram.org/bot{token}/sendDocument",
        data=bytes(body),
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
    )
    try:
        with urllib.request.urlopen(request, timeout=180) as response:
            response_text = response.read().decode("utf-8", errors="replace")[:500]
            ok = 200 <= response.status < 300
            log(f"Telegram sendDocument: status={response.status}, ok={ok}, response={response_text}")
            return ok
    except (urllib.error.URLError, TimeoutError, OSError) as exc:
        log(f"Telegram sendDocument: ошибка: {exc}", "ERROR")
        return False


# ---------------------------------------------------------------------------
# WB API
# ---------------------------------------------------------------------------


class RateLimiter:
    def __init__(self, min_interval_seconds: float):
        self.min_interval_seconds = max(float(min_interval_seconds), 0.0)
        self.last_request_at = 0.0

    def wait(self) -> None:
        elapsed = time.monotonic() - self.last_request_at
        wait_for = self.min_interval_seconds - elapsed
        if wait_for > 0:
            time.sleep(wait_for)
        self.last_request_at = time.monotonic()


class WBClient:
    def __init__(self, main_token: str, finance_token: str):
        self.main_token = main_token.strip()
        self.finance_token = finance_token.strip()
        self.session = requests.Session()
        self.supplies_limiter = RateLimiter(os.getenv("WB_SUPPLIES_INTERVAL_SECONDS", "2.05"))
        self.finance_limiter = RateLimiter(os.getenv("WB_FINANCE_INTERVAL_SECONDS", "60.1"))
        self.analytics_limiter = RateLimiter(os.getenv("WB_ANALYTICS_INTERVAL_SECONDS", "20.1"))
        self.retry_delays = [5, 10, 20, 40, 80, 120]

    def request_json(
        self,
        method: str,
        url: str,
        *,
        token: str,
        params: Optional[Dict[str, Any]] = None,
        json_body: Any = None,
        limiter: Optional[RateLimiter] = None,
        timeout: int = 180,
    ) -> Any:
        headers = {"Authorization": token}
        if json_body is not None:
            headers["Content-Type"] = "application/json"

        last_error = ""
        for attempt in range(1, len(self.retry_delays) + 2):
            if limiter:
                limiter.wait()
            try:
                response = self.session.request(
                    method,
                    url,
                    headers=headers,
                    params=params,
                    json=json_body,
                    timeout=timeout,
                )
                if response.status_code == 204:
                    return None
                if response.status_code == 429:
                    retry_after = response.headers.get("Retry-After", "")
                    delay = (
                        int(retry_after)
                        if str(retry_after).isdigit()
                        else self.retry_delays[min(attempt - 1, len(self.retry_delays) - 1)]
                    )
                    log(f"WB API 429: {method} {url}; ожидание {delay} сек", "WARN")
                    time.sleep(delay)
                    continue
                if response.status_code in {500, 502, 503, 504}:
                    delay = self.retry_delays[min(attempt - 1, len(self.retry_delays) - 1)]
                    log(f"WB API {response.status_code}: повтор через {delay} сек", "WARN")
                    time.sleep(delay)
                    continue
                if response.status_code >= 400:
                    raise RuntimeError(
                        f"HTTP {response.status_code}: {method} {url}; {response.text[:1500]}"
                    )
                if not response.text.strip():
                    return None
                return response.json()
            except (requests.RequestException, ValueError, RuntimeError) as exc:
                last_error = str(exc)
                if attempt <= len(self.retry_delays):
                    delay = self.retry_delays[attempt - 1]
                    log(f"WB API: {exc}; повтор через {delay} сек", "WARN")
                    time.sleep(delay)
                    continue
                raise RuntimeError(f"WB API: запрос не выполнен: {last_error}") from exc
        raise RuntimeError(f"WB API: запрос не выполнен: {last_error}")

    @staticmethod
    def unwrap(data: Any) -> Any:
        if isinstance(data, dict) and "data" in data:
            return data["data"]
        if isinstance(data, dict) and "result" in data:
            return data["result"]
        return data

    @staticmethod
    def rows(data: Any, keys: Sequence[str]) -> List[Dict[str, Any]]:
        raw = WBClient.unwrap(data)
        if isinstance(raw, list):
            return [row for row in raw if isinstance(row, dict)]
        if isinstance(raw, dict):
            for key in keys:
                value = raw.get(key)
                if isinstance(value, list):
                    return [row for row in value if isinstance(row, dict)]
        return []

    def list_supplies(self, date_from: dt.date, date_to: dt.date) -> List[Dict[str, Any]]:
        by_id: Dict[int, Dict[str, Any]] = {}
        for date_type in ("createDate", "supplyDate", "factDate"):
            offset = 0
            while True:
                body = {
                    "dates": [
                        {
                            "from": date_from.isoformat(),
                            "till": date_to.isoformat(),
                            "type": date_type,
                        }
                    ],
                    "statusIDs": ALL_SUPPLY_STATUS_IDS,
                }
                data = self.request_json(
                    "POST",
                    SUPPLIES_LIST_URL,
                    token=self.main_token,
                    params={"limit": 1000, "offset": offset},
                    json_body=body,
                    limiter=self.supplies_limiter,
                )
                rows = self.rows(data, ("supplies", "items", "rows"))
                log(f"Поставки: type={date_type}, offset={offset}, получено={len(rows)}")
                for row in rows:
                    supply_id = to_int(first_nonempty(row.get("supplyID"), row.get("supplyId"), row.get("id")), 0)
                    if supply_id:
                        by_id[supply_id] = {**by_id.get(supply_id, {}), **row}
                if len(rows) < 1000:
                    break
                offset += 1000
        return list(by_id.values())

    def get_supply_details(self, supply_id: int) -> Dict[str, Any]:
        data = self.request_json(
            "GET",
            SUPPLY_DETAILS_URL.format(supply_id=supply_id),
            token=self.main_token,
            params={"isPreorderID": "false"},
            limiter=self.supplies_limiter,
        )
        raw = self.unwrap(data)
        return raw if isinstance(raw, dict) else {}

    def get_supply_goods(self, supply_id: int) -> List[Dict[str, Any]]:
        result: List[Dict[str, Any]] = []
        offset = 0
        while True:
            data = self.request_json(
                "GET",
                SUPPLY_GOODS_URL.format(supply_id=supply_id),
                token=self.main_token,
                params={"limit": 1000, "offset": offset, "isPreorderID": "false"},
                limiter=self.supplies_limiter,
            )
            rows = self.rows(data, ("goods", "items", "products", "rows"))
            result.extend(rows)
            if len(rows) < 1000:
                break
            offset += 1000
        return result

    def get_finance_details(
        self,
        date_from: dt.date,
        date_to: dt.date,
        page_limit: int = 100000,
    ) -> List[Dict[str, Any]]:
        fields = [
            "rrdId",
            "giId",
            "nmId",
            "vendorCode",
            "title",
            "techSize",
            "sku",
            "docTypeName",
            "quantity",
            "sellerOperName",
            "orderDt",
            "saleDt",
            "rrDate",
            "shkId",
            "srid",
            "officeName",
        ]
        limit = max(1, min(int(page_limit), 100000))
        result: List[Dict[str, Any]] = []
        rrd_id = 0
        page = 0
        while True:
            page += 1
            body = {
                "dateFrom": date_from.isoformat(),
                "dateTo": date_to.isoformat(),
                "limit": limit,
                "rrdId": rrd_id,
                "period": "daily",
                "fields": fields,
            }
            data = self.request_json(
                "POST",
                FINANCE_SALES_DETAILED_URL,
                token=self.finance_token,
                json_body=body,
                limiter=self.finance_limiter,
            )
            rows = self.rows(data, ("items", "rows", "details"))
            log(f"Финансы: страница={page}, rrdId={rrd_id}, строк={len(rows)}")
            if not rows:
                break
            result.extend(rows)
            next_rrd_id = to_int(first_nonempty(rows[-1].get("rrdId"), rows[-1].get("rrd_id")), 0)
            if next_rrd_id <= rrd_id:
                raise RuntimeError(
                    f"Финансы: пагинация остановилась, old={rrd_id}, new={next_rrd_id}"
                )
            rrd_id = next_rrd_id
            if len(rows) < limit:
                break
        return result

    def get_current_warehouse_remains(self) -> List[Dict[str, Any]]:
        params = {
            "locale": "ru",
            "groupByBrand": "true",
            "groupBySubject": "true",
            "groupBySa": "true",
            "groupByNm": "true",
            "groupByBarcode": "true",
            "groupBySize": "true",
        }
        create = self.request_json(
            "GET",
            WAREHOUSE_REMAINS_CREATE_URL,
            token=self.main_token,
            params=params,
            limiter=self.analytics_limiter,
            timeout=120,
        )
        raw = self.unwrap(create)
        task_id = ""
        if isinstance(raw, dict):
            task_id = str(first_nonempty(raw.get("taskId"), raw.get("taskID")))
        if not task_id:
            raise RuntimeError(f"warehouse_remains: нет taskId, ответ={str(create)[:1000]}")
        log(f"warehouse_remains: taskId={task_id}")

        status_url = WAREHOUSE_REMAINS_STATUS_URL.format(task_id=task_id)
        for attempt in range(1, 61):
            status_data = self.request_json(
                "GET",
                status_url,
                token=self.main_token,
                limiter=RateLimiter(3.0),
                timeout=60,
            )
            status_raw = self.unwrap(status_data)
            status = ""
            if isinstance(status_raw, dict):
                status = str(status_raw.get("status", "")).strip().lower()
            log(f"warehouse_remains: status={status or 'unknown'} ({attempt}/60)")
            if status == "done":
                break
            if status in {"failed", "error", "canceled", "cancelled"}:
                raise RuntimeError(f"warehouse_remains завершился статусом {status}")
            time.sleep(7)
        else:
            raise RuntimeError("warehouse_remains не подготовился за отведённое время")

        download = self.request_json(
            "GET",
            WAREHOUSE_REMAINS_DOWNLOAD_URL.format(task_id=task_id),
            token=self.main_token,
            limiter=self.analytics_limiter,
            timeout=180,
        )
        rows = self.rows(download, ("items", "rows", "stocks"))
        log(f"warehouse_remains: получено товаров={len(rows)}")
        return rows


# ---------------------------------------------------------------------------
# Нормализация исходных данных
# ---------------------------------------------------------------------------


@dataclass
class IncidentInfo:
    warehouse: str
    incident_date: Optional[dt.date]
    source: str
    last_order_date: Optional[dt.date]
    active_order_days: int
    active_orders: int
    quiet_days: int
    confidence: str
    warning: str = ""


def supply_id_from(row: Dict[str, Any]) -> int:
    return to_int(first_nonempty(row.get("supplyID"), row.get("supplyId"), row.get("id")), 0)


def build_supply_lots(
    client: WBClient,
    supplies: Sequence[Dict[str, Any]],
    date_from: dt.date,
    date_to: dt.date,
) -> Tuple[pd.DataFrame, List[str]]:
    rows: List[Dict[str, Any]] = []
    warnings: List[str] = []
    total = len(supplies)
    for index, shallow in enumerate(supplies, start=1):
        supply_id = supply_id_from(shallow)
        if not supply_id:
            continue
        try:
            details = client.get_supply_details(supply_id)
            goods = client.get_supply_goods(supply_id)
            merged = {**shallow, **details}
            supply_date = parse_date(
                first_nonempty(
                    merged.get("factDate"),
                    merged.get("supplyDate"),
                    merged.get("updatedDate"),
                    merged.get("createDate"),
                )
            )
            if not supply_date or not (date_from <= supply_date <= date_to):
                continue
            warehouse = canonical_warehouse(
                first_nonempty(
                    merged.get("actualWarehouseName"),
                    merged.get("warehouseName"),
                    merged.get("transitWarehouseName"),
                )
            )
            status_id = to_int(merged.get("statusID"), 0)
            if not goods:
                warnings.append(f"Поставка {supply_id}: Supplies API не вернул товары")
                continue
            log(f"Поставка {supply_id}: товары={len(goods)} ({index}/{total})")
            for good in goods:
                barcode = normalize_barcode(first_nonempty(good.get("barcode"), good.get("barCode")))
                if not barcode:
                    warnings.append(f"Поставка {supply_id}: строка без barcode")
                    continue
                planned = to_int(good.get("quantity"), 0)
                accepted = to_int(good.get("acceptedQuantity"), 0)
                unloading = to_int(good.get("unloadingQuantity"), 0)
                ready = to_int(good.get("readyForSaleQuantity"), 0)
                rows.append(
                    {
                        "supply_id": supply_id,
                        "supply_barcode_key": f"{supply_id}|{barcode}",
                        "supply_date": supply_date,
                        "acceptance_warehouse": warehouse,
                        "status_id": status_id,
                        "status_name": STATUS_NAMES.get(status_id, f"Статус {status_id}"),
                        "vendor_code": str(
                            first_nonempty(
                                good.get("vendorCode"),
                                good.get("supplierArticle"),
                                good.get("article"),
                            )
                        ).strip(),
                        "nm_id": to_int(first_nonempty(good.get("nmID"), good.get("nmId")), 0),
                        "barcode": barcode,
                        "tech_size": str(first_nonempty(good.get("techSize"), good.get("size"))).strip(),
                        "planned_qty": planned,
                        "accepted_qty": accepted,
                        "unloading_qty": unloading,
                        "ready_qty_at_acceptance": ready,
                        "acceptance_source": "Supplies API acceptedQuantity",
                    }
                )
        except Exception as exc:
            message = f"Поставка {supply_id}: {exc}"
            warnings.append(message)
            log(message, "WARN")
    if not rows:
        return pd.DataFrame(), warnings
    df = pd.DataFrame(rows)
    numeric_cols = [
        "supply_id",
        "status_id",
        "nm_id",
        "planned_qty",
        "accepted_qty",
        "unloading_qty",
        "ready_qty_at_acceptance",
    ]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    return df, warnings


def finance_movement_kind(row: Dict[str, Any]) -> str:
    doc_type = normalize_text(first_nonempty(row.get("docTypeName"), row.get("doc_type_name")))
    operation = normalize_text(first_nonempty(row.get("sellerOperName"), row.get("supplier_oper_name")))
    if operation:
        if doc_type == "ПРОДАЖА" and operation == "ПРОДАЖА":
            return "sale"
        if doc_type == "ВОЗВРАТ" and operation == "ВОЗВРАТ":
            return "return"
        return "other"
    if doc_type == "ПРОДАЖА":
        return "sale"
    if doc_type == "ВОЗВРАТ":
        return "return"
    return "other"


def aggregate_finance(
    rows: Sequence[Dict[str, Any]],
) -> Tuple[pd.DataFrame, Set[str], Set[str], List[Dict[str, Any]]]:
    aggregate: Dict[Tuple[int, str], Dict[str, Any]] = {}
    sold_srids: Set[str] = set()
    returned_srids: Set[str] = set()
    sale_operations: List[Dict[str, Any]] = []
    seen_rrd: Set[int] = set()

    for row in rows:
        rrd_id = to_int(first_nonempty(row.get("rrdId"), row.get("rrd_id")), 0)
        if rrd_id and rrd_id in seen_rrd:
            continue
        if rrd_id:
            seen_rrd.add(rrd_id)

        gi_id = to_int(first_nonempty(row.get("giId"), row.get("gi_id")), 0)
        barcode = normalize_barcode(first_nonempty(row.get("sku"), row.get("barcode")))
        movement = finance_movement_kind(row)
        if not gi_id or not barcode or movement == "other":
            continue
        quantity = max(to_int(row.get("quantity"), 1), 0)
        srid = normalize_srid(row.get("srid"))
        key = (gi_id, barcode)
        item = aggregate.setdefault(
            key,
            {
                "supply_id": gi_id,
                "barcode": barcode,
                "gross_sales": 0,
                "returns": 0,
                "first_sale_date": None,
                "last_sale_date": None,
                "finance_vendor_code": str(first_nonempty(row.get("vendorCode"), row.get("sa_name"))).strip(),
                "finance_nm_id": to_int(first_nonempty(row.get("nmId"), row.get("nm_id")), 0),
                "finance_title": str(row.get("title", "")).strip(),
                "finance_tech_size": str(first_nonempty(row.get("techSize"), row.get("ts_name"))).strip(),
            },
        )
        event_date = parse_date(
            first_nonempty(
                row.get("saleDt"),
                row.get("sale_dt"),
                row.get("rrDate"),
                row.get("rr_dt"),
            )
        )
        if movement == "sale":
            item["gross_sales"] += quantity
            if srid:
                sold_srids.add(srid)
            if event_date:
                if item["first_sale_date"] is None or event_date < item["first_sale_date"]:
                    item["first_sale_date"] = event_date
                if item["last_sale_date"] is None or event_date > item["last_sale_date"]:
                    item["last_sale_date"] = event_date
        elif movement == "return":
            item["returns"] += quantity
            if srid:
                returned_srids.add(srid)

        sale_operations.append(
            {
                "rrd_id": rrd_id,
                "supply_id": gi_id,
                "barcode": barcode,
                "movement": movement,
                "quantity": quantity,
                "srid": srid,
                "event_date": event_date,
                "office_name": str(row.get("officeName", "")).strip(),
            }
        )

    result_rows: List[Dict[str, Any]] = []
    for item in aggregate.values():
        item["net_sales"] = item["gross_sales"] - item["returns"]
        result_rows.append(item)
    return pd.DataFrame(result_rows), sold_srids, returned_srids, sale_operations


def flatten_current_stocks(raw_items: Sequence[Dict[str, Any]], snapshot_date: dt.date) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for item in raw_items:
        barcode = normalize_barcode(item.get("barcode"))
        if not barcode:
            continue
        warehouses = item.get("warehouses") or []
        common = {
            "snapshot_date": snapshot_date,
            "vendor_code": str(item.get("vendorCode", "")).strip(),
            "nm_id": to_int(item.get("nmId"), 0),
            "barcode": barcode,
            "tech_size": str(item.get("techSize", "")).strip(),
            "subject_name": str(item.get("subjectName", "")).strip(),
            "brand": str(item.get("brand", "")).strip(),
        }
        if not warehouses:
            rows.append(
                {
                    **common,
                    "warehouse_raw": "",
                    "warehouse": "",
                    "quantity": 0,
                    "in_way_to_client": 0,
                    "in_way_from_client": 0,
                    "physical_total": 0,
                    "stock_type": "empty",
                }
            )
            continue
        for wh in warehouses:
            raw_name = str(wh.get("warehouseName", "")).strip()
            normalized_name = normalize_text(raw_name)
            quantity = to_int(wh.get("quantity"), 0)
            if normalized_name in TOTAL_WAREHOUSE_NAMES:
                continue
            if normalized_name in IN_WAY_TO_NAMES:
                row_qty, in_to, in_from, stock_type = 0, quantity, 0, "in_way_to"
            elif normalized_name in IN_WAY_FROM_NAMES:
                row_qty, in_to, in_from, stock_type = 0, 0, quantity, "in_way_from"
            else:
                row_qty, in_to, in_from, stock_type = quantity, 0, 0, "warehouse"
            rows.append(
                {
                    **common,
                    "warehouse_raw": raw_name,
                    "warehouse": canonical_warehouse(raw_name),
                    "quantity": row_qty,
                    "in_way_to_client": in_to,
                    "in_way_from_client": in_from,
                    "physical_total": row_qty + in_to + in_from,
                    "stock_type": stock_type,
                }
            )
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# История из недельных файлов S3
# ---------------------------------------------------------------------------


def load_weekly_excels(
    storage: S3Storage,
    prefix: str,
    date_from: dt.date,
    date_to: dt.date,
    label: str,
) -> Tuple[pd.DataFrame, List[str]]:
    warnings: List[str] = []
    frames: List[pd.DataFrame] = []
    keys = [key for key in storage.list_keys(prefix) if key.lower().endswith(".xlsx")]
    log(f"{label}: недельных файлов найдено={len(keys)}")
    for key in keys:
        try:
            data = storage.read_bytes(key)
            df = pd.read_excel(io.BytesIO(data), sheet_name=0)
            if df.empty:
                continue
            df["_source_key"] = key
            frames.append(df)
        except Exception as exc:
            warnings.append(f"{label}: не прочитан {key}: {exc}")
    if not frames:
        return pd.DataFrame(), warnings
    combined = pd.concat(frames, ignore_index=True, sort=False)
    return combined, warnings


def normalize_orders_history(
    raw: pd.DataFrame,
    date_from: dt.date,
    date_to: dt.date,
) -> pd.DataFrame:
    if raw.empty:
        return pd.DataFrame()
    date_col = dataframe_column(raw, ["date", "Дата", "Дата заказа"])
    warehouse_col = dataframe_column(raw, ["warehouseName", "Склад"])
    barcode_col = dataframe_column(raw, ["barcode", "Баркод"])
    srid_col = dataframe_column(raw, ["srid"])
    cancel_col = dataframe_column(raw, ["isCancel", "Отмена"])
    income_col = dataframe_column(raw, ["incomeID", "giId", "Номер поставки"])
    vendor_col = dataframe_column(raw, ["supplierArticle", "Артикул продавца"])
    nm_col = dataframe_column(raw, ["nmId", "nmID", "Артикул WB"])

    required = {"date": date_col, "warehouse": warehouse_col, "barcode": barcode_col}
    missing = [name for name, col in required.items() if col is None]
    if missing:
        raise ValueError(f"В истории заказов нет колонок: {', '.join(missing)}")

    rows = pd.DataFrame(
        {
            "order_date": raw[date_col].map(parse_date),
            "warehouse": raw[warehouse_col].map(canonical_warehouse),
            "barcode": raw[barcode_col].map(normalize_barcode),
            "srid": raw[srid_col].map(normalize_srid) if srid_col else "",
            "is_cancel": raw[cancel_col].map(to_bool) if cancel_col else False,
            "income_id": raw[income_col].map(to_int) if income_col else 0,
            "vendor_code": raw[vendor_col].astype(str).str.strip() if vendor_col else "",
            "nm_id": raw[nm_col].map(to_int) if nm_col else 0,
        }
    )
    rows = rows[
        rows["order_date"].notna()
        & rows["order_date"].map(lambda value: date_from <= value <= date_to)
        & rows["warehouse"].astype(str).ne("")
        & rows["barcode"].astype(str).ne("")
    ].copy()
    if rows.empty:
        return rows

    rows["order_key"] = rows.apply(
        lambda row: row["srid"]
        if row["srid"]
        else f"{row['order_date']}|{row['warehouse']}|{row['barcode']}|{row.name}",
        axis=1,
    )
    rows = rows.drop_duplicates(subset=["order_key"], keep="last")
    return rows


def normalize_stock_history(
    raw: pd.DataFrame,
    date_from: dt.date,
    date_to: dt.date,
) -> pd.DataFrame:
    if raw.empty:
        return pd.DataFrame()
    ref_date_col = dataframe_column(raw, ["Дата запроса", "snapshot_date", "Дата"])
    collected_col = dataframe_column(raw, ["Дата сбора", "Дата последнего изменения"])
    warehouse_col = dataframe_column(raw, ["Склад", "warehouseName"])
    barcode_col = dataframe_column(raw, ["Баркод", "barcode"])
    quantity_col = dataframe_column(raw, ["Доступно для продажи", "quantity"])
    in_to_col = dataframe_column(raw, ["В пути к клиенту", "inWayToClient"])
    in_from_col = dataframe_column(raw, ["В пути от клиента", "inWayFromClient"])
    vendor_col = dataframe_column(raw, ["Артикул продавца", "vendorCode"])
    nm_col = dataframe_column(raw, ["Артикул WB", "nmId", "nmID"])

    required = {
        "Дата запроса": ref_date_col,
        "Склад": warehouse_col,
        "Баркод": barcode_col,
        "Доступно для продажи": quantity_col,
    }
    missing = [name for name, col in required.items() if col is None]
    if missing:
        raise ValueError(f"В истории остатков нет колонок: {', '.join(missing)}")

    rows = pd.DataFrame(
        {
            "snapshot_date": raw[ref_date_col].map(parse_date),
            "collected_date": raw[collected_col].map(parse_date) if collected_col else None,
            "warehouse": raw[warehouse_col].map(canonical_warehouse),
            "barcode": raw[barcode_col].map(normalize_barcode),
            "quantity": raw[quantity_col].map(to_int),
            "in_way_to_client": raw[in_to_col].map(to_int) if in_to_col else 0,
            "in_way_from_client": raw[in_from_col].map(to_int) if in_from_col else 0,
            "vendor_code": raw[vendor_col].astype(str).str.strip() if vendor_col else "",
            "nm_id": raw[nm_col].map(to_int) if nm_col else 0,
        }
    )
    rows = rows[
        rows["snapshot_date"].notna()
        & rows["snapshot_date"].map(lambda value: date_from <= value <= date_to)
        & rows["warehouse"].astype(str).ne("")
        & rows["barcode"].astype(str).ne("")
    ].copy()
    if rows.empty:
        return rows
    rows["physical_total"] = (
        rows["quantity"] + rows["in_way_to_client"] + rows["in_way_from_client"]
    )
    rows = rows.drop_duplicates(
        subset=["snapshot_date", "warehouse", "barcode"], keep="last"
    )
    return rows


# ---------------------------------------------------------------------------
# Инциденты и оценка уничтоженного товара
# ---------------------------------------------------------------------------


def resolve_incidents(
    incident_warehouses: Sequence[str],
    manual_dates: Dict[str, Any],
    orders: pd.DataFrame,
    stocks_history: pd.DataFrame,
    current_stocks: pd.DataFrame,
    date_to: dt.date,
    shutdown_confirm_days: int,
    min_active_order_days: int,
    min_active_orders: int,
    detection_lookback_days: int,
    min_active_barcodes: int,
    min_stock_at_stop: int,
) -> Dict[str, IncidentInfo]:
    """Автоматически определяет склады с резкой устойчивой остановкой заказов.

    Если ``incident_warehouses`` пуст, анализируются все реальные склады,
    встречающиеся в истории заказов и остатков. Ручные даты из
    ``manual_dates`` всегда имеют приоритет и добавляют склад в анализ.

    Автоинцидент подтверждается, только если одновременно:
    - до остановки склад был активен в коротком окне;
    - после последнего неотменённого заказа прошло достаточно дней;
    - до остановки был положительный остаток;
    - после остановки WB продолжает показывать положительный остаток.

    Это снижает риск принять обычный низкий спрос за пожар/остановку склада.
    """
    result: Dict[str, IncidentInfo] = {}
    manual_canonical: Dict[str, Any] = {
        canonical_warehouse(key): value for key, value in manual_dates.items()
        if canonical_warehouse(key)
    }

    observed_order_end = date_to
    if not orders.empty and orders["order_date"].notna().any():
        observed_order_end = max(orders["order_date"].dropna().tolist())

    forced = {
        canonical_warehouse(value)
        for value in incident_warehouses
        if canonical_warehouse(value)
    }
    candidates: Set[str] = set(forced) | set(manual_canonical.keys())
    if not forced:
        if not orders.empty and "warehouse" in orders.columns:
            candidates.update(str(v) for v in orders["warehouse"].dropna().tolist())
        if not stocks_history.empty and "warehouse" in stocks_history.columns:
            candidates.update(str(v) for v in stocks_history["warehouse"].dropna().tolist())
        if not current_stocks.empty and "warehouse" in current_stocks.columns:
            candidates.update(str(v) for v in current_stocks["warehouse"].dropna().tolist())

    candidates = {
        canonical_warehouse(value)
        for value in candidates
        if canonical_warehouse(value)
        and canonical_warehouse(value) not in {"В пути к клиенту", "В пути от клиента"}
    }

    rejected: List[str] = []
    for warehouse in sorted(candidates):
        warehouse_orders = (
            orders[(orders["warehouse"] == warehouse) & (~orders["is_cancel"])].copy()
            if not orders.empty else pd.DataFrame()
        )
        order_dates = (
            sorted(set(warehouse_orders["order_date"].dropna().tolist()))
            if not warehouse_orders.empty else []
        )
        last_order_date = max(order_dates) if order_dates else None
        active_order_days = len(order_dates)
        active_orders = len(warehouse_orders)
        quiet_days = (observed_order_end - last_order_date).days if last_order_date else 0

        manual_value = manual_canonical.get(warehouse)
        if manual_value not in {None, "", "auto", "AUTO"}:
            incident_date = parse_date(manual_value)
            if not incident_date:
                raise ValueError(f"Некорректная дата инцидента для {warehouse}: {manual_value}")
            orders_after = (
                int((warehouse_orders["order_date"] >= incident_date).sum())
                if not warehouse_orders.empty else 0
            )
            warning = ""
            confidence = "Высокая"
            if orders_after > 0:
                warning = f"После ручной даты есть неотменённые заказы: {orders_after}"
                confidence = "Средняя"
            result[warehouse] = IncidentInfo(
                warehouse=warehouse,
                incident_date=incident_date,
                source="Дата задана вручную",
                last_order_date=last_order_date,
                active_order_days=active_order_days,
                active_orders=active_orders,
                quiet_days=max((observed_order_end - incident_date).days, 0),
                confidence=confidence,
                warning=warning,
            )
            continue

        if last_order_date is None:
            rejected.append(f"{warehouse}: нет заказов")
            continue

        incident_date = last_order_date + dt.timedelta(days=1)
        lookback_start = last_order_date - dt.timedelta(days=max(detection_lookback_days, 1) - 1)
        recent_orders = warehouse_orders[
            (warehouse_orders["order_date"] >= lookback_start)
            & (warehouse_orders["order_date"] <= last_order_date)
        ].copy()
        recent_order_days = int(recent_orders["order_date"].nunique()) if not recent_orders.empty else 0
        recent_order_count = len(recent_orders)
        recent_barcodes = int(recent_orders["barcode"].nunique()) if not recent_orders.empty else 0

        stock_before = 0
        stock_after = 0
        if not stocks_history.empty:
            before_rows = stocks_history[
                (stocks_history["warehouse"] == warehouse)
                & (stocks_history["snapshot_date"] <= incident_date)
            ]
            if not before_rows.empty:
                last_snapshot = max(before_rows["snapshot_date"].dropna().tolist())
                stock_before = int(
                    before_rows[before_rows["snapshot_date"] == last_snapshot]["quantity"].sum()
                )
            after_rows = stocks_history[
                (stocks_history["warehouse"] == warehouse)
                & (stocks_history["snapshot_date"] >= incident_date)
            ]
            if not after_rows.empty:
                stock_after = int(after_rows["quantity"].max())

        current_stock = 0
        if not current_stocks.empty:
            current_stock = int(
                current_stocks[
                    (current_stocks["warehouse"] == warehouse)
                    & (current_stocks["stock_type"] == "warehouse")
                ]["quantity"].sum()
            )
        stock_after = max(stock_after, current_stock)

        reasons: List[str] = []
        if quiet_days < shutdown_confirm_days:
            reasons.append(f"тишина только {quiet_days} дн.")
        if recent_order_days < min_active_order_days:
            reasons.append(f"активных дней до остановки {recent_order_days}")
        if recent_order_count < min_active_orders:
            reasons.append(f"заказов до остановки {recent_order_count}")
        if recent_barcodes < min_active_barcodes:
            reasons.append(f"активных баркодов {recent_barcodes}")
        if stock_before < min_stock_at_stop:
            reasons.append(f"остаток перед остановкой {stock_before}")
        if stock_after <= 0:
            reasons.append("после остановки положительный остаток не подтверждён")

        if reasons:
            rejected.append(f"{warehouse}: " + "; ".join(reasons))
            continue

        confidence = "Высокая" if quiet_days >= max(shutdown_confirm_days * 2, 7) else "Средняя"
        result[warehouse] = IncidentInfo(
            warehouse=warehouse,
            incident_date=incident_date,
            source=(
                f"Авто: последний заказ {last_order_date.isoformat()}, затем {quiet_days} дней без заказов; "
                f"до остановки {recent_order_count} заказов по {recent_barcodes} баркодам; "
                f"остаток перед/после {stock_before}/{stock_after}"
            ),
            last_order_date=last_order_date,
            active_order_days=active_order_days,
            active_orders=active_orders,
            quiet_days=quiet_days,
            confidence=confidence,
            warning="Дата определяется по резкой остановке заказов, а не по официальному акту WB",
        )

    log(
        "Автоопределение проблемных складов: "
        + (", ".join(sorted(result)) if result else "подтверждённых складов нет")
    )
    if rejected:
        log(f"Автоопределение: отклонено кандидатов={len(rejected)}")
    return result


def estimate_destroyed_stock(
    incidents: Dict[str, IncidentInfo],
    stock_history: pd.DataFrame,
    orders: pd.DataFrame,
    sold_srids: Set[str],
    current_stocks: pd.DataFrame,
) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for warehouse, incident in incidents.items():
        if not incident.incident_date:
            rows.append(
                {
                    "warehouse": warehouse,
                    "incident_date": None,
                    "incident_source": incident.source,
                    "confidence": incident.confidence,
                    "warning": incident.warning,
                    "last_order_date": incident.last_order_date,
                    "snapshot_date": None,
                    "barcode": "",
                    "vendor_code": "",
                    "nm_id": 0,
                    "snapshot_available": 0,
                    "orders_after_snapshot": 0,
                    "confirmed_sales_after_snapshot": 0,
                    "potential_destroyed_conservative": 0,
                    "potential_destroyed_confirmed_only": 0,
                    "current_stock_still_shown": 0,
                    "calculation_status": "Инцидент не подтверждён",
                }
            )
            continue

        wh_history = stock_history[
            (stock_history["warehouse"] == warehouse)
            & (stock_history["snapshot_date"] <= incident.incident_date)
            & (stock_history["quantity"] > 0)
        ].copy() if not stock_history.empty else pd.DataFrame()

        snapshot_date: Optional[dt.date] = None
        snapshot_rows = pd.DataFrame()
        fallback_current = False
        if not wh_history.empty:
            snapshot_date = max(wh_history["snapshot_date"].tolist())
            snapshot_rows = wh_history[wh_history["snapshot_date"] == snapshot_date].copy()
        else:
            current_rows = current_stocks[
                (current_stocks["warehouse"] == warehouse)
                & (current_stocks["stock_type"] == "warehouse")
                & (current_stocks["quantity"] > 0)
            ].copy() if not current_stocks.empty else pd.DataFrame()
            if not current_rows.empty:
                fallback_current = True
                snapshot_date = max(current_rows["snapshot_date"].tolist())
                snapshot_rows = current_rows.rename(
                    columns={"snapshot_date": "_snapshot_date"}
                ).copy()
                snapshot_rows["snapshot_date"] = snapshot_date

        if snapshot_rows.empty:
            rows.append(
                {
                    "warehouse": warehouse,
                    "incident_date": incident.incident_date,
                    "incident_source": incident.source,
                    "confidence": "Низкая",
                    "warning": (incident.warning + "; нет снимка остатков").strip("; "),
                    "last_order_date": incident.last_order_date,
                    "snapshot_date": None,
                    "barcode": "",
                    "vendor_code": "",
                    "nm_id": 0,
                    "snapshot_available": 0,
                    "orders_after_snapshot": 0,
                    "confirmed_sales_after_snapshot": 0,
                    "potential_destroyed_conservative": 0,
                    "potential_destroyed_confirmed_only": 0,
                    "current_stock_still_shown": 0,
                    "calculation_status": "Нет исторического или текущего остатка",
                }
            )
            continue

        end_order_date = incident.last_order_date or (
            incident.incident_date - dt.timedelta(days=1)
        )
        for barcode, group in snapshot_rows.groupby("barcode", dropna=False):
            barcode_text = normalize_barcode(barcode)
            snapshot_available = int(group["quantity"].sum())
            vendor_code = str(first_nonempty(*group.get("vendor_code", pd.Series(dtype=str)).tolist()))
            nm_id = max([to_int(value, 0) for value in group.get("nm_id", pd.Series(dtype=int)).tolist()] or [0])

            matching_orders = orders[
                (orders["warehouse"] == warehouse)
                & (orders["barcode"] == barcode_text)
                & (~orders["is_cancel"])
                & (orders["order_date"] > snapshot_date)
                & (orders["order_date"] <= end_order_date)
            ].copy() if not orders.empty and snapshot_date else pd.DataFrame()

            if matching_orders.empty:
                orders_after = 0
                confirmed_sales = 0
            else:
                orders_after = len(matching_orders)
                confirmed_sales = int(
                    matching_orders["srid"].map(lambda value: value in sold_srids if value else False).sum()
                )

            conservative = max(snapshot_available - orders_after, 0)
            confirmed_only = max(snapshot_available - confirmed_sales, 0)
            current_stale = 0
            if not current_stocks.empty:
                current_stale = int(
                    current_stocks[
                        (current_stocks["warehouse"] == warehouse)
                        & (current_stocks["barcode"] == barcode_text)
                        & (current_stocks["stock_type"] == "warehouse")
                    ]["quantity"].sum()
                )

            confidence = incident.confidence
            warning = incident.warning
            status = "Расчёт по последнему историческому снимку"
            if fallback_current:
                confidence = "Низкая"
                warning = (warning + "; использован текущий остаток вместо снимка до инцидента").strip("; ")
                status = "Резервная оценка по текущему остатку"

            rows.append(
                {
                    "warehouse": warehouse,
                    "incident_date": incident.incident_date,
                    "incident_source": incident.source,
                    "confidence": confidence,
                    "warning": warning,
                    "last_order_date": incident.last_order_date,
                    "snapshot_date": snapshot_date,
                    "barcode": barcode_text,
                    "vendor_code": vendor_code,
                    "nm_id": nm_id,
                    "snapshot_available": snapshot_available,
                    "orders_after_snapshot": orders_after,
                    "confirmed_sales_after_snapshot": confirmed_sales,
                    "potential_destroyed_conservative": conservative,
                    "potential_destroyed_confirmed_only": confirmed_only,
                    "current_stock_still_shown": current_stale,
                    "calculation_status": status,
                }
            )
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Итоговые реестры
# ---------------------------------------------------------------------------


def combine_lots_with_finance(lots: pd.DataFrame, finance: pd.DataFrame) -> pd.DataFrame:
    if lots.empty:
        return lots
    finance_cols = [
        "supply_id",
        "barcode",
        "gross_sales",
        "returns",
        "net_sales",
        "first_sale_date",
        "last_sale_date",
        "finance_vendor_code",
        "finance_nm_id",
        "finance_title",
        "finance_tech_size",
    ]
    if finance.empty:
        merged = lots.copy()
        for col in finance_cols[2:]:
            merged[col] = 0 if col in {"gross_sales", "returns", "net_sales", "finance_nm_id"} else ""
    else:
        merged = lots.merge(finance[finance_cols], on=["supply_id", "barcode"], how="left")
    for col in ("gross_sales", "returns", "net_sales", "finance_nm_id"):
        merged[col] = pd.to_numeric(merged[col], errors="coerce").fillna(0).astype(int)
    for col in (
        "first_sale_date",
        "last_sale_date",
        "finance_vendor_code",
        "finance_title",
        "finance_tech_size",
    ):
        merged[col] = merged[col].fillna("")
    merged["vendor_code"] = merged.apply(
        lambda row: row["vendor_code"] or row["finance_vendor_code"], axis=1
    )
    merged["nm_id"] = merged.apply(
        lambda row: row["nm_id"] or row["finance_nm_id"], axis=1
    )
    merged["tech_size"] = merged.apply(
        lambda row: row["tech_size"] or row["finance_tech_size"], axis=1
    )
    merged["title"] = merged["finance_title"]
    merged["expected_unsold"] = merged["accepted_qty"] - merged["net_sales"]
    merged["lot_status"] = merged.apply(
        lambda row: (
            "Продаж больше принятого — проверить"
            if row["expected_unsold"] < 0
            else "Партия расчётно продана"
            if row["expected_unsold"] == 0
            else "Партия продана частично"
            if row["net_sales"] > 0
            else "Продаж по поставке не найдено"
        ),
        axis=1,
    )
    return merged


def build_warehouse_summary(
    lot_registry: pd.DataFrame,
    current_stocks: pd.DataFrame,
    destroyed: pd.DataFrame,
    incidents: Dict[str, IncidentInfo],
) -> pd.DataFrame:
    acceptance = pd.DataFrame()
    if not lot_registry.empty:
        acceptance = (
            lot_registry.groupby("acceptance_warehouse", dropna=False)
            .agg(
                supplies_count=("supply_id", "nunique"),
                sku_lots_count=("supply_barcode_key", "nunique"),
                accepted_qty=("accepted_qty", "sum"),
                gross_sales=("gross_sales", "sum"),
                returns=("returns", "sum"),
                net_sales=("net_sales", "sum"),
                expected_unsold=("expected_unsold", "sum"),
            )
            .reset_index()
            .rename(columns={"acceptance_warehouse": "warehouse"})
        )

    stock_summary = pd.DataFrame()
    if not current_stocks.empty:
        stock_summary = (
            current_stocks.groupby("warehouse", dropna=False)
            .agg(
                current_available=("quantity", "sum"),
                in_way_to_client=("in_way_to_client", "sum"),
                in_way_from_client=("in_way_from_client", "sum"),
                current_physical_total=("physical_total", "sum"),
                current_barcodes=("barcode", "nunique"),
            )
            .reset_index()
        )

    destroyed_summary = pd.DataFrame()
    if not destroyed.empty:
        valid = destroyed[destroyed["barcode"].astype(str).ne("")].copy()
        if not valid.empty:
            destroyed_summary = (
                valid.groupby("warehouse", dropna=False)
                .agg(
                    incident_date=("incident_date", "max"),
                    potential_destroyed=("potential_destroyed_conservative", "sum"),
                    destroyed_upper_bound=("potential_destroyed_confirmed_only", "sum"),
                    current_stale_stock=("current_stock_still_shown", "sum"),
                    incident_barcodes=("barcode", "nunique"),
                )
                .reset_index()
            )

    warehouses: Set[str] = set()
    for frame in (acceptance, stock_summary, destroyed_summary):
        if not frame.empty:
            warehouses.update(str(value) for value in frame["warehouse"].dropna().tolist())
    warehouses.update(incidents.keys())
    base = pd.DataFrame({"warehouse": sorted(value for value in warehouses if value)})
    for frame in (acceptance, stock_summary, destroyed_summary):
        if not frame.empty:
            base = base.merge(frame, on="warehouse", how="left")

    numeric_cols = [
        "supplies_count",
        "sku_lots_count",
        "accepted_qty",
        "gross_sales",
        "returns",
        "net_sales",
        "expected_unsold",
        "current_available",
        "in_way_to_client",
        "in_way_from_client",
        "current_physical_total",
        "current_barcodes",
        "potential_destroyed",
        "destroyed_upper_bound",
        "current_stale_stock",
        "incident_barcodes",
    ]
    for col in numeric_cols:
        if col not in base.columns:
            base[col] = 0
        base[col] = pd.to_numeric(base[col], errors="coerce").fillna(0).astype(int)

    base["is_incident_warehouse"] = base["warehouse"].isin(incidents.keys())
    base["incident_detected"] = base["warehouse"].map(
        lambda value: bool(incidents.get(value) and incidents[value].incident_date)
    )
    base["incident_source"] = base["warehouse"].map(
        lambda value: incidents[value].source if value in incidents else ""
    )
    base["incident_confidence"] = base["warehouse"].map(
        lambda value: incidents[value].confidence if value in incidents else ""
    )
    base["comment"] = base.apply(
        lambda row: (
            "В пути не привязан API к конкретному реальному складу"
            if row["warehouse"] in {"В пути к клиенту", "В пути от клиента"}
            else "Проблемный склад: текущий остаток вынесен отдельно"
            if row["incident_detected"]
            else "Принято и продано относятся к складу приёмки; текущий остаток — к текущему складу WB"
        ),
        axis=1,
    )
    return base.sort_values(
        by=["incident_detected", "potential_destroyed", "accepted_qty"],
        ascending=[False, False, False],
    )


def build_sku_summary(
    lot_registry: pd.DataFrame,
    current_stocks: pd.DataFrame,
    destroyed: pd.DataFrame,
    incidents: Dict[str, IncidentInfo],
) -> pd.DataFrame:
    if lot_registry.empty:
        return pd.DataFrame()
    base = (
        lot_registry.groupby("barcode", dropna=False)
        .agg(
            vendor_code=("vendor_code", "first"),
            title=("title", "first"),
            nm_id=("nm_id", "max"),
            tech_size=("tech_size", "first"),
            supplies_count=("supply_id", "nunique"),
            accepted_qty=("accepted_qty", "sum"),
            gross_sales=("gross_sales", "sum"),
            returns=("returns", "sum"),
            net_sales=("net_sales", "sum"),
            expected_unsold=("expected_unsold", "sum"),
        )
        .reset_index()
    )

    incident_warehouses = {
        warehouse for warehouse, info in incidents.items() if info.incident_date
    }
    stock_rows: List[Dict[str, Any]] = []
    if not current_stocks.empty:
        for barcode, group in current_stocks.groupby("barcode", dropna=False):
            warehouse_rows = group[group["stock_type"] == "warehouse"]
            in_to = int(group["in_way_to_client"].sum())
            in_from = int(group["in_way_from_client"].sum())
            active = int(
                warehouse_rows[~warehouse_rows["warehouse"].isin(incident_warehouses)][
                    "quantity"
                ].sum()
            )
            stale = int(
                warehouse_rows[warehouse_rows["warehouse"].isin(incident_warehouses)][
                    "quantity"
                ].sum()
            )
            stock_rows.append(
                {
                    "barcode": normalize_barcode(barcode),
                    "current_active_stock": active,
                    "current_stale_incident_stock": stale,
                    "in_way_to_client": in_to,
                    "in_way_from_client": in_from,
                    "current_wb_total_including_stale": active + stale + in_to + in_from,
                }
            )
    stock_df = pd.DataFrame(stock_rows)
    if not stock_df.empty:
        base = base.merge(stock_df, on="barcode", how="left")

    destroyed_df = pd.DataFrame()
    if not destroyed.empty:
        valid = destroyed[destroyed["barcode"].astype(str).ne("")].copy()
        if not valid.empty:
            destroyed_df = (
                valid.groupby("barcode", dropna=False)
                .agg(
                    potential_destroyed=("potential_destroyed_conservative", "sum"),
                    destroyed_upper_bound=("potential_destroyed_confirmed_only", "sum"),
                )
                .reset_index()
            )
            base = base.merge(destroyed_df, on="barcode", how="left")

    numeric_cols = [
        "supplies_count",
        "accepted_qty",
        "gross_sales",
        "returns",
        "net_sales",
        "expected_unsold",
        "current_active_stock",
        "current_stale_incident_stock",
        "in_way_to_client",
        "in_way_from_client",
        "current_wb_total_including_stale",
        "potential_destroyed",
        "destroyed_upper_bound",
    ]
    for col in numeric_cols:
        if col not in base.columns:
            base[col] = 0
        base[col] = pd.to_numeric(base[col], errors="coerce").fillna(0).astype(int)

    base["covered_after_sales"] = (
        base["current_active_stock"]
        + base["in_way_to_client"]
        + base["in_way_from_client"]
        + base["potential_destroyed"]
    )
    base["other_potential_loss"] = (
        base["expected_unsold"] - base["covered_after_sales"]
    ).clip(lower=0)
    base["potential_loss_total"] = (
        base["potential_destroyed"] + base["other_potential_loss"]
    )
    base["status"] = base.apply(
        lambda row: (
            "Есть потенциально уничтоженный товар"
            if row["potential_destroyed"] > 0
            else "Есть прочая необъяснимая разница"
            if row["other_potential_loss"] > 0
            else "Партии расчётно закрыты"
            if row["expected_unsold"] <= 0
            else "Расхождений по консервативной формуле нет"
        ),
        axis=1,
    )
    return base.sort_values(
        by=["potential_loss_total", "accepted_qty"], ascending=[False, False]
    )


# ---------------------------------------------------------------------------
# Оборачиваемость непроданных партий
# ---------------------------------------------------------------------------


def build_unsold_lots_aging(
    lot_registry: pd.DataFrame,
    as_of_date: dt.date,
    aging_threshold_days: int = 120,
) -> pd.DataFrame:
    """Сформировать список только открытых партий.

    Полностью проданные строки (expected_unsold <= 0) не попадают в таблицу.
    Возраст считается от даты приёмки до даты отчёта. Количество непроданного
    остатка относится к строгому ключу supplyID + barcode.
    """
    columns = [
        "aging_alert",
        "age_group",
        "age_days",
        "supply_id",
        "supply_barcode_key",
        "supply_date",
        "acceptance_warehouse",
        "vendor_code",
        "title",
        "nm_id",
        "barcode",
        "tech_size",
        "accepted_qty",
        "gross_sales",
        "returns",
        "net_sales",
        "expected_unsold",
        "first_sale_date",
        "last_sale_date",
        "lot_status",
    ]
    if lot_registry.empty:
        return pd.DataFrame(columns=columns)

    result = lot_registry.copy()
    result["expected_unsold"] = pd.to_numeric(
        result.get("expected_unsold", 0), errors="coerce"
    ).fillna(0).astype(int)
    result = result[result["expected_unsold"] > 0].copy()
    if result.empty:
        return pd.DataFrame(columns=columns)

    result["supply_date"] = pd.to_datetime(
        result["supply_date"], errors="coerce"
    ).dt.date
    result["age_days"] = result["supply_date"].apply(
        lambda value: max((as_of_date - value).days, 0) if value else 0
    )
    threshold = max(int(aging_threshold_days), 1)
    result["aging_alert"] = result["age_days"].apply(
        lambda days: f"ДА — {threshold}+ дней" if days >= threshold else "Нет"
    )

    def age_group(days: int) -> str:
        if days >= threshold:
            return f"{threshold}+ дней"
        if days >= 90:
            return "90–119 дней" if threshold >= 120 else f"90–{threshold - 1} дней"
        if days >= 60:
            return "60–89 дней"
        if days >= 30:
            return "30–59 дней"
        return "До 30 дней"

    result["age_group"] = result["age_days"].apply(age_group)
    for col in columns:
        if col not in result.columns:
            result[col] = ""
    result = result[columns]
    return result.sort_values(
        by=["age_days", "expected_unsold", "supply_date"],
        ascending=[False, False, True],
    ).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def rename_for_excel(df: pd.DataFrame, mapping: Dict[str, str]) -> pd.DataFrame:
    result = df.copy()
    result = result.rename(columns=mapping)
    ordered = [value for value in mapping.values() if value in result.columns]
    remaining = [col for col in result.columns if col not in ordered]
    return result[ordered + remaining]


def build_workbook(
    lot_registry: pd.DataFrame,
    warehouse_summary: pd.DataFrame,
    sku_summary: pd.DataFrame,
    incident_detail: pd.DataFrame,
    current_stocks: pd.DataFrame,
    unsold_aging: pd.DataFrame,
    aging_threshold_days: int,
    warnings: Sequence[str],
    store: str,
    date_from: dt.date,
    date_to: dt.date,
) -> bytes:
    lot_excel = rename_for_excel(
        lot_registry,
        {
            "supply_barcode_key": "Уникальный ключ поставка+баркод",
            "supply_id": "Номер поставки / giId",
            "supply_date": "Дата приёмки",
            "acceptance_warehouse": "Склад приёмки",
            "status_name": "Статус поставки",
            "vendor_code": "Артикул продавца",
            "title": "Название",
            "nm_id": "Артикул WB",
            "barcode": "Баркод",
            "tech_size": "Размер",
            "planned_qty": "Заявлено",
            "accepted_qty": "Фактически принято",
            "gross_sales": "Продажи",
            "returns": "Возвраты",
            "net_sales": "Чистые продажи",
            "expected_unsold": "Расчётно не продано",
            "first_sale_date": "Первая продажа",
            "last_sale_date": "Последняя продажа",
            "lot_status": "Статус партии",
            "acceptance_source": "Источник приёмки",
        },
    )
    warehouse_excel = rename_for_excel(
        warehouse_summary,
        {
            "warehouse": "Склад",
            "supplies_count": "Поставок принято",
            "sku_lots_count": "Строк поставка+баркод",
            "accepted_qty": "Фактически принято",
            "gross_sales": "Продажи",
            "returns": "Возвраты",
            "net_sales": "Чистые продажи",
            "expected_unsold": "Расчётно не продано",
            "current_available": "Текущий остаток WB",
            "in_way_to_client": "В пути к клиенту",
            "in_way_from_client": "В пути от клиента",
            "current_physical_total": "Текущий физический контур",
            "potential_destroyed": "Потенциально уничтожено",
            "destroyed_upper_bound": "Верхняя оценка уничтоженного",
            "current_stale_stock": "Сгоревший остаток, который WB ещё показывает",
            "incident_date": "Дата инцидента",
            "incident_detected": "Инцидент подтверждён расчётом",
            "incident_source": "Источник даты инцидента",
            "incident_confidence": "Надёжность",
            "comment": "Комментарий",
        },
    )
    sku_excel = rename_for_excel(
        sku_summary,
        {
            "vendor_code": "Артикул продавца",
            "title": "Название",
            "nm_id": "Артикул WB",
            "barcode": "Баркод",
            "tech_size": "Размер",
            "supplies_count": "Поставок",
            "accepted_qty": "Фактически принято за период",
            "gross_sales": "Продажи",
            "returns": "Возвраты",
            "net_sales": "Чистые продажи",
            "expected_unsold": "Расчётно не продано",
            "current_active_stock": "Текущий рабочий остаток",
            "current_stale_incident_stock": "Остаток на проблемных складах в API",
            "in_way_to_client": "В пути к клиенту",
            "in_way_from_client": "В пути от клиента",
            "current_wb_total_including_stale": "Полный остаток WB включая проблемные склады",
            "potential_destroyed": "Потенциально уничтожено",
            "destroyed_upper_bound": "Верхняя оценка уничтоженного",
            "other_potential_loss": "Прочая потенциальная потеря",
            "potential_loss_total": "Всего потенциальная потеря",
            "status": "Статус",
        },
    )
    incident_excel = rename_for_excel(
        incident_detail,
        {
            "warehouse": "Склад",
            "incident_date": "Дата инцидента",
            "incident_source": "Как определена дата",
            "confidence": "Надёжность",
            "warning": "Предупреждение",
            "last_order_date": "Последняя дата заказов",
            "snapshot_date": "Дата снимка остатков",
            "vendor_code": "Артикул продавца",
            "nm_id": "Артикул WB",
            "barcode": "Баркод",
            "snapshot_available": "Доступно на снимке",
            "orders_after_snapshot": "Неотменённых заказов после снимка",
            "confirmed_sales_after_snapshot": "Из них подтверждено продажей по srid",
            "potential_destroyed_conservative": "Потенциально уничтожено",
            "potential_destroyed_confirmed_only": "Верхняя оценка",
            "current_stock_still_shown": "WB всё ещё показывает на остатке",
            "calculation_status": "Статус расчёта",
        },
    )
    stocks_excel = rename_for_excel(
        current_stocks,
        {
            "snapshot_date": "Дата снимка",
            "warehouse": "Склад",
            "warehouse_raw": "Склад как вернул WB",
            "vendor_code": "Артикул продавца",
            "nm_id": "Артикул WB",
            "barcode": "Баркод",
            "tech_size": "Размер",
            "quantity": "Доступно для продажи",
            "in_way_to_client": "В пути к клиенту",
            "in_way_from_client": "В пути от клиента",
            "physical_total": "Физический контур",
            "stock_type": "Тип строки",
        },
    )

    unsold_excel = rename_for_excel(
        unsold_aging,
        {
            "aging_alert": "Залежалый остаток",
            "age_group": "Группа возраста",
            "age_days": "Дней с даты приёмки",
            "supply_id": "Номер поставки / giId",
            "supply_barcode_key": "Уникальный ключ поставка+баркод",
            "supply_date": "Дата приёмки",
            "acceptance_warehouse": "Склад приёмки",
            "vendor_code": "Артикул продавца",
            "title": "Название",
            "nm_id": "Артикул WB",
            "barcode": "Баркод",
            "tech_size": "Размер",
            "accepted_qty": "Фактически принято",
            "gross_sales": "Продажи",
            "returns": "Возвраты",
            "net_sales": "Чистые продажи",
            "expected_unsold": "Количество не продано",
            "first_sale_date": "Первая продажа",
            "last_sale_date": "Последняя продажа",
            "lot_status": "Статус партии",
        },
    )
    stale_unsold_excel = unsold_excel[
        pd.to_numeric(
            unsold_excel.get("Дней с даты приёмки", 0), errors="coerce"
        ).fillna(0) >= max(int(aging_threshold_days), 1)
    ].copy()

    methodology_rows = [
        ["Параметр", "Значение / методика"],
        ["Версия", VERSION],
        ["Магазин", store],
        ["Период", f"{date_from.isoformat()} — {date_to.isoformat()}"],
        ["Ключ партии", "Номер поставки / giId + barcode"],
        ["Продажи", "Только финансовые операции Продажа; возвраты вычитаются"],
        ["Расчётно не продано", "Фактически принято − чистые продажи"],
        [
            "Непроданные партии",
            "На отдельный лист попадают только строки supplyID + barcode, где расчётно не продано больше нуля; полностью проданные партии исключаются",
        ],
        [
            "Залежалый остаток",
            f"Отдельный лист содержит непроданные партии возрастом {max(int(aging_threshold_days), 1)} дней и более",
        ],
        ["Текущий физический контур", "Остаток + в пути к клиенту + в пути от клиента"],
        [
            "Потенциально уничтожено",
            "Последний доступный остаток проблемного склада перед остановкой минус неотменённые заказы после снимка до остановки",
        ],
        [
            "Верхняя оценка уничтоженного",
            "Последний остаток минус только заказы, подтверждённые продажей через srid",
        ],
        [
            "Сгоревший остаток в API",
            "Количество, которое WB продолжает показывать на текущем проблемном складе; в рабочий остаток оно не включается",
        ],
        [
            "Внутренние перемещения",
            "WB не раскрывает перемещение партии между складами, поэтому склад приёмки и текущий склад нельзя строго свести один к одному",
        ],
        [
            "Остатки до начала периода",
            "Текущий остаток может включать товар, принятый до даты начала. Поэтому итоговая потеря по товару считается консервативно",
        ],
    ]
    if warnings:
        methodology_rows.append(["Предупреждения", ""])
        methodology_rows.extend([["WARN", item] for item in warnings])
    methodology = pd.DataFrame(methodology_rows[1:], columns=methodology_rows[0])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name
    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            warehouse_excel.to_excel(writer, sheet_name="Сводка по складам", index=False)
            sku_excel.to_excel(writer, sheet_name="Сводка по товарам", index=False)
            lot_excel.to_excel(writer, sheet_name="Реестр поставок", index=False)
            unsold_excel.to_excel(writer, sheet_name="Непроданные партии", index=False)
            stale_unsold_excel.to_excel(writer, sheet_name="Непродано 120+ дней", index=False)
            incident_excel.to_excel(writer, sheet_name="Проблемные склады", index=False)
            stocks_excel.to_excel(writer, sheet_name="Текущие остатки", index=False)
            methodology.to_excel(writer, sheet_name="Методика", index=False)

        workbook = load_workbook(path)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        incident_fill = PatternFill("solid", fgColor="F4CCCC")
        warning_fill = PatternFill("solid", fgColor="FFF2CC")
        good_fill = PatternFill("solid", fgColor="D9EAD3")

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=False)
            for column_cells in sheet.columns:
                letter = get_column_letter(column_cells[0].column)
                max_length = 0
                for cell in column_cells[: min(len(column_cells), 5000)]:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, min(len(value), 70))
                sheet.column_dimensions[letter].width = max(10, min(max_length + 2, 45))

        for sheet_name in ("Сводка по складам", "Сводка по товарам", "Проблемные склады"):
            sheet = workbook[sheet_name]
            headers = {str(cell.value): cell.column for cell in sheet[1]}
            for header in (
                "Потенциально уничтожено",
                "Всего потенциальная потеря",
                "Прочая потенциальная потеря",
                "Сгоревший остаток, который WB ещё показывает",
            ):
                col = headers.get(header)
                if col:
                    letter = get_column_letter(col)
                    sheet.conditional_formatting.add(
                        f"{letter}2:{letter}{sheet.max_row}",
                        CellIsRule(operator="greaterThan", formula=["0"], fill=incident_fill),
                    )
            status_col = headers.get("Статус")
            if status_col:
                for row_num in range(2, sheet.max_row + 1):
                    value = str(sheet.cell(row=row_num, column=status_col).value or "")
                    if "Расхождений" in value or "закрыты" in value:
                        sheet.cell(row=row_num, column=status_col).fill = good_fill
                    elif "потер" in value.lower() or "уничтож" in value.lower():
                        sheet.cell(row=row_num, column=status_col).fill = incident_fill

        for sheet_name in ("Непроданные партии", "Непродано 120+ дней"):
            sheet = workbook[sheet_name]
            headers = {str(cell.value): cell.column for cell in sheet[1]}
            days_col = headers.get("Дней с даты приёмки")
            unsold_col = headers.get("Количество не продано")
            alert_col = headers.get("Залежалый остаток")
            if sheet.max_row >= 2 and days_col:
                letter = get_column_letter(days_col)
                sheet.conditional_formatting.add(
                    f"{letter}2:{letter}{sheet.max_row}",
                    CellIsRule(
                        operator="greaterThanOrEqual",
                        formula=[str(max(int(aging_threshold_days), 1))],
                        fill=incident_fill,
                    ),
                )
            if sheet.max_row >= 2 and unsold_col:
                letter = get_column_letter(unsold_col)
                sheet.conditional_formatting.add(
                    f"{letter}2:{letter}{sheet.max_row}",
                    CellIsRule(operator="greaterThan", formula=["0"], fill=warning_fill),
                )
            if sheet.max_row >= 2 and alert_col:
                for row_num in range(2, sheet.max_row + 1):
                    value = str(sheet.cell(row=row_num, column=alert_col).value or "")
                    if value.startswith("ДА"):
                        sheet.cell(row=row_num, column=alert_col).fill = incident_fill

        methodology_sheet = workbook["Методика"]
        methodology_sheet.column_dimensions["A"].width = 32
        methodology_sheet.column_dimensions["B"].width = 120
        for row in methodology_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            if str(row[0].value or "") == "WARN":
                for cell in row:
                    cell.fill = warning_fill

        # Баркоды всегда сохраняем как текст.
        for sheet_name in ("Сводка по товарам", "Реестр поставок", "Непроданные партии", "Непродано 120+ дней", "Проблемные склады", "Текущие остатки"):
            sheet = workbook[sheet_name]
            barcode_col = next(
                (cell.column for cell in sheet[1] if str(cell.value) == "Баркод"),
                None,
            )
            if barcode_col:
                for row_num in range(2, sheet.max_row + 1):
                    sheet.cell(row=row_num, column=barcode_col).number_format = "@"

        workbook.save(path)
        return Path(path).read_bytes()
    finally:
        try:
            os.remove(path)
        except OSError:
            pass


# ---------------------------------------------------------------------------
# Запуск
# ---------------------------------------------------------------------------


def resolve_tokens(store: str) -> Tuple[str, str, str, str]:
    store = store.upper().strip()
    if store not in STORE_MAIN_TOKEN_ENV:
        raise ValueError(f"Неизвестный магазин {store}")
    main_env = STORE_MAIN_TOKEN_ENV[store]
    finance_env = STORE_FINANCE_TOKEN_ENV[store]
    main_token = os.getenv(main_env, "").strip()
    finance_token = os.getenv(finance_env, "").strip() or main_token
    if not main_token:
        raise ValueError(f"Не задан {main_env}")
    return main_token, finance_token, main_env, finance_env


def resolve_storage() -> S3Storage:
    access = os.getenv("YC_ACCESS_KEY_ID", "").strip()
    secret = os.getenv("YC_SECRET_ACCESS_KEY", "").strip()
    bucket = os.getenv("YC_BUCKET_NAME", "").strip()
    endpoint = os.getenv("YC_ENDPOINT_URL", "").strip() or "https://storage.yandexcloud.net"
    missing = [
        name
        for name, value in (
            ("YC_ACCESS_KEY_ID", access),
            ("YC_SECRET_ACCESS_KEY", secret),
            ("YC_BUCKET_NAME", bucket),
        )
        if not value
    ]
    if missing:
        raise ValueError("Не заданы: " + ", ".join(missing))
    return S3Storage(access, secret, bucket, endpoint)


def run(args: argparse.Namespace) -> int:
    load_report_env()
    date_from = parse_date(args.date_from)
    date_to = parse_date(args.date_to) if args.date_to else today_msk()
    if not date_from or not date_to:
        raise ValueError("Нужны корректные даты YYYY-MM-DD")
    if date_from > date_to:
        raise ValueError("Дата начала больше даты окончания")

    store = args.store.upper().strip()
    main_token, finance_token, main_env, finance_env = resolve_tokens(store)
    storage = resolve_storage()
    client = WBClient(main_token, finance_token)

    incident_warehouses = [
        canonical_warehouse(value)
        for value in str(args.incident_warehouses or "").split(",")
        if canonical_warehouse(value)
    ]
    manual_dates = parse_json_object(args.incident_dates_json, "incident-dates-json")

    log(f"VERSION: {VERSION}")
    log(f"Магазин: {store}")
    log(f"Период: {date_from} — {date_to}")
    log(f"Основной токен: {main_env}; finance token: {finance_env if os.getenv(finance_env) else main_env}")
    log("Поиск проблемных складов: автоматически по всем складам" if not incident_warehouses else f"Поиск ограничен складами: {', '.join(incident_warehouses)}")

    if args.dry_run:
        return 0

    # 1. Поставки и принятые товары.
    supplies = client.list_supplies(date_from, date_to)
    log(f"Уникальных поставок API: {len(supplies)}")
    lots, warnings = build_supply_lots(client, supplies, date_from, date_to)
    if lots.empty:
        raise RuntimeError("Не удалось сформировать реестр поставок")
    log(
        f"Реестр поставок: строк supplyID+barcode={len(lots)}, принято={int(lots['accepted_qty'].sum())}"
    )

    # 2. Финансы и точные продажи по giId + barcode.
    finance_raw = client.get_finance_details(date_from, date_to, args.finance_page_limit)
    finance, sold_srids, returned_srids, _ = aggregate_finance(finance_raw)
    tracked_ids = set(lots["supply_id"].astype(int).tolist())
    tracked_barcodes = set(lots["barcode"].astype(str).tolist())
    if not finance.empty:
        finance = finance[
            finance["supply_id"].isin(tracked_ids)
            & finance["barcode"].isin(tracked_barcodes)
        ].copy()
    lot_registry = combine_lots_with_finance(lots, finance)
    log(
        f"Продажи: gross={int(lot_registry['gross_sales'].sum())}, "
        f"returns={int(lot_registry['returns'].sum())}, "
        f"net={int(lot_registry['net_sales'].sum())}"
    )

    # 3. Текущие остатки.
    current_raw = client.get_current_warehouse_remains()
    current_stocks = flatten_current_stocks(current_raw, date_to)
    log(
        f"Текущие остатки: строк={len(current_stocks)}, "
        f"физический контур={int(current_stocks['physical_total'].sum()) if not current_stocks.empty else 0}"
    )

    # 4. История заказов и остатков из Yandex Object Storage.
    orders_prefix = f"Отчёты/Заказы/{store}/Недельные/"
    stocks_prefix = f"Отчёты/Остатки/{store}/Недельные/"
    raw_orders, order_warnings = load_weekly_excels(
        storage, orders_prefix, date_from, date_to, "История заказов"
    )
    raw_stocks, stock_warnings = load_weekly_excels(
        storage, stocks_prefix, date_from, date_to, "История остатков"
    )
    warnings.extend(order_warnings)
    warnings.extend(stock_warnings)

    try:
        orders = normalize_orders_history(raw_orders, date_from, date_to)
    except Exception as exc:
        warnings.append(f"История заказов не нормализована: {exc}")
        orders = pd.DataFrame()
    try:
        stocks_history = normalize_stock_history(raw_stocks, date_from, date_to)
    except Exception as exc:
        warnings.append(f"История остатков не нормализована: {exc}")
        stocks_history = pd.DataFrame()
    log(f"История: заказов={len(orders)}, снимков остатков={len(stocks_history)}")

    # 5. Инциденты и потенциально уничтоженный товар.
    incidents = resolve_incidents(
        incident_warehouses,
        manual_dates,
        orders,
        stocks_history,
        current_stocks,
        date_to,
        args.shutdown_confirm_days,
        args.min_active_order_days,
        args.min_active_orders,
        args.incident_detection_lookback_days,
        args.min_active_barcodes,
        args.min_stock_at_stop,
    )
    incident_detail = estimate_destroyed_stock(
        incidents,
        stocks_history,
        orders,
        sold_srids,
        current_stocks,
    )

    # 6. Сводки.
    warehouse_summary = build_warehouse_summary(
        lot_registry, current_stocks, incident_detail, incidents
    )
    sku_summary = build_sku_summary(
        lot_registry, current_stocks, incident_detail, incidents
    )

    # 7. Оборачиваемость только открытых партий.
    unsold_aging = build_unsold_lots_aging(
        lot_registry,
        date_to,
        args.aging_threshold_days,
    )
    stale_unsold_count = int(
        (pd.to_numeric(unsold_aging.get("age_days", 0), errors="coerce").fillna(0)
         >= max(int(args.aging_threshold_days), 1)).sum()
    ) if not unsold_aging.empty else 0
    stale_unsold_qty = int(
        pd.to_numeric(
            unsold_aging.loc[
                pd.to_numeric(unsold_aging["age_days"], errors="coerce").fillna(0)
                >= max(int(args.aging_threshold_days), 1),
                "expected_unsold",
            ],
            errors="coerce",
        ).fillna(0).sum()
    ) if not unsold_aging.empty else 0
    log(
        f"Непроданные партии: строк={len(unsold_aging)}, "
        f"старше порога={stale_unsold_count}, непродано в них={stale_unsold_qty}"
    )

    # 8. Excel.
    xlsx = build_workbook(
        lot_registry,
        warehouse_summary,
        sku_summary,
        incident_detail,
        current_stocks,
        unsold_aging,
        args.aging_threshold_days,
        warnings,
        store,
        date_from,
        date_to,
    )
    filename = (
        f"Реестр_поставок_продаж_остатков_потерь_{store}_"
        f"{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
    )
    month = date_to.strftime("%Y-%m")
    output_key = f"Отчёты/Контроль потерь/{store}/{month}/{filename}"
    current_key = f"Отчёты/Контроль потерь/{store}/Текущий_реестр.xlsx"
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    storage.write_bytes(output_key, xlsx, content_type)
    storage.write_bytes(current_key, xlsx, content_type)
    log(f"Excel сохранён: {output_key}")

    accepted_total = int(lot_registry["accepted_qty"].sum())
    net_sales_total = int(lot_registry["net_sales"].sum())
    current_active = int(sku_summary["current_active_stock"].sum()) if not sku_summary.empty else 0
    in_way_to = int(sku_summary["in_way_to_client"].sum()) if not sku_summary.empty else 0
    in_way_from = int(sku_summary["in_way_from_client"].sum()) if not sku_summary.empty else 0
    destroyed_total = int(sku_summary["potential_destroyed"].sum()) if not sku_summary.empty else 0
    other_loss = int(sku_summary["other_potential_loss"].sum()) if not sku_summary.empty else 0
    stale_total = int(sku_summary["current_stale_incident_stock"].sum()) if not sku_summary.empty else 0

    caption = (
        f"Реестр WB {store}: {date_from.isoformat()} — {date_to.isoformat()}\n"
        f"Принято: {accepted_total} шт.\n"
        f"Чистые продажи: {net_sales_total} шт.\n"
        f"Рабочий остаток: {current_active} шт.\n"
        f"В пути к клиенту/от клиента: {in_way_to}/{in_way_from} шт.\n"
        f"Потенциально уничтожено: {destroyed_total} шт.\n"
        f"WB ещё показывает на проблемных складах: {stale_total} шт.\n"
        f"Прочая потенциальная разница: {other_loss} шт.\n"
        f"Непроданных партий: {len(unsold_aging)}\n"
        f"Непродано {max(int(args.aging_threshold_days), 1)}+ дней: "
        f"{stale_unsold_qty} шт. в {stale_unsold_count} партиях"
    )

    if not args.no_telegram:
        if not send_telegram_document(filename, xlsx, caption):
            return 2
    else:
        local_path = Path(args.local_output or f"/tmp/{filename}")
        local_path.parent.mkdir(parents=True, exist_ok=True)
        local_path.write_bytes(xlsx)
        log(f"Telegram отключён. Локальная копия: {local_path}")

    log(
        f"Готово: принято={accepted_total}, net_sales={net_sales_total}, "
        f"active_stock={current_active}, destroyed={destroyed_total}, other_loss={other_loss}"
    )
    return 0


# ---------------------------------------------------------------------------
# Самопроверка
# ---------------------------------------------------------------------------


def self_test() -> int:
    assert canonical_warehouse("Рязань (Тюшевское)") == "Рязань"
    assert canonical_warehouse("В пути до получателей") == "В пути к клиенту"
    assert normalize_barcode(8681217250383.0) == "8681217250383"

    finance_raw = [
        {
            "rrdId": 1,
            "giId": 100,
            "sku": "111",
            "docTypeName": "Продажа",
            "sellerOperName": "Продажа",
            "quantity": 1,
            "srid": "A",
        },
        {
            "rrdId": 2,
            "giId": 100,
            "sku": "111",
            "docTypeName": "Продажа",
            "sellerOperName": "Продажа",
            "quantity": 1,
            "srid": "B",
        },
        {
            "rrdId": 3,
            "giId": 100,
            "sku": "111",
            "docTypeName": "Возврат",
            "sellerOperName": "Возврат",
            "quantity": 1,
            "srid": "B",
        },
    ]
    finance, sold, returned, _ = aggregate_finance(finance_raw)
    assert int(finance.iloc[0]["gross_sales"]) == 2
    assert int(finance.iloc[0]["returns"]) == 1
    assert int(finance.iloc[0]["net_sales"]) == 1
    assert sold == {"A", "B"}
    assert returned == {"B"}

    lots = pd.DataFrame(
        [
            {
                "supply_id": 100,
                "supply_barcode_key": "100|111",
                "supply_date": dt.date(2026, 1, 10),
                "acceptance_warehouse": "Коледино",
                "status_id": 5,
                "status_name": "Принято",
                "vendor_code": "X",
                "nm_id": 1,
                "barcode": "111",
                "tech_size": "0",
                "planned_qty": 10,
                "accepted_qty": 10,
                "unloading_qty": 10,
                "ready_qty_at_acceptance": 10,
                "acceptance_source": "test",
            }
        ]
    )
    registry = combine_lots_with_finance(lots, finance)
    assert int(registry.iloc[0]["expected_unsold"]) == 9
    aging = build_unsold_lots_aging(registry, dt.date(2026, 8, 5), 120)
    assert len(aging) == 1
    assert int(aging.iloc[0]["age_days"]) == 207
    assert int(aging.iloc[0]["expected_unsold"]) == 9
    sold_out = registry.copy()
    sold_out["expected_unsold"] = 0
    assert build_unsold_lots_aging(sold_out, dt.date(2026, 8, 5), 120).empty

    current = pd.DataFrame(
        [
            {
                "snapshot_date": dt.date(2026, 8, 5),
                "warehouse": "Коледино",
                "warehouse_raw": "Коледино",
                "vendor_code": "X",
                "nm_id": 1,
                "barcode": "111",
                "tech_size": "0",
                "quantity": 4,
                "in_way_to_client": 0,
                "in_way_from_client": 0,
                "physical_total": 4,
                "stock_type": "warehouse",
            },
            {
                "snapshot_date": dt.date(2026, 8, 5),
                "warehouse": "В пути к клиенту",
                "warehouse_raw": "В пути до получателей",
                "vendor_code": "X",
                "nm_id": 1,
                "barcode": "111",
                "tech_size": "0",
                "quantity": 0,
                "in_way_to_client": 2,
                "in_way_from_client": 0,
                "physical_total": 2,
                "stock_type": "in_way_to",
            },
        ]
    )
    sku = build_sku_summary(registry, current, pd.DataFrame(), {})
    assert int(sku.iloc[0]["other_potential_loss"]) == 3
    print("SELF-TEST OK")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="WB: реестр поставок, продаж, остатков и потенциально уничтоженного товара"
    )
    parser.add_argument("--store", default=os.getenv("WB_STORE", "TOPFACE"))
    parser.add_argument(
        "--date-from",
        default=os.getenv("WB_LOSS_REGISTRY_DATE_FROM", "2026-01-01"),
    )
    parser.add_argument(
        "--date-to",
        default=os.getenv("WB_LOSS_REGISTRY_DATE_TO", ""),
    )
    parser.add_argument(
        "--incident-warehouses",
        default=os.getenv("WB_INCIDENT_WAREHOUSES", DEFAULT_INCIDENT_WAREHOUSES),
        help="Необязательное ограничение списка складов. Пусто = автоматический анализ всех складов",
    )
    parser.add_argument(
        "--incident-dates-json",
        default=os.getenv("WB_INCIDENT_DATES_JSON", "{}"),
    )
    parser.add_argument(
        "--shutdown-confirm-days",
        type=int,
        default=int(os.getenv("WB_SHUTDOWN_CONFIRM_DAYS", "3")),
    )
    parser.add_argument(
        "--min-active-order-days",
        type=int,
        default=int(os.getenv("WB_MIN_ACTIVE_ORDER_DAYS", "2")),
    )
    parser.add_argument(
        "--min-active-orders",
        type=int,
        default=int(os.getenv("WB_MIN_ACTIVE_ORDERS", "5")),
    )
    parser.add_argument(
        "--incident-detection-lookback-days",
        type=int,
        default=int(os.getenv("WB_INCIDENT_DETECTION_LOOKBACK_DAYS", "14")),
        help="Окно активности до остановки склада, дней",
    )
    parser.add_argument(
        "--min-active-barcodes",
        type=int,
        default=int(os.getenv("WB_MIN_ACTIVE_BARCODES", "3")),
        help="Минимум разных баркодов с заказами до остановки",
    )
    parser.add_argument(
        "--min-stock-at-stop",
        type=int,
        default=int(os.getenv("WB_MIN_STOCK_AT_STOP", "10")),
        help="Минимальный суммарный остаток склада перед остановкой",
    )
    parser.add_argument(
        "--finance-page-limit",
        type=int,
        default=int(os.getenv("WB_FINANCE_PAGE_LIMIT", "100000")),
    )
    parser.add_argument(
        "--aging-threshold-days",
        type=int,
        default=int(os.getenv("WB_AGING_THRESHOLD_DAYS", "120")),
        help="Порог залежалой непроданной партии в днях; по умолчанию 120",
    )
    parser.add_argument("--no-telegram", action="store_true")
    parser.add_argument("--local-output", default="")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    if args.self_test:
        return self_test()
    try:
        return run(args)
    except KeyboardInterrupt:
        log("Остановлено пользователем", "ERROR")
        return 130
    except Exception as exc:
        log(f"Критическая ошибка: {exc}", "ERROR")
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
