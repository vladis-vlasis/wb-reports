#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Отдельный сводный отчёт Ozon FBO: остатки, оборачиваемость, кластеры и прогноз OOS.

Это САМОСТОЯТЕЛЬНЫЙ скрипт. Он не запускает рекламу, финансы, отзывы и прочие
модули большого сборщика.

Главный результат в Object Storage:
  Сводные отчёты/<STORE>/Остатки_и_оборачиваемость_YYYY-MM-DD.xlsx
  Сводные отчёты/<STORE>/Последний.xlsx

Что считает:
- среднесуточные продажи за 7 дней по артикулу и по кластеру;
- доступный остаток FBO сейчас;
- товары в пути внутри Ozon и возвраты покупателей на склад;
- все активные FBO-заявки продавца (и уже переданные Ozon, и ещё ожидающие отгрузки);
- запас в днях сейчас и с учётом пути/активных заявок;
- процент выкупа за 30 дней без ранних отмен и без обычных ClientReturn;
- прогноз дефицита по кластерам;
- норматив кросс-дока из «Москва — Кавказский» из Базы знаний Ozon + 4 дня на сборку;
- консервативный срок до доступности товара: после прибытия добавляем до 5 рабочих дней на приёмку/размещение Ozon.

Важно по кросс-доку:
с 16.02.2026 Ozon для кросс-докинговых поставок не возвращает
storage_warehouse.arrival_date. Поэтому прогноз строится по сроку маршрута из
официальной Базы знаний Ozon и фактическому статусу/слоту существующей заявки.
Если срок маршрута не удалось прочитать ни из БЗ, ни из кэша/override, скрипт
НЕ выдумывает срок и помечает прогноз как «Нет срока маршрута».
"""
from __future__ import annotations

import argparse
import io
import json
import logging
import math
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from html import unescape
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple
from urllib.parse import urljoin
from zoneinfo import ZoneInfo

import boto3
import pandas as pd
import requests
from botocore.client import Config
from botocore.exceptions import ClientError
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_VERSION = "OZON_STOCK_TURNOVER_SUMMARY_STANDALONE_V1_6_20260807"
MOSCOW_TZ = ZoneInfo("Europe/Moscow")
OZON_API_BASE = "https://api-seller.ozon.ru"
DEFAULT_BUCKET = "ozon-assist"
SELLER_MIN_INTERVAL_SECONDS = 0.55
CROSSDOCK_ORIGIN = "Москва — Кавказский"
CROSSDOCK_KB_URL = "https://seller-edu.ozon.ru/fbo/crossdoking/delivery-time"
# Публичная копия XLSX из статьи MPSTATS, где источником указан Ozon Knowledge Base.
# Через публичный API Яндекс Диска GitHub Runner может получить бинарный XLSX без браузера.
CROSSDOCK_XLSX_PUBLIC_KEY = "https://disk.yandex.lt/i/Az_zYiNu-3KHBg"
YANDEX_PUBLIC_DOWNLOAD_API = "https://cloud-api.yandex.net/v1/disk/public/resources/download"
PREPARATION_DAYS = 4
# Ozon/профильные источники указывают 2–5 рабочих дней на приёмку и размещение
# после прибытия на конечный склад. Для OOS-прогноза используем консервативно 5.
FINAL_ACCEPTANCE_WORKDAYS = 5
CROSSDOCK_CACHE_DAYS = 30

TERMINAL_SUPPLY_STATES = {
    "COMPLETED", "CANCELLED", "CANCELED", "OVERDUE", "REJECTED", "ARCHIVED",
    "REJECTED_AT_SUPPLY_WAREHOUSE", "REPORTS_CONFIRMATION_AWAITING", "REPORT_REJECTED"
}

# Резервное сопоставление физического FBO-склада с кластером. Оно применяется
# только когда API не вернул кластер. Название склада НИКОГДА не становится
# отдельным «кластером» само по себе.
STATIC_WAREHOUSE_CLUSTER_PATTERNS = [
    (("ПЕТРОВСК", "ПУШКИНО", "ДОМОДЕД", "ХОРУГВ", "ЖУКОВСК", "СОФЬИНО", "ИСТРА", "ГРИВНО", "ПАВЛОВСК"), "Москва, МО и Дальние регионы"),
    (("СПБ", "САНКТ", "КОЛПИНО", "ШУШАР", "БУГРЫ"), "Санкт-Петербург и СЗО"),
    (("РОСТОВ",), "Ростов"),
    (("КРАСНОДАР", "НОВОРОССИЙСК", "АДЫГЕЙСК"), "Краснодар"),
    (("КРАСНОЯРСК",), "Красноярск"),
    (("НЕВИННОМЫССК",), "Невинномысск"),
    (("КАЛИНИНГРАД",), "Калининград"),
    (("ПЕРМ",), "Пермь"),
    (("САМАР",), "Самара"),
    (("ЕКАТЕРИНБУРГ",), "Екатеринбург"),
    (("КАЗАН",), "Казань"),
    (("МАХАЧКАЛ",), "Махачкала"),
    (("НОВОСИБИР",), "Новосибирск"),
    (("САРАТОВ",), "Саратов"),
    (("ХАБАРОВСК", "ВЛАДИВОСТОК"), "Дальний Восток"),
    (("ОМСК",), "Омск"),
    (("ТВЕР",), "Тверь"),
    (("ТЮМЕН",), "Тюмень"),
    (("УФА",), "Уфа"),
    (("ЯРОСЛАВ",), "Ярославль"),
    (("ВОЛГОГРАД",), "Волгоград"),
    (("ВОРОНЕЖ",), "Воронеж"),
    (("НИЖНИЙ", "НОВГОРОД"), "Нижний Новгород"),
    (("ОРЕНБУРГ",), "Оренбург"),
    (("АСТАН",), "Казахстан"),
    (("ЕРЕВАН",), "Армения"),
]


ACTIVE_STATES = {
    "DATA_FILLING",
    "READY_TO_SUPPLY",
    "ACCEPTED_AT_SUPPLY_WAREHOUSE",
    "IN_TRANSIT",
    "ACCEPTANCE_AT_STORAGE_WAREHOUSE",
    "REPORTS_CONFIRMATION_AWAITING",
    "REPORT_REJECTED",
}
PHYSICAL_STATES = {
    "ACCEPTED_AT_SUPPLY_WAREHOUSE",
    "IN_TRANSIT",
    "ACCEPTANCE_AT_STORAGE_WAREHOUSE",
}
PRE_HANDOFF_STATES = {"DATA_FILLING", "READY_TO_SUPPLY"}
STATUS_RU = {
    "DATA_FILLING": "Заполняются данные",
    "READY_TO_SUPPLY": "Готова к отгрузке",
    "ACCEPTED_AT_SUPPLY_WAREHOUSE": "Принята в точке отгрузки",
    "IN_TRANSIT": "В пути",
    "ACCEPTANCE_AT_STORAGE_WAREHOUSE": "Приёмка на складе назначения",
    "REPORTS_CONFIRMATION_AWAITING": "Ожидается подтверждение актов",
    "REPORT_REJECTED": "Спор по акту",
    "COMPLETED": "Завершена",
    "CANCELLED": "Отменена",
    "OVERDUE": "Просрочена",
}
REFUSAL_MARKERS = (
    "отказался при вручении", "отказалась при вручении", "отказ при вручении",
    "не забрал", "не забрала", "не забрали", "истек срок хранения",
    "истёк срок хранения", "срок хранения", "отказ от получения",
    "отказался от получения", "отказалась от получения", "не пришел",
    "не пришёл", "не пришла", "не востребован", "невыкуп",
)


def load_report_env() -> None:
    raw = os.getenv("REPORT_ENV", "") or ""
    for line in raw.splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        if line.startswith("export "):
            line = line[7:].strip()
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def parse_date(value: str) -> date:
    return datetime.strptime(value.strip(), "%Y-%m-%d").date()


def resolve_target_date(value: str = "") -> date:
    forced = (value or os.getenv("OZON_TARGET_DATE", "") or "").strip()
    if forced:
        return parse_date(forced)
    return datetime.now(MOSCOW_TZ).date() - timedelta(days=1)


def to_ozon_datetime(d: date, end: bool = False) -> str:
    return f"{d.isoformat()}T{'23:59:59.999Z' if end else '00:00:00.000Z'}"


def norm_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    s = str(value).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def num(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        if isinstance(value, str):
            value = value.replace("\xa0", "").replace(" ", "").replace(",", ".")
        x = float(value)
        return x if math.isfinite(x) else default
    except Exception:
        return default


def int_qty(value: Any) -> int:
    return max(0, int(round(num(value))))


def extract_items(data: Any, paths: Sequence[Sequence[str]]) -> List[Dict[str, Any]]:
    for path in paths:
        cur = data
        ok = True
        for key in path:
            if not isinstance(cur, Mapping) or key not in cur:
                ok = False
                break
            cur = cur[key]
        if ok and isinstance(cur, list):
            return [dict(x) for x in cur if isinstance(x, Mapping)]
    if isinstance(data, list):
        return [dict(x) for x in data if isinstance(x, Mapping)]
    return []


def extract_cursor(data: Any) -> str:
    for path in (("result", "cursor"), ("cursor",), ("result", "last_id"), ("last_id",),
                 ("result", "next_cursor"), ("next_cursor",)):
        cur = data
        ok = True
        for key in path:
            if not isinstance(cur, Mapping) or key not in cur:
                ok = False
                break
            cur = cur[key]
        if ok and cur not in (None, ""):
            return str(cur)
    return ""


def extract_total(data: Any) -> Optional[int]:
    for path in (("result", "total"), ("total",), ("result", "total_count"), ("total_count",)):
        cur = data
        ok = True
        for key in path:
            if not isinstance(cur, Mapping) or key not in cur:
                ok = False
                break
            cur = cur[key]
        if ok:
            try:
                return int(cur)
            except Exception:
                pass
    return None


def walk_dicts(obj: Any) -> Iterable[Mapping[str, Any]]:
    if isinstance(obj, Mapping):
        yield obj
        for v in obj.values():
            yield from walk_dicts(v)
    elif isinstance(obj, list):
        for v in obj:
            yield from walk_dicts(v)


def deep_first(obj: Any, keys: Sequence[str]) -> Any:
    keyset = {k.lower() for k in keys}
    for d in walk_dicts(obj):
        for k, v in d.items():
            if str(k).lower() in keyset and v not in (None, ""):
                return v
    return None


def parse_any_date(value: Any) -> Optional[date]:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    for candidate in (text[:10], text):
        try:
            return datetime.fromisoformat(candidate.replace("Z", "+00:00")).date()
        except Exception:
            pass
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(candidate, fmt).date()
            except Exception:
                pass
    return None


def norm_name(value: Any) -> str:
    s = str(value or "").lower().replace("ё", "е")
    s = re.sub(r"[^a-zа-я0-9]+", " ", s)
    stop = {"кластер", "макролокальный", "макрорегион", "регион", "склад", "озон", "ozon"}
    return " ".join(x for x in s.split() if x not in stop).strip()


def similarity(a: str, b: str) -> float:
    a, b = norm_name(a), norm_name(b)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if a in b or b in a:
        return 0.92
    ta, tb = set(a.split()), set(b.split())
    jaccard = len(ta & tb) / max(1, len(ta | tb))
    return max(jaccard, SequenceMatcher(None, a, b).ratio())


def conservative_days(value: Any) -> Optional[int]:
    """Берём верхнюю границу срока: '2–4' -> 4, 'до 7 дней' -> 7."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, (int, float)):
        x = int(math.ceil(float(value)))
        return x if 0 <= x <= 60 else None
    text = str(value).lower().replace("−", "-").replace("–", "-").replace("—", "-")
    nums = [int(x) for x in re.findall(r"(?<!\d)(\d{1,2})(?!\d)", text)]
    nums = [x for x in nums if 0 <= x <= 60]
    return max(nums) if nums else None


def static_cluster_from_warehouse(value: Any) -> str:
    raw = str(value or "").upper().replace("Ё", "Е")
    raw = re.sub(r"[^A-ZА-Я0-9]+", "_", raw)
    for tokens, cluster in STATIC_WAREHOUSE_CLUSTER_PATTERNS:
        # Для Нижнего Новгорода требуем оба слова, для остальных достаточно одного маркера.
        if tokens == ("НИЖНИЙ", "НОВГОРОД"):
            if all(token in raw for token in tokens):
                return cluster
        elif any(token in raw for token in tokens):
            return cluster
    return ""


def add_workdays(start: date, workdays: int) -> date:
    result = start
    added = 0
    while added < max(0, int(workdays)):
        result += timedelta(days=1)
        if result.weekday() < 5:
            added += 1
    return result


class OzonApiError(RuntimeError):
    def __init__(self, path: str, status: int, message: str):
        super().__init__(f"POST {path}: HTTP {status}: {message}")
        self.path = path
        self.status = status


class OzonClient:
    def __init__(self, client_id: str, api_key: str):
        self.session = requests.Session()
        self.session.headers.update({
            "Client-Id": str(client_id),
            "Api-Key": str(api_key),
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": f"ozon-assist/{SCRIPT_VERSION}",
        })
        self.last_at = 0.0

    def _throttle(self) -> None:
        wait = SELLER_MIN_INTERVAL_SECONDS - (time.monotonic() - self.last_at)
        if wait > 0:
            time.sleep(wait)
        self.last_at = time.monotonic()

    def post(self, path: str, payload: Mapping[str, Any], retries: int = 6) -> Dict[str, Any]:
        url = OZON_API_BASE + path
        for attempt in range(1, retries + 1):
            try:
                self._throttle()
                r = self.session.post(url, json=dict(payload), timeout=180)
            except requests.RequestException as exc:
                if attempt == retries:
                    raise OzonApiError(path, 0, str(exc)) from exc
                time.sleep(min(60, 2 ** attempt))
                continue
            if r.status_code == 200:
                try:
                    return r.json()
                except Exception as exc:
                    raise OzonApiError(path, 200, "Ответ не JSON") from exc
            if r.status_code in {429, 500, 502, 503, 504} and attempt < retries:
                try:
                    sec = int(float(r.headers.get("Retry-After", "")))
                except Exception:
                    sec = min(60, 2 ** attempt)
                logging.warning("%s: HTTP %s, повтор через %s сек", path, r.status_code, max(1, sec))
                time.sleep(max(1, sec))
                continue
            try:
                body = r.json()
                msg = body.get("message") or body.get("error") or body.get("code") or r.text[:1000]
            except Exception:
                msg = r.text[:1000]
            raise OzonApiError(path, r.status_code, str(msg))
        raise OzonApiError(path, 0, "Неизвестная ошибка")

    def offset_pages(self, path: str, payload: Dict[str, Any], paths: Sequence[Sequence[str]],
                     limit: int = 1000, max_pages: int = 200) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        offset = int(payload.get("offset", 0) or 0)
        for _ in range(max_pages):
            body = dict(payload)
            body["limit"] = min(limit, int(body.get("limit", limit)))
            body["offset"] = offset
            data = self.post(path, body)
            got = extract_items(data, paths)
            rows.extend(got)
            total = extract_total(data)
            if not got or len(got) < body["limit"] or (total is not None and len(rows) >= total):
                break
            offset += len(got)
        return rows

    def cursor_pages(self, path: str, payload: Dict[str, Any], paths: Sequence[Sequence[str]],
                     cursor_field: str = "cursor", limit: int = 1000, max_pages: int = 200) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        cursor = str(payload.get(cursor_field, "") or "")
        for _ in range(max_pages):
            body = dict(payload)
            body["limit"] = min(limit, int(body.get("limit", limit)))
            body[cursor_field] = cursor
            data = self.post(path, body)
            got = extract_items(data, paths)
            rows.extend(got)
            nxt = extract_cursor(data)
            if not got or not nxt or nxt == cursor:
                break
            cursor = nxt
        return rows


class Storage:
    def __init__(self, access_key: str, secret_key: str, bucket: str, endpoint: str):
        self.bucket = bucket
        self.client = boto3.client(
            "s3", endpoint_url=endpoint,
            aws_access_key_id=access_key, aws_secret_access_key=secret_key,
            region_name="ru-central1",
            config=Config(signature_version="s3v4", read_timeout=300, connect_timeout=60,
                          retries={"max_attempts": 7, "mode": "standard"}),
        )

    def exists(self, key: str) -> bool:
        try:
            self.client.head_object(Bucket=self.bucket, Key=key)
            return True
        except ClientError as exc:
            if str(exc.response.get("Error", {}).get("Code", "")) in {"404", "NoSuchKey", "NotFound"}:
                return False
            raise

    def read_json(self, key: str) -> Dict[str, Any]:
        if not self.exists(key):
            return {}
        try:
            raw = self.client.get_object(Bucket=self.bucket, Key=key)["Body"].read()
            return json.loads(raw.decode("utf-8"))
        except Exception:
            return {}

    def upload(self, key: str, data: bytes, content_type: str) -> None:
        self.client.put_object(Bucket=self.bucket, Key=key, Body=data, ContentType=content_type)

    def upload_json(self, key: str, value: Any) -> None:
        self.upload(key, json.dumps(value, ensure_ascii=False, indent=2, default=str).encode("utf-8"), "application/json")


@dataclass
class ClusterMaps:
    cluster_names: List[str]
    warehouse_to_cluster: Dict[str, str]
    warehouse_name_to_cluster: Dict[str, str]
    macro_to_cluster: Dict[str, str]
    raw_endpoint: str

    def canonical(self, name: Any) -> str:
        text = str(name or "").strip()
        if not text:
            return ""
        best_name, best_score = text, 0.0
        for candidate in self.cluster_names:
            score = similarity(text, candidate)
            if score > best_score:
                best_name, best_score = candidate, score
        return best_name if best_score >= 0.62 else text


class CrossdockSLAProvider:
    """Сроки кросс-дока из Москва — Кавказский.

    Приоритет источников:
    1) явный OZON_CROSSDOCK_SLA_JSON;
    2) свежий кэш в Object Storage (30 дней);
    3) XLSX, ссылка на который опубликована MPSTATS со ссылкой на БЗ Ozon;
    4) прямое чтение страницы БЗ Ozon;
    5) старый кэш.

    В расчёте используем верхнюю границу диапазона, то есть консервативный срок.
    """
    def __init__(self, storage: Storage, store: str):
        self.storage = storage
        self.store = store
        self.cache_key = f"Сводные отчёты/{store}/Справочники/Сроки_кроссдока_Кавказский.json"
        self.rows: Dict[str, int] = {}
        self.source = ""
        self.updated_at = ""
        self.raw_rows: Dict[str, int] = {}

    def _set(self, mapping: Mapping[str, Any], source: str, updated_at: Optional[str] = None) -> Dict[str, int]:
        rows: Dict[str, int] = {}
        for k, v in mapping.items():
            d = conservative_days(v)
            if str(k).strip() and d is not None:
                rows[str(k).strip()] = d
        # Добавляем агрегаты по кластерам из названий конечных складов.
        cluster_max: Dict[str, int] = {}
        for key, days in rows.items():
            cluster = static_cluster_from_warehouse(key)
            if cluster:
                cluster_max[cluster] = max(cluster_max.get(cluster, 0), days)
        for cluster, days in cluster_max.items():
            rows.setdefault(cluster, days)
        self.raw_rows = dict(rows)
        self.rows = rows
        self.source = source
        self.updated_at = updated_at or datetime.now(MOSCOW_TZ).isoformat()
        return self.rows

    @staticmethod
    def _cache_is_fresh(cached: Mapping[str, Any]) -> bool:
        stamp = str(cached.get("updated_at") or "")
        try:
            dt = datetime.fromisoformat(stamp.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=MOSCOW_TZ)
            return datetime.now(MOSCOW_TZ) - dt.astimezone(MOSCOW_TZ) <= timedelta(days=CROSSDOCK_CACHE_DAYS)
        except Exception:
            return False

    def _save_cache(self) -> None:
        if not self.rows:
            return
        self.storage.upload_json(self.cache_key, {
            "source": self.source,
            "url": CROSSDOCK_KB_URL,
            "xlsx_public_key": CROSSDOCK_XLSX_PUBLIC_KEY,
            "updated_at": self.updated_at,
            "origin": CROSSDOCK_ORIGIN,
            "sla_days": self.rows,
            "note": "В расчёте используется верхняя граница диапазона срока кросс-дока.",
        })

    def load(self) -> Dict[str, int]:
        # Явный override всегда имеет наивысший приоритет.
        raw = os.getenv("OZON_CROSSDOCK_SLA_JSON", "").strip()
        if raw:
            try:
                obj = json.loads(raw)
                if isinstance(obj, Mapping) and obj:
                    result = self._set(obj, "Ручной норматив OZON_CROSSDOCK_SLA_JSON")
                    self._save_cache()
                    logging.info("Сроки кросс-дока: используется ручной норматив (%s маршрутов)", len(result))
                    return result
            except Exception as exc:
                logging.warning("Некорректный OZON_CROSSDOCK_SLA_JSON: %s", exc)

        cached = self.storage.read_json(self.cache_key)
        cached_map = cached.get("sla_days") if isinstance(cached, Mapping) else None
        if isinstance(cached_map, Mapping) and cached_map and self._cache_is_fresh(cached):
            result = self._set(cached_map, str(cached.get("source") or "Кэш сроков кросс-дока"), str(cached.get("updated_at") or ""))
            logging.info("Сроки кросс-дока: свежий кэш (%s маршрутов)", len(result))
            return result

        # Основной автоматический источник — XLSX из материала, где источником указана БЗ Ozon.
        try:
            mapping = self._load_from_yandex_public_xlsx()
            if mapping:
                result = self._set(mapping, "XLSX сроков кросс-дока из БЗ Ozon (публичная копия MPSTATS/Yandex Disk)")
                self._save_cache()
                logging.info("Сроки кросс-дока: XLSX успешно прочитан (%s маршрутов)", len(result))
                return result
        except Exception as exc:
            logging.warning("Не удалось загрузить XLSX сроков кросс-дока: %s", exc)

        # Прямой сайт часто отвечает 403 GitHub Runner'у, поэтому это только резерв.
        try:
            mapping = self._load_from_ozon_kb()
            if mapping:
                result = self._set(mapping, "База знаний Ozon — Москва — Кавказский")
                self._save_cache()
                return result
        except Exception as exc:
            logging.warning("Прямая БЗ Ozon недоступна: %s", exc)

        if isinstance(cached_map, Mapping) and cached_map:
            result = self._set(cached_map, str(cached.get("source") or "Старый кэш сроков кросс-дока"), str(cached.get("updated_at") or ""))
            logging.warning("Сроки кросс-дока: используется старый кэш (%s маршрутов)", len(result))
            return result

        logging.warning("Сроки кросс-дока не получены. Прогноз OOS будет рассчитан только там, где есть фактическая дата/слот поставки.")
        return {}

    def lookup(self, cluster: str) -> Tuple[Optional[int], str]:
        if not cluster or not self.rows:
            return None, self.source or "Срок не найден"
        # Сначала точное/нормализованное совпадение.
        target = norm_name(cluster)
        for key, days in self.rows.items():
            if norm_name(key) == target:
                return int(days), self.source
        best_key, best_score = "", 0.0
        for key in self.rows:
            score = similarity(cluster, key)
            if score > best_score:
                best_key, best_score = key, score
        if best_key and best_score >= 0.60:
            return int(self.rows[best_key]), self.source
        return None, self.source or "Срок не найден"

    def as_dataframe(self) -> pd.DataFrame:
        rows = [{
            "Точка отправления": CROSSDOCK_ORIGIN,
            "Кластер / конечный склад": key,
            "Срок кросс-дока, дней (верхняя граница)": days,
            "Источник": self.source,
            "Обновлено": self.updated_at,
        } for key, days in sorted(self.rows.items())]
        return pd.DataFrame(rows)

    def _load_from_yandex_public_xlsx(self) -> Dict[str, int]:
        api = requests.get(
            YANDEX_PUBLIC_DOWNLOAD_API,
            params={"public_key": CROSSDOCK_XLSX_PUBLIC_KEY},
            headers={"User-Agent": "Mozilla/5.0 (OzonAssist)"},
            timeout=45,
        )
        api.raise_for_status()
        href = str((api.json() if api.content else {}).get("href") or "").strip()
        if not href:
            raise RuntimeError("Яндекс Диск не вернул ссылку download.href")
        fr = requests.get(href, headers={"User-Agent": "Mozilla/5.0 (OzonAssist)"}, timeout=90)
        fr.raise_for_status()
        if len(fr.content) < 1000:
            raise RuntimeError("Слишком маленький ответ вместо XLSX")
        book = pd.ExcelFile(io.BytesIO(fr.content))
        mapping: Dict[str, int] = {}
        for sheet in book.sheet_names:
            table = pd.read_excel(book, sheet_name=sheet, header=None)
            mapping.update(self._extract_from_table(table))
        return mapping

    def _load_from_ozon_kb(self) -> Dict[str, int]:
        headers = {"User-Agent": "Mozilla/5.0 (compatible; OzonAssist/1.0)"}
        r = requests.get(CROSSDOCK_KB_URL, headers=headers, timeout=45, allow_redirects=True)
        r.raise_for_status()
        html = r.text
        tables: List[pd.DataFrame] = []
        try:
            tables.extend(pd.read_html(io.StringIO(html)))
        except Exception:
            pass
        links = set(re.findall(r"""(?:href|src)=["']([^"']+\.(?:xlsx?|csv)(?:\?[^"']*)?)["']""", html, flags=re.I))
        for href in links:
            try:
                url = urljoin(r.url, unescape(href))
                fr = requests.get(url, headers=headers, timeout=45)
                fr.raise_for_status()
                if re.search(r"\.csv(?:\?|$)", url, re.I):
                    tables.append(pd.read_csv(io.BytesIO(fr.content)))
                else:
                    book = pd.ExcelFile(io.BytesIO(fr.content))
                    for sheet in book.sheet_names:
                        tables.append(pd.read_excel(book, sheet_name=sheet, header=None))
            except Exception:
                continue
        mapping: Dict[str, int] = {}
        for table in tables:
            mapping.update(self._extract_from_table(table))
        return mapping

    def _extract_from_table(self, table: pd.DataFrame) -> Dict[str, int]:
        if table is None or table.empty:
            return {}
        df = table.copy().fillna("")
        result: Dict[str, int] = {}

        # 1) Ищем строку/маршрут с Кавказским и пару «назначение — дни» в той же строке.
        # 2) Ищем матричную форму: строка = Кавказский, заголовки столбцов = назначения.
        for idx in range(len(df)):
            vals = [str(x).strip() for x in df.iloc[idx].tolist()]
            origin_cols = [j for j, x in enumerate(vals) if "кавказ" in norm_name(x) and ("моск" in norm_name(x) or "мск" in norm_name(x))]
            if not origin_cols:
                continue
            # Заголовки пробуем взять из нескольких строк выше.
            candidates: List[List[str]] = []
            for h in range(max(0, idx - 5), idx):
                candidates.append([str(x).strip() for x in df.iloc[h].tolist()])
            if not candidates:
                candidates.append([str(x) for x in df.columns])
            headers = max(candidates, key=lambda hs: sum(bool(norm_name(x)) for x in hs))
            for j, value in enumerate(vals):
                if j in origin_cols:
                    continue
                days = conservative_days(value)
                dest = headers[j].strip() if j < len(headers) else ""
                if days is not None and dest and not dest.isdigit() and "unnamed" not in dest.lower():
                    result[dest] = days

        # 3) Обратная ориентация матрицы: Кавказский может быть заголовком СТОЛБЦА,
        # а конечные склады — строками. Это важно, потому что формат XLSX Ozon менялся.
        for origin_row in range(min(len(df), 12)):
            for origin_col in range(df.shape[1]):
                cell = str(df.iat[origin_row, origin_col]).strip()
                ncell = norm_name(cell)
                if "кавказ" not in ncell or not ("моск" in ncell or "мск" in ncell):
                    continue
                for r in range(origin_row + 1, len(df)):
                    days = conservative_days(df.iat[r, origin_col])
                    if days is None:
                        continue
                    # Название назначения ищем слева в той же строке — обычно это
                    # один из первых текстовых столбцов матрицы.
                    dest = ""
                    for c in range(0, origin_col):
                        candidate = str(df.iat[r, c]).strip()
                        if not candidate or conservative_days(candidate) is not None:
                            continue
                        nc = norm_name(candidate)
                        if len(nc) >= 3 and "кавказ" not in nc:
                            dest = candidate
                            break
                    if dest:
                        result[dest] = days

        # Строчная форма: в окрестности Кавказского ищем отдельные поля назначения и срока.
        for idx in range(len(df)):
            vals = [str(x).strip() for x in df.iloc[idx].tolist()]
            if not any("кавказ" in norm_name(x) for x in vals):
                continue
            textual = [x for x in vals if x and conservative_days(x) is None]
            numeric = [(x, conservative_days(x)) for x in vals if conservative_days(x) is not None]
            for dest in textual:
                nd = norm_name(dest)
                if "кавказ" in nd or len(nd) < 3:
                    continue
                # Последнее числовое значение в строке обычно срок; берём консервативно максимум.
                if numeric:
                    result.setdefault(dest, max(d for _, d in numeric if d is not None))

        return result


class ReportBuilder:
    def __init__(self, store: str, target_date: date, client: OzonClient, storage: Storage):
        self.store = store.upper()
        self.target_date = target_date
        self.client = client
        self.storage = storage
        self.offer_by_sku: Dict[str, str] = {}
        self.name_by_sku: Dict[str, str] = {}
        self.sku_ids: List[int] = []
        self.cluster_maps = ClusterMaps([], {}, {}, {}, "")
        self.sla = CrossdockSLAProvider(storage, self.store)
        # FBO-отправления за 30 дней загружаем один раз и переиспользуем
        # для продаж за 7 дней и расчёта выкупа.
        self._postings_30d_cache: Optional[List[Dict[str, Any]]] = None
        self._posting_qty_scale: int = 1
        self._posting_qty_detection: Dict[str, Any] = {}
        self._unknown_warehouses: set[str] = set()
        self._supply_diagnostics: Dict[str, Any] = {}

    # ---------- справочники ----------
    def fetch_products(self) -> None:
        logging.info("1/8 Справочник товаров")
        items = self.client.cursor_pages(
            "/v3/product/list", {"filter": {"visibility": "ALL"}, "limit": 1000, "last_id": ""},
            [("result", "items"), ("items",)], cursor_field="last_id")
        offers = [str(x.get("offer_id")) for x in items if x.get("offer_id") not in (None, "")]
        for i in range(0, len(offers), 100):
            batch = offers[i:i+100]
            data = self.client.post("/v3/product/info/list", {"offer_id": batch, "product_id": [], "sku": []})
            for item in extract_items(data, [("items",), ("result", "items")]):
                offer = str(item.get("offer_id") or item.get("offerId") or "").strip()
                name = str(item.get("name") or item.get("product_name") or "").strip()
                for key in ("sku", "fbo_sku"):
                    sku = norm_id(item.get(key))
                    if sku:
                        if sku.isdigit():
                            self.sku_ids.append(int(sku))
                        if offer:
                            self.offer_by_sku[sku] = offer
                        if name:
                            self.name_by_sku[sku] = name
        self.sku_ids = sorted(set(self.sku_ids))
        logging.info("Товаров SKU в справочнике: %s", len(self.sku_ids))

    def fetch_clusters(self) -> None:
        logging.info("2/8 Кластеры FBO")
        names: List[str] = []
        wh_id: Dict[str, str] = {}
        wh_name: Dict[str, str] = {}
        macro: Dict[str, str] = {}

        def add_name(value: Any) -> str:
            name = str(value or "").strip()
            if name and name not in names:
                names.append(name)
            return name

        def parse_cluster_tree(node: Any, inherited: str = "", parent_key: str = "") -> None:
            if isinstance(node, list):
                for item in node:
                    parse_cluster_tree(item, inherited, parent_key)
                return
            if not isinstance(node, Mapping):
                return

            explicit = str(node.get("macrolocal_cluster_name") or node.get("cluster_name") or "").strip()
            if not explicit and parent_key in {"clusters", "macrolocal_clusters", "logistic_clusters"}:
                explicit = str(node.get("name") or "").strip()
            if not explicit and any(k in node for k in ("warehouses", "logistic_clusters")):
                explicit = str(node.get("name") or "").strip()
            cname = add_name(explicit or inherited) if (explicit or inherited) else ""

            if cname:
                for key in ("macrolocal_cluster_id", "macro_cluster_id", "cluster_id", "id"):
                    cid = norm_id(node.get(key))
                    if cid and parent_key not in {"warehouses", "warehouse"}:
                        macro.setdefault(cid, cname)

            # Если узел явно складской — связываем его с наследованным кластером.
            is_warehouse = parent_key in {"warehouses", "warehouse"} or "warehouse_id" in node or "warehouse_name" in node
            if is_warehouse and cname:
                wid = norm_id(node.get("warehouse_id") or node.get("id"))
                wname = str(node.get("warehouse_name") or node.get("name") or "").strip()
                if wid:
                    wh_id[wid] = cname
                if wname:
                    wh_name[norm_name(wname)] = cname

            for key, value in node.items():
                if isinstance(value, (Mapping, list)):
                    parse_cluster_tree(value, cname or inherited, str(key))

        response: Any = {}
        try:
            response = self.client.post("/v2/cluster/list", {})
            parse_cluster_tree(response)
        except OzonApiError as exc:
            logging.warning("/v2/cluster/list недоступен: %s", exc)

        # Дополнительный актуальный метод складов Ozon. Не делаем его обязательным:
        # структура бета/основной версии могла меняться.
        try:
            wresp = self.client.post("/v1/warehouse/ozon/list", {})
            parse_cluster_tree(wresp)
        except OzonApiError as exc:
            if exc.status not in {400, 403, 404, 405, 422}:
                logging.warning("/v1/warehouse/ozon/list: %s", exc)

        # Добавляем известные названия кластеров, чтобы canonical() мог нормализовать
        # статический warehouse fallback даже если /v2/cluster/list вернул новую схему.
        for _tokens, cluster in STATIC_WAREHOUSE_CLUSTER_PATTERNS:
            add_name(cluster)

        self.cluster_maps = ClusterMaps(names, wh_id, wh_name, macro, "/v2/cluster/list + /v1/warehouse/ozon/list")
        logging.info("Кластеров/нормализованных названий: %s, складов с API-привязкой: %s", len(names), len(wh_id) + len(wh_name))
        if isinstance(response, Mapping) and not wh_id and not wh_name:
            result = response.get("result")
            if isinstance(result, Mapping):
                logging.warning("Кластерный API не дал привязку складов. Поля result: %s", list(result.keys())[:30])
            elif isinstance(result, list):
                first_keys = list(result[0].keys()) if result and isinstance(result[0], Mapping) else []
                logging.warning("Кластерный API: result=list[%s], поля первого элемента: %s", len(result), first_keys)

    # ---------- остатки ----------
    def fetch_stock_warehouses(self) -> pd.DataFrame:
        logging.info("3/8 Остатки FBO по складам")
        rows = self.client.offset_pages(
            "/v2/analytics/stock_on_warehouses",
            {"limit": 1000, "offset": 0, "warehouse_type": "ALL"},
            [("result", "rows"), ("rows",), ("result", "items")], limit=1000)
        out: List[Dict[str, Any]] = []
        for r in rows:
            sku = norm_id(r.get("sku"))
            if not sku:
                continue
            offer = str(r.get("item_code") or r.get("offer_id") or self.offer_by_sku.get(sku, "")).strip()
            name = str(r.get("item_name") or r.get("name") or self.name_by_sku.get(sku, "")).strip()
            wh_name = str(r.get("warehouse_name") or "").strip()
            wh_id = norm_id(r.get("warehouse_id"))
            cluster = str(r.get("cluster_name") or r.get("macrolocal_cluster_name") or r.get("cluster") or "").strip()
            source = "API остатков" if cluster else ""
            if not cluster and wh_id:
                cluster = self.cluster_maps.warehouse_to_cluster.get(wh_id, "")
                if cluster:
                    source = "Справочник API"
            if not cluster and wh_name:
                cluster = self.cluster_maps.warehouse_name_to_cluster.get(norm_name(wh_name), "")
                if cluster:
                    source = "Справочник API"
            if not cluster and wh_name:
                cluster = static_cluster_from_warehouse(wh_name)
                if cluster:
                    source = "Резерв по названию склада"
            if cluster:
                cluster = self.cluster_maps.canonical(cluster)
            else:
                cluster = "Кластер не определён"
                if wh_name:
                    self._unknown_warehouses.add(wh_name)
            out.append({
                "SKU Ozon": sku,
                "Артикул": offer or f"SKU {sku}",
                "Название": name,
                "Кластер": cluster,
                "Склад Ozon": wh_name,
                "Источник кластера": source or "Не определён",
                "Доступно сейчас, шт.": int_qty(r.get("free_to_sell_amount")),
                "Подтверждено Ozon в поставках, шт.": int_qty(r.get("promised_amount")),
                "Зарезервировано Ozon, шт.": int_qty(r.get("reserved_amount")),
            })
            if offer:
                self.offer_by_sku[sku] = offer
            if name:
                self.name_by_sku[sku] = name
        logging.info("Строк остатков: %s; складов без кластера: %s", len(out), len(self._unknown_warehouses))
        return pd.DataFrame(out)

    def fetch_transit_stocks(self) -> pd.DataFrame:
        logging.info("4/8 Возвраты покупателей и перемещения внутри Ozon")
        if not self.sku_ids:
            return pd.DataFrame()
        out: List[Dict[str, Any]] = []
        for i in range(0, len(self.sku_ids), 100):
            batch = self.sku_ids[i:i+100]
            try:
                rows = self.client.offset_pages(
                    "/v1/analytics/stocks", {"skus": batch, "limit": 1000, "offset": 0},
                    [("result", "rows"), ("rows",), ("result", "items"), ("items",)], limit=1000)
            except OzonApiError:
                rows = self.client.offset_pages(
                    "/v1/analytics/stocks", {"skus": [str(x) for x in batch], "limit": 1000, "offset": 0},
                    [("result", "rows"), ("rows",), ("result", "items"), ("items",)], limit=1000)
            for r in rows:
                sku = norm_id(r.get("sku"))
                if not sku:
                    continue
                cluster = str(r.get("cluster_name") or r.get("cluster") or "").strip()
                wh_name = str(r.get("warehouse_name") or "").strip()
                wh_id = norm_id(r.get("warehouse_id"))
                if not cluster and wh_id:
                    cluster = self.cluster_maps.warehouse_to_cluster.get(wh_id, "")
                if not cluster and wh_name:
                    cluster = self.cluster_maps.warehouse_name_to_cluster.get(norm_name(wh_name), "")
                out.append({
                    "SKU Ozon": sku,
                    "Артикул": self.offer_by_sku.get(sku, f"SKU {sku}"),
                    "Кластер": self.cluster_maps.canonical(cluster) if cluster else "",
                    "Возврат покупателя в пути, шт.": int_qty(r.get("return_from_customer_stock_count")),
                    "Перемещение внутри Ozon, шт.": int_qty(r.get("transit_stock_count")),
                })
        return pd.DataFrame(out)

    # ---------- продажи ----------
    def _fetch_postings_v3(self, start: date, end: date) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        seen_postings: set[str] = set()
        seen_cursors: set[str] = set()
        cursor = ""
        for page in range(1, 501):
            payload: Dict[str, Any] = {
                "filter": {"since": to_ozon_datetime(start), "to": to_ozon_datetime(end, True)},
                "limit": 100,
                "sort_dir": "ASC",
                "with": {"analytics_data": True, "financial_data": True, "legal_info": False},
            }
            if cursor:
                payload["cursor"] = cursor
            data = self.client.post("/v3/posting/fbo/list", payload)
            postings = extract_items(data, [("result", "postings"), ("postings",), ("result", "items"), ("items",)])
            added = 0
            for p in postings:
                key = str(p.get("posting_number") or p.get("order_id") or p.get("order_number") or "").strip()
                if not key:
                    key = json.dumps(p, ensure_ascii=False, sort_keys=True, default=str)[:500]
                if key in seen_postings:
                    continue
                seen_postings.add(key)
                rows.append(p)
                added += 1
            root = data.get("result", {}) if isinstance(data, Mapping) and isinstance(data.get("result"), Mapping) else {}
            next_cursor = str(
                (data.get("cursor") if isinstance(data, Mapping) else "") or
                root.get("cursor") or root.get("next_cursor") or
                deep_first(data, ["cursor", "next_cursor"]) or ""
            ).strip()
            has_next_val = (data.get("has_next") if isinstance(data, Mapping) else None)
            if has_next_val is None:
                has_next_val = root.get("has_next") if isinstance(root, Mapping) else None
            has_next = bool(has_next_val) if has_next_val is not None else bool(next_cursor and len(postings) >= 100)
            logging.info("FBO v3: страница %s, получено %s, новых %s, итого %s", page, len(postings), added, len(rows))
            if not postings or added == 0:
                break
            if not has_next:
                # Если страница полная, но Ozon не дал признака/курсора, считаем это
                # подозрительным и переключаемся на v2 fallback, пока v2 ещё доступен.
                if len(postings) >= 100 and not next_cursor:
                    raise RuntimeError("FBO v3: полная страница без cursor/has_next")
                break
            if not next_cursor:
                raise RuntimeError("FBO v3 сообщил продолжение, но не вернул cursor")
            if next_cursor == cursor or next_cursor in seen_cursors:
                raise RuntimeError("FBO v3 вернул повторяющийся cursor")
            seen_cursors.add(next_cursor)
            cursor = next_cursor
        else:
            raise RuntimeError("FBO v3: превышен защитный лимит 500 страниц")
        return rows

    def _fetch_postings_v2_fallback(self, start: date, end: date) -> List[Dict[str, Any]]:
        """Временный fallback до отключения v2 31.08.2026. Нужен только если v3
        изменил схему cursor. Защита от повторяющейся страницы обязательна."""
        rows: List[Dict[str, Any]] = []
        seen: set[str] = set()
        offset = 0
        for page in range(1, 501):
            payload = {
                "dir": "ASC",
                "filter": {"since": to_ozon_datetime(start), "to": to_ozon_datetime(end, True), "status": ""},
                "limit": 100,
                "offset": offset,
                "translit": False,
                "with": {"analytics_data": True, "financial_data": True},
            }
            data = self.client.post("/v2/posting/fbo/list", payload)
            postings = extract_items(data, [("result",), ("result", "postings"), ("postings",)])
            # /v2 обычно result=list; extract_items умеет list только на верхнем уровне,
            # поэтому обрабатываем этот вариант отдельно.
            if isinstance(data, Mapping) and isinstance(data.get("result"), list):
                postings = [dict(x) for x in data["result"] if isinstance(x, Mapping)]
            added = 0
            for p in postings:
                key = str(p.get("posting_number") or p.get("order_id") or p.get("order_number") or "").strip()
                if not key:
                    key = json.dumps(p, ensure_ascii=False, sort_keys=True, default=str)[:500]
                if key in seen:
                    continue
                seen.add(key)
                rows.append(p)
                added += 1
            logging.info("FBO v2 fallback: страница %s, получено %s, новых %s, итого %s", page, len(postings), added, len(rows))
            if not postings or added == 0 or len(postings) < 100:
                break
            offset += len(postings)
        return rows

    def fetch_postings(self, start: date, end: date) -> List[Dict[str, Any]]:
        try:
            rows = self._fetch_postings_v3(start, end)
            logging.info("FBO postings: основной источник /v3/posting/fbo/list, строк %s", len(rows))
            return rows
        except Exception as exc:
            logging.warning("FBO v3 не удалось надёжно прочитать (%s). Пробуем временный v2 fallback.", exc)
            try:
                rows = self._fetch_postings_v2_fallback(start, end)
                logging.info("FBO postings: fallback /v2/posting/fbo/list, строк %s", len(rows))
                return rows
            except Exception as exc2:
                raise RuntimeError(f"Не удалось получить FBO-отправления ни v3, ни v2: v3={exc}; v2={exc2}") from exc2

    def _detect_posting_quantity_scale(self, postings: Sequence[Mapping[str, Any]]) -> None:
        vals: List[float] = []
        for p in postings:
            products = p.get("products") if isinstance(p.get("products"), list) else []
            for product in products:
                if not isinstance(product, Mapping):
                    continue
                raw = num(product.get("quantity"), 0.0)
                if raw > 0:
                    vals.append(raw)
                if len(vals) >= 5000:
                    break
            if len(vals) >= 5000:
                break
        if not vals:
            self._posting_qty_scale = 1
            self._posting_qty_detection = {"sample": 0, "scale": 1, "reason": "нет quantity"}
            return
        multiples = sum(1 for x in vals if x >= 1000 and abs(x / 1000 - round(x / 1000)) < 1e-9)
        ratio = multiples / len(vals)
        sorted_vals = sorted(vals)
        median = sorted_vals[len(sorted_vals)//2]
        scale = 1000 if len(vals) >= 5 and ratio >= 0.80 and median >= 1000 else 1
        self._posting_qty_scale = scale
        self._posting_qty_detection = {
            "sample": len(vals), "scale": scale, "ratio_multiples_1000": round(ratio, 3),
            "median_raw": median, "min_raw": min(vals), "max_raw": max(vals),
        }
        logging.info("FBO quantity: масштаб=%s, sample=%s, доля кратных 1000=%.1f%%, медиана raw=%s", scale, len(vals), ratio*100, median)

    def _posting_product_qty(self, product: Mapping[str, Any]) -> int:
        raw = num(product.get("quantity"), 1.0)
        if self._posting_qty_scale == 1000:
            raw = raw / 1000.0
        q = int(round(raw))
        return max(1, q)

    def postings_30d(self) -> List[Dict[str, Any]]:
        if self._postings_30d_cache is None:
            start = self.target_date - timedelta(days=29)
            logging.info("Загрузка FBO-отправлений за 30 дней — один раз для продаж и выкупа")
            self._postings_30d_cache = self.fetch_postings(start, self.target_date)
            self._detect_posting_quantity_scale(self._postings_30d_cache)
            logging.info("FBO-отправлений за 30 дней получено: %s", len(self._postings_30d_cache))
        return self._postings_30d_cache

    @staticmethod
    def _posting_date(posting: Mapping[str, Any]) -> Optional[date]:
        return parse_any_date(
            posting.get("created_at") or
            posting.get("in_process_at") or
            posting.get("shipment_date")
        )

    def sales_7d_by_cluster(self) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
        logging.info("5/8 Продажи за 7 дней по кластерам")
        start = self.target_date - timedelta(days=6)
        all_postings = self.postings_30d()
        postings = [
            p for p in all_postings
            if (self._posting_date(p) is not None and start <= self._posting_date(p) <= self.target_date)
        ]
        logging.info("Из них попало в окно последних 7 дней: %s", len(postings))

        rows: List[Dict[str, Any]] = []
        for p in postings:
            status = str(p.get("status") or "").lower()
            if status == "cancelled":
                continue
            financial = p.get("financial_data") if isinstance(p.get("financial_data"), Mapping) else {}
            analytics = p.get("analytics_data") if isinstance(p.get("analytics_data"), Mapping) else {}
            cluster = str(
                financial.get("cluster_to") or analytics.get("cluster_to") or
                p.get("cluster_to") or deep_first(p, ["cluster_to"]) or "Кластер не определён"
            ).strip()
            cluster = self.cluster_maps.canonical(cluster)
            products = p.get("products") if isinstance(p.get("products"), list) else []
            for product in products:
                if not isinstance(product, Mapping):
                    continue
                sku = norm_id(product.get("sku") or product.get("product_id"))
                if not sku:
                    continue
                offer = str(product.get("offer_id") or self.offer_by_sku.get(sku, f"SKU {sku}")).strip()
                name = str(product.get("name") or self.name_by_sku.get(sku, "")).strip()
                qty = self._posting_product_qty(product)
                rows.append({"Артикул": offer, "SKU Ozon": sku, "Название": name, "Кластер": cluster, "Заказано, шт.": qty})
        df = pd.DataFrame(rows)
        if df.empty:
            return pd.DataFrame(columns=["Артикул", "Кластер", "Заказано за 7 дней, шт.", "Среднесуточные продажи за 7 дней, шт."]), postings
        agg = df.groupby(["Артикул", "Кластер"], as_index=False).agg(
            **{"Заказано за 7 дней, шт.": ("Заказано, шт.", "sum")})
        agg["Среднесуточные продажи за 7 дней, шт."] = (agg["Заказано за 7 дней, шт."] / 7).round(1)
        return agg, postings

    # ---------- выкуп ----------
    def fetch_returns(self, start: date, end: date) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        last_id: Any = 0
        seen_last_ids: set[str] = set()
        for page in range(1, 101):
            payload = {
                "filter": {"logistic_return_date": {"time_from": to_ozon_datetime(start), "time_to": to_ozon_datetime(end, True)}},
                "limit": 500, "last_id": last_id,
            }
            data = self.client.post("/v1/returns/list", payload)
            got = extract_items(data, [("returns",), ("result", "returns"), ("result", "items"), ("items",)])
            rows.extend(got)
            nxt = deep_first(data, ["last_id"])
            logging.info("Возвраты: страница %s, получено %s, итого %s", page, len(got), len(rows))
            if not got or nxt in (None, "", 0, last_id):
                break
            nxt_s = str(nxt)
            if nxt_s in seen_last_ids:
                logging.warning("Возвраты: повторился last_id, пагинация остановлена")
                break
            seen_last_ids.add(nxt_s)
            last_id = nxt
        else:
            logging.warning("Возвраты: достигнут защитный лимит 100 страниц")
        return rows

    def buyout_30d(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        logging.info("6/8 Процент выкупа за 30 дней")
        start = self.target_date - timedelta(days=29)
        postings = self.postings_30d()
        logging.info("FBO-отправления повторно не скачиваем: используем кэш из шага 5 (%s шт.)", len(postings))
        returns = self.fetch_returns(start, self.target_date)
        logging.info("Возвратов за 30 дней получено: %s", len(returns))
        bought: Dict[str, int] = {}
        refused: Dict[str, int] = {}
        audit_rows: List[Dict[str, Any]] = []
        for p in postings:
            if str(p.get("status") or "").lower() != "delivered":
                continue
            for product in p.get("products") or []:
                if not isinstance(product, Mapping):
                    continue
                sku = norm_id(product.get("sku") or product.get("product_id"))
                if not sku:
                    continue
                offer = str(product.get("offer_id") or self.offer_by_sku.get(sku, f"SKU {sku}")).strip()
                qty = self._posting_product_qty(product)
                bought[offer] = bought.get(offer, 0) + qty
        for r in returns:
            rtype = str(r.get("type") or r.get("return_type") or "").lower()
            reason = str(r.get("return_reason_name") or r.get("reason_name") or r.get("reason") or "").lower()
            include = rtype == "cancellation" and any(m in reason for m in REFUSAL_MARKERS)
            product = r.get("product") if isinstance(r.get("product"), Mapping) else {}
            sku = norm_id(product.get("sku") or r.get("sku"))
            offer = str(product.get("offer_id") or self.offer_by_sku.get(sku, f"SKU {sku}" if sku else "")).strip()
            qty = max(1, int_qty(product.get("quantity") or r.get("quantity") or 1))
            audit_rows.append({
                "Дата возврата": str(deep_first(r, ["return_date", "logistic_return_date", "final_moment"]) or "")[:10],
                "Артикул": offer,
                "SKU Ozon": sku,
                "Тип возврата": r.get("type") or r.get("return_type"),
                "Причина": r.get("return_reason_name") or r.get("reason_name") or r.get("reason"),
                "Количество, шт.": qty,
                "Считаем невыкупом": "Да" if include else "Нет",
            })
            if include and offer:
                refused[offer] = refused.get(offer, 0) + qty
        result: List[Dict[str, Any]] = []
        for offer in sorted(set(bought) | set(refused)):
            b, f = bought.get(offer, 0), refused.get(offer, 0)
            denom = b + f
            result.append({
                "Артикул": offer,
                "Выкуплено за 30 дней, шт.": b,
                "Невыкуплено после доставки покупателю, шт.": f,
                "Процент выкупа за 30 дней, %": round(b / denom * 100, 1) if denom else None,
            })
        return pd.DataFrame(result), pd.DataFrame(audit_rows)

    # ---------- поставки ----------
    @staticmethod
    def _extract_timeslot(obj: Any) -> Tuple[Optional[date], Optional[date]]:
        pairs: List[Tuple[date, date]] = []
        for d in walk_dicts(obj):
            if d.get("from") and d.get("to"):
                a, b = parse_any_date(d.get("from")), parse_any_date(d.get("to"))
                if a and b:
                    pairs.append((a, b))
        if not pairs:
            return None, None
        pairs.sort(key=lambda x: x[0])
        return pairs[0]

    def _bundle_items(self, bundle_id: str) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        last_id = ""
        for _ in range(100):
            data = None
            last_exc: Optional[Exception] = None
            for payload in (
                {"bundle_ids": [bundle_id], "limit": 1000, "last_id": last_id},
                {"bundle_ids": [bundle_id], "limit": 1000, "last_id": last_id, "is_asc": True, "sort_field": "SKU", "query": ""},
            ):
                try:
                    data = self.client.post("/v1/supply-order/bundle", payload)
                    break
                except OzonApiError as exc:
                    last_exc = exc
                    if exc.status in {400, 404, 422}:
                        continue
                    raise
            if data is None:
                if last_exc:
                    raise last_exc
                break
            got = extract_items(data, [("items",), ("result", "items")])
            rows.extend(got)
            has_next = bool(data.get("has_next")) if isinstance(data, Mapping) else False
            if isinstance(data, Mapping) and isinstance(data.get("result"), Mapping):
                has_next = has_next or bool(data["result"].get("has_next"))
            nxt = extract_cursor(data)
            if not got or not has_next or not nxt or nxt == last_id:
                break
            last_id = nxt
        return rows

    def _supply_list_v3(self) -> List[Dict[str, Any]]:
        variants = [
            {"filter": {}, "limit": 100, "last_id": "", "sort_by": 1, "sort_direction": 2},
            {"filter": {"states": []}, "limit": 100, "last_id": "", "sort_by": 1, "sort_direction": 2},
            {"filter": {}, "limit": 100, "last_id": ""},
            {"limit": 100, "last_id": ""},
            {"filter": {"created_date_from": (self.target_date - timedelta(days=180)).isoformat(),
                        "created_date_to": (self.target_date + timedelta(days=60)).isoformat()},
             "limit": 100, "last_id": "", "sort_by": 1, "sort_direction": 2},
        ]
        errors: List[str] = []
        for base in variants:
            try:
                rows: List[Dict[str, Any]] = []
                last_id = str(base.get("last_id") or "")
                seen_ids: set[str] = set()
                for page in range(1, 101):
                    payload = dict(base)
                    payload["last_id"] = last_id
                    data = self.client.post("/v3/supply-order/list", payload)
                    got = extract_items(data, [("result", "items"), ("items",), ("result", "orders"), ("orders",)])
                    if isinstance(data, Mapping) and isinstance(data.get("result"), list):
                        got = [dict(x) for x in data["result"] if isinstance(x, Mapping)]
                    added = 0
                    for item in got:
                        oid = norm_id(item.get("supply_order_id") or item.get("order_id") or item.get("id"))
                        key = oid or json.dumps(item, ensure_ascii=False, sort_keys=True, default=str)[:500]
                        if key in seen_ids:
                            continue
                        seen_ids.add(key)
                        rows.append(item)
                        added += 1
                    nxt = str(deep_first(data, ["last_id", "cursor", "next_cursor"]) or "").strip()
                    has_next = bool(deep_first(data, ["has_next"]))
                    logging.info("Supply list v3: страница %s, получено %s, новых %s, итого %s", page, len(got), added, len(rows))
                    if not got or added == 0:
                        break
                    if not has_next and (len(got) < 100 or not nxt):
                        break
                    if not nxt or nxt == last_id:
                        break
                    last_id = nxt
                return rows
            except OzonApiError as exc:
                errors.append(str(exc))
                if exc.status not in {400, 404, 405, 422}:
                    raise
        logging.warning("/v3/supply-order/list не принял варианты запроса: %s", " | ".join(errors[-3:]))
        return []

    def _supply_list_legacy(self) -> List[Dict[str, Any]]:
        """Резерв для кабинетов, где v3/list не возвращает ранее созданные заявки.
        Старый метод давно помечен устаревшим, поэтому используется только после пустого v3."""
        for states in ([], sorted(ACTIVE_STATES)):
            try:
                rows: List[Dict[str, Any]] = []
                page = 1
                for _ in range(100):
                    payload = {"page": page, "page_size": 100, "states": states}
                    data = self.client.post("/v1/supply-order/list", payload)
                    got = extract_items(data, [("supply_orders",), ("result", "supply_orders"), ("items",)])
                    rows.extend(got)
                    has_next = bool(data.get("has_next")) if isinstance(data, Mapping) else False
                    if not got or not has_next:
                        break
                    page += 1
                if rows:
                    logging.warning("Поставки получены через legacy /v1/supply-order/list: %s", len(rows))
                    self._supply_diagnostics["list_source"] = "/v1/supply-order/list fallback"
                    return rows
            except OzonApiError as exc:
                if exc.status not in {400, 403, 404, 405, 422}:
                    logging.warning("Legacy supply list: %s", exc)
        return []

    def _legacy_supply_items(self, oid: str) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        page = 1
        for _ in range(100):
            try:
                data = self.client.post("/v1/supply-order/items", {
                    "supply_order_id": int(oid) if str(oid).isdigit() else oid,
                    "page": page,
                    "page_size": 100,
                })
            except OzonApiError:
                return rows
            got = extract_items(data, [("items",), ("result", "items")])
            rows.extend(got)
            has_next = bool(data.get("has_next")) if isinstance(data, Mapping) else False
            if not got or not has_next:
                break
            page += 1
        return rows

    @staticmethod
    def _supply_state(obj: Mapping[str, Any]) -> str:
        return str(obj.get("supply_state") or obj.get("state") or obj.get("status") or "").upper().strip()

    @staticmethod
    def _is_active_supply_state(state: str) -> bool:
        if not state:
            return True  # неизвестный новый статус не теряем раньше времени
        return not any(token in state for token in TERMINAL_SUPPLY_STATES)

    def _supply_detail(self, oid: str) -> Mapping[str, Any]:
        errors: List[Exception] = []
        for payload in ({"order_id": oid}, {"supply_order_id": oid}):
            try:
                return self.client.post("/v3/supply-order/get", payload)
            except OzonApiError as exc:
                errors.append(exc)
                if exc.status not in {400, 404, 422}:
                    raise
        # Последний резерв — legacy get, если кабинет его ещё поддерживает.
        try:
            return self.client.post("/v1/supply-order/get", {
                "supply_order_id": int(oid) if str(oid).isdigit() else oid
            })
        except Exception:
            raise errors[-1] if errors else RuntimeError("Нет ответа supply-order/get")

    def supply_orders(self) -> pd.DataFrame:
        logging.info("7/8 Активные FBO-заявки и поставки")
        # Счётчик нужен только как диагностика и не влияет на расчёт.
        try:
            counters = self.client.post("/v1/supply-order/status/counter", {})
            self._supply_diagnostics["status_counter_available"] = True
            logging.info("Supply status counter: ответ получен")
        except Exception as exc:
            self._supply_diagnostics["status_counter_available"] = False
            logging.warning("Supply status counter недоступен: %s", exc)

        list_items = self._supply_list_v3()
        if list_items:
            self._supply_diagnostics["list_source"] = "/v3/supply-order/list"
        else:
            list_items = self._supply_list_legacy()
        self._supply_diagnostics["list_total"] = len(list_items)
        logging.info("Заявок supply-order/list получено без раннего фильтра по статусу: %s", len(list_items))
        if not list_items:
            logging.warning("Поставки не найдены ни v3, ни legacy API. Отчёт продолжит работу без seller supply-order.")
            return pd.DataFrame()

        status_counts: Dict[str, int] = {}
        active_short: List[Dict[str, Any]] = []
        for item in list_items:
            state = self._supply_state(item) or "НЕИЗВЕСТНО"
            status_counts[state] = status_counts.get(state, 0) + 1
            if self._is_active_supply_state(state):
                active_short.append(item)
        self._supply_diagnostics["status_counts"] = status_counts
        self._supply_diagnostics["active_short"] = len(active_short)
        logging.info("Статусы заявок: %s", ", ".join(f"{k}={v}" for k,v in sorted(status_counts.items())))
        logging.info("Активных/нефинальных заявок после локального фильтра: %s", len(active_short))

        seen: set[str] = set()
        rows: List[Dict[str, Any]] = []
        for idx, short in enumerate(active_short, 1):
            oid = norm_id(short.get("supply_order_id") or short.get("order_id") or short.get("id"))
            if not oid or oid in seen:
                continue
            seen.add(oid)
            try:
                detail = self._supply_detail(oid)
            except Exception as exc:
                logging.warning("Заявка %s: детали недоступны: %s", oid, exc)
                detail = short

            root = detail.get("result", detail) if isinstance(detail, Mapping) else detail
            order_root = root if isinstance(root, Mapping) else short
            order_state = self._supply_state(order_root) or self._supply_state(short)
            if not self._is_active_supply_state(order_state):
                continue
            order_slot_from, _ = self._extract_timeslot(order_root)

            # Ищем все реальные supply-узлы по bundle_id; если структура изменилась,
            # не теряем root/short, где bundle_id может лежать напрямую.
            supply_nodes: List[Mapping[str, Any]] = []
            for node in walk_dicts(order_root):
                if node.get("bundle_id") not in (None, ""):
                    supply_nodes.append(node)
            if not supply_nodes and isinstance(short, Mapping) and short.get("bundle_id") not in (None, ""):
                supply_nodes = [short]
            # В некоторых ответах supplies есть, но bundle_id может быть выше/ниже.
            if not supply_nodes:
                supplied = order_root.get("supplies") if isinstance(order_root, Mapping) else None
                if isinstance(supplied, list):
                    supply_nodes = [x for x in supplied if isinstance(x, Mapping)]
            if not supply_nodes:
                logging.warning("Заявка %s: не найден supply/bundle в деталях", oid)
                continue

            unique_nodes: List[Mapping[str, Any]] = []
            seen_bundle_nodes: set[str] = set()
            for node in supply_nodes:
                k = norm_id(node.get("bundle_id")) or norm_id(node.get("supply_id") or node.get("id")) or str(id(node))
                if k in seen_bundle_nodes:
                    continue
                seen_bundle_nodes.add(k)
                unique_nodes.append(node)

            for supply in unique_nodes:
                state = self._supply_state(supply) or order_state
                if not self._is_active_supply_state(state):
                    continue
                supply_slot_from, _ = self._extract_timeslot(supply)
                slot_from = order_slot_from or supply_slot_from
                bundle_id = norm_id(supply.get("bundle_id"))
                storage_wh = supply.get("storage_warehouse") if isinstance(supply.get("storage_warehouse"), Mapping) else {}
                dropoff_wh = supply.get("drop_off_warehouse") if isinstance(supply.get("drop_off_warehouse"), Mapping) else {}
                storage_id = norm_id(supply.get("storage_warehouse_id") or storage_wh.get("warehouse_id") or storage_wh.get("id"))
                dropoff_id = norm_id(supply.get("dropoff_warehouse_id") or supply.get("drop_off_warehouse_id") or dropoff_wh.get("warehouse_id") or dropoff_wh.get("id") or deep_first(order_root,["dropoff_warehouse_id"]))
                storage_name = str(storage_wh.get("name") or supply.get("storage_warehouse_name") or deep_first(supply,["storage_warehouse_name"]) or "").strip()
                dropoff_name = str(dropoff_wh.get("name") or deep_first(order_root,["dropoff_warehouse_name"]) or "").strip()
                macro_id = norm_id(supply.get("macrolocal_cluster_id") or deep_first(supply,["macrolocal_cluster_id","cluster_id"]) or deep_first(order_root,["macrolocal_cluster_id","cluster_id"]))
                cluster = self.cluster_maps.macro_to_cluster.get(macro_id, "")
                if not cluster and storage_id:
                    cluster = self.cluster_maps.warehouse_to_cluster.get(storage_id, "")
                if not cluster and storage_name:
                    cluster = self.cluster_maps.warehouse_name_to_cluster.get(norm_name(storage_name), "") or static_cluster_from_warehouse(storage_name)
                # Новая модель кросс-дока может отдавать название целевого кластера без конечного склада.
                if not cluster:
                    cluster = str(deep_first(supply,["macrolocal_cluster_name","cluster_name"]) or deep_first(order_root,["macrolocal_cluster_name","cluster_name"]) or "").strip()
                cluster = self.cluster_maps.canonical(cluster) if cluster else "Кластер не определён"

                route = "Кросс-док"
                if dropoff_id and storage_id and dropoff_id == storage_id:
                    route = "Прямая"
                elif storage_name and dropoff_name and similarity(storage_name, dropoff_name) >= 0.9:
                    route = "Прямая"
                arrival_api = parse_any_date(storage_wh.get("arrival_date") or supply.get("arrival_date"))

                products: List[Dict[str, Any]] = []
                if bundle_id:
                    try:
                        products = self._bundle_items(bundle_id)
                    except Exception as exc:
                        logging.warning("Состав поставки bundle=%s: %s", bundle_id, exc)
                if not products:
                    for key in ("products", "items"):
                        val = supply.get(key)
                        if isinstance(val, list):
                            products = [dict(x) for x in val if isinstance(x, Mapping)]
                            if products:
                                break
                if not products:
                    products = self._legacy_supply_items(oid)
                if not products:
                    logging.warning("Заявка %s bundle=%s: состав товаров не получен", oid, bundle_id)
                    continue

                for product in products:
                    sku = norm_id(product.get("sku") or product.get("product_id"))
                    if not sku:
                        continue
                    offer = str(product.get("contractor_item_code") or product.get("offer_id") or self.offer_by_sku.get(sku, f"SKU {sku}")).strip()
                    qty = int_qty(product.get("quantity"))
                    availability, eta_basis, sla_days = self._estimate_supply_eta(
                        state=state, route=route, cluster=cluster, slot_from=slot_from,
                        arrival_api=arrival_api, detail=order_root, supply=supply)
                    rows.append({
                        "Артикул": offer,
                        "SKU Ozon": sku,
                        "Название": str(product.get("name") or self.name_by_sku.get(sku, "")),
                        "Кластер назначения": cluster,
                        "Количество, шт.": qty,
                        "Статус": STATUS_RU.get(state, state or "Неизвестно"),
                        "Статус API": state,
                        "Уже передано Ozon": "Да" if state in PHYSICAL_STATES else "Нет",
                        "Тип поставки": route,
                        "Точка отгрузки": dropoff_name or (CROSSDOCK_ORIGIN if route == "Кросс-док" else ""),
                        "Склад назначения": storage_name,
                        "Слот отгрузки / приёмки": slot_from,
                        "Дата прибытия от API": arrival_api,
                        "Прогноз доступности к продаже": availability,
                        "Основание прогноза": eta_basis,
                        "Срок кросс-дока из Кавказского, дней": sla_days,
                        "Приёмка Ozon после прибытия, рабочих дней": FINAL_ACCEPTANCE_WORKDAYS,
                        "ID заявки": oid,
                        "ID поставки": norm_id(supply.get("supply_id") or supply.get("id")),
                    })
            if idx % 20 == 0:
                logging.info("Обработано активных заявок: %s/%s", idx, len(active_short))

        self._supply_diagnostics["product_rows"] = len(rows)
        logging.info("Строк товаров в активных поставках: %s", len(rows))
        return pd.DataFrame(rows)

    def _estimate_supply_eta(self, state: str, route: str, cluster: str, slot_from: Optional[date],
                             arrival_api: Optional[date], detail: Mapping[str, Any], supply: Mapping[str, Any]) -> Tuple[Optional[date], str, Optional[int]]:
        # Возвращаем не просто прибытие машины, а дату, когда товар консервативно
        # должен стать доступен к продаже после приёмки Ozon.
        if state == "ACCEPTANCE_AT_STORAGE_WAREHOUSE":
            avail = add_workdays(self.target_date, FINAL_ACCEPTANCE_WORKDAYS)
            return avail, f"Уже на приёмке конечного склада + до {FINAL_ACCEPTANCE_WORKDAYS} раб. дней на размещение", 0
        if arrival_api:
            avail = add_workdays(max(self.target_date, arrival_api), FINAL_ACCEPTANCE_WORKDAYS)
            return avail, f"Дата прибытия API + до {FINAL_ACCEPTANCE_WORKDAYS} раб. дней на приёмку", None
        if route == "Прямая":
            if slot_from:
                avail = add_workdays(max(self.target_date, slot_from), FINAL_ACCEPTANCE_WORKDAYS)
                return avail, f"Слот прямой поставки + до {FINAL_ACCEPTANCE_WORKDAYS} раб. дней на приёмку", 0
            return None, "Нет даты прямой поставки в API", None

        sla_days, source = self.sla.lookup(cluster)
        if sla_days is None:
            return None, "Нет подтверждённого срока маршрута Москва — Кавказский → кластер", None

        if state in PHYSICAL_STATES:
            # 4 дня на сборку уже НЕ добавляем. Если известен слот фактической/плановой
            # передачи Ozon, учитываем уже прошедшие дни маршрута. Если нормативный
            # срок уже истёк, консервативно считаем, что прибытие не раньше даты отчёта.
            if slot_from:
                arrival = max(self.target_date, slot_from + timedelta(days=sla_days))
                basis = f"Слот передачи Ozon + до {sla_days} дн. кросс-дока + до {FINAL_ACCEPTANCE_WORKDAYS} раб. дней приёмки ({source})"
            else:
                arrival = self.target_date + timedelta(days=sla_days)
                basis = f"Поставка уже передана Ozon; без даты передачи закладываем до {sla_days} дн. кросс-дока + до {FINAL_ACCEPTANCE_WORKDAYS} раб. дней приёмки ({source})"
            avail = add_workdays(arrival, FINAL_ACCEPTANCE_WORKDAYS)
            return avail, basis, sla_days

        # Заявка создана, но товар ещё у продавца. Реальный будущий слот сильнее
        # стандартных +4 дней; если слота нет — добавляем 4 дня на сборку.
        handoff = max(self.target_date, slot_from) if slot_from else self.target_date + timedelta(days=PREPARATION_DAYS)
        arrival = handoff + timedelta(days=sla_days)
        avail = add_workdays(arrival, FINAL_ACCEPTANCE_WORKDAYS)
        if slot_from:
            basis = f"Существующий слот + до {sla_days} дн. кросс-дока + до {FINAL_ACCEPTANCE_WORKDAYS} раб. дней приёмки ({source})"
        else:
            basis = f"{PREPARATION_DAYS} дн. на сборку + до {sla_days} дн. кросс-дока + до {FINAL_ACCEPTANCE_WORKDAYS} раб. дней приёмки ({source})"
        return avail, basis, sla_days

    # ---------- сборка ----------
    def build(self) -> Dict[str, pd.DataFrame]:
        self.fetch_products()
        self.fetch_clusters()
        sla_map = self.sla.load()
        logging.info("Сроков кросс-дока из Кавказского загружено: %s", len(sla_map))
        stocks = self.fetch_stock_warehouses()
        transit = self.fetch_transit_stocks()
        sales_cluster, _ = self.sales_7d_by_cluster()
        buyout, buyout_audit = self.buyout_30d()
        supplies = self.supply_orders()
        logging.info("8/8 Формирование сводки")

        by_article = self._build_by_article(stocks, transit, sales_cluster, buyout, supplies)
        by_cluster = self._build_by_cluster(stocks, transit, sales_cluster, supplies)
        method = self._methodology()
        cluster_ref = self._cluster_reference()
        diag = self._diagnostics(stocks, supplies)
        return {
            "По артикулам": by_article,
            "По кластерам": by_cluster,
            "Поставки и прогноз": supplies,
            "Расчёт выкупа": buyout_audit,
            "Нормативы кросс-дока": self.sla.as_dataframe(),
            "Диагностика": diag,
            "Справочник кластеров": cluster_ref,
            "Методика": method,
        }

    def _diagnostics(self, stocks: pd.DataFrame, supplies: pd.DataFrame) -> pd.DataFrame:
        rows: List[Dict[str, Any]] = [
            {"Проверка": "Версия кода", "Значение": SCRIPT_VERSION},
            {"Проверка": "Масштаб quantity в FBO postings", "Значение": self._posting_qty_scale},
            {"Проверка": "Диагностика quantity", "Значение": json.dumps(self._posting_qty_detection, ensure_ascii=False)},
            {"Проверка": "Строк остатков", "Значение": 0 if stocks is None else len(stocks)},
            {"Проверка": "Складов без определённого кластера", "Значение": len(self._unknown_warehouses)},
            {"Проверка": "Строк товаров в активных поставках", "Значение": 0 if supplies is None else len(supplies)},
            {"Проверка": "Диагностика поставок", "Значение": json.dumps(self._supply_diagnostics, ensure_ascii=False, default=str)},
            {"Проверка": "Источник сроков кросс-дока", "Значение": self.sla.source or "Не получен"},
            {"Проверка": "Количество нормативов кросс-дока", "Значение": len(self.sla.rows)},
        ]
        for wh in sorted(self._unknown_warehouses):
            rows.append({"Проверка": "Склад без кластера", "Значение": wh})
        return pd.DataFrame(rows)

    def _build_by_article(self, stocks: pd.DataFrame, transit: pd.DataFrame, sales_cluster: pd.DataFrame,
                          buyout: pd.DataFrame, supplies: pd.DataFrame) -> pd.DataFrame:
        articles = set()
        for df in (stocks, transit, sales_cluster, buyout, supplies):
            if df is not None and not df.empty and "Артикул" in df.columns:
                articles |= set(str(x) for x in df["Артикул"].dropna() if str(x).strip())
        rows: List[Dict[str, Any]] = []
        for article in sorted(articles):
            s = stocks[stocks["Артикул"] == article] if not stocks.empty else pd.DataFrame()
            t = transit[transit["Артикул"] == article] if not transit.empty else pd.DataFrame()
            sc = sales_cluster[sales_cluster["Артикул"] == article] if not sales_cluster.empty else pd.DataFrame()
            sp = supplies[supplies["Артикул"] == article] if not supplies.empty else pd.DataFrame()
            b = buyout[buyout["Артикул"] == article] if not buyout.empty else pd.DataFrame()
            current = int(s["Доступно сейчас, шт."].sum()) if not s.empty else 0
            ret = int(t["Возврат покупателя в пути, шт."].sum()) if not t.empty else 0
            internal = int(t["Перемещение внутри Ozon, шт."].sum()) if not t.empty else 0
            physical = int(sp.loc[sp["Уже передано Ozon"] == "Да", "Количество, шт."].sum()) if not sp.empty else 0
            requests = int(sp.loc[sp["Уже передано Ozon"] == "Нет", "Количество, шт."].sum()) if not sp.empty else 0
            supply_api_qty = int(sp["Количество, шт."].sum()) if not sp.empty else 0
            promised = int(s["Подтверждено Ozon в поставках, шт."].sum()) if (not s.empty and "Подтверждено Ozon в поставках, шт." in s.columns) else 0
            # promised_amount может дублировать supply-order. Поэтому используем его только
            # как резерв на НЕПОКРЫТУЮ часть, если API заявок вернул меньше товара.
            promised_fallback = max(0, promised - supply_api_qty)
            future = current + ret + internal + supply_api_qty + promised_fallback
            sold7 = int(sc["Заказано за 7 дней, шт."].sum()) if not sc.empty else 0
            avg = round(sold7 / 7, 1)
            days_now = int(round(current / avg)) if avg > 0 else None
            days_future = int(round(future / avg)) if avg > 0 else None
            name = ""
            if not s.empty and "Название" in s.columns:
                name = next((str(x) for x in s["Название"] if str(x).strip()), "")
            if not name:
                skus = set(s["SKU Ozon"].astype(str)) if not s.empty else set()
                for sku in skus:
                    if self.name_by_sku.get(sku):
                        name = self.name_by_sku[sku]
                        break
            pct = None
            bought = refused = 0
            if not b.empty:
                bought = int(b["Выкуплено за 30 дней, шт."].sum())
                refused = int(b["Невыкуплено после доставки покупателю, шт."].sum())
                pct = round(bought / (bought + refused) * 100, 1) if bought + refused else None
            rows.append({
                "Артикул": article,
                "Среднесуточные продажи за 7 дней, шт.": avg,
                "Запас, дней: сейчас (с учётом пути и заявок)": f"{days_now if days_now is not None else '—'} ({days_future if days_future is not None else '—'})",
                "Остаток, шт.: сейчас (с учётом пути и заявок)": f"{current} ({future})",
                "Процент выкупа за 30 дней, %": pct,
                "Название товара": name,
                "Заказано за 7 дней, шт.": sold7,
                "Доступно к продаже сейчас, шт.": current,
                "Возвраты покупателей в пути на склад Ozon, шт.": ret,
                "Перемещение между складами Ozon, шт.": internal,
                "Поставки уже переданы Ozon, шт.": physical,
                "Активные заявки, ещё не переданы Ozon, шт.": requests,
                "Подтверждено Ozon, но не найдено в API заявок, шт.": promised_fallback,
                "Итого с учётом пути и заявок, шт.": future,
                "Запас сейчас, дней": days_now,
                "Запас с учётом пути и заявок, дней": days_future,
                "Выкуплено за 30 дней, шт.": bought,
                "Невыкуплено после доставки покупателю, шт.": refused,
                "Дата отчёта": self.target_date,
            })
        df = pd.DataFrame(rows)
        if not df.empty:
            df = df.sort_values(["Среднесуточные продажи за 7 дней, шт.", "Запас сейчас, дней"], ascending=[False, True], na_position="last").reset_index(drop=True)
        return df

    def _build_by_cluster(self, stocks: pd.DataFrame, transit: pd.DataFrame, sales: pd.DataFrame,
                          supplies: pd.DataFrame) -> pd.DataFrame:
        pairs = set()
        for df, ccol in ((stocks, "Кластер"), (sales, "Кластер"), (supplies, "Кластер назначения")):
            if df is not None and not df.empty and "Артикул" in df.columns and ccol in df.columns:
                pairs |= set((str(a), str(c)) for a, c in zip(df["Артикул"], df[ccol]) if str(a).strip() and str(c).strip())
        rows: List[Dict[str, Any]] = []
        for article, cluster_raw in sorted(pairs):
            cluster = self.cluster_maps.canonical(cluster_raw)
            s = stocks[(stocks["Артикул"] == article) & (stocks["Кластер"].map(self.cluster_maps.canonical) == cluster)] if not stocks.empty else pd.DataFrame()
            sale = sales[(sales["Артикул"] == article) & (sales["Кластер"].map(self.cluster_maps.canonical) == cluster)] if not sales.empty else pd.DataFrame()
            sp = supplies[(supplies["Артикул"] == article) & (supplies["Кластер назначения"].map(self.cluster_maps.canonical) == cluster)] if not supplies.empty else pd.DataFrame()
            # Внутренний транзит/возврат относим к кластеру только когда сам API
            # вернул склад/кластер. Пустой кластер не размазываем пропорционально.
            t = pd.DataFrame()
            if transit is not None and not transit.empty and "Кластер" in transit.columns:
                t = transit[(transit["Артикул"] == article) & (transit["Кластер"].astype(str).str.strip() != "") & (transit["Кластер"].map(self.cluster_maps.canonical) == cluster)]
            current = int(s["Доступно сейчас, шт."].sum()) if not s.empty else 0
            cluster_returns = int(t["Возврат покупателя в пути, шт."].sum()) if not t.empty else 0
            cluster_internal = int(t["Перемещение внутри Ozon, шт."].sum()) if not t.empty else 0
            sold7 = int(sale["Заказано за 7 дней, шт."].sum()) if not sale.empty else 0
            avg = round(sold7 / 7, 1)
            future_supply = int(sp["Количество, шт."].sum()) if not sp.empty else 0
            promised = int(s["Подтверждено Ozon в поставках, шт."].sum()) if (not s.empty and "Подтверждено Ozon в поставках, шт." in s.columns) else 0
            promised_fallback = max(0, promised - future_supply)
            future_total = current + cluster_returns + cluster_internal + future_supply + promised_fallback
            days_now = int(round(current / avg)) if avg > 0 else None
            days_future = int(round(future_total / avg)) if avg > 0 else None
            depletion = self.target_date + timedelta(days=math.ceil(current / avg)) if avg > 0 else None

            # Существующие поставки событиями по ETA.
            events: List[Tuple[date, int, str]] = []
            if not sp.empty:
                for _, r in sp.iterrows():
                    eta = r.get("Прогноз доступности к продаже")
                    if isinstance(eta, pd.Timestamp):
                        eta = eta.date()
                    elif not isinstance(eta, date):
                        eta = parse_any_date(eta)
                    if eta:
                        events.append((eta, int_qty(r.get("Количество, шт.")), str(r.get("Статус") or "")))
            events.sort(key=lambda x: x[0])

            sla_days, sla_source = self.sla.lookup(cluster)
            normal_lead = PREPARATION_DAYS + sla_days if sla_days is not None else None
            hypothetical_eta = None
            if normal_lead is not None:
                arrival = self.target_date + timedelta(days=normal_lead)
                hypothetical_eta = add_workdays(arrival, FINAL_ACCEPTANCE_WORKDAYS)

            # Если supply-order не раскрыл всю поставку, но stock API показывает promised_amount,
            # не теряем этот товар. Для неизвестного статуса используем консервативный ETA
            # новой поставки: +4 дня на сборку + кросс-док + приёмка Ozon.
            if promised_fallback > 0 and hypothetical_eta is not None:
                events.append((hypothetical_eta, promised_fallback, "Подтверждено Ozon, детали заявки не получены"))
                events.sort(key=lambda x: x[0])

            oos_days_existing, nearest_eta, nearest_qty, nearest_status = self._simulate_oos(current, avg, events)
            if events:
                forecast_eta = nearest_eta
                forecast_qty = nearest_qty
                forecast_status = nearest_status
                oos_days = oos_days_existing
                forecast_basis = "Учитываем существующую заявку/поставку"
            else:
                forecast_eta = hypothetical_eta
                forecast_qty = 0
                forecast_status = "Новой заявки ещё нет"
                if avg > 0 and hypothetical_eta and depletion:
                    oos_days = max(0, int(math.ceil((hypothetical_eta - depletion).total_seconds() / 86400)))
                else:
                    oos_days = None
                forecast_basis = f"Если начать сборку сегодня: {PREPARATION_DAYS} дн. сборка + кросс-док + до {FINAL_ACCEPTANCE_WORKDAYS} раб. дней приёмки" if hypothetical_eta else "Нет срока маршрута"

            if avg <= 0:
                will_make = "Продаж за 7 дней нет"
            elif forecast_eta is None or depletion is None:
                will_make = "Нет данных"
            else:
                will_make = "Да" if forecast_eta <= depletion else "Нет"

            rows.append({
                "Артикул": article,
                "Кластер": cluster,
                "Среднесуточные продажи за 7 дней, шт.": avg,
                "Запас, дней: сейчас (с учётом заявок)": f"{days_now if days_now is not None else '—'} ({days_future if days_future is not None else '—'})",
                "Остаток, шт.: сейчас (с учётом заявок)": f"{current} ({future_total})",
                "Заказано за 7 дней, шт.": sold7,
                "Остаток сейчас в кластере, шт.": current,
                "Возвраты покупателей в пути в кластер, шт.": cluster_returns,
                "Перемещение внутри Ozon в кластер, шт.": cluster_internal,
                "В активных поставках в кластер, шт.": future_supply,
                "Подтверждено Ozon, но не найдено в API заявок, шт.": promised_fallback,
                "Остаток с учётом пути и поставок, шт.": future_total,
                "Запас сейчас, дней": days_now,
                "Запас с учётом заявок, дней": days_future,
                "Дата ожидаемого окончания остатка": depletion if depletion else None,
                "Ближайшее пополнение, шт.": forecast_qty,
                "Статус ближайшего пополнения": forecast_status,
                "Прогноз доступности пополнения к продаже": forecast_eta,
                "Срок кросс-дока из Москвы — Кавказский, дней": sla_days,
                "Срок до прибытия новой поставки: 4 дня сборки + кросс-док, дней": normal_lead,
                "Дополнительно на приёмку Ozon, рабочих дней": FINAL_ACCEPTANCE_WORKDAYS,
                "Источник срока кросс-дока": sla_source,
                "Успеет пополнение до дефицита": will_make,
                "Прогноз дней без остатка": oos_days,
                "Основание прогноза": forecast_basis,
                "Дата отчёта": self.target_date,
            })
        df = pd.DataFrame(rows)
        if not df.empty:
            # Критичные кластеры сверху: сначала OOS, затем короткий запас.
            rank = df["Успеет пополнение до дефицита"].map({"Нет": 0, "Нет данных": 1, "Да": 2, "Продаж за 7 дней нет": 3}).fillna(4)
            df = df.assign(_rank=rank).sort_values(["_rank", "Запас сейчас, дней", "Среднесуточные продажи за 7 дней, шт."], ascending=[True, True, False], na_position="last").drop(columns="_rank").reset_index(drop=True)
        return df

    def _simulate_oos(self, current: int, avg: float, events: List[Tuple[date, int, str]]) -> Tuple[Optional[int], Optional[date], int, str]:
        if not events:
            return None, None, 0, ""
        nearest_eta, nearest_qty, nearest_status = events[0]
        if avg <= 0:
            return 0, nearest_eta, nearest_qty, nearest_status
        inv = float(current)
        cursor = self.target_date
        oos_days = 0.0
        for eta, qty, _status in events:
            if eta < cursor:
                eta = cursor
            delta = (eta - cursor).days
            need = avg * delta
            if inv >= need:
                inv -= need
            else:
                covered = inv / avg if avg > 0 else delta
                oos_days += max(0.0, delta - covered)
                inv = 0.0
            inv += qty
            cursor = eta
        return int(math.ceil(oos_days)), nearest_eta, nearest_qty, nearest_status

    def _cluster_reference(self) -> pd.DataFrame:
        rows = []
        for wh, cluster in sorted(self.cluster_maps.warehouse_name_to_cluster.items()):
            rows.append({"Источник": "API", "Склад / правило": wh, "ID макролокального кластера": "", "Кластер": cluster})
        for mid, cluster in sorted(self.cluster_maps.macro_to_cluster.items()):
            rows.append({"Источник": "API", "Склад / правило": "", "ID макролокального кластера": mid, "Кластер": cluster})
        for tokens, cluster in STATIC_WAREHOUSE_CLUSTER_PATTERNS:
            rows.append({"Источник": "Резерв по названию склада", "Склад / правило": " + ".join(tokens), "ID макролокального кластера": "", "Кластер": cluster})
        return pd.DataFrame(rows)

    def _methodology(self) -> pd.DataFrame:
        return pd.DataFrame([
            {"Показатель": "Среднесуточные продажи за 7 дней", "Как считаем": "Заказанные FBO-штуки за последние 7 календарных дней / 7. quantity FBO автоматически проверяется на масштаб ×1000 и нормализуется только если структура данных это подтверждает."},
            {"Показатель": "Остаток сейчас", "Как считаем": "free_to_sell_amount — реально доступно к продаже на FBO-складах Ozon."},
            {"Показатель": "Остаток в скобках по артикулу", "Как считаем": "Остаток сейчас + возвраты покупателей в пути + внутренние перемещения Ozon + все активные заявки продавца, включая ещё не переданные Ozon."},
            {"Показатель": "Кластер", "Как считаем": "Сначала кластер из API, затем справочник склад→кластер, затем резерв по географическому названию склада. Физическое имя РФЦ никогда не используется как отдельный кластер."},
            {"Показатель": "promised_amount", "Как считаем": "Используем как резерв: прибавляем только ту часть promised_amount, которая не покрыта найденными supply-order. Так поставка не теряется при неполном API и не считается дважды."},
            {"Показатель": "Процент выкупа", "Как считаем": "Delivered / (Delivered + явный отказ/неполучение после прибытия покупателю). Ранние отмены и ClientReturn после покупки исключены."},
            {"Показатель": "Срок кросс-дока", "Как считаем": f"Верхняя граница маршрута от точки «{CROSSDOCK_ORIGIN}» до конечного склада/кластера. Источник — таблица сроков из БЗ Ozon (публичная копия XLSX), кэш обновляется не чаще чем раз в {CROSSDOCK_CACHE_DAYS} дней."},
            {"Показатель": "Новая поставка без заявки", "Как считаем": f"{PREPARATION_DAYS} календарных дня на сборку + срок кросс-дока + до {FINAL_ACCEPTANCE_WORKDAYS} рабочих дней на приёмку/размещение Ozon."},
            {"Показатель": "Существующая заявка", "Как считаем": f"Если товар уже передан Ozon, {PREPARATION_DAYS} дня на сборку повторно не добавляем. Если есть слот — считаем от слота. После прибытия консервативно добавляем до {FINAL_ACCEPTANCE_WORKDAYS} рабочих дней на приёмку."},
            {"Показатель": "Прогноз дней без остатка", "Как считаем": "Сравниваем текущий запас в днях с датой доступности ближайшего известного пополнения; если заявки нет — с нормативным сроком новой поставки."},
        ])

    # ---------- Excel / S3 ----------
    def save(self, sheets: Dict[str, pd.DataFrame]) -> Tuple[str, str]:
        data = self._xlsx(sheets)
        prefix = f"Сводные отчёты/{self.store}"
        dated = f"{prefix}/Остатки_и_оборачиваемость_{self.target_date.isoformat()}.xlsx"
        latest = f"{prefix}/Последний.xlsx"
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        self.storage.upload(dated, data, content_type)
        self.storage.upload(latest, data, content_type)
        status_key = f"{prefix}/Служебные/Последний_запуск.json"
        self.storage.upload_json(status_key, {
            "status": "OK",
            "script_version": SCRIPT_VERSION,
            "store": self.store,
            "date": self.target_date.isoformat(),
            "created_at": datetime.now(MOSCOW_TZ).isoformat(),
            "file": dated,
            "rows_articles": len(sheets.get("По артикулам", [])),
            "rows_clusters": len(sheets.get("По кластерам", [])),
            "rows_supplies": len(sheets.get("Поставки и прогноз", [])),
            "crossdock_sla_source": self.sla.source,
            "crossdock_sla_updated_at": self.sla.updated_at,
            "posting_quantity_scale": self._posting_qty_scale,
            "unknown_warehouses": sorted(self._unknown_warehouses),
            "supply_diagnostics": self._supply_diagnostics,
        })
        return dated, latest

    def _xlsx(self, sheets: Dict[str, pd.DataFrame]) -> bytes:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            for sheet_name, source in sheets.items():
                df = source.copy() if source is not None else pd.DataFrame()
                if df.empty:
                    df = pd.DataFrame([{"Статус": "Нет данных"}])
                # Даты оставляем датами — Excel их нормально форматирует.
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                ws = writer.book[sheet_name[:31]]
                ws.auto_filter.ref = ws.dimensions
                ws.freeze_panes = "A2"
                fill = PatternFill("solid", fgColor="1F4E78")
                font = Font(color="FFFFFF", bold=True)
                for cell in ws[1]:
                    cell.fill = fill
                    cell.font = font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                for col_idx, col in enumerate(df.columns, 1):
                    header = str(col)
                    vals = [len(header)] + [len(str(x)) for x in df[col].dropna().astype(str).head(100)]
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(52, max(10, max(vals, default=10) + 2))
                    if "Среднесуточные" in header or "%" in header:
                        fmt = "0.0"
                    elif any(x in header for x in ("шт.", "дней")) and "Запас, дней:" not in header:
                        fmt = "#,##0"
                    elif "Дата" in header or "Прогноз" in header or "Слот" in header:
                        fmt = "DD.MM.YYYY"
                    else:
                        fmt = None
                    if fmt:
                        for cell in ws[get_column_letter(col_idx)][1:]:
                            cell.number_format = fmt
                if sheet_name in {"По артикулам", "По кластерам"}:
                    ws.freeze_panes = "E2"
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=min(5, ws.max_column)):
                        for cell in row:
                            if cell.column in {1, 3, 4}:
                                cell.font = Font(bold=True)
        return buf.getvalue()


def env_first(*names: str, default: str = "") -> str:
    """Возвращает первое непустое значение переменной окружения из списка."""
    for name in names:
        value = (os.getenv(name, "") or "").strip()
        if value:
            return value
    return default


def required_env(store: str, suffix: str) -> str:
    options = [f"OZON_{suffix}_{store}", f"OZON_{store}_{suffix}"]
    value = env_first(*options)
    if value:
        return value
    raise RuntimeError(f"Не найден секрет: {' / '.join(options)}")


def main() -> int:
    load_report_env()
    parser = argparse.ArgumentParser(description="Ozon — отдельный сводный отчёт остатков и оборачиваемости")
    parser.add_argument("--store", default=os.getenv("OZON_STORE", "FINICK"))
    parser.add_argument("--target-date", default=os.getenv("OZON_TARGET_DATE", ""))
    args = parser.parse_args()
    store = args.store.upper().strip()
    target = resolve_target_date(args.target_date)

    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    logging.info("=== Остатки и оборачиваемость | %s | %s ===", store, target)

    client_id = required_env(store, "CLIENT_ID")
    api_key = required_env(store, "API_KEY")
    # Поддерживаем все варианты имён, которые использовались в предыдущих
    # Ozon-сборщиках и в GitHub Secrets. Это позволяет запускать отдельный
    # отчёт без создания новых секретов.
    access = env_first("OZON_YC_ACCESS_KEY_ID", "YC_ACCESS_KEY_ID", "AWS_ACCESS_KEY_ID")
    secret = env_first("OZON_YC_SECRET_ACCESS_KEY", "YC_SECRET_ACCESS_KEY", "AWS_SECRET_ACCESS_KEY")
    if not access or not secret:
        raise RuntimeError(
            "Не найдены ключи Object Storage. Поддерживаются: "
            "OZON_YC_ACCESS_KEY_ID/SECRET_ACCESS_KEY, "
            "YC_ACCESS_KEY_ID/YC_SECRET_ACCESS_KEY или AWS_ACCESS_KEY_ID/AWS_SECRET_ACCESS_KEY"
        )
    bucket = env_first("OZON_YC_BUCKET", "YC_BUCKET_NAME", default=DEFAULT_BUCKET) or DEFAULT_BUCKET
    endpoint = env_first(
        "OZON_YC_ENDPOINT_URL", "YC_ENDPOINT_URL",
        default="https://storage.yandexcloud.net"
    )

    storage = Storage(access, secret, bucket, endpoint)
    client = OzonClient(client_id, api_key)
    builder = ReportBuilder(store, target, client, storage)
    sheets = builder.build()
    dated, latest = builder.save(sheets)
    logging.info("ГОТОВО: s3://%s/%s", bucket, dated)
    logging.info("Последняя версия: s3://%s/%s", bucket, latest)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception:
        logging.exception("Ошибка формирования сводного отчёта")
        raise
