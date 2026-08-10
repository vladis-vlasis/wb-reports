#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ozon FBO Daily Data Collector v20

Назначение:
- поддержка магазинов TOPFACE и FINICK;
- единое недельное хранение всех пользовательских отчётов, как в WB;
- ежедневный целевой день всегда «вчера», независимо от времени запуска;
- два автоматических запуска: ночью (EARLY) и после 15:00 МСК (FINAL);
- проверка покрытия последних 60 дней и дозагрузка пропусков;
- отдельный ручной режим history для первичной посуточной загрузки 60 дней;
- рекламная статистика за каждый день и отдельное исследование 14-дневного лага
  в служебных файлах без перезаписи снимков;
- базовые отчёты Seller API, финансы, незавершённые поставки и рекламные отчёты Performance API;
- экономное хранение: успешные RAW JSON по умолчанию не сохраняются;
- справочники товаров/складов хранятся как один актуальный снимок, без ежедневного размножения.

Правило целевой даты:
- OZON_TARGET_DATE / --target-date задана явно -> используем её;
- иначе всегда используется предыдущий календарный день по Москве;
- запуск до 15:00 МСК маркируется EARLY, после 15:00 — FINAL.

Режимы:
- test: один день без ремонта 60-дневной истории;
- daily: вчера + ограниченная дозагрузка пропусков;
- history: все доступные пропуски последних 60 дней с контролем времени;
- archive: ZIP отчётов строго за одну выбранную дату;
- archive_all: один ZIP со всеми накопленными пользовательскими и служебными файлами выбранного магазина или сразу обоих магазинов.

Необязательные методы не останавливают базовые выгрузки. Ошибки и диагностика
фиксируются в служебных файлах. Полные RAW-ответы сохраняются только при
OZON_SAVE_RAW_RESPONSES=1 (по умолчанию выключено).
"""
from __future__ import annotations

import argparse
import io
import json
import logging
import os
import re
import sys
import tempfile
import time
import traceback
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple
from zoneinfo import ZoneInfo

import boto3
import pandas as pd
import requests
from botocore.client import Config
from botocore.exceptions import ClientError
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter

SCRIPT_VERSION = "OZON_FBO_BASIC_PREMIUM_FINANCE_BY_SKU_V20_20260809"

ALLOWED_STORES = {"TOPFACE", "FINICK"}
MOSCOW_TZ = ZoneInfo("Europe/Moscow")
OZON_API_BASE = "https://api-seller.ozon.ru"
OZON_PERFORMANCE_API_BASE = "https://api-performance.ozon.ru"
DEFAULT_BUCKET = "ozon-assist"
DEFAULT_STORE = "TOPFACE"
DEFAULT_HISTORY_DAYS = 60
DEFAULT_REPAIR_DAYS_PER_RUN = 5
MAX_EXCEL_CELL = 32000
AD_LAG_DAYS = 14
MAX_RUNTIME_MINUTES = 220
API_PAUSE_SECONDS = 0.18
SELLER_MIN_INTERVAL_SECONDS = 0.55
PERFORMANCE_MIN_INTERVAL_SECONDS = 0.12

# Версии схемы заставляют history повторно загрузить дни после изменения логики разбора.
REPORT_SCHEMA_VERSIONS: Dict[str, int] = {
    "orders": 4,              # замена дневного среза целиком + financial_data
    "returns": 3,             # замена дневного среза целиком
    "funnel": 4,              # пакетные запросы метрик без ложного присвоения
    "finance_accruals": 1,
    "finance_by_sku": 1,     # управленческий финансовый отчёт по SKU/артикулу
    "realization_report": 1, # официальный отчёт реализации за последний закрытый месяц
    "ad_statistics": 4,       # замена дневного среза и campaign_id
    "ad_product_statistics": 2,
    "ad_orders": 2,
}

FUNNEL_METRICS: List[str] = [
    "hits_view_search",
    "hits_view_pdp",
    "hits_tocart",
    "ordered_units",
    "revenue",
    "cancellations",
    "delivered_units",
    "returns",
    "session_view",
    "conv_tocart",
    "conv_tocart_from_search",
    "conv_order",
    "position_category",
]

# Приоритет фактической даты строки. Техническая дата запуска используется только
# для снимков состояния и никогда не должна распределять события по неделям.
REPORT_DATE_CANDIDATES: Dict[str, List[str]] = {
    "orders": ["Дата", "created_at", "in_process_at"],
    "returns": ["Дата", "logistic.return_date", "return_date", "logistic.final_moment"],
    "funnel": ["Дата", "day"],
    "finance_accruals": ["Дата", "date", "operation_date", "accrual_date"],
    "finance_by_sku": ["Дата"],
    "ad_statistics": ["Дата", "date", "day", "statDate"],
    "ad_product_statistics": ["Дата", "date", "day", "statDate"],
    "ad_orders": ["Дата", "date", "day", "statDate"],
}

# Метрики, которые базовая аналитика Ozon стабильно возвращает и название которых
# можно считать подтверждённым даже без метаданных колонок в ответе.
FUNNEL_ALWAYS_CONFIRMED = {"ordered_units", "revenue"}


# ---------------------------------------------------------------------------
# Environment / dates
# ---------------------------------------------------------------------------

def load_report_env() -> Dict[str, str]:
    """Загружает многострочный REPORT_ENV без перезаписи уже заданных env."""
    raw = os.getenv("REPORT_ENV", "") or ""
    loaded: Dict[str, str] = {}
    for line in raw.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[len("export "):].strip()
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        key, value = key.strip(), value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value
            loaded[key] = value
    return loaded


def parse_iso_date(value: str) -> date:
    return datetime.strptime(value.strip(), "%Y-%m-%d").date()


def resolve_target_date(forced: str = "", now_msk: Optional[datetime] = None) -> date:
    forced = (forced or os.getenv("OZON_TARGET_DATE", "") or "").strip()
    if forced:
        return parse_iso_date(forced)
    now_msk = now_msk or datetime.now(MOSCOW_TZ)
    return now_msk.date() - timedelta(days=1)


def resolve_request_phase(now_msk: Optional[datetime] = None) -> str:
    now_msk = now_msk or datetime.now(MOSCOW_TZ)
    return "FINAL" if now_msk.hour >= 15 else "EARLY"


def iso_week(d: date) -> Tuple[int, int]:
    y, w, _ = d.isocalendar()
    return int(y), int(w)


def week_filename(prefix: str, d: date) -> str:
    year, week = iso_week(d)
    return f"{prefix}_{year}-W{week:02d}.xlsx"


def to_ozon_datetime(d: date, end: bool = False) -> str:
    suffix = "23:59:59.999Z" if end else "00:00:00.000Z"
    return f"{d.isoformat()}T{suffix}"


def mode_period(mode: str, target_date: date) -> Tuple[date, date]:
    """Любой отдельный API-сбор выполняется строго за один календарный день."""
    return target_date, target_date


def history_start(target_date: date, days: int = DEFAULT_HISTORY_DAYS) -> date:
    return target_date - timedelta(days=max(1, int(days)) - 1)


def iter_days(start: date, end: date) -> Iterable[date]:
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def iter_date_chunks(start: date, end: date, chunk_days: int = 1) -> Iterable[Tuple[date, date]]:
    """Разбивает диапазон, чтобы не упираться в MAX_OFFSET_EXCEEDED."""
    current = start
    while current <= end:
        chunk_end = min(end, current + timedelta(days=max(1, chunk_days) - 1))
        yield current, chunk_end
        current = chunk_end + timedelta(days=1)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def batched(values: Sequence[Any], size: int) -> Iterable[List[Any]]:
    """Разбивает список на пакеты допустимого размера API."""
    size = max(1, int(size))
    for idx in range(0, len(values), size):
        yield list(values[idx:idx + size])

def safe_json(value: Any) -> str:
    try:
        text = json.dumps(value, ensure_ascii=False, default=str, separators=(",", ":"))
    except Exception:
        text = str(value)
    return text[:MAX_EXCEL_CELL]


def scalarize(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return safe_json(value)
    return value


def to_number(value: Any) -> Optional[float]:
    """Безопасно переводит денежные и числовые значения API в float."""
    if value in (None, "", "null"):
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
        return float(cleaned)
    except (TypeError, ValueError):
        return None



def parse_json_value(value: Any) -> Any:
    """Возвращает dict/list из JSON-строки, не меняя уже разобранные значения."""
    if isinstance(value, (dict, list, tuple)):
        return value
    if value in (None, "", "null"):
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if text[:1] in {"{", "["}:
            try:
                return json.loads(text)
            except Exception:
                return None
    return None


def money_amount(value: Any) -> Optional[float]:
    """Извлекает рублёвую сумму из числа или объекта Ozon {'amount': ...}."""
    if isinstance(value, Mapping):
        for key in ("amount", "value", "units"):
            if key in value:
                num = to_number(value.get(key))
                if num is not None:
                    return num
        return None
    return to_number(value)


def infer_finance_quantity(seller_price: Any, sale_amount: Any) -> float:
    """Пытается восстановить количество единиц в финансовой операции."""
    price = money_amount(seller_price)
    amount = money_amount(sale_amount)
    if price not in (None, 0) and amount is not None:
        ratio = abs(amount / price)
        nearest = round(ratio)
        if nearest >= 1 and abs(ratio - nearest) <= 0.02:
            return float(nearest)
    return 1.0


def round_money(value: Any) -> int:
    """Управленческие суммы показываем в целых рублях."""
    number = to_number(value) or 0.0
    return int(round(number))


def round_percent(value: Any, digits: int = 1) -> Optional[float]:
    number = to_number(value)
    return None if number is None else round(number, digits)


def safe_ratio(numerator: Any, denominator: Any, multiplier: float = 100.0) -> Optional[float]:
    n = to_number(numerator)
    d = to_number(denominator)
    if n is None or d in (None, 0):
        return None
    return n / d * multiplier


def flatten_record(record: Mapping[str, Any], prefix: str = "") -> Dict[str, Any]:
    """Плоские скаляры; вложенные списки/словари сохраняются JSON для сохранения всех данных."""
    out: Dict[str, Any] = {}
    for key, value in record.items():
        name = f"{prefix}{key}" if prefix else str(key)
        if isinstance(value, dict):
            # Делаем и плоские дочерние поля, и исходный JSON для полной сохранности.
            out[name] = safe_json(value)
            for child_key, child_val in value.items():
                child_name = f"{name}.{child_key}"
                out[child_name] = scalarize(child_val)
        else:
            out[name] = scalarize(value)
    return out


def records_to_df(records: Iterable[Mapping[str, Any]], extra: Optional[Dict[str, Any]] = None) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for rec in records:
        if not isinstance(rec, Mapping):
            rows.append({"value": scalarize(rec)})
            continue
        row = flatten_record(rec)
        if extra:
            row.update(extra)
        rows.append(row)
    return pd.DataFrame(rows)


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(c).strip() for c in out.columns]
    return out


def first_existing(df: pd.DataFrame, candidates: Sequence[str]) -> Optional[str]:
    lookup = {str(c).strip().lower(): c for c in df.columns}
    for candidate in candidates:
        actual = lookup.get(candidate.strip().lower())
        if actual is not None:
            return actual
    return None


def ensure_date_column(df: pd.DataFrame, column: str, value: date) -> pd.DataFrame:
    out = df.copy()
    out[column] = value.isoformat()
    return out


def dedupe_merge(old: pd.DataFrame, new: pd.DataFrame, keys: Sequence[str]) -> pd.DataFrame:
    if old is None or old.empty:
        combined = new.copy()
    elif new is None or new.empty:
        combined = old.copy()
    else:
        combined = pd.concat([old, new], ignore_index=True, sort=False)
    if combined.empty:
        return combined
    valid = [k for k in keys if k in combined.columns]
    if valid:
        combined = combined.drop_duplicates(subset=valid, keep="last")
    else:
        combined = combined.drop_duplicates(keep="last")
    return combined.reset_index(drop=True)


def dataframe_unique_count(df: pd.DataFrame, keys: Sequence[str]) -> int:
    if df is None or df.empty:
        return 0
    valid = [k for k in keys if k in df.columns]
    if not valid:
        return int(len(df.drop_duplicates()))
    return int(len(df.drop_duplicates(subset=valid)))


def normalize_identifier(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def extract_items(data: Any, paths: Sequence[Sequence[str]]) -> List[Dict[str, Any]]:
    """Извлекает список из нескольких известных форматов ответа Ozon."""
    for path in paths:
        cur = data
        ok = True
        for key in path:
            if not isinstance(cur, Mapping) or key not in cur:
                ok = False
                break
            cur = cur[key]
        if ok and isinstance(cur, list):
            return [x for x in cur if isinstance(x, Mapping)]
    if isinstance(data, list):
        return [x for x in data if isinstance(x, Mapping)]
    return []


def extract_cursor(data: Any) -> str:
    candidates = [
        ("result", "cursor"), ("cursor",), ("result", "last_id"), ("last_id",),
        ("result", "next_cursor"), ("next_cursor",),
    ]
    for path in candidates:
        cur = data
        ok = True
        for key in path:
            if not isinstance(cur, Mapping) or key not in cur:
                ok = False
                break
            cur = cur[key]
        if ok and cur not in (None, ""):
            return str(cur)
    return ""


def extract_total(data: Any) -> Optional[int]:
    for path in [("result", "total"), ("total",), ("result", "total_count"), ("total_count",)]:
        cur = data
        ok = True
        for key in path:
            if not isinstance(cur, Mapping) or key not in cur:
                ok = False
                break
            cur = cur[key]
        if ok:
            try:
                return int(cur)
            except Exception:
                pass
    return None


# ---------------------------------------------------------------------------
# S3
# ---------------------------------------------------------------------------

class S3Storage:
    def __init__(self, access_key: str, secret_key: str, bucket: str, endpoint: str):
        self.bucket = bucket
        self.client = boto3.client(
            "s3",
            endpoint_url=endpoint,
            aws_access_key_id=access_key,
            aws_secret_access_key=secret_key,
            region_name="ru-central1",
            config=Config(signature_version="s3v4", read_timeout=300, connect_timeout=60,
                          retries={"max_attempts": 7, "mode": "standard"}),
        )

    def ensure_bucket(self) -> None:
        self.client.head_bucket(Bucket=self.bucket)

    def exists(self, key: str) -> bool:
        try:
            self.client.head_object(Bucket=self.bucket, Key=key)
            return True
        except ClientError as exc:
            code = str(exc.response.get("Error", {}).get("Code", ""))
            if code in {"404", "NoSuchKey", "NotFound"}:
                return False
            raise

    def read_bytes(self, key: str) -> bytes:
        return self.client.get_object(Bucket=self.bucket, Key=key)["Body"].read()

    def read_excel(self, key: str) -> pd.DataFrame:
        if not self.exists(key):
            return pd.DataFrame()
        return pd.read_excel(io.BytesIO(self.read_bytes(key)))

    def read_json(self, key: str, default: Any = None) -> Any:
        if not self.exists(key):
            return {} if default is None else default
        try:
            return json.loads(self.read_bytes(key).decode("utf-8"))
        except Exception:
            logging.exception("Не удалось прочитать JSON: s3://%s/%s", self.bucket, key)
            return {} if default is None else default

    def upload_bytes(self, key: str, data: bytes, content_type: Optional[str] = None) -> None:
        kwargs: Dict[str, Any] = {"Bucket": self.bucket, "Key": key, "Body": data}
        if content_type:
            kwargs["ContentType"] = content_type
        self.client.put_object(**kwargs)

    def upload_file(self, local_path: str, key: str) -> None:
        self.client.upload_file(local_path, self.bucket, key)

    def upload_json(self, key: str, value: Any) -> None:
        self.upload_bytes(key, json.dumps(value, ensure_ascii=False, indent=2, default=str).encode("utf-8"),
                          "application/json")

    def list_keys(self, prefix: str) -> List[str]:
        return [obj["Key"] for obj in self.list_objects(prefix)]

    def list_objects(self, prefix: str) -> List[Dict[str, Any]]:
        out: List[Dict[str, Any]] = []
        token: Optional[str] = None
        while True:
            kwargs: Dict[str, Any] = {"Bucket": self.bucket, "Prefix": prefix}
            if token:
                kwargs["ContinuationToken"] = token
            resp = self.client.list_objects_v2(**kwargs)
            for item in resp.get("Contents", []):
                out.append({
                    "Key": item.get("Key", ""),
                    "Size": int(item.get("Size", 0) or 0),
                    "LastModified": item.get("LastModified"),
                    "ETag": str(item.get("ETag", "") or "").strip('"'),
                })
            if not resp.get("IsTruncated"):
                break
            token = resp.get("NextContinuationToken")
        return out


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------

class OzonApiError(RuntimeError):
    def __init__(self, method: str, path: str, status: int, message: str, response_text: str = ""):
        super().__init__(f"{method} {path}: HTTP {status}: {message}")
        self.method = method
        self.path = path
        self.status = status
        self.response_text = response_text


class OzonSellerClient:
    def __init__(self, client_id: str, api_key: str, timeout: int = 180):
        self.base = OZON_API_BASE.rstrip("/")
        self.timeout = timeout
        self.session = requests.Session()
        self._last_request_at = 0.0
        self.session.headers.update({
            "Client-Id": str(client_id),
            "Api-Key": str(api_key),
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": f"ozon-assist/{SCRIPT_VERSION}",
        })


    def _throttle(self) -> None:
        elapsed = time.monotonic() - self._last_request_at
        wait = SELLER_MIN_INTERVAL_SECONDS - elapsed
        if wait > 0:
            time.sleep(wait)
        self._last_request_at = time.monotonic()

    def post(self, path: str, payload: Mapping[str, Any], retries: int = 6) -> Dict[str, Any]:
        url = self.base + path
        last_error = ""
        for attempt in range(1, retries + 1):
            try:
                self._throttle()
                resp = self.session.post(url, json=payload, timeout=self.timeout)
            except requests.RequestException as exc:
                last_error = str(exc)
                if attempt == retries:
                    raise OzonApiError("POST", path, 0, last_error) from exc
                time.sleep(min(60, 2 ** attempt))
                continue

            if resp.status_code == 200:
                try:
                    return resp.json()
                except Exception as exc:
                    raise OzonApiError("POST", path, 200, "Ответ не является JSON", resp.text[:2000]) from exc

            text = resp.text[:6000]
            if resp.status_code in {429, 500, 502, 503, 504} and attempt < retries:
                wait = resp.headers.get("Retry-After")
                try:
                    seconds = max(1, int(float(wait))) if wait else min(90, 2 ** attempt)
                except Exception:
                    seconds = min(90, 2 ** attempt)
                logging.warning("Повтор %s/%s %s после HTTP %s через %s сек", attempt, retries, path,
                                resp.status_code, seconds)
                time.sleep(seconds)
                continue

            try:
                body = resp.json()
                msg = body.get("message") or body.get("error") or body.get("code") or text
            except Exception:
                msg = text
            raise OzonApiError("POST", path, resp.status_code, str(msg), text)

        raise OzonApiError("POST", path, 0, last_error or "Неизвестная ошибка")

    def get(self, path: str, params: Optional[Mapping[str, Any]] = None, retries: int = 6) -> Dict[str, Any]:
        url = self.base + path
        last_error = ""
        for attempt in range(1, retries + 1):
            try:
                self._throttle()
                resp = self.session.get(url, params=dict(params or {}), timeout=self.timeout)
            except requests.RequestException as exc:
                last_error = str(exc)
                if attempt == retries:
                    raise OzonApiError("GET", path, 0, last_error) from exc
                time.sleep(min(60, 2 ** attempt))
                continue

            if resp.status_code == 200:
                try:
                    return resp.json()
                except Exception as exc:
                    raise OzonApiError("GET", path, 200, "Ответ не является JSON", resp.text[:2000]) from exc

            text = resp.text[:6000]
            if resp.status_code in {429, 500, 502, 503, 504} and attempt < retries:
                wait = resp.headers.get("Retry-After")
                try:
                    seconds = max(1, int(float(wait))) if wait else min(90, 2 ** attempt)
                except Exception:
                    seconds = min(90, 2 ** attempt)
                time.sleep(seconds)
                continue
            try:
                body = resp.json()
                msg = body.get("message") or body.get("error") or body.get("code") or text
            except Exception:
                msg = text
            raise OzonApiError("GET", path, resp.status_code, str(msg), text)
        raise OzonApiError("GET", path, 0, last_error or "Неизвестная ошибка")

    def cursor_pages(
        self,
        path: str,
        payload: Dict[str, Any],
        item_paths: Sequence[Sequence[str]],
        limit: int = 1000,
        max_pages: int = 500,
        cursor_field: str = "cursor",
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        all_items: List[Dict[str, Any]] = []
        raw_pages: List[Dict[str, Any]] = []
        cursor = str(payload.get(cursor_field, "") or "")
        for page in range(1, max_pages + 1):
            body = dict(payload)
            body["limit"] = min(limit, int(body.get("limit", limit)))
            body[cursor_field] = cursor
            data = self.post(path, body)
            raw_pages.append(data)
            items = extract_items(data, item_paths)
            all_items.extend(items)
            next_cursor = extract_cursor(data)
            if not items or not next_cursor or next_cursor == cursor:
                break
            cursor = next_cursor
            time.sleep(0.12)
        return all_items, raw_pages

    def offset_pages(
        self,
        path: str,
        payload: Dict[str, Any],
        item_paths: Sequence[Sequence[str]],
        limit: int = 1000,
        max_pages: int = 500,
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        all_items: List[Dict[str, Any]] = []
        raw_pages: List[Dict[str, Any]] = []
        offset = int(payload.get("offset", 0) or 0)
        for page in range(1, max_pages + 1):
            body = dict(payload)
            body["limit"] = min(limit, int(body.get("limit", limit)))
            body["offset"] = offset
            data = self.post(path, body)
            raw_pages.append(data)
            items = extract_items(data, item_paths)
            all_items.extend(items)
            total = extract_total(data)
            if not items or len(items) < body["limit"] or (total is not None and len(all_items) >= total):
                break
            offset += len(items)
            time.sleep(0.12)
        return all_items, raw_pages



class OzonPerformanceClient:
    """Клиент Performance API.

    Авторизация:
    POST /api/client/token
    client_id + client_secret + grant_type=client_credentials.

    Методы чтения сделаны с fallback-вариантами, потому что Ozon постепенно
    переводит рекламные инструменты между версиями API. Все успешные и
    неуспешные ответы сохраняются сборщиком в raw JSON.
    """

    def __init__(
        self,
        client_id: str,
        client_secret: str,
        timeout: int = 180,
    ):
        self.base = OZON_PERFORMANCE_API_BASE.rstrip("/")
        self.client_id = str(client_id)
        self.client_secret = str(client_secret)
        self.timeout = timeout
        self.session = requests.Session()
        self.session.headers.update({
            "Accept": "application/json",
            "User-Agent": f"ozon-assist/{SCRIPT_VERSION}",
        })
        self.access_token = ""
        self.token_expires_at = 0.0
        self._last_request_at = 0.0


    def _throttle(self) -> None:
        elapsed = time.monotonic() - self._last_request_at
        wait = PERFORMANCE_MIN_INTERVAL_SECONDS - elapsed
        if wait > 0:
            time.sleep(wait)
        self._last_request_at = time.monotonic()

    def _ensure_token(self) -> None:
        if self.access_token and time.time() < self.token_expires_at - 60:
            return

        url = self.base + "/api/client/token"
        payload = {
            "client_id": self.client_id,
            "client_secret": self.client_secret,
            "grant_type": "client_credentials",
        }
        self._throttle()
        response = self.session.post(
            url,
            json=payload,
            headers={"Content-Type": "application/json"},
            timeout=self.timeout,
        )
        if response.status_code != 200:
            raise OzonApiError(
                "POST",
                "/api/client/token",
                response.status_code,
                response.text[:2000],
                response.text[:6000],
            )

        try:
            body = response.json()
        except Exception as exc:
            raise OzonApiError(
                "POST",
                "/api/client/token",
                200,
                "Ответ token не является JSON",
                response.text[:2000],
            ) from exc

        token = (
            body.get("access_token")
            or body.get("accessToken")
            or body.get("token")
        )
        if not token:
            raise OzonApiError(
                "POST",
                "/api/client/token",
                200,
                "В ответе отсутствует access_token",
                safe_json(body),
            )

        self.access_token = str(token)
        expires_in = body.get("expires_in") or body.get("expiresIn") or 1800
        try:
            expires_seconds = int(expires_in)
        except Exception:
            expires_seconds = 1800
        self.token_expires_at = time.time() + max(300, expires_seconds)
        self.session.headers.update({
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": "application/json",
        })

    def request(
        self,
        method: str,
        path: str,
        *,
        payload: Optional[Mapping[str, Any]] = None,
        params: Optional[Mapping[str, Any]] = None,
        retries: int = 6,
    ) -> Any:
        self._ensure_token()
        url = self.base + path
        last_error = ""

        for attempt in range(1, retries + 1):
            try:
                self._throttle()
                response = self.session.request(
                    method.upper(),
                    url,
                    json=dict(payload) if payload is not None else None,
                    params=dict(params) if params is not None else None,
                    timeout=self.timeout,
                )
            except requests.RequestException as exc:
                last_error = str(exc)
                if attempt == retries:
                    raise OzonApiError(method.upper(), path, 0, last_error) from exc
                time.sleep(min(60, 2 ** attempt))
                continue

            if response.status_code in {200, 201, 202}:
                if not response.text.strip():
                    return {}
                try:
                    return response.json()
                except Exception:
                    return {
                        "_content_type": response.headers.get("Content-Type", ""),
                        "_text": response.text,
                    }

            if response.status_code == 401 and attempt < retries:
                self.access_token = ""
                self.token_expires_at = 0.0
                self._ensure_token()
                continue

            if response.status_code in {429, 500, 502, 503, 504} and attempt < retries:
                retry_after = response.headers.get("Retry-After")
                try:
                    seconds = int(float(retry_after)) if retry_after else min(90, 2 ** attempt)
                except Exception:
                    seconds = min(90, 2 ** attempt)
                logging.warning(
                    "Performance API повтор %s/%s %s после HTTP %s через %s сек",
                    attempt,
                    retries,
                    path,
                    response.status_code,
                    seconds,
                )
                time.sleep(max(1, seconds))
                continue

            text_body = response.text[:6000]
            try:
                body = response.json()
                message = (
                    body.get("message")
                    or body.get("error")
                    or body.get("code")
                    or text_body
                )
            except Exception:
                message = text_body
            raise OzonApiError(
                method.upper(),
                path,
                response.status_code,
                str(message),
                text_body,
            )

        raise OzonApiError(method.upper(), path, 0, last_error or "Неизвестная ошибка")

    def get(self, path: str, params: Optional[Mapping[str, Any]] = None) -> Any:
        return self.request("GET", path, params=params)

    def post(self, path: str, payload: Mapping[str, Any]) -> Any:
        return self.request("POST", path, payload=payload)

    def try_variants(
        self,
        variants: Sequence[Tuple[str, str, Optional[Mapping[str, Any]], Optional[Mapping[str, Any]]]],
    ) -> Tuple[Any, Dict[str, Any]]:
        """Пробует варианты method/path/payload/params до первого успеха."""
        attempts: List[Dict[str, Any]] = []
        last_exc: Optional[Exception] = None

        for method, path, payload, params in variants:
            try:
                data = self.request(
                    method,
                    path,
                    payload=payload,
                    params=params,
                )
                meta = {
                    "chosen_method": method,
                    "chosen_path": path,
                    "chosen_payload": payload,
                    "chosen_params": params,
                    "attempts": attempts,
                }
                return data, meta
            except OzonApiError as exc:
                last_exc = exc
                attempts.append({
                    "method": method,
                    "path": path,
                    "payload": payload,
                    "params": params,
                    "status": exc.status,
                    "error": str(exc),
                    "response_text": exc.response_text,
                })
                if exc.status in {400, 404, 405, 422}:
                    continue
                raise

        if last_exc:
            raise last_exc
        raise RuntimeError("Не задано ни одного варианта Performance API")

# ---------------------------------------------------------------------------
# Report definitions
# ---------------------------------------------------------------------------

@dataclass
class ReportResult:
    code: str
    title: str
    folder: str
    filename_prefix: str
    df: pd.DataFrame = field(default_factory=pd.DataFrame)
    raw: Any = None
    date_column: str = "Дата"
    keys: Sequence[str] = field(default_factory=list)
    snapshot: bool = False
    optional: bool = False
    method_paths: List[str] = field(default_factory=list)
    status: str = "OK"
    message: str = ""
    s3_key: str = ""
    local_path: str = ""
    skip_standard_save: bool = False
    save_audit: Dict[str, Any] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Collector
# ---------------------------------------------------------------------------

class OzonFboCollector:
    def __init__(
        self,
        client: OzonSellerClient,
        storage: S3Storage,
        store: str,
        target_date: date,
        mode: str,
        workdir: Path,
        performance_client: Optional[OzonPerformanceClient] = None,
        fast_history: bool = False,
    ):
        self.client = client
        self.performance_client = performance_client
        self.fast_history = bool(fast_history)
        self.storage = storage
        self.store = store.upper().strip()
        self.target_date = target_date
        self.mode = str(mode or "test").strip().lower()
        self.period_from, self.period_to = mode_period(self.mode, target_date)
        self.workdir = workdir
        self.workdir.mkdir(parents=True, exist_ok=True)
        self.run_id = datetime.now(MOSCOW_TZ).strftime("%Y%m%d_%H%M%S")
        self.results: List[ReportResult] = []
        self.errors: List[Dict[str, Any]] = []
        self.catalog_items: List[Dict[str, Any]] = []
        self.catalog_ids: List[int] = []
        self.offer_ids: List[str] = []
        self.sku_ids: List[int] = []
        self.campaign_ids_override: List[str] = []
        self.active_campaign_ids_override: List[str] = []
        # Справочник для понятных пользовательских отчётов: SKU -> артикул/название.
        self.reference_sku_to_offer: Dict[str, str] = {}
        self.reference_sku_to_name: Dict[str, str] = {}
        self.request_phase = resolve_request_phase()
        self.request_started_at = datetime.now(MOSCOW_TZ)

    @property
    def base_prefix(self) -> str:
        return "Отчёты"

    def log_error(self, code: str, path: str, exc: Exception, optional: bool) -> None:
        row = {
            "run_id": self.run_id,
            "report": code,
            "path": path,
            "optional": optional,
            "error_type": type(exc).__name__,
            "error": str(exc),
            "traceback": traceback.format_exc(limit=8),
        }
        if isinstance(exc, OzonApiError):
            row["http_status"] = exc.status
            row["response_text"] = exc.response_text
        self.errors.append(row)
        level = logging.WARNING if optional else logging.ERROR
        logging.log(level, "%s: %s", code, exc)


    def _postprocess_report_df(self, code: str, df: pd.DataFrame) -> pd.DataFrame:
        df = normalize_columns(df if df is not None else pd.DataFrame())
        if df.empty:
            return df

        if code == "orders" and "created_at" in df.columns:
            parsed = pd.to_datetime(df["created_at"], errors="coerce", utc=True)
            df["Дата"] = parsed.dt.strftime("%Y-%m-%d")
        elif code == "returns":
            for col in ("logistic.return_date", "return_date", "logistic.final_moment"):
                if col in df.columns:
                    parsed = pd.to_datetime(df[col], errors="coerce", utc=True)
                    if parsed.notna().any():
                        df["Дата"] = parsed.dt.strftime("%Y-%m-%d")
                        break
        elif code in {"ad_statistics", "ad_product_statistics", "ad_orders"}:
            if "Дата" not in df.columns:
                for col in ("date", "day", "statDate"):
                    if col in df.columns:
                        df["Дата"] = df[col].astype(str).str[:10]
                        break
            if "campaign_id" not in df.columns:
                for col in ("campaignId", "campaign", "id"):
                    if col in df.columns:
                        df["campaign_id"] = df[col].map(normalize_identifier)
                        break
            else:
                df["campaign_id"] = df["campaign_id"].map(normalize_identifier)
            if "status" in df.columns:
                df = df[df["status"].astype(str).str.upper() != "ERROR"].copy()
            if "Дата" in df.columns and "campaign_id" in df.columns:
                df = df[
                    df["Дата"].astype(str).str.match(r"^\d{4}-\d{2}-\d{2}$", na=False)
                    & df["campaign_id"].astype(str).str.len().gt(0)
                ].copy()

            if code == "ad_statistics" and not df.empty:
                source_map = {
                    "Показы": "views", "Клики": "clicks", "Расход, ₽": "moneySpent",
                    "Рекламные заказы, шт.": "orders", "Рекламная выручка, ₽": "ordersMoney",
                }
                for target, source in source_map.items():
                    if source in df.columns:
                        df[target] = df[source].map(to_number)
                if {"Показы", "Клики"}.issubset(df.columns):
                    df["CTR, %"] = [safe_ratio(c, v) for c, v in zip(df["Клики"], df["Показы"])]
                if {"Расход, ₽", "Клики"}.issubset(df.columns):
                    df["CPC, ₽"] = [safe_ratio(s, c, multiplier=1.0) for s, c in zip(df["Расход, ₽"], df["Клики"])]
                if {"Рекламные заказы, шт.", "Клики"}.issubset(df.columns):
                    df["CR клик → заказ, %"] = [safe_ratio(o, c) for o, c in zip(df["Рекламные заказы, шт."], df["Клики"])]
                if {"Расход, ₽", "Рекламные заказы, шт."}.issubset(df.columns):
                    df["CPO, ₽"] = [safe_ratio(s, o, multiplier=1.0) for s, o in zip(df["Расход, ₽"], df["Рекламные заказы, шт."])]
                if {"Расход, ₽", "Рекламная выручка, ₽"}.issubset(df.columns):
                    df["ДРР, %"] = [safe_ratio(s, r) for s, r in zip(df["Расход, ₽"], df["Рекламная выручка, ₽"])]

        df["_schema_version"] = int(REPORT_SCHEMA_VERSIONS.get(code, 1))
        return df.reset_index(drop=True)

    def _run(self, code: str, title: str, folder: str, filename_prefix: str, fn,
             date_column: str, keys: Sequence[str], snapshot: bool = False,
             optional: bool = False) -> ReportResult:
        logging.info("=== %s ===", title)
        result = ReportResult(code=code, title=title, folder=folder,
                              filename_prefix=filename_prefix, date_column=date_column,
                              keys=list(keys), snapshot=snapshot, optional=optional)
        try:
            df, raw, methods = fn()
            result.df = self._postprocess_report_df(code, df)
            result.raw = raw
            result.method_paths = list(methods)
            if result.df.empty:
                result.status = "EMPTY"
                result.message = "Метод доступен, но строки отсутствуют"
            logging.info("%s: %s строк", title, len(result.df))
        except Exception as exc:
            self.log_error(code, title, exc, optional)
            result.status = "SKIPPED_OPTIONAL" if optional else "ERROR"
            result.message = str(exc)
        self.results.append(result)
        return result

    def fetch_product_catalog(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        items, raw = self.client.cursor_pages(
            "/v3/product/list",
            {"filter": {"visibility": "ALL"}, "limit": 1000, "last_id": ""},
            [("result", "items"), ("items",)],
            cursor_field="last_id",
        )
        # cursor_pages expects generic cursor extraction and body field last_id works here.
        self.catalog_items = items
        ids: List[int] = []
        offers: List[str] = []
        for item in items:
            try:
                if item.get("product_id") is not None:
                    ids.append(int(item["product_id"]))
            except Exception:
                pass
            if item.get("offer_id") not in (None, ""):
                offers.append(str(item["offer_id"]))
        self.catalog_ids = sorted(set(ids))
        self.offer_ids = sorted(set(offers))
        df = records_to_df(items, {"Дата снимка": self.target_date.isoformat()})
        return df, raw, ["/v3/product/list"]

    def _batch(self, values: Sequence[Any], size: int) -> Iterable[List[Any]]:
        for i in range(0, len(values), size):
            yield list(values[i:i + size])

    def fetch_product_info(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        if not self.offer_ids and not self.catalog_ids:
            return pd.DataFrame(), [], ["/v3/product/info/list"]
        rows: List[Dict[str, Any]] = []
        raw: List[Any] = []
        # v3 accepts offer_id/product_id lists. Use offer_id to keep seller identity stable.
        source = self.offer_ids or self.catalog_ids
        for batch in self._batch(source, 100):
            payload = {"offer_id": batch, "product_id": [], "sku": []} if self.offer_ids else {
                "offer_id": [], "product_id": batch, "sku": []}
            data = self.client.post("/v3/product/info/list", payload)
            raw.append(data)
            batch_items = extract_items(data, [("items",), ("result", "items")])
            rows.extend(batch_items)
            for item in batch_items:
                offer = str(item.get("offer_id") or item.get("offerId") or "").strip()
                name = str(item.get("name") or item.get("product_name") or "").strip()
                for candidate in ("sku", "fbo_sku", "fbs_sku"):
                    try:
                        value = item.get(candidate)
                        if value not in (None, "", 0):
                            sku_text = normalize_identifier(value)
                            self.sku_ids.append(int(float(value)))
                            if offer:
                                self.reference_sku_to_offer[sku_text] = offer
                            if name:
                                self.reference_sku_to_name[sku_text] = name
                    except Exception:
                        pass
            time.sleep(0.12)
        self.sku_ids = sorted(set(self.sku_ids))
        return records_to_df(rows, {"Дата снимка": self.target_date.isoformat()}), raw, ["/v3/product/info/list"]

    def fetch_prices(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        items, raw = self.client.cursor_pages(
            "/v5/product/info/prices",
            {"filter": {"visibility": "ALL"}, "limit": 1000, "cursor": ""},
            [("items",), ("result", "items")],
        )
        df = records_to_df(items, {"Дата снимка": self.target_date.isoformat()})
        # Расчётный аналог поддержки скидки, только когда поля доступны.
        seller_col = first_existing(df, ["price.marketing_seller_price", "marketing_seller_price", "price"])
        buyer_col = first_existing(df, ["price.marketing_price", "marketing_price", "customer_price"])
        if seller_col and buyer_col:
            seller = pd.to_numeric(df[seller_col], errors="coerce")
            buyer = pd.to_numeric(df[buyer_col], errors="coerce")
            df["Расчётная поддержка скидки Ozon, руб"] = (seller - buyer).clip(lower=0)
            df["Расчётная поддержка скидки Ozon, %"] = ((seller - buyer) / seller.replace(0, pd.NA) * 100).clip(lower=0)
        return df, raw, ["/v5/product/info/prices"]

    # ------------------------- stocks -------------------------
    def fetch_stocks(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        items, raw = self.client.cursor_pages(
            "/v4/product/info/stocks",
            {"filter": {"visibility": "ALL"}, "limit": 1000, "cursor": ""},
            [("items",), ("result", "items")],
        )
        rows: List[Dict[str, Any]] = []
        for item in items:
            base = {k: v for k, v in item.items() if k != "stocks"}
            stocks = item.get("stocks")
            if isinstance(stocks, list) and stocks:
                for stock in stocks:
                    row = dict(base)
                    if isinstance(stock, Mapping):
                        row.update({f"stock.{k}": v for k, v in stock.items()})
                    else:
                        row["stock"] = stock
                    rows.append(row)
            else:
                rows.append(item)
        return records_to_df(rows, {"Дата снимка": self.target_date.isoformat()}), raw, ["/v4/product/info/stocks"]

    def fetch_stock_on_warehouses(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        # Расширенный складской разрез. Не во всех кабинетах/версиях доступен.
        items, raw = self.client.offset_pages(
            "/v2/analytics/stock_on_warehouses",
            {"limit": 1000, "offset": 0, "warehouse_type": "ALL"},
            [("result", "rows"), ("rows",), ("result", "items")],
        )
        return records_to_df(items, {"Дата снимка": self.target_date.isoformat()}), raw, ["/v2/analytics/stock_on_warehouses"]

    def fetch_turnover(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        data = self.client.post("/v1/analytics/turnover/stocks", {"limit": 1000, "offset": 0})
        items = extract_items(data, [("items",), ("result", "items"), ("result", "rows"), ("rows",)])
        # Если метод offset-пагинируемый, добираем страницы.
        if len(items) >= 1000:
            items, raw = self.client.offset_pages(
                "/v1/analytics/turnover/stocks", {"limit": 1000, "offset": 0},
                [("items",), ("result", "items"), ("result", "rows"), ("rows",)]
            )
        else:
            raw = [data]
        return records_to_df(items, {"Дата снимка": self.target_date.isoformat()}), raw, ["/v1/analytics/turnover/stocks"]
    def fetch_analytics_stocks(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        """Дополнительная аналитика остатков FBO по SKU, пакетами до 100."""
        endpoint = "/v1/analytics/stocks"
        skus = [int(x) for x in self.sku_ids if str(x).isdigit()]
        if not skus:
            return pd.DataFrame(), {"status": "NO_SKUS"}, [endpoint]

        rows: List[Dict[str, Any]] = []
        raw: List[Any] = []
        last_exc: Optional[Exception] = None
        for sku_batch in batched(sorted(set(skus)), 100):
            variants = [
                {"skus": sku_batch, "limit": 1000, "offset": 0},
                {"skus": [str(x) for x in sku_batch], "limit": 1000, "offset": 0},
            ]
            batch_ok = False
            for payload in variants:
                try:
                    items, pages = self.client.offset_pages(
                        endpoint, payload,
                        [("result", "rows"), ("rows",), ("result", "items"), ("items",)],
                        limit=1000, max_pages=100,
                    )
                    raw.append({"request": payload, "responses": pages})
                    rows.extend(items)
                    batch_ok = True
                    break
                except OzonApiError as exc:
                    last_exc = exc
                    raw.append({"request": payload, "status": "ERROR", "error": str(exc)})
                    if exc.status in {400, 404, 422}:
                        continue
                    raise
            if not batch_ok and last_exc:
                raise last_exc
            time.sleep(API_PAUSE_SECONDS)

        return records_to_df(rows, {"Дата снимка": self.target_date.isoformat()}), raw, [endpoint]

    # ------------------------- orders/returns -------------------------
    def fetch_fbo_postings(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        """FBO-заказы с полным financial_data и продуктовым финансовым разрезом."""
        all_postings: List[Dict[str, Any]] = []
        raw_all: List[Any] = []
        for day_from, day_to in iter_date_chunks(self.period_from, self.period_to, chunk_days=1):
            items, raw = self.client.offset_pages(
                "/v2/posting/fbo/list",
                {
                    "dir": "ASC",
                    "filter": {
                        "since": to_ozon_datetime(day_from),
                        "to": to_ozon_datetime(day_to, end=True),
                        "status": "",
                    },
                    "limit": 1000,
                    "offset": 0,
                    "translit": True,
                    "with": {"analytics_data": True, "financial_data": True},
                },
                [("result",), ("result", "postings"), ("postings",)],
                max_pages=100,
            )
            raw_all.extend(raw)
            all_postings.extend(items)
            logging.info("Заказы FBO: получен день %s, отправлений %s", day_from, len(items))

        def product_signature(value: Mapping[str, Any]) -> Tuple[str, str, str]:
            return (
                str(value.get("sku") or value.get("product_id") or ""),
                str(value.get("offer_id") or value.get("offerId") or ""),
                str(value.get("name") or ""),
            )

        rows: List[Dict[str, Any]] = []
        for posting in all_postings:
            base = {k: v for k, v in posting.items() if k not in {"products", "financial_data"}}
            financial_data = posting.get("financial_data") if isinstance(posting.get("financial_data"), Mapping) else {}
            financial_products = financial_data.get("products") if isinstance(financial_data, Mapping) else []
            if not isinstance(financial_products, list):
                financial_products = []
            financial_base = {k: v for k, v in financial_data.items() if k != "products"} if isinstance(financial_data, Mapping) else {}

            products = posting.get("products")
            if not isinstance(products, list) or not products:
                products = [{}]

            used_fin: set[int] = set()
            for idx, product in enumerate(products):
                product = product if isinstance(product, Mapping) else {"value": product}
                row = dict(base)
                row.update({f"product.{k}": v for k, v in product.items()})
                row.update({f"financial_data.{k}": v for k, v in financial_base.items()})

                matched: Optional[Mapping[str, Any]] = None
                sig = product_signature(product)
                for fin_idx, fin in enumerate(financial_products):
                    if fin_idx in used_fin or not isinstance(fin, Mapping):
                        continue
                    if any(sig) and product_signature(fin) == sig:
                        matched = fin
                        used_fin.add(fin_idx)
                        break
                if matched is None and idx < len(financial_products) and isinstance(financial_products[idx], Mapping):
                    matched = financial_products[idx]
                    used_fin.add(idx)
                if matched is not None:
                    row.update({f"financial_product.{k}": v for k, v in matched.items()})

                seller_price = to_number(
                    product.get("price")
                    or (matched or {}).get("price")
                    or (matched or {}).get("seller_price")
                )
                buyer_price = to_number(
                    (matched or {}).get("client_price")
                    or (matched or {}).get("customer_price")
                    or (matched or {}).get("buyer_price")
                )
                quantity = to_number(product.get("quantity") or (matched or {}).get("quantity")) or 1.0
                row["Цена продажи, ₽"] = seller_price
                row["Цена покупателя, ₽"] = buyer_price
                row["Количество, шт."] = quantity
                if seller_price is not None and buyer_price is not None:
                    support = seller_price - buyer_price
                    row["Поддержка скидки Ozon, ₽"] = support
                    row["Поддержка скидки Ozon, %"] = safe_ratio(support, seller_price)
                row["Комиссия Ozon, ₽"] = to_number((matched or {}).get("commission_amount"))
                row["Комиссия Ozon, %"] = to_number((matched or {}).get("commission_percent"))
                row["Выплата продавцу, ₽"] = to_number((matched or {}).get("payout"))
                row["Финансовые услуги"] = safe_json((matched or {}).get("item_services", {}))
                rows.append(row)

        df = records_to_df(rows, {
            "Дата обновления снимка": self.target_date.isoformat(),
            "Период с": self.period_from.isoformat(),
            "Период по": self.period_to.isoformat(),
        })
        return df, raw_all, ["/v2/posting/fbo/list"]

    def fetch_returns(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        since = self.period_from
        payload = {
            "filter": {
                "logistic_return_date": {"time_from": to_ozon_datetime(since),
                                         "time_to": to_ozon_datetime(self.period_to, end=True)},
            },
            "limit": 500,
            "last_id": 0,
        }
        # Метод имеет разные форматы пагинации между версиями. Сначала один вызов,
        # затем при наличии last_id продолжаем вручную.
        rows: List[Dict[str, Any]] = []
        raw: List[Any] = []
        last_id: Any = 0
        for _ in range(500):
            body = dict(payload)
            body["last_id"] = last_id
            data = self.client.post("/v1/returns/list", body)
            raw.append(data)
            items = extract_items(data, [("returns",), ("result", "returns"), ("result", "items"), ("items",)])
            rows.extend(items)
            new_last = None
            for path in [("last_id",), ("result", "last_id")]:
                cur = data
                ok = True
                for key in path:
                    if not isinstance(cur, Mapping) or key not in cur:
                        ok = False
                        break
                    cur = cur[key]
                if ok:
                    new_last = cur
                    break
            if not items or new_last in (None, "", 0, last_id):
                break
            last_id = new_last
            time.sleep(0.12)
        return records_to_df(rows, {"Дата обновления снимка": self.target_date.isoformat()}), raw, ["/v1/returns/list"]

    # ------------------------- analytics funnel -------------------------

    def fetch_funnel(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        """Получает доступные метрики воронки за 1 день без шквала запросов.

        v17 делал отдельный запрос на каждую из 13 метрик, из-за чего один день
        мог порождать десятки 429 и занимать минуту. В v18 базовые метрики
        запрашиваются одним пакетом, расширенные — вторым. Позиционное
        сопоставление используется только когда число значений в каждой строке
        точно равно числу запрошенных метрик. Иначе расширенный ответ сохраняется
        только в raw JSON и не попадает в пользовательский Excel.
        """
        endpoint = "/v1/analytics/data"
        dimensions = ["day", "sku"]
        raw: List[Any] = []
        merged: Dict[Tuple[str, str], Dict[str, Any]] = {}
        verification: Dict[str, str] = {}
        rejected: Dict[str, str] = {}

        def declared_metric_names(value: Any) -> List[str]:
            found: List[str] = []
            def walk(obj: Any) -> None:
                if isinstance(obj, Mapping):
                    for key, child in obj.items():
                        key_l = str(key).lower()
                        if key_l in {"metric_names", "metrics_names", "columns", "headers", "fields"}:
                            if isinstance(child, list):
                                for item in child:
                                    candidate = None
                                    if isinstance(item, str):
                                        candidate = item
                                    elif isinstance(item, Mapping):
                                        for field in ("name", "key", "id", "metric"):
                                            if isinstance(item.get(field), str):
                                                candidate = item.get(field)
                                                break
                                    if candidate in FUNNEL_METRICS and candidate not in found:
                                        found.append(candidate)
                        walk(child)
                elif isinstance(obj, list):
                    for child in obj:
                        walk(child)
            walk(value)
            return found

        def parse_dimensions(item: Mapping[str, Any]) -> Tuple[str, str, str]:
            dims = item.get("dimensions") or item.get("dimension") or []
            day_value = self.target_date.isoformat()
            sku_value = ""
            sku_name = ""
            if isinstance(dims, list):
                if len(dims) > 0:
                    d = dims[0]
                    day_value = str((d.get("id") or d.get("name")) if isinstance(d, Mapping) else d)[:10]
                if len(dims) > 1:
                    s = dims[1]
                    if isinstance(s, Mapping):
                        sku_value = normalize_identifier(s.get("id") or s.get("value"))
                        sku_name = str(s.get("name") or "")
                    else:
                        sku_value = normalize_identifier(s)
            return day_value, sku_value, sku_name

        def request_group(metrics: List[str], label: str, strict: bool) -> None:
            payload = {
                "date_from": self.period_from.isoformat(),
                "date_to": self.period_to.isoformat(),
                "dimension": dimensions,
                "metrics": metrics,
                "filters": [],
                "sort": [{"key": metrics[0], "order": "DESC"}],
                "limit": 1000,
                "offset": 0,
            }
            try:
                items, pages = self.client.offset_pages(
                    endpoint, payload,
                    [("result", "data"), ("data",), ("result", "rows"), ("rows",)],
                    max_pages=20,
                )
            except OzonApiError as exc:
                raw.append({"group": label, "request": payload, "status": "ERROR", "error": str(exc)})
                for metric in metrics:
                    rejected[metric] = f"HTTP {exc.status}: пакет {label} недоступен"
                if strict:
                    raise
                return

            declared: List[str] = []
            for page in pages:
                for name in declared_metric_names(page):
                    if name not in declared:
                        declared.append(name)

            lengths = []
            for item in items:
                values = item.get("metrics")
                if isinstance(values, list):
                    lengths.append(len(values))
            exact_shape = bool(lengths) and all(length == len(metrics) for length in lengths)

            mapped_names: List[str] = []
            mapping_reason = ""
            if len(declared) == len(metrics) and set(declared).issubset(set(metrics)):
                mapped_names = declared
                mapping_reason = "explicit_response_metadata"
            elif exact_shape:
                mapped_names = list(metrics)
                mapping_reason = "exact_response_shape"
            elif len(metrics) == 1:
                mapped_names = list(metrics)
                mapping_reason = "single_metric_request"

            raw.append({
                "group": label,
                "request": payload,
                "declared_metrics": declared,
                "exact_shape": exact_shape,
                "mapped_metrics": mapped_names,
                "rows": len(items),
                "responses": pages,
            })

            if not mapped_names:
                for metric in metrics:
                    rejected[metric] = (
                        f"Ответ пакета {label} не сопоставлен: длины={sorted(set(lengths))}, "
                        f"ожидалось={len(metrics)}, metadata={declared}"
                    )
                return

            for item in items:
                day_value, sku_value, sku_name = parse_dimensions(item)
                values = item.get("metrics")
                if not isinstance(values, list):
                    values = [item.get(name) for name in mapped_names]
                if len(values) < len(mapped_names):
                    continue
                key = (day_value, sku_value)
                row = merged.setdefault(key, {"Дата": day_value, "sku": sku_value})
                if sku_name:
                    row["Название товара"] = sku_name
                for idx, metric in enumerate(mapped_names):
                    row[metric] = scalarize(values[idx])
                    verification[metric] = mapping_reason

        base_metrics = ["ordered_units", "revenue"]
        advanced_metrics = [m for m in FUNNEL_METRICS if m not in base_metrics]
        request_group(base_metrics, "base", strict=True)
        request_group(advanced_metrics, "advanced", strict=False)

        rows = list(merged.values())
        df = records_to_df(rows, {
            "Период с": self.period_from.isoformat(),
            "Период по": self.period_to.isoformat(),
            "Доступные метрики API": ", ".join(verification),
            "Проверка метрик": safe_json(verification),
            "Отклонённые метрики": safe_json(rejected),
        })

        aliases = {
            "hits_view_search": "Показы/просмотры в поиске",
            "hits_view_pdp": "Просмотры карточки",
            "hits_tocart": "Добавления в корзину",
            "ordered_units": "Заказы, шт.",
            "revenue": "Сумма заказов, ₽",
            "cancellations": "Отмены, шт.",
            "delivered_units": "Доставлено, шт.",
            "returns": "Возвраты, шт.",
            "session_view": "Сессии просмотра",
            "conv_tocart": "Конверсия в корзину Ozon, %",
            "conv_tocart_from_search": "Конверсия из поиска в корзину Ozon, %",
            "conv_order": "Конверсия в заказ Ozon, %",
            "position_category": "Позиция в категории",
        }
        for source, target_name in aliases.items():
            if source in df.columns:
                df[target_name] = df[source]

        def add_calc(target_name: str, numerator: str, denominator: str) -> None:
            if numerator in df.columns and denominator in df.columns:
                df[target_name] = [safe_ratio(n, d) for n, d in zip(df[numerator], df[denominator])]

        add_calc("Конверсия карточка → корзина, %", "hits_tocart", "hits_view_pdp")
        add_calc("Конверсия карточка → заказ, %", "ordered_units", "hits_view_pdp")
        add_calc("Конверсия корзина → заказ, %", "ordered_units", "hits_tocart")
        add_calc("Конверсия поиск → карточка, %", "hits_view_pdp", "hits_view_search")

        if "ordered_units" in df.columns and "delivered_units" in df.columns:
            df["Процент доставки, %"] = [
                safe_ratio(delivered, ordered)
                for delivered, ordered in zip(df["delivered_units"], df["ordered_units"])
            ]
        if "ordered_units" in df.columns and "returns" in df.columns:
            net: List[Optional[float]] = []
            pct: List[Optional[float]] = []
            for ordered, returned in zip(df["ordered_units"], df["returns"]):
                ordered_num = to_number(ordered)
                returned_num = to_number(returned) or 0.0
                value = None if ordered_num is None else ordered_num - returned_num
                net.append(value)
                pct.append(safe_ratio(value, ordered_num))
            df["Чисто выкуплено, шт."] = net
            df["Процент выкупа, %"] = pct

        return df, {"requests": raw, "confirmed": verification, "rejected": rejected}, [endpoint]

    def fetch_supply_orders(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        """Только незавершённые FBO-поставки: запланированные, готовые и в пути.

        Ozon менял enum статусов. Поэтому код пробует несколько совместимых
        наборов, затем одиночные статусы. Финально дополнительно отбрасывает
        отменённые и завершённые строки по фактическому статусу ответа.
        """
        endpoint_list = "/v3/supply-order/list"
        active_state_groups: List[List[Any]] = [
            [
                "DATA_FILLING",
                "READY_TO_SUPPLY",
                "ACCEPTED_AT_SUPPLY_WAREHOUSE",
                "IN_TRANSIT",
                "ACCEPTANCE_AT_STORAGE_WAREHOUSE",
            ],
            [
                "CREATED",
                "PROCESSING",
                "READY_TO_SUPPLY",
                "IN_TRANSIT",
                "ACCEPTED_PARTIALLY",
            ],
            [
                "DRAFT",
                "PLANNED",
                "READY_FOR_SHIPMENT",
                "SHIPPED",
                "IN_TRANSIT",
                "PARTIALLY_ACCEPTED",
            ],
        ]
        all_state_candidates = list(dict.fromkeys(
            state for group in active_state_groups for state in group
        ))

        created_from = (self.target_date - timedelta(days=180)).isoformat()
        created_to = (self.target_date + timedelta(days=30)).isoformat()

        attempts: List[Any] = []
        items: List[Dict[str, Any]] = []
        chosen_states: List[Any] = []
        chosen_payload: Optional[Dict[str, Any]] = None
        last_exc: Optional[Exception] = None

        def run_with_states(states: List[Any]) -> Tuple[List[Dict[str, Any]], List[Any], Dict[str, Any]]:
            base_filter = {
                "states": states,
                "created_date_from": created_from,
                "created_date_to": created_to,
            }
            # В protobuf-версии v3 строковые enum могут преобразовываться в 0,
            # который API запрещает. Поэтому сначала пробуем допустимые числовые enum.
            variants = [
                {"filter": base_filter, "limit": 100, "last_id": "", "sort_by": 1, "sort_direction": 2},
                {"filter": base_filter, "limit": 100, "last_id": "", "sort_by": 1, "sort_direction": 1},
                {"filter": base_filter, "limit": 100, "last_id": "", "sort_by": 2, "sort_direction": 2},
                {"filter": base_filter, "limit": 100, "last_id": "", "sort_by": "CREATION_DATE", "sort_direction": "DESC"},
                {"filter": base_filter, "limit": 100, "last_id": "", "sort_by": "CREATED_AT", "sort_direction": "DESC"},
            ]
            local_last: Optional[Exception] = None
            for payload in variants:
                try:
                    got, pages = self.client.cursor_pages(
                        endpoint_list,
                        payload,
                        [("result", "items"), ("items",)],
                        limit=100,
                        cursor_field="last_id",
                    )
                    return got, pages, payload
                except OzonApiError as exc:
                    local_last = exc
                    message = str(exc).lower()
                    if exc.status == 400 and (
                        "sort" in message
                        or "state" in message
                        or "validation" in message
                        or "enum" in message
                    ):
                        continue
                    raise
            if local_last:
                raise local_last
            raise RuntimeError("Не найден рабочий payload /v3/supply-order/list")

        for states in active_state_groups:
            try:
                got, pages, payload = run_with_states(states)
                attempts.append({
                    "states": states,
                    "status": "OK",
                    "request": payload,
                    "responses": pages,
                })
                items.extend(got)
                chosen_states = list(states)
                chosen_payload = payload
                break
            except Exception as exc:
                last_exc = exc
                attempts.append({
                    "states": states,
                    "status": "ERROR",
                    "error": str(exc),
                })

        # При несовпадении enum пробуем статусы по одному.
        if chosen_payload is None:
            successful_states: List[Any] = []
            successful_items: List[Dict[str, Any]] = []
            for state in all_state_candidates:
                try:
                    got, pages, payload = run_with_states([state])
                    attempts.append({
                        "states": [state],
                        "status": "OK",
                        "request": payload,
                        "responses": pages,
                    })
                    successful_states.append(state)
                    successful_items.extend(got)
                    chosen_payload = payload
                    time.sleep(API_PAUSE_SECONDS)
                except Exception as exc:
                    attempts.append({
                        "states": [state],
                        "status": "ERROR",
                        "error": str(exc),
                    })
            if successful_states:
                chosen_states = successful_states
                items = successful_items

        if chosen_payload is None:
            if last_exc:
                raise last_exc
            return pd.DataFrame(), {"attempts": attempts}, [endpoint_list]

        # Дедупликация заявок.
        deduped_items: List[Dict[str, Any]] = []
        seen_ids: set[str] = set()
        for item in items:
            order_id = (
                item.get("order_id")
                or item.get("supply_order_id")
                or item.get("id")
            )
            signature = str(order_id or safe_json(item))
            if signature in seen_ids:
                continue
            seen_ids.add(signature)
            deduped_items.append(item)

        def status_text(value: Mapping[str, Any]) -> str:
            candidates = [
                value.get("state"),
                value.get("status"),
                value.get("supply_order_state"),
                value.get("order_state"),
            ]
            return " ".join(str(x) for x in candidates if x not in (None, "")).upper()

        # Повторная защита от отменённых/завершённых поставок.
        inactive_markers = (
            "CANCEL", "CANCELED", "CANCELLED", "COMPLETE", "COMPLETED",
            "CLOSED", "FINISHED", "REJECTED",
        )
        active_items = [
            item for item in deduped_items
            if not any(marker in status_text(item) for marker in inactive_markers)
        ]

        rows: List[Dict[str, Any]] = []
        raw_details: List[Any] = []
        for index, item in enumerate(active_items, start=1):
            order_id = (
                item.get("order_id")
                or item.get("supply_order_id")
                or item.get("id")
            )
            base = dict(item)
            if not order_id:
                rows.append(base)
                continue
            try:
                detail = self.client.post("/v3/supply-order/get", {"order_id": order_id})
                raw_details.append({"order_id": order_id, "response": detail})
                root = detail.get("result", detail) if isinstance(detail, Mapping) else {}
                products = []
                if isinstance(root, Mapping):
                    products = (
                        root.get("items")
                        or root.get("products")
                        or root.get("supply_order_items")
                        or []
                    )
                if isinstance(products, list) and products:
                    for product in products:
                        row = dict(base)
                        if isinstance(product, Mapping):
                            row.update({f"product.{k}": v for k, v in product.items()})
                        row["detail"] = safe_json(root)
                        rows.append(row)
                else:
                    row = dict(base)
                    row["detail"] = safe_json(root)
                    rows.append(row)
            except Exception as exc:
                self.log_error("supply_detail", f"order_id={order_id}", exc, optional=True)
                base["detail_error"] = str(exc)
                rows.append(base)

            if index % 20 == 0:
                logging.info("Поставки FBO: обработано %s/%s", index, len(active_items))
            time.sleep(API_PAUSE_SECONDS)

        df = records_to_df(rows, {
            "Дата снимка": self.target_date.isoformat(),
            "Статусы запроса": ",".join(map(str, chosen_states)),
        })
        raw = {
            "selected_states": chosen_states,
            "chosen_list_payload": chosen_payload,
            "attempts": attempts,
            "details": raw_details,
        }
        return df, raw, [endpoint_list, "/v3/supply-order/get"]

    def fetch_warehouses(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        """Старый /v1/warehouse/list отключён. Справочник строим из отчёта остатков по складам."""
        source_result = next(
            (r for r in self.results if r.code == "stocks_warehouses" and not r.df.empty),
            None,
        )
        if source_result is None:
            return pd.DataFrame(), {"source": "stocks_warehouses", "rows": 0}, [
                "/v2/analytics/stock_on_warehouses"
            ]

        df = source_result.df.copy()
        candidates = [
            "warehouse_id", "warehouse_name", "cluster_name",
            "warehouse", "cluster", "region",
        ]
        cols = [c for c in candidates if c in df.columns]
        if not cols:
            return pd.DataFrame(), {"source": "stocks_warehouses", "rows": len(df)}, [
                "/v2/analytics/stock_on_warehouses"
            ]

        result = df[cols].drop_duplicates().reset_index(drop=True)
        result["Дата снимка"] = self.target_date.isoformat()
        return result, {
            "source": "stocks_warehouses",
            "rows": len(df),
        }, ["/v2/analytics/stock_on_warehouses"]


    # ------------------------- finance -------------------------
    def _product_article_map(self) -> Tuple[Dict[str, str], Dict[str, str]]:
        """Возвращает отображения SKU/product_id -> offer_id."""
        sku_to_offer: Dict[str, str] = dict(self.reference_sku_to_offer)
        product_to_offer: Dict[str, str] = {}
        for result in self.results:
            if result.code not in {"products", "product_info"} or result.df.empty:
                continue
            df = result.df
            offer_col = first_existing(df, ["offer_id", "offerId", "product.offer_id"])
            if not offer_col:
                continue
            sku_col = first_existing(df, ["sku", "fbo_sku", "product.sku"])
            product_col = first_existing(df, ["product_id", "id", "product.id"])
            for _, row in df.iterrows():
                offer = str(row.get(offer_col, "") or "").strip()
                if not offer or offer.lower() == "nan":
                    continue
                if sku_col:
                    value = str(row.get(sku_col, "") or "").strip()
                    if value and value.lower() != "nan":
                        sku_to_offer[value] = offer
                if product_col:
                    value = str(row.get(product_col, "") or "").strip()
                    if value and value.lower() != "nan":
                        product_to_offer[value] = offer
        return sku_to_offer, product_to_offer

    def fetch_finance_accruals(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        """Начисления за конкретный день.

        Метод новый и в разных кабинетах встречается с cursor/без cursor.
        Код пробует совместимые варианты и не расширяет запрос больше одного дня.
        """
        endpoint = "/v1/finance/accrual/by-day"
        day = self.target_date.isoformat()
        variants: List[Dict[str, Any]] = [
            {"date": day, "limit": 1000, "cursor": ""},
            {"date_from": day, "date_to": day, "limit": 1000, "cursor": ""},
            {"date": day},
        ]
        raw: List[Any] = []
        last_exc: Optional[Exception] = None

        for initial in variants:
            rows: List[Dict[str, Any]] = []
            cursor = str(initial.get("cursor", "") or "")
            try:
                for page in range(1, 1000):
                    body = dict(initial)
                    if "cursor" in body:
                        body["cursor"] = cursor
                    data = self.client.post(endpoint, body)
                    raw.append({"request": body, "response": data})
                    items = extract_items(
                        data,
                        [
                            ("accruals",),
                            ("result", "accruals"),
                            ("result", "items"),
                            ("items",),
                            ("rows",),
                        ],
                    )
                    rows.extend(items)
                    next_cursor = extract_cursor(data)
                    if not items or not next_cursor or next_cursor == cursor:
                        break
                    cursor = next_cursor
                    time.sleep(API_PAUSE_SECONDS)

                df = records_to_df(rows, {"Дата": day})
                return df, raw, [endpoint]
            except OzonApiError as exc:
                last_exc = exc
                raw.append({"request": initial, "status": "ERROR", "error": str(exc)})
                if exc.status in {400, 404, 422}:
                    continue
                raise

        if last_exc:
            raise last_exc
        return pd.DataFrame(), raw, [endpoint]

    def build_finance_by_sku(self, accrual_df: pd.DataFrame) -> Tuple[pd.DataFrame, Any, List[str]]:
        """Строит понятный управленческий финансовый отчёт по SKU.

        Источник — уже полученный /v1/finance/accrual/by-day, поэтому второй API-запрос
        не выполняется. Все расходы отображаются положительными числами; возврат/сторно
        расхода уменьшает итог. Поддержка Ozon (аналог СПП) контролируется двумя способами:
        разницей seller_price - sale_price и суммой bonus + coinvestment.
        """
        if accrual_df is None or accrual_df.empty:
            return pd.DataFrame(), {"source": "finance_accruals", "rows": 0}, ["/v1/finance/accrual/by-day"]

        aggregates: Dict[Tuple[str, str], Dict[str, Any]] = {}

        def get_bucket(day: str, sku: str) -> Dict[str, Any]:
            key = (day, sku)
            if key not in aggregates:
                offer = self.reference_sku_to_offer.get(sku, "") if sku else ""
                name = self.reference_sku_to_name.get(sku, "") if sku else ""
                aggregates[key] = {
                    "Дата": day,
                    "Артикул продавца": offer or ("Нераспределённые расходы магазина" if not sku else ""),
                    "SKU Ozon": sku,
                    "Название товара": name,
                    "Реализовано по финансовым данным, шт.": 0.0,
                    "Сумма по цене продавца, ₽": 0.0,
                    "Сумма реализации после поддержки Ozon, ₽": 0.0,
                    "Бонус Ozon, ₽": 0.0,
                    "Доплата Ozon по программам лояльности, ₽": 0.0,
                    "Поддержка Ozon по разнице цен, ₽": 0.0,
                    "Комиссия Ozon, ₽": 0.0,
                    "Логистика Ozon, ₽": 0.0,
                    "Другие расходы Ozon по товару, ₽": 0.0,
                    "Нераспределённые расходы Ozon, ₽": 0.0,
                    "Финансовых операций, шт.": 0,
                }
            return aggregates[key]

        def add_signed_expense(bucket: Dict[str, Any], column: str, signed_amount: Any) -> None:
            amount = money_amount(signed_amount)
            if amount is not None:
                # В API удержания обычно отрицательные. В управленческом отчёте расходы
                # показываем положительными, а возвраты/сторно — отрицательными.
                bucket[column] += -amount

        def parse_posting_product(product: Mapping[str, Any], bucket: Dict[str, Any]) -> None:
            delivery = product.get("delivery") if isinstance(product.get("delivery"), Mapping) else {}
            if delivery:
                total_delivery = money_amount(delivery.get("total_accrued"))
                if total_delivery is None:
                    total_delivery = 0.0
                    services = delivery.get("services") if isinstance(delivery.get("services"), list) else []
                    for service in services:
                        if isinstance(service, Mapping):
                            total_delivery += money_amount(service.get("accrued")) or 0.0
                add_signed_expense(bucket, "Логистика Ozon, ₽", total_delivery)

            commission = product.get("commission") if isinstance(product.get("commission"), Mapping) else {}
            if commission:
                seller_price = money_amount(commission.get("seller_price"))
                sale_price = money_amount(commission.get("sale_price"))
                sale_amount = money_amount(commission.get("sale_amount"))
                bonus = money_amount(commission.get("bonus")) or 0.0
                coinvestment = money_amount(commission.get("coinvestment")) or 0.0
                quantity = infer_finance_quantity(commission.get("seller_price"), commission.get("sale_amount"))

                if seller_price is not None:
                    bucket["Реализовано по финансовым данным, шт."] += quantity
                    bucket["Сумма по цене продавца, ₽"] += seller_price * quantity
                elif sale_amount is not None:
                    bucket["Реализовано по финансовым данным, шт."] += quantity
                    bucket["Сумма по цене продавца, ₽"] += sale_amount

                if sale_price is not None:
                    bucket["Сумма реализации после поддержки Ozon, ₽"] += sale_price * quantity
                bucket["Бонус Ozon, ₽"] += bonus * quantity
                bucket["Доплата Ozon по программам лояльности, ₽"] += coinvestment * quantity
                if seller_price is not None and sale_price is not None:
                    bucket["Поддержка Ozon по разнице цен, ₽"] += (seller_price - sale_price) * quantity

                commission_value = commission.get("commission")
                if commission_value is None:
                    commission_value = commission.get("sale_commission")
                add_signed_expense(bucket, "Комиссия Ozon, ₽", commission_value)

        for _, row in accrual_df.iterrows():
            day = str(row.get("Дата") or row.get("date") or self.target_date.isoformat())[:10]
            attributed = False

            posting = parse_json_value(row.get("posting"))
            if isinstance(posting, Mapping):
                products = posting.get("products") if isinstance(posting.get("products"), list) else []
                for product in products:
                    if not isinstance(product, Mapping):
                        continue
                    sku = normalize_identifier(product.get("sku"))
                    if not sku:
                        continue
                    bucket = get_bucket(day, sku)
                    bucket["Финансовых операций, шт."] += 1
                    parse_posting_product(product, bucket)
                    attributed = True

            item_fees = parse_json_value(row.get("item_fees"))
            if isinstance(item_fees, Mapping):
                fee_groups = item_fees.get("fees") if isinstance(item_fees.get("fees"), list) else []
                for group in fee_groups:
                    if not isinstance(group, Mapping):
                        continue
                    sku = normalize_identifier(group.get("sku"))
                    if not sku:
                        continue
                    bucket = get_bucket(day, sku)
                    bucket["Финансовых операций, шт."] += 1
                    fees = group.get("fees") if isinstance(group.get("fees"), list) else []
                    for fee in fees:
                        if isinstance(fee, Mapping):
                            add_signed_expense(bucket, "Другие расходы Ozon по товару, ₽", fee.get("accrued"))
                    attributed = True

            # Некоторые версии ответа раскрывают item_fees.fees отдельной колонкой.
            if not attributed:
                fees_flat = parse_json_value(row.get("item_fees.fees"))
                if isinstance(fees_flat, list):
                    for group in fees_flat:
                        if not isinstance(group, Mapping):
                            continue
                        sku = normalize_identifier(group.get("sku"))
                        if not sku:
                            continue
                        bucket = get_bucket(day, sku)
                        bucket["Финансовых операций, шт."] += 1
                        for fee in group.get("fees", []) if isinstance(group.get("fees"), list) else []:
                            if isinstance(fee, Mapping):
                                add_signed_expense(bucket, "Другие расходы Ozon по товару, ₽", fee.get("accrued"))
                        attributed = True

            # Нераспределённые начисления обязательно сохраняем отдельно, а не размазываем по SKU.
            non_item = parse_json_value(row.get("non_item_fee"))
            if isinstance(non_item, Mapping):
                bucket = get_bucket(day, "")
                bucket["Финансовых операций, шт."] += 1
                add_signed_expense(bucket, "Нераспределённые расходы Ozon, ₽", non_item.get("accrued"))
                attributed = True
            elif row.get("non_item_fee.accrued") not in (None, ""):
                bucket = get_bucket(day, "")
                bucket["Финансовых операций, шт."] += 1
                add_signed_expense(bucket, "Нераспределённые расходы Ozon, ₽", row.get("non_item_fee.accrued"))
                attributed = True

            if not attributed:
                total_amount = money_amount(row.get("total_amount"))
                if total_amount is None:
                    total_amount = money_amount(row.get("total_amount.amount"))
                if total_amount not in (None, 0):
                    bucket = get_bucket(day, "")
                    bucket["Финансовых операций, шт."] += 1
                    add_signed_expense(bucket, "Нераспределённые расходы Ozon, ₽", total_amount)

        output: List[Dict[str, Any]] = []
        for (_, _), row in sorted(aggregates.items(), key=lambda x: (x[0][0], x[0][1])):
            qty = float(row["Реализовано по финансовым данным, шт."] or 0.0)
            seller_sum = float(row["Сумма по цене продавца, ₽"] or 0.0)
            sale_sum = float(row["Сумма реализации после поддержки Ozon, ₽"] or 0.0)
            bonus = float(row["Бонус Ozon, ₽"] or 0.0)
            coinvestment = float(row["Доплата Ozon по программам лояльности, ₽"] or 0.0)
            support_by_diff = float(row["Поддержка Ozon по разнице цен, ₽"] or 0.0)
            support_components = bonus + coinvestment
            commission_cost = float(row["Комиссия Ozon, ₽"] or 0.0)
            logistics_cost = float(row["Логистика Ozon, ₽"] or 0.0)
            other_cost = float(row["Другие расходы Ozon по товару, ₽"] or 0.0)
            unallocated_cost = float(row["Нераспределённые расходы Ozon, ₽"] or 0.0)
            total_cost = commission_cost + logistics_cost + other_cost + unallocated_cost
            support_check = support_by_diff - support_components

            result_row = {
                "Дата": row["Дата"],
                "Артикул продавца": row["Артикул продавца"],
                "SKU Ozon": row["SKU Ozon"],
                "Название товара": row["Название товара"],
                "Реализовано по финансовым данным, шт.": int(round(qty)),
                "Средняя цена продавца до поддержки Ozon, ₽": round_money(seller_sum / qty) if qty else 0,
                "Средняя цена реализации после поддержки Ozon, ₽": round_money(sale_sum / qty) if qty else 0,
                "Сумма по цене продавца, ₽": round_money(seller_sum),
                "Сумма реализации после поддержки Ozon, ₽": round_money(sale_sum),
                "Поддержка Ozon (аналог СПП), ₽": round_money(support_by_diff),
                "Поддержка Ozon (аналог СПП), %": round_percent(safe_ratio(support_by_diff, seller_sum)),
                "Бонус Ozon, ₽": round_money(bonus),
                "Доплата Ozon по программам лояльности, ₽": round_money(coinvestment),
                "Проверка СПП: расхождение, ₽": round_money(support_check),
                "СПП сверено": "Да" if seller_sum > 0 and abs(support_check) <= max(1.0, seller_sum * 0.001) else ("Нет данных" if seller_sum <= 0 else "Нет"),
                "Комиссия Ozon, ₽": round_money(commission_cost),
                "Комиссия Ozon, % от цены продавца": round_percent(safe_ratio(commission_cost, seller_sum)),
                "Логистика Ozon, ₽": round_money(logistics_cost),
                "Другие расходы Ozon по товару, ₽": round_money(other_cost),
                "Нераспределённые расходы Ozon, ₽": round_money(unallocated_cost),
                "Всего расходов Ozon, ₽": round_money(total_cost),
                "Расходы Ozon, % от цены продавца": round_percent(safe_ratio(total_cost, seller_sum)),
                "Остаток после расходов Ozon, ₽": round_money(seller_sum - total_cost),
                "Финансовых операций, шт.": int(row["Финансовых операций, шт."] or 0),
            }
            output.append(result_row)

        df = pd.DataFrame(output)
        return df, {
            "source": "finance_accruals",
            "source_rows": int(len(accrual_df)),
            "result_rows": int(len(df)),
            "note": "Управленческий отчёт построен без дополнительного запроса API",
        }, ["/v1/finance/accrual/by-day"]

    def fetch_realization_report(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        """Официальный отчёт реализации за последний полностью закрытый месяц."""
        endpoint = "/v2/finance/realization"
        first_day = self.target_date.replace(day=1)
        month_day = first_day - timedelta(days=1)
        payload = {"month": month_day.month, "year": month_day.year}
        data = self.client.post(endpoint, payload)
        rows = extract_items(data, [("result", "rows"), ("rows",)])
        out: List[Dict[str, Any]] = []
        month_label = f"{month_day.year:04d}-{month_day.month:02d}"
        for source in rows:
            if not isinstance(source, Mapping):
                continue
            item = source.get("item") if isinstance(source.get("item"), Mapping) else {}
            delivery = source.get("delivery_commission") if isinstance(source.get("delivery_commission"), Mapping) else {}
            returned = source.get("return_commission") if isinstance(source.get("return_commission"), Mapping) else {}
            seller_price = to_number(source.get("seller_price_per_instance"))
            sale_price = to_number(delivery.get("price_per_instance"))
            support = None
            if seller_price is not None and sale_price is not None:
                support = seller_price - sale_price
            out.append({
                "Месяц отчёта": month_label,
                "Артикул продавца": item.get("offer_id") or "",
                "SKU Ozon": normalize_identifier(item.get("sku")),
                "Название товара": item.get("name") or "",
                "Штрихкод": item.get("barcode") or "",
                "Цена продавца, ₽": seller_price,
                "Цена реализации, ₽": sale_price,
                "Поддержка Ozon (аналог СПП), ₽ за шт.": support,
                "Поддержка Ozon (аналог СПП), %": safe_ratio(support, seller_price),
                "Реализовано, шт.": to_number(delivery.get("quantity")) or 0,
                "Сумма реализации, ₽": to_number(delivery.get("amount")) or 0,
                "Бонус Ozon по продажам, ₽": to_number(delivery.get("bonus")) or 0,
                "Доплата Ozon по лояльности, ₽": to_number(delivery.get("bank_coinvestment")) or 0,
                "Компенсация Ozon по продажам, ₽": to_number(delivery.get("compensation")) or 0,
                "Стандартная комиссия по продажам, ₽": to_number(delivery.get("standard_fee")) or 0,
                "Итого по продажам, ₽": to_number(delivery.get("total")) or 0,
                "Возвращено, шт.": to_number(returned.get("quantity")) or 0,
                "Сумма возвратов, ₽": to_number(returned.get("amount")) or 0,
                "Бонус Ozon по возвратам, ₽": to_number(returned.get("bonus")) or 0,
                "Доплата Ozon по лояльности при возврате, ₽": to_number(returned.get("bank_coinvestment")) or 0,
                "Компенсация Ozon по возвратам, ₽": to_number(returned.get("compensation")) or 0,
                "Итого по возвратам, ₽": to_number(returned.get("total")) or 0,
                "Комиссия, %": round_percent((to_number(source.get("commission_ratio")) or 0) * 100 if (to_number(source.get("commission_ratio")) or 0) <= 1.5 else to_number(source.get("commission_ratio"))),
            })
        return pd.DataFrame(out), {"request": payload, "response": data}, [endpoint]

    def collect_realization_report_if_needed(self) -> Optional[ReportResult]:
        """Обновляет месячный отчёт реализации только при смене закрытого месяца.

        В дневной archive этот отчёт не включается: он относится ко всему месяцу,
        а не к выбранному дню. В обычных запусках хранится один актуальный файл.
        """
        if self.mode == "archive":
            logging.info("Отчёт реализации: пропущен в дневном archive (месячный отчёт)")
            return None

        first_day = self.target_date.replace(day=1)
        month_day = first_day - timedelta(days=1)
        expected_month = f"{month_day.year:04d}-{month_day.month:02d}"
        key = (
            f"{self.base_prefix}/Отчёт реализации/{self.store}/"
            "Отчёт_реализации_последний_закрытый_месяц.xlsx"
        )
        if self.storage.exists(key):
            try:
                old = self.storage.read_excel(key)
                if not old.empty and "Месяц отчёта" in old.columns:
                    months = set(old["Месяц отчёта"].dropna().astype(str).str[:7])
                    if expected_month in months:
                        logging.info(
                            "Отчёт реализации: %s уже сохранён, повторный API-запрос не нужен",
                            expected_month,
                        )
                        return None
            except Exception as exc:
                logging.warning("Не удалось проверить кэш отчёта реализации: %s", exc)

        return self._run(
            "realization_report", "Отчёт реализации Ozon — последний закрытый месяц",
            "Отчёт реализации", "Отчёт_реализации_последний_закрытый_месяц",
            self.fetch_realization_report, "Месяц отчёта",
            ["Месяц отчёта", "SKU Ozon"],
            snapshot=True, optional=True,
        )

    # ------------------------- Performance API -------------------------

    def _require_performance(self) -> OzonPerformanceClient:
        if self.performance_client is None:
            raise RuntimeError(
                "Не заданы OZON_PERFORMANCE_CLIENT_ID и OZON_PERFORMANCE_CLIENT_SECRET"
            )
        return self.performance_client

    @staticmethod
    def _extract_performance_items(data: Any) -> List[Dict[str, Any]]:
        if isinstance(data, list):
            return [
                dict(item) if isinstance(item, Mapping) else {"value": item}
                for item in data
            ]
        if not isinstance(data, Mapping):
            return []

        paths = [
            ("list",),
            ("items",),
            ("campaigns",),
            ("products",),
            ("rows",),
            ("result",),
            ("result", "items"),
            ("result", "campaigns"),
            ("result", "products"),
            ("data",),
            ("data", "items"),
            ("data", "campaigns"),
            ("data", "products"),
        ]
        items = extract_items(data, paths)
        if items:
            return items

        # Иногда один объект возвращается без массива.
        if any(
            key in data
            for key in ("id", "campaignId", "campaign_id", "sku", "productId")
        ):
            return [dict(data)]
        return []

    @staticmethod
    def _find_object_lists(value: Any) -> List[Dict[str, Any]]:
        """Ищет списки рекламных объектов в произвольной структуре ответа."""
        found: List[Dict[str, Any]] = []

        def walk(node: Any, path: str = "") -> None:
            if isinstance(node, list):
                for item in node:
                    if isinstance(item, Mapping):
                        keys = {str(k).lower() for k in item.keys()}
                        if keys.intersection({
                            "sku", "productid", "product_id", "offerid", "offer_id",
                            "objectid", "object_id", "bid", "price", "status"
                        }):
                            row = dict(item)
                            row.setdefault("_source_path", path)
                            found.append(row)
                        walk(item, path)
            elif isinstance(node, Mapping):
                for key, child in node.items():
                    child_path = f"{path}.{key}" if path else str(key)
                    walk(child, child_path)

        walk(value)
        return found

    def fetch_ad_campaigns(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        client = self._require_performance()
        data, meta = client.try_variants([
            ("GET", "/api/client/campaign", None, None),
            ("GET", "/api/client/campaigns", None, None),
            ("POST", "/api/client/campaign", {}, None),
        ])
        items = self._extract_performance_items(data)
        df = records_to_df(items, {
            "Дата снимка": self.target_date.isoformat(),
        })
        raw = {"meta": meta, "response": data}
        return df, raw, [meta["chosen_path"]]

    def _campaign_ids_from_results(self) -> List[str]:
        if self.campaign_ids_override:
            return sorted(set(str(x) for x in self.campaign_ids_override if str(x).strip()))
        result = next(
            (r for r in self.results if r.code == "ad_campaigns" and not r.df.empty),
            None,
        )
        if result is None:
            return []

        candidates = [
            "id", "campaignId", "campaign_id", "campaign.id", "advCampaignId"
        ]
        ids: List[str] = []
        for column in candidates:
            if column not in result.df.columns:
                continue
            for value in result.df[column].dropna().tolist():
                text_value = str(value).strip()
                if text_value and text_value.lower() != "nan":
                    ids.append(text_value)
        return sorted(set(ids))

    def _active_campaign_ids_from_results(self) -> List[str]:
        if self.active_campaign_ids_override:
            return sorted(set(str(x) for x in self.active_campaign_ids_override if str(x).strip()))
        result = next(
            (r for r in self.results if r.code == "ad_campaigns" and not r.df.empty),
            None,
        )
        if result is None:
            return self._campaign_ids_from_results()
        id_col = next((c for c in ("id", "campaignId", "campaign_id") if c in result.df.columns), None)
        state_col = next((c for c in ("state", "status", "campaignState") if c in result.df.columns), None)
        if not id_col:
            return self._campaign_ids_from_results()
        work = result.df.copy()
        if state_col:
            state = work[state_col].astype(str).str.upper()
            mask = state.str.contains("RUNNING|ACTIVE", regex=True, na=False) & ~state.str.contains("INACTIVE", regex=True, na=False)
            work = work[mask]
        ids = [normalize_identifier(v) for v in work[id_col].dropna().tolist()]
        return sorted(set(v for v in ids if v)) or self._campaign_ids_from_results()

    def fetch_ad_products_bids(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        """Получает товары и ставки по кампаниям.

        В v13 ошибка одной кампании больше не прерывает весь отчёт. Сначала
        читаем карточку кампании, затем пробуем актуальные object-endpoint'ы.
        """
        client = self._require_performance()
        campaign_ids = self._campaign_ids_from_results()
        if not campaign_ids:
            return pd.DataFrame(), {"warning": "campaign_ids is empty"}, []

        rows: List[Dict[str, Any]] = []
        diagnostics: List[Dict[str, Any]] = []
        raw: List[Any] = []
        used_paths: List[str] = []

        for index, campaign_id in enumerate(campaign_ids, start=1):
            campaign_rows: List[Dict[str, Any]] = []
            campaign_raw: Dict[str, Any] = {"campaign_id": campaign_id}

            # 1. Карточка кампании часто уже содержит objects/products.
            try:
                detail, detail_meta = client.try_variants([
                    ("GET", f"/api/client/campaign/{campaign_id}", None, None),
                ])
                used_paths.append(detail_meta["chosen_path"])
                campaign_raw["campaign_detail"] = detail
                campaign_raw["campaign_detail_meta"] = detail_meta
                campaign_rows.extend(self._find_object_lists(detail))
            except Exception as exc:
                campaign_raw["campaign_detail_error"] = str(exc)

            # 2. Отдельные методы объектов.
            object_data = None
            object_meta = None
            try:
                object_data, object_meta = client.try_variants([
                    ("GET", f"/api/client/campaign/{campaign_id}/objects", None, None),
                    ("GET", f"/api/client/campaign/{campaign_id}/object", None, None),
                    ("GET", f"/api/client/campaign/{campaign_id}/objects/info", None, None),
                    ("GET", f"/api/client/campaign/{campaign_id}/products", None, None),
                    ("GET", f"/api/client/campaign/{campaign_id}/v2/products", None, None),
                    ("GET", f"/campaign/{campaign_id}/v2/products", None, None),
                ])
                used_paths.append(object_meta["chosen_path"])
                campaign_raw["objects_response"] = object_data
                campaign_raw["objects_meta"] = object_meta
                extracted = self._extract_performance_items(object_data)
                if not extracted:
                    extracted = self._find_object_lists(object_data)
                campaign_rows.extend(extracted)
            except Exception as exc:
                campaign_raw["objects_error"] = str(exc)

            # 3. Конкурентные ставки — необязательны.
            competitive_data: Any = {}
            competitive_items: List[Dict[str, Any]] = []
            try:
                competitive_data, competitive_meta = client.try_variants([
                    ("GET", f"/api/client/campaign/{campaign_id}/objects/bids/competitive", None, None),
                    ("GET", f"/api/client/campaign/{campaign_id}/bids/competitive", None, None),
                    ("GET", f"/api/client/campaign/{campaign_id}/competitive-bids", None, None),
                ])
                used_paths.append(competitive_meta["chosen_path"])
                campaign_raw["competitive_response"] = competitive_data
                campaign_raw["competitive_meta"] = competitive_meta
                competitive_items = self._extract_performance_items(competitive_data)
                if not competitive_items:
                    competitive_items = self._find_object_lists(competitive_data)
            except Exception as exc:
                campaign_raw["competitive_error"] = str(exc)

            # Дедупликация объектов внутри кампании.
            seen: set[str] = set()
            unique_rows: List[Dict[str, Any]] = []
            for item in campaign_rows:
                key_value = (
                    item.get("sku")
                    or item.get("productId")
                    or item.get("product_id")
                    or item.get("offerId")
                    or item.get("offer_id")
                    or item.get("objectId")
                    or item.get("object_id")
                )
                signature = f"{key_value}|{safe_json(item)}"
                if signature in seen:
                    continue
                seen.add(signature)
                unique_rows.append(item)

            competitive_by_key: Dict[str, Dict[str, Any]] = {}
            for item in competitive_items:
                key_value = (
                    item.get("sku")
                    or item.get("productId")
                    or item.get("product_id")
                    or item.get("offerId")
                    or item.get("offer_id")
                    or item.get("objectId")
                    or item.get("object_id")
                )
                if key_value is not None:
                    competitive_by_key[str(key_value)] = item

            for item in unique_rows:
                row: Dict[str, Any] = {
                    "Дата снимка": self.target_date.isoformat(),
                    "campaign_id": campaign_id,
                }
                row.update(item)
                key_value = (
                    item.get("sku")
                    or item.get("productId")
                    or item.get("product_id")
                    or item.get("offerId")
                    or item.get("offer_id")
                    or item.get("objectId")
                    or item.get("object_id")
                )
                competitive = competitive_by_key.get(str(key_value), {})
                for field, value in competitive.items():
                    if field not in row:
                        row[f"competitive.{field}"] = value
                rows.append(row)

            if not unique_rows:
                diagnostics.append({
                    "Дата снимка": self.target_date.isoformat(),
                    "campaign_id": campaign_id,
                    "status": "NO_OBJECTS",
                    "campaign_detail_error": campaign_raw.get("campaign_detail_error", ""),
                    "objects_error": campaign_raw.get("objects_error", ""),
                    "competitive_error": campaign_raw.get("competitive_error", ""),
                })

            raw.append(campaign_raw)

            if index % 20 == 0:
                logging.info(
                    "Performance API: товары и ставки %s/%s кампаний, строк %s",
                    index,
                    len(campaign_ids),
                    len(rows),
                )
            time.sleep(0.10)

        if rows:
            df = records_to_df(rows, {"Дата снимка": self.target_date.isoformat()})
        else:
            # Не скрываем результат: сохраняем диагностический Excel.
            df = records_to_df(diagnostics, {
                "Дата снимка": self.target_date.isoformat(),
            })

        return df, {"campaigns": raw, "diagnostics": diagnostics}, sorted(set(used_paths))

    def _poll_statistics_report(
        self,
        client: OzonPerformanceClient,
        report_id: str,
        max_attempts: int = 60,
    ) -> Tuple[Any, List[Any], str]:
        raw: List[Any] = []
        variants = [
            f"/api/client/statistics/{report_id}",
            f"/api/client/statistics/report/{report_id}",
            f"/api/client/statistics/{report_id}/data",
        ]
        last_path = variants[0]

        for attempt in range(1, max_attempts + 1):
            for path in variants:
                try:
                    data = client.get(path)
                    last_path = path
                    raw.append({
                        "attempt": attempt,
                        "path": path,
                        "response": data,
                    })

                    status = ""
                    if isinstance(data, Mapping):
                        status = str(
                            data.get("status")
                            or data.get("state")
                            or data.get("reportStatus")
                            or ""
                        ).upper()

                    if status in {"ERROR", "FAILED", "CANCELLED"}:
                        raise RuntimeError(
                            f"Performance report {report_id}: status={status}"
                        )

                    items = self._extract_performance_items(data)
                    if items:
                        return data, raw, path

                    if status in {
                        "OK", "DONE", "READY", "COMPLETED", "SUCCESS"
                    }:
                        return data, raw, path

                except OzonApiError as exc:
                    if exc.status in {404, 405}:
                        continue
                    raise

            time.sleep(min(15, 2 + attempt // 5))

        raise RuntimeError(
            f"Performance report {report_id} не готов после {max_attempts} проверок"
        )


    def fetch_ad_statistics(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        """Дневная статистика рекламы с обязательным campaign_id.

        Диагностические ошибки сохраняются только в raw JSON. В пользовательский
        Excel попадают исключительно строки кампаний, поэтому ERROR-строки больше
        не могут смешиваться со статистикой и ломать дедупликацию.
        """
        client = self._require_performance()
        campaign_ids = self._campaign_ids_from_results()
        date_from = self.period_from.isoformat()
        date_to = self.period_to.isoformat()
        raw: List[Any] = []
        used_paths: List[str] = []

        def normalize_items(items: Sequence[Mapping[str, Any]], extra: Optional[Mapping[str, Any]] = None) -> List[Dict[str, Any]]:
            rows: List[Dict[str, Any]] = []
            for item in items:
                row = dict(item)
                if extra:
                    for key, value in extra.items():
                        row.setdefault(key, value)
                campaign_id = (
                    row.get("campaign_id") or row.get("campaignId") or
                    row.get("campaign") or row.get("id")
                )
                day_value = row.get("Дата") or row.get("date") or row.get("day") or row.get("statDate")
                campaign_id = normalize_identifier(campaign_id)
                day_value = str(day_value or "")[:10]
                if not campaign_id or not day_value:
                    continue
                row["campaign_id"] = campaign_id
                row["Дата"] = day_value
                rows.append(row)
            return rows

        direct_variants = [
            ("GET", "/api/client/statistics/daily/json", None, {"dateFrom": date_from, "dateTo": date_to}),
            ("GET", "/api/client/statistics/daily", None, {"dateFrom": date_from, "dateTo": date_to}),
        ]
        try:
            data, meta = client.try_variants(direct_variants)
            items = self._extract_performance_items(data)
            if not items and isinstance(data, Mapping):
                items = self._find_object_lists(data)
            rows = normalize_items(items)
            raw.append({"mode": "daily_direct", "meta": meta, "response": data, "normalized_rows": len(rows)})
            used_paths.append(meta["chosen_path"])
            if rows:
                return records_to_df(rows, {
                    "Дата запроса": datetime.now(MOSCOW_TZ).isoformat(),
                    "Период с": date_from,
                    "Период по": date_to,
                }), raw, sorted(set(used_paths))
        except Exception as exc:
            raw.append({"mode": "daily_direct", "status": "ERROR", "error": str(exc)})
            logging.warning("Performance daily/json недоступен: %s", exc)

        if not campaign_ids:
            return pd.DataFrame(), {"warning": "campaign_ids is empty", "attempts": raw}, used_paths

        all_rows: List[Dict[str, Any]] = []
        raw_batches: List[Any] = []
        for batch_no, campaign_batch in enumerate(batched(campaign_ids, 10), start=1):
            int_ids: List[Any] = []
            for value in campaign_batch:
                try:
                    int_ids.append(int(value))
                except Exception:
                    int_ids.append(value)
            payload_variants = [
                {"campaigns": int_ids, "dateFrom": date_from, "dateTo": date_to, "groupBy": "DATE"},
                {"campaigns": [str(x) for x in campaign_batch], "dateFrom": date_from, "dateTo": date_to, "groupBy": "DATE"},
            ]
            batch_errors: List[Any] = []
            success = False
            for payload in payload_variants:
                try:
                    data, meta = client.try_variants([
                        ("POST", "/api/client/statistics/json", payload, None),
                        ("POST", "/api/client/statistics", payload, None),
                    ])
                    used_paths.append(meta["chosen_path"])
                    items = self._extract_performance_items(data)
                    if items:
                        rows = normalize_items(items, {"statistics_batch_no": batch_no})
                        all_rows.extend(rows)
                        raw_batches.append({"batch_no": batch_no, "request": payload, "meta": meta, "response": data, "mode": "direct", "normalized_rows": len(rows)})
                        success = True
                        break
                    report_id = None
                    if isinstance(data, Mapping):
                        report_id = data.get("UUID") or data.get("uuid") or data.get("reportId") or data.get("report_id") or data.get("id")
                    if not report_id:
                        batch_errors.append({"request": payload, "error": "Ответ без строк и без report_id", "response": data})
                        continue
                    final_data, poll_raw, poll_path = self._poll_statistics_report(client, str(report_id))
                    used_paths.append(poll_path)
                    rows = normalize_items(self._extract_performance_items(final_data), {
                        "statistics_batch_no": batch_no,
                        "report_id": str(report_id),
                    })
                    all_rows.extend(rows)
                    raw_batches.append({
                        "batch_no": batch_no, "request": payload, "meta": meta,
                        "create_response": data, "report_id": report_id,
                        "poll": poll_raw, "final_response": final_data,
                        "mode": "async", "normalized_rows": len(rows),
                    })
                    success = True
                    break
                except Exception as exc:
                    batch_errors.append({"request": payload, "error": str(exc)})
            if not success:
                raw_batches.append({"batch_no": batch_no, "status": "ERROR", "errors": batch_errors})
            logging.info("Performance API: пакет %s, успех=%s, строк всего=%s", batch_no, success, len(all_rows))
            time.sleep(API_PAUSE_SECONDS)

        raw.extend(raw_batches)
        if not all_rows:
            return pd.DataFrame(), {"attempts": raw, "message": "Строк кампаний не получено"}, sorted(set(used_paths))
        df = records_to_df(all_rows, {
            "Дата запроса": datetime.now(MOSCOW_TZ).isoformat(),
            "Период с": date_from,
            "Период по": date_to,
        })
        return df, {"attempts": raw}, sorted(set(used_paths))

    def fetch_ad_product_statistics(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        """Статистика рекламы на уровне кампания × товар за один день."""
        client = self._require_performance()
        campaign_ids = self._active_campaign_ids_from_results()
        if not campaign_ids:
            return pd.DataFrame(), {"warning": "active campaign_ids is empty"}, []

        rows: List[Dict[str, Any]] = []
        raw: List[Any] = []
        used_paths: List[str] = []
        date_from = self.period_from.isoformat()
        date_to = self.period_to.isoformat()

        logging.info("Performance product statistics: проверяем только активные кампании: %s", len(campaign_ids))
        empty_streak = 0
        for index, campaign_id in enumerate(campaign_ids, start=1):
            params = {
                "campaignId": campaign_id,
                "dateFrom": date_from,
                "dateTo": date_to,
            }
            try:
                data, meta = client.try_variants([
                    ("GET", "/api/client/statistics/campaign/product", None, params),
                    ("GET", "/api/client/statistics/campaign/product/json", None, params),
                ])
                used_paths.append(meta["chosen_path"])
                items = self._extract_performance_items(data)
                if not items:
                    items = self._find_object_lists(data)
                for item in items:
                    row = dict(item)
                    row.setdefault("campaign_id", campaign_id)
                    row.setdefault("Дата", self.target_date.isoformat())
                    rows.append(row)
                raw.append({
                    "campaign_id": campaign_id,
                    "meta": meta,
                    "response": data,
                })
                if items:
                    empty_streak = 0
                else:
                    empty_streak += 1
            except Exception as exc:
                raw.append({
                    "campaign_id": campaign_id,
                    "status": "ERROR",
                    "error": str(exc),
                })

            if index % 20 == 0:
                logging.info(
                    "Performance product statistics: %s/%s кампаний, строк %s",
                    index, len(campaign_ids), len(rows)
                )
            time.sleep(API_PAUSE_SECONDS)

        if not rows and raw:
            errors = [x for x in raw if x.get("status") == "ERROR"]
            if len(errors) == len(raw):
                raise RuntimeError(
                    "Не удалось получить статистику товаров ни по одной кампании: "
                    + safe_json(errors[:3])
                )

        return records_to_df(rows, {
            "Дата запроса": datetime.now(MOSCOW_TZ).isoformat(),
            "Период с": date_from,
            "Период по": date_to,
        }), raw, sorted(set(used_paths))

    def build_ad_orders_report(self) -> Tuple[pd.DataFrame, Any, List[str]]:
        """Производный отчёт рекламных заказов из товарной статистики."""
        source = next(
            (r for r in self.results if r.code == "ad_product_statistics" and not r.df.empty),
            None,
        )
        if source is None:
            return pd.DataFrame(), {"source": "ad_product_statistics", "rows": 0}, []

        df = source.df.copy()
        candidate_keywords = (
            "order", "sale", "revenue", "gmv", "sold", "purchase",
            "заказ", "продаж", "выруч",
        )
        identity = [
            c for c in df.columns
            if str(c).lower() in {
                "дата", "date", "day", "campaign_id", "campaignid",
                "sku", "productid", "product_id", "offerid", "offer_id"
            }
        ]
        metrics = [
            c for c in df.columns
            if any(k in str(c).lower() for k in candidate_keywords)
        ]
        selected = list(dict.fromkeys(identity + metrics))
        if not selected:
            selected = list(df.columns)
        out = df[selected].copy()
        return out, {
            "source": "ad_product_statistics",
            "source_rows": len(df),
            "selected_columns": selected,
        }, list(source.method_paths)

    def save_ad_lag_service(self) -> Dict[str, Any]:
        """Сохраняет неизменяемые снимки 14-дневного рекламного лага."""
        if self.performance_client is None:
            return {"status": "SKIPPED", "reason": "Performance credentials missing"}

        original_from, original_to = self.period_from, self.period_to
        request_at = datetime.now(MOSCOW_TZ)
        lag_from = self.target_date - timedelta(days=AD_LAG_DAYS - 1)
        lag_to = self.target_date
        self.period_from = lag_from
        self.period_to = lag_to
        try:
            df, raw, methods = self.fetch_ad_statistics()
        finally:
            self.period_from, self.period_to = original_from, original_to

        if df.empty:
            return {"status": "EMPTY", "rows": 0, "methods": methods}

        work = df.copy()
        date_col = self._detect_row_date_column(work, "Дата")
        if date_col:
            parsed = pd.to_datetime(work[date_col], errors="coerce", utc=True)
            work["Дата статистики"] = parsed.dt.date.astype("string")
            lag = (request_at.date() - parsed.dt.date).apply(
                lambda x: x.days if pd.notna(x) else pd.NA
            )
            work["Лаг, дней"] = lag
        else:
            work["Дата статистики"] = ""
            work["Лаг, дней"] = pd.NA

        work["Дата и время запроса МСК"] = request_at.isoformat()
        work["Фаза запроса"] = resolve_request_phase(request_at)
        filename = week_filename("Рекламный_лаг", request_at.date())
        key = (
            f"{self.base_prefix}/Служебные/{self.store}/Рекламный лаг/"
            f"Недельные/{filename}"
        )
        old = self.storage.read_excel(key)
        merged = pd.concat([old, work], ignore_index=True, sort=False)
        sort_candidates = [
            c for c in (
                "Дата статистики",
                "campaign_id",
                "campaignId",
                "sku",
                "productId",
                "Дата и время запроса МСК",
            )
            if c in merged.columns
        ]
        if sort_candidates:
            merged = merged.sort_values(sort_candidates, kind="stable").reset_index(drop=True)
        self.storage.upload_bytes(
            key,
            self._excel_bytes(merged, "Рекламный лаг"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        raw_key = ""
        save_raw = str(os.getenv("OZON_SAVE_RAW_RESPONSES", "0")).strip().lower() in {
            "1", "true", "yes", "on"
        }
        if save_raw:
            raw_key = (
                f"{self.base_prefix}/Служебные/{self.store}/Рекламный лаг/"
                f"Сырые ответы/{request_at:%Y-%m-%d}/{self.run_id}.json"
            )
            self.storage.upload_json(raw_key, raw)
        return {
            "status": "OK",
            "rows_added": len(work),
            "total_rows": len(merged),
            "s3_key": key,
            "raw_key": raw_key,
            "period_from": lag_from.isoformat(),
            "period_to": lag_to.isoformat(),
            "methods": methods,
        }

    # ------------------------- collect/save -------------------------
    def collect_snapshot_reports(self) -> None:
        """Снимки состояния: динамические — недельно, статические — только latest."""
        self._run(
            "products", "Справочник товаров", "Товары", "Товары",
            self.fetch_product_catalog, "Дата снимка",
            ["Дата снимка", "product_id", "offer_id"],
            snapshot=True,
        )
        self._run(
            "product_info", "Подробная информация о товарах",
            "Информация о товарах", "Информация_о_товарах",
            self.fetch_product_info, "Дата снимка",
            ["Дата снимка", "id", "product_id", "offer_id"],
            snapshot=True, optional=True,
        )
        self._run(
            "prices", "Цены", "Цены", "Цены",
            self.fetch_prices, "Дата снимка",
            ["Дата снимка", "product_id", "offer_id"],
        )
        self._run(
            "stocks", "Остатки FBO", "Остатки", "Остатки",
            self.fetch_stocks, "Дата снимка",
            ["Дата снимка", "product_id", "offer_id", "stock.warehouse_id", "stock.type"],
        )
        self._run(
            "stocks_warehouses", "Остатки по складам",
            "Остатки по складам", "Остатки_по_складам",
            self.fetch_stock_on_warehouses, "Дата снимка",
            ["Дата снимка", "sku", "warehouse_name", "warehouse_id"],
            optional=True,
        )
        self._run(
            "turnover", "Оборачиваемость Ozon",
            "Оборачиваемость Ozon", "Оборачиваемость_Ozon",
            self.fetch_turnover, "Дата снимка",
            ["Дата снимка", "sku", "offer_id", "product_id"],
            optional=True,
        )
        self._run(
            "analytics_stocks", "Аналитика остатков Ozon",
            "Аналитика остатков", "Аналитика_остатков",
            self.fetch_analytics_stocks, "Дата снимка",
            ["Дата снимка", "sku", "offer_id", "product_id", "warehouse_id"],
            optional=True,
        )

        if self.performance_client is not None:
            self._run(
                "ad_campaigns", "Реклама — кампании",
                "Реклама/Кампании", "Рекламные_кампании",
                self.fetch_ad_campaigns, "Дата снимка",
                ["Дата снимка", "id", "campaignId", "campaign_id"],
                optional=True,
            )
            self._run(
                "ad_products_bids", "Реклама — товары и ставки",
                "Реклама/Товары и ставки", "Реклама_товары_и_ставки",
                self.fetch_ad_products_bids, "Дата снимка",
                ["Дата снимка", "campaign_id", "sku", "productId", "offerId", "id"],
                optional=True,
            )
        else:
            logging.warning("Performance API отключён: рекламные снимки пропущены")

        self._run(
            "supplies", "Поставки FBO — запланированные и в пути",
            "Поставки", "Поставки",
            self.fetch_supply_orders, "Дата снимка",
            ["Дата снимка", "order_id", "supply_order_id", "product.sku", "product.offer_id"],
            optional=True,
        )
        self._run(
            "warehouses", "Справочник складов",
            "Склады", "Склады",
            self.fetch_warehouses, "Дата снимка",
            ["Дата снимка", "warehouse_id", "name"],
            snapshot=True, optional=True,
        )
        self.collect_realization_report_if_needed()

    def collect_optional_event_reports(self) -> None:
        """Дополнительная товарная рекламная аналитика."""
        if self.performance_client is not None:
            self._run(
                "ad_product_statistics", "Реклама — статистика по товарам",
                "Реклама/По товарам", "Реклама_по_товарам",
                self.fetch_ad_product_statistics, "Дата",
                ["Дата", "campaign_id", "sku", "productId", "product_id"],
                optional=True,
            )
            self._run(
                "ad_orders", "Реклама — заказы",
                "Реклама/Заказы", "Рекламные_заказы",
                self.build_ad_orders_report, "Дата",
                ["Дата", "campaign_id", "sku", "productId", "product_id"],
                optional=True,
            )


    def collect_event_reports(self, include_heavy: bool = True) -> None:
        """Событийные отчёты строго за self.target_date."""
        self.period_from = self.target_date
        self.period_to = self.target_date

        self._run(
            "orders", "Заказы FBO", "Заказы", "Заказы",
            self.fetch_fbo_postings, "created_at",
            ["posting_number", "product.sku", "product.offer_id", "product.name"],
        )
        self._run(
            "returns", "Возвраты", "Возвраты", "Возвраты",
            self.fetch_returns, "logistic.return_date",
            ["id"],
            optional=True,
        )
        self._run(
            "funnel", "Воронка продаж", "Воронка продаж", "Воронка_продаж",
            self.fetch_funnel, "Дата",
            ["Дата", "sku"],
        )
        finance_result = self._run(
            "finance_accruals", "Финансовые начисления — технический источник",
            "Финансовые начисления", "Финансовые_начисления",
            self.fetch_finance_accruals, "Дата",
            ["Дата", "accrual_id", "posting.posting_number", "posting.products.sku", "sku"],
            optional=True,
        )
        if finance_result.status in {"OK", "EMPTY"}:
            self._run(
                "finance_by_sku", "Финансы по артикулам — управленческий отчёт",
                "Финансы по артикулам", "Финансы_по_артикулам",
                lambda: self.build_finance_by_sku(finance_result.df), "Дата",
                ["Дата", "SKU Ozon"],
                optional=True,
            )

        if self.performance_client is not None:
            self._run(
                "ad_statistics", "Реклама — дневная статистика",
                "Реклама/Статистика", "Реклама_статистика",
                self.fetch_ad_statistics, "Дата",
                ["Дата", "campaign_id"],
                optional=True,
            )

        if include_heavy:
            self.collect_optional_event_reports()

    def collect_all(self) -> None:
        """Совместимость для test/archive: снимки + событийные отчёты."""
        self.collect_snapshot_reports()
        self.collect_event_reports()

    @staticmethod
    def _clean_excel_value(value: Any) -> Any:
        """Удаляет управляющие символы XML, запрещённые внутри XLSX."""
        if value is None:
            return value
        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub("", value)
        return value

    def _sanitize_dataframe_for_excel(self, df: pd.DataFrame) -> pd.DataFrame:
        clean = df.copy()
        clean.columns = [
            ILLEGAL_CHARACTERS_RE.sub("", str(col)) for col in clean.columns
        ]
        object_cols = clean.select_dtypes(include=["object", "string"]).columns
        for col in object_cols:
            clean[col] = clean[col].map(self._clean_excel_value)
        return clean

    def _excel_bytes(self, df: pd.DataFrame, sheet_name: str = "Данные") -> bytes:
        buffer = io.BytesIO()
        df = self._sanitize_dataframe_for_excel(df)
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            safe_sheet = re.sub(r"[\\/*?:\[\]]", "_", sheet_name)[:31] or "Данные"
            safe_sheet = ILLEGAL_CHARACTERS_RE.sub("", safe_sheet)
            df.to_excel(writer, sheet_name=safe_sheet, index=False)
            ws = writer.book[safe_sheet]
            ws.freeze_panes = "A2"
            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col_idx, col in enumerate(df.columns, start=1):
                sample = [len(str(col))] + [len(str(x)) for x in df[col].dropna().astype(str).head(150)]
                ws.column_dimensions[get_column_letter(col_idx)].width = min(42, max(10, max(sample, default=10) + 2))
                header = str(col)
                # Понятное отображение чисел: штуки/показы/клики — целые;
                # управленческие рубли — целые; точные бухгалтерские суммы — с копейками.
                if header == "SKU Ozon":
                    number_format = "0"
                elif "шт." in header or header in {"Показы", "Клики", "Финансовых операций"}:
                    number_format = "#,##0"
                elif "%" in header:
                    number_format = "0.0"
                elif "₽" in header:
                    if sheet_name.startswith("Отчёт реализации"):
                        number_format = "#,##0.00"
                    elif any(token in header for token in ("CPC", "CPO")):
                        number_format = "#,##0.00"
                    else:
                        number_format = "#,##0"
                else:
                    number_format = None
                if number_format:
                    for cell in ws[get_column_letter(col_idx)][1:]:
                        cell.number_format = number_format
        return buffer.getvalue()


    def _date_candidates_for_result(self, result: ReportResult) -> List[str]:
        candidates = list(REPORT_DATE_CANDIDATES.get(result.code, []))
        candidates.extend([
            result.date_column, "Дата", "day", "created_at", "in_process_at",
            "shipment_date", "logistic.return_date", "return_date", "posting_date",
            "Дата заказа", "Дата создания", "Дата события", "Дата снимка",
        ])
        return list(dict.fromkeys([c for c in candidates if c]))

    def _detect_row_date_column(self, df: pd.DataFrame, preferred: str = "", code: str = "") -> Optional[str]:
        candidates = list(REPORT_DATE_CANDIDATES.get(code, []))
        candidates.extend([
            preferred, "Дата", "day", "created_at", "in_process_at", "shipment_date",
            "logistic.return_date", "return_date", "posting_date", "Дата заказа",
            "Дата создания", "Дата события", "Дата снимка",
        ])
        for col in dict.fromkeys(c for c in candidates if c):
            if col in df.columns:
                parsed = pd.to_datetime(df[col], errors="coerce", utc=True)
                if parsed.notna().any():
                    return col
        return None

    def _filter_existing_to_partition_week(
        self, result: ReportResult, old: pd.DataFrame, partition_date: date
    ) -> pd.DataFrame:
        if old is None or old.empty:
            return pd.DataFrame()
        old = normalize_columns(old)

        # Старую воронку до v17 нельзя сохранять: именно в ней ordered_units/revenue
        # были ошибочно подписаны как hits_view_*. History пересоберёт её за 60 дней.
        if result.code == "funnel":
            if "_schema_version" not in old.columns:
                return pd.DataFrame()
            versions = pd.to_numeric(old["_schema_version"], errors="coerce")
            if not versions.ge(REPORT_SCHEMA_VERSIONS["funnel"]).any():
                return pd.DataFrame()

        old = self._postprocess_report_df(result.code, old)
        date_col = self._detect_row_date_column(old, result.date_column, result.code)
        if not date_col:
            return pd.DataFrame()
        parsed = pd.to_datetime(old[date_col], errors="coerce", utc=True)
        iso = partition_date.isocalendar()
        mask = parsed.map(
            lambda x: bool(pd.notna(x) and x.date().isocalendar()[:2] == iso[:2])
        )
        filtered = old[mask].copy()
        return filtered.reset_index(drop=True)

    def _slice_for_dates(self, result: ReportResult, df: pd.DataFrame, dates: set[date]) -> pd.DataFrame:
        if df is None or df.empty:
            return pd.DataFrame()
        date_col = self._detect_row_date_column(df, result.date_column, result.code)
        if not date_col:
            return pd.DataFrame()
        parsed = pd.to_datetime(df[date_col], errors="coerce", utc=True).dt.date
        return df[parsed.isin(dates)].copy()

    def _save_weekly_partitioned(self, result: ReportResult, df: pd.DataFrame) -> None:
        """Разносит строки по фактической дате и проверяет целостность записи."""
        df = self._postprocess_report_df(result.code, df)
        date_col = self._detect_row_date_column(df, result.date_column, result.code)
        if not date_col:
            partitions = [(self.target_date, df)]
        else:
            work = df.copy()
            parsed = pd.to_datetime(work[date_col], errors="coerce", utc=True)
            work["_partition_date"] = parsed.dt.date
            valid = work[work["_partition_date"].notna()].copy()
            invalid = work[work["_partition_date"].isna()].drop(columns=["_partition_date"], errors="ignore")
            partitions: List[Tuple[date, pd.DataFrame]] = []
            if not valid.empty:
                valid["_week_key"] = valid["_partition_date"].map(lambda d: iso_week(d))
                for _, part in valid.groupby("_week_key", dropna=False):
                    part_date = part["_partition_date"].iloc[0]
                    partitions.append((part_date, part.drop(columns=["_partition_date", "_week_key"], errors="ignore")))
            if not invalid.empty:
                partitions.append((self.target_date, invalid))

        written_keys: List[str] = []
        audit_parts: List[Dict[str, Any]] = []
        for partition_date, part_df in partitions:
            filename = week_filename(result.filename_prefix, partition_date)
            key = f"{self.base_prefix}/{result.folder}/{self.store}/Недельные/{filename}"
            old_raw = self.storage.read_excel(key)
            old = self._filter_existing_to_partition_week(result, old_raw, partition_date)
            part_df = self._postprocess_report_df(result.code, part_df)

            part_dates: set[date] = set()
            part_date_col = self._detect_row_date_column(part_df, result.date_column, result.code)
            if part_date_col:
                part_dates = set(pd.to_datetime(part_df[part_date_col], errors="coerce", utc=True).dropna().dt.date)
            if not part_dates:
                part_dates = {partition_date}

            # Ремонт/повторный запуск должен ЗАМЕНЯТЬ дневной срез, а не
            # дописывать его поверх старой схемы. Именно из-за сохранения старых
            # строк v17 получал unique_saved > unique_received.
            if old is not None and not old.empty:
                old_date_col = self._detect_row_date_column(old, result.date_column, result.code)
                if old_date_col:
                    old_dates = pd.to_datetime(old[old_date_col], errors="coerce", utc=True).dt.date
                    old = old[~old_dates.isin(part_dates)].copy()
            merged = dedupe_merge(old, part_df, result.keys)

            expected_unique = dataframe_unique_count(part_df, result.keys)
            saved_slice = self._slice_for_dates(result, merged, part_dates)
            saved_unique = dataframe_unique_count(saved_slice, result.keys)
            integrity = expected_unique == saved_unique

            self.storage.upload_bytes(
                key, self._excel_bytes(merged, result.title),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            written_keys.append(key)
            audit_parts.append({
                "key": key,
                "dates": sorted(d.isoformat() for d in part_dates),
                "rows_received": int(len(part_df)),
                "unique_received": expected_unique,
                "rows_in_week_after_save": int(len(merged)),
                "unique_saved_for_dates": saved_unique,
                "integrity": integrity,
            })
            if not integrity:
                logging.error(
                    "Нарушение целостности %s: получено unique=%s, сохранено unique=%s, key=%s",
                    result.code, expected_unique, saved_unique, key,
                )
            logging.info("Сохранено: s3://%s/%s (%s строк)", self.storage.bucket, key, len(merged))

        result.s3_key = "; ".join(written_keys)
        result.save_audit = {
            "integrity": all(p["integrity"] for p in audit_parts) if audit_parts else True,
            "partitions": audit_parts,
        }

    def save_results(self) -> None:
        archive_local_files: List[Tuple[str, str]] = []
        for result in self.results:
            if result.skip_standard_save:
                continue
            if result.status in {"ERROR", "SKIPPED_OPTIONAL"}:
                if self.mode != "archive":
                    continue
                df = pd.DataFrame([{
                    "Статус": result.status,
                    "Отчёт": result.title,
                    "Дата": self.target_date.isoformat(),
                    "Период с": self.period_from.isoformat(),
                    "Период по": self.period_to.isoformat(),
                    "Ошибка": result.message,
                    "Методы": ", ".join(result.method_paths),
                }])
            else:
                df = result.df.copy()
            if df.empty:
                # В archive режиме сохраняем и пустой отчёт, чтобы было видно, что метод отработал.
                if self.mode != "archive":
                    continue
                df = pd.DataFrame([{
                    "Статус": "Нет строк",
                    "Дата": self.target_date.isoformat(),
                    "Методы": ", ".join(result.method_paths),
                }])

            if self.mode == "archive":
                local_name = f"{result.filename_prefix}_{self.target_date.isoformat()}.xlsx"
                local_path = self.workdir / local_name
                local_path.write_bytes(self._excel_bytes(df, result.title))
                result.local_path = str(local_path)
                archive_local_files.append((str(local_path), f"{result.folder}/{local_name}"))
                continue

            if result.snapshot:
                key = f"{self.base_prefix}/{result.folder}/{self.store}/{result.filename_prefix}.xlsx"
                merged = df
            else:
                self._save_weekly_partitioned(result, df)
                # Сырые ответы сохраняются ниже.
                merged = None
                key = result.s3_key

            if result.snapshot:
                self.storage.upload_bytes(key, self._excel_bytes(merged, result.title),
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                result.s3_key = key
                logging.info("Сохранено: s3://%s/%s (%s строк)", self.storage.bucket, key, len(merged))

            # Полные RAW JSON очень быстро раздувают Object Storage.
            # По умолчанию НЕ сохраняем. Включать только временно для отладки:
            # OZON_SAVE_RAW_RESPONSES=1.
            save_raw = str(os.getenv("OZON_SAVE_RAW_RESPONSES", "0")).strip().lower() in {
                "1", "true", "yes", "on"
            }
            if save_raw and result.raw is not None:
                raw_key = (
                    f"{self.base_prefix}/Служебные/{self.store}/Сырые ответы/"
                    f"{self.target_date.isoformat()}/{result.code}_{self.run_id}.json"
                )
                self.storage.upload_json(raw_key, result.raw)

        if self.mode == "archive":
            archive_dir = self.workdir / f"Архив_{self.store}_{self.target_date.isoformat()}"
            archive_dir.mkdir(exist_ok=True)
            zip_name = f"Архив_всех_отчётов_{self.store}_{self.target_date.isoformat()}_{self.run_id}.zip"
            zip_path = self.workdir / zip_name
            manifest = self.build_manifest()
            manifest_path = self.workdir / "manifest.json"
            manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
            with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as zf:
                for local_path, archive_name in archive_local_files:
                    zf.write(local_path, archive_name)
                zf.write(manifest_path, "Служебные/manifest.json")
                if self.errors:
                    err_path = self.workdir / "errors.json"
                    err_path.write_text(json.dumps(self.errors, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
                    zf.write(err_path, "Служебные/errors.json")
            archive_key = (f"{self.base_prefix}/Архив/{self.store}/{self.target_date:%Y/%m}/"
                           f"{zip_name}")
            self.storage.upload_file(str(zip_path), archive_key)
            logging.info("Архив сохранён: s3://%s/%s", self.storage.bucket, archive_key)
            # Удобный указатель на последний архив выбранного дня.
            self.storage.upload_json(
                f"{self.base_prefix}/Архив/{self.store}/Последний_архив.json",
                {"date": self.target_date.isoformat(), "key": archive_key, "run_id": self.run_id,
                 "created_at": datetime.now(MOSCOW_TZ).isoformat()},
            )

    def build_manifest(self) -> Dict[str, Any]:
        return {
            "script_version": SCRIPT_VERSION,
            "run_id": self.run_id,
            "store": self.store,
            "mode": self.mode,
            "target_date": self.target_date.isoformat(),
            "period_from": self.period_from.isoformat(),
            "period_to": self.period_to.isoformat(),
            "started_at": datetime.now(MOSCOW_TZ).isoformat(),
            "reports": [
                {
                    "code": r.code,
                    "title": r.title,
                    "status": r.status,
                    "rows": len(r.df),
                    "methods": r.method_paths,
                    "s3_key": r.s3_key,
                    "message": r.message,
                    "save_audit": r.save_audit,
                } for r in self.results
            ],
            "errors_count": len(self.errors),
        }

    def save_diagnostics(self) -> None:
        manifest = self.build_manifest()
        base = f"{self.base_prefix}/Служебные/{self.store}"
        self.storage.upload_json(f"{base}/Запуски/{self.target_date.isoformat()}_{self.run_id}.json", manifest)
        self.storage.upload_json(f"{base}/Последний_запуск.json", manifest)
        if self.errors:
            self.storage.upload_json(f"{base}/Ошибки_API/{self.target_date.isoformat()}_{self.run_id}.json", self.errors)
        required_errors = [r for r in self.results if r.status == "ERROR" and not r.optional]
        if required_errors:
            names = ", ".join(r.title for r in required_errors)
            if self.mode == "daily":
                raise RuntimeError(
                    "Есть ошибки обязательных отчётов в daily: "
                    + names
                    + ". См. Служебные/Ошибки_API"
                )
            logging.error(
                "В режиме %s есть ошибки обязательных отчётов: %s. "
                "Архив/тест сохранён с diagnostics и workflow не прерывается.",
                self.mode,
                names,
            )



# ---------------------------------------------------------------------------
# Coverage / orchestration
# ---------------------------------------------------------------------------

EVENT_REPORT_CODES = [
    "orders",
    "returns",
    "funnel",
    "finance_accruals",
    "finance_by_sku",
    "ad_statistics",
    "ad_product_statistics",
    "ad_orders",
]

EVENT_REPORT_LOCATIONS: Dict[str, Tuple[str, str]] = {
    "orders": ("Заказы", "Заказы"),
    "returns": ("Возвраты", "Возвраты"),
    "funnel": ("Воронка продаж", "Воронка_продаж"),
    "finance_accruals": ("Финансовые начисления", "Финансовые_начисления"),
    "finance_by_sku": ("Финансы по артикулам", "Финансы_по_артикулам"),
    "ad_statistics": ("Реклама/Статистика", "Реклама_статистика"),
    "ad_product_statistics": ("Реклама/По товарам", "Реклама_по_товарам"),
    "ad_orders": ("Реклама/Заказы", "Рекламные_заказы"),
}


class CoverageTracker:
    """Служебный индекс полноты посуточных данных за последние 60 дней."""

    def __init__(self, storage: S3Storage, store: str, target_date: date):
        self.storage = storage
        self.store = store
        self.target_date = target_date
        self.key = (
            f"Отчёты/Служебные/{store}/Покрытие данных/"
            "Покрытие_60_дней.json"
        )
        raw = storage.read_json(self.key, default={})
        self.data: Dict[str, Any] = dict(raw) if isinstance(raw, Mapping) else {}
        self.data.setdefault("store", store)
        self.data.setdefault("history_days", DEFAULT_HISTORY_DAYS)
        self.data.setdefault("reports", {})

    def _entry(self, code: str, day: date) -> Optional[Mapping[str, Any]]:
        reports = self.data.get("reports", {})
        report = reports.get(code, {}) if isinstance(reports, Mapping) else {}
        value = report.get(day.isoformat()) if isinstance(report, Mapping) else None
        return value if isinstance(value, Mapping) else None

    def is_covered(self, code: str, day: date) -> bool:
        entry = self._entry(code, day)
        if not entry:
            return False
        expected_schema = int(REPORT_SCHEMA_VERSIONS.get(code, 1))
        actual_schema = int(entry.get("schema_version", 1) or 1)
        if actual_schema < expected_schema:
            return False
        status = str(entry.get("status", "")).lower()
        if status == "complete":
            if entry.get("save_integrity") is False:
                return False
            location = EVENT_REPORT_LOCATIONS.get(code)
            if location:
                folder, prefix = location
                expected_key = (
                    f"Отчёты/{folder}/{self.store}/Недельные/"
                    f"{week_filename(prefix, day)}"
                )
                if not self.storage.exists(expected_key):
                    logging.warning("Покрытие отмечено complete, но файл отсутствует: %s", expected_key)
                    return False
        return status in {
            "complete",
            "empty",
            "unavailable",
            "no_data",
        }

    def day_is_covered(self, day: date, active_codes: Sequence[str]) -> bool:
        return all(self.is_covered(code, day) for code in active_codes)

    def mark_result(self, code: str, day: date, result: ReportResult) -> None:
        reports = self.data.setdefault("reports", {})
        report = reports.setdefault(code, {})
        status = "error"

        save_integrity = bool(result.save_audit.get("integrity", True))
        if result.status == "OK" and save_integrity:
            status = "complete"
        elif result.status == "OK" and not save_integrity:
            status = "corrupted"
        elif result.status == "EMPTY":
            status = "empty"
        elif result.status == "SKIPPED_OPTIONAL":
            text = (result.message or "").lower()
            if any(token in text for token in (
                "there is no data",
                "no data",
                "not found",
                "http 404",
                "permission",
                "forbidden",
                "unsupported",
                "недоступ",
            )):
                status = "unavailable"
            else:
                status = "error"
        elif result.status == "ERROR":
            status = "error"

        report[day.isoformat()] = {
            "status": status,
            "rows": int(len(result.df)),
            "report_status": result.status,
            "message": result.message,
            "methods": list(result.method_paths),
            "updated_at": datetime.now(MOSCOW_TZ).isoformat(),
            "schema_version": int(REPORT_SCHEMA_VERSIONS.get(code, 1)),
            "save_integrity": save_integrity,
            "save_audit": result.save_audit,
        }

    def mark_missing_reports(
        self,
        day: date,
        active_codes: Sequence[str],
        present_results: Sequence[ReportResult],
    ) -> None:
        present = {r.code for r in present_results}
        reports = self.data.setdefault("reports", {})
        for code in active_codes:
            if code in present:
                continue
            report = reports.setdefault(code, {})
            report[day.isoformat()] = {
                "status": "unavailable",
                "rows": 0,
                "report_status": "NOT_RUN",
                "message": "Метод отключён для текущей конфигурации",
                "methods": [],
                "updated_at": datetime.now(MOSCOW_TZ).isoformat(),
                "schema_version": int(REPORT_SCHEMA_VERSIONS.get(code, 1)),
            }

    def dates_to_process(
        self,
        mode: str,
        active_codes: Sequence[str],
        repair_days_per_run: int,
    ) -> List[date]:
        mode = str(mode).lower()
        if mode in {"test", "archive"}:
            return [self.target_date]

        all_days = list(iter_days(history_start(self.target_date), self.target_date))
        missing = [
            day for day in all_days
            if not self.day_is_covered(day, active_codes)
        ]

        if mode == "history":
            return missing

        # daily: вчера всегда обновляем заново, затем ограниченно ремонтируем старые даты.
        ordered: List[date] = [self.target_date]
        old_missing = [d for d in missing if d != self.target_date]
        ordered.extend(old_missing[:max(0, int(repair_days_per_run))])
        return list(dict.fromkeys(ordered))

    def save(self) -> None:
        self.data["updated_at"] = datetime.now(MOSCOW_TZ).isoformat()
        self.data["period_from"] = history_start(self.target_date).isoformat()
        self.data["period_to"] = self.target_date.isoformat()
        self.storage.upload_json(self.key, self.data)


def collector_result_summary(results: Sequence[ReportResult]) -> List[Dict[str, Any]]:
    return [
        {
            "code": r.code,
            "title": r.title,
            "status": r.status,
            "rows": len(r.df),
            "methods": list(r.method_paths),
            "s3_key": r.s3_key,
            "message": r.message,
            "save_audit": r.save_audit,
        }
        for r in results
    ]


def campaign_ids_from_collector(collector: OzonFboCollector) -> List[str]:
    return collector._campaign_ids_from_results()


def copy_reference_context(
    source: OzonFboCollector,
    target: OzonFboCollector,
) -> None:
    target.catalog_items = list(source.catalog_items)
    target.catalog_ids = list(source.catalog_ids)
    target.offer_ids = list(source.offer_ids)
    target.sku_ids = list(source.sku_ids)
    target.campaign_ids_override = campaign_ids_from_collector(source)
    target.active_campaign_ids_override = source._active_campaign_ids_from_results()
    target.reference_sku_to_offer = dict(source.reference_sku_to_offer)
    target.reference_sku_to_name = dict(source.reference_sku_to_name)


def rebuild_finance_by_sku_from_storage(
    client: OzonSellerClient,
    storage: S3Storage,
    store: str,
    target_date: date,
    workdir: Path,
) -> Dict[str, Any]:
    """Быстро пересобирает 60 дней «Финансов по артикулам» без повторных запросов истории.

    Берёт уже сохранённые недельные «Финансовые начисления», агрегирует их по
    дата × SKU и перезаписывает понятные управленческие недельные файлы.
    Из API запрашивается только актуальный справочник товаров для связи SKU -> артикул.
    """
    collector = OzonFboCollector(
        client, storage, store, target_date, "test", workdir / "finance_rebuild",
        performance_client=None, fast_history=True,
    )
    mapping_errors: List[str] = []
    try:
        collector.fetch_product_catalog()
        collector.fetch_product_info()
    except Exception as exc:
        mapping_errors.append(str(exc))
        logging.warning("Не удалось обновить справочник SKU -> артикул: %s", exc)

    start = history_start(target_date)
    weeks = sorted({iso_week(day) for day in iter_days(start, target_date)})
    files_done: List[Dict[str, Any]] = []
    missing_sources: List[str] = []
    total_source_rows = 0
    total_result_rows = 0

    for year, week in weeks:
        week_day = date.fromisocalendar(year, week, 1)
        source_key = (
            f"Отчёты/Финансовые начисления/{store}/Недельные/"
            f"Финансовые_начисления_{year}-W{week:02d}.xlsx"
        )
        source = storage.read_excel(source_key)
        if source is None or source.empty:
            missing_sources.append(source_key)
            continue

        # Ограничиваем первый/последний неполный недельный файл нашим 60-дневным окном.
        date_col = collector._detect_row_date_column(source, "Дата", "finance_accruals")
        if date_col:
            parsed = pd.to_datetime(source[date_col], errors="coerce", utc=True).dt.date
            source = source[(parsed >= start) & (parsed <= target_date)].copy()
        if source.empty:
            continue

        result_df, meta, _ = collector.build_finance_by_sku(source)
        result_df = collector._postprocess_report_df("finance_by_sku", result_df)
        target_key = (
            f"Отчёты/Финансы по артикулам/{store}/Недельные/"
            f"Финансы_по_артикулам_{year}-W{week:02d}.xlsx"
        )
        storage.upload_bytes(
            target_key,
            collector._excel_bytes(result_df, "Финансы по артикулам"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        files_done.append({
            "неделя": f"{year}-W{week:02d}",
            "источник": source_key,
            "файл": target_key,
            "строк_источника": int(len(source)),
            "строк_результата": int(len(result_df)),
        })
        total_source_rows += int(len(source))
        total_result_rows += int(len(result_df))
        logging.info(
            "Финансы по артикулам %s-W%02d: %s -> %s строк",
            year, week, len(source), len(result_df),
        )

    info = {
        "status": "OK",
        "store": store,
        "period_from": start.isoformat(),
        "period_to": target_date.isoformat(),
        "weeks_processed": len(files_done),
        "source_rows": total_source_rows,
        "result_rows": total_result_rows,
        "files": files_done,
        "missing_sources": missing_sources,
        "mapping_errors": mapping_errors,
        "created_at": datetime.now(MOSCOW_TZ).isoformat(),
    }
    storage.upload_json(
        f"Отчёты/Служебные/{store}/Финансы по артикулам/Последняя_пересборка.json",
        info,
    )
    return info


def create_archive_all(
    storage: S3Storage,
    stores: Sequence[str],
    target_date: date,
    workdir: Path,
) -> Dict[str, Any]:
    """ZIP только из канонических файлов текущего 60-дневного контура.

    Исключает legacy-файлы без папки «Недельные», старые ошибочные недели за
    пределами 60 дней, сырые ответы и предыдущие архивы.
    """
    selected = {str(s).upper() for s in stores}
    objects = storage.list_objects("Отчёты/")
    included: List[Dict[str, Any]] = []
    excluded: List[Dict[str, Any]] = []
    period_start = history_start(target_date)

    event_prefixes = {
        "Возвраты", "Воронка продаж", "Заказы", "Финансовые начисления", "Финансы по артикулам",
        "Реклама/Статистика", "Реклама/По товарам", "Реклама/Заказы",
        "Поисковые запросы", "Поисковые запросы по товарам",
    }

    def parse_week_from_key(key: str) -> Optional[date]:
        match = re.search(r"_(\d{4})-W(\d{2})\.xlsx$", key)
        if not match:
            return None
        year, week = int(match.group(1)), int(match.group(2))
        try:
            return date.fromisocalendar(year, week, 1)
        except ValueError:
            return None

    def belongs(key: str) -> Tuple[bool, str]:
        if key.startswith("Отчёты/Архив/") or key.startswith("Отчёты/Архивы/"):
            return False, "previous_archive"
        if "/Сырые ответы/" in key or key.lower().endswith(".zip"):
            return False, "raw_or_zip"
        if not key.lower().endswith((".xlsx", ".json", ".csv")):
            return False, "unsupported_extension"
        if not any(f"/{store}/" in key for store in selected):
            return False, "other_store"

        relative = key[len("Отчёты/"):] if key.startswith("Отчёты/") else key
        is_service = relative.startswith("Служебные/")
        is_weekly = "/Недельные/" in key
        is_current_realization = (
            relative.startswith("Отчёт реализации/")
            and key.endswith("/Отчёт_реализации_последний_закрытый_месяц.xlsx")
        )
        if not is_service and not is_weekly and not is_current_realization:
            return False, "legacy_non_weekly"

        if is_weekly:
            week_monday = parse_week_from_key(key)
            first_section = relative.split("/", 1)[0]
            # Для вложенной рекламы восстанавливаем два сегмента.
            if relative.startswith("Реклама/"):
                first_section = "/".join(relative.split("/")[:2])
            if first_section in event_prefixes and week_monday is not None:
                week_sunday = week_monday + timedelta(days=6)
                if week_sunday < period_start or week_monday > target_date:
                    return False, "outside_60_days"
        return True, ""

    candidates: List[Dict[str, Any]] = []
    for obj in objects:
        key = str(obj.get("Key", ""))
        ok, reason = belongs(key)
        if ok:
            candidates.append(obj)
        elif reason not in {"other_store", "unsupported_extension", "previous_archive", "raw_or_zip"}:
            excluded.append({"Ключ": key, "Причина": reason})

    candidates.sort(key=lambda x: str(x.get("Key", "")))
    workdir.mkdir(parents=True, exist_ok=True)
    stores_label = "ALL" if selected == ALLOWED_STORES else "_".join(sorted(selected))
    run_id = datetime.now(MOSCOW_TZ).strftime("%Y%m%d_%H%M%S")
    zip_name = f"Архив_всех_отчётов_{stores_label}_{target_date.isoformat()}_{run_id}.zip"
    zip_path = workdir / zip_name

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as zf:
        for index, obj in enumerate(candidates, start=1):
            key = str(obj["Key"])
            try:
                data = storage.read_bytes(key)
            except Exception as exc:
                included.append({"Ключ": key, "Статус": "ERROR", "Ошибка": str(exc), "Размер, байт": int(obj.get("Size", 0) or 0)})
                continue
            archive_name = key[len("Отчёты/"):] if key.startswith("Отчёты/") else key
            zf.writestr(archive_name, data)
            included.append({
                "Ключ": key, "Файл в архиве": archive_name, "Статус": "OK",
                "Размер, байт": len(data), "Изменён": str(obj.get("LastModified") or ""),
            })
            if index % 25 == 0:
                logging.info("Архивировано %s/%s файлов", index, len(candidates))

        manifest_bytes = io.BytesIO()
        with pd.ExcelWriter(manifest_bytes, engine="openpyxl") as writer:
            pd.DataFrame(included).to_excel(writer, index=False, sheet_name="Файлы")
            pd.DataFrame(excluded).to_excel(writer, index=False, sheet_name="Исключено")
            pd.DataFrame([{
                "Магазины": ", ".join(sorted(selected)),
                "Дата создания": datetime.now(MOSCOW_TZ).isoformat(),
                "Период данных с": period_start.isoformat(),
                "Период данных по": target_date.isoformat(),
                "Файлов найдено": len(candidates),
                "Файлов добавлено": sum(1 for x in included if x.get("Статус") == "OK"),
                "Ошибок": sum(1 for x in included if x.get("Статус") == "ERROR"),
                "Legacy исключено": len(excluded),
                "Сырые ответы исключены": True,
                "Предыдущие архивы исключены": True,
            }]).to_excel(writer, index=False, sheet_name="Сводка")
        zf.writestr("Манифест_архива.xlsx", manifest_bytes.getvalue())

    archive_key = f"Отчёты/Архивы/{stores_label}/{target_date:%Y/%m}/{zip_name}"
    storage.upload_file(str(zip_path), archive_key)
    latest_key = f"Отчёты/Архивы/{stores_label}/Последний_архив.json"
    result = {
        "status": "OK", "stores": sorted(selected), "period_from": period_start.isoformat(),
        "period_to": target_date.isoformat(), "files_found": len(candidates),
        "files_added": sum(1 for x in included if x.get("Статус") == "OK"),
        "legacy_excluded": len(excluded),
        "errors": [x for x in included if x.get("Статус") == "ERROR"],
        "local_path": str(zip_path), "s3_key": archive_key,
        "created_at": datetime.now(MOSCOW_TZ).isoformat(),
    }
    storage.upload_json(latest_key, result)
    logging.info("Общий архив сохранён: s3://%s/%s", storage.bucket, archive_key)
    return result

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def env_first(*names: str, default: str = "") -> str:
    for name in names:
        val = os.getenv(name)
        if val not in (None, ""):
            return str(val)
    return default


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Ozon FBO: ежедневный сбор продаж, аналитики и рекламы")
    p.add_argument(
        "--mode",
        choices=["test", "daily", "history", "finance_rebuild", "archive", "archive_all"],
        default=env_first("OZON_MODE", default="test"),
    )
    p.add_argument("--target-date", default=env_first("OZON_TARGET_DATE"))
    p.add_argument("--store", default=env_first("OZON_STORE", "STORE", default=DEFAULT_STORE))
    p.add_argument("--bucket", default=env_first("OZON_YC_BUCKET", "YC_BUCKET_NAME", default=DEFAULT_BUCKET))
    p.add_argument("--workdir", default="output_ozon_v19")
    p.add_argument(
        "--repair-days-per-run",
        type=int,
        default=int(env_first("OZON_REPAIR_DAYS_PER_RUN", default=str(DEFAULT_REPAIR_DAYS_PER_RUN))),
        help="Сколько старых пропущенных дней ремонтировать дополнительно в daily",
    )
    return p


def main() -> int:
    load_report_env()
    args = build_parser().parse_args()
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    logging.info("VERSION: %s", SCRIPT_VERSION)
    save_raw_enabled = str(os.getenv("OZON_SAVE_RAW_RESPONSES", "0")).strip().lower() in {
        "1", "true", "yes", "on"
    }
    logging.info(
        "RAW JSON в Object Storage: %s",
        "ВКЛЮЧЕНЫ (debug)" if save_raw_enabled else "ВЫКЛЮЧЕНЫ",
    )

    store = str(args.store).upper().strip() or DEFAULT_STORE
    if store not in ALLOWED_STORES and not (args.mode == "archive_all" and store == "ALL"):
        raise RuntimeError(f"Неизвестный магазин: {store}. Допустимо: {sorted(ALLOWED_STORES)} или ALL для archive_all")

    access_key = env_first("OZON_YC_ACCESS_KEY_ID", "YC_ACCESS_KEY_ID", "AWS_ACCESS_KEY_ID")
    secret_key = env_first("OZON_YC_SECRET_ACCESS_KEY", "YC_SECRET_ACCESS_KEY", "AWS_SECRET_ACCESS_KEY")
    endpoint = env_first("OZON_YC_ENDPOINT_URL", "YC_ENDPOINT_URL", default="https://storage.yandexcloud.net")
    if not access_key or not secret_key:
        raise RuntimeError("Не заданы ключи Yandex Object Storage")

    target = resolve_target_date(args.target_date)
    phase = resolve_request_phase()
    logging.info("Магазин: %s", store)
    logging.info("Режим: %s", args.mode)
    logging.info("Фаза запуска: %s", phase)
    logging.info("Целевая дата: %s", target)
    logging.info("Контроль покрытия: %s — %s", history_start(target), target)
    logging.info("Бакет: %s", args.bucket)

    storage = S3Storage(access_key, secret_key, args.bucket, endpoint)
    storage.ensure_bucket()

    workdir = Path(args.workdir) / store
    workdir.mkdir(parents=True, exist_ok=True)
    if args.mode == "archive_all":
        selected_stores = sorted(ALLOWED_STORES) if store == "ALL" else [store]
        info = create_archive_all(storage, selected_stores, target, workdir)
        print(json.dumps(info, ensure_ascii=False, indent=2, default=str))
        return 0

    client_id = env_first(f"OZON_CLIENT_ID_{store}", "OZON_CLIENT_ID")
    api_key = env_first(f"OZON_API_KEY_{store}", "OZON_API_KEY")
    performance_client_id = env_first(
        f"OZON_PERFORMANCE_CLIENT_ID_{store}",
        "OZON_PERFORMANCE_CLIENT_ID",
    )
    performance_client_secret = env_first(
        f"OZON_PERFORMANCE_CLIENT_SECRET_{store}",
        "OZON_PERFORMANCE_CLIENT_SECRET",
    )
    if not client_id or not api_key:
        raise RuntimeError(f"Не заданы OZON_CLIENT_ID_{store} и/или OZON_API_KEY_{store}")
    seller_client = OzonSellerClient(client_id, api_key)

    if args.mode == "finance_rebuild":
        info = rebuild_finance_by_sku_from_storage(
            seller_client, storage, store, target, workdir
        )
        print(json.dumps(info, ensure_ascii=False, indent=2, default=str))
        return 0

    performance_client: Optional[OzonPerformanceClient] = None
    if performance_client_id and performance_client_secret:
        performance_client = OzonPerformanceClient(
            performance_client_id,
            performance_client_secret,
        )
        logging.info("Performance API: ключи найдены, рекламные отчёты включены")
    else:
        logging.warning("Performance API: ключи не найдены, рекламные отчёты пропущены")

    started_monotonic = time.monotonic()

    # Архив — один единый проход строго за выбранный день.
    if args.mode == "archive":
        collector = OzonFboCollector(
            seller_client,
            storage,
            store,
            target,
            args.mode,
            workdir,
            performance_client=performance_client,
        )
        collector.collect_all()
        collector.save_results()
        collector.save_diagnostics()
        logging.info("Архив завершён")
        return 0

    # 1. Текущие снимки состояния.
    snapshot = OzonFboCollector(
        seller_client,
        storage,
        store,
        target,
        args.mode,
        workdir / "snapshot",
        performance_client=performance_client,
    )
    snapshot.collect_snapshot_reports()
    snapshot.save_results()

    # 2. Служебный рекламный лаг — отдельно от основных отчётов.
    lag_info: Dict[str, Any] = {}
    try:
        lag_info = snapshot.save_ad_lag_service()
    except Exception as exc:
        snapshot.log_error("ad_lag", "Рекламный лаг", exc, optional=True)
        lag_info = {"status": "ERROR", "error": str(exc)}

    # 3. Проверка полноты и посуточная дозагрузка.
    # История должна быстро восстановить ключевые управленческие данные.
    # Нулевые экспериментальные методы не запускаем 60 раз подряд.
    active_codes = ["orders", "returns", "funnel", "finance_accruals"]
    if performance_client is not None:
        active_codes.append("ad_statistics")
    if args.mode != "history" and performance_client is not None:
        active_codes.extend(["ad_product_statistics", "ad_orders"])

    coverage = CoverageTracker(storage, store, target)
    process_dates = coverage.dates_to_process(
        args.mode,
        active_codes,
        args.repair_days_per_run,
    )
    logging.info(
        "Посуточная очередь: %s дней (%s)",
        len(process_dates),
        ", ".join(d.isoformat() for d in process_dates[:10])
        + ("..." if len(process_dates) > 10 else ""),
    )

    event_runs: List[Dict[str, Any]] = []
    for index, day in enumerate(process_dates, start=1):
        elapsed_minutes = (time.monotonic() - started_monotonic) / 60
        if elapsed_minutes >= MAX_RUNTIME_MINUTES:
            logging.warning(
                "Достигнут лимит времени %s минут. Остаток будет догружен следующим запуском.",
                MAX_RUNTIME_MINUTES,
            )
            break

        logging.info(
            "=== Посуточная выгрузка %s/%s: %s ===",
            index,
            len(process_dates),
            day,
        )
        event_collector = OzonFboCollector(
            seller_client,
            storage,
            store,
            day,
            "test",
            workdir / f"events_{day.isoformat()}",
            performance_client=performance_client,
            fast_history=(args.mode == "history"),
        )
        copy_reference_context(snapshot, event_collector)
        event_collector.collect_event_reports(include_heavy=(args.mode != "history"))
        event_collector.save_results()

        for result in event_collector.results:
            if result.code in active_codes:
                coverage.mark_result(result.code, day, result)
        coverage.mark_missing_reports(day, active_codes, event_collector.results)
        coverage.save()

        event_runs.append({
            "date": day.isoformat(),
            "run_id": event_collector.run_id,
            "reports": collector_result_summary(event_collector.results),
            "errors": event_collector.errors,
        })

        elapsed = time.monotonic() - started_monotonic
        avg = elapsed / max(1, index)
        eta = avg * max(0, len(process_dates) - index)
        logging.info(
            "Прогресс history: %s/%s, прошло %.1f мин, ETA %.1f мин",
            index, len(process_dates), elapsed / 60, eta / 60,
        )

        # Небольшая пауза между днями снижает риск 429 на длинном history.
        time.sleep(API_PAUSE_SECONDS)

    # 4. Дорогие экспериментальные методы в history проверяем один раз.
    optional_probe_reports: List[Dict[str, Any]] = []
    if args.mode == "history":
        logging.info("=== Однократная проверка товарной рекламной аналитики ===")
        probe = OzonFboCollector(
            seller_client, storage, store, target, "test",
            workdir / "optional_probe",
            performance_client=performance_client,
            fast_history=True,
        )
        copy_reference_context(snapshot, probe)
        probe.collect_optional_event_reports()
        probe.save_results()
        optional_probe_reports = collector_result_summary(probe.results)

    # 5. Итоговый служебный отчёт запуска.
    summary = {
        "script_version": SCRIPT_VERSION,
        "store": store,
        "mode": args.mode,
        "phase": phase,
        "target_date": target.isoformat(),
        "history_from": history_start(target).isoformat(),
        "history_to": target.isoformat(),
        "started_at": snapshot.request_started_at.isoformat(),
        "finished_at": datetime.now(MOSCOW_TZ).isoformat(),
        "snapshot_reports": collector_result_summary(snapshot.results),
        "event_runs": event_runs,
        "optional_probe_reports": optional_probe_reports,
        "ad_lag": lag_info,
        "coverage_key": coverage.key,
        "snapshot_errors": snapshot.errors,
    }
    summary_key = (
        f"Отчёты/Служебные/{store}/Запуски/"
        f"{target.isoformat()}_{phase}_{snapshot.run_id}.json"
    )
    storage.upload_json(summary_key, summary)
    storage.upload_json(f"Отчёты/Служебные/{store}/Последний_запуск.json", summary)

    # Сохраняем обычную диагностику снимков, но не прерываем history из-за optional.
    snapshot.save_diagnostics()

    ok = sum(r.status == "OK" for r in snapshot.results)
    empty = sum(r.status == "EMPTY" for r in snapshot.results)
    skipped = sum(r.status == "SKIPPED_OPTIONAL" for r in snapshot.results)
    failed = sum(r.status == "ERROR" for r in snapshot.results)
    logging.info(
        "Готово. Снимки OK=%s, EMPTY=%s, optional skipped=%s, errors=%s, "
        "обработано дней=%s, summary=%s",
        ok,
        empty,
        skipped,
        failed,
        len(event_runs),
        summary_key,
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        logging.exception("Критическая ошибка: %s", exc)
        raise
