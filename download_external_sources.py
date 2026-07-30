# VERSION: EXTERNAL_SOURCES_TORGSTAT_ABC_WB_ENTRY_POINTS_V7_20260730
"""Загрузка внешних источников в Yandex Object Storage.

Источники:
1. Торгстат — АБС-анализ.
2. Wildberries — «Портрет покупателя → Точки входа».

Имя файла в репозитории: download_external_sources.py

GitHub Secrets:
- TORGSTAT_ABC_CURL: один актуальный Copy as cURL запроса выгрузки Торгстат.
- WB_ENTRY_POINTS_CURLS: несколько Copy as cURL подряд из кабинета WB. Код сам
  распознаёт запросы по URL. Достаточно запросов /file-manager/download и
  /tokensjrpc; список отчётов и конечный URL файла код умеет построить сам.
- REPORT_ENV или отдельные YC_ACCESS_KEY_ID, YC_SECRET_ACCESS_KEY,
  YC_BUCKET_NAME, YC_ENDPOINT_URL.

Режим auto:
- ежедневно: Торгстат АБС за вчера;
- по понедельникам: Торгстат АБС за закрытую неделю и Точки входа ВБ за
  закрытую неделю понедельник–воскресенье.
"""
from __future__ import annotations

import argparse
import datetime as dt
import io
import json
import os
import re
import shlex
import sys
import tempfile
import time
import uuid
import zipfile
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

import boto3
import requests
from openpyxl import load_workbook

VERSION = "EXTERNAL_SOURCES_TORGSTAT_ABC_WB_ENTRY_POINTS_V7_20260730"
DEFAULT_REPORTS_ROOT = "Отчёты"
DEFAULT_ABC_FOLDER = "ABC"
DEFAULT_WB_ENTRY_FOLDER = "Точки входа"
DEFAULT_STORE = "TOPFACE"
DEFAULT_TZ_OFFSET_HOURS = 3
WB_REPORT_TYPE = "CUSTOMER_PROFILE_ENTRY_POINTS_REPORT_V2"
WB_MANAGER_BASE = "https://seller-content.wildberries.ru/ns/analytics-api/content-analytics/api/v1/file-manager"
WB_TOKEN_URL = "https://seller-content.wildberries.ru/ns/suppliers-auth-tokens/suppliers-portal-core/api/v1/tokensjrpc"
WB_FILE_BASE = "https://downloads-content-analytics.wildberries.ru/api/v1/file-manager/download"

START_KEYS = {
    "datefrom", "fromdate", "begindate", "startdate", "datestart", "periodstart",
    "date_from", "from_date", "begin_date", "start_date", "date_start", "period_start",
    "from", "start", "begin", "dtfrom", "dt_from", "dfrom", "date1", "startperiod",
    "period[from]", "filter[datefrom]", "filter[from]", "filter[startdate]",
}
END_KEYS = {
    "dateto", "todate", "enddate", "dateend", "periodend",
    "date_to", "to_date", "end_date", "date_end", "period_end",
    "to", "end", "dtto", "dt_to", "dto", "date2", "endperiod",
    "period[to]", "filter[dateto]", "filter[to]", "filter[enddate]",
}
DATE_RE = re.compile(r"(?:\d{4}-\d{2}-\d{2}|\d{2}\.\d{2}\.\d{4}|\d{2}/\d{2}/\d{4})")


@dataclass
class CurlRequest:
    url: str
    method: str = "GET"
    headers: Dict[str, str] | None = None
    body: bytes | None = None

    def __post_init__(self) -> None:
        if self.headers is None:
            self.headers = {}


def log(message: str) -> None:
    print(message, flush=True)


def fail(message: str, code: int = 1) -> None:
    print(f"ERROR: {message}", file=sys.stderr, flush=True)
    raise SystemExit(code)


def load_report_env() -> None:
    """Загрузить KEY=VALUE из общего секрета REPORT_ENV, не затирая отдельные secrets."""
    raw = os.environ.get("REPORT_ENV", "") or ""
    if not raw.strip():
        log("report_env: REPORT_ENV пустой/не передан; использую отдельные env/secrets")
        return

    normalized = raw.replace("\r\n", "\n").replace("\r", "\n")
    if "\n" not in normalized and "\\n" in normalized:
        normalized = normalized.replace("\\n", "\n")

    loaded: List[str] = []
    skipped: List[str] = []
    bad_lines = 0
    for line in normalized.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[len("export "):].strip()
        if "=" not in line:
            bad_lines += 1
            continue
        key, value = line.split("=", 1)
        key = key.strip().lstrip("\ufeff")
        value = value.strip().strip('"').strip("'")
        if not key:
            bad_lines += 1
            continue
        if (os.environ.get(key) or "").strip():
            skipped.append(key)
            continue
        os.environ[key] = value
        loaded.append(key)

    interesting = [
        "YC_ACCESS_KEY_ID", "YC_SECRET_ACCESS_KEY", "YC_BUCKET_NAME", "YC_ENDPOINT_URL",
        "TORGSTAT_ABC_CURL", "WB_ENTRY_POINTS_CURLS",
    ]
    state = ", ".join(f"{key}={'set' if (os.environ.get(key) or '').strip() else 'empty'}" for key in interesting)
    log(f"report_env: loaded_keys={loaded}, skipped_existing={skipped}, bad_lines={bad_lines}")
    log(f"env_state: {state}")


def now_local() -> dt.datetime:
    return dt.datetime.now(dt.UTC) + dt.timedelta(hours=DEFAULT_TZ_OFFSET_HOURS)


def today_local() -> dt.date:
    return now_local().date()


def parse_date(value: str) -> dt.date:
    value = (value or "").strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(value, fmt).date()
        except ValueError:
            pass
    fail(f"Неверный формат даты: {value}. Используй YYYY-MM-DD или DD.MM.YYYY")


def fmt_iso(value: dt.date) -> str:
    return value.strftime("%Y-%m-%d")


def fmt_dmy(value: dt.date) -> str:
    return value.strftime("%d.%m.%Y")


def previous_full_week(reference: Optional[dt.date] = None) -> Tuple[dt.date, dt.date]:
    reference = reference or today_local()
    last_sunday = reference - dt.timedelta(days=reference.weekday() + 1)
    return last_sunday - dt.timedelta(days=6), last_sunday


def torgstat_periods(mode: str, date_from: str, date_to: str) -> List[Tuple[dt.date, dt.date, str]]:
    mode = mode.lower().strip()
    today = today_local()
    yesterday = today - dt.timedelta(days=1)
    if mode == "custom":
        if not date_from or not date_to:
            fail("mode=custom требует date_from и date_to")
        return [(parse_date(date_from), parse_date(date_to), "custom")]
    if mode == "daily":
        target = parse_date(date_from) if date_from else yesterday
        return [(target, target, "daily")]
    if mode == "weekly":
        start, end = previous_full_week(today)
        return [(start, end, "weekly")]
    if mode == "auto":
        result = [(yesterday, yesterday, "daily")]
        if today.weekday() == 0:
            start, end = previous_full_week(today)
            result.append((start, end, "weekly"))
        return result
    fail(f"Неизвестный mode={mode}")


def wb_periods(mode: str, date_from: str, date_to: str) -> List[Tuple[dt.date, dt.date, str]]:
    mode = mode.lower().strip()
    today = today_local()
    yesterday = today - dt.timedelta(days=1)
    if mode == "custom":
        if not date_from or not date_to:
            fail("mode=custom требует date_from и date_to")
        return [(parse_date(date_from), parse_date(date_to), "custom")]
    if mode == "daily":
        target = parse_date(date_from) if date_from else yesterday
        return [(target, target, "daily")]
    if mode == "weekly":
        start, end = previous_full_week(today)
        return [(start, end, "weekly")]
    if mode == "auto":
        if today.weekday() != 0:
            log("WB entry-points: auto — сегодня не понедельник, недельную выгрузку пропускаю")
            return []
        start, end = previous_full_week(today)
        return [(start, end, "weekly")]
    fail(f"Неизвестный mode={mode}")


def split_curl_commands(raw: str) -> List[str]:
    """Разделить multiline secret с несколькими Copy as cURL."""
    raw = (raw or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not raw:
        return []
    starts = [m.start() for m in re.finditer(r"(?mi)^\s*curl(?:\.exe)?\s+", raw)]
    if not starts:
        return [raw]
    starts.append(len(raw))
    return [raw[starts[i]:starts[i + 1]].strip() for i in range(len(starts) - 1)]


def clean_multiline_curl(raw: str) -> str:
    raw = (raw or "").strip()
    if not raw:
        fail("Пустой Copy as cURL")
    raw = raw.replace("\\\r\n", " ").replace("\\\n", " ")
    raw = raw.replace("\r\n", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", raw).strip()


def parse_curl(raw: str) -> CurlRequest:
    text = clean_multiline_curl(raw)
    text = re.sub(r"\s+[\^`]\s+", " ", text)
    try:
        parts = shlex.split(text, posix=True)
    except ValueError as exc:
        fail(f"Copy as cURL не распознан: {exc}")
    if parts and parts[0].lower() in {"curl", "curl.exe"}:
        parts = parts[1:]

    url: Optional[str] = None
    method = "GET"
    headers: Dict[str, str] = {}
    body_parts: List[str] = []
    i = 0
    while i < len(parts):
        item = parts[i]
        if item in {"-X", "--request"} and i + 1 < len(parts):
            method = parts[i + 1].upper()
            i += 2
            continue
        if item in {"-H", "--header"} and i + 1 < len(parts):
            header = parts[i + 1]
            if ":" in header:
                key, value = header.split(":", 1)
                headers[key.strip()] = value.strip()
            i += 2
            continue
        if item in {"-b", "--cookie", "--cookie-raw"} and i + 1 < len(parts):
            headers["cookie"] = parts[i + 1].strip()
            i += 2
            continue
        if item == "--url" and i + 1 < len(parts):
            url = parts[i + 1]
            i += 2
            continue
        if item in {"--data", "--data-raw", "--data-binary", "--data-ascii", "-d"} and i + 1 < len(parts):
            body_parts.append(parts[i + 1])
            method = "POST" if method == "GET" else method
            i += 2
            continue
        if item.startswith(("http://", "https://")):
            url = item
            i += 1
            continue
        i += 1

    if not url:
        fail("В Copy as cURL не найден URL")
    body = "&".join(body_parts).encode("utf-8") if body_parts else None
    return CurlRequest(url=url, method=method, headers=headers, body=body)


def header_get(headers: Dict[str, str], name: str) -> str:
    for key, value in headers.items():
        if key.lower() == name.lower():
            return value
    return ""


def header_set(headers: Dict[str, str], name: str, value: str) -> None:
    for key in list(headers):
        if key.lower() == name.lower():
            headers.pop(key, None)
    headers[name] = value


def is_valid_jwt_like(value: str) -> bool:
    value = (value or "").strip()
    return bool(re.fullmatch(r"[A-Za-z0-9_-]+\.[A-Za-z0-9_-]+\.[A-Za-z0-9_-]+", value))


def best_wb_auth_headers(requests_: Sequence[CurlRequest]) -> Dict[str, str]:
    """Собрать актуальные auth/cookie из всех cURL и игнорировать повреждённые JWT."""
    result: Dict[str, str] = {}
    for req in requests_:
        value = header_get(req.headers or {}, "authorizev3")
        if value and is_valid_jwt_like(value):
            result["authorizev3"] = value
        cookie = header_get(req.headers or {}, "cookie")
        if cookie and len(cookie) > len(result.get("cookie", "")):
            result["cookie"] = cookie
        for name in ("user-agent", "accept-language", "root-version"):
            value = header_get(req.headers or {}, name)
            if value:
                result[name] = value
    if not result.get("authorizev3"):
        fail("WB_ENTRY_POINTS_CURLS: не найден целый authorizev3. Скопируй cURL без ручных вставок внутрь токена")
    if not result.get("cookie"):
        fail("WB_ENTRY_POINTS_CURLS: не найдены cookies")
    return result


def apply_wb_auth(req: CurlRequest, auth: Dict[str, str]) -> CurlRequest:
    headers = dict(req.headers or {})
    for key, value in auth.items():
        header_set(headers, key, value)
    return CurlRequest(req.url, req.method, headers, req.body)


def sanitized_headers(
    headers: Dict[str, str],
    *,
    keep_content_type: bool = True,
    keep_download_token: bool = False,
) -> Dict[str, str]:
    """Удалить браузерные/транспортные заголовки перед requests.

    x-download-token запрещён для обычных запросов, но обязателен при скачивании
    готового XLSX с downloads-content-analytics.wildberries.ru.
    """
    output: Dict[str, str] = {}
    blocked = {"host", "content-length", "connection", "accept-encoding"}
    if not keep_download_token:
        blocked.add("x-download-token")

    for key, value in (headers or {}).items():
        low = key.lower().strip()
        if low.startswith(":") or low in blocked:
            continue
        if not keep_content_type and low == "content-type":
            continue
        output[key] = value
    return output


def request_json(session: requests.Session, req: CurlRequest, timeout: int = 180) -> Any:
    headers = sanitized_headers(req.headers or {})
    response = session.request(req.method, req.url, headers=headers, data=req.body, timeout=timeout, allow_redirects=True)
    log(f"HTTP {req.method} {urlparse(req.url).path}: status={response.status_code}, bytes={len(response.content):,}")
    if response.status_code >= 400:
        fail(f"HTTP {response.status_code}: {response.text[:1000]}")
    try:
        return response.json()
    except Exception:
        return {"raw": response.text[:3000]}


def norm_key(key: str) -> str:
    return re.sub(r"[^a-z0-9_\[\]]+", "", str(key).lower())


def format_like(original: Any, new_date: dt.date) -> str:
    text = str(original)
    if re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", text):
        return fmt_dmy(new_date)
    if re.fullmatch(r"\d{2}/\d{2}/\d{4}", text):
        return new_date.strftime("%d/%m/%Y")
    return fmt_iso(new_date)


def replace_dates_json(obj: Any, start: dt.date, end: dt.date) -> Tuple[Any, int]:
    changed = 0
    if isinstance(obj, dict):
        output: Dict[str, Any] = {}
        for key, value in obj.items():
            normalized = norm_key(key)
            if normalized in START_KEYS or "datefrom" in normalized or "startdate" in normalized or normalized.endswith("from"):
                output[key] = format_like(value, start)
                changed += 1
            elif normalized in END_KEYS or "dateto" in normalized or "enddate" in normalized or normalized.endswith("to"):
                output[key] = format_like(value, end)
                changed += 1
            else:
                output[key], nested = replace_dates_json(value, start, end)
                changed += nested
        return output, changed
    if isinstance(obj, list):
        output_list = []
        for value in obj:
            new_value, nested = replace_dates_json(value, start, end)
            output_list.append(new_value)
            changed += nested
        return output_list, changed
    return obj, 0


def replace_date_tokens(text: str, start: dt.date, end: dt.date) -> Tuple[str, int]:
    matches = list(DATE_RE.finditer(text))
    if not matches:
        return text, 0
    distinct: List[str] = []
    for match in matches:
        value = match.group(0)
        if value not in distinct:
            distinct.append(value)
    replacements: Dict[str, str] = {}
    if distinct:
        replacements[distinct[0]] = format_like(distinct[0], start)
    if len(distinct) > 1:
        replacements[distinct[1]] = format_like(distinct[1], end)
    output = text
    for old, new in replacements.items():
        output = output.replace(old, new)
    return output, len(replacements)


def update_request_dates(req: CurlRequest, start: dt.date, end: dt.date) -> CurlRequest:
    url = req.url
    body = req.body
    changed = 0

    parsed = urlparse(url)
    query = parse_qsl(parsed.query, keep_blank_values=True)
    new_query: List[Tuple[str, str]] = []
    for key, value in query:
        normalized = norm_key(key)
        if normalized in START_KEYS or "datefrom" in normalized or "startdate" in normalized or normalized.endswith("from"):
            new_query.append((key, format_like(value, start)))
            changed += 1
        elif normalized in END_KEYS or "dateto" in normalized or "enddate" in normalized or normalized.endswith("to"):
            new_query.append((key, format_like(value, end)))
            changed += 1
        else:
            new_query.append((key, value))
    if query:
        url = urlunparse(parsed._replace(query=urlencode(new_query, doseq=True)))

    if body:
        text = body.decode("utf-8", errors="replace")
        try:
            data = json.loads(text)
            data, body_changed = replace_dates_json(data, start, end)
            if body_changed:
                text = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
            changed += body_changed
        except Exception:
            text, body_changed = replace_date_tokens(text, start, end)
            changed += body_changed
        body = text.encode("utf-8")

    if changed == 0:
        url, url_changed = replace_date_tokens(url, start, end)
        changed += url_changed
    if changed == 0:
        fail("Не нашёл даты в TORGSTAT_ABC_CURL")
    log(f"date_replace: changed_fields={changed}, period={fmt_dmy(start)}-{fmt_dmy(end)}")
    return CurlRequest(url, req.method, dict(req.headers or {}), body)


def looks_like_zip(content: bytes) -> bool:
    return bool(content and content[:2] == b"PK")


def looks_like_xlsx(content: bytes) -> bool:
    """Быстрая проверка достаточно крупного ZIP/XLSX-ответа."""
    return bool(looks_like_zip(content) and len(content) > 1000)


def is_real_xlsx(content: bytes) -> bool:
    """Проверить, что ZIP является непосредственно XLSX, а не внешним архивом."""
    if not looks_like_zip(content):
        return False
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            names = set(archive.namelist())
            return "[Content_Types].xml" in names and "xl/workbook.xml" in names
    except (zipfile.BadZipFile, OSError):
        return False


def unwrap_wb_xlsx(content: bytes) -> Tuple[bytes, str]:
    """Извлечь XLSX из ответа WB или вернуть прямой XLSX."""
    if is_real_xlsx(content):
        return content, "direct_xlsx"

    if not looks_like_zip(content):
        preview = content[:300].decode("utf-8", errors="replace")
        fail(f"WB вернул не ZIP/XLSX: {preview}")

    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            members = [
                name
                for name in archive.namelist()
                if not name.endswith("/") and name.lower().endswith(".xlsx")
            ]
            if not members:
                names_preview = ", ".join(archive.namelist()[:20])
                fail(f"В архиве WB нет XLSX. Содержимое: {names_preview}")

            member = max(members, key=lambda name: archive.getinfo(name).file_size)
            inner = archive.read(member)
    except zipfile.BadZipFile as exc:
        fail(f"WB вернул повреждённый ZIP: {exc}")

    if not is_real_xlsx(inner):
        fail(f"Файл {member!r} внутри архива WB не является корректным XLSX")

    return inner, member


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower().replace("ё", "е"))


def validate_torgstat_xlsx(content: bytes) -> Tuple[str, int]:
    if not looks_like_xlsx(content):
        fail(f"Торгстат вернул не XLSX: {content[:300].decode('utf-8', errors='replace')}")
    path = write_temp_xlsx(content)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 30), values_only=True), 1):
                headers = [normalize_header(value) for value in row if value is not None]
                has_gp = any("валов" in header and "приб" in header for header in headers)
                has_article = any(token in header for header in headers for token in ("артикул wb", "артикул вб", "nm id", "nmid"))
                if has_gp and has_article:
                    return ws.title, row_index
        fail("XLSX Торгстата не содержит колонок валовой прибыли и артикула WB")
    finally:
        safe_remove(path)


def validate_wb_entry_xlsx(content: bytes) -> Tuple[str, int, int]:
    if not is_real_xlsx(content):
        fail("WB вернул ZIP, который не является корректным XLSX")
    path = write_temp_xlsx(content)
    try:
        # У файлов WB неверный dimension в XML, поэтому read_only=False обязателен.
        wb = load_workbook(path, read_only=False, data_only=True)
        for ws in wb.worksheets:
            for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 30), values_only=True), 1):
                headers = {normalize_header(value) for value in row if value is not None}
                required = {"раздел", "точка входа", "показы", "переходы в карточку", "заказы"}
                if required.issubset(headers):
                    return ws.title, row_index, ws.max_row
        fail("XLSX WB не похож на отчёт «Точки входа»: не найдены ожидаемые колонки")
    finally:
        safe_remove(path)


def write_temp_xlsx(content: bytes) -> str:
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        handle.write(content)
        return handle.name


def safe_remove(path: str) -> None:
    try:
        os.remove(path)
    except OSError:
        pass


def s3_client():
    endpoint = os.environ.get("YC_ENDPOINT_URL") or os.environ.get("S3_ENDPOINT_URL") or "https://storage.yandexcloud.net"
    region = os.environ.get("YC_REGION") or os.environ.get("AWS_DEFAULT_REGION") or "ru-central1"
    access = os.environ.get("YC_ACCESS_KEY_ID") or os.environ.get("AWS_ACCESS_KEY_ID")
    secret = os.environ.get("YC_SECRET_ACCESS_KEY") or os.environ.get("AWS_SECRET_ACCESS_KEY")
    if not access or not secret:
        fail("Нет YC_ACCESS_KEY_ID/YC_SECRET_ACCESS_KEY")
    return boto3.client("s3", endpoint_url=endpoint, region_name=region, aws_access_key_id=access, aws_secret_access_key=secret)


def upload_to_s3(content: bytes, key: str) -> None:
    bucket = os.environ.get("YC_BUCKET_NAME") or os.environ.get("S3_BUCKET") or os.environ.get("AWS_BUCKET")
    if not bucket:
        fail("Нет YC_BUCKET_NAME")
    s3_client().put_object(
        Bucket=bucket,
        Key=key,
        Body=content,
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    log(f"uploaded: s3://{bucket}/{key}")


def timestamp_suffix() -> str:
    return now_local().strftime("%Y-%m-%d_%H-%M")


def torgstat_key(start: dt.date, end: dt.date, reports_root: str, abc_folder: str) -> str:
    filename = f"wb_abc_report_goods__{fmt_dmy(start)}-{fmt_dmy(end)}__at_{timestamp_suffix()}.xlsx"
    return f"{reports_root.rstrip('/')}/{abc_folder.strip('/') or DEFAULT_ABC_FOLDER}/{filename}"


def wb_entry_key(store: str, start: dt.date, end: dt.date, reports_root: str, folder: str) -> str:
    filename = f"wb_entry_points__{fmt_dmy(start)}-{fmt_dmy(end)}__at_{timestamp_suffix()}.xlsx"
    return f"{reports_root.rstrip('/')}/{folder.strip('/') or DEFAULT_WB_ENTRY_FOLDER}/{filename}"


def download_torgstat_period(raw_curl: str, start: dt.date, end: dt.date, reports_root: str, abc_folder: str, dry_run: bool) -> str:
    req = update_request_dates(parse_curl(raw_curl), start, end)
    key = torgstat_key(start, end, reports_root, abc_folder)
    log(f"Torgstat target_key: {key}")
    if dry_run:
        return key
    headers = sanitized_headers(req.headers or {})
    if not header_get(headers, "cookie") and not header_get(headers, "authorization"):
        fail("TORGSTAT_ABC_CURL не содержит cookie/authorization")
    response = requests.request(req.method, req.url, headers=headers, data=req.body, timeout=180, allow_redirects=True)
    log(f"Torgstat response: status={response.status_code}, bytes={len(response.content):,}")
    if response.status_code >= 400:
        fail(f"Торгстат HTTP {response.status_code}: {response.text[:1000]}")
    content = response.content
    if not looks_like_xlsx(content):
        try:
            data = response.json()
            link = find_first_url(data)
            if link:
                second = requests.get(link, headers=headers, timeout=180, allow_redirects=True)
                content = second.content
        except Exception:
            pass
    sheet, row = validate_torgstat_xlsx(content)
    log(f"Torgstat XLSX OK: sheet={sheet}, header_row={row}")
    upload_to_s3(content, key)
    return key


def find_first_url(obj: Any) -> Optional[str]:
    if isinstance(obj, dict):
        for key, value in obj.items():
            if isinstance(value, str) and value.startswith(("http://", "https://")) and any(token in key.lower() for token in ("url", "link", "download", "file")):
                return value
            found = find_first_url(value)
            if found:
                return found
    elif isinstance(obj, list):
        for value in obj:
            found = find_first_url(value)
            if found:
                return found
    return None


def classify_wb_requests(raw: str) -> Tuple[CurlRequest, CurlRequest, CurlRequest, Optional[CurlRequest], Dict[str, str]]:
    commands = split_curl_commands(raw)
    if not commands:
        fail("WB_ENTRY_POINTS_CURLS пустой")
    requests_ = [parse_curl(command) for command in commands]
    create_req: Optional[CurlRequest] = None
    list_req: Optional[CurlRequest] = None
    token_req: Optional[CurlRequest] = None
    file_req: Optional[CurlRequest] = None

    for req in requests_:
        parsed = urlparse(req.url)
        path = parsed.path.lower()
        host = parsed.netloc.lower()
        if "downloads-content-analytics.wildberries.ru" in host:
            file_req = req
        elif path.endswith("/file-manager/downloads") or "/file-manager/downloads" in path:
            list_req = req
        elif path.endswith("/file-manager/download") and "seller-content.wildberries.ru" in host:
            create_req = req
        elif path.endswith("/tokensjrpc") or path.endswith("/tokens/rpc"):
            token_req = req

    if not create_req:
        fail("WB_ENTRY_POINTS_CURLS: не найден cURL /file-manager/download")
    if not token_req:
        fail("WB_ENTRY_POINTS_CURLS: не найден cURL /tokensjrpc")

    auth = best_wb_auth_headers(requests_)
    create_req = apply_wb_auth(create_req, auth)
    token_req = apply_wb_auth(token_req, auth)

    if not list_req:
        list_req = CurlRequest(
            url=f"{WB_MANAGER_BASE}/downloads?report_types={WB_REPORT_TYPE}",
            method="GET",
            headers=dict(create_req.headers or {}),
            body=None,
        )
    else:
        list_req = apply_wb_auth(list_req, auth)

    if file_req:
        file_req = apply_wb_auth(file_req, auth)
    return create_req, list_req, token_req, file_req, auth


def prepare_wb_create_request(template: CurlRequest, start: dt.date, end: dt.date, report_id: str) -> CurlRequest:
    try:
        payload = json.loads((template.body or b"{}").decode("utf-8"))
    except Exception as exc:
        fail(f"Не удалось прочитать JSON /file-manager/download: {exc}")
    payload["id"] = report_id
    payload["reportType"] = WB_REPORT_TYPE
    params = payload.setdefault("params", {})
    params["startDate"] = fmt_iso(start)
    params["endDate"] = fmt_iso(end)
    params.setdefault("nmsIDs", [])
    params.setdefault("subjectIDs", [])
    params.setdefault("brandNames", [])
    params.setdefault("tagIds", [])
    params.setdefault("vendorCodes", [])
    headers = dict(template.headers or {})
    header_set(headers, "content-type", "application/json")
    return CurlRequest(template.url, "POST", headers, json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8"))


def prepare_token_request(template: CurlRequest) -> CurlRequest:
    payload = {"method": "generateToken", "params": {"team": "content-analytics"}, "jsonrpc": "2.0", "id": f"json-rpc_{uuid.uuid4().hex[:10]}"}
    headers = dict(template.headers or {})
    header_set(headers, "content-type", "application/json")
    return CurlRequest(template.url or WB_TOKEN_URL, "POST", headers, json.dumps(payload, separators=(",", ":")).encode("utf-8"))


def find_dict_by_id(obj: Any, report_id: str) -> Optional[Dict[str, Any]]:
    if isinstance(obj, dict):
        for key in ("id", "reportId", "reportID", "uuid"):
            if str(obj.get(key, "")) == report_id:
                return obj
        for value in obj.values():
            found = find_dict_by_id(value, report_id)
            if found:
                return found
    elif isinstance(obj, list):
        for value in obj:
            found = find_dict_by_id(value, report_id)
            if found:
                return found
    return None


def report_state(entry: Optional[Dict[str, Any]]) -> str:
    if not entry:
        return "not_found"
    for key in ("status", "state", "reportStatus", "fileStatus"):
        if key in entry:
            return str(entry[key]).strip().lower()
    for key in ("isReady", "ready", "completed"):
        if entry.get(key) is True:
            return "ready"
    return "found"


def extract_download_token(obj: Any) -> Optional[str]:
    if isinstance(obj, dict):
        for key, value in obj.items():
            if "token" in key.lower() and isinstance(value, str) and len(value.strip()) > 50:
                return value.strip()
        if "result" in obj and isinstance(obj["result"], str) and len(obj["result"].strip()) > 50:
            return obj["result"].strip()
        for value in obj.values():
            found = extract_download_token(value)
            if found:
                return found
    elif isinstance(obj, list):
        for value in obj:
            found = extract_download_token(value)
            if found:
                return found
    elif isinstance(obj, str) and len(obj.strip()) > 100:
        return obj.strip()
    return None


def build_file_headers(file_template: Optional[CurlRequest], auth: Dict[str, str], token: str) -> Dict[str, str]:
    headers = dict(file_template.headers or {}) if file_template else {}
    for key, value in auth.items():
        header_set(headers, key, value)
    header_set(headers, "accept", "*/*")
    header_set(headers, "origin", "https://seller.wildberries.ru")
    header_set(headers, "referer", "https://seller.wildberries.ru/")
    header_set(headers, "x-download-token", token)

    prepared = sanitized_headers(headers, keep_download_token=True)
    if not header_get(prepared, "x-download-token"):
        fail("WB download: внутренний сбой — x-download-token потерян перед запросом")
    return prepared


def try_download_wb_file(
    session: requests.Session,
    report_id: str,
    token_req_template: CurlRequest,
    file_template: Optional[CurlRequest],
    auth: Dict[str, str],
) -> Tuple[Optional[bytes], int, str]:
    token_data = request_json(session, prepare_token_request(token_req_template))
    token = extract_download_token(token_data)
    if not token:
        fail(f"WB tokensjrpc: не найден download token в ответе: {str(token_data)[:1000]}")
    url = f"{WB_FILE_BASE}/{report_id}"
    headers = build_file_headers(file_template, auth, token)
    log(
        "WB file request: "
        f"x-download-token={'set' if header_get(headers, 'x-download-token') else 'missing'}, "
        f"cookie={'set' if header_get(headers, 'cookie') else 'missing'}"
    )
    response = session.get(url, headers=headers, timeout=180, allow_redirects=True)
    content_type = response.headers.get("content-type", "")
    log(f"WB file GET: status={response.status_code}, content-type={content_type}, bytes={len(response.content):,}")
    if response.status_code == 200 and looks_like_xlsx(response.content):
        xlsx_content, source = unwrap_wb_xlsx(response.content)
        if source == "direct_xlsx":
            log(f"WB payload: direct XLSX, bytes={len(xlsx_content):,}")
        else:
            log(
                "WB payload: outer ZIP extracted, "
                f"member={source!r}, xlsx_bytes={len(xlsx_content):,}"
            )
        return xlsx_content, response.status_code, "ready"

    preview = response.text[:500] if not looks_like_xlsx(response.content) else ""
    return None, response.status_code, preview


def download_wb_entry_period(
    raw_curls: str,
    store: str,
    start: dt.date,
    end: dt.date,
    reports_root: str,
    folder: str,
    dry_run: bool,
    max_wait_seconds: int,
) -> str:
    create_template, list_template, token_template, file_template, auth = classify_wb_requests(raw_curls)
    report_id = str(uuid.uuid4())
    key = wb_entry_key(store, start, end, reports_root, folder)
    log(f"WB entry-points report_id={report_id}")
    log(f"WB entry-points target_key: {key}")
    if dry_run:
        prepared = prepare_wb_create_request(create_template, start, end, report_id)
        log(f"WB dry_run: create_url={prepared.url}, token_url={token_template.url}")
        return key

    session = requests.Session()
    create_data = request_json(session, prepare_wb_create_request(create_template, start, end, report_id))
    log(f"WB create response: {str(create_data)[:500]}")

    deadline = time.monotonic() + max_wait_seconds
    attempt = 0
    last_state = "unknown"
    last_error = ""
    while time.monotonic() < deadline:
        attempt += 1
        try:
            list_data = request_json(session, list_template)
            entry = find_dict_by_id(list_data, report_id)
            last_state = report_state(entry)
            log(f"WB poll #{attempt}: state={last_state}")
            failed_tokens = ("fail", "error", "cancel", "reject")
            if any(token in last_state for token in failed_tokens):
                fail(f"WB не сформировал отчёт {report_id}: state={last_state}, entry={entry}")
        except SystemExit:
            raise
        except Exception as exc:
            last_error = str(exc)
            log(f"WB list warning: {last_error}")

        # Через первые 10 секунд пробуем файл независимо от названия статуса WB.
        if attempt >= 2 or last_state in {"ready", "done", "success", "completed", "created", "finished"}:
            try:
                content, status, preview = try_download_wb_file(session, report_id, token_template, file_template, auth)
                if content:
                    sheet, header_row, rows = validate_wb_entry_xlsx(content)
                    log(f"WB XLSX OK: sheet={sheet}, header_row={header_row}, rows={rows}")
                    upload_to_s3(content, key)
                    return key
                if status in {401, 403}:
                    fail(
                        "WB download: сервер отклонил x-download-token или сессию "
                        f"(HTTP {status}). Ответ: {preview}"
                    )
                last_error = f"download status={status}, response={preview}"
            except SystemExit:
                raise
            except Exception as exc:
                last_error = str(exc)
                log(f"WB download waiting: {last_error}")

        time.sleep(10)

    fail(f"WB отчёт не был скачан за {max_wait_seconds // 60} мин. Последний state={last_state}; {last_error}")


def self_test() -> None:
    commands = split_curl_commands("curl 'https://a.test/x' \\\n -H 'accept: */*'\n\ncurl 'https://b.test/y' -d '{\"x\":1}'")
    assert len(commands) == 2
    assert parse_curl(commands[1]).method == "POST"
    start, end = previous_full_week(dt.date(2026, 7, 30))
    assert start == dt.date(2026, 7, 20) and end == dt.date(2026, 7, 26)
    payload = {"params": {"startDate": "2026-07-13", "endDate": "2026-07-19"}}
    replaced, count = replace_dates_json(payload, start, end)
    assert count == 2 and replaced["params"]["startDate"] == "2026-07-20"
    token_response = {"jsonrpc": "2.0", "result": {"token": "x" * 120}}
    assert extract_download_token(token_response) == "x" * 120

    ordinary = sanitized_headers({"x-download-token": "secret", "accept": "*/*"})
    assert header_get(ordinary, "x-download-token") == ""
    file_headers = sanitized_headers(
        {"x-download-token": "secret", "accept": "*/*"},
        keep_download_token=True,
    )
    assert header_get(file_headers, "x-download-token") == "secret"

    fake_xlsx_buffer = io.BytesIO()
    with zipfile.ZipFile(fake_xlsx_buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "<Types/>")
        archive.writestr("xl/workbook.xml", "<workbook/>")
    fake_xlsx = fake_xlsx_buffer.getvalue()

    outer_buffer = io.BytesIO()
    with zipfile.ZipFile(outer_buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("report.xlsx", fake_xlsx)

    extracted, member = unwrap_wb_xlsx(outer_buffer.getvalue())
    assert member == "report.xlsx"
    assert extracted == fake_xlsx
    assert is_real_xlsx(extracted)
    log("self-test: OK")


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Загрузка внешних источников: Торгстат АБС-анализ и Точки входа ВБ")
    parser.add_argument("--source", default="all", choices=["all", "torgstat_abc", "wb_entry_points"])
    parser.add_argument("--mode", default="auto", choices=["auto", "daily", "weekly", "custom"])
    parser.add_argument("--store", default=DEFAULT_STORE)
    parser.add_argument("--date-from", default="")
    parser.add_argument("--date-to", default="")
    parser.add_argument("--reports-root", default=os.environ.get("REPORTS_ROOT", DEFAULT_REPORTS_ROOT))
    parser.add_argument("--abc-folder", default=os.environ.get("ABC_FOLDER", DEFAULT_ABC_FOLDER))
    parser.add_argument("--wb-entry-folder", default=os.environ.get("WB_ENTRY_POINTS_FOLDER", DEFAULT_WB_ENTRY_FOLDER))
    parser.add_argument("--wb-max-wait-minutes", type=int, default=20)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args(argv)

    log(f"VERSION: {VERSION}")
    if args.self_test:
        self_test()
        return 0

    load_report_env()
    results: List[str] = []

    if args.source in {"all", "torgstat_abc"}:
        raw_torgstat = os.environ.get("TORGSTAT_ABC_CURL", "")
        if not raw_torgstat.strip():
            fail("Не задан secret TORGSTAT_ABC_CURL")
        periods = torgstat_periods(args.mode, args.date_from, args.date_to)
        log(f"Torgstat periods={[(fmt_dmy(a), fmt_dmy(b), label) for a, b, label in periods]}")
        for start, end, label in periods:
            if start > end:
                fail(f"Некорректный период {start}—{end}")
            log(f"--- Torgstat {label}: {fmt_dmy(start)}-{fmt_dmy(end)} ---")
            results.append(download_torgstat_period(raw_torgstat, start, end, args.reports_root, args.abc_folder, args.dry_run))
            time.sleep(1)

    if args.source in {"all", "wb_entry_points"}:
        periods = wb_periods(args.mode, args.date_from, args.date_to)
        if periods:
            raw_wb = os.environ.get("WB_ENTRY_POINTS_CURLS", "")
            if not raw_wb.strip():
                fail("Не задан secret WB_ENTRY_POINTS_CURLS")
            log(f"WB entry-points periods={[(fmt_dmy(a), fmt_dmy(b), label) for a, b, label in periods]}")
            for start, end, label in periods:
                if start > end:
                    fail(f"Некорректный период {start}—{end}")
                log(f"--- WB entry-points {label}: {fmt_dmy(start)}-{fmt_dmy(end)} ---")
                results.append(download_wb_entry_period(
                    raw_wb,
                    args.store,
                    start,
                    end,
                    args.reports_root,
                    args.wb_entry_folder,
                    args.dry_run,
                    max(1, args.wb_max_wait_minutes) * 60,
                ))

    log("DONE")
    for key in results:
        log(f"RESULT_KEY={key}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
